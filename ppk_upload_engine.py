"""
ppk_upload_engine.py — Engine upload dokumen persiapan pengadaan untuk role PPK ke SPSE.
"""

import json
import re
import os
import tempfile
import urllib.request
import requests
from bs4 import BeautifulSoup
import spse_browser
from config import SPSE_BASE_URL, POKJA_ROOT

BASE_URL = SPSE_BASE_URL.rstrip("/")
_LPSE = BASE_URL.rstrip("/").rsplit("/", 1)[-1]  # "tapinkab"


_UPLOAD_NAME_KEYS = ("fileName", "file_name", "nama_file", "name", "filename")
_UPLOAD_ID_KEYS = ("fileId", "file_id", "id")
_UPLOAD_VERSION_KEYS = ("versi", "version", "fileVersion", "file_version")


def _normalized_upload_name(name: str) -> str:
    """Normalisasi nama file untuk membandingkan variasi output SPSE."""
    import unicodedata

    value = unicodedata.normalize("NFKC", os.path.basename(str(name or "")))
    return re.sub(r"\s+", " ", value).strip().casefold()


def _compact_upload_name(name: str) -> str:
    """Bentuk tolerant: abaikan spasi, titik, strip, dan underscore."""
    return "".join(char for char in _normalized_upload_name(name) if char.isalnum())


def _upload_names_match(left: str, right: str) -> bool:
    return (
        _normalized_upload_name(left) == _normalized_upload_name(right)
        or _compact_upload_name(left) == _compact_upload_name(right)
    )


def _first_upload_value(item: dict, keys: tuple[str, ...]):
    if not isinstance(item, dict):
        return None
    for key in keys:
        value = item.get(key)
        if value is not None and str(value).strip() not in {"", "null", "None"}:
            return value
    return None


def _upload_response_entries(result: dict) -> list[dict]:
    """Ambil kandidat entry dari variasi envelope JSON SPSE."""
    response = result.get("response") if isinstance(result, dict) else None
    candidates = []
    for value in (
        result.get("files") if isinstance(result, dict) else None,
        response.get("files") if isinstance(response, dict) else None,
        response.get("result", {}).get("files")
        if isinstance(response, dict) and isinstance(response.get("result"), dict)
        else None,
        response.get("data", {}).get("files")
        if isinstance(response, dict) and isinstance(response.get("data"), dict)
        else None,
        response.get("result", {}).get("data", {}).get("files")
        if isinstance(response, dict)
        and isinstance(response.get("result"), dict)
        and isinstance(response.get("result", {}).get("data"), dict)
        else None,
    ):
        if isinstance(value, list):
            candidates.extend(item for item in value if isinstance(item, dict))

    # Be tolerant of endpoint responses with a single file entry at root.
    if not candidates:
        for item in (response, response.get("result") if isinstance(response, dict) else None,
                     response.get("data") if isinstance(response, dict) else None):
            if isinstance(item, dict) and (
                _first_upload_value(item, _UPLOAD_NAME_KEYS) is not None
                or _first_upload_value(item, _UPLOAD_VERSION_KEYS) is not None
            ):
                candidates.append(item)
    return candidates


def _match_upload_response(
    result: dict,
    file_name: str,
    file_id=None,
    path: str | None = None,
) -> tuple[dict | None, object | None, str | None]:
    """Cocokkan response submit yang variasinya berbeda-beda di SPSE."""
    entries = _upload_response_entries(result)
    matched = None
    for item in entries:
        names = [_first_upload_value(item, (_key,)) for _key in _UPLOAD_NAME_KEYS]
        names = [value for value in names if value is not None]
        ids = [_first_upload_value(item, (_key,)) for _key in _UPLOAD_ID_KEYS]
        ids = [value for value in ids if value is not None]
        item_path = _first_upload_value(item, ("path", "filePath", "file_path"))
        if (
            any(_upload_names_match(value, file_name) for value in names)
            or any(file_id is not None and str(value) == str(file_id) for value in ids)
            or (path and item_path and str(item_path) == str(path))
        ):
            matched = item
            break

    # Jika server hanya mengembalikan satu file dan success=true, entry tunggal
    # adalah hasil upload walau nama di-normalisasi atau field namanya berbeda.
    if matched is None and result.get("ok") is True and len(entries) == 1:
        matched = entries[0]
    version = _first_upload_value(matched, _UPLOAD_VERSION_KEYS)
    if version is None:
        return None, None, (
            "Submit tidak menemukan file/versi yang cocok: "
            + json.dumps(result.get("response"), ensure_ascii=False, default=str)[:800]
        )
    return matched, version, None


_cdp_eval_lock = __import__("threading").Lock()
_CDP_PORT = 9222
_PPK_FETCH_STATE = {"ok": False, "status": None, "reason": "not_fetched"}


def get_ppk_fetch_state() -> dict:
    """Status fetch terakhir agar UI membedakan kosong vs session invalid."""
    return dict(_PPK_FETCH_STATE)


def _cdp_eval(js: str, timeout: int = 30) -> tuple[bool, object, str]:
    """
    Jalankan JS di tab SPSE via pure WebSocket CDP (tanpa Playwright connect_over_cdp).
    Cari tab SPSE dari /json endpoint, connect langsung ke tab WS, eval JS.
    Return (ok, result_value, error_msg).
    """
    import asyncio, json as _json

    async def _run_ws():
        import websockets
        # Ambil daftar tab dari CDP HTTP
        import urllib.request as _ur
        try:
            tabs = _json.loads(_ur.urlopen(f"http://localhost:{_CDP_PORT}/json", timeout=3).read())
        except Exception as e:
            raise RuntimeError(f"CDP HTTP tidak aktif: {e}")

        page_tabs = [t for t in tabs if t.get("type") == "page"]
        if not page_tabs:
            raise RuntimeError("Tidak ada tab page di CDP")

        # Pilih tab SPSE paketnontender, fallback tab SPSE lain
        tab = next(
            (t for t in page_tabs if "paketnontender" in t.get("url", "") and "spse.inaproc.id" in t.get("url", "")),
            next((t for t in page_tabs if "spse.inaproc.id" in t.get("url", "")), page_tabs[0])
        )
        ws_url = tab.get("webSocketDebuggerUrl", "")
        if not ws_url:
            raise RuntimeError(f"Tab tidak punya webSocketDebuggerUrl: {tab.get('url')}")

        async with websockets.connect(ws_url, open_timeout=5) as ws:
            cmd = _json.dumps({"id": 1, "method": "Runtime.evaluate", "params": {
                "expression": js, "returnByValue": True, "awaitPromise": True
            }})
            await ws.send(cmd)
            while True:
                msg = _json.loads(await asyncio.wait_for(ws.recv(), timeout=timeout))
                if msg.get("id") == 1:
                    break
        res = msg.get("result", {})
        exc = res.get("exceptionDetails")
        if exc:
            raise RuntimeError(exc.get("text", "JS exception"))
        return res.get("result", {}).get("value")

    with _cdp_eval_lock:
        try:
            import threading
            result_holder = [None, None]  # [value, exception]
            def _in_thread():
                loop = asyncio.ProactorEventLoop()  # Windows: wajib Proactor untuk WS
                asyncio.set_event_loop(loop)
                try:
                    result_holder[0] = loop.run_until_complete(_run_ws())
                except Exception as e:
                    result_holder[1] = e
                finally:
                    loop.close()
            t = threading.Thread(target=_in_thread, daemon=True)
            t.start()
            t.join(timeout=timeout + 5)
            if t.is_alive():
                return False, None, "CDP eval timeout"
            if result_holder[1]:
                raise result_holder[1]
            return True, result_holder[0], ""
        except Exception as e:
            return False, None, str(e)

_SUBMIT_ENDPOINTS = {
    "kak":     "spekPpkSubmit",
    "kontrak": "uploadSskkSubmit",
    "uraian":  "uploadUraianSubmit",
    "lainnya": "lainnyaPpkSubmit",
}

UPLOAD_TARGET_LABELS = {
    "kak": "KAK / Spesifikasi",
    "kontrak": "Rancangan Kontrak",
    "uraian": "Uraian Singkat",
    "lainnya": "Informasi Lainnya",
    "nd": "Nota Dinas PPK",
}


def upload_target_label(jenis: str) -> str:
    """Nama kategori tujuan SPSE untuk log upload yang konsisten."""
    return UPLOAD_TARGET_LABELS.get(str(jenis or "").strip().lower(), str(jenis or "-"))

_DELETE_ENDPOINTS = {
    "kak":     "hapusspekppk",
    "kontrak": "hapussskkattachment",
    "uraian":  "hapusuraianattachment",
    "lainnya": "hapuslainnyappk",
}

def _headers(referer: str = "") -> dict:
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": referer or BASE_URL,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
    }

def get_cookies_from_browser() -> str:
    """Ambil cookie SPSE via spse_browser"""
    return spse_browser.get_spse_cookies()



def fetch_paket_ppk() -> list[dict]:
    """
    Ambil daftar paket non-tender PPK via CDP WebSocket.
    Cookie HttpOnly tidak bisa diakses Python — fetch dari dalam browser.
    """
    import time as _time
    ts = int(_time.time() * 1000)
    js = f"""
    (async () => {{
        const r = await fetch('/{_LPSE}/dt/paketppknontender?draw=1&start=0&length=200&_={ts}',
                              {{credentials:'include'}});
        if (!r.ok) return {{ok:false, status:r.status, rows:[]}};
        const j = await r.json();
        return {{ok:true, status:r.status, rows:(j.data || []).map(row => ({{
            kode_paket: String(row[0]),
            nama_paket: String(row[1]),
            status:     String(row[2]),
        }}))}};
    }})()
    """
    global _PPK_FETCH_STATE
    ok, val, err = _cdp_eval(js, timeout=20)
    if not ok:
        _PPK_FETCH_STATE = {"ok": False, "status": None, "reason": "cdp_error", "error": err}
        return []
    if not isinstance(val, dict):
        _PPK_FETCH_STATE = {"ok": False, "status": None, "reason": "invalid_response"}
        return []
    status = val.get("status")
    if not val.get("ok"):
        server_statuses = (500, 502, 503, 504)
        _PPK_FETCH_STATE = {
            "ok": False,
            "status": status,
            "reason": (
                "auth_error" if status in (401, 403)
                else "server_error" if status in server_statuses
                else "http_error"
            ),
        }
        return []
    _PPK_FETCH_STATE = {"ok": True, "status": status, "reason": "ok"}
    rows = val.get("rows") or []
    # Tab PPK hanya menampilkan paket aktif berstatus Draft.
    return [p for p in rows if str(p.get("status", "")).strip().lower() == "draft"]

def fetch_detail_paket(kode_paket: str) -> dict:
    """
    Scrape detail paket PPK dari edit?step=1 dan step=2.
    Return dict berisi: kode_rup, mak, nilai_pagu, nilai_hps, tahun_anggaran,
    sumber_dana, lokasi, jenis_kontrak, nama_ppk, instansi, satker.
    """
    cookie_str = get_cookies_from_browser()
    if not cookie_str:
        return {}
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Cookie": cookie_str,
        "Referer": f"{BASE_URL}/paketnontender/{kode_paket}/edit?step=1",
    }
    result: dict = {}

    # ── Step 1 ────────────────────────────────────────────────────────────────
    try:
        r1 = requests.get(f"{BASE_URL}/paketnontender/{kode_paket}/edit?step=1", headers=headers, timeout=15)
        soup1 = BeautifulSoup(r1.text, "html.parser")

        # Metadata family paket. Nilai ini sengaja diambil dari label/select
        # SPSE, bukan ditebak dari nama paket.
        for el in soup1.find_all(["input", "select", "textarea"]):
            name = (el.get("name") or el.get("id") or "").lower()
            value = el.get("value", "")
            if el.name == "select":
                opt = el.find("option", selected=True)
                value = opt.get_text(" ", strip=True) if opt else value
            if not value:
                value = el.get_text(" ", strip=True)
            if not value:
                continue
            if any(k in name for k in ("jenis_pengadaan", "jenispekerjaan", "jenis_pekerjaan", "kategori_pengadaan", "tipe_pengadaan", "package_type")):
                result.setdefault("jenis_pengadaan", value)
            if any(k in name for k in ("jenis_pl", "jenispl")):
                result["jenis_pl"] = value
        for label in soup1.find_all(["label", "th", "td"]):
            label_text = label.get_text(" ", strip=True).lower()
            if not any(k in label_text for k in ("jenis pengadaan", "jenis pekerjaan", "kategori pengadaan")):
                continue
            sibling = label.find_next_sibling()
            if sibling:
                value = sibling.get_text(" ", strip=True)
                if value:
                    result.setdefault("jenis_pengadaan", value)
        # Beberapa versi SPSE merender detail sebagai tabel key/value tanpa
        # input name. Tangkap pasangan sel secara eksplisit.
        for tr in soup1.find_all("tr"):
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
            if len(cells) < 2:
                continue
            key = cells[0].lower().rstrip(":")
            value = cells[1]
            if key in {"jenis pengadaan", "jenis pekerjaan", "kategori pengadaan"} and value:
                result.setdefault("jenis_pengadaan", value)

        # Tabel RUP (kolom: kode_rup, nama_paket_rup, sumber_dana)
        for tbl in soup1.find_all("table"):
            rows = tbl.find_all("tr")
            for tr in rows:
                cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if len(cells) >= 2 and re.match(r"^\d{8,}$", cells[0]):
                    result["kode_rup"] = cells[0]
                    break

        # Tabel Anggaran (kolom: tahun, sumber_dana, kode_rekening, nilai_pagu, nama_ppk)
        tbl_rows_all = []
        for tbl in soup1.find_all("table"):
            for tr in tbl.find_all("tr"):
                cells = [td.get_text(separator=" ", strip=True) for td in tr.find_all(["td", "th"])]
                tbl_rows_all.append(cells)

        for cells in tbl_rows_all:
            if len(cells) >= 3 and re.match(r"^\d{4}$", cells[0]):  # tahun anggaran
                result["tahun_anggaran"] = cells[0]
                if len(cells) > 1: result["sumber_dana"] = cells[1]
                if len(cells) > 2: result["mak"] = cells[2].rstrip(".")
                if len(cells) > 3:
                    # "Rp. 44.950.000,00" → hapus titik ribuan, ambil sebelum koma
                    raw = cells[3].replace("Rp.", "").replace("Rp", "").strip()
                    raw = raw.replace(".", "").split(",")[0]
                    if raw.isdigit(): result["nilai_pagu"] = int(raw)
                if len(cells) > 4: result["nama_ppk"] = cells[4]
                break

        # K/L/PD (Instansi) & Satker
        for label in soup1.find_all(["label", "th", "td"]):
            txt = label.get_text(strip=True)
            if "K/L/PD" in txt or "Instansi" in txt:
                sib = label.find_next_sibling()
                if sib: result["instansi"] = sib.get_text(strip=True)
            if "Satuan Kerja" in txt:
                sib = label.find_next_sibling()
                if sib: result["satker"] = sib.get_text(strip=True)

        # Lokasi: input textlokasi
        inp_lok = soup1.find("input", {"id": "textlokasi"})
        if inp_lok:
            result["lokasi"] = inp_lok.get("value", "")
        sel_kab = soup1.find("select", {"id": "kabupaten"})
        if sel_kab:
            sel_opt = sel_kab.find("option", {"selected": True})
            if sel_opt:
                kab_txt = sel_opt.get_text(strip=True)
                lok = result.get("lokasi", "")
                result["lokasi"] = f"{kab_txt}, {lok}".strip(", ") if lok else kab_txt

    except Exception as e:
        result["_error_step1"] = str(e)

    # ── Step 2: nilai HPS + jenis kontrak ─────────────────────────────────────
    try:
        r2 = requests.get(f"{BASE_URL}/paketnontender/{kode_paket}/edit?step=2", headers=headers, timeout=15)
        soup2 = BeautifulSoup(r2.text, "html.parser")

        # Nilai HPS dari tabel row "Nilai HPS"
        for tr in soup2.find_all("tr"):
            cells = [td.get_text(separator=" ", strip=True) for td in tr.find_all(["td", "th"])]
            if cells and "Nilai HPS" in cells[0] and len(cells) > 1:
                raw = cells[1].replace("Rp.", "").replace("Rp", "").strip()
                raw = raw.split()[0].replace(".", "").split(",")[0]
                if raw.isdigit(): result["nilai_hps"] = int(raw)
                break

        # Jenis kontrak
        sel_k = soup2.find("select", {"name": "kontrak_pembayaran"})
        if sel_k:
            opt = sel_k.find("option", {"selected": True})
            if opt: result["jenis_kontrak"] = opt.get_text(strip=True)

    except Exception as e:
        result["_error_step2"] = str(e)

    return result


def upload_dokumen(
    kode_paket: str,
    jenis: str,
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
    cookies: dict = None,
    log_fn=None
) -> dict:
    """
    Upload dokumen PPK via CDP page.evaluate (5 langkah).
    Cookie PPK ber-flag HttpOnly — seluruh flow dijalankan dari dalam browser.
    File bytes dikirim sebagai base64, di-decode di browser lalu di-upload.
    """
    import base64

    def _log(msg):
        if log_fn:
            log_fn(msg)

    sub_endpoint = _SUBMIT_ENDPOINTS.get(jenis)
    if not sub_endpoint:
        return {"ok": False, "error": f"Jenis '{jenis}' tidak dikenal."}

    _log(f"Langkah 1-5: Upload via browser (CDP)...")
    b64 = base64.b64encode(file_bytes).decode()

    js = f"""
    (async () => {{
        const lpse = "{_LPSE}";
        const kode = "{kode_paket}";
        const mime = "{mime_type}";
        const fname = "{file_name}";
        const subEndpoint = "{sub_endpoint}";
        const b64 = "{b64}";

        const binaryStr = atob(b64);
        const bytes = new Uint8Array(binaryStr.length);
        for (let i = 0; i < binaryStr.length; i++) bytes[i] = binaryStr.charCodeAt(i);

        await fetch('/' + lpse + '/otorisasiDataPaketPlUpload?id=' + kode, {{credentials:'include'}});

        const formData = new FormData();
        formData.append('input[uploadSignedUrlReq][0][contentType]', mime);
        formData.append('input[uploadSignedUrlReq][0][identifier]', '');
        formData.append('input[uploadSignedUrlReq][0][fileName]', fname);
        formData.append('input[uploadSignedUrlReq][0][isPublic]', 'false');
        formData.append('isArchieve', 'true');

        const r2 = await fetch('/' + lpse + '/paketnontender/' + kode + '/getSignedUrl', {{
            method: 'POST', credentials: 'include', body: formData
        }});
        if (!r2.ok) return {{ok: false, error: 'getSignedUrl HTTP ' + r2.status}};
        const res2 = await r2.json();
        const fileId = res2?.result?.data?.fileId;
        const signedUrl = res2?.result?.data?.signedUrl;
        const path = res2?.result?.data?.path || res2?.path;
        if (!fileId || !signedUrl || !path) return {{ok: false, error: 'getSignedUrl data tidak lengkap: ' + JSON.stringify(res2)}};

        const r3 = await fetch(signedUrl, {{
            method: 'PUT', headers: {{'Content-Type': mime}}, body: bytes
        }});
        if (!r3.ok) return {{ok: false, error: 'PUT storage HTTP ' + r3.status}};

        for (let i = 0; i < 5; i++) {{
            const fd4 = new FormData();
            fd4.append('input', fileId);
            const r4 = await fetch('/' + lpse + '/uploadCheckStatus', {{
                method: 'POST', credentials: 'include', body: fd4
            }});
            let d4 = null;
            try {{ d4 = await r4.json(); }} catch(e) {{ break; }}  // parse error = lanjut
            if (!d4 || d4?.errors === null) break;  // null errors = sukses
            if (d4?.errors) return {{ok: false, error: 'checkStatus error: ' + JSON.stringify(d4.errors)}};
            const st = d4?.data?.status;
            if (!st || st === 'UPLOAD_SUCCESS') break;
            if (st === 'UPLOAD_FAILED') return {{ok: false, error: 'Upload dinyatakan gagal server'}};
            await new Promise(r => setTimeout(r, 300));
        }}

        const fd5 = new FormData();
        fd5.append('id', kode);
        fd5.append('path', path);
        fd5.append('fileId', fileId);
        const r5 = await fetch('/' + lpse + '/dokumennontender/' + kode + '/' + subEndpoint, {{
            method: 'POST', credentials: 'include', body: fd5
        }});
        if (!r5.ok) return {{ok: false, error: 'submit HTTP ' + r5.status}};
        let res5 = null;
        try {{ res5 = await r5.json(); }} catch(e) {{
            return {{ok: false, status: r5.status, error: 'Submit response bukan JSON'}};
        }}
        const fileLists = [
            res5?.files,
            res5?.result?.files,
            res5?.data?.files,
            res5?.result?.data?.files
        ].filter(Array.isArray);
        const files = fileLists.flat();
        // Matching nama/ID/version dilakukan di Python agar fallback dan
        // regresinya bisa diuji tanpa browser. SPSE kadang mengembalikan
        // nama yang sudah dinormalisasi atau hanya satu entry.
        return {{
            ok: res5?.success === true,
            status: r5.status,
            path,
            fileId,
            files,
            response: res5
        }};
    }})()
    """
    ok, result, err = _cdp_eval(js, timeout=120)
    if not ok:
        return {"ok": False, "error": err}
    if result and result.get("ok"):
        _entry, _versi, _match_error = _match_upload_response(
            result, file_name, result.get("fileId"), result.get("path")
        )
        if _entry is None or _versi is None:
            result["ok"] = False
            result["error"] = _match_error or "Submit tidak menemukan file/versi yang cocok"
        else:
            result["entry"] = _entry
            result["versi"] = _versi
        if result.get("ok"):
            _log(f"✅ {file_name} → {upload_target_label(jenis)} berhasil diupload")
        else:
            _log(
                f"❌ {file_name} → {upload_target_label(jenis)} gagal: "
                f"{result.get('error') or 'respons tidak cocok'}"
            )
    else:
        _log(
            f"❌ {file_name} → {upload_target_label(jenis)} gagal: "
            f"{result.get('error') if result else err}"
        )
    return result or {"ok": False, "error": err}

def upload_nota_dinas(kode_paket: str, file_bytes: bytes, file_name: str, mime_type: str, log_fn=None) -> dict:
    """
    Upload Nota Dinas via GCS flow + submit ke uploadAttachment step=1.
    """
    import base64

    def _log(msg):
        if log_fn: log_fn(msg)

    _log(f"Langkah 1-5: Upload ND via browser (CDP)...")
    b64 = base64.b64encode(file_bytes).decode()

    js = f"""
    (async () => {{
        const lpse = "{_LPSE}";
        const kode = "{kode_paket}";
        const mime = "{mime_type}";
        const fname = "{file_name}";
        const b64 = "{b64}";

        const binaryStr = atob(b64);
        const bytes = new Uint8Array(binaryStr.length);
        for (let i = 0; i < binaryStr.length; i++) bytes[i] = binaryStr.charCodeAt(i);

        await fetch('/' + lpse + '/otorisasiDataPaketPlUpload?id=' + kode, {{credentials:'include'}});

        const formData = new FormData();
        formData.append('input[uploadSignedUrlReq][0][contentType]', mime);
        formData.append('input[uploadSignedUrlReq][0][identifier]', '');
        formData.append('input[uploadSignedUrlReq][0][fileName]', fname);
        formData.append('input[uploadSignedUrlReq][0][isPublic]', 'false');
        formData.append('isArchieve', 'true');

        const r2 = await fetch('/' + lpse + '/paketnontender/' + kode + '/getSignedUrl', {{
            method: 'POST', credentials: 'include', body: formData
        }});
        if (!r2.ok) return {{ok: false, error: 'getSignedUrl HTTP ' + r2.status}};
        const res2 = await r2.json();
        const fileId = res2?.result?.data?.fileId;
        const signedUrl = res2?.result?.data?.signedUrl;
        const path = res2?.result?.data?.path || res2?.path;
        if (!fileId || !signedUrl || !path) return {{ok: false, error: 'getSignedUrl data tidak lengkap: ' + JSON.stringify(res2)}};

        const r3 = await fetch(signedUrl, {{
            method: 'PUT', headers: {{'Content-Type': mime}}, body: bytes
        }});
        if (!r3.ok) return {{ok: false, error: 'PUT storage HTTP ' + r3.status}};

        for (let i = 0; i < 5; i++) {{
            const fd4 = new FormData();
            fd4.append('input', fileId);
            const r4 = await fetch('/' + lpse + '/uploadCheckStatus', {{
                method: 'POST', credentials: 'include', body: fd4
            }});
            let d4 = null;
            try {{ d4 = await r4.json(); }} catch(e) {{ break; }}
            if (!d4 || d4?.errors === null) break;
            if (d4?.errors) return {{ok: false, error: 'checkStatus error: ' + JSON.stringify(d4.errors)}};
            const st = d4?.data?.status;
            if (!st || st === 'UPLOAD_SUCCESS') break;
            if (st === 'UPLOAD_FAILED') return {{ok: false, error: 'Upload dinyatakan gagal server'}};
            await new Promise(r => setTimeout(r, 300));
        }}

        const fd5 = new FormData();
        fd5.append('id', kode);
        fd5.append('step', '1');
        fd5.append('path', path);
        fd5.append('fileId', fileId);
        const r5 = await fetch('/' + lpse + '/paketnontender/' + kode + '/uploadAttachment', {{
            method: 'POST', credentials: 'include', body: fd5
        }});
        if (!r5.ok) return {{ok: false, error: 'submit HTTP ' + r5.status}};

        return {{ok: true, path, fileId}};
    }})()
    """
    ok, result, err = _cdp_eval(js, timeout=120)
    if not ok:
        return {"ok": False, "error": err}
    if result and result.get("ok"):
        _log(f"✅ {file_name} berhasil diupload")
    else:
        _log(f"❌ {file_name} gagal: {result.get('error') if result else err}")
    return result or {"ok": False, "error": err}

def pilih_pp(kode_paket: str, pp_id: str = "74177", log_fn=None) -> bool:
    """
    Step 3 SPSE: pilih PP via /pilihpp -> submit_pp.
    Ambil token dari halaman /pilihpp, POST ke /submit_pp dengan field ppId (bukan pp_id).
    """
    js = f"""
    (async () => {{
        const tokenResp = await fetch('/{_LPSE}/paketnontender/{kode_paket}/pilihpp', {{credentials:'include'}});
        const html = await tokenResp.text();
        const tokenMatch = html.match(/authenticityToken[^>]*value="([^"]+)"/);
        const token = tokenMatch ? tokenMatch[1] : '';
        const fd = new FormData();
        fd.append('authenticityToken', token);
        fd.append('ppId', '{pp_id}');
        const r = await fetch('/{_LPSE}/paketnontender/{kode_paket}/submit_pp', {{
            method: 'POST', credentials: 'include', body: fd
        }});
        return {{status: r.status, ok: r.ok, url: r.url}};
    }})()
    """
    ok, val, _ = _cdp_eval(js, timeout=20)
    if not ok or not val:
        return False
    return val.get("ok") is True or val.get("status") in (200, 302)


def kirim_email_pp(kode_paket: str, path: str, file_id: str, log_fn=None) -> bool:
    """
    Kirim email pemberitahuan ke PP_ISRANDRIA (pp_id=74177) setelah ND terupload.
    Wajib panggil pilih_pp() dulu sebelum fungsi ini.
    """
    js = f"""
    (async () => {{
        const fd = new FormData();
        fd.append('id', '{kode_paket}');
        fd.append('pp_id', '74177');
        fd.append('path', '{path}');
        fd.append('fileId', '{file_id}');
        const r = await fetch('/{_LPSE}/paketnontender/{kode_paket}/submitrekirimpesanpp?pp_id=74177', {{
            method: 'POST', credentials: 'include', body: fd, redirect: 'manual'
        }});
        return {{status: r.status}};
    }})()
    """
    ok, val, _ = _cdp_eval(js, timeout=20)
    if not ok or not val: return False
    st = val.get("status")
    return st in (0, 200, 302)


def simpan_dan_membuat_paket(kode_paket: str, log_fn=None) -> dict:
    """Submit final step 3 SPSE setelah PP, Nota Dinas, dan email siap.

    Token diambil ulang dari halaman step 3 agar tidak memakai token lama
    setelah rangkaian upload/email sebelumnya.
    """
    import json as _json

    def _log(msg):
        if log_fn:
            log_fn(msg)

    kode_js = _json.dumps(str(kode_paket))
    js = f"""
    (async () => {{
        const lpse = {_json.dumps(_LPSE)};
        const kode = {kode_js};
        const editUrl = '/' + lpse + '/paketnontender/' + kode + '/edit?step=3';
        const submitPath = '/' + lpse + '/paketnontender/' + kode + '/simpanpartthree';
        const pageResp = await fetch(editUrl, {{credentials: 'include', cache: 'no-store'}});
        if (!pageResp.ok) return {{ok: false, error: 'Gagal membaca step 3 HTTP ' + pageResp.status}};

        const html = await pageResp.text();
        const doc = new DOMParser().parseFromString(html, 'text/html');
        const form = doc.querySelector('#formTambahPaket');
        if (!form) return {{ok: false, error: 'Form #formTambahPaket tidak ditemukan pada step 3'}};

        const fd = new FormData(form);
        fd.set('step', fd.get('step') || '3');
        fd.set('flow', fd.get('flow') || '1');
        fd.set('simpan', 'simpan');
        const response = await fetch(submitPath, {{
            method: 'POST', credentials: 'include', body: fd, redirect: 'follow'
        }});
        const responseText = await response.text();
        const finalUrl = response.url || '';
        const followed = !finalUrl.includes('/simpanpartthree');
        const success = response.ok && followed;
        return {{
            ok: success,
            status: response.status,
            finalUrl,
            error: success ? '' : ('Submit final tidak terkonfirmasi (HTTP ' + response.status + ', URL ' + finalUrl + ')'),
            preview: responseText.slice(0, 300)
        }};
    }})()
    """
    ok, val, err = _cdp_eval(js, timeout=30)
    if not ok or not val:
        _log(f"❌ Simpan dan Membuat Paket gagal: {err or 'respons kosong'}")
        return {"ok": False, "error": err or "Respons kosong dari submit final"}
    if val.get("ok"):
        _log("✅ Simpan dan Membuat Paket berhasil")
    else:
        _log(f"❌ Simpan dan Membuat Paket gagal: {val.get('error', 'respons tidak terkonfirmasi')}")
    return val

_LIST_ENDPOINTS = {
    "kak":     "spekppk",
    "kontrak": "uploadsskk",
    "uraian":  "uploaduraian",
    "lainnya": "lainnyappk",
    "nd":      "notadinasppk",
}

def list_semua_dokumen(kode_paket: str) -> dict[str, list[dict]]:
    """
    Ambil dokumen semua jenis via cdp_eval + fetch() + DOMParser.
    Tidak membuka tab baru — fetch dari tab SPSE yang sudah aktif.
    """
    import json as _json
    endpoints = {k: v for k, v in _LIST_ENDPOINTS.items() if k != "nd"}
    lpse = _LPSE
    js = f"""
(async () => {{
  const lpse = {_json.dumps(lpse)};
  const kode = {_json.dumps(kode_paket)};
  const endpoints = {_json.dumps(endpoints)};
  const results = {{}};
  await Promise.all(Object.entries(endpoints).map(async ([jenis, ep]) => {{
    try {{
      const r = await fetch('/' + lpse + '/dokumennontender/' + kode + '/' + ep, {{credentials: 'include'}});
      const html = await r.text();
      const parser = new DOMParser();
      const doc = parser.parseFromString(html, 'text/html');
      const rows = doc.querySelectorAll('#files tbody tr');
      const docs = [];
      rows.forEach(tr => {{
        const a = tr.querySelector('td a');
        const rem = tr.querySelector('.removeDok');
        const versi = rem ? parseInt(rem.getAttribute('versi') || '0') : 0;
        if (a && a.textContent.trim()) {{
          docs.push({{nama_file: a.textContent.trim(), url_dl: a.href || '', versi: versi}});
        }}
      }});
      results[jenis] = docs;
    }} catch(e) {{ results[jenis] = []; }}
  }}));
  return results;
}})()
"""
    ok, val, _ = _cdp_eval(js, timeout=20)
    return val if ok and isinstance(val, dict) else {}


def list_dokumen_cdp(kode_paket: str, jenis: str) -> dict:
    """Ambil satu kategori via CDP dan bedakan error halaman dari daftar kosong."""
    import json as _json

    endpoint = _LIST_ENDPOINTS.get(jenis)
    if not endpoint:
        return {"ok": False, "status": 0, "documents": [], "error": f"Jenis '{jenis}' tidak dikenal."}
    js = f"""
(async () => {{
  const url = '/{_LPSE}/dokumennontender/' + {_json.dumps(kode_paket)} + '/{endpoint}';
  const r = await fetch(url, {{credentials: 'include'}});
  const html = await r.text();
  if (!r.ok) return {{ok: false, status: r.status, documents: [], error: 'List HTTP ' + r.status}};
  const doc = new DOMParser().parseFromString(html, 'text/html');
  const table = doc.querySelector('#files');
  if (!table) return {{ok: false, status: r.status, documents: [], error: 'Tabel #files tidak ditemukan'}};
  const documents = [];
  table.querySelectorAll('tbody tr').forEach(tr => {{
    const a = tr.querySelector('td a');
    const rem = tr.querySelector('.removeDok');
    if (a && a.textContent.trim()) documents.push({{
      nama_file: a.textContent.trim(),
      url_dl: a.href || '',
      versi: rem ? parseInt(rem.getAttribute('versi') || '', 10) : null
    }});
  }});
  return {{ok: true, status: r.status, documents}};
}})()
"""
    ok, value, err = _cdp_eval(js, timeout=20)
    if not ok or not isinstance(value, dict):
        return {"ok": False, "status": 0, "documents": [], "error": err or "Respons list kosong"}
    value.setdefault("documents", [])
    value.setdefault("status", 0)
    value.setdefault("ok", False)
    return value


def list_bulk_semua_paket(kode_list: list[str]) -> dict[str, dict[str, list[dict]]]:
    """
    Ambil dokumen semua jenis untuk N paket via cdp_eval + fetch() paralel.
    Tidak membuka tab baru — semua fetch dari tab SPSE yang sudah aktif.
    """
    import json as _json
    endpoints = {k: v for k, v in _LIST_ENDPOINTS.items() if k != "nd"}
    lpse = _LPSE
    js = f"""
(async () => {{
  const lpse = {_json.dumps(lpse)};
  const kode_list = {_json.dumps(kode_list)};
  const endpoints = {_json.dumps(endpoints)};
  const out = {{}};
  const maxPackagesPerBatch = 8;
  for (let start = 0; start < kode_list.length; start += maxPackagesPerBatch) {{
    const batch = kode_list.slice(start, start + maxPackagesPerBatch);
    await Promise.all(batch.map(async kode => {{
      out[kode] = {{}};
      await Promise.all(Object.entries(endpoints).map(async ([jenis, ep]) => {{
      try {{
        const r = await fetch('/' + lpse + '/dokumennontender/' + kode + '/' + ep, {{credentials: 'include'}});
        const html = await r.text();
        const parser = new DOMParser();
        const doc = parser.parseFromString(html, 'text/html');
        const rows = doc.querySelectorAll('#files tbody tr');
        const docs = [];
        rows.forEach(tr => {{
          const a = tr.querySelector('td a');
          const rem = tr.querySelector('.removeDok');
          const versi = rem ? parseInt(rem.getAttribute('versi') || '0') : 0;
          if (a && a.textContent.trim()) {{
            docs.push({{nama_file: a.textContent.trim(), url_dl: a.href || '', versi: versi}});
          }}
        }});
        out[kode][jenis] = docs;
      }} catch(e) {{ out[kode][jenis] = []; }}
      }}));
    }}));
  }}
  return out;
}})()
"""
    ok, val, _ = _cdp_eval(js, timeout=30)
    return val if ok and isinstance(val, dict) else {}


def list_dokumen(kode_paket: str, jenis: str, cookies: dict = None) -> list[dict]:
    """
    Ambil daftar dokumen terunggah via Playwright goto (tabel diisi JS, bukan SSR).
    Navigate ke endpoint, tunggu #files tbody terisi, parse .removeDok.
    """
    endpoint = _LIST_ENDPOINTS.get(jenis)
    if not endpoint:
        return []

    url = f"https://spse.inaproc.id/{_LPSE}/dokumennontender/{kode_paket}/{endpoint}"
    import subprocess, sys, json as _json, tempfile, os

    script = f"""
import sys, json
from playwright.sync_api import sync_playwright

url = {_json.dumps(url)}
with sync_playwright() as pw:
    browser = pw.chromium.connect_over_cdp("http://localhost:9222")
    ctx = browser.contexts[0]
    # Buka halaman baru agar tidak ganggu tab user
    page = ctx.new_page()
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=15000)
        # Tunggu jQuery populate #files tbody (via uploadFlow successData)
        # Max 8 detik; kalau tidak ada row, kembalikan []
        try:
            page.wait_for_function(
                "() => document.querySelectorAll('#files tbody tr td').length > 0",
                timeout=8000
            )
        except Exception:
            pass  # timeout = tidak ada file
        result = page.evaluate('''() => {{
            const rows = document.querySelectorAll("#files tbody tr");
            const out = [];
            rows.forEach(tr => {{
                const a = tr.querySelector("td a");
                const rem = tr.querySelector(".removeDok");
                const versi = rem ? parseInt(rem.getAttribute("versi") || "0") : 0;
                if (a && a.textContent.trim()) {{
                    out.push({{
                        nama_file: a.textContent.trim(),
                        url_dl: a.href || "",
                        versi: versi,
                    }});
                }}
            }});
            return out;
        }}''')
        print(json.dumps({{"ok": True, "value": result}}))
    except Exception as e:
        print(json.dumps({{"ok": False, "error": str(e)}}))
    finally:
        page.close()
"""
    fd, path = tempfile.mkstemp(suffix="_list_dok.py", prefix="pokja_")
    os.write(fd, script.encode())
    os.close(fd)
    try:
        proc = subprocess.run(
            [sys.executable, path],
            capture_output=True, timeout=30
        )
        stdout = proc.stdout.decode(errors="replace").strip()
        if stdout:
            resp = json.loads(stdout)
            if resp.get("ok"):
                return resp.get("value") or []
        return []
    except Exception:
        return []
    finally:
        os.unlink(path)


def download_existing_document(doc: dict) -> dict:
    """Backup bytes dokumen lama via tab SPSE sebelum replacement berisiko."""
    import base64 as _base64
    import json as _json
    import mimetypes as _mimetypes

    url = (doc or {}).get("url_dl") or (doc or {}).get("url")
    if not url:
        return {"ok": False, "error": "URL download dokumen lama tidak tersedia"}
    js = f"""
(async () => {{
  try {{
    const r = await fetch({_json.dumps(url)}, {{credentials: 'include'}});
    if (!r.ok) return {{ok: false, error: 'Download HTTP ' + r.status}};
    const bytes = new Uint8Array(await r.arrayBuffer());
    let binary = '';
    const chunk = 0x8000;
    for (let i = 0; i < bytes.length; i += chunk) {{
      binary += String.fromCharCode(...bytes.subarray(i, i + chunk));
    }}
    return {{ok: true, b64: btoa(binary), mime: r.headers.get('content-type') || ''}};
  }} catch (e) {{ return {{ok: false, error: String(e)}}; }}
}})()
"""
    ok, value, err = _cdp_eval(js, timeout=120)
    if not ok or not isinstance(value, dict) or not value.get("ok"):
        return {"ok": False, "error": (value or {}).get("error") if isinstance(value, dict) else err}
    try:
        data = _base64.b64decode(value.get("b64") or "", validate=True)
    except Exception as exc:
        return {"ok": False, "error": f"Backup base64 tidak valid: {exc}"}
    name = str((doc or {}).get("nama_file") or "dokumen.pdf")
    return {
        "ok": True,
        "file_bytes": data,
        "mime_type": value.get("mime") or _mimetypes.guess_type(name)[0] or "application/octet-stream",
        "file_name": name,
    }


def hapus_dokumen(kode_paket: str, jenis: str, versi: int, cookies: dict = None) -> bool:
    """
    Hapus dokumen via CDP (cookie HttpOnly).
    """
    del_endpoint = _DELETE_ENDPOINTS.get(jenis)
    if not del_endpoint:
        return False

    js = f"""
    (async () => {{
        const fd = new FormData();
        fd.append('versi', '{versi}');
        const r = await fetch('/{_LPSE}/dokumennontender/{kode_paket}/{del_endpoint}',
                              {{method:'POST', credentials:'include', body:fd}});
        return r.ok || r.status === 302;
    }})()
    """
    ok, val, _ = _cdp_eval(js, timeout=15)
    return bool(ok and val)


_PPK_REPLACE_JENIS = frozenset({"kak", "kontrak", "uraian", "lainnya"})


def _upload_slot(name: str) -> int | None:
    """Ambil slot numerik dari nama file, mis. ``11. Diskresi.pdf`` -> 11."""
    match = re.match(r"^\s*(\d+)\s*\.\s+", os.path.basename(str(name or "")))
    return int(match.group(1)) if match else None


def _replacement_candidates(existing_docs: list[dict] | None, new_name: str) -> list[dict]:
    """Cari dokumen lama pada slot yang sama, atau exact name bila tanpa nomor."""
    docs = [doc for doc in (existing_docs or []) if isinstance(doc, dict)]
    new_slot = _upload_slot(new_name)
    if new_slot is not None:
        return [
            doc for doc in docs
            if _upload_slot(doc.get("nama_file")) == new_slot
            and doc.get("versi") is not None
        ]
    return [
        doc for doc in docs
        if _upload_names_match(doc.get("nama_file"), new_name)
        and doc.get("versi") is not None
    ]


def verifikasi_dokumen_terunggah(
    kode_paket: str,
    jenis: str,
    file_name: str,
    versi: int | None = None,
    retries: int = 3,
    delay: float = 0.35,
) -> dict:
    """Pastikan file terlihat pada endpoint list SPSE sebelum replacement."""
    import time

    if versi is None:
        return {
            "verified": False,
            "versi": None,
            "documents": [],
            "error": "Versi hasil upload kosong; verifikasi dibatalkan",
        }
    last_docs: list[dict] = []
    last_error = ""
    for attempt in range(max(1, retries)):
        try:
            listed = list_dokumen_cdp(kode_paket, jenis)
            last_docs = list(listed.get("documents", []) or [])
            if not listed.get("ok"):
                last_error = listed.get("error", "Endpoint list SPSE gagal")
                last_docs = []
                continue
        except Exception as exc:
            last_docs = []
            last_error = str(exc)
        for doc in last_docs:
            if _upload_names_match(doc.get("nama_file"), file_name):
                verified_version = doc.get("versi")
                if verified_version is not None and str(verified_version) == str(versi):
                    return {
                        "verified": True,
                        "versi": verified_version,
                        "document": doc,
                    }
        if attempt + 1 < max(1, retries):
            time.sleep(delay)
    return {
        "verified": False,
        "versi": None,
        "documents": last_docs,
        "error": last_error or f"Dokumen {file_name} belum muncul pada daftar SPSE",
    }


def _is_duplicate_upload_error(result: dict) -> bool:
    """True hanya untuk penolakan nama file duplikat dari endpoint SPSE."""
    if not isinstance(result, dict):
        return False
    text = json.dumps(result, ensure_ascii=False, default=str).casefold()
    markers = (
        "file dengan nama yang sama",
        "nama yang sama telah ada",
        "duplicate",
        "already exists",
    )
    return any(marker in text for marker in markers)


def upload_dokumen_dengan_replace(
    kode_paket: str,
    jenis: str,
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
    existing_docs: list[dict] | None = None,
    cookies: dict = None,
    log_fn=None,
    upload_fn=None,
    delete_fn=None,
    verify_fn=None,
    backup_fn=None,
) -> dict:
    """Upload dokumen lalu hapus versi lama pada slot/nama yang sama.

    Hanya empat kategori PPK yang boleh auto-replace. Nota Dinas sengaja
    diproteksi di helper ini agar tidak ikut terhapus bila caller keliru.
    Kegagalan upload tidak pernah memicu penghapusan. Jalur nama duplikat
    membuat backup via CDP sebelum delete dan mencoba rollback jika retry/
    verifikasi gagal.
    """
    uploader = upload_fn or upload_dokumen
    deleter = delete_fn or hapus_dokumen
    result = uploader(
        kode_paket=kode_paket,
        jenis=jenis,
        file_bytes=file_bytes,
        file_name=file_name,
        mime_type=mime_type,
        cookies=cookies,
        log_fn=log_fn,
    ) or {"ok": False, "error": "Upload mengembalikan hasil kosong."}
    result = dict(result)
    result.setdefault("replaced_versions", [])
    result.setdefault("replacement_errors", [])
    pre_replaced_versions = []
    verifier = verify_fn or verifikasi_dokumen_terunggah
    backupper = backup_fn or download_existing_document

    def _restore_backups(_backups, _delete_new_version=None):
        """Pulihkan dokumen lama; dipakai hanya saat jalur duplicate gagal."""
        if _delete_new_version is not None:
            try:
                deleter(kode_paket, jenis, _delete_new_version, cookies=cookies)
            except Exception as exc:
                result["replacement_errors"].append(
                    f"versi retry {_delete_new_version} gagal dihapus saat rollback: {exc}"
                )
        for _backup in _backups:
            _restored = uploader(
                kode_paket=kode_paket,
                jenis=jenis,
                file_bytes=_backup["file_bytes"],
                file_name=_backup["file_name"],
                mime_type=_backup["mime_type"],
                cookies=cookies,
                log_fn=log_fn,
            ) or {"ok": False, "error": "Respons restore kosong"}
            if not _restored.get("ok"):
                result["replacement_errors"].append(
                    f"restore {_backup['file_name']} gagal: {_restored.get('error', 'unknown')}"
                )

    # Endpoint SPSE menolak nama duplikat sebelum versi lama bisa diganti.
    # Hanya pada error eksplisit ini versi lama boleh dihapus lalu upload
    # dicoba ulang. Error upload lain tetap tidak boleh menghapus data lama.
    if (
        not result.get("ok")
        and jenis in _PPK_REPLACE_JENIS
        and _is_duplicate_upload_error(result)
    ):
        _old_docs = list(existing_docs or [])
        _candidates = _replacement_candidates(_old_docs, file_name)
        # Cache session setelah upload hanya menyimpan nama/versi. Ambil
        # metadata URL terbaru sebelum backup bila URL belum tersedia.
        if _candidates and any(not (doc.get("url_dl") or doc.get("url")) for doc in _candidates):
            try:
                _old_docs = list_dokumen(kode_paket, jenis)
                _candidates = _replacement_candidates(_old_docs, file_name)
            except Exception as exc:
                result["replacement_errors"].append(
                    f"daftar dokumen lama tidak terbaca: {exc}"
                )
        if not _candidates:
            try:
                _old_docs = list_dokumen(kode_paket, jenis)
                _candidates = _replacement_candidates(_old_docs, file_name)
            except Exception as exc:
                result["replacement_errors"].append(
                    f"daftar dokumen lama tidak terbaca: {exc}"
                )
        if not _candidates:
            result["replacement_errors"].append(
                "SPSE menolak nama duplikat, tetapi versi lama tidak ditemukan"
            )
            return result

        _backups = []
        for _old_doc in _candidates:
            try:
                _backup = backupper(_old_doc) or {"ok": False}
            except Exception as exc:
                _backup = {"ok": False, "error": str(exc)}
            if not _backup.get("ok"):
                result["replacement_errors"].append(
                    f"backup {_old_doc.get('nama_file', _old_doc.get('versi'))} gagal: "
                    f"{_backup.get('error', 'respons backup tidak valid')}"
                )
                result["error"] = (
                    result.get("error", "Upload ditolak SPSE")
                    + "; dokumen lama dipertahankan karena backup gagal"
                )
                return result
            _backups.append({
                "file_name": _backup.get("file_name") or _old_doc.get("nama_file"),
                "file_bytes": _backup.get("file_bytes"),
                "mime_type": _backup.get("mime_type") or "application/octet-stream",
                "versi": _old_doc.get("versi"),
            })

        for _old_doc in _candidates:
            _old_versi = _old_doc.get("versi")
            try:
                _deleted = deleter(
                    kode_paket, jenis, _old_versi, cookies=cookies
                )
            except Exception as exc:
                _deleted = False
                result["replacement_errors"].append(
                    f"{_old_doc.get('nama_file', _old_versi)}: {exc}"
                )
            if _deleted:
                pre_replaced_versions.append(_old_versi)
            elif _old_versi not in pre_replaced_versions:
                result["replacement_errors"].append(
                    f"{_old_doc.get('nama_file', _old_versi)}: gagal dihapus"
                )

        if len(pre_replaced_versions) != len(_candidates):
            _restore_backups(
                [backup for backup in _backups if backup["versi"] in pre_replaced_versions]
            )
            result["error"] = (
                result.get("error", "Upload ditolak SPSE")
                + "; versi lama belum seluruhnya terhapus"
            )
            return result

        _retry = uploader(
            kode_paket=kode_paket,
            jenis=jenis,
            file_bytes=file_bytes,
            file_name=file_name,
            mime_type=mime_type,
            cookies=cookies,
            log_fn=log_fn,
        ) or {"ok": False, "error": "Retry upload mengembalikan hasil kosong."}
        _retry = dict(_retry)
        _retry.setdefault("replaced_versions", [])
        _retry.setdefault("replacement_errors", [])
        _retry["replaced_versions"] = list(pre_replaced_versions) + list(
            _retry.get("replaced_versions", [])
        )
        _retry["replacement_errors"] = list(result["replacement_errors"]) + list(
            _retry.get("replacement_errors", [])
        )
        if not _retry.get("ok"):
            result = _retry
            _restore_backups(
                [backup for backup in _backups if backup["versi"] in pre_replaced_versions]
            )
            return result
        result = _retry
        try:
            _verification = verifier(
                kode_paket, jenis, file_name, result.get("versi")
            ) or {"verified": False}
        except Exception as exc:
            _verification = {"verified": False, "error": str(exc)}
        if not _verification.get("verified"):
            _restore_backups(
                [backup for backup in _backups if backup["versi"] in pre_replaced_versions],
                _delete_new_version=result.get("versi"),
            )
            result["ok"] = False
            result["verified"] = False
            result["versi"] = None
            result["error"] = (
                _verification.get("error")
                or "Retry upload belum terverifikasi; dokumen lama dipulihkan"
            )
            return result
        result["verified"] = True
        result["versi"] = _verification.get("versi", result.get("versi"))
        existing_docs = [
            doc for doc in _old_docs
            if doc.get("versi") not in set(pre_replaced_versions)
        ]

    if not result.get("ok") or jenis not in _PPK_REPLACE_JENIS:
        return result

    if not result.get("verified"):
        try:
            verification = verifier(
                kode_paket, jenis, file_name, result.get("versi")
            ) or {"verified": False}
        except Exception as exc:
            verification = {"verified": False, "error": str(exc)}
        if not verification.get("verified"):
            result["ok"] = False
            result["verified"] = False
            result["versi"] = None
            result["error"] = (
                verification.get("error")
                or "Upload belum terverifikasi pada daftar dokumen SPSE"
            )
            return result
        result["verified"] = True
        result["versi"] = verification.get("versi", result.get("versi"))

    if existing_docs is None:
        try:
            existing_docs = list_dokumen(kode_paket, jenis)
        except Exception as exc:
            result["replacement_errors"].append(
                f"daftar dokumen lama tidak terbaca: {exc}"
            )
            existing_docs = []

    for old_doc in _replacement_candidates(existing_docs, file_name):
        versi = old_doc.get("versi")
        error_added = False
        try:
            deleted = deleter(kode_paket, jenis, versi, cookies=cookies)
        except Exception as exc:
            deleted = False
            error_added = True
            result["replacement_errors"].append(
                f"{old_doc.get('nama_file', versi)}: {exc}"
            )
        if deleted:
            result["replaced_versions"].append(versi)
        elif not error_added:
            result["replacement_errors"].append(
                f"{old_doc.get('nama_file', versi)}: gagal dihapus"
            )
    return result


def hapus_semua_dokumen(kode_paket: str, versi_map: dict = None) -> dict:
    """
    Hapus semua dokumen (kak/kontrak/uraian/lainnya).
    versi_map: {jenis: [versi, ...]} dari session_state (lebih cepat).
    Kalau tidak ada, fallback ke list_dokumen (Playwright goto).
    Return {"dihapus": N, "gagal": N}.
    """
    jenis_list = [k for k in _LIST_ENDPOINTS if k != "nd"]

    # Kumpulkan semua (jenis, versi) yang perlu dihapus
    to_delete = []
    if versi_map:
        for jenis, versis in versi_map.items():
            for v in versis:
                to_delete.append({"jenis": jenis, "versi": v})
    else:
        for jenis in jenis_list:
            docs = list_dokumen(kode_paket, jenis)
            for doc in docs:
                if doc.get("versi") is not None:
                    to_delete.append({"jenis": jenis, "versi": doc["versi"]})

    if not to_delete:
        return {"dihapus": 0, "gagal": 0}

    # Hapus semua dalam 1 CDP call (paralel di browser)
    import json as _json
    del_ep = _DELETE_ENDPOINTS
    lpse   = _LPSE

    js = f"""
    (async () => {{
        const lpse = "{lpse}";
        const kode = "{kode_paket}";
        const toDelete = {_json.dumps(to_delete)};
        const deleteEp = {_json.dumps(del_ep)};
        let dihapus = 0, gagal = 0;
        await Promise.all(toDelete.map(async ({{jenis, versi}}) => {{
            const ep = deleteEp[jenis];
            if (!ep) {{ gagal++; return; }}
            const fd = new FormData();
            fd.append('versi', String(versi));
            const r = await fetch(`/${{lpse}}/dokumennontender/${{kode}}/${{ep}}`,
                                  {{method:'POST', credentials:'include', body:fd}});
            if (r.ok || r.status === 302) dihapus++;
            else gagal++;
        }}));
        return {{dihapus, gagal}};
    }})()
    """
    ok, val, err = _cdp_eval(js, timeout=30)
    if not ok or not val:
        return {"dihapus": 0, "gagal": len(to_delete)}
    return val


PPK_PL_V2_BASE = os.path.join(
    POKJA_ROOT,
    "Paket Experiment - Pengadaan Langsung",
    "V2 - Template PPK PL",
)
PPK_PL_LEGACY_BASE = os.environ.get(
    "POKJA_PPK_PL_LEGACY_BASE",
    os.path.join(
        POKJA_ROOT,
        "@ Pejabat Pengadaan 2026",
        "@ Dinas Perdagangan",
        "1 PERENCANAAN PENGADAAN",
        "Dokumen Upload PPK PL",
    ),
)
PPK_WORKFLOW_REGISTRY = {
    "JKK": {
        "label": "PL Konsultansi (JKK)",
        # JKK dan PK berbagi root paket V2; family dipisahkan oleh metadata.
        "root": os.environ.get("POKJA_PPK_JKK_BASE", PPK_PL_V2_BASE),
        "mapping": {"1.": "kak", "9.": "kak", "12.": "kak", "13.": "kak", "14.": "kak", "15.": "kak", "2.": "uraian", "3.": "kontrak", "4.": "kontrak", "5.": "kontrak", "6.": "kontrak", "8.": "nd", "11.": "lainnya"},
    },
    "PK": {
        "label": "PL Pekerjaan Konstruksi (PK)",
        # Canonical PPK V2 berada satu root dengan folder paket bernomor.
        # Tetap beri override agar PC lain/struktur lama tidak terkunci.
        "root": os.environ.get("POKJA_PPK_PK_BASE", PPK_PL_V2_BASE),
        # PK memakai seluruh dokumen rancangan kontrak SPPBJ/SPMK/R_SPK/SUK 3-6.
        "mapping": {"1.": "kak", "9.": "kak", "12.": "kak", "13.": "kak", "14.": "kak", "15.": "kak", "2.": "uraian", "3.": "kontrak", "4.": "kontrak", "5.": "kontrak", "6.": "kontrak", "8.": "nd", "11.": "lainnya"},
    },
}

PPK_WORKFLOW_CODES = frozenset({"PK", "JKK"})
PPK_WORKFLOW_REGISTRY_PATH = os.path.join(
    PPK_PL_V2_BASE,
    ".ppk_workflow_registry.json",
)


def _ppk_workflow_registry_path(path=None) -> str:
    """Ambil path registry, dengan override lokal per komputer bila ada."""
    if path is not None:
        return os.fspath(path)
    override = str(os.environ.get("POKJA_PPK_WORKFLOW_REGISTRY") or "").strip()
    return override or PPK_WORKFLOW_REGISTRY_PATH


def _normalize_ppk_workflow_registry(registry) -> dict[str, str]:
    """Normalisasi registry; hanya kode paket dan workflow PK/JKK yang valid."""
    if not isinstance(registry, dict):
        return {}
    normalized = {}
    for code, workflow in registry.items():
        package_code = str(code or "").strip()
        workflow_code = str(workflow or "").strip().upper()
        if package_code and workflow_code in PPK_WORKFLOW_CODES:
            normalized[package_code] = workflow_code
    return normalized


def load_ppk_workflow_registry(path=None) -> dict[str, str]:
    """Baca registry workflow PPK; file hilang/rusak dianggap registry kosong."""
    registry_path = _ppk_workflow_registry_path(path)
    try:
        with open(registry_path, "r", encoding="utf-8") as handle:
            raw = json.load(handle)
    except (OSError, UnicodeError, json.JSONDecodeError, TypeError):
        return {}
    return _normalize_ppk_workflow_registry(raw)


def save_ppk_workflow_registry(registry, path=None) -> dict[str, str]:
    """Simpan registry valid secara atomic dan kembalikan isi yang disimpan."""
    normalized = _normalize_ppk_workflow_registry(registry)
    registry_path = os.path.abspath(_ppk_workflow_registry_path(path))
    parent = os.path.dirname(registry_path) or os.curdir
    os.makedirs(parent, exist_ok=True)
    temporary_path = None
    try:
        fd, temporary_path = tempfile.mkstemp(
            prefix=f".{os.path.basename(registry_path)}.",
            suffix=".tmp",
            dir=parent,
        )
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(normalized, handle, ensure_ascii=False, indent=2, sort_keys=True)
            handle.write("\n")
        os.replace(temporary_path, registry_path)
        temporary_path = None
    finally:
        if temporary_path:
            try:
                os.unlink(temporary_path)
            except OSError:
                pass
    return normalized


def ppk_workflow_config(workflow: str | None) -> dict:
    key = str(workflow or "").upper().strip()
    if key not in PPK_WORKFLOW_REGISTRY:
        raise KeyError(f"Workflow PPK tidak dikenal: {workflow}")
    cfg = dict(PPK_WORKFLOW_REGISTRY[key])
    root = os.path.normpath(cfg["root"])
    cfg["transitional"] = False
    if not os.path.isdir(root) and os.path.isdir(PPK_PL_LEGACY_BASE):
        cfg["root"] = PPK_PL_LEGACY_BASE
        cfg["transitional"] = True
    return cfg


def _workflow_from_explicit_metadata(value) -> str | None:
    """Map nilai field jenis pekerjaan eksplisit ke workflow PPK."""
    normalized = str(value or "").strip().upper()
    if not normalized:
        return None
    if normalized in {"JKK", "PLJKK"} or "KONSULTAN" in normalized:
        return "JKK"
    if normalized in {"PK", "PLPK", "KONSTRUKSI"} or "PEKERJAAN KONSTRUKSI" in normalized:
        return "PK"
    return None


def resolve_ppk_workflow(metadata: dict | None, registry=None) -> dict:
    """Resolve workflow dari registry paket, lalu metadata eksplisit.

    Nama paket sengaja tidak digunakan: nomenklatur rekening/uraian SPSE
    dapat menyerupai konstruksi tetapi bukan penentu family PPK.
    """
    row = metadata if isinstance(metadata, dict) else {}
    registry_data = (
        load_ppk_workflow_registry() if registry is None
        else _normalize_ppk_workflow_registry(registry)
    )
    package_code = str(row.get("kode_paket") or "").strip()
    mapped_workflow = registry_data.get(package_code)
    if mapped_workflow:
        return {"status": "resolved", "workflow": mapped_workflow, "source": "registry"}

    for key in (
        "workflow", "ppk_workflow", "jenis_pl", "jenis_pengadaan",
        "jenis_pekerjaan", "kategori_pengadaan", "tipe_pengadaan",
        "package_type", "procurement_type",
    ):
        workflow = _workflow_from_explicit_metadata(row.get(key))
        if workflow:
            return {"status": "resolved", "workflow": workflow, "source": "metadata"}

    return {"status": "ambiguous", "workflow": None, "source": "metadata_missing"}


PPK_MODE_REGISTRY = {
    "PPK - Konsultan": {
        "workflow": "JKK",
        "label": "PPK - Konsultan",
    },
    "PPK - Pekerjaan Konstruksi": {
        "workflow": "PK",
        "label": "PPK - Pekerjaan Konstruksi",
    },
}


def ppk_mode_config(mode: str) -> dict:
    """Ambil konfigurasi family dari submode PPK yang dipilih UI."""
    key = str(mode or "").strip()
    if key not in PPK_MODE_REGISTRY:
        raise KeyError(f"Mode PPK tidak dikenal: {mode}")
    return dict(PPK_MODE_REGISTRY[key])


def filter_paket_ppk_by_workflow(
    rows: list[dict], workflow: str, details_by_code: dict | None = None,
    registry: dict | None = None,
) -> list[dict]:
    """Filter paket PPK berdasarkan registry atau metadata eksplisit.

    Paket yang belum memiliki mapping registry/metadata tetap dikeluarkan agar
    mode Konsultan/PK tidak tercampur.
    """
    expected = str(workflow or "").strip().upper()
    if expected not in PPK_WORKFLOW_REGISTRY:
        return []
    details = details_by_code or {}
    registry_data = (
        load_ppk_workflow_registry() if registry is None
        else _normalize_ppk_workflow_registry(registry)
    )
    result = []
    for row in rows or []:
        code = str(row.get("kode_paket") or "")
        metadata = dict(row)
        extra = details.get(code) or details.get(row.get("kode_paket")) or {}
        if isinstance(extra, dict):
            metadata.update(extra)
        resolved = resolve_ppk_workflow(metadata, registry=registry_data)
        if resolved.get("workflow") == expected:
            result.append(row)
    return result


# Backward-compatible alias untuk modul lama yang masih mengimpor konstanta ini.
PPK_PL_BASE = PPK_PL_LEGACY_BASE


# Struktur output generator PPK PL V2. Resolver di bawah hanya dipakai oleh
# alur upload Streamlit; folder legacy tetap didukung jika subfolder tahap
# belum dibuat.
PPK_DOCUMENT_DRAFT_ROOT = "0. Draft Dokumen PPK"
PPK_DOCUMENT_STAGE_MARKER = "._ppk_stage.txt"
PPK_DOCUMENT_STAGE_DIRS = {
    "UPLOAD AWAL": "01. Upload Awal",
    "BERKONTRAK": "02. Berkontrak",
}
PPK_DOCUMENT_STAGE_ALIASES = {
    "UPLOAD AWAL": "UPLOAD AWAL",
    "BERKONTRAK": "BERKONTRAK",
    # Nilai lama dari workbook sebelum dua tahap disederhanakan.
    "SPPBJ FINAL": "BERKONTRAK",
    "SPK FINAL": "BERKONTRAK",
    "SPMK FINAL": "BERKONTRAK",
}


def normalize_ppk_document_stage(value: object) -> str:
    """Normalisasi tahap workbook ke dua nilai routing yang didukung."""
    normalized = re.sub(r"\s+", " ", str(value or "").strip().upper())
    return PPK_DOCUMENT_STAGE_ALIASES.get(normalized, "UPLOAD AWAL")


def _ppk_package_root(candidate: str) -> str:
    """Naik ke root paket jika user menempel path subfolder draft/tahap."""
    path = os.path.normpath(os.path.expandvars(os.path.expanduser(str(candidate or ""))))
    leaf = os.path.basename(path).casefold()
    draft_leaf = PPK_DOCUMENT_DRAFT_ROOT.casefold()
    stage_leaves = {name.casefold() for name in PPK_DOCUMENT_STAGE_DIRS.values()}

    if leaf in stage_leaves:
        parent = os.path.dirname(path)
        if os.path.basename(parent).casefold() == draft_leaf:
            return os.path.dirname(parent)
    if leaf == draft_leaf:
        return os.path.dirname(path)
    return path


def _find_ppk_master_workbook(package_folder: str) -> str | None:
    """Cari workbook canonical; suffix ``(1)`` dipertahankan untuk legacy."""
    candidates = (
        "0. Master_Data_PL_PPK.xlsm",
        "0. Master_Data_PL_PPK (1).xlsm",
    )
    for name in candidates:
        path = os.path.join(package_folder, name)
        if os.path.isfile(path):
            return path
    return None


def _read_ppk_stage_marker(package_folder: str) -> str | None:
    """Baca tahap hasil generator jika workbook tidak ikut disalin ke paket."""
    marker = os.path.join(package_folder, PPK_DOCUMENT_STAGE_MARKER)
    try:
        with open(marker, "r", encoding="utf-8") as handle:
            raw = re.sub(r"\s+", " ", handle.readline().strip().upper())
        return normalize_ppk_document_stage(raw) if raw in PPK_DOCUMENT_STAGE_ALIASES else None
    except (OSError, UnicodeError):
        return None


def read_ppk_document_stage(package_folder: str) -> tuple[str, str | None, str]:
    """Baca ``Tahap Dokumen`` tanpa mengubah workbook.

    Return ``(stage, workbook_path, source)``. Setiap kegagalan baca sengaja
    menjadi ``UPLOAD AWAL`` agar upload awal tidak salah masuk folder kontrak.
    """
    package_root = _ppk_package_root(package_folder)
    workbook_path = _find_ppk_master_workbook(package_root)
    marker_stage = _read_ppk_stage_marker(package_root)
    if not workbook_path:
        return marker_stage or "UPLOAD AWAL", None, (
            "marker" if marker_stage else "default_missing_workbook"
        )

    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheets = []
        if "Master Data" in workbook.sheetnames:
            sheets.append(workbook["Master Data"])
        sheets.extend(ws for ws in workbook.worksheets if ws not in sheets)
        for worksheet in sheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    label = str(cell.value or "").strip().casefold()
                    if label != "tahap dokumen":
                        continue
                    value = worksheet.cell(row=cell.row, column=cell.column + 1).value
                    return normalize_ppk_document_stage(value), workbook_path, "workbook"
        if marker_stage:
            return marker_stage, workbook_path, "marker_missing_field"
        return "UPLOAD AWAL", workbook_path, "default_missing_field"
    except Exception:
        if marker_stage:
            return marker_stage, workbook_path, "marker_unreadable_workbook"
        return "UPLOAD AWAL", workbook_path, "default_unreadable_workbook"
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def resolve_ppk_upload_folder(package_folder: str, stage: object = None) -> dict:
    """Resolve folder upload PPK dari root paket dan tahap workbook.

    Jika folder tahap tersedia, hanya folder itu yang dipakai. Jika belum ada,
    fallback ke root paket legacy; folder ``0. Draft Dokumen PPK`` sendiri
    tidak pernah dipakai sebagai fallback karena dapat berisi arsip/root.
    """
    package_root = _ppk_package_root(package_folder)
    workbook_stage, workbook_path, stage_source = read_ppk_document_stage(package_root)
    resolved_stage = normalize_ppk_document_stage(stage) if stage is not None else workbook_stage
    stage_folder = os.path.join(
        package_root,
        PPK_DOCUMENT_DRAFT_ROOT,
        PPK_DOCUMENT_STAGE_DIRS[resolved_stage],
    )
    if os.path.isdir(stage_folder):
        return {
            "folder": os.path.normpath(stage_folder),
            "package_folder": os.path.normpath(package_root),
            "stage": resolved_stage,
            "source": "stage",
            "workbook": workbook_path,
            "stage_source": "explicit" if stage is not None else stage_source,
        }
    return {
        "folder": os.path.normpath(package_root),
        "package_folder": os.path.normpath(package_root),
        "stage": resolved_stage,
        "source": "legacy",
        "workbook": workbook_path,
        "stage_source": "explicit" if stage is not None else stage_source,
    }


def folder_number(folder_name: str) -> int | None:
    """Ambil nomor urut dari prefix folder PPK, mis. ``9. Nama`` → 9."""
    m = re.match(r"^\s*(\d+)\s*\.\s*", str(folder_name or ""))
    return int(m.group(1)) if m else None


def next_folder_number(subfolders: list[str] | None = None) -> int:
    """Nomor paket berikutnya berdasarkan folder fisik PPK, bukan jumlah API."""
    folders = subfolders if subfolders is not None else list_subfolder_ppk()
    return max((folder_number(f) or 0 for f in folders), default=0) + 1

FILE_PREFIX_MAP = {
    "1.":  "kak",      # KAK / Spesifikasi Teknis
    "9.":  "kak",      # List Personil (masuk KAK)
    "12.": "kak",      # Spesifikasi Teknis manual (masuk KAK)
    "13.": "kak",      # Gambar (masuk KAK)
    "14.": "kak",      # RK3 (masuk KAK)
    "15.": "kak",      # TKDN (masuk KAK)
    "2.":  "uraian",   # Uraian Singkat
    "3.":  "kontrak",  # SPPBJ (masuk Rancangan Kontrak)
    "4.":  "kontrak",  # SPMK (masuk Rancangan Kontrak)
    "5.":  "kontrak",  # Rancangan Kontrak (SPK)
    "6.":  "kontrak",  # SUK (masuk Rancangan Kontrak)
    "8.":  "nd",       # Nota Dinas PPK
    "11.": "lainnya",  # Diskresi / Informasi Lainnya
    # 7. HPS, 10. Survey — tidak diupload
}


def _windows_filesystem_path(path: str) -> str:
    """Tambahkan prefix extended-length untuk path file Windows yang panjang."""
    path = os.fspath(path)
    if os.name != "nt" or path.startswith(("\\\\?\\", "\\\\.\\")):
        return path
    absolute = os.path.abspath(path)
    if len(absolute) < 240:
        return absolute
    if absolute.startswith("\\\\"):
        return "\\\\?\\UNC\\" + absolute[2:]
    return "\\\\?\\" + absolute


def scan_folder(folder_path: str, pdf_only: bool = False, workflow: str = "JKK") -> list[dict]:
    """
    Scan folder → return list file yang akan diupload.
    [{"path": str, "nama": str, "jenis": str, "mime": str}]
    Skip file yang prefixnya tidak ada di mapping family aktif.
    """
    import os, mimetypes
    hasil = []
    mapping = ppk_workflow_config(workflow)["mapping"]
    scan_path = _windows_filesystem_path(folder_path)
    if not os.path.isdir(scan_path):
        return []
    for fname in sorted(os.listdir(scan_path)):
        fpath = _windows_filesystem_path(os.path.join(scan_path, fname))
        if not os.path.isfile(fpath):
            continue
        extension = os.path.splitext(fname)[1].lower()
        # Bulk upload umumnya PDF, tetapi List Personil/Alat canonical
        # memang boleh tetap DOCX bila PDF-nya belum dibuat. Preferensi PDF
        # ditangani oleh select_bulk_upload_files() agar tidak dobel upload.
        if pdf_only and extension != ".pdf":
            if not (fname.startswith("9. ") and extension in {".doc", ".docx"}):
                continue
        jenis = None
        for prefix, j in mapping.items():
            if fname.startswith(prefix + " ") or fname == prefix.rstrip(".") + ".pdf":
                jenis = j
                break
        if not jenis:
            continue
        mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
        hasil.append({"path": fpath, "nama": fname, "jenis": jenis, "mime": mime})
    return hasil


def select_bulk_upload_files(files: list[dict], pdf_only: bool = True) -> list[dict]:
    """Pilih dokumen bulk bernomor yang siap dikirim ke kategori SPSE.

    Saat ``pdf_only=False``, perilaku legacy dipertahankan: semua file
    bernomor yang terdaftar dikembalikan. Saat mode PDF aktif, PDF terbaru
    diprioritaskan; DOCX hanya fallback untuk List Personil/Alat nomor 9.
    """
    if not pdf_only:
        return [
            item for item in (files or [])
            if item.get("jenis") != "nd"
            and re.match(r"^(?:[1-9]|11|12|13|14|15)\.\s", str(item.get("nama") or ""))
        ]

    candidates = []
    for item in files or []:
        name = str(item.get("nama") or "")
        if item.get("jenis") == "nd":
            continue
        match = re.match(r"^(\d+)\.\s", name)
        if not match or int(match.group(1)) not in set(range(1, 10)) | {11, 12, 13, 14, 15}:
            continue
        extension = os.path.splitext(name)[1].lower()
        if extension == ".pdf":
            priority = 0
        elif match.group(1) == "9" and extension in {".doc", ".docx"}:
            priority = 1
        else:
            continue
        try:
            modified_ns = os.stat(item.get("path", "")).st_mtime_ns
        except (OSError, TypeError, ValueError):
            modified_ns = 0
        candidates.append((int(match.group(1)), priority, -modified_ns, name.casefold(), item))

    selected = {}
    for number, priority, newest_key, name_key, item in sorted(candidates, key=lambda row: row[:4]):
        selected.setdefault(number, (priority, newest_key, name_key, item))
    return [selected[number][3] for number in sorted(selected)]


def list_subfolder_ppk(workflow: str = "JKK") -> list[str]:
    """
    List subfolder di root family PPK yang bukan _* atau .*
    Return sorted list nama folder.
    """
    import os
    root = ppk_workflow_config(workflow)["root"]
    if not os.path.isdir(root):
        return []
    package_folders = []
    for name in os.listdir(root):
        if not os.path.isdir(os.path.join(root, name)):
            continue
        if name.startswith('_') or name.startswith('.'):
            continue
        # Root V2 juga berisi donor template (Konstruksi/Perencanaan/
        # Pengawasan). Tab 2 hanya boleh menampilkan folder paket bernomor.
        if not re.match(r"^\s*\d+\s*\.\s*", name):
            continue
        package_folders.append(name)

    def _package_sort_key(name: str):
        match = re.match(r"^\s*(\d+)\s*\.\s*", name)
        return (int(match.group(1)) if match else 10**9, name.casefold())

    return sorted(package_folders, key=_package_sort_key)


def auto_match_folder(
    nama_paket_spse: str,
    subfolder_list: list[str] | None = None,
    workflow: str | None = None,
) -> str | None:
    """
    Fuzzy match nama paket SPSE ke subfolder PPK PL.
    Jika workflow diberikan, gunakan root folder family tersebut. Ini menjaga
    paket JKK dan PK tidak saling tertukar; argumen list tetap dipertahankan
    untuk kompatibilitas caller lama.
    """
    import re
    from difflib import SequenceMatcher

    if subfolder_list is None:
        # Caller tanpa daftar eksplisit tetap dapat meminta root family.
        subfolder_list = list_subfolder_ppk(workflow) if workflow else []
    subfolder_list = subfolder_list or []

    STRIP_PREFIXES = [
        'Belanja Jasa Konsultansi Perencanaan Arsitektur-Jasa Arsitektur Lainnya ',
        'Belanja Jasa Konsultansi Perencanaan Arsitektur-Jasa Arsitektur ',
        'Belanja Jasa Konsultansi Perencanaan Rekayasa-Jasa Desain Rekayasa untuk Konstruksi ',
        'Belanja Jasa Konsultansi Perencanaan ',
        'Belanja Jasa Konsultansi ',
        'Belanja Jasa ',
    ]

    def _strip(nama):
        for pfx in STRIP_PREFIXES:
            if nama.startswith(pfx):
                return nama[len(pfx):]
        return nama

    def _strip_num(folder):
        return re.sub(r'^\d+\.\s*', '', folder)

    def _folder_num(folder):
        m = re.match(r'^(\d+)\.', folder)
        return int(m.group(1)) if m else 0

    target = _strip(nama_paket_spse).lower().strip()
    best, best_score, best_num = None, 0.0, -1
    for folder in subfolder_list:
        candidate = _strip_num(folder).lower().strip()
        # substring → score 1.0, tapi tetap bandingkan semua (jangan early return)
        if target in candidate or candidate in target:
            score = 1.0
        else:
            score = SequenceMatcher(None, target, candidate).ratio()
        fnum = _folder_num(folder)
        # ambil score tertinggi; tie-break: nomor folder terbesar (versi terbaru)
        if score > best_score or (score == best_score and fnum > best_num):
            best_score = score
            best = folder
            best_num = fnum
    return best if best_score > 0.65 else None


def upload_dari_folder(
    kode_paket: str,
    folder_path: str,
    log_fn=None,
    pdf_only: bool = False,
    workflow: str = "JKK",
) -> dict:
    """
    Upload semua file yang cocok dari folder ke SPSE.
    Return {"results": [{"jenis", "nama", "ok", "error"}], "total_ok": int, "total_err": int}
    """
    def _log(msg):
        if log_fn: log_fn(msg)

    files = scan_folder(folder_path, pdf_only=pdf_only, workflow=workflow)
    if not files:
        return {"results": [], "total_ok": 0, "total_err": 0, "error": "Tidak ada file yang cocok di folder ini."}

    from concurrent.futures import ThreadPoolExecutor, as_completed

    nd_files = [f for f in files if f["jenis"] == "nd"]
    non_nd_files = [f for f in files if f["jenis"] != "nd"]

    results = []
    nd_result = None

    def _upload_one(f):
        # log_fn TIDAK dipanggil dari sini (Streamlit tidak thread-safe)
        # — kumpulkan log di result, flush di main thread
        logs = []
        try:
            with open(f["path"], "rb") as fh:
                file_bytes = fh.read()
            logs.append(f"⬆️ Upload [{f['jenis']}] {f['nama']}...")
            res = upload_dokumen(
                kode_paket=kode_paket,
                jenis=f["jenis"],
                file_bytes=file_bytes,
                file_name=f["nama"],
                mime_type=f["mime"],
                log_fn=None,  # no Streamlit call in thread
            )
            ok = res.get("ok", False)
            logs.append(f"  {'✅' if ok else '❌'} {f['nama']} {'berhasil' if ok else 'gagal: ' + res.get('error','')}")
            return {"jenis": f["jenis"], "nama": f["nama"], "ok": ok, "error": res.get("error", ""), "versi": res.get("versi"), "_logs": logs}
        except Exception as e:
            logs.append(f"  ❌ Exception: {e}")
            return {"jenis": f["jenis"], "nama": f["nama"], "ok": False, "error": str(e), "versi": None, "_logs": logs}

    # Upload non-ND paralel, max 3 worker
    with ThreadPoolExecutor(max_workers=3) as ex:
        futures = {ex.submit(_upload_one, f): f for f in non_nd_files}
        for fut in as_completed(futures):
            r = fut.result()
            for msg in r.pop("_logs", []):
                _log(msg)  # flush log di main thread (aman untuk Streamlit)
            results.append(r)

    # Pilih PP (step 3) sebelum upload ND
    _log("🔗 Pilih PP (step 3)...")
    pp_ok = pilih_pp(kode_paket, log_fn=log_fn)
    if not pp_ok:
        _log("⚠️ Pilih PP gagal — lanjut upload ND tapi email mungkin gagal")

    # Upload ND setelah PP dipilih
    for f in nd_files:
        try:
            with open(f["path"], "rb") as fh:
                file_bytes = fh.read()
            _log(f"⬆️ Upload [nd] {f['nama']}...")
            res = upload_nota_dinas(
                kode_paket=kode_paket,
                file_bytes=file_bytes,
                file_name=f["nama"],
                mime_type=f["mime"],
                log_fn=log_fn,
            )
            ok = res.get("ok", False)
            if ok:
                nd_result = res
                _log(f"  ✅ {f['nama']} berhasil")
            else:
                _log(f"  ❌ {f['nama']} gagal: {res.get('error')}")
            results.append({"jenis": f["jenis"], "nama": f["nama"], "ok": ok, "error": res.get("error", ""), "versi": res.get("versi")})
        except Exception as e:
            _log(f"  ❌ Exception: {e}")
            results.append({"jenis": f["jenis"], "nama": f["nama"], "ok": False, "error": str(e), "versi": None})

    if nd_result:
        _log("📧 Kirim email pemberitahuan ke PP...")
        email_ok = kirim_email_pp(kode_paket, nd_result["path"], nd_result["fileId"], log_fn)
        if email_ok:
            _log("✅ Email ke PP berhasil dikirim")
        else:
            _log("⚠️ Email ke PP gagal — upload ND ok tapi email gagal")

    total_ok = sum(1 for r in results if r["ok"])
    total_err = len(results) - total_ok
    return {"results": results, "total_ok": total_ok, "total_err": total_err}


def upload_dokumen_dari_folder(
    kode_paket: str,
    folder_path: str,
    log_fn=None,
    pdf_only: bool = False,
    workflow: str = "JKK",
) -> dict:
    """Pilih PP lalu upload dokumen PPK 1-9/11-15; Nota Dinas dikecualikan."""
    files = select_bulk_upload_files(
        scan_folder(folder_path, pdf_only=pdf_only, workflow=workflow),
        pdf_only=pdf_only,
    )
    if not files:
        return {
            "results": [],
            "total_ok": 0,
            "total_err": 0,
            "error": "Tidak ada dokumen bernomor 1-9/11-15 (selain Nota Dinas) yang cocok di folder ini.",
        }

    def _log(msg):
        if log_fn:
            log_fn(msg)

    _log("🔗 Memilih Pejabat Pengadaan...")
    pp_ok = pilih_pp(kode_paket, log_fn=log_fn)
    if not pp_ok:
        _log("❌ Pemilihan Pejabat Pengadaan gagal; dokumen belum diupload.")
        return {
            "results": [],
            "total_ok": 0,
            "total_err": len(files),
            "pp_ok": False,
            "error": "Pemilihan Pejabat Pengadaan gagal; dokumen belum diupload.",
        }

    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Snapshot existing per kategori sebelum upload paralel. Snapshot ini
    # sengaja tidak mencakup Nota Dinas karena ND memakai flow email terpisah.
    existing_by_jenis = {}
    for jenis in sorted({item["jenis"] for item in files} & _PPK_REPLACE_JENIS):
        existing_by_jenis[jenis] = list_dokumen(kode_paket, jenis)

    def _upload_one(file_info):
        try:
            with open(file_info["path"], "rb") as fh:
                file_bytes = fh.read()
            result = upload_dokumen_dengan_replace(
                kode_paket=kode_paket,
                jenis=file_info["jenis"],
                file_bytes=file_bytes,
                file_name=file_info["nama"],
                mime_type=file_info["mime"],
                existing_docs=existing_by_jenis.get(file_info["jenis"], []),
            )
            return {
                "jenis": file_info["jenis"],
                "nama": file_info["nama"],
                "ok": result.get("ok", False),
                "verified": result.get("verified", False),
                "error": result.get("error", ""),
                "versi": result.get("versi"),
                "replaced_versions": result.get("replaced_versions", []),
                "replacement_errors": result.get("replacement_errors", []),
            }
        except Exception as exc:
            return {
                "jenis": file_info["jenis"],
                "nama": file_info["nama"],
                "ok": False,
                "verified": False,
                "error": str(exc),
                "versi": None,
                "replaced_versions": [],
                "replacement_errors": [],
            }

    results = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(_upload_one, f): f for f in files}
        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            status = "✅" if result["ok"] else "❌"
            suffix = "" if result["ok"] else f": {result['error']}"
            _log(
                f"{status} {result['nama']} → "
                f"{upload_target_label(result['jenis'])}{suffix}"
            )
            for warning in result.get("replacement_errors", []):
                _log(f"  ⚠️ Replace: {warning}")

    total_ok = sum(1 for result in results if result["ok"] and result.get("verified", False))
    return {
        "results": results,
        "total_ok": total_ok,
        "total_err": len(results) - total_ok,
        "pp_ok": True,
    }


def upload_nota_dinas_dan_email(
    kode_paket: str,
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
    log_fn=None,
) -> dict:
    """Upload Nota Dinas lalu kirim email; pembuatan paket dilakukan tombol terpisah."""
    def _log(msg):
        if log_fn:
            log_fn(msg)

    _log(f"📨 Upload Nota Dinas: {file_name}...")
    upload_result = upload_nota_dinas(
        kode_paket=kode_paket,
        file_bytes=file_bytes,
        file_name=file_name,
        mime_type=mime_type,
        log_fn=log_fn,
    )
    if not upload_result.get("ok"):
        return {
            "ok": False,
            "upload_ok": False,
            "email_ok": False,
            "upload": upload_result,
            "error": upload_result.get("error", "Upload Nota Dinas gagal."),
        }

    _log("📧 Mengirim email pemberitahuan ke Pejabat Pengadaan...")
    email_ok = kirim_email_pp(
        kode_paket,
        upload_result.get("path", ""),
        upload_result.get("fileId", ""),
        log_fn=log_fn,
    )
    if email_ok:
        _log("✅ Nota Dinas terupload dan email berhasil dikirim")
    else:
        _log("⚠️ Nota Dinas terupload, tetapi email gagal dikirim")
    return {
        "ok": bool(email_ok),
        "upload_ok": True,
        "email_ok": bool(email_ok),
        "upload": upload_result,
        "error": "" if email_ok else "Email gagal dikirim; paket belum dibuat.",
    }
