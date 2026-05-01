import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (Advanced Concatenation Logic)...")

groups = []
current_group = None

with pdfplumber.open(pdf_path) as pdf:
    for i, page in enumerate(pdf.pages):
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 10: continue
            
            # Deteksi baris baru (biasanya ada Nomor di kolom 0 atau teks baru di Klasifikasi)
            is_new_row = False
            if row[0] and str(row[0]).strip().endswith("."): is_new_row = True
            if row[1] and any(k in str(row[1]) for k in ["Arsitektur", "Rekayasa", "Sipil", "Mekanikal", "Spesialis"]): 
                if "Permen" not in str(row[1]) and "Undang" not in str(row[1]):
                    is_new_row = True
            
            if is_new_row:
                if current_group: groups.append(current_group)
                current_group = [str(c) if c else "" for c in row]
            else:
                if current_group:
                    # Sambungkan teks ke group yang sedang berjalan
                    for idx in range(len(row)):
                        if row[idx]:
                            current_group[idx] += "\n" + str(row[idx])
    
    if current_group: groups.append(current_group)

# Sekarang olah grup-grup tersebut menjadi baris data
final_data = []
last_klas = ""

for g in groups:
    # Klasifikasi
    if g[1].strip():
        k_clean = g[1].replace("\n", " ").strip()
        if not any(x in k_clean for x in ["Permen", "Undang", "1999"]):
            last_klas = k_clean
            
    # Kode SBU Baru (Kolom 7)
    kode_text = g[7]
    kodes = re.findall(r'[A-Z]{2}\d{3}', kode_text)
    
    # Subklasifikasi Baru (Kolom 6)
    sub_text = g[6]
    # Pecah berdasarkan pola "1. ", "2. ", dst atau newline jika tidak ada angka
    if "1." in sub_text:
        subs = re.split(r'\n\d+\.\s*|\d+\.\s*', sub_text)
        subs = [s.replace("\n", " ").strip() for s in subs if s.strip()]
    else:
        subs = [s.strip() for s in sub_text.split("\n") if s.strip()]
        
    # KBLI 2020 (Kolom 8)
    kbli_text = g[8]
    kblis = re.findall(r'\d{5}', kbli_text)
    
    # Lingkup (Kolom 5)
    lingkup = g[5].replace("\n", " ").strip()
    
    # Mapping
    max_len = max(len(kodes), len(subs), len(kblis))
    for i in range(max_len):
        kode = kodes[i] if i < len(kodes) else (kodes[0] if kodes else "")
        name = subs[i] if i < len(subs) else (subs[0] if subs else "")
        kbli = kblis[i] if i < len(kblis) else (kblis[0] if kblis else "")
        
        # Bersihkan 'name' dari sisa potongan kalimat di PDF
        name = name.strip(",").strip(";").strip()
        
        if kode:
            final_data.append({
                "Klasifikasi": last_klas,
                "Kode SBU": kode,
                "Subklasifikasi": name,
                "KBLI 2020": kbli,
                "Lingkup Pekerjaan": lingkup
            })

df = pd.DataFrame(final_data).drop_duplicates()
df = df.sort_values(["Klasifikasi", "Kode SBU"])

print(f"Selesai! Berhasil mengolah {len(df)} data SBU utuh.")

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

headers = ["Klasifikasi", "Kode SBU", "Subklasifikasi", "KBLI 2020", "Lingkup Pekerjaan"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

for r, row in enumerate(df.values, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
        ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 80

wb.save(excel_path)
print("File template berhasil diperbarui dengan data utuh.")
