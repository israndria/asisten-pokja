import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (V4 - Mapping Mode)...")

groups = []
current_group = None

with pdfplumber.open(pdf_path) as pdf:
    for i, page in enumerate(pdf.pages):
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 10: continue
            
            # Deteksi baris baru berdasarkan Nomor di kolom 0 atau Klasifikasi di kolom 1
            is_new_row = False
            if row[0] and str(row[0]).strip().endswith("."): is_new_row = True
            if row[1] and any(k in str(row[1]) for k in ["Arsitektur", "Rekayasa", "Sipil", "Mekanikal", "Spesialis", "Keterampilan", "Lainnya", "Terintegrasi"]): 
                if "Permen" not in str(row[1]) and "Undang" not in str(row[1]):
                    is_new_row = True
            
            if is_new_row:
                if current_group: groups.append(current_group)
                current_group = [str(c) if c else "" for c in row]
            else:
                if current_group:
                    for idx in range(len(row)):
                        if row[idx]:
                            current_group[idx] += "\n" + str(row[idx])
    
    if current_group: groups.append(current_group)

final_rows = []
last_klas = ""

for g in groups:
    # 1. Klasifikasi
    if g[1].strip():
        k_clean = g[1].replace("\n", " ").strip()
        if not any(x in k_clean for x in ["Permen", "Undang", "1999"]):
            last_klas = k_clean
            
    # 2. Data LAMA (Kolom 2, 3, 4)
    kbli_old = g[2].strip()
    kode_old = g[3].strip()
    nama_old = g[4].replace("\n", " ").strip()
    
    # 3. Data BARU (Kolom 6, 7, 8)
    # Kita pecah berdasarkan kode SBU baru yang ditemukan
    kode_new_text = g[7]
    kodes_new = re.findall(r'[A-Z]{2}\d{3}', kode_new_text)
    
    nama_new_text = g[6]
    if "1." in nama_new_text:
        names_new = re.split(r'\n\d+\.\s*|\d+\.\s*', nama_new_text)
        names_new = [n.replace("\n", " ").strip() for n in names_new if n.strip()]
    else:
        names_new = [n.strip() for n in nama_new_text.split("\n") if n.strip()]
        
    kbli_new_text = g[8]
    kblis_new = re.findall(r'\d{5}', kbli_new_text)
    
    # Sinkronisasi baris mapping
    max_count = max(len(kodes_new), len(names_new), len(kblis_new), 1)
    
    for i in range(max_count):
        k_new = kodes_new[i] if i < len(kodes_new) else (kodes_new[0] if kodes_new else "")
        n_new = names_new[i] if i < len(names_new) else (names_new[0] if names_new else "")
        kb_new = kblis_new[i] if i < len(kblis_new) else (kblis_new[0] if kblis_new else "")
        
        # Bersihkan '1.' di awal nama
        n_new = re.sub(r'^\d+\.\s*', '', n_new).strip(";").strip(",").strip()

        final_rows.append({
            "Klasifikasi": last_klas,
            "KBLI 2017 (Lama)": kbli_old,
            "Kode SBU (Lama)": kode_old,
            "Nama SBU (Lama)": nama_old,
            "Nama SBU (Baru)": n_new,
            "Kode SBU (Baru)": k_new,
            "KBLI 2020 (Baru)": kb_new
        })

df = pd.DataFrame(final_rows).drop_duplicates()

# Sortir berdasarkan Klasifikasi dan Kode Baru
df = df.sort_values(["Klasifikasi", "Kode SBU (Baru)"])

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

headers = list(df.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

for r, row in enumerate(df.values, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(vertical="top")

# Lebar kolom otomatis secukupnya
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

wb.save(excel_path)
print(f"Mapping SBU Berhasil! {len(df)} baris tersimpan di {excel_path}")
