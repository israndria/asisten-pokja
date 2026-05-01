import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (V8 - Assembler Logic)...")

def clean_and_join(lines):
    if not lines: return ""
    # Gabungkan semua baris fisik jadi satu string utuh
    text = " ".join([str(l).strip() for l in lines if l]).strip()
    # Bersihkan sampah header/footer
    trash = ["Subklasifikasi", "Klasifikasi", "KBLI 2015", "UU 18/1999", "Undang – Undang", "PP No.5", "Keterangan", "Halaman"]
    for t in trash:
        text = re.sub(t + r".*?(\s|$)", " ", text, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', text).strip()

def split_by_pattern(text, ptype):
    if not text: return [""]
    if ptype == "code":
        # Split berdasarkan Kode SBU (misal: BG001, AR002)
        matches = re.findall(r'[A-Z]{2}\d{3}', text)
        return matches if matches else [text]
    elif ptype == "kbli":
        # Split berdasarkan KBLI 5 digit
        matches = re.findall(r'\d{5}', text)
        return matches if matches else [text]
    elif ptype == "name_new":
        # Split berdasarkan nomor urut "1. ", "2. " atau titik koma ";"
        items = re.split(r'\d+\.\s+|;', text)
        return [i.strip() for i in items if i.strip()]
    return [text]

final_rows = []
last_klas = "Lainnya"

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 9: continue
            
            # 1. Update Klasifikasi
            raw_klas = str(row[1]) if row[1] else ""
            for k in ["Arsitektur", "Rekayasa", "Sipil", "Mekanikal", "Elektrikal", "Spesialis", "Keterampilan", "Lainnya", "Terintegrasi"]:
                if k.lower() in raw_klas.lower():
                    last_klas = k
                    break
            
            # 2. Ambil & Rakit Teks Seluruh Kolom
            # Kolom: 2:KBLI Old, 3:SBU Old, 4:Nama Old, 6:Nama New, 7:SBU New, 8:KBLI New
            kb_old_raw = clean_and_join(str(row[2]).split("\n"))
            kd_old_raw = clean_and_join(str(row[3]).split("\n"))
            nm_old_raw = clean_and_join(str(row[4]).split("\n"))
            nm_new_raw = clean_and_join(str(row[6]).split("\n"))
            kd_new_raw = clean_and_join(str(row[7]).split("\n"))
            kb_new_raw = clean_and_join(str(row[8]).split("\n"))

            # 3. Split Berdasarkan Pola
            kb_old_list = split_by_pattern(kb_old_raw, "kbli")
            kd_old_list = split_by_pattern(kd_old_raw, "code")
            nm_old_list = [nm_old_raw] # Nama lama biasanya satu kesatuan
            
            nm_new_list = split_by_pattern(nm_new_raw, "name_new")
            kd_new_list = split_by_pattern(kd_new_raw, "code")
            kb_new_list = split_by_pattern(kb_new_raw, "kbli")

            # 4. Tentukan Jumlah Baris Terbanyak (Biasanya di sisi Baru)
            max_lines = max(len(nm_new_list), len(kd_new_list), len(kb_new_list))
            if max_lines == 0: continue

            # 5. Zipping & Mapping
            for i in range(max_lines):
                # Ambil data ke-i dengan fallback ke elemen terakhir (Forward Fill)
                # Untuk sisi LAMA
                cur_kb_old = kb_old_list[i] if i < len(kb_old_list) else kb_old_list[-1]
                cur_kd_old = kd_old_list[i] if i < len(kd_old_list) else kd_old_list[-1]
                cur_nm_old = nm_old_list[0] # Selalu pakai nama lama yang sama
                
                # Untuk sisi BARU
                cur_nm_new = nm_new_list[i] if i < len(nm_new_list) else nm_new_list[-1]
                cur_kd_new = kd_new_list[i] if i < len(kd_new_list) else kd_new_list[-1]
                cur_kb_new = kb_new_list[i] if i < len(kb_new_list) else kb_new_list[-1]

                # Filter validitas: minimal ada kode baru
                if re.search(r'[A-Z]{2}\d{3}', cur_kd_new) and len(cur_nm_new) > 3:
                    final_rows.append({
                        "Klasifikasi": last_klas,
                        "KBLI 2017 (Lama)": cur_kb_old,
                        "Kode SBU (Lama)": cur_kd_old,
                        "Nama SBU (Lama)": cur_nm_old,
                        "Nama SBU (Baru)": cur_nm_new,
                        "Kode SBU (Baru)": cur_kd_new,
                        "KBLI 2020 (Baru)": cur_kb_new
                    })

df = pd.DataFrame(final_rows).drop_duplicates()

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Style Header
headers = list(df.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Style Data
for r, row in enumerate(df.values, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Column Widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

wb.save(excel_path)
print(f"Mapping SBU Selesai! Berhasil merakit {len(df)} pemetaan rapi.")
