import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (V5 - Ultra Clean Mapping)...")

def clean_text(text):
    if not text: return ""
    # Hapus sampah yang sering muncul berulang di sel
    text = text.replace("Klasifikasi", "").replace("Subklasifikasi", "").replace("KBLI 2015", "").replace("UU 18/1999", "")
    text = text.replace("Kode", "").replace("KBLI", "")
    # Gabungkan baris baru jadi spasi, hapus spasi ganda
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_codes(text, pattern):
    if not text: return []
    return re.findall(pattern, text)

groups = []
current_group = None

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 10: continue
            
            # Deteksi awal baris baru (Ada Nomor atau Kode Klasifikasi Utama)
            is_new = False
            if row[0] and str(row[0]).strip().endswith("."): is_new = True
            
            # Jika sel pertama (No) ada isinya, berarti record baru
            if is_new:
                if current_group: groups.append(current_group)
                current_group = [str(c) if c else "" for c in row]
            else:
                if current_group:
                    for idx in range(len(row)):
                        if row[idx]:
                            current_group[idx] += " " + str(row[idx])
    
    if current_group: groups.append(current_group)

final_rows = []
last_klas = ""

for g in groups:
    # 1. Identifikasi Klasifikasi
    raw_klas = g[1].strip()
    for k in ["Arsitektur", "Sipil", "Mekanikal", "Elektrikal", "Spesialis", "Keterampilan", "Lainnya", "Terintegrasi"]:
        if k in raw_klas:
            last_klas = k
            break
            
    # 2. Data LAMA
    # Ambil semua KBLI 2017 (5 digit)
    kblis_old = " ".join(re.findall(r'\d{5}', g[2]))
    # Ambil Kode SBU Lama (e.g. BG004, SI001)
    kodes_old = " ".join(re.findall(r'[A-Z]{2}\d{3}', g[3]))
    # Nama SBU Lama
    nama_old = clean_text(g[4])
    
    # 3. Data BARU (Pecah per SBU Baru jika ada banyak)
    # Nama SBU Baru biasanya dipisah angka 1., 2. dst atau baris
    nama_new_raw = g[6]
    # Pisahkan berdasarkan pola angka "1. ", "2. "
    names_new = re.split(r'\d+\.\s*', nama_new_raw)
    names_new = [clean_text(n) for n in names_new if n.strip()]
    if not names_new: names_new = [clean_text(nama_new_raw)]
    
    # Kode SBU Baru
    kodes_new = extract_codes(g[7], r'[A-Z]{2}\d{3}')
    # KBLI 2020 Baru
    kblis_new = extract_codes(g[8], r'\d{5}')
    
    # Sinkronisasi baris mapping
    # Kita buat baris sebanyak jumlah SBU Baru yang ditemukan
    max_count = max(len(kodes_new), len(names_new), len(kblis_new), 1)
    
    for i in range(max_count):
        k_new = kodes_new[i] if i < len(kodes_new) else (kodes_new[0] if kodes_new else "")
        n_new = names_new[i] if i < len(names_new) else (names_new[0] if names_new else "")
        kb_new = kblis_new[i] if i < len(kblis_new) else (kblis_new[0] if kblis_new else "")
        
        # Hanya masukkan jika ada data minimal (Kode Baru atau Nama Baru)
        if k_new or n_new:
            final_rows.append({
                "Klasifikasi": last_klas,
                "KBLI 2017 (Lama)": kblis_old,
                "Kode SBU (Lama)": kodes_old,
                "Nama SBU (Lama)": nama_old,
                "Nama SBU (Baru)": n_new,
                "Kode SBU (Baru)": k_new,
                "KBLI 2020 (Baru)": kb_new
            })

# Buat DataFrame dan hilangkan duplikasi akibat looping
df = pd.DataFrame(final_rows).drop_duplicates()

# Sortir agar rapi
df = df.sort_values(["Klasifikasi", "Kode SBU (Baru)"])

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Header styling
headers = list(df.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Data styling
for r, row in enumerate(df.values, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Lebar kolom
widths = [15, 15, 15, 45, 45, 15, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb.save(excel_path)
print(f"Mapping SBU Berhasil! {len(df)} baris tersimpan dengan kondisi bersih.")
