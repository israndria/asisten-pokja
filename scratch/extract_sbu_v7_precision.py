import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (V7 - Precision Splitter)...")

def clean_val(val):
    if not val: return ""
    # Hapus nomor urut 1. 2. dst di awal kalimat
    val = re.sub(r'^\d+\.\s*', '', str(val).strip())
    # Hapus sampah spesifik
    trash = ["Subklasifikasi", "Klasifikasi", "KBLI 2015", "UU 18/1999", "Undang – Undang", "PP No.5", "Keterangan"]
    for t in trash:
        val = val.replace(t, "")
    return val.strip()

final_rows = []
last_klas = ""

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 9: continue
            
            # Deteksi Klasifikasi Utama (Biasanya di baris tersendiri atau kolom 1)
            raw_klas = str(row[1]) if row[1] else ""
            for k in ["Arsitektur", "Rekayasa", "Sipil", "Mekanikal", "Elektrikal", "Spesialis", "Keterampilan", "Lainnya", "Terintegrasi"]:
                if k.lower() in raw_klas.lower():
                    last_klas = k
                    break
            
            # Ambil data per baris di dalam sel (Split by \n)
            # Kolom: 2:KBLI Old, 3:SBU Old, 4:Nama Old, 6:Nama New, 7:SBU New, 8:KBLI New
            cols_to_parse = [2, 3, 4, 6, 7, 8]
            cell_lines = []
            for idx in cols_to_parse:
                lines = str(row[idx]).split("\n") if row[idx] else [""]
                lines = [clean_val(l) for l in lines if l.strip()]
                if not lines: lines = [""]
                cell_lines.append(lines)
            
            # Tentukan berapa baris mapping yang ada di sel ini
            max_lines = max(len(l) for l in cell_lines)
            
            # Jika baris ini kosong melompong, lewati
            if max_lines == 1 and not any(cell_lines): continue
            
            # Jodohkan data (Zipping with Forward Fill)
            for i in range(max_lines):
                # Ambil item ke-i, jika tidak ada ambil yang terakhir (Forward Fill)
                sub_data = []
                for lines in cell_lines:
                    val = lines[i] if i < len(lines) else lines[-1]
                    sub_data.append(val)
                
                # Filter: Harus ada Kode SBU (Lama atau Baru) untuk dianggap baris valid
                kb_old, kd_old, nm_old, nm_new, kd_new, kb_new = sub_data
                
                if re.search(r'[A-Z]{2}\d{3}', kd_old) or re.search(r'[A-Z]{2}\d{3}', kd_new):
                    final_rows.append({
                        "Klasifikasi": last_klas,
                        "KBLI 2017 (Lama)": kb_old,
                        "Kode SBU (Lama)": kd_old,
                        "Nama SBU (Lama)": nm_old,
                        "Nama SBU (Baru)": nm_new,
                        "Kode SBU (Baru)": kd_new,
                        "KBLI 2020 (Baru)": kb_new
                    })

df = pd.DataFrame(final_rows).drop_duplicates()

# Sortir
df = df.sort_values(["Klasifikasi", "Kode SBU (Baru)"])

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Header
headers = list(df.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Data
for r, row in enumerate(df.values, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(vertical="top")

# Column Widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

wb.save(excel_path)
print(f"Mapping SBU Selesai! Berhasil mengekstrak {len(df)} pemetaan presisi.")
