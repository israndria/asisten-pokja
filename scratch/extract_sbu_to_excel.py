import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

all_data = []

print("Memulai ekstraksi PDF...")
with pdfplumber.open(pdf_path) as pdf:
    for i, page in enumerate(pdf.pages):
        tables = page.extract_tables()
        if not tables:
            continue
            
        for table in tables:
            for row in table:
                # Bersihkan row dari None dan newline
                clean_row = [str(cell).replace("\n", " ").strip() if cell else "" for cell in row]
                
                # Filter: Kita cari baris yang punya Kode SBU Baru (biasanya di kolom index 7 atau 8)
                # Berdasarkan preview, kolom 7 adalah Subklasifikasi Baru, 8 adalah Kode Baru, 9 adalah KBLI 2020
                # Tapi urutannya bisa geser jika ada merger cell. Kita ambil semua dulu.
                if len(clean_row) >= 10:
                    all_data.append(clean_row)
        
        if (i+1) % 10 == 0:
            print(f"  Sudah memproses {i+1} halaman...")

# Convert ke DataFrame
df_raw = pd.DataFrame(all_data)

# Pembersihan data (Sangat Penting karena PDF sering berantakan)
# Kita cari baris yang mengandung kode SBU (2 huruf + 3 angka, misal AR001)
def find_sbu_code(row):
    import re
    text = " ".join(row)
    match = re.search(r'[A-Z]{2}\d{3}', text)
    return match.group(0) if match else None

# Simpan hasil akhir yang relevan saja
final_rows = []
last_klasifikasi = ""

for index, row in df_raw.iterrows():
    # Kolom 1 biasanya Klasifikasi
    klas = row[1] if row[1] else last_klasifikasi
    if klas and "Permen" not in klas and "Klasifikasi" not in klas:
        last_klasifikasi = klas
        
    # Kode Baru biasanya di kolom 7 atau 8
    kode_baru = ""
    import re
    # Cari di kolom 7 atau 8
    for col_idx in [7, 8]:
        m = re.search(r'[A-Z]{2}\d{3}', str(row[col_idx]))
        if m:
            kode_baru = m.group(0)
            break
            
    if kode_baru:
        # Nama Subklas ada di kolom 6 (biasanya)
        nama_subklas = row[6]
        # KBLI 2020 ada di kolom 8 atau 9
        kbli = row[8] if kode_baru == row[7] else row[9]
        # Lingkup ada di kolom 5
        lingkup = row[5]
        
        final_rows.append({
            "Klasifikasi": last_klasifikasi,
            "Kode SBU (Baru)": kode_baru,
            "Subklasifikasi": nama_subklas,
            "KBLI 2020": kbli,
            "Lingkup Pekerjaan": lingkup
        })

df_final = pd.DataFrame(final_rows).drop_duplicates()

print(f"Ekstraksi selesai. Menemukan {len(df_final)} SBU.")

# Tulis ke Excel
print(f"Menulis ke {excel_path}...")
wb = openpyxl.load_workbook(excel_path, keep_vba=True)

if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Header
headers = list(df_final.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data
for r, row in enumerate(df_final.values, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)

# Styling
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 60

# Freeze panes
ws.freeze_panes = "A2"

wb.save(excel_path)
print("SELESAI! Sheet DATABASE SBU 2022 telah dibuat.")
