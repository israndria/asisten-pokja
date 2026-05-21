"""
Copy sheet dari Paket 1 ke semua target menggunakan openpyxl.
Sheet yang dicopy: @ Evaluasi, satu_data, 5. HPS, 6. Penawaran,
  database_reviu, database_dokpil, 7.2 Dengan Nego, list_reviu, list_dokpil

Catatan: openpyxl copy sheet antar workbook = copy cell values + formats.
Formula referensi cross-sheet di-copy apa adanya (string formula).
VBA macro tidak ikut (sudah di-inject terpisah via inject_pl.py).
"""
import sys, io, shutil
import openpyxl
from openpyxl import load_workbook
from copy import copy
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

POKJA_ROOT = Path("D:/Dokumen/@ POKJA 2026")
PL_ROOT    = POKJA_ROOT / "@ Pejabat Pengadaan 2026/@ Pengadaan Langsung JKK"
DEV        = POKJA_ROOT / "Paket Experiment - Pengadaan Langsung - Konsultan Konstuksi/Development/0. BAPLJKK - Template.xlsm"
SOURCE     = PL_ROOT / "1. PLJKK - Konsultan Perencanaan Paket 1/0. BAPLJKK - Konsultan Perencanaan Paket 1.xlsm"

SHEETS_TO_COPY = ["@ Evaluasi","satu_data","5. HPS","6. Penawaran",
                  "database_reviu","database_dokpil","7.2 Dengan Nego",
                  "list_reviu","list_dokpil"]

EVAL_FORMULAS = {
    (8,  3): '=RIGHT(C4,4)',
    (14, 3): '=RIGHT(C10,4)',
    (20, 3): '=RIGHT(C16,4)',
    (30, 3): '=LEFT(C29,2)',
    (31, 3): '=terbilang(C30)',
    (32, 3): '=TRIM(MID(C29,4,FIND(" ",C29,4)-4))',
}

TARGETS = [DEV] + sorted(PL_ROOT.glob("*/0. BAPLJKK*.xlsm"))
TARGETS = [t for t in TARGETS if "Paket 1\\" not in str(t)]


def copy_sheet(src_ws, dst_wb, sheet_name):
    """Copy worksheet src_ws ke dst_wb dengan nama sheet_name."""
    # Hapus sheet lama jika ada
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]

    # Buat sheet baru di dst
    dst_ws = dst_wb.create_sheet(title=sheet_name)

    # Copy dimensi kolom
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden

    # Copy dimensi baris
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
        dst_ws.row_dimensions[row_num].hidden = row_dim.hidden

    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Copy tab color
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = copy(src_ws.sheet_properties.tabColor)

    # Copy visibility
    dst_ws.sheet_state = src_ws.sheet_state

    # Copy cells — skip MergedCell (read-only proxy, nilai dari master cell)
    from openpyxl.cell.cell import MergedCell
    for row in src_ws.iter_rows():
        for src_cell in row:
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    return dst_ws


def main():
    print(f"Source: {SOURCE.name}")
    src_wb = load_workbook(str(SOURCE), keep_vba=True)
    src_sheets = {name: src_wb[name] for name in SHEETS_TO_COPY if name in src_wb.sheetnames}
    print(f"Sheet available di source: {list(src_sheets.keys())}")

    for target in TARGETS:
        if not target.exists():
            print(f"SKIP: {target.name}")
            continue
        print(f"\n{target.name}")

        dst_wb = load_workbook(str(target), keep_vba=True)
        ok = err = 0

        for sh_name, src_ws in src_sheets.items():
            try:
                copy_sheet(src_ws, dst_wb, sh_name)
                print(f"  copy OK: {sh_name}")
                ok += 1
            except Exception as e:
                print(f"  copy ERR {sh_name}: {e}")
                err += 1

        # Fix formula @ Evaluasi
        if "@ Evaluasi" in dst_wb.sheetnames:
            ws_eval = dst_wb["@ Evaluasi"]
            for (row, col), formula in EVAL_FORMULAS.items():
                ws_eval.cell(row=row, column=col).value = formula
            print(f"  formula fixed")

        # Verify
        has_eval = "@ Evaluasi" in dst_wb.sheetnames
        print(f"  sheets: {dst_wb.sheetnames}")
        print(f"  has @ Evaluasi: {has_eval}")

        dst_wb.save(str(target))
        print(f"  SAVED ({ok} OK, {err} err)")

    print("\nSelesai.")


if __name__ == "__main__":
    main()
