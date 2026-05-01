import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

print("Memulai ekstraksi PDF (V6 - Surgeon Mode)...")

def ultra_clean(text):
    if not text: return ""
    # Daftar sampah yang harus dibuang
    trash = [
        r"Undang\s*–\s*Undang.*2021", 
        r"PP\s*No\.5.*2021", 
        r"Subklasifik\w*", 
        r"\basi\b", 
        r"Klasifikasi",
        r"KBLI 2015",
        r"Halaman \d+",
        r"LAMPIRAN",
        r"Permen PUPR",
        r"Salinan",
        r"-\d+-"
    ]
    for pattern in trash:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)
    
    # Bersihkan spasi ganda dan karakter aneh di awal/akhir
    text = re.sub(r'\s+', ' ', text).strip()
    text = text.strip(";").strip(",").strip(".")
    return text

def is_valid_code(text, pattern):
    if not text: return False
    return bool(re.search(pattern, str(text)))

final_rows = []
last_klas = "Lainnya"
last_kbli_old = ""
last_kbli_new = ""

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        table = page.extract_table()
        if not table: continue
        
        for row in table:
            if len(row) < 10: continue
            
            # Kolom Mapping:
            # 1: Klasifikasi, 2: KBLI 2017, 3: Kode SBU Lama, 4: Nama SBU Lama
            # 6: Nama SBU Baru, 7: Kode SBU Baru, 8: KBLI 2020
            
            raw_klas = str(row[1]) if row[1] else ""
            kb_old = ultra_clean(row[2])
            kd_old = ultra_clean(row[3])
            nm_old = ultra_clean(row[4])
            nm_new = ultra_clean(row[6])
            kd_new = ultra_clean(row[7])
            kb_new = ultra_clean(row[8])

            # 1. Update Klasifikasi jika ada
            for k in ["Arsitektur", "Rekayasa", "Sipil", "Mekanikal", "Elektrikal", "Spesialis", "Keterampilan", "Lainnya", "Terintegrasi"]:
                if k.lower() in raw_klas.lower():
                    last_klas = k
                    break

            # 2. Forward Fill KBLI
            if kb_old: last_kbli_old = kb_old
            if kb_new: last_kbli_new = kb_new

            # 3. Validasi: Minimal ada Kode SBU Baru atau Nama SBU Baru yang valid
            # Serta hindari baris yang isinya cuma "Sipil" dsb (Header sampah)
            if not kd_new and not nm_new: continue
            if "Undang" in nm_new or "asi" == nm_new.lower(): continue
            if len(nm_new) < 3 and not kd_new: continue

            # 4. Simpan Data
            # Kita hanya ambil baris yang benar-benar punya Kode SBU (Lama atau Baru)
            if is_valid_code(kd_old, r'[A-Z]{2}\d{3}') or is_valid_code(kd_new, r'[A-Z]{2}\d{3}'):
                final_rows.append({
                    "Klasifikasi": last_klas,
                    "KBLI 2017 (Lama)": last_kbli_old,
                    "Kode SBU (Lama)": kd_old,
                    "Nama SBU (Lama)": nm_old,
                    "Nama SBU (Baru)": nm_new,
                    "Kode SBU (Baru)": kd_new,
                    "KBLI 2020 (Baru)": last_kbli_new
                })

df = pd.DataFrame(final_rows)

# Pembersihan tahap akhir: Gabungkan baris-baris yang merupakan kelanjutan
# (Sering terjadi Nama SBU terpotong jadi 2 baris di PDF)
cleaned_data = []
prev_row = None

for _, row in df.iterrows():
    # Jika baris sekarang tidak punya Kode Baru tapi punya Nama Baru, 
    # kemungkinan ini adalah sambungan Nama dari baris sebelumnya
    if prev_row is not None and not row["Kode SBU (Baru)"] and row["Nama SBU (Baru)"]:
        # Cek apakah baris sebelumnya punya Kode tapi Nama-nya mungkin belum lengkap
        if prev_row["Kode SBU (Baru)"]:
            prev_row["Nama SBU (Baru)"] += " " + row["Nama SBU (Baru)"]
            prev_row["Nama SBU (Lama)"] += " " + row["Nama SBU (Lama)"]
            continue
    
    if prev_row is not None:
        cleaned_data.append(prev_row)
    prev_row = row.to_dict()

if prev_row: cleaned_data.append(prev_row)

df_final = pd.DataFrame(cleaned_data)

# Hapus baris yang masih mengandung sampah spesifik
df_final = df_final[~df_final['Nama SBU (Baru)'].str.contains("Undang", na=False)]
df_final = df_final[df_final['Nama SBU (Baru)'].str.len() > 5]

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Styling
headers = list(df_final.columns)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

for r, row in enumerate(df_final.values, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Auto-width manual
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

wb.save(excel_path)
print(f"Mapping SBU Selesai! {len(df_final)} baris mapping bersih tersimpan.")
