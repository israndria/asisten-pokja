import pdfplumber
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment

pdf_path = r"D:\Download\PermenPUPR 8 2022 SBU.pdf"
excel_path = r"d:\Dokumen\@ POKJA 2026\Paket Experiment\0. BAPK - Template.xlsm"
sheet_name = "DATABASE SBU 2022"

all_rows = []

print("Memulai ekstraksi PDF (Improved Logic)...")
with pdfplumber.open(pdf_path) as pdf:
    last_klasifikasi = ""
    
    for i, page in enumerate(pdf.pages):
        tables = page.extract_tables()
        if not tables:
            continue
            
        for table in tables:
            for row in table:
                if len(row) < 10: continue
                
                # Clean each cell
                clean_row = [str(c).strip() if c else "" for c in row]
                
                # 1. Update Klasifikasi (Kolom 1)
                # Hanya ambil jika mengandung kata kunci klasifikasi asli
                klas_raw = clean_row[1].replace("\n", " ")
                if any(k in klas_raw for k in ["Arsitektur", "Rekayasa", "Penataan Ruang", "Sipil", "Mekanikal", "Spesialis", "Lainnya", "Terintegrasi"]):
                    if "Permen" not in klas_raw and "Undang" not in klas_raw and "1999" not in klas_raw:
                        last_klasifikasi = klas_raw
                
                # 2. Cari Kode SBU Baru (Kolom 7)
                # Kolom 7 sering berisi multi-line codes (AR001\nAR002)
                kode_cell = clean_row[7]
                kodes = re.findall(r'[A-Z]{2}\d{3}', kode_cell)
                
                if kodes:
                    # Ambil data pendukung
                    subklas_new_cell = clean_row[6] # Nama Subklas Baru
                    kbli_cell = clean_row[8]        # KBLI 2020
                    lingkup = clean_row[5].replace("\n", " ") # Lingkup Pekerjaan
                    
                    # Pecah jika ada multi-line
                    subklas_list = subklas_new_cell.split("\n")
                    kbli_list = kbli_cell.split("\n")
                    
                    for idx, k in enumerate(kodes):
                        # Coba pasangkan nama dan kbli sesuai index jika memungkinkan
                        s_name = subklas_list[idx] if idx < len(subklas_list) else subklas_list[0]
                        k_val = kbli_list[idx] if idx < len(kbli_list) else kbli_list[0]
                        
                        # Bersihkan s_name dari angka urut "1.", "2."
                        s_name = re.sub(r'^\d+\.\s*', '', s_name.replace("\n", " ")).strip()
                        k_val = k_val.replace("\n", " ").strip()
                        
                        # Jika k_val isinya malah teks panjang (Penyetaraan...), kosongkan
                        if len(k_val) > 10: k_val = ""

                        all_rows.append({
                            "Klasifikasi": last_klasifikasi,
                            "Kode SBU": k,
                            "Subklasifikasi": s_name,
                            "KBLI 2020": k_val,
                            "Lingkup Pekerjaan": lingkup
                        })

        if (i+1) % 20 == 0:
            print(f"  Halaman {i+1} selesai...")

df_final = pd.DataFrame(all_rows).drop_duplicates()

# Sortir agar rapi
df_final = df_final.sort_values(["Klasifikasi", "Kode SBU"])

print(f"Ekstraksi selesai. Menemukan {len(df_final)} baris data bersih.")

# Tulis ke Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# Header
headers = ["Klasifikasi", "Kode SBU", "Subklasifikasi", "KBLI 2020", "Lingkup Pekerjaan"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Data
for r, row in enumerate(df_final.values, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
        if c != 5: # Selain lingkup, bungkus teks
             ws.cell(row=r, column=c).alignment = Alignment(vertical="top")
        else:
             ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

# Widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 80

wb.save(excel_path)
print(f"Update Selesai! File: {excel_path}")
