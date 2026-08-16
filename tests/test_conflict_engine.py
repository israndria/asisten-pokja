import conflict_engine
from datetime import date


class _Result:
    data = []


class _Table:
    def __init__(self, db, name):
        self.db = db
        self.name = name
        self.filters = []
        self.operation = "select"
        self.rows = None

    def select(self, *_args):
        return self

    def delete(self):
        self.operation = "delete"
        return self

    def eq(self, key, value):
        self.filters.append((key, value))
        return self

    def in_(self, key, values):
        self.filters.append((key, set(values)))
        return self

    def limit(self, *_args):
        return self

    def upsert(self, rows, on_conflict):
        self.operation = "upsert"
        self.rows = rows
        self.conflict_keys = on_conflict.split(",")
        return self

    def execute(self):
        current = self.db.setdefault(self.name, [])

        def matches(row):
            return all(
                row.get(key) in value if isinstance(value, set) else row.get(key) == value
                for key, value in self.filters
            )

        if self.operation == "delete":
            self.db[self.name] = [
                row for row in current
                if not matches(row)
            ]
        elif self.operation == "upsert":
            for row in self.rows or []:
                match = next(
                    (
                        old for old in current
                        if all(old.get(key) == row.get(key) for key in self.conflict_keys)
                    ),
                    None,
                )
                if match is None:
                    current.append(dict(row))
                else:
                    match.update(row)
        result = _Result()
        result.data = [row for row in self.db.get(self.name, []) if matches(row)]
        return result


class _DB(dict):
    def table(self, name):
        return _Table(self, name)


def test_sync_from_pdf_replaces_stale_personnel_and_ignores_equipment(monkeypatch):
    db = _DB(
        paket_personil=[
            {
                "kode_tender": "T-1",
                "peserta_id": "P-1",
                "nama_personil": "Personel Lama",
            },
            {
                "kode_tender": "T-1",
                "peserta_id": "P-2",
                "nama_personil": "Personel Peserta Lain",
            },
        ]
    )
    monkeypatch.setattr(conflict_engine, "_sb", lambda: db)

    result = conflict_engine.sync_from_pdf(
        "T-1",
        "P-1",
        "CV Contoh",
        [
            "Budi Santoso (Tenaga Ahli)",
            "BUDI SANTOSO (Cadangan)",
            "Citra Dewi (Petugas K3)",
        ],
    )

    rows = db["paket_personil"]
    assert result == {"personil": 2}
    assert {row["nama_personil"] for row in rows} == {
        "Budi Santoso",
        "Citra Dewi",
        "Personel Peserta Lain",
    }
    assert "paket_alat" not in db


def test_tender_assignment_window_uses_gcal_sppbj_plus_excel_duration(monkeypatch):
    db = _DB(
        draft_paket=[{
            "kode_tender": "T-1",
            "nama_tender": "Paket Contoh",
            "folder_dibuat": "01. Paket Contoh",
            "kode_pokja": "001",
        }]
    )
    monkeypatch.setattr(conflict_engine, "_sb", lambda: db)
    monkeypatch.setattr(
        conflict_engine,
        "_get_tgl_sppbj_gcal",
        lambda _nama, kode_tender="": date(2026, 8, 16),
    )
    monkeypatch.setattr(
        conflict_engine,
        "_read_tender_duration_excel",
        lambda _kode, _row: 30,
    )

    assert conflict_engine._get_jadwal_paket("T-1") == (
        date(2026, 8, 16),
        date(2026, 9, 15),
    )


def test_parse_hari_jangka_accepts_numeric_excel_value():
    assert conflict_engine._parse_hari_jangka(45) == 45


def test_sync_coverage_uses_current_year_only(monkeypatch):
    db = _DB(
        draft_paket=[
            {"kode_tender": "CURRENT-1", "nomor_surat_dinas": "001/2026"},
            {"kode_tender": "CURRENT-2", "nomor_surat_dinas": "002/2026"},
            {"kode_tender": "OLD-1", "nomor_surat_dinas": "003/2025"},
            {"kode_tender": "UNKNOWN-1", "nomor_surat_dinas": ""},
            {"kode_tender": "_err_123", "nomor_surat_dinas": None},
        ],
        paket_personil=[
            {"kode_tender": "CURRENT-1"},
            {"kode_tender": "OLD-1"},
        ],
    )
    monkeypatch.setattr(conflict_engine, "_sb", lambda: db)

    assert conflict_engine.get_sync_coverage(2026) == {
        "tahun": 2026,
        "aktif": 2,
        "personil": 1,
        "lengkap": 1,
        "belum_lengkap": 1,
    }


def test_conflict_uses_only_winner_and_manual_excel_personnel(monkeypatch, tmp_path):
    from openpyxl import Workbook

    packages = {
        "10129575000": "CV Pemenang Satu",
        "10131057000": "CV Pemenang Dua",
    }
    workbooks = {}
    for code, provider in packages.items():
        path = tmp_path / f"{code}.xlsx"
        wb = Workbook()
        master = wb.active
        master.title = "@ Master Data"
        master["C4"] = code
        input_ba = wb.create_sheet("0. Input BA")
        input_ba["G7"] = provider
        input_ba["G13"] = "Budi Manual (Tenaga Ahli)"
        wb.save(path)
        workbooks[code] = path

    db = _DB(
        draft_paket=[{"kode_tender": code} for code in packages],
        tender_peserta=[
            {"kode_tender": code, "nama_peserta": provider, "is_pemenang": True}
            for code, provider in packages.items()
        ] + [
            {"kode_tender": code, "nama_peserta": "CV Peserta Kalah", "is_pemenang": False}
            for code in packages
        ],
        paket_personil=[
            {
                "kode_tender": code,
                "peserta_id": f"loser-{code}",
                "nama_penyedia": "CV Peserta Kalah",
                "nama_personil": "Budi Manual",
            }
            for code in packages
        ] + [
            {
                "kode_tender": code,
                "peserta_id": f"winner-{code}",
                "nama_penyedia": provider,
                "nama_personil": "Nama Parser Lama",
            }
            for code, provider in packages.items()
        ],
    )
    monkeypatch.setattr(conflict_engine, "_sb", lambda: db)
    monkeypatch.setattr(
        conflict_engine,
        "_get_tender_row",
        lambda code: {"folder_dibuat": str(tmp_path)},
    )
    monkeypatch.setattr(
        conflict_engine,
        "_iter_tender_workbooks",
        lambda code, _row: iter([str(workbooks[code])]),
    )
    monkeypatch.setattr(
        conflict_engine,
        "_get_jadwal_paket",
        lambda _code: (date(2026, 8, 16), date(2026, 9, 15)),
    )

    conflicts = conflict_engine.get_konflik_personil()

    assert [row["nama_personil_display"] for row in conflicts] == ["Budi Manual"]
    assert {
        entry["nama_penyedia"]
        for entry in conflicts[0]["paket"]
    } == set(packages.values())
    assert "Nama Parser Lama" not in str(conflicts)


def test_excel_personnel_requires_matching_master_code(monkeypatch, tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "wrong.xlsx"
    wb = Workbook()
    master = wb.active
    master.title = "@ Master Data"
    master["C4"] = "99999999999"
    wb.create_sheet("0. Input BA")["G13"] = "Budi Dari Paket Lain"
    wb.save(path)

    monkeypatch.setattr(
        conflict_engine,
        "_iter_tender_workbooks",
        lambda _code, _row: iter([str(path)]),
    )

    assert conflict_engine._read_excel_personil("10129575000", {}) == {
        "found": False,
        "provider": "",
        "personil": [],
    }
