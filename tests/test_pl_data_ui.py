"""Regression gate folder lokal/workbook untuk daftar operasional PL."""

from unittest.mock import patch
from datetime import date, datetime

import parse_kak_pl
import pl_engine
import pl_data_ui
import pl_ui_helpers
import zipfile
from pl_data_ui import (
    _filter_pl_family,
    filter_paket_draft_live,
    _hydrate_provider_from_excel,
    filter_paket_siap_dijadwalkan,
    filter_paket_kirim_undangan_dpp,
    format_pl_announce_log,
    get_reviu_full_pdf_path,
    filter_local_pl_rows,
    overlay_live_tahap_spse,
    select_rows_by_checkbox_state,
    get_manual_ba_date,
)
from ui_pl_pk import (
    PLPK_TAB_LABELS,
    active_rows,
    provider_identity_available,
    provider_selection_status_caption,
)
from ui_pl_jadwal import (
    _absolute_schedule_seed_for_selection,
    _absolute_schedule_seed_signature,
    _format_jadwal_submit_result,
    _validasi_perubahan_jadwal,
    filter_paket_penandatanganan_kontrak,
    filter_paket_sudah_tayang,
    filter_paket_upload_terlambat,
    is_late_upload_scan_current,
)
from ui_pl_common import (
    apply_pl7_historical_backfill,
    mark_pl7_action_success,
    summarize_pl7_action_status,
)


def test_pl8_evaluation_checkbox_defaults_checked_once_and_keep_manual_choice():
    state = {
        "pl8_do_eval_admin": False,
        "pl8_do_eval_teknis": False,
        "pl8_do_eval_harga": False,
    }

    pl_ui_helpers.ensure_pl_evaluation_checkbox_defaults(state)

    assert state["pl8_do_eval_admin"] is True
    assert state["pl8_do_eval_teknis"] is True
    assert state["pl8_do_eval_harga"] is True

    state["pl8_do_eval_teknis"] = False
    pl_ui_helpers.ensure_pl_evaluation_checkbox_defaults(state)

    assert state["pl8_do_eval_admin"] is True
    assert state["pl8_do_eval_teknis"] is False
    assert state["pl8_do_eval_harga"] is True


def test_select_rows_by_checkbox_state_keeps_source_order_and_ignores_unchecked():
    rows = [
        {"kode_paket": "P-2", "nama_paket": "dua"},
        {"kode_paket": "P-1", "nama_paket": "satu"},
        {"kode_paket": "", "nama_paket": "kosong"},
    ]
    state = {"pl10_jkk_P-1": True, "pl10_jkk_P-2": False}

    result = select_rows_by_checkbox_state(rows, state, "pl10_jkk_")

    assert result == [{"kode_paket": "P-1", "nama_paket": "satu"}]


def test_get_manual_ba_date_keeps_evaluation_and_selection_dates_separate():
    evaluasi = date(2026, 9, 1)
    pemilihan = date(2026, 9, 3)

    assert get_manual_ba_date("evaluasi", evaluasi, pemilihan) == evaluasi
    assert get_manual_ba_date("hasil", evaluasi, pemilihan) == pemilihan
    assert get_manual_ba_date("lainnya", evaluasi, pemilihan) is None


def test_pl_setup_draft_filter_excludes_active_or_published_status():
    assert pl_engine.is_paket_draft({"status": "draft"}) is True
    assert pl_engine.is_paket_draft({"status": "paket sedang berjalan"}) is False
    assert pl_engine.is_paket_draft({"status": "draft", "tahap_spse": "Upload Dokumen Penawaran"}) is False
    assert pl_engine.is_paket_draft({"status": "", "tahap_spse": "Upload Dokumen Penawaran"}) is False
    assert pl_engine.is_paket_draft({"status": "", "tahap_spse": ""}) is False


def test_pl_operational_filter_requires_active_status_or_stage():
    assert pl_engine.is_paket_berjalan({"status": "draft"}) is False
    assert pl_engine.is_paket_berjalan({"status": "", "tahap_spse": ""}) is False
    assert pl_engine.is_paket_berjalan({"status": "berjalan"}) is True
    assert pl_engine.is_paket_berjalan({"status": "evaluasi"}) is True
    assert pl_engine.is_paket_berjalan({"status": "draft", "tahap_spse": "Upload Dokumen Penawaran"}) is True
    assert pl_engine.is_paket_berjalan({"status": "selesai"}) is False
    assert pl_engine.is_paket_berjalan({"status": "berjalan", "tahap_spse": "Penandatanganan Kontrak"}) is False


def test_pl_operational_gate_excludes_pre_upload_and_ambiguous_stage():
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Upload Dokumen Penawaran",
    }) is False
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Upload Dokumen Penawaran",
        "tgl_pembukaan": "2026-08-26",
    }, today=date(2026, 8, 26)) is True
    assert pl_engine.is_paket_operasional_eligible({
        "status": "draft",
        "tgl_pembukaan": "2026-08-26",
    }, today=date(2026, 8, 26)) is False
    assert pl_engine.is_paket_operasional_eligible({
        "status": "draft",
        "tgl_pembukaan": "2026-08-27",
    }, today=date(2026, 8, 26)) is False
    assert pl_engine.is_paket_operasional_eligible({
        "status": "draft",
        "tahap_spse": "",
    }) is False
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Tidak Ada Jadwal",
    }) is True
    assert pl_engine.is_paket_operasional_eligible({
        "status": "draft",
        "tahap_spse": "Pembukaan Dokumen Penawaran",
    }) is True
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Evaluasi Penawaran",
    }) is True
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Klarifikasi Teknis dan Negosiasi",
    }) is True
    assert pl_engine.is_paket_operasional_eligible({
        "status": "paket sedang berjalan",
        "tahap_spse": "Penandatanganan Kontrak",
    }) is False


def test_filter_paket_penandatanganan_kontrak_uses_live_stage_or_status():
    rows = [
        {"kode_paket": "stage", "tahap_spse": "Penandatanganan Kontrak", "status": "berjalan"},
        {"kode_paket": "status", "tahap_spse": "", "status": "Penandatanganan Kontrak"},
        {"kode_paket": "done", "tahap_spse": "Paket Sudah Selesai", "status": "selesai"},
        {"kode_paket": "active", "tahap_spse": "Evaluasi Penawaran", "status": "berjalan"},
    ]

    assert [row["kode_paket"] for row in filter_paket_penandatanganan_kontrak(rows)] == [
        "stage",
        "status",
    ]


def test_filter_paket_sudah_tayang_is_independent_and_keeps_prestart_until_t5():
    rows = [
        {"kode_paket": "prestart", "status": "draft", "tahap_spse": "Paket Belum Dilaksanakan"},
        {"kode_paket": "upload", "status": "draft", "tahap_spse": "Upload Dokumen Penawaran"},
        {"kode_paket": "contract", "status": "berjalan", "tahap_spse": "Penandatanganan Kontrak"},
        {"kode_paket": "announced", "status": "draft", "tahap_spse": "Pengumuman"},
        {"kode_paket": "session", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "draft", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "done", "status": "selesai", "tahap_spse": "Paket Sudah Selesai"},
    ]

    assert [row["kode_paket"] for row in filter_paket_sudah_tayang(
        rows, {"session": {"status": "sudah diumumkan"}}
    )] == ["prestart", "upload", "contract", "announced", "session"]


def test_filter_paket_upload_terlambat_requires_zero_participant_and_past_t1():
    now = datetime(2026, 8, 29, 16, 30)
    rows = [
        {"kode_paket": "past-zero", "nama_paket": "Past zero"},
        {"kode_paket": "past-one", "nama_paket": "Past one"},
        {"kode_paket": "future-zero", "nama_paket": "Future zero"},
        {"kode_paket": "error-zero", "nama_paket": "Error zero"},
        {"kode_paket": "short-zero", "nama_paket": "Short zero"},
    ]
    schedule = lambda deadline: [
        {"mulai": now.replace(hour=8), "selesai": deadline},
        {"mulai": deadline, "selesai": deadline},
        {"mulai": deadline, "selesai": deadline},
        {"mulai": deadline, "selesai": deadline},
        {"mulai": deadline, "selesai": deadline},
    ]
    schedules = {
        "past-zero": schedule(datetime(2026, 8, 29, 12, 0)),
        "past-one": schedule(datetime(2026, 8, 29, 12, 0)),
        "future-zero": schedule(datetime(2026, 8, 29, 18, 0)),
        "error-zero": schedule(datetime(2026, 8, 29, 12, 0)),
        "short-zero": schedule(datetime(2026, 8, 29, 12, 0))[:4],
    }
    statuses = {
        "past-zero": {"jumlah": 0, "error": None},
        "past-one": {"jumlah": 1, "error": None},
        "future-zero": {"jumlah": 0, "error": None},
        "error-zero": {"jumlah": 0, "error": "timeout"},
        "short-zero": {"jumlah": 0, "error": None},
    }

    result = filter_paket_upload_terlambat(
        rows, statuses, schedules, now=now
    )

    assert [row["kode_paket"] for row in result] == ["past-zero"]
    assert result[0]["_upload_deadline"] == datetime(2026, 8, 29, 12, 0)


def test_filter_paket_upload_terlambat_accepts_spse_datetime_string():
    now = datetime(2026, 8, 29, 16, 30)
    schedule = [
        {"mulai": "29-08-2026 08:00", "selesai": "29-08-2026 12:00"}
        for _ in range(5)
    ]
    result = filter_paket_upload_terlambat(
        [{"kode_paket": "P-1"}],
        {"P-1": {"jumlah": "0", "error": None}},
        {"P-1": schedule},
        now=now,
    )

    assert [row["kode_paket"] for row in result] == ["P-1"]


def test_late_upload_scan_rejects_legacy_session_state():
    assert is_late_upload_scan_current({"schema_version": 2}) is True
    assert is_late_upload_scan_current({"schema_version": 1}) is False
    assert is_late_upload_scan_current({"errors": {}}) is False
    assert is_late_upload_scan_current(None) is False


def _schedule_for_test(t3_finish, t4_start):
    return [
        {"mulai": datetime(2026, 8, 26, 21, 45), "selesai": datetime(2026, 9, 2, 10, 0)},
        {"mulai": datetime(2026, 9, 2, 10, 1), "selesai": datetime(2026, 9, 2, 11, 5)},
        {"mulai": datetime(2026, 9, 2, 11, 6), "selesai": t3_finish},
        {"mulai": t4_start, "selesai": datetime(2026, 9, 3, 16, 15)},
        {"mulai": datetime(2026, 9, 3, 16, 30), "selesai": datetime(2026, 9, 11, 17, 1)},
    ]


def test_absolute_schedule_seed_uses_first_selected_code_not_stale_loaded_entry():
    stale = _schedule_for_test(
        datetime(2026, 9, 3, 9, 0), datetime(2026, 9, 3, 9, 0)
    )
    selected = _schedule_for_test(
        datetime(2026, 9, 3, 16, 0), datetime(2026, 9, 3, 9, 0)
    )

    loaded = {
        "stale-package": {"jadwal": stale},
        "selected-package": {"jadwal": selected},
    }

    assert _absolute_schedule_seed_for_selection(["selected-package"], loaded) is selected


def test_validation_accepts_unchanged_existing_t3_t4_overlap():
    schedule = _schedule_for_test(
        datetime(2026, 9, 3, 16, 0), datetime(2026, 9, 3, 9, 0)
    )

    assert _validasi_perubahan_jadwal(schedule, schedule) == []


def test_validation_leaves_cross_stage_overlap_to_spse_warning():
    current = _schedule_for_test(
        datetime(2026, 9, 3, 9, 0), datetime(2026, 9, 3, 9, 0)
    )
    proposed = _schedule_for_test(
        datetime(2026, 9, 3, 16, 0), datetime(2026, 9, 3, 9, 0)
    )

    assert _validasi_perubahan_jadwal(current, proposed) == []


def test_absolute_seed_signature_changes_only_for_new_selection_or_live_schedule():
    schedule = _schedule_for_test(
        datetime(2026, 9, 3, 16, 0), datetime(2026, 9, 3, 9, 0)
    )
    same = _absolute_schedule_seed_signature(["selected-package"], schedule)

    assert same == _absolute_schedule_seed_signature(["selected-package"], schedule)
    assert same != _absolute_schedule_seed_signature(["other-package"], schedule)
    changed = [dict(row) for row in schedule]
    changed[4]["selesai"] = datetime(2026, 9, 11, 17, 2)
    assert same != _absolute_schedule_seed_signature(["selected-package"], changed)


def test_successful_schedule_result_uses_folder_label_not_code_or_http_status():
    result = _format_jadwal_submit_result(
        {
            "kode_paket": "10999983000",
            "nomor_urut": 66,
            "nama_paket": "Konsultan Pengawasan Paket 20",
        },
        True,
        "HTTP 302",
    )

    assert result == "✅ 66. Konsultan Pengawasan Paket 20 — Berhasil"
    assert "10999983000" not in result
    assert "HTTP 302" not in result

def test_filter_paket_siap_dijadwalkan_keeps_unpublished_draft_and_hides_tayang():
    rows = [
        {"kode_paket": "prestart", "status": "draft", "tahap_spse": "Paket Belum Dilaksanakan"},
        {"kode_paket": "pending", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "announced", "status": "draft", "tahap_spse": "Pengumuman"},
        {"kode_paket": "scheduled", "status": "draft", "tahap_spse": "Upload Dokumen Penawaran"},
        {"kode_paket": "not-draft", "status": "berjalan", "tahap_spse": ""},
    ]

    assert [row["kode_paket"] for row in filter_paket_siap_dijadwalkan(rows)] == [
        "pending",
    ]


def test_filter_paket_draft_live_hides_published_and_prestart_live_stage():
    rows = [
        {"kode_paket": "draft", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "published", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "prestart", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "running", "status": "berjalan", "tahap_spse": ""},
    ]
    live_tahap = {
        "published": "Upload Dokumen Penawaran",
        "prestart": "Paket Belum Dilaksanakan",
    }

    assert [row["kode_paket"] for row in filter_paket_draft_live(
        rows,
        live_tahap,
        live_status_ok=True,
    )] == ["draft"]


def test_filter_paket_draft_live_vetoes_live_code_even_when_stage_is_blank():
    rows = [
        {"kode_paket": "live-blank-stage", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "draft", "status": "draft", "tahap_spse": ""},
    ]

    assert [row["kode_paket"] for row in filter_paket_draft_live(
        rows,
        {"live-blank-stage": ""},
        live_status_ok=True,
    )] == ["draft"]


def test_filter_paket_draft_live_fails_closed_when_live_verification_fails():
    rows = [{"kode_paket": "draft-local", "status": "draft", "tahap_spse": ""}]

    assert filter_paket_draft_live(rows, {}, live_status_ok=False) == []


def test_filter_paket_kirim_undangan_dpp_hides_live_tayang_from_draft_cache():
    rows = [
        {"kode_paket": "pending", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "live", "status": "draft", "tahap_spse": ""},
        {"kode_paket": "stage", "status": "draft", "tahap_spse": "Upload Dokumen Penawaran"},
        {"kode_paket": "running", "status": "berjalan", "tahap_spse": ""},
    ]

    assert [row["kode_paket"] for row in filter_paket_kirim_undangan_dpp(
        rows, {"live": {"status": "sudah diumumkan"}}
    )] == ["pending"]


def test_filter_paket_kirim_undangan_dpp_fails_closed_when_live_sync_fails():
    rows = [{"kode_paket": "draft-local", "status": "draft", "tahap_spse": ""}]
    assert filter_paket_kirim_undangan_dpp(rows, live_status_ok=False) == []


def test_filter_paket_kirim_undangan_dpp_hides_exact_local_reviu_full(tmp_path):
    folder = tmp_path / "36. PLPK - Paket Jalan"
    full_dir = folder / "6. BA Reviu Lengkap"
    full_dir.mkdir(parents=True)
    full = full_dir / "2. Isi Reviu Fix Full - PLPK36.pdf"
    full.write_bytes(b"%PDF-1.7\n")
    row = {
        "kode_paket": "36",
        "jenis_pl": "PK",
        "status": "draft",
        "tahap_spse": "",
        "_folder_lokal": str(folder),
    }

    assert get_reviu_full_pdf_path(row) == full
    # Keberadaan Isi Reviu Full hanya bukti lokal/indikator; bukan gate
    # undangan. Gate daftar tetap status tayang live SPSE.
    assert filter_paket_kirim_undangan_dpp([row]) == [row]

    wrong = full_dir / "2. Isi Reviu Fix Full - PLPK999.pdf"
    full.unlink()
    wrong.write_bytes(b"%PDF-1.7\n")
    assert get_reviu_full_pdf_path(row) is None
    assert filter_paket_kirim_undangan_dpp([row]) == [row]


def test_filter_paket_penandatanganan_kontrak_limits_t5_to_six_hour_window():
    now = datetime(2026, 8, 27, 18, 0)
    rows = [
        {"kode_paket": "recent", "tahap_spse": "Penandatanganan Kontrak"},
        {"kode_paket": "upcoming", "status": "Penandatanganan Kontrak"},
        {"kode_paket": "old", "tahap_spse": "Penandatanganan Kontrak"},
        {"kode_paket": "unknown", "tahap_spse": "Penandatanganan Kontrak"},
    ]
    schedules = {
        "recent": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 27, 16, 0)}],
        "upcoming": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 27, 23, 59)}],
        "old": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 23, 16, 0)}],
        "unknown": [],
    }

    result = filter_paket_penandatanganan_kontrak(
        rows,
        schedule_loader=lambda code: schedules[code],
        now=now,
        window_hours=6,
    )

    assert [row["kode_paket"] for row in result] == ["recent", "upcoming"]


def test_filter_paket_sudah_tayang_hides_packages_with_t5_started_over_three_days_ago():
    rows = [
        {
            "kode_paket": "old-disdag",
            "tahap_spse": "Penandatanganan Kontrak",
            "status": "paket sedang berjalan",
            "satker": "Dinas Perdagangan",
        },
        {
            "kode_paket": "boundary",
            "tahap_spse": "Penandatanganan Kontrak",
            "status": "paket sedang berjalan",
        },
        {
            "kode_paket": "recent",
            "tahap_spse": "Penandatanganan Kontrak",
            "status": "paket sedang berjalan",
        },
        {
            "kode_paket": "future",
            "tahap_spse": "Paket Belum Dilaksanakan",
            "status": "draft",
        },
        {
            "kode_paket": "unreadable",
            "tahap_spse": "Penandatanganan Kontrak",
            "status": "paket sedang berjalan",
        },
    ]
    schedules = {
        "old-disdag": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 27, 16, 45)}],
        "boundary": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 28, 21, 30)}],
        "recent": [{}, {}, {}, {}, {"mulai": datetime(2026, 8, 30, 16, 45)}],
        "future": [{}, {}, {}, {}, {"mulai": datetime(2026, 9, 2, 15, 46)}],
        "unreadable": [],
    }

    result = filter_paket_sudah_tayang(
        rows,
        schedule_loader=lambda code: schedules[code],
        now=datetime(2026, 8, 31, 21, 30),
    )

    assert [row["kode_paket"] for row in result] == [
        "boundary",
        "recent",
        "future",
        "unreadable",
    ]


def test_plpk_active_rows_excludes_draft_ambiguous_and_completed():
    rows = [
        {"kode_paket": "D", "status": "draft"},
        {"kode_paket": "A", "status": "berjalan", "tahap_spse": "Evaluasi Penawaran"},
        {"kode_paket": "X", "status": ""},
        {"kode_paket": "S", "status": "selesai"},
    ]

    filtered, duplicate_count = active_rows(lambda _kind: rows, pl_engine)

    assert [row["kode_paket"] for row in filtered] == ["A"]
    assert duplicate_count == 0


def test_live_tahap_overrides_stale_local_stage_before_operational_gate():
    state = {
        "pl_umumkan_status": {
            "10999940000": {
                "status": "sudah diumumkan",
                "tahap_spse": "Penandatanganan Kontrak",
            },
            "10999940001": {
                "status": "sudah diumumkan",
                "tahap_spse": "Evaluasi Penawaran",
            },
        }
    }
    rows = [
        {
            "kode_paket": "10999940000",
            "status": "paket sedang berjalan",
            "tahap_spse": "Tidak Ada Jadwal",
        },
        {
            "kode_paket": "10999940001",
            "status": "paket sedang berjalan",
            "tahap_spse": "Tidak Ada Jadwal",
        },
    ]

    with patch.object(pl_data_ui.st, "session_state", state):
        live_rows = overlay_live_tahap_spse(rows)

    assert live_rows[0]["tahap_spse"] == "Penandatanganan Kontrak"
    assert live_rows[1]["tahap_spse"] == "Evaluasi Penawaran"
    assert pl_engine.is_paket_operasional_eligible(live_rows[0]) is False
    assert pl_engine.is_paket_operasional_eligible(live_rows[1]) is True


def test_live_tahap_hides_tayang_from_provider_draft_candidates():
    state = {
        "pl_umumkan_status": {
            "10999940002": {
                "status": "sudah diumumkan",
                "tahap_spse": "Upload Dokumen Penawaran",
            }
        }
    }
    rows = [
        {
            "kode_paket": "10999940002",
            "status": "draft",
            "tahap_spse": "",
        }
    ]

    with patch.object(pl_data_ui.st, "session_state", state):
        live_rows = overlay_live_tahap_spse(rows)

    assert live_rows[0]["tahap_spse"] == "Upload Dokumen Penawaran"
    assert [row for row in live_rows if pl_engine.is_paket_draft(row)] == []


def test_provider_identity_allows_name_fallback_without_npwp():
    assert provider_identity_available({"npwp_penyedia": "123"}) is True
    assert provider_identity_available({"nama_penyedia": "CV. Nama Penyedia"}) is True
    assert provider_identity_available({"npwp_penyedia": "", "nama_penyedia": ""}) is False


def test_provider_selection_existing_other_is_still_success_caption():
    caption = provider_selection_status_caption(
        {"kode_paket": "PK-56"},
        {"PK-56": {"ok": True, "status": "sudah_terpilih_lain", "nama": "CAKRAWALA BANUA"}},
    )

    assert caption == "✅ Penyedia sudah dipilih: CAKRAWALA BANUA"


def test_format_pl_announce_log_contains_code_stage_http_and_outcome():
    lines = format_pl_announce_log(
        [
            {
                "kode_paket": "PK-56",
                "paket": "56. Paket Rabat Beton",
                "ok": True,
                "status_code": 302,
                "pesan": "Berhasil diumumkan",
                "stage": "GET /edit + POST /pengumumanpp",
            },
            {
                "kode_paket": "PK-57",
                "paket": "57. Paket Drainase",
                "ok": False,
                "status_code": 404,
                "error": "GET edit paket gagal",
                "stage": "GET /edit",
            },
        ],
        family="PK",
        started_at=datetime(2026, 8, 28, 10, 30),
    )

    text = "\n".join(lines)
    assert "Family: PK" in text
    assert "PK-56" in text and "HTTP  : 302" in text and "BERHASIL" in text
    assert "PK-57" in text and "HTTP  : 404" in text and "GAGAL" in text
    assert "GET /edit" in text


def test_plpk_tab_five_and_six_are_swapped():
    assert PLPK_TAB_LABELS[4] == "5️⃣ Buat & Monitor Jadwal"
    assert PLPK_TAB_LABELS[5] == "6️⃣ Pilih Penyedia & Umumkan"


def test_pl7_action_status_is_per_package_and_family_specific():
    state = {}
    rows = [
        {"kode_paket": "PK-1", "nama_paket": "Paket Jalan", "nomor_urut": 81},
        {"kode_paket": "PK-2", "nama_paket": "Paket Jembatan", "nomor_urut": 82},
    ]

    mark_pl7_action_success(state, "PK", "PK-1", "download")
    mark_pl7_action_success(state, "PK", "PK-1", "parse")
    mark_pl7_action_success(state, "PK", "PK-1", "parse_eval")
    mark_pl7_action_success(state, "JKK", "PK-2", "hps")

    pk_summary = summarize_pl7_action_status(state=state, rows=rows, family="PK", label_fn=pl_ui_helpers._pl_label)
    jkk_summary = summarize_pl7_action_status(state=state, rows=rows, family="JKK", label_fn=pl_ui_helpers._pl_label)

    assert pk_summary[0]["Download Kualifikasi"] == "✅"
    assert pk_summary[0]["Parse & Populate"] == "✅"
    assert pk_summary[0]["Parse @ Evaluasi"] == "✅"
    assert pk_summary[0]["Update HPS"] == "—"
    assert pk_summary[0]["Status Paket"] == "Sudah ada aksi sukses"
    assert pk_summary[1]["Status Paket"] == "Belum diproses"
    assert jkk_summary[0]["Status Paket"] == "Belum diproses"
    assert jkk_summary[1]["Update HPS"] == "✅"


def test_pl7_historical_backfill_is_explicit_idempotent_and_does_not_touch_new_or_excluded_packages():
    state = {}
    rows = [
        {"kode_paket": "10999962000", "nama_paket": "Konsultan Pengawasan Paket 10"},
        {"kode_paket": "10999965000", "nama_paket": "Konsultan Pengawasan Paket 11"},
        {"kode_paket": "10999980000", "nama_paket": "Konsultan Pengawasan Paket 19"},
        {"kode_paket": "10999983000", "nama_paket": "Konsultan Pengawasan Paket 20"},
        {"kode_paket": "JKK-NEW", "nama_paket": "Konsultan Pengawasan Paket Baru"},
    ]

    assert apply_pl7_historical_backfill(state, "JKK", rows) == 1
    assert apply_pl7_historical_backfill(state, "JKK", rows) == 0

    summary = summarize_pl7_action_status(
        state=state, rows=rows, family="JKK", label_fn=pl_ui_helpers._pl_label
    )
    assert all(summary[0][key] == "✅" for key in (
        "Download Kualifikasi", "Parse & Populate", "Update HPS"
    ))
    assert all(row["Status Paket"] == "Belum diproses" for row in summary[1:])


def test_pl_label_falls_back_to_number_in_physical_folder(tmp_path):
    folder = tmp_path / "56. PLPK - Paket Rabat Beton"
    folder.mkdir()

    label = pl_ui_helpers._pl_label({
        "nama_paket": "Paket Rabat Beton",
        "_folder_lokal": str(folder),
    })

    assert label == "56. Paket Rabat Beton"


def test_pl_label_uses_planned_number_before_folder_exists():
    label = pl_ui_helpers._pl_label({
        "nama_paket": "Belanja Modal Bangunan Gedung",
        "_display_nomor_urut": 80,
    })

    assert label == "80. Belanja Modal Bangunan Gedung"


def test_pl_family_loader_isolated_and_fail_closed():
    rows = [
        {"kode_paket": "PK-1", "jenis_pl": "PK"},
        {"kode_paket": "JKK-1", "jenis_pl": "JKK"},
        {"kode_paket": "UNKNOWN-1", "jenis_pl": ""},
    ]

    assert [row["kode_paket"] for row in _filter_pl_family(rows, "PK")] == ["PK-1"]
    assert [row["kode_paket"] for row in _filter_pl_family(rows, "JKK")] == ["JKK-1"]
    assert _filter_pl_family(rows, "unexpected-family") == []


def test_provider_sync_private_alias_matches_canonical_helper():
    assert (
        pl_ui_helpers._sinkronkan_identitas_penyedia_pl
        is pl_ui_helpers.sinkronkan_identitas_penyedia_pl
    )


def test_plpk_provider_uses_c77_c78_and_overrides_stale_cache(monkeypatch, tmp_path):
    folder = tmp_path / "30. PLPK - Paket Fisik"
    folder.mkdir()
    workbook = folder / "0. BAPLPK - Paket Fisik.xlsm"
    workbook.write_bytes(b"placeholder")

    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (str(folder), "30"),
    )
    monkeypatch.setattr(
        pl_ui_helpers,
        "_cari_xlsm_pl",
        lambda candidate: str(workbook) if candidate == str(folder) else None,
    )

    class FakeSheet:
        values = {
            "C51": "1 Unit",
            "C52": "1 Unit",
            "C77": "CV. HARAPAN GROUP",
            "C78": "0503336018733000",
        }

        def __getitem__(self, cell):
            return type("Cell", (), {"value": self.values[cell]})()

    class FakeWorkbook:
        def __getitem__(self, sheet):
            assert sheet == "@ Master Data"
            return FakeSheet()

        def close(self):
            pass

    import openpyxl
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: FakeWorkbook())

    row = {
        "kode_paket": "PK-30",
        "nama_paket": "Paket fisik",
        "jenis_pl": "PK",
        "folder_dibuat": True,
        "nama_penyedia": "1 Unit",
        "npwp_penyedia": "1 Unit",
    }

    result = _hydrate_provider_from_excel([row])

    assert result[0]["nama_penyedia"] == "CV. HARAPAN GROUP"
    assert result[0]["npwp_penyedia"] == "0503336018733000"


def test_pljkk_provider_keeps_c51_c52_source(monkeypatch, tmp_path):
    folder = tmp_path / "10. PLJKK - Paket Konsultan"
    folder.mkdir()
    workbook = folder / "0. BAPLJKK - Paket Konsultan.xlsm"
    workbook.write_bytes(b"placeholder")

    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (str(folder), "10"),
    )
    monkeypatch.setattr(
        pl_ui_helpers,
        "_cari_xlsm_pl",
        lambda candidate: str(workbook) if candidate == str(folder) else None,
    )

    class FakeSheet:
        values = {
            "C51": "CV. RAHMAT BANUA KONSULTAN",
            "C52": "911052223733000",
            "C77": "CV. WRONG",
            "C78": "0000000000000000",
        }

        def __getitem__(self, cell):
            return type("Cell", (), {"value": self.values[cell]})()

    class FakeWorkbook:
        def __getitem__(self, sheet):
            assert sheet == "@ Master Data"
            return FakeSheet()

        def close(self):
            pass

    import openpyxl
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: FakeWorkbook())

    row = {
        "kode_paket": "JKK-10",
        "nama_paket": "Paket konsultan",
        "jenis_pl": "JKK",
        "folder_dibuat": True,
    }

    result = _hydrate_provider_from_excel([row])

    assert result[0]["nama_penyedia"] == "CV. RAHMAT BANUA KONSULTAN"
    assert result[0]["npwp_penyedia"] == "911052223733000"


def test_provider_uses_xml_fast_path_without_openpyxl(monkeypatch, tmp_path):
    folder = tmp_path / "10. PLJKK - Paket Konsultan"
    folder.mkdir()
    workbook = folder / "0. BAPLJKK - Paket Konsultan.xlsm"
    with zipfile.ZipFile(workbook, "w") as archive:
        archive.writestr(
            "xl/workbook.xml",
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="@ Master Data" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Target="worksheets/sheet1.xml"/></Relationships>',
        )
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<sheetData>"
            '<row r="51"><c r="C51" t="inlineStr"><is><t>CV. XML CEPAT</t></is></c></row>'
            '<row r="52"><c r="C52" t="inlineStr"><is><t>1234567890123456</t></is></c></row>'
            "</sheetData></worksheet>",
        )

    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (str(folder), "10"),
    )
    monkeypatch.setattr(
        pl_ui_helpers,
        "_cari_xlsm_pl",
        lambda candidate: str(workbook) if candidate == str(folder) else None,
    )
    import openpyxl
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("OpenPyXL tidak boleh dipanggil pada fast path")
        ),
    )
    pl_engine.read_master_cells_cached.cache_clear()

    result = _hydrate_provider_from_excel([{
        "kode_paket": "JKK-XML",
        "nama_paket": "Paket XML",
        "jenis_pl": "JKK",
    }])

    assert result[0]["nama_penyedia"] == "CV. XML CEPAT"
    assert result[0]["npwp_penyedia"] == "1234567890123456"


def test_row_without_local_folder_is_hidden(monkeypatch, tmp_path):
    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (None, ""),
    )

    rows = [{
        "kode_paket": "PK-NO-FOLDER",
        "nama_paket": "Paket belum dibuat",
        "jenis_pl": "PK",
        "folder_dibuat": None,
    }]

    assert filter_local_pl_rows(rows) == []


def test_physical_folder_and_workbook_are_authoritative(monkeypatch, tmp_path):
    folder = tmp_path / "13. PLPK - Paket Fisik"
    folder.mkdir()
    workbook = folder / "0. BAPLPK - Paket Fisik.xlsm"
    workbook.write_bytes(b"placeholder")

    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (str(folder), "13"),
    )
    monkeypatch.setattr(
        pl_ui_helpers,
        "_cari_xlsm_pl",
        lambda candidate: str(workbook) if candidate == str(folder) else None,
    )

    row = {
        "kode_paket": "PK-FISIK",
        "nama_paket": "Paket fisik",
        "jenis_pl": "PK",
        # DB status boleh stale/kosong; disk tetap sumber gate.
        "folder_dibuat": None,
    }

    result = filter_local_pl_rows([row])

    assert len(result) == 1
    assert result[0]["_folder_lokal"] == str(folder)
    assert result[0]["_xlsm_lokal"] == str(workbook)


def test_folder_without_workbook_is_hidden(monkeypatch, tmp_path):
    folder = tmp_path / "14. PLPK - Tanpa Workbook"
    folder.mkdir()

    monkeypatch.setattr(
        parse_kak_pl,
        "_resolve_folder_pl",
        lambda *_args, **_kwargs: (str(folder), "14"),
    )
    monkeypatch.setattr(pl_ui_helpers, "_cari_xlsm_pl", lambda _candidate: None)

    row = {
        "kode_paket": "PK-NO-XLSM",
        "nama_paket": "Paket tanpa workbook",
        "jenis_pl": "PK",
        "folder_dibuat": True,
    }

    assert filter_local_pl_rows([row]) == []


def test_stale_number_cannot_resolve_other_package(monkeypatch, tmp_path):
    other_folder = tmp_path / "15. PLPK - Paket Lain"
    other_folder.mkdir()
    other_workbook = other_folder / "0. BA Paket Lain.xlsm"
    other_workbook.write_bytes(b"placeholder")

    def fake_resolve(_nomor, _nama, _jenis, *, is_ulang=False, strict_name=False):
        assert strict_name is True
        return (None, "")

    monkeypatch.setattr(parse_kak_pl, "_resolve_folder_pl", fake_resolve)
    monkeypatch.setattr(
        pl_ui_helpers,
        "_cari_xlsm_pl",
        lambda _candidate: str(other_workbook),
    )

    row = {
        "kode_paket": "PK-STALE-NUMBER",
        "nama_paket": "Paket yang belum dibuat",
        "jenis_pl": "PK",
        "nomor_urut": "15",
        "folder_dibuat": True,
    }

    assert filter_local_pl_rows([row]) == []


def test_strict_resolver_accepts_officially_truncated_folder(monkeypatch, tmp_path):
    root = tmp_path / "pk"
    root.mkdir()
    package_name = "Belanja Modal Bangunan Gedung Pertokoan Koperasi Pasar Pembangunan Gapura Pintu Gerbang Pasar Binuang Jalan Bay Pass"
    full_name = f"28. PLPK - {package_name}"
    truncated_name = pl_engine.truncate_nama_folder(str(root), full_name)
    folder = root / truncated_name
    folder.mkdir()

    import config

    monkeypatch.setattr(config, "OUTPUT_DIR_PL_PK", str(root))

    resolved, number = parse_kak_pl._resolve_folder_pl(
        28,
        package_name,
        "PK",
        strict_name=True,
    )

    assert resolved == str(folder)
    assert number == "28"


def test_strict_resolver_accepts_unique_short_physical_name(monkeypatch, tmp_path):
    import config

    root = tmp_path / "plpk"
    folder = root / "28. PLPK - Belanja Modal Bangunan Gedung Pertokoan-Koperasi-Pasar Pembangunan"
    folder.mkdir(parents=True)
    monkeypatch.setattr(config, "OUTPUT_DIR_PL_PK", str(root))

    resolved, nomor = parse_kak_pl._resolve_folder_pl(
        28,
        "Belanja Modal Bangunan Gedung Pertokoan/Koperasi/Pasar Pembangunan Gapura Pintu Gerbang",
        "PK",
        strict_name=True,
    )

    assert resolved == str(folder)
    assert nomor == "28"


def test_strict_resolver_rejects_ambiguous_short_physical_name(monkeypatch, tmp_path):
    import config

    root = tmp_path / "plpk"
    (root / "28. PLPK - Belanja Modal Pembangunan A").mkdir(parents=True)
    (root / "28. PLPK - Belanja Modal Pembangunan B").mkdir(parents=True)
    monkeypatch.setattr(config, "OUTPUT_DIR_PL_PK", str(root))

    resolved, nomor = parse_kak_pl._resolve_folder_pl(
        28,
        "Belanja Modal Pembangunan",
        "PK",
        strict_name=True,
    )

    assert resolved is None
    assert nomor == ""

def test_status_peserta_tab1_lazy_loads_only_when_requested():
    state = {}
    kode = ("1001", "1002")
    status = {"1001": {"jumlah": 1}, "1002": {"jumlah": 0}}

    with patch.object(pl_data_ui.st, "session_state", state), patch.object(
        pl_data_ui.st, "button", side_effect=[False, True]
    ), patch.object(
        pl_data_ui, "fetch_status_semua_paket_cached", return_value=status
    ) as fetch:
        awal = pl_data_ui.load_status_peserta_on_demand(
            kode, "test_status", "test_status_button"
        )
        assert awal == {}
        fetch.assert_not_called()

        setelah_muat = pl_data_ui.load_status_peserta_on_demand(
            kode, "test_status", "test_status_button"
        )

    assert setelah_muat == status
    assert state["test_status"] == status
    fetch.assert_called_once_with(kode)


def test_provider_selection_status_auto_syncs_once_per_signature():
    state = {}
    provider_items = (("PK-1", "CV. Target", "0012345678901234"),)
    status = {"PK-1": {"ok": True, "status": "sudah_terpilih", "nama": "CV. Target"}}

    class _Spinner:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    with patch.object(pl_data_ui.st, "session_state", state), patch.object(
        pl_data_ui.st, "button", return_value=False
    ), patch.object(pl_data_ui.st, "spinner", return_value=_Spinner()), patch.object(
        pl_data_ui, "fetch_status_pilih_penyedia_cached", return_value=status
    ) as fetch:
        first = pl_data_ui.load_status_pilih_penyedia_on_demand(
            provider_items, "provider_status", "provider_status_button"
        )
        second = pl_data_ui.load_status_pilih_penyedia_on_demand(
            provider_items, "provider_status", "provider_status_button"
        )

    assert first == status
    assert second == status
    assert state["provider_status_signature"] == provider_items
    fetch.assert_called_once_with(provider_items)


def test_paket_umumkan_status_uses_session_then_spse_fields():
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {"kode_paket": "34", "status": "draft"},
        {"34": {"status": "sudah diumumkan"}},
    ) is True
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {"kode_paket": "35", "status": "berjalan"},
        {},
    ) is True
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {"kode_paket": "36", "status": "draft", "tahap_spse": "Pengumuman"},
        {},
    ) is True
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {"kode_paket": "36b", "status": "draft", "tahap_spse": "Upload Dokumen Penawaran"},
        {},
    ) is True
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {"kode_paket": "37", "status": "draft", "tahap_spse": ""},
        {},
    ) is False
    assert pl_data_ui.is_paket_sudah_diumumkan(
        {
            "kode_paket": "34",
            "status": "draft",
            "tahap_spse": "Paket Belum Dilaksanakan",
        },
        {},
    ) is True


def test_mark_paket_sudah_diumumkan_ke_session_state():
    state = {}
    with patch.object(pl_data_ui.st, "session_state", state):
        pl_data_ui.mark_paket_sudah_diumumkan(
            "34", {"status_code": 302, "location": "/nontender/34"}
        )

    assert state["pl_umumkan_status"]["34"] == {
        "status": "sudah diumumkan",
        "status_code": 302,
        "location": "/nontender/34",
    }


def test_mark_tahap_spse_sudah_diumumkan_menyimpan_status_batch():
    state = {}
    with patch.object(pl_data_ui.st, "session_state", state):
        jumlah = pl_data_ui.mark_tahap_spse_sudah_diumumkan({
            "34": "Paket Belum Dilaksanakan",
            "35": "Upload Dokumen Penawaran",
            "36": "",
        })

    assert jumlah == 2
    assert state["pl_umumkan_status"]["34"]["tahap_spse"] == "Paket Belum Dilaksanakan"
    assert state["pl_umumkan_status"]["35"]["status"] == "sudah diumumkan"


def test_update_hps_paket_pl_backs_up_workbook_before_official_writer(tmp_path, monkeypatch):
    workbook = tmp_path / "0. BAPLPK - Uji.xlsm"
    workbook.write_bytes(b"macro-workbook")
    calls = {}

    class Engine:
        @staticmethod
        def _find_xlsm(kode_paket):
            calls["kode"] = kode_paket
            return str(workbook)

    def fake_writer(kode_paket, excel_path, progress_cb=None):
        calls["writer"] = (kode_paket, excel_path, progress_cb)
        return {"ok": True, "pesan": "HPS ditulis", "count": 3}

    import hps_engine
    monkeypatch.setattr(hps_engine, "scrape_hps_pl_ke_excel", fake_writer)

    logs = []
    result = pl_ui_helpers.update_hps_paket_pl("PK-1", Engine, logs.append)

    assert result["ok"] is True
    assert result["count"] == 3
    assert calls["kode"] == "PK-1"
    assert calls["writer"] == ("PK-1", str(workbook), logs.append)
    backup = pl_ui_helpers.pathlib.Path(result["backup_path"])
    assert backup.is_file()
    assert backup.read_bytes() == workbook.read_bytes()
    assert (tmp_path / "_PROMPT_AUDIT_PERUBAHAN_HPS_AGY.md").is_file()
    assert result["hps_prompt_path"].endswith("_PROMPT_AUDIT_PERUBAHAN_HPS_AGY.md")
    assert any(line.startswith("Backup HPS:") for line in logs)


def test_refresh_evaluasi_pl_only_runs_evaluation_macro_without_master_writer(
    tmp_path, monkeypatch
):
    workbook = tmp_path / "0. BAPLPK - Uji.xlsm"
    workbook.write_bytes(b"macro-workbook")
    calls = []

    class FakePythoncom:
        def CoUninitialize(self):
            calls.append("CoUninitialize")

    class FakePywintypes:
        com_error = RuntimeError

    class FakeWorkbook:
        ReadOnly = False

        def Save(self):
            calls.append("Save")

        def Close(self, SaveChanges):
            calls.append(("Close", SaveChanges))

    class FakeExcel:
        def __init__(self):
            self.Workbooks = self
            self.book = FakeWorkbook()

        def Open(self, *args, **kwargs):
            calls.append(("Open", args, kwargs))
            return self.book

        def Run(self, macro, *args):
            calls.append(("Run", macro, args))

        def Quit(self):
            calls.append("Quit")

    fake_excel = FakeExcel()
    monkeypatch.setattr(
        pl_ui_helpers,
        "_open_excel_for_pl_action",
        lambda: (FakePythoncom(), FakePywintypes(), fake_excel),
        raising=False,
    )

    class Engine:
        @staticmethod
        def _find_xlsm(kode_paket):
            return str(workbook)

    logs = []
    result = pl_ui_helpers.refresh_evaluasi_pl_only("PK-1", Engine, logs.append)

    assert result["ok"] is True
    assert result["workbook"] == str(workbook)
    assert [call[1] for call in calls if isinstance(call, tuple) and call[0] == "Run"] == [
        "ModDraftPaketPL.SetSilentPL",
        "ModDraftPaketPL.IsiEvaluasiPLStandalone",
        "ModDraftPaketPL.SetSilentPL",
    ]
    assert not any(
        isinstance(call, tuple)
        and call[0] == "Run"
        and "IsiDataPLByKode" in call[1]
        for call in calls
    )
    assert ("Close", False) in calls
    assert "Quit" in calls


def test_update_hps_paket_pl_shortens_backup_for_long_workbook_path(tmp_path, monkeypatch):
    filename = "0. BAPLJKK " + ("x" * 30) + ".xlsm"
    package_dir = tmp_path
    while len(str(package_dir / filename)) < 245:
        package_dir = package_dir / ("s" * 10)
    package_dir.mkdir(parents=True)
    workbook = package_dir / filename
    assert len(str(workbook)) > 240
    workbook.write_bytes(b"macro-workbook")

    class Engine:
        @staticmethod
        def _find_xlsm(kode_paket):
            return str(workbook)

    import hps_engine
    monkeypatch.setattr(
        hps_engine,
        "scrape_hps_pl_ke_excel",
        lambda kode_paket, excel_path, progress_cb=None: {
            "ok": True,
            "pesan": "HPS ditulis",
            "count": 1,
        },
    )

    result = pl_ui_helpers.update_hps_paket_pl("PK-LONG", Engine)

    assert result["ok"] is True
    assert len(result["backup_path"]) <= 240
    assert result["backup_path"] != str(workbook)
    assert pl_ui_helpers.pathlib.Path(result["backup_path"]).is_file()


def test_dokpil_resolver_uses_only_package_root(tmp_path):
    folder = tmp_path / "41. PLPK - Paket Fisik"
    folder.mkdir()
    root_pdf = folder / "dokpil_Paket Fisik.pdf"
    root_pdf.write_bytes(b"root")
    nested = folder / "2. Rancangan Kontrak"
    nested.mkdir()
    (nested / "dokpil_Paket Fisik - subfolder.pdf").write_bytes(b"nested")

    result = pl_ui_helpers._find_dokpil_pdf_root(str(folder))

    assert result["status"] == "found"
    assert result["path"] == str(root_pdf)


def test_dokpil_resolver_does_not_guess_ambiguous_root_files(tmp_path):
    folder = tmp_path / "61. PLJKK - Paket Konsultan"
    folder.mkdir()
    (folder / "dokpil_Paket Konsultan.pdf").write_bytes(b"one")
    (folder / "dokpil_Paket Konsultan v2.pdf").write_bytes(b"two")
    (folder / "_backup_dokpil_lama.pdf").write_bytes(b"backup")

    result = pl_ui_helpers._find_dokpil_pdf_root(str(folder))

    assert result["status"] == "ambiguous"
    assert len(result["candidates"]) == 2
    assert all("backup" not in path.casefold() for path in result["candidates"])


def test_sync_live_paket_umumkan_status_marks_and_throttles(monkeypatch):
    state = {}
    calls = []

    monkeypatch.setattr(pl_data_ui.time, "monotonic", lambda: 100.0)
    monkeypatch.setattr(pl_data_ui.st, "session_state", state)

    import pl_engine
    import spse_browser

    monkeypatch.setattr(spse_browser, "get_spse_cookies", lambda: "cookie")
    monkeypatch.setattr(
        pl_engine,
        "_fetch_tahap_spse",
        lambda *_args, **_kwargs: (calls.append(True) or {"74": "Upload Dokumen Penawaran"}),
    )

    first = pl_data_ui.sync_live_paket_umumkan_status("test_tayang")
    second = pl_data_ui.sync_live_paket_umumkan_status("test_tayang")

    assert first == {"ok": True, "cached": False, "count": 1}
    assert second == {"ok": True, "cached": True, "count": 0}
    assert len(calls) == 1
    assert state["test_tayang:tahap_map"] == {"74": "Upload Dokumen Penawaran"}
    assert pl_data_ui.is_paket_sudah_diumumkan({"kode_paket": "74", "status": "draft"}) is True


def test_sync_live_paket_umumkan_status_fails_closed_on_fetch_error(monkeypatch):
    state = {}

    monkeypatch.setattr(pl_data_ui.time, "monotonic", lambda: 100.0)
    monkeypatch.setattr(pl_data_ui.st, "session_state", state)

    import pl_engine
    import spse_browser

    monkeypatch.setattr(spse_browser, "get_spse_cookies", lambda: "cookie")
    monkeypatch.setattr(
        pl_engine,
        "_fetch_tahap_spse",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(TimeoutError("SPSE timeout")),
    )

    result = pl_data_ui.sync_live_paket_umumkan_status("test_tayang_error")

    assert result["ok"] is False
    assert "tahap_map" not in state
    assert filter_paket_draft_live(
        [{"kode_paket": "draft-local", "status": "draft", "tahap_spse": ""}],
        {},
        live_status_ok=result["ok"],
    ) == []


def test_verifikasi_package_label_uses_folder_number_and_name_only():
    row = {
        "kode_paket": "10999967000",
        "kode_unik": "KODE-INTERNAL-67",
        "nomor_urut": 67,
        "nama_paket": "Konsultan Pengawasan Paket 12",
    }

    assert pl_ui_helpers._pl_label(row) == "67. Konsultan Pengawasan Paket 12"
    assert pl_ui_helpers._pl_checkbox_label(row) == r"67\. Konsultan Pengawasan Paket 12"
    assert "10999967000" not in pl_ui_helpers._pl_checkbox_label(row)
    assert "KODE-INTERNAL-67" not in pl_ui_helpers._pl_checkbox_label(row)


def test_provider_verification_location_is_short_and_dpp_location_stays_detailed():
    from pl_kirimpesan_engine import TEMPAT_OPTIONS
    from verifikasi_penyedia_pl import TEMPAT_DEFAULT

    assert TEMPAT_DEFAULT == (
        "Kantor UKPBJ Kabupaten Tapin, Jl. Datu Suban RT. 01, "
        "Kelurahan Rangda Malingkung, Kecamatan Tapin Utara, Rantau, "
        "Kabupaten Tapin"
    )
    assert "Ruang Aula Rapat Lantai 2" not in TEMPAT_DEFAULT
    assert "Ruang Aula Rapat Lantai 2" in TEMPAT_OPTIONS["Kantor UKPBJ Kabupaten Tapin"]
