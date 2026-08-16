"""
Conflict detection engine — personil lintas paket.

Flow:
1. sync_from_supabase(kode_tender)  — ambil peserta_identitas, upsert ke paket_personil
2. sync_from_pdf(kode_tender, peserta_id, personel_list)  — dari kualifikasi_parser
3. get_konflik_personil()  — query: nama muncul di >1 kode_tender dari draft_paket
"""

import re
from datetime import date, timedelta
from config import sb as _sb

# ------------------------------------------------------------------
# Normalisasi nama personil
# ------------------------------------------------------------------

_RE_GELAR = re.compile(
    r"\b(Ir|Dr|H|Hj|S\.T|S\.E|S\.H|S\.Kom|A\.Md|M\.T|M\.Si|M\.M|Ph\.D|ST|AMd|SE|SH|MT)\.?\b",
    re.IGNORECASE,
)
_RE_NONALPHA = re.compile(r"[^a-z\s]")


_DEFAULT_PROVIDER_NAMES = {
    "",
    "-",
    "belum ada pemenang",
    "belum ada kontrak",
    "belum ada penyedia",
    "null",
    "none",
}
_EXCEL_ERROR_VALUES = {"#n/a", "#value!", "#ref!", "#div/0!", "#name?", "#null!", "#num!"}


def _normalize_nama(raw: str) -> str:
    """Lowercase + strip gelar Indonesia + hapus tanda baca + normalisasi spasi."""
    if not raw:
        return ""
    s = raw.lower()
    s = _RE_GELAR.sub("", s)
    s = _RE_NONALPHA.sub("", s)
    # Hapus spasi berlebih
    s = " ".join(s.split())
    return s


def _normalize_provider(raw: str) -> str:
    """Normalisasi nama penyedia untuk mencocokkan scraper, Excel, dan cache."""
    value = str(raw or "").strip().lower()
    if value in _DEFAULT_PROVIDER_NAMES:
        return ""
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def _excel_text(value) -> str:
    """Ubah error formula Excel menjadi kosong, bukan nama/provider."""
    text = str(value or "").strip()
    return "" if text.lower() in _EXCEL_ERROR_VALUES else text


def _is_true(value) -> bool:
    """Normalisasi flag boolean dari Supabase/DataTables."""
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "t", "yes", "y"}
    return bool(value)


def _get_winner_provider_map(kode_tenders: list[str]) -> dict[str, dict[str, str]]:
    """Ambil pemenang authoritative dari hasil scraper V20 di Supabase.

    ``tender_peserta.is_pemenang`` diprioritaskan. Kolom pemenang di ``tender``
    hanya fallback untuk paket lama yang belum memiliki flag peserta.
    """
    codes = {str(k).strip() for k in kode_tenders if str(k).strip()}
    result: dict[str, dict[str, str]] = {code: {} for code in codes}
    if not codes:
        return result

    try:
        peserta_rows = (
            _sb().table("tender_peserta")
            .select("kode_tender,nama_peserta,is_pemenang")
            .in_("kode_tender", list(codes))
            .execute().data or []
        )
    except Exception:
        peserta_rows = []

    for row in peserta_rows:
        if not _is_true(row.get("is_pemenang")):
            continue
        code = str(row.get("kode_tender") or "").strip()
        provider = str(row.get("nama_peserta") or "").strip()
        key = _normalize_provider(provider)
        if code in result and key:
            result[code].setdefault(key, provider)

    missing = [code for code, providers in result.items() if not providers]
    if not missing:
        return result

    try:
        tender_rows = (
            _sb().table("tender")
            .select("kode_tender,nama_pemenang,pemenang_berkontrak")
            .in_("kode_tender", missing)
            .execute().data or []
        )
    except Exception:
        tender_rows = []

    for row in tender_rows:
        code = str(row.get("kode_tender") or "").strip()
        if code not in result or result[code]:
            continue
        for field in ("nama_pemenang", "pemenang_berkontrak"):
            provider = str(row.get(field) or "").strip()
            key = _normalize_provider(provider)
            if key:
                result[code].setdefault(key, provider)

    return result


# ------------------------------------------------------------------
# Jadwal pelaksanaan paket
# ------------------------------------------------------------------

def _parse_hari_jangka(jangka_waktu: str) -> int | None:
    """Ambil angka pertama dari string jangka_waktu, asumsikan hari."""
    if not jangka_waktu:
        return None
    m = re.search(r"(\d+)", str(jangka_waktu))
    return int(m.group(1)) if m else None


def _get_tender_row(kode_tender: str) -> dict:
    """Ambil metadata paket Tender untuk resolver workbook/GCal."""
    try:
        rows = _sb().table("draft_paket") \
            .select("nama_tender,folder_dibuat,kode_pokja,data_snapshot") \
            .eq("kode_tender", kode_tender).limit(1).execute().data or []
        return rows[0] if rows else {}
    except Exception:
        return {}


def _iter_tender_workbooks(kode_tender: str, row: dict):
    """Yield workbook Tender kandidat tanpa pernah membuka/simpan Excel."""
    import os
    from config import TENDER_ROOT

    folders = []
    folder_raw = str(row.get("folder_dibuat") or "").strip()
    if folder_raw:
        folders.append(folder_raw if os.path.isabs(folder_raw)
                       else os.path.join(TENDER_ROOT, folder_raw))
    if not folders:
        fallback = _resolve_folder_paket(str(row.get("kode_pokja") or ""))
        if fallback:
            folders.append(fallback)

    seen = set()
    for folder in folders:
        try:
            for name in sorted(os.listdir(folder)):
                if not name.lower().endswith((".xlsm", ".xlsx")):
                    continue
                path = os.path.join(folder, name)
                if os.path.isfile(path) and path not in seen:
                    seen.add(path)
                    yield path
        except OSError:
            continue


def _read_tender_duration_excel(kode_tender: str, row: dict) -> int | None:
    """Baca masa pelaksanaan dari ``@ Master Data!C12`` secara read-only."""
    import re as _re
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    kode_digits = _re.sub(r"\D", "", str(kode_tender or ""))
    for path in _iter_tender_workbooks(kode_tender, row):
        wb = None
        try:
            wb = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_vba=str(path).lower().endswith(".xlsm"),
            )
            ws = wb["@ Master Data"]
            kode_excel = _re.sub(r"\D", "", str(ws["C4"].value or ""))
            if kode_digits and kode_excel and kode_digits != kode_excel:
                continue
            return _parse_hari_jangka(ws["C12"].value)
        except Exception:
            continue
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    # Fallback hanya untuk workbook lama yang belum bisa ditemukan di disk.
    snapshot = row.get("data_snapshot") or {}
    if isinstance(snapshot, dict):
        return _parse_hari_jangka(snapshot.get("r12") or snapshot.get("jangka_waktu"))
    return None


def _read_excel_personil(kode_tender: str, row: dict) -> dict:
    """Baca personil pemenang final dari ``0. Input BA`` secara read-only.

    Kolom G adalah kolom peserta terpilih pada template Tender. G7 berisi
    penyedia pemenang, sedangkan G13:G14 berisi personil yang sudah dipilih/
    dikoreksi manual. Workbook wajib lolos validasi ``@ Master Data!C4``.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return {"found": False, "provider": "", "personil": []}

    kode_digits = re.sub(r"\D", "", str(kode_tender or ""))
    if not kode_digits:
        return {"found": False, "provider": "", "personil": []}

    for path in _iter_tender_workbooks(kode_tender, row):
        wb = None
        try:
            wb = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_vba=str(path).lower().endswith(".xlsm"),
            )
            master = wb["@ Master Data"]
            kode_excel = re.sub(r"\D", "", str(master["C4"].value or ""))
            if kode_excel != kode_digits:
                continue

            ws = wb["0. Input BA"]
            provider = _excel_text(ws["G7"].value)
            personil = []
            for row_number in (13, 14):
                nama = _parse_nama(_excel_text(ws.cell(row_number, 7).value))
                if nama and _normalize_nama(nama):
                    personil.append(nama)
            return {"found": True, "provider": provider, "personil": personil}
        except Exception:
            continue
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    return {"found": False, "provider": "", "personil": []}


def _get_tgl_sppbj_gcal(nama_tender: str, kode_tender: str = ""):
    """Ambil tanggal mulai event SPPBJ dari Google Calendar."""
    if not nama_tender:
        return None
    try:
        from gcal_helper import get_tanggal_ba_dari_gcal
        return (get_tanggal_ba_dari_gcal(nama_tender, kode_paket=kode_tender) or {}).get("sppbj")
    except Exception:
        # GCal offline/token expired tidak boleh mematikan dashboard konflik.
        return None


def _get_jadwal_paket(kode_tender: str) -> tuple:
    """
    Return perkiraan (tgl_mulai, tgl_selesai) sebagai date.

    Tender: mulai = event SPPBJ di GCal, selesai = mulai + masa pelaksanaan
    dari workbook ``@ Master Data!C12``. PL tetap memakai metadata Supabase.
    """
    # Coba PL dulu
    rows_pl = _sb().table("draft_paket_pl")\
        .select("tgl_penetapan,jangka_waktu")\
        .eq("kode_paket", kode_tender)\
        .limit(1).execute().data or []
    if rows_pl:
        r = rows_pl[0]
        tgl_str = r.get("tgl_penetapan")
        hari = _parse_hari_jangka(r.get("jangka_waktu", ""))
        if tgl_str and hari:
            try:
                tgl_mulai = date.fromisoformat(tgl_str)
                tgl_selesai = tgl_mulai + timedelta(days=hari)
                return (tgl_mulai, tgl_selesai)
            except ValueError:
                pass
        return (None, None)

    row = _get_tender_row(kode_tender)
    tgl_mulai = _get_tgl_sppbj_gcal(
        str(row.get("nama_tender") or ""), kode_tender=kode_tender
    )
    hari = _read_tender_duration_excel(kode_tender, row)
    if tgl_mulai and hari:
        return (tgl_mulai, tgl_mulai + timedelta(days=hari))
    return (tgl_mulai, None)


def _parse_nama(raw: str) -> str:
    """Ambil nama saja dari string 'Nama (Jabatan)' atau 'Nama'."""
    if not raw:
        return ""
    m = re.match(r"^(.+?)\s*\(", raw.strip())
    return m.group(1).strip() if m else raw.strip()


def _parse_posisi(raw: str) -> str:
    """Ambil posisi/jabatan dari string 'Nama (Jabatan)'."""
    if not raw:
        return ""
    m = re.search(r"\((.+?)\)", raw.strip())
    return m.group(1).strip() if m else ""


def _upsert_personil_batch(rows: list[dict]) -> None:
    if not rows:
        return
    _sb().table("paket_personil").upsert(rows, on_conflict="kode_tender,peserta_id,nama_personil").execute()


def sync_from_supabase(kode_tender: str, log=print) -> dict:
    """
    Ambil semua peserta_identitas untuk kode_tender,
    upsert personil ke paket_personil.
    Sumber = 'supabase'.
    """
    rows_p = _sb().table("peserta_identitas")\
        .select("peserta_id,nama_perusahaan,personel_1,personel_2")\
        .eq("kode_tender", kode_tender).execute().data or []

    upsert_p = []
    for row in rows_p:
        pid   = row["peserta_id"]
        nama_penyedia = row.get("nama_perusahaan", "")
        for key in ("personel_1", "personel_2"):
            raw = row.get(key, "")
            if not raw:
                continue
            upsert_p.append({
                "kode_tender":  kode_tender,
                "peserta_id":   pid,
                "nama_penyedia": nama_penyedia,
                "nama_personil": _parse_nama(raw),
                "posisi":        key.replace("_", " ").title(),
                "sumber":        "supabase",
            })
    _upsert_personil_batch(upsert_p)
    log(f"sync_from_supabase {kode_tender}: {len(upsert_p)} personil")
    return {"personil": len(upsert_p)}


def sync_from_pdf(
    kode_tender: str,
    peserta_id: str,
    nama_penyedia: str,
    personel_list: list[str],
    log=print,
) -> dict:
    """
    Ganti data personil peserta dari hasil kualifikasi_parser (PDF).
    Sumber = 'pdf'. Penggantian mencegah personil stale tetap dianggap aktif.
    """
    if not str(peserta_id or "").strip():
        log(f"sync_from_pdf {kode_tender}: peserta_id kosong, skip personil")
        return {"personil": 0}
    if not personel_list:
        log(f"sync_from_pdf {kode_tender} {peserta_id}: personil kosong, data lama dipertahankan")
        return {"personil": 0}

    upsert_p = []
    seen = set()
    for i, raw in enumerate(personel_list):
        nama = _parse_nama(raw)
        norm = _normalize_nama(nama)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        upsert_p.append({
            "kode_tender":  kode_tender,
            "peserta_id":   peserta_id,
            "nama_penyedia": nama_penyedia,
            "nama_personil": nama,
            "posisi":        _parse_posisi(raw) or f"Personel {i+1}",
            "sumber":        "pdf",
        })

    if not upsert_p:
        log(f"sync_from_pdf {kode_tender} {peserta_id}: personil tidak valid, data lama dipertahankan")
        return {"personil": 0}

    _sb().table("paket_personil")\
        .delete()\
        .eq("kode_tender", kode_tender)\
        .eq("peserta_id", peserta_id)\
        .execute()
    _upsert_personil_batch(upsert_p)
    log(f"sync_from_pdf {kode_tender} {peserta_id}: {len(upsert_p)} personil")
    return {"personil": len(upsert_p)}


def _resolve_folder_paket(kode_pokja: str) -> str | None:
    """Cari folder paket di TENDER_ROOT yang mengandung 'Pokja {kode_pokja zero-padded}'."""
    from config import TENDER_ROOT
    if not kode_pokja:
        return None
    import os
    _pokja_digits = re.sub(r"\D", "", str(kode_pokja or ""))
    _pokja_base = _pokja_digits.lstrip("0") or "0"
    label_re = re.compile(
        rf"Pokja\s*[-]?\s*0*{re.escape(_pokja_base)}(?!\d)",
        re.IGNORECASE,
    )
    try:
        for d in os.listdir(TENDER_ROOT):
            full = os.path.join(TENDER_ROOT, d)
            if os.path.isdir(full) and label_re.search(d):
                return full
    except Exception:
        pass
    return None


def sync_from_doktek_folder(kode_tender: str, log=print) -> dict:
    """
    Fallback: peserta yang personel_1 masih kosong di peserta_identitas
    → cari DoktekFull_*.pdf di folder paket lokal → parse personil
    → ganti data paket_personil via sync_from_pdf().
    """
    import os
    import glob

    # Ambil peserta yang personil atau alat belum lengkap
    rows = _sb().table("peserta_identitas")\
        .select("peserta_id,nama_perusahaan")\
        .eq("kode_tender", kode_tender)\
        .execute().data or []

    synced_peserta = {
        r.get("peserta_id")
        for r in (_sb().table("paket_personil")
                  .select("peserta_id")
                  .eq("kode_tender", kode_tender)
                  .execute().data or [])
        if r.get("peserta_id")
    }
    kosong = [
        r for r in rows
        if r.get("peserta_id") and r["peserta_id"] not in synced_peserta
    ]
    if not kosong:
        log(f"sync_from_doktek_folder {kode_tender}: personil sudah lengkap, skip")
        return {"personil": 0}

    # Ambil kode_pokja dari draft_paket (satu paket, satu kode_pokja)
    dp_rows = _sb().table("draft_paket")\
        .select("kode_pokja")\
        .eq("kode_tender", kode_tender)\
        .limit(1).execute().data or []
    kode_pokja = (dp_rows[0].get("kode_pokja") or "") if dp_rows else ""

    folder_paket = _resolve_folder_paket(kode_pokja)
    if not folder_paket:
        log(f"sync_from_doktek_folder {kode_tender}: folder paket Pokja {kode_pokja} tidak ditemukan")
        return {"personil": 0}

    total_p = 0
    for row in kosong:
        pid = row["peserta_id"]
        nama_penyedia = row.get("nama_perusahaan", "")

        # Cari DoktekFull_*.pdf: coba subfolder "1. Dokumen Penawaran\*nama*" dulu, lalu glob recursive
        pdf_path = None
        # Subfolder dokumen penawaran — coba match nama perusahaan (substring case-insensitive)
        dok_dir = os.path.join(folder_paket, "1. Dokumen Penawaran")
        if os.path.isdir(dok_dir):
            nama_lower = nama_penyedia.lower()
            for sub in os.listdir(dok_dir):
                sub_full = os.path.join(dok_dir, sub)
                if os.path.isdir(sub_full) and nama_lower[:6] in sub.lower():
                    # Cari DoktekFull_*.pdf di subfolder ini
                    hits = glob.glob(os.path.join(sub_full, "*DoktekFull_*.pdf"))
                    if hits:
                        pdf_path = hits[0]
                        break

        # Fallback: glob recursive seluruh folder paket
        if not pdf_path:
            hits = glob.glob(os.path.join(folder_paket, "**", "*DoktekFull_*.pdf"), recursive=True)
            # Jika banyak, coba match nama perusahaan via path
            if len(hits) > 1:
                nama_lower = nama_penyedia.lower()
                matched = [h for h in hits if nama_lower[:6] in h.lower()]
                hits = matched if matched else hits
            if hits:
                pdf_path = hits[0]

        if not pdf_path:
            log(f"  [{nama_penyedia}] DoktekFull_*.pdf tidak ditemukan, skip")
            continue

        log(f"  [{nama_penyedia}] parse dari {os.path.basename(pdf_path)}")
        try:
            # Lazy import hindari circular
            from dokumen_teknis_engine import parse_personel
            personel_list  = parse_personel(pdf_path)
        except Exception as e:
            log(f"  [{nama_penyedia}] parse error: {e}")
            continue

        res = sync_from_pdf(kode_tender, pid, nama_penyedia, personel_list, log=log)
        total_p += res["personil"]

    log(f"sync_from_doktek_folder {kode_tender}: {total_p} personil (dari {len(kosong)} peserta kosong)")
    return {"personil": total_p}


def _get_aktif_kode_tender() -> list[str]:
    """Semua kode_tender di draft_paket = paket aktif."""
    rows = _sb().table("draft_paket").select("kode_tender").execute().data or []
    return [
        r["kode_tender"] for r in rows
        if r.get("kode_tender") and str(r["kode_tender"]) != "10096884000"
    ]


def _get_kode_tender_tahun_berjalan(tahun: int | None = None) -> set[str]:
    """Kode paket tahun berjalan berdasarkan tahun pada nomor surat dinas."""
    target = int(tahun or date.today().year)
    rows = _sb().table("draft_paket").select(
        "kode_tender,nomor_surat_dinas"
    ).execute().data or []
    kode = set()
    for row in rows:
        kt = str(row.get("kode_tender") or "").strip()
        if not kt or kt.startswith("_err_") or kt == "10096884000":
            continue
        tahun_surat = re.findall(
            r"(?<!\d)(20\d{2})(?!\d)",
            str(row.get("nomor_surat_dinas") or ""),
        )
        if tahun_surat and int(tahun_surat[-1]) == target:
            kode.add(kt)
    return kode


def sync_new_paket(log=print) -> dict:
    """
    Hanya sync paket yang belum ada di paket_personil sama sekali.
    Jauh lebih cepat dari loop semua paket aktif.
    """
    aktif = set(_get_aktif_kode_tender())
    if not aktif:
        return {"synced": 0}

    # Paket dianggap tersync jika data personilnya sudah tersedia.
    tersync_p = set(
        r["kode_tender"]
        for r in (_sb().table("paket_personil").select("kode_tender").execute().data or [])
    )

    belum = aktif - tersync_p
    if not belum:
        log("sync_new_paket: semua paket sudah tersync, skip")
        return {"synced": 0}

    log(f"sync_new_paket: {len(belum)} paket belum tersync → {sorted(belum)}")
    total = 0
    for kt in belum:
        r = sync_from_doktek_folder(kt, log=log)
        total += r.get("personil", 0)
    return {"synced": len(belum), "personil": total}


def get_sync_coverage(tahun: int | None = None) -> dict:
    """Coverage personil paket tahun berjalan tanpa membaca PDF."""
    target = int(tahun or date.today().year)
    aktif = _get_kode_tender_tahun_berjalan(target)
    p = {
        r["kode_tender"]
        for r in (_sb().table("paket_personil").select("kode_tender").execute().data or [])
    }
    return {
        "tahun": target,
        "aktif": len(aktif),
        "personil": len(aktif & p),
        "lengkap": len(aktif & p),
        "belum_lengkap": len(aktif - p),
    }


def _overlap(mulai_a, selesai_a, mulai_b, selesai_b) -> bool:
    """True jika dua range tanggal overlap, atau salah satu None (konservatif)."""
    if mulai_a is None or selesai_a is None or mulai_b is None or selesai_b is None:
        return True  # data tidak lengkap = asumsi overlap
    return max(mulai_a, mulai_b) <= min(selesai_a, selesai_b)


def get_konflik_personil(kode_tender_target: str | None = None) -> list[dict]:
    """
    Return personil pemenang yang muncul di >1 paket aktif dengan jadwal overlap.
    Jika kode_tender_target diisi, filter hanya konflik yang melibatkan paket itu.

    Sumber personil utama = workbook ``0. Input BA`` kolom G (hasil koreksi
    manual). ``paket_personil`` hanya fallback ketika workbook paket belum
    ditemukan; baris fallback tetap dibatasi pada penyedia pemenang scraper.

    Return: [{
        "nama_personil": str (normalized),
        "nama_personil_display": str (nama asli pertama),
        "paket": [{"kode_tender":..., "nama_penyedia":..., "peserta_id":...,
                   "tgl_mulai": date|None, "tgl_selesai": date|None}]
    }]
    """
    aktif = _get_aktif_kode_tender()
    if not aktif:
        return []

    winner_map = _get_winner_provider_map(aktif)
    if not any(winner_map.values()):
        return []

    rows = _sb().table("paket_personil")\
        .select("kode_tender,peserta_id,nama_penyedia,nama_personil")\
        .in_("kode_tender", aktif).execute().data or []

    # Satu kali lookup raw per paket. Workbook dibaca setiap render agar edit
    # manual user langsung tercermin tanpa cache stale.
    personil_by_package: dict[str, list[dict]] = {}
    raw_by_package: dict[str, list[dict]] = {}
    for row in rows:
        code = str(row.get("kode_tender") or "").strip()
        if code:
            raw_by_package.setdefault(code, []).append(row)

    for code in aktif:
        winners = winner_map.get(str(code), {})
        if not winners:
            continue

        workbook = _read_excel_personil(str(code), _get_tender_row(str(code)))
        if workbook["found"]:
            provider = str(workbook.get("provider") or "").strip()
            provider_key = _normalize_provider(provider)
            # Workbook valid tetapi provider kosong/tidak sama dengan hasil
            # scraper = jangan membuat false positive dari workbook yang salah.
            if provider_key not in winners:
                continue
            personil_by_package[str(code)] = [
                {
                    "kode_tender": str(code),
                    "peserta_id": None,
                    "nama_penyedia": provider or winners[provider_key],
                    "nama_personil": nama,
                }
                for nama in workbook.get("personil", [])
            ]
            continue

        # Fallback legacy: hanya cache personil milik provider pemenang.
        personil_by_package[str(code)] = [
            row for row in raw_by_package.get(str(code), [])
            if _normalize_provider(row.get("nama_penyedia")) in winners
        ]

    # Cache jadwal per kode_tender (lazy)
    _jadwal_cache: dict[str, tuple] = {}

    def _jadwal(kt: str) -> tuple:
        if kt not in _jadwal_cache:
            _jadwal_cache[kt] = _get_jadwal_paket(kt)
        return _jadwal_cache[kt]

    # Group by nama_personil NORMALIZED, simpan display name pertama
    from collections import defaultdict
    grouped: dict[str, list] = defaultdict(list)
    display_map: dict[str, str] = {}  # normalized → nama asli pertama
    for r in (
        entry
        for package_rows in personil_by_package.values()
        for entry in package_rows
    ):
        nama_raw = r["nama_personil"] or ""
        norm = _normalize_nama(nama_raw)
        if not norm:
            continue
        if norm not in display_map:
            display_map[norm] = nama_raw
        # Tambah info jadwal ke entry
        mulai, selesai = _jadwal(r["kode_tender"])
        grouped[norm].append({
            "kode_tender": r["kode_tender"],
            "nama_penyedia": r["nama_penyedia"],
            "peserta_id": r["peserta_id"],
            "tgl_mulai": mulai,
            "tgl_selesai": selesai,
        })

    konflik = []
    for norm, entries in grouped.items():
        # Deduplikasi per kode_tender (ambil entry pertama per kode)
        seen_kt: dict[str, dict] = {}
        for e in entries:
            if e["kode_tender"] not in seen_kt:
                seen_kt[e["kode_tender"]] = e
        unique_entries = list(seen_kt.values())
        if len(unique_entries) <= 1:
            continue
        if kode_tender_target and kode_tender_target not in seen_kt:
            continue

        # Cek overlap jadwal antar semua kombinasi paket
        ada_overlap = False
        kt_list = unique_entries
        for i in range(len(kt_list)):
            for j in range(i + 1, len(kt_list)):
                ea, eb = kt_list[i], kt_list[j]
                if _overlap(ea["tgl_mulai"], ea["tgl_selesai"], eb["tgl_mulai"], eb["tgl_selesai"]):
                    ada_overlap = True
                    break
            if ada_overlap:
                break

        if not ada_overlap:
            continue

        # Satu baris per paket; variasi posisi/nama lama dalam paket yang sama
        # tidak boleh membuat konflik tampil berulang.
        konflik.append({
            "nama_personil": norm,
            "nama_personil_display": display_map[norm],
            "paket": unique_entries,
        })

    return konflik
