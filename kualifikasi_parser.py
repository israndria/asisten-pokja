"""
Parser data kualifikasi peserta tender.

Sumber data:
1. HTML /kualifikasi/{id}/preview  — requests + BeautifulSoup (data utama)
2. DOM Playwright                  — KSWP status (img logo verified, JS-rendered)
3. PDF via OCR (pytesseract+fitz)  — KSWP validity text + Nilai Kinerja
4. pdfplumber                      — Formulir Isian (SKP double-check)
"""

import os
import re
import io
import glob as _glob

import requests
from bs4 import BeautifulSoup

import spse_browser
from config import SPSE_BASE_URL
from person_name_utils import clean_person_name, format_equipment_entry

# ── Tesseract ──────────────────────────────────────────────────────────────────
_TESSERACT_CANDIDATES = [
    r"D:\Tesseract OCR\tesseract.exe",
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Tesseract-OCR\tesseract.exe",
]
_TESSERACT_PATH = next((p for p in _TESSERACT_CANDIDATES if os.path.exists(p)), _TESSERACT_CANDIDATES[0])

def _ocr_image(img):
    """OCR PIL Image → string. Lazy import pytesseract."""
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = _TESSERACT_PATH
    return pytesseract.image_to_string(img, lang="ind+eng")


def _pdf_to_text(pdf_path: str) -> str:
    """
    Extract teks dari PDF.
    Kalau halaman berbasis gambar → OCR via pytesseract.
    """
    import fitz
    from PIL import Image, ImageEnhance

    doc = fitz.open(pdf_path)
    parts = []
    for page in doc:
        text = page.get_text().strip()
        if text:
            parts.append(text)
            continue
        # Gambar — render resolusi tinggi lalu OCR
        mat = fitz.Matrix(2.5, 2.5)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        parts.append(_ocr_image(img))
    return "\n".join(parts)


def _pdf_footer_text(pdf_path: str) -> str:
    """
    OCR 25% bawah halaman pertama PDF (area nilai kinerja).
    Dengan enhance kontras agar teks kecil terbaca.
    """
    import fitz
    from PIL import Image, ImageEnhance

    doc = fitz.open(pdf_path)
    page = doc[0]
    imgs = page.get_images(full=True)
    if imgs:
        # Ambil gambar konten terbesar (bukan watermark)
        largest = max(imgs, key=lambda x: doc.extract_image(x[0])["image"].__len__())
        raw = doc.extract_image(largest[0])["image"]
        img = Image.open(io.BytesIO(raw))
    else:
        mat = fitz.Matrix(3, 3)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes("png")))

    w, h = img.size
    footer = img.crop((0, int(h * 0.75), w, h))
    footer = ImageEnhance.Contrast(footer).enhance(2.0)
    footer = footer.resize((footer.width * 2, footer.height * 2), Image.LANCZOS)
    return _ocr_image(footer)


# ── Helpers HTTP ──────────────────────────────────────────────────────────────

def _headers():
    cookie = spse_browser.get_spse_cookies()
    return {"Cookie": cookie, "User-Agent": "Mozilla/5.0", "Referer": SPSE_BASE_URL}


def _normalize_npwp_source(value: str) -> str:
    """Validate NPWP while preserving representation supplied by SPSE."""
    text = re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()
    if not text or text in {"-", "—", "N/A", "NA"}:
        return ""
    if not re.fullmatch(r"[0-9.\- ]+", text):
        return ""
    digits = re.sub(r"[^0-9]", "", text)
    return text if len(digits) in {15, 16} else ""


# ── Cari file di folder peserta ────────────────────────────────────────────────

def _find_file(folder: str, *patterns: str) -> str | None:
    """Cari file pertama yang cocok salah satu pattern (case-insensitive) di folder."""
    try:
        files = os.listdir(folder)
    except Exception:
        return None
    for pat in patterns:
        pat_lower = pat.lower()
        for f in files:
            if pat_lower in f.lower():
                return os.path.join(folder, f)
    return None


def _normalisasi_spasi(value: str) -> str:
    """Rapatkan whitespace hasil HTML/OCR agar regex stabil."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _parse_sbu_requirement_keywords(value: str) -> set[str]:
    """Ambil kode subklasifikasi, KBLI, dan nama SBU dari requirement."""
    text = _normalisasi_spasi(value).upper()
    if not text:
        return set()

    keywords = set(re.findall(r"\b[A-Z]{2}\d{3}\b", text))
    keywords.update(
        match.group(1)
        for match in re.finditer(r"\bKBLI(?:\s+20\d{2})?\s*[:\-]?\s*(\d{5})\b", text)
    )

    # Fallback nama untuk requirement lama yang tidak memuat KBLI numerik.
    name = re.search(
        r"\b[A-Z]{2}\d{3}\b(?:\s*\([^)]*\))?\s+(.+?)(?:\s+KBLI\b|$)",
        text,
    )
    if name:
        phrase = _normalisasi_spasi(name.group(1))
        keywords.add(phrase)
        words = phrase.split()
        if len(words) >= 2:
            keywords.add(" ".join(words[-2:]))
        if words:
            keywords.add(words[-1])
    return keywords


def _find_nearest_xlsm(folder: str) -> str | None:
    """Cari workbook paket pada folder peserta atau ancestor-nya."""
    current = os.path.abspath(folder or "")
    while current and os.path.isdir(current):
        try:
            for name in sorted(os.listdir(current)):
                if name.lower().endswith(".xlsm"):
                    return os.path.join(current, name)
        except Exception:
            pass
        parent = os.path.dirname(current)
        if parent == current:
            break
        current = parent
    return None


def _ambil_requirement_lokal(folder_peserta: str) -> str:
    """Baca requirement SBU dari list_dokpil workbook paket bila tersedia."""
    workbook = _find_nearest_xlsm(folder_peserta)
    if not workbook:
        return ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(workbook, read_only=True, data_only=True, keep_vba=True)
        try:
            ws = wb["list_dokpil"]
            # AQ = SBU 1, AS = SBU gabungan. Satu workbook = satu paket.
            for row in range(2, min(ws.max_row or 2, 30) + 1):
                for col in (43, 45):
                    value = ws.cell(row, col).value
                    if value and "SBU" in str(value).upper():
                        return _normalisasi_spasi(str(value))
        finally:
            wb.close()
    except Exception:
        return ""
    return ""


# ── Parse halaman HTML /preview ───────────────────────────────────────────────

def _ambil_syarat_sbu_keywords(
    kode_tender: str,
    folder_peserta: str | None = None,
) -> list[str] | None:
    """Ambil token syarat SBU dengan prioritas workbook Dokpil lokal.

    ``draft_paket`` dapat tertinggal dari Dokpil yang sudah dipakai paket.
    Karena itu workbook lokal menjadi sumber pertama; Supabase hanya fallback.
    """
    local_requirement = _ambil_requirement_lokal(folder_peserta or "")
    if local_requirement:
        keywords = _parse_sbu_requirement_keywords(local_requirement)
        if keywords:
            return sorted(keywords)

    try:
        from config import sb as _sb_fn
        row = _sb_fn().table("draft_paket").select("sbu_baru,sbu_lama") \
                      .eq("kode_tender", kode_tender).maybe_single().execute()
        if not row or not row.data:
            return None
        keywords: set[str] = set()
        for field in ("sbu_baru", "sbu_lama"):
            val = (row.data.get(field) or "").strip()
            if not val:
                continue
            keywords.update(_parse_sbu_requirement_keywords(val))
        return list(keywords) if keywords else None
    except Exception:
        return None


def parse_preview_html(kualifikasi_id: str, syarat_keywords: list[str] | None = None) -> dict:
    """
    Scrape /kualifikasi/{id}/preview via requests.
    Return dict dengan semua field yang bisa diambil dari HTML.
    """
    base = SPSE_BASE_URL.rstrip("/")
    url = f"{base}/kualifikasi/{kualifikasi_id}/preview"
    try:
        r = requests.get(url, headers=_headers(), timeout=15)
        if r.status_code != 200:
            return {"ok": False, "pesan": f"HTTP {r.status_code}"}
        r.encoding = r.apparent_encoding or "utf-8"
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        return {"ok": False, "pesan": str(e)}

    tables = soup.find_all("table")
    hasil = {"ok": True}

    def _tbl(idx) -> list[list[str]]:
        if idx >= len(tables):
            return []
        return [
            [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            for tr in tables[idx].find_all("tr")
            if tr.find_all(["td", "th"])
        ]

    def _cell(rows, label_contains, col=1, default=""):
        for row in rows:
            if row and label_contains.lower() in row[0].lower():
                return row[col] if len(row) > col else default
        return default

    # Tabel 0: IDENTITAS
    t0 = _tbl(0)
    hasil["nama"] = _cell(t0, "Nama")
    hasil["npwp"] = _normalize_npwp_source(_cell(t0, "NPWP"))
    hasil["alamat"] = _cell(t0, "Alamat")
    hasil["email"] = _cell(t0, "Email")

    # Tabel 1: IZIN USAHA (NIB, SS, SBU)
    t1 = _tbl(1)
    nib_row = next((r for r in t1 if r and ("Nomor Induk Berusaha" in r[0] or "NIB" == r[0].strip())), None)

    def _pick_izin(jenis_str, keywords):
        """
        Pilih baris izin usaha yang cocok keyword syarat paket.
        Kalau keyword None atau tidak ada match → fallback ke baris pertama.
        Return (row, tidak_sesuai_flag).
        """
        rows = [r for r in t1 if r and jenis_str in r[0]]
        if not rows:
            return None, False
        if keywords:
            for r in rows:
                klas = (r[5] if len(r) > 5 else "").upper()
                if any(kw in klas for kw in keywords):
                    return r, False   # match sesuai syarat
            # Tidak ada match → fallback pertama, tandai tidak_sesuai
            return rows[0], True
        return rows[0], False         # perilaku lama

    ss_row,  ss_tidak_sesuai  = _pick_izin("Sertifikat Standar",    syarat_keywords)
    sbu_row, sbu_tidak_sesuai = _pick_izin("Sertifikat Badan Usaha", syarat_keywords)

    if nib_row and len(nib_row) >= 3:
        hasil["nib_nomor"]  = nib_row[1]
        hasil["nib_berlaku"] = nib_row[2]

    if ss_row and len(ss_row) >= 4:
        hasil["ss_nomor"]       = ss_row[1]
        hasil["ss_berlaku"]     = ss_row[2]
        hasil["ss_instansi"]    = ss_row[3] if len(ss_row) > 3 else ""
        hasil["ss_kualifikasi"] = ss_row[4] if len(ss_row) > 4 else ""
        hasil["ss_klasifikasi"] = ss_row[5] if len(ss_row) > 5 else ""

    if sbu_row and len(sbu_row) >= 4:
        hasil["sbu_nomor"]       = sbu_row[1]   # PBUMKU
        hasil["sbu_berlaku"]     = sbu_row[2]
        hasil["sbu_instansi"]    = sbu_row[3] if len(sbu_row) > 3 else ""
        hasil["sbu_kualifikasi"] = sbu_row[4] if len(sbu_row) > 4 else ""
        sbu_klas = sbu_row[5] if len(sbu_row) > 5 else ""
        hasil["sbu_klasifikasi"] = sbu_klas
        # Bentuk label ringkas: "F42911 - KONSTRUKSI BANGUNAN..." → ambil kode + nama pendek
        if sbu_klas:
            kode_sbu = sbu_klas.split(" - ")[0].strip() if " - " in sbu_klas else sbu_klas
            nama_sbu = sbu_klas.split(" - ", 1)[1].strip()[:80] if " - " in sbu_klas else ""
            hasil["sbu_subklas_label"] = f"{kode_sbu} - {nama_sbu}" if nama_sbu else kode_sbu
        else:
            hasil["sbu_subklas_label"] = ""

    # Flag kesesuaian syarat paket
    hasil["sbu_tidak_sesuai"] = sbu_tidak_sesuai
    hasil["ss_tidak_sesuai"]  = ss_tidak_sesuai

    # Tabel 2: AKTA
    t2 = _tbl(2)
    pendirian_idx = next((i for i, r in enumerate(t2) if r and "Akta Pendirian" in r[0]), None)
    perubahan_idx = next((i for i, r in enumerate(t2) if r and "Akta Perubahan" in r[0]), None)

    def _akta_block(rows, start_idx):
        data = {}
        if start_idx is None:
            return data
        for row in rows[start_idx + 1:]:
            if not row:
                continue
            label = row[0].lower()
            val   = row[1] if len(row) > 1 else ""
            if "nomor" in label:
                data["nomor"] = val
            elif "tanggal" in label:
                data["tanggal"] = val
            elif "notaris" in label:
                data["notaris"] = val
            elif "akta" in label:  # header baru = berhenti
                break
        return data

    hasil["akta_pendirian"] = _akta_block(t2, pendirian_idx)
    hasil["akta_perubahan"] = _akta_block(t2, perubahan_idx)

    # Tabel 3: MANAJERIAL (pemilik/direktur)
    t3 = _tbl(3)
    pemilik_list = []
    for row in t3[1:]:  # skip header
        if row and len(row) >= 1 and row[0] and row[0] != "No data available in table":
            pemilik_list.append(row[0])
    hasil["pemilik"] = pemilik_list

    # Tabel PERSONEL MANAJERIAL — id: "table-tenaga-ahli"
    # Kolom terverifikasi: Nama | Tanggal Lahir | NPWP | Pendidikan | Pengalaman Kerja | Profesi/Keahlian
    # col[0]=Nama, col[5]=Profesi/Keahlian (jabatan)
    _tbl_personel = soup.find("table", id="table-tenaga-ahli")
    if _tbl_personel:
        _rows_per_orang = [
            [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            for tr in _tbl_personel.find_all("tr")
            if tr.find_all(["td", "th"])
        ]
    else:
        _rows_per_orang = []

    personel_list = []
    for row in _rows_per_orang[1:]:
        if not row or row[0] == "No data available in table":
            continue
        nama_p  = row[0].strip()
        jabatan = row[5].strip() if len(row) > 5 else ""
        if not nama_p or nama_p.lower() in ("nama", "no", "no."):
            continue
        nama_p = clean_person_name(nama_p)
        if not nama_p:
            continue
        personel_list.append(nama_p)
    hasil["personel_list"] = personel_list

    # Tabel 5: PENGALAMAN KERJA
    t5 = _tbl(5)
    pengalaman = []
    for row in t5[1:]:
        if not row or row[0] == "No data available in table":
            continue
        pengalaman.append({
            "nama":      row[0] if len(row) > 0 else "",
            "lokasi":    row[1] if len(row) > 1 else "",
            "instansi":  row[2] if len(row) > 2 else "",
            "tgl_mulai": row[4] if len(row) > 4 else "",
            "tgl_selesai": row[5] if len(row) > 5 else "",
            "nilai":     row[6] if len(row) > 6 else "",
            "nomor":     row[7] if len(row) > 7 else "",
        })
    hasil["pengalaman"] = pengalaman

    # Tabel PERALATAN UTAMA — id: "table-peralatan"
    # Kolom terverifikasi: Nama Alat | Jumlah | Kapasitas | Merk/Tipe | Tahun | Kondisi | Lokasi | Status | Bukti
    # col[0]=Nama Alat, col[1]=Jumlah
    _tbl_peralatan = soup.find("table", id="table-peralatan")
    if _tbl_peralatan:
        _rows_alat = [
            [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            for tr in _tbl_peralatan.find_all("tr")
            if tr.find_all(["td", "th"])
        ]
    else:
        _rows_alat = []

    peralatan_list = []
    for row in _rows_alat[1:]:
        if not row or row[0] == "No data available in table":
            continue
        nama_alat = row[0].strip()
        jumlah    = row[1].strip() if len(row) > 1 else ""
        if not nama_alat or nama_alat.lower() in ("nama alat", "peralatan", "no", "no."):
            continue
        peralatan_list.append(format_equipment_entry(nama_alat, jumlah))
    hasil["peralatan_list"] = peralatan_list

    # Tabel 6: PEKERJAAN SEDANG BERJALAN → hitung JP → SKP
    t6 = _tbl(6)
    jp_berjalan = [
        r for r in t6[1:]
        if r and r[0] and r[0] != "No data available in table"
    ]
    hasil["pekerjaan_berjalan"] = jp_berjalan
    hasil["jp_preview"] = len(jp_berjalan)
    hasil["skp_preview"] = 5 - len(jp_berjalan)

    return hasil


_BULAN_ID = (
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
)


def _tanggal_id_from_iso(value: str) -> str:
    match = re.fullmatch(r"(\d{4})[-/]([01]\d)[-/]([0-3]\d)", value or "")
    if not match:
        return value
    year, month, day = match.groups()
    month_num = int(month)
    return f"{int(day)} {_BULAN_ID[month_num - 1]} {year}" if 1 <= month_num <= 12 else value


def _pdf_candidates(folder: str, kind: str) -> list[str]:
    try:
        names = sorted(os.listdir(folder))
    except Exception:
        return []
    result = []
    for name in names:
        lower = name.lower()
        if not lower.endswith(".pdf"):
            continue
        if kind == "sbu" and ("sbu" in lower or "sertifikat badan usaha" in lower):
            result.append(os.path.join(folder, name))
        elif kind == "ss" and (
            lower.startswith("ss ")
            or "sertifikat standar" in lower
            or "sertifikat_standar" in lower
            or "serifikat standar" in lower  # typo umum pada nama upload
        ):
            result.append(os.path.join(folder, name))
    return result


def _parse_local_sbu_pdf(folder: str) -> dict:
    """Ambil identitas SBU dari PDF peserta; return kosong bila tidak terbaca."""
    for pdf in _pdf_candidates(folder, "sbu"):
        try:
            text = _normalisasi_spasi(_pdf_to_text(pdf))
        except Exception:
            continue
        upper = text.upper()
        filename_upper = os.path.basename(pdf).upper()
        metadata_text = f"{upper} {filename_upper}"
        if "SERTIFIKAT BADAN USAHA" not in upper and "PB-UMKU" not in upper:
            continue

        number = ""
        match = re.search(r"\bPB[- ]?UMKU\s*:\s*(\d{10,})", upper)
        if match:
            number = match.group(1)
        if not number:
            match = re.search(r"\bNOMOR\s*:\s*(\d{12,})", upper)
            if match:
                number = match.group(1)

        code_match = re.search(r"\b([A-Z]{2}\d{3})\b", metadata_text)
        kbli_match = re.search(
            r"\bKBLI\b[^\d]{0,20}(4\d{4})(?:\s*-\s*|\s+)",
            metadata_text,
        )
        label = ""
        if kbli_match:
            label_match = re.search(
                rf"\b{kbli_match.group(1)}\s*-\s*(.+?)(?=\s+(?:\d+\.\s+)?LOKASI|\s+TELAH\s+MEMENUHI|\s+DITERBITKAN|\s+MASA\s+BERLAKU|$)",
                upper,
            )
            if label_match:
                label = _normalisasi_spasi(label_match.group(1)).title()
                label = re.sub(r"\s+Indonesia\s+\(Kbli\)$", "", label, flags=re.IGNORECASE)
        if not label and code_match:
            known = {"BG002": "Konstruksi Gedung Perkantoran"}
            label = known.get(code_match.group(1), "")
        if not code_match and kbli_match and kbli_match.group(1) == "41012":
            code = "BG002"
        else:
            code = code_match.group(1) if code_match else ""

        validity_match = re.search(
            r"MASA\s+BERLAKU\s*S\.?D\.?\s*:\s*(\d{4}[-/]\d{2}[-/]\d{2})",
            upper,
        )
        validity = _tanggal_id_from_iso(validity_match.group(1)) if validity_match else ""
        qualification = "Kecil" if re.search(r"\bKECIL\b", upper) else ""
        if number or code or kbli_match:
            kbli = kbli_match.group(1) if kbli_match else ""
            display = f"{code} - {label}" if code and label else (code or label)
            return {
                "sbu_nomor": number,
                "sbu_berlaku": validity,
                "sbu_kualifikasi": qualification,
                "sbu_klasifikasi": f"{display}|KBLI {kbli}" if display and kbli else display,
                "sbu_subklas_label": display,
                "sbu_kbli": kbli,
                "sbu_source": os.path.basename(pdf),
            }
    return {}


def _parse_local_ss_pdf(folder: str) -> dict:
    """Ambil nomor/status Sertifikat Standar dari PDF peserta."""
    for pdf in _pdf_candidates(folder, "ss"):
        try:
            text = _normalisasi_spasi(_pdf_to_text(pdf))
        except Exception:
            continue
        upper = text.upper()
        if "SERTIFIKAT STANDAR" not in upper:
            continue
        match = re.search(r"SERTIFIKAT\s+STANDAR\s*:\s*(\d{10,})", upper)
        number = match.group(1) if match else ""
        if not number:
            continue
        if re.search(r"\bBELUM\s+TERVERIFIKASI\b", upper):
            status = "Belum Terverifikasi"
        elif "TELAH TERVERIFIKASI" in upper or "TELAH MEMENUHI PERSYARATAN" in upper:
            status = "Terverifikasi"
        else:
            status = ""
        return {
            "ss_nomor": number,
            "ss_status": status,
            "ss_source": os.path.basename(pdf),
        }
    return {}


def _merge_local_license_evidence(
    html_data: dict,
    folder_peserta: str,
    syarat_keywords: list[str] | None,
    log_cb=None,
) -> dict:
    """Gunakan bukti PDF lokal sebagai koreksi data preview yang stale/salah."""
    local_sbu = _parse_local_sbu_pdf(folder_peserta)
    local_ss = _parse_local_ss_pdf(folder_peserta)
    conflicts = []

    def _apply(field, value, label):
        if not value:
            return
        before = html_data.get(field)
        if before and _normalisasi_spasi(str(before)).upper() != _normalisasi_spasi(str(value)).upper():
            conflicts.append(f"{label}: {before} -> {value}")
        html_data[field] = value

    for field in (
        "sbu_nomor", "sbu_berlaku", "sbu_kualifikasi", "sbu_klasifikasi",
        "sbu_subklas_label", "sbu_kbli",
    ):
        _apply(field, local_sbu.get(field), field)
    _apply("ss_nomor", local_ss.get("ss_nomor"), "ss_nomor")

    if local_ss.get("ss_status"):
        html_data["ss_status"] = local_ss["ss_status"]
        html_data["ss_instansi"] = "OSS"
    if local_sbu.get("sbu_subklas_label"):
        evidence = " ".join(
            str(local_sbu.get(key) or "")
            for key in ("sbu_subklas_label", "sbu_kbli")
        ).upper()
        matches_requirement = not syarat_keywords or any(
            keyword.upper() in evidence for keyword in syarat_keywords
        )
        html_data["sbu_tidak_sesuai"] = not matches_requirement
    if local_ss.get("ss_nomor") and local_ss.get("ss_status"):
        html_data["ss_tidak_sesuai"] = False
    if conflicts and log_cb:
        log_cb(
            "  ⚠️ Preview SPSE berbeda dengan PDF izin lokal; "
            f"pakai PDF: {'; '.join(conflicts[:4])}"
        )
    return html_data


# ── KSWP: DOM Playwright (fallback primer) ─────────────────────────────────────

def get_kswp_from_dom(kualifikasi_id: str) -> str:
    """
    Baca status KSWP dari DOM /preview yang sudah di-render JS.
    Return: 'VALID' | 'TIDAK VALID' | 'TIDAK DIKETAHUI'
    """
    base = SPSE_BASE_URL.rstrip("/")
    url  = f"{base}/kualifikasi/{kualifikasi_id}/preview"

    async def _check():
        page = await spse_browser._connect_cdp_async(url, navigate=False)
        return await page.evaluate("""
        () => {
            var span = document.getElementById('kswpS');
            if (!span) return 'TIDAK DIKETAHUI';
            var img = span.querySelector('img[src*="verified"]');
            return img ? 'VALID' : 'TIDAK VALID';
        }
        """)

    try:
        return spse_browser._run(_check())
    except Exception:
        return "TIDAK DIKETAHUI"


# ── KSWP: dari file PDF ────────────────────────────────────────────────────────

def parse_kswp_pdf(pdf_path: str) -> str:
    """
    Baca status KSWP dari file PDF (teks atau scan).
    Return: 'VALID' | 'TIDAK VALID' | 'TIDAK DIKETAHUI'
    """
    try:
        text = _pdf_to_text(pdf_path).lower()
        if "kswp valid" in text or (
            "status kswp" in text and "valid" in text.split("status kswp", 1)[-1][:50]
        ):
            return "VALID"
        if "tidak valid" in text or "not valid" in text:
            return "TIDAK VALID"
        if "valid" in text:
            return "VALID"
        return "TIDAK DIKETAHUI"
    except Exception:
        return "TIDAK DIKETAHUI"


def get_kswp_status(kualifikasi_id: str, folder_peserta: str) -> str:
    """
    Strategi double-check KSWP:
    1. Cari file KSWP*.pdf di folder peserta → parse teks/OCR
    2. Fallback: DOM Playwright logo verified
    """
    pdf = _find_file(folder_peserta, "kswp", "konfirmasi validasi")
    if pdf:
        result = parse_kswp_pdf(pdf)
        if result != "TIDAK DIKETAHUI":
            return result
    # Fallback DOM
    return get_kswp_from_dom(kualifikasi_id)


# ── Nilai Kinerja: dari file PDF ──────────────────────────────────────────────

_KINERJA_KATEGORI = [
    ("SANGAT BAIK", ["sangat baik"]),
    ("BAIK",        ["baik"]),
    ("CUKUP",       ["cukup"]),
    ("BURUK",       ["buruk", "kurang"]),
]

def parse_kinerja_pdf(pdf_path: str) -> dict:
    """
    Baca nilai kinerja dari PDF (teks atau scan).
    Return: {"ada": bool, "nilai": str, "kategori": str}
    """
    try:
        # OCR seluruh dokumen
        full_text = _pdf_to_text(pdf_path).lower()
        # OCR footer (area nilai total)
        footer_text = _pdf_footer_text(pdf_path).lower()
        combined = full_text + "\n" + footer_text

        # Cari angka nilai (pola: "nilai kinerja ... 2.00" atau "2 (baik)" atau "2,00")
        nilai_str = ""
        # Pola 1: "nilai kinerja ... angka"
        m = re.search(r"nilai\s+kinerja\D{0,40}(\d[\d.,]+)", combined)
        if m:
            nilai_str = m.group(1).replace(",", ".")
        # Pola 2: "angka (baik/cukup/...)"
        if not nilai_str:
            m = re.search(r"(\d[\d.,]+)\s*\(\s*(baik|cukup|buruk|sangat)", combined)
            if m:
                nilai_str = m.group(1).replace(",", ".")
        # Pola 3: angka desimal standalone di area footer (misal "2.00" atau "2,00")
        if not nilai_str:
            for m in re.finditer(r"\b([0-3][.,]\d{2})\b", footer_text):
                nilai_str = m.group(1).replace(",", ".")
                break

        # Cari kategori — cek combined lalu footer khusus
        kategori = ""
        for kat, keywords in _KINERJA_KATEGORI:
            if any(kw in combined for kw in keywords):
                kategori = kat
                break

        return {"ada": True, "nilai": nilai_str or "-", "kategori": kategori or "-"}
    except Exception as e:
        return {"ada": False, "nilai": "-", "kategori": "-", "error": str(e)}


def get_kinerja(folder_peserta: str) -> dict:
    """
    Cari file kinerja di folder peserta lalu parse.
    Return: {"ada": bool, "nilai": str, "kategori": str}
    """
    pdf = _find_file(folder_peserta, "kinerja", "penilaian kinerja", "lembar penilaian")
    if not pdf:
        return {"ada": False, "nilai": "-", "kategori": "-"}
    return parse_kinerja_pdf(pdf)


# ── SKP: double-check dari Formulir Isian Kualifikasi PDF ─────────────────────

def parse_skp_formulir(pdf_path: str) -> dict:
    """
    Parse tabel 'PEKERJAAN SEDANG DILAKSANAKAN' dari Formulir Isian.
    Return: {"ok": bool, "jp": int, "skp": int, "baris": list}
    """
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "\n".join(
                (p.extract_text() or "") for p in pdf.pages
            )

        # Cari section pekerjaan sedang berjalan
        lower = full_text.lower()
        marker = "pekerjaan yang sedang dilaksanakan"
        if marker not in lower:
            marker = "pekerjaan sedang"
        if marker not in lower:
            return {"ok": False, "jp": 0, "skp": 5, "baris": []}

        section = full_text[lower.index(marker):]
        lines = [l.strip() for l in section.split("\n") if l.strip()]

        # Baris data: skip header, ambil sampai "nihil" atau bagian baru
        baris_data = []
        header_passed = False
        for line in lines[1:]:
            ll = line.lower()
            if not header_passed:
                if any(k in ll for k in ["no.", "nama paket", "bidang"]):
                    header_passed = True
                continue
            if "nihil" in ll:
                break
            if any(k in ll for k in ["modal kerja", "surat keterangan", "demikian"]):
                break
            if line and not line.isdigit():
                baris_data.append(line)

        jp = len(baris_data) if baris_data else 0
        return {"ok": True, "jp": jp, "skp": 5 - jp, "baris": baris_data}
    except Exception as e:
        return {"ok": False, "jp": 0, "skp": 5, "baris": [], "error": str(e)}


def get_skp(folder_peserta: str, jp_preview: int) -> dict:
    """
    Hitung SKP dari jp_preview (tabel pekerjaan berjalan di HTML /preview).
    Formulir Isian PDF tidak dipakai karena sering tidak konsisten.
    Return: {"skp": int, "jp": int, "catatan": int (JP pekerjaan berjalan), "berbeda": bool}
    """
    skp = 5 - jp_preview
    return {
        "skp": skp,
        "jp": jp_preview,
        "catatan": jp_preview,  # jumlah pekerjaan berjalan (P), bukan SKP
        "berbeda": False,
    }


# ── Entry point utama: parse semua data 1 peserta ─────────────────────────────

def parse_peserta_lengkap(
    kualifikasi_id: str,
    folder_peserta: str,
    progress_cb=None,
    kode_tender: str = "",
) -> dict:
    """
    Parse semua data kualifikasi 1 peserta dari semua sumber.

    Return dict lengkap siap di-fill ke Excel:
    {
        "ok": bool,
        "nama", "npwp", "alamat",
        "nib_nomor", "nib_berlaku",
        "ss_nomor", "ss_berlaku", "ss_kualifikasi",
        "sbu_nomor", "sbu_berlaku", "sbu_kualifikasi", "sbu_klasifikasi",
        "pengalaman": [...],
        "pemilik": [...],
        "akta_pendirian": {"nomor","tanggal","notaris"},
        "akta_perubahan": {"nomor","tanggal","notaris"},
        "skp": int, "skp_catatan": str, "skp_berbeda": bool,
        "kswp_status": str,
        "kinerja_ada": bool, "kinerja_nilai": str, "kinerja_kategori": str,
    }
    """
    def _log(msg):
        if progress_cb:
            progress_cb(msg)

    _log(f"[Parser] Fetch HTML preview kualifikasi {kualifikasi_id}...")
    syarat_keywords = (
        _ambil_syarat_sbu_keywords(kode_tender, folder_peserta)
        if kode_tender else None
    )
    html_data = parse_preview_html(kualifikasi_id, syarat_keywords=syarat_keywords)
    if not html_data.get("ok"):
        return {"ok": False, "pesan": html_data.get("pesan", "Gagal fetch preview")}

    _merge_local_license_evidence(
        html_data, folder_peserta, syarat_keywords, log_cb=_log
    )

    _log("[Parser] Cek KSWP...")
    kswp = get_kswp_status(kualifikasi_id, folder_peserta)

    _log("[Parser] Cek Nilai Kinerja...")
    kinerja = get_kinerja(folder_peserta)

    _log("[Parser] Hitung SKP (double-check)...")
    skp_data = get_skp(folder_peserta, html_data.get("jp_preview", 0))

    return {
        "ok": True,
        # Identitas
        "nama":    html_data.get("nama", ""),
        "npwp":    html_data.get("npwp", ""),
        "alamat":  html_data.get("alamat", ""),
        # NIB
        "nib_nomor":   html_data.get("nib_nomor", ""),
        "nib_berlaku": html_data.get("nib_berlaku", ""),
        # Sertifikat Standar
        "ss_nomor":      html_data.get("ss_nomor", ""),
        "ss_berlaku":    html_data.get("ss_berlaku", ""),
        "ss_kualifikasi": html_data.get("ss_kualifikasi", ""),
        # Status PDF OSS lebih kuat daripada label instansi di preview SPSE.
        "ss_terverifikasi": (
            html_data.get("ss_status") if html_data.get("ss_status")
            else ("Terverifikasi" if html_data.get("ss_nomor") and "OSS" in html_data.get("ss_instansi", "")
            else ("Belum Terverifikasi" if html_data.get("ss_nomor")
            else "Tidak Menyampaikan"))
        ),
        # SBU
        "sbu_nomor":      html_data.get("sbu_nomor", ""),
        "sbu_berlaku":    html_data.get("sbu_berlaku", ""),
        "sbu_kualifikasi": html_data.get("sbu_kualifikasi", "Kecil"),
        "sbu_klasifikasi": html_data.get("sbu_klasifikasi", ""),
        "sbu_subklas_label": html_data.get("sbu_subklas_label", ""),
        # Pengalaman
        "pengalaman": html_data.get("pengalaman", []),
        # Pemilik
        "pemilik": html_data.get("pemilik", []),
        # Akta
        "akta_pendirian": html_data.get("akta_pendirian", {}),
        "akta_perubahan": html_data.get("akta_perubahan", {}),
        # SKP
        "skp":          skp_data["skp"],
        "skp_jp":       skp_data["jp"],
        "skp_catatan":  skp_data["catatan"],
        "skp_berbeda":  skp_data["berbeda"],
        # KSWP
        "kswp_status": kswp,
        # Kinerja
        "kinerja_ada":       kinerja["ada"],
        "kinerja_nilai":     kinerja["nilai"],
        "kinerja_kategori":  kinerja["kategori"],
        # Personel & Peralatan dari /preview (kosong jika tidak diinput peserta)
        "personel_list":     html_data.get("personel_list", []),
        "peralatan_list":    html_data.get("peralatan_list", []),
        # Flag kesesuaian syarat paket (True = tidak match, fallback baris pertama)
        "sbu_tidak_sesuai":  html_data.get("sbu_tidak_sesuai", False),
        "ss_tidak_sesuai":   html_data.get("ss_tidak_sesuai", False),
    }
