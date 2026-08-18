"""
Mode Pengadaan Langsung — Tab 0: Draft Paket PL
Input manual paket PL (JKK atau PK), simpan ke Supabase tabel draft_paket_pl.
Juga berisi fungsi scrape otomatis dari SPSE /dt/paketpp.
"""

import os
import re
import hashlib
from functools import lru_cache
from datetime import date, datetime, timezone, timedelta
from config import sb as _sb

BASE_URL = "https://spse.inaproc.id/tapinkab"

SATKER_LIST = [
    "Dinas Perdagangan",
    "Dinas Pekerjaan Umum Dan Penataan Ruang (Bina Marga)",
    "Dinas Pekerjaan Umum Dan Penataan Ruang (PUPR)",
    "Kecamatan CLU",
    "Dinas Perizinan Terpadu Satu Pintu",
    "Lainnya",
]

STATUS_LIST = ["draft", "undangan", "evaluasi", "negosiasi", "selesai"]

_BULAN_INDONESIA = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4,
    "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
    "september": 9, "oktober": 10, "november": 11, "desember": 12,
}


def _hydrate_nomor_urut_folder(rows: list[dict]) -> list[dict]:
    """Isi nomor urut display dari prefix folder fisik untuk row lama.

    Tidak menulis balik ke Supabase; hanya melengkapi row yang nomor_urut-nya
    kosong agar seluruh tab tetap menampilkan nomor folder yang sebenarnya.
    """
    try:
        from config import OUTPUT_DIR_PL_JKK, OUTPUT_DIR_PL_PK, sanitasi_nama_folder
        roots = (OUTPUT_DIR_PL_JKK, OUTPUT_DIR_PL_PK)
        folders = []
        for root in roots:
            if not os.path.isdir(root):
                continue
            for name in os.listdir(root):
                path = os.path.join(root, name)
                m = re.match(r"^\s*(\d+)\.\s*(.*)$", name)
                if os.path.isdir(path) and m:
                    label = re.sub(r"^PL(?:JKK|PK)\s*-\s*", "", m.group(2), flags=re.I)
                    folders.append((int(m.group(1)), set(sanitasi_nama_folder(label).lower().split())))
        for row in rows:
            if row.get("nomor_urut"):
                continue
            words = set(sanitasi_nama_folder(row.get("nama_paket") or "").lower().split())
            if not words:
                continue
            best = max(folders, key=lambda item: len(words & item[1]) / max(len(words), 1), default=None)
            if best and len(words & best[1]) / max(len(words), 1) >= 0.6:
                row["nomor_urut"] = best[0]
    except Exception:
        pass
    return rows


def _normalisasi_tanggal_excel(value) -> str:
    """Normalisasi C21 Excel ke ISO agar bisa dipakai UI dan SPSE."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    # XML worksheet menyimpan tanggal sebagai serial Excel (bukan datetime).
    # C21 memang field tanggal; konversi serial umum tanpa membuka workbook.
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            serial = float(value)
            if 1 <= serial <= 100000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).date().isoformat()
        except (OverflowError, ValueError):
            pass
    text = str(value or "").strip()
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        try:
            serial = float(text)
            if 1 <= serial <= 100000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).date().isoformat()
        except (OverflowError, ValueError):
            pass
    iso = re.match(r"^(\d{4}-\d{2}-\d{2})(?:[T ]|$)", text)
    if iso:
        return iso.group(1)
    m = re.fullmatch(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", text)
    if m and m.group(2).lower() in _BULAN_INDONESIA:
        return date(int(m.group(3)), _BULAN_INDONESIA[m.group(2).lower()], int(m.group(1))).isoformat()
    return text


def _read_master_data_xml(
    xlsm: str,
    wanted: tuple[str, ...] | None = None,
) -> tuple[str, ...]:
    """Baca sel Master Data langsung dari XML XLSM (tanpa OpenPyXL)."""
    import posixpath
    import zipfile
    import xml.etree.ElementTree as ET

    def local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    with zipfile.ZipFile(xlsm, "r") as zf:
        names = set(zf.namelist())
        workbook_path = "xl/workbook.xml"
        rels_path = "xl/_rels/workbook.xml.rels"
        if workbook_path not in names or rels_path not in names:
            raise ValueError("Workbook XML tidak lengkap")

        workbook = ET.fromstring(zf.read(workbook_path))
        sheet_rel_id = ""
        for node in workbook.iter():
            if local(node.tag) == "sheet" and node.attrib.get("name") == "@ Master Data":
                sheet_rel_id = next(
                    (value for key, value in node.attrib.items() if local(key) == "id"),
                    "",
                )
                break
        if not sheet_rel_id:
            raise KeyError("Sheet @ Master Data tidak ditemukan")

        rels = ET.fromstring(zf.read(rels_path))
        target = ""
        for node in rels.iter():
            rel_id = next((value for key, value in node.attrib.items() if local(key) == "Id"), "")
            if local(node.tag) != "Relationship" or rel_id != sheet_rel_id:
                continue
            target = node.attrib.get("Target", "")
            break
        if not target:
            raise KeyError("Relasi worksheet Master Data tidak ditemukan")
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = posixpath.join("xl", target)
        target = posixpath.normpath(target)
        if not target.startswith("xl/") or target not in names:
            raise ValueError(f"Worksheet XML tidak valid: {target}")

        shared_strings = []
        shared_path = "xl/sharedStrings.xml"
        if shared_path in names:
            shared_root = ET.fromstring(zf.read(shared_path))
            for si in shared_root.iter():
                if local(si.tag) == "si":
                    shared_strings.append("".join(
                        (text.text or "")
                        for text in si.iter()
                        if local(text.tag) == "t"
                    ))

        sheet_root = ET.fromstring(zf.read(target))
        cells = {}
        requested = tuple(wanted or ("F2", "C20", "C21"))
        for cell in sheet_root.iter():
            if local(cell.tag) != "c" or cell.attrib.get("r") not in requested:
                continue
            cell_type = cell.attrib.get("t", "")
            value_node = next((child for child in cell if local(child.tag) == "v"), None)
            inline_node = next((child for child in cell if local(child.tag) == "is"), None)
            if cell_type == "inlineStr" and inline_node is not None:
                value = "".join(
                    (text.text or "")
                    for text in inline_node.iter()
                    if local(text.tag) == "t"
                )
            elif value_node is None:
                value = ""
            else:
                raw = value_node.text or ""
                if cell_type == "s":
                    try:
                        value = shared_strings[int(raw)]
                    except (ValueError, IndexError):
                        value = ""
                elif cell_type == "b":
                    value = raw == "1"
                else:
                    value = raw
            cells[cell.attrib["r"]] = value

        if wanted is not None:
            return tuple(str(cells.get(cell) or "").strip() for cell in requested)
        return (
            str(cells.get("F2") or "").strip(),
            str(cells.get("C20") or "").strip(),
            _normalisasi_tanggal_excel(cells.get("C21")),
        )


def _enrich_manual_excel_pl(row: dict) -> dict:
    """Excel Master Data menjadi sumber utama metadata surat PL."""
    try:
        from parse_kak_pl import _resolve_folder_pl
        folder, _ = _resolve_folder_pl(
            row.get("nomor_urut"), row.get("nama_paket") or "",
            row.get("jenis_pl") or "JKK", is_ulang=bool(row.get("is_ulang")),
        )
        if not folder:
            return row
        xlsm = next(
            (os.path.join(folder, f) for f in os.listdir(folder)
             if f.lower().endswith(".xlsm") and "backup" not in f.lower()),
            "",
        )
        if not xlsm:
            return row
        stat = os.stat(xlsm)
        kode_unik, nomor_dokpil, tgl_dokpil = read_master_data_cached(
            xlsm, stat.st_mtime_ns, stat.st_size
        )
        if kode_unik:
            row["kode_unik"] = kode_unik
        if nomor_dokpil:
            row["nomor_dokpil"] = nomor_dokpil
        # Tanggal kosong di Excel sengaja menimpa cache Supabase: user wajib
        # mengisi panel INPUT TANGGAL DOKPIL sebelum upload/merge dokumen.
        row["tgl_dokpil"] = tgl_dokpil
    except Exception:
        pass
    return row


@lru_cache(maxsize=256)
def read_master_data_cached(xlsm: str, mtime_ns: int, size: int) -> tuple[str, str, str]:
    """Baca field Master Data sekali per versi file workbook.

    Key menyertakan mtime + size, jadi edit workbook otomatis menghasilkan
    cache key baru tanpa mengubah source-of-truth Excel.
    """
    try:
        return _read_master_data_xml(xlsm)
    except Exception:
        # Fallback kompatibilitas untuk workbook XML non-standar/corrupt.
        from openpyxl import load_workbook
        wb = load_workbook(xlsm, read_only=True, data_only=True, keep_vba=False)
        try:
            ws = wb["@ Master Data"]
            return (
                str(ws["F2"].value or "").strip(),
                str(ws["C20"].value or "").strip(),
                _normalisasi_tanggal_excel(ws["C21"].value),
            )
        finally:
            wb.close()


@lru_cache(maxsize=512)
def read_master_cells_cached(
    xlsm: str,
    mtime_ns: int,
    size: int,
    cells: tuple[str, ...],
) -> dict[str, str]:
    """Baca sel Master Data ringan; fallback OpenPyXL untuk workbook legacy."""
    del mtime_ns, size
    requested = tuple(dict.fromkeys(str(cell).strip() for cell in cells if cell))
    if not requested:
        return {}
    try:
        values = _read_master_data_xml(xlsm, requested)
    except Exception:
        from openpyxl import load_workbook

        wb = load_workbook(xlsm, read_only=True, data_only=True, keep_vba=False)
        try:
            ws = wb["@ Master Data"]
            values = tuple(ws[cell].value for cell in requested)
        finally:
            wb.close()
    return {
        cell: str(value or "").strip()
        for cell, value in zip(requested, values)
    }


def load_draft_pl() -> list[dict]:
    """Ambil paket PL JKK (jenis_pl=JKK, jenis_kontrak=Lumsum/Waktu Penugasan), urut terbaru dulu."""
    try:
        rows = (
            _sb()
            .table("draft_paket_pl")
            .select("*")
            .eq("jenis_pl", "JKK")
            .in_("jenis_kontrak", ["Lumsum", "Waktu Penugasan"])
            .order("diambil_pada", desc=True)
            .execute()
            .data or []
        )
        rows = _hydrate_nomor_urut_folder(rows)
        hasil = [_enrich_manual_excel_pl(row) for row in rows]
        for row in hasil:
            if row.get("nama_ppk"):
                row["nama_ppk"] = _lookup_nama_ppk_lengkap(row["nama_ppk"])
            # Lokasi authoritative berasal dari viewdraftpl. Hanya fallback
            # bila row lama belum punya nilai sama sekali.
            if not str(row.get("lokasi") or "").strip():
                row["lokasi"] = "Kabupaten Tapin"
        return hasil
    except Exception as e:
        return []


_TAHAP_SELESAI_KEYWORDS = ("penandatanganan kontrak", "paket sudah selesai", "sudah selesai", "selesai")
_STATUS_DITARIK_SPSE = "ditarik_spse"


def is_paket_ditarik(r: dict) -> bool:
    """True bila row tidak lagi ada di daftar paket PP SPSE.

    State ini sengaja disimpan sebagai status, bukan menghapus row Supabase,
    agar histori/folder lokal tetap aman dan paket dapat hidup lagi bila SPSE
    mengembalikan kode yang sama.
    """
    status = str(r.get("status") or "").strip().lower()
    return any(marker in status for marker in (
        _STATUS_DITARIK_SPSE,
        "withdrawn",
        "retired",
        "ditarik ppk",
        "dihapus ppk",
    ))

def is_paket_selesai(r: dict) -> bool:
    """
    True jika paket sudah selesai dari sisi PP.
    Sumber: kolom tahap_spse (dt/pengadaan-pp). Fallback ke status lama.
    """
    tahap = (r.get("tahap_spse") or "").lower()
    if tahap:
        return any(k in tahap for k in _TAHAP_SELESAI_KEYWORDS)
    return any(k in (r.get("status") or "").lower() for k in _TAHAP_SELESAI_KEYWORDS)


def is_paket_draft(r: dict) -> bool:
    """True hanya untuk paket yang masih berstatus Draft di SPSE.

    Tab setup awal tidak boleh mengubah atau mengunggah paket yang sudah
    masuk status berjalan atau tahap SPSE lain. Row tanpa status juga ditolak
    agar fase setup tidak mengambil keputusan dari data yang ambigu.
    """
    status = str(r.get("status") or "").strip().lower()
    tahap = str(r.get("tahap_spse") or "").strip()
    if tahap:
        return False
    return status == "draft"


def is_paket_berjalan(r: dict) -> bool:
    """True hanya untuk paket aktif/berjalan pada tab operasional.

    Status atau tahap SPSE wajib memberi bukti bahwa row bukan draft. Row
    kosong/ambigu dan paket selesai ditolak agar Tab 6–10 fail-closed.
    """
    status = str(r.get("status") or "").strip()
    tahap = str(r.get("tahap_spse") or "").strip()
    if is_paket_ditarik(r) or (not status and not tahap):
        return False
    return not is_paket_draft(r) and not is_paket_selesai(r)


def buang_duplikat_paket_lama(rows: list[dict]) -> tuple[list[dict], int]:
    """
    Jika ada >1 row dgn nama_paket sama (paket di-ulang → kode baru, row lama nyangkut),
    simpan hanya row kode_paket TERBARU (string terbesar; kode ulang 1089xxx > lama 1086xxx).
    Return: (rows_terfilter, jumlah_dibuang).
    """
    by_nama: dict[str, dict] = {}
    for r in rows:
        nama = (r.get("nama_paket") or "").strip()
        if not nama:
            # tanpa nama → loloskan apa adanya (pakai kode sebagai key unik)
            by_nama[r.get("kode_paket") or id(r)] = r
            continue
        prev = by_nama.get(nama)
        if prev is None or str(r.get("kode_paket") or "") > str(prev.get("kode_paket") or ""):
            by_nama[nama] = r
    hasil = list(by_nama.values())
    return hasil, len(rows) - len(hasil)


_SUFFIX_ULANG = " (PL - Ulang)"


def nama_folder_dengan_suffix_ulang(
    output_base: str,
    nama_folder: str,
    paksa_suffix: bool = False,
) -> str:
    """
    Generate nama folder final, auto-tambah ' (PL - Ulang)' bila folder dgn
    nama sama SUDAH ADA di disk (paket di-ulang → simpan folder lama).

    - paksa_suffix=True → selalu tempel suffix (override manual user).
    - Idempoten: kalau nama sudah berakhiran suffix, tidak ditumpuk.
    - Bila folder bersuffix juga sudah ada → tambah angka: ' (PL - Ulang 2)', dst.

    Return: nama folder (bukan path penuh).
    """
    import os as _os
    nama = (nama_folder or "").strip()
    if not nama:
        return nama

    sudah_bersuffix = nama.endswith(_SUFFIX_ULANG.strip()) or _SUFFIX_ULANG.strip() in nama
    target_polos = _os.path.join(output_base, nama)
    perlu_suffix = paksa_suffix or _os.path.exists(target_polos)

    if not perlu_suffix or sudah_bersuffix:
        return nama

    kandidat = f"{nama}{_SUFFIX_ULANG}"
    if not _os.path.exists(_os.path.join(output_base, kandidat)):
        return kandidat
    # Folder ulang juga sudah ada → cari angka kosong
    n = 2
    while _os.path.exists(_os.path.join(output_base, f"{nama}{_SUFFIX_ULANG.rstrip(')')}{n})")):
        n += 1
    return f"{nama}{_SUFFIX_ULANG.rstrip(')')}{n})"


def nomor_folder_tertinggi(output_base: str) -> int:
    """Scan folder di output_base, ambil nomor terbesar dari prefix 'N. PLJKK/PLPK ...'."""
    import os as _os, re as _re
    maks = 0
    if not _os.path.isdir(output_base):
        return maks
    for nama in _os.listdir(output_base):
        if _os.path.isdir(_os.path.join(output_base, nama)):
            m = _re.match(r'^(\d+)\.', nama)
            if m:
                maks = max(maks, int(m.group(1)))
    return maks


def truncate_nama_folder(output_base: str, nama_folder: str, max_path: int = 247) -> str:
    """
    Potong nama_folder agar semua path di dalamnya (subfolder + file Excel) ≤ max_path.
    Constraint terberat: file Excel = output_base/folder/0. BAPLJKK - {suffix}.xlsm
    di mana suffix = bagian nama_folder setelah prefix "N. PLJKK - ".
    Formula: len(base) + 1 + len(folder) + 1 + 14 + len(suffix) + 6 ≤ max_path
    suffix ≈ folder - len_prefix → 2*folder terlibat → pakai max_nama = (max_path - konstanta) / 2.
    max_path=247 — threshold WinError 206 aktual di sistem ini.
    """
    import os as _os, re as _re
    # Hitung panjang prefix nomor folder (e.g. "16. PLJKK - " = 12, "999. PLJKK - " = 15)
    _m = _re.match(r'^(\d+\.\s+PL(?:JKK|PK)\s+-\s+)', nama_folder)
    _prefix_len = len(_m.group(1)) if _m else 12  # fallback 12
    # Formula: len(base)+1 + len(folder)+1 + len("0. BAPLJKK - ")+len(suffix)+len(".xlsm")
    #        = len(base)+1 + folder+1 + 14 + (folder-prefix_len) + 6
    #        = len(base) + 2 + 2*folder + 20 - prefix_len
    # Agar ≤ max_path: 2*folder ≤ max_path - len(base) - 2 - 20 + prefix_len
    _base_len = len(output_base)
    max_nama = (max_path - _base_len - 2 - 20 + _prefix_len) // 2

    if len(nama_folder) <= max_nama:
        return nama_folder

    # Pisahkan suffix ulang agar tidak terpotong
    suffix = ""
    _sfx_ulang = " (PL - Ulang"
    if _sfx_ulang in nama_folder:
        idx = nama_folder.index(_sfx_ulang)
        suffix = nama_folder[idx:]
        nama_folder = nama_folder[:idx]

    # Sisakan ruang untuk suffix
    batas = max_nama - len(suffix)
    if batas <= 0:
        return (nama_folder[:max_nama]).rstrip()

    # Potong di batas kata
    terpotong = nama_folder[:batas].rstrip()
    if len(nama_folder) > batas and " " in terpotong:
        terpotong = terpotong.rsplit(" ", 1)[0]

    return terpotong + suffix


def simpan_paket_pl(data: dict) -> dict:
    """
    Upsert satu paket PL ke draft_paket_pl.
    data harus memiliki key 'kode_paket'.
    Return: {"ok": True} atau {"ok": False, "error": str}
    """
    if not data.get("kode_paket"):
        return {"ok": False, "error": "kode_paket wajib diisi"}
    data = dict(data)
    # Dua nama kolom historis dipakai lintas flow. Jaga keduanya selalu sama
    # agar workbook donor tidak mempertahankan tanggal paket sebelumnya.
    tgl_buka = data.get("tgl_buka_penawaran") or data.get("tgl_pembukaan")
    if tgl_buka:
        data["tgl_buka_penawaran"] = tgl_buka
        data["tgl_pembukaan"] = tgl_buka
    data.setdefault("diambil_pada", datetime.now(timezone.utc).isoformat())
    try:
        _sb().table("draft_paket_pl").upsert(data).execute()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def hapus_paket_pl(kode_paket: str) -> dict:
    """Hapus satu baris dari draft_paket_pl berdasarkan kode_paket."""
    try:
        _sb().table("draft_paket_pl").delete().eq("kode_paket", kode_paket).execute()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def update_status(kode_paket: str, status: str) -> dict:
    """Update kolom status paket PL."""
    try:
        _sb().table("draft_paket_pl").update({"status": status}).eq("kode_paket", kode_paket).execute()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def tandai_folder_dibuat(kode_paket: str) -> dict:
    """Set folder_dibuat=True dan folder_dibuat_pada=now."""
    try:
        query = _sb().table("draft_paket_pl").update({
            "folder_dibuat": True,
            "folder_dibuat_pada": datetime.now(timezone.utc).isoformat(),
        }).eq("kode_paket", kode_paket)
        # Minta row hasil update. Mock lama mungkin belum punya select().
        try:
            query = query.select("kode_paket")
        except (AttributeError, TypeError):
            pass
        response = query.execute()
        data = getattr(response, "data", None)
        if data is None and isinstance(response, dict):
            data = response.get("data")
        if isinstance(data, (list, tuple, dict)) and not data:
            return {"ok": False, "error": f"Paket {kode_paket} tidak ditemukan saat menandai folder_dibuat."}
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ============================================================
# Scrape otomatis dari SPSE
# ============================================================

def _parse_hps_dari_edit(html: str) -> str:
    """Ekstrak nilai HPS dari halaman nontender/{kode}/edit."""
    import re
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    teks = soup.get_text(" ", strip=True)
    m = re.search(r"Nilai HPS\s*Rp\.\s*([\d.,]+)", teks)
    return f"Rp. {m.group(1)}" if m else ""


def _parse_jenis_kontrak_dari_edit(html: str) -> str:
    """Ekstrak Jenis Kontrak dari halaman nontender/{kode}/edit."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    teks = soup.get_text(" ", strip=True)
    m = re.search(r"Jenis Kontrak\s+([\w\s]+?)(?:Dokumen|Jadwal|Survey|\Z)", teks)
    if m:
        return m.group(1).strip()
    return ""


def _parse_metode_pengadaan_dari_edit(html: str) -> str:
    """Ekstrak Metode Pengadaan dari halaman nontender/{kode}/edit."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    teks = soup.get_text(" ", strip=True)
    m = re.search(r"Metode Pengadaan\s+(.+?)(?:Kualifikasi Usaha|Jenis Kontrak|Dokumen|Jadwal|\Z)", teks)
    if m:
        return m.group(1).strip()
    return ""


def _derive_jenis_pl_dari_metode(metode: str, nama_paket: str) -> str:
    """Derive jenis_pl dari metode_pengadaan (lebih akurat dari nama saja)."""
    if metode:
        m_lower = metode.lower()
        if "barang" in m_lower:
            return "PK"
        # "Non Konstruksi" maupun "Konstruksi" → JKK (LDK/checklist identik)
        if "konsultan" in m_lower:
            return "JKK"
    # Fallback ke keyword nama
    nama_lower = nama_paket.lower()
    if any(k in nama_lower for k in ["konsultan", "perencanaan", "pengawasan", "supervisi", "manajemen konstruksi"]):
        return "JKK"
    return "PK"


def _fetch_tahap_spse(cookie_str: str, base_url: str, log_fn=None) -> dict:
    """
    Fetch status TAHAPAN paket dari dashboard /home → dt/pengadaan-pp?status=1.
    Berbeda dgn dt/paketpp (status PP basi: Draft/Berjalan), endpoint ini
    nunjukin tahapan REAL: "Penandatanganan Kontrak" = selesai dari sisi PP.

    Wajib POST + authenticityToken (CSRF dari HTML /home) + X-Requested-With.
    Paket draft/ulang yang belum sampai tahapan TIDAK muncul → tidak ada di map.

    Return: {kode_paket: tahap_str, ...}  (kosong kalau gagal — non-fatal)
    """
    import requests
    import re as _re

    def log(msg):
        if log_fn:
            log_fn(msg)

    base = base_url.rstrip("/")
    headers = {"Cookie": cookie_str, "User-Agent": "Mozilla/5.0", "Referer": base + "/home"}

    # 1. Ambil authenticityToken dari HTML /home
    try:
        r_home = requests.get(f"{base}/home", headers=headers, timeout=15)
        m = _re.search(r"authenticityToken\s*=\s*'([0-9a-f]+)'", r_home.text)
        token = m.group(1) if m else ""
    except Exception as e:
        log(f"  [tahap] gagal fetch /home: {e}")
        return {}

    if not token:
        log("  [tahap] authenticityToken tidak ditemukan — lewati status tahapan")
        return {}

    # 2. POST dt/pengadaan-pp?status=1
    hdr = dict(headers)
    hdr["X-Requested-With"] = "XMLHttpRequest"
    hdr["Content-Type"] = "application/x-www-form-urlencoded"
    payload = {
        "draw": "1", "start": "0", "length": "200",
        "authenticityToken": token,
        "search[value]": "", "search[regex]": "false",
    }
    tahap_map = {}
    try:
        r = requests.post(f"{base}/dt/pengadaan-pp?status=1", headers=hdr, data=payload, timeout=20)
        data = r.json().get("data", [])
        for row in data:
            kode  = str(row[0])   # col[0] = kode_paket
            tahap = str(row[2]) if len(row) > 2 else ""   # col[2] = tahap/status
            if kode and tahap:
                tahap_map[kode] = tahap
        log(f"  [tahap] {len(tahap_map)} paket punya status tahapan (dt/pengadaan-pp)")
    except Exception as e:
        log(f"  [tahap] gagal POST dt/pengadaan-pp: {e}")

    return tahap_map


# Kolom yang disalin dari paket lama ke paket ulang
_KOLOM_FALLBACK = [
    "mak", "kode_rup", "nilai_pagu", "sub_kegiatan",
    "sumber_anggaran", "dpa_nomor", "lokasi", "sbu_baru", "sbu_lama",
]


def _copy_data_dari_paket_lama(
    kode_paket_baru: str,
    nama_paket: str,
    semua_rows: list[dict],
    log_fn=None,
) -> bool:
    """
    Salin field bolong dari paket lama (nama_paket sama, kode_paket lebih kecil).

    Dipanggil setelah upsert selesai pada serap_paket_pl_dari_spse.
    Tidak akan memanggil load_draft_pl() lagi (menerima semua_rows dari luar)
    sehingga tidak ada loop tak terhingga.

    Return: True jika ada data yang disalin, False jika tidak perlu.
    """
    def log(msg):
        if log_fn:
            log_fn(msg)

    if not nama_paket or not kode_paket_baru:
        return False

    nama_bersih = nama_paket.strip()

    # Cari baris paket baru dari semua_rows (sudah di-load setelah upsert)
    row_baru = next(
        (r for r in semua_rows if str(r.get("kode_paket") or "") == kode_paket_baru),
        None,
    )
    if row_baru is None:
        # Fallback: ambil langsung dari DB (paket baru mungkin tidak ada di semua_rows lama)
        try:
            res = _sb().table("draft_paket_pl").select("*").eq("kode_paket", kode_paket_baru).maybe_single().execute()
            row_baru = (res.data or {}) if res else {}
        except Exception:
            return False

    # Cek apakah ada field yang masih kosong
    field_kosong = [k for k in _KOLOM_FALLBACK if not (row_baru or {}).get(k)]
    if not field_kosong:
        return False  # semua sudah terisi, tidak perlu copy

    # Cari kandidat paket lama: nama sama, kode lebih kecil
    kandidat_lama = [
        r for r in semua_rows
        if (r.get("nama_paket") or "").strip() == nama_bersih
        and str(r.get("kode_paket") or "") < kode_paket_baru
        and str(r.get("kode_paket") or "") != kode_paket_baru
    ]
    if not kandidat_lama:
        return False

    # Paket lama terbaru (kode terbesar di antara kandidat)
    paket_lama = max(kandidat_lama, key=lambda r: str(r.get("kode_paket") or ""))
    kode_lama = paket_lama.get("kode_paket")

    # Ambil nilai dari paket lama hanya untuk field yang masih kosong di paket baru
    update_data = {}
    for k in field_kosong:
        nilai_lama = paket_lama.get(k)
        if nilai_lama:
            update_data[k] = nilai_lama

    if not update_data:
        return False

    log(f"  [ulang] salin dari {kode_lama} → {kode_paket_baru}: {list(update_data.keys())}")
    try:
        _sb().table("draft_paket_pl").update(update_data).eq("kode_paket", kode_paket_baru).execute()
        # Mark paket lama selesai — sudah digantikan oleh paket ulang
        _sb().table("draft_paket_pl").update(
            {"tahap_spse": "Paket Sudah Selesai"}
        ).eq("kode_paket", kode_lama).execute()
        log(f"  [ulang] {kode_lama} di-mark selesai (digantikan {kode_paket_baru})")
        return True
    except Exception as e:
        log(f"  [ulang] GAGAL update {kode_paket_baru}: {e}")
        return False


def _parse_nama_ppk_dari_view(html: str) -> str:
    """Parse nama PPK dari halaman view /nontender/{kode}."""
    from bs4 import BeautifulSoup as _BS
    soup = _BS(html, "html.parser")
    for tr in soup.find_all("tr"):
        tds = tr.find_all(["td", "th"])
        if len(tds) >= 2 and tds[0].get_text(strip=True) == "PPK":
            return tds[1].get_text(strip=True)
    return ""


@lru_cache(maxsize=128)
def _lookup_nama_ppk_lengkap(nama_ppk: str) -> str:
    """Enrich nama PPK SPSE (sering tanpa gelar) dari master_kpa.

    Tidak membuat tabel baru: master_kpa sudah menjadi sumber referensi SK/NIP
    dan memuat nama resmi beserta gelar. Jika tidak ada kecocokan, nama SPSE
    dikembalikan apa adanya.
    """
    nama_asli = str(nama_ppk or "").strip()
    if not nama_asli:
        return ""
    nama_dasar = nama_asli.split(",", 1)[0].strip()
    try:
        rows = (
            _sb().table("master_kpa")
            .select("nama")
            .ilike("nama", f"%{nama_dasar}%")
            .limit(10)
            .execute()
            .data or []
        )
        dasar_norm = re.sub(r"\s+", " ", nama_dasar).casefold()
        for row in rows:
            kandidat = str(row.get("nama") or "").strip()
            kandidat_dasar = re.sub(r"\s+", " ", kandidat.split(",", 1)[0]).casefold()
            if kandidat_dasar == dasar_norm:
                return kandidat
    except Exception:
        pass
    return nama_asli


def _scrape_viewdraftpl(kode_paket: str, headers: dict, base_url: str) -> dict:
    """Scrape sumber_anggaran + lokasi dari /nontender/{kode}/viewdraftpl."""
    import requests
    from bs4 import BeautifulSoup
    result = {"sumber_anggaran": "", "lokasi": ""}
    _SUMBER_VALID = {
        "APBD", "APBDP", "APBD-P", "APBN", "DAK", "BLU", "BLUD",
        "APBD Provinsi",
    }
    try:
        r = requests.get(f"{base_url}nontender/{kode_paket}/viewdraftpl",
                         headers=headers, timeout=15)
        if r.status_code != 200:
            return result
        soup = BeautifulSoup(r.text, "html.parser")
        for tr in soup.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if not tds:
                continue
            label = tds[0].get_text(strip=True)
            if label == "Lokasi Pekerjaan" and len(tds) >= 2:
                result["lokasi"] = tds[1].get_text(strip=True)
        # Sumber dana: cari cell yang nilainya tepat salah satu _SUMBER_VALID
        for td in soup.find_all(["td", "th"]):
            txt = td.get_text(strip=True)
            if txt in _SUMBER_VALID:
                result["sumber_anggaran"] = "APBDP" if txt == "APBD-P" else txt
                break
    except Exception:
        pass
    return result


def serap_paket_pl_dari_spse(cookie_str: str, base_url: str, log_fn=None) -> dict:
    """
    Scrape daftar paket non-tender dari SPSE /dt/paketpp,
    fetch detail tiap paket dari /nontender/{kode}/edit,
    upsert ke Supabase draft_paket_pl.

    cookie_str : hasil get_spse_cookies()
    base_url   : SPSE_BASE_URL (diakhiri /)
    log_fn     : callable(str) untuk log progres, opsional
    Returns    : {"ok": True, "scraped": N, "errors": [...]}
    """
    import requests

    def log(msg):
        if log_fn:
            log_fn(msg)

    headers = {"Cookie": cookie_str, "User-Agent": "Mozilla/5.0"}

    # 1. Fetch daftar paket
    try:
        resp = requests.get(f"{base_url}dt/paketpp", headers=headers, timeout=15)
        rows = resp.json().get("data", [])
    except Exception as e:
        return {"ok": False, "scraped": 0, "errors": [f"Gagal fetch dt/paketpp: {e}"]}

    log(f"Ditemukan {len(rows)} paket di SPSE")

    # 1b. Fetch status TAHAPAN (Penandatanganan Kontrak dll) dari dt/pengadaan-pp
    tahap_map = _fetch_tahap_spse(cookie_str, base_url, log_fn)

    # 1c. Ambil kode_paket yang sudah selesai dari DB (fallback jika tidak muncul di tahap_map)
    _selesai_db = set()
    try:
        _res = _sb().table("draft_paket_pl").select("kode_paket,tahap_spse").execute()
        for _r in (_res.data or []):
            if any(k in (_r.get("tahap_spse") or "").lower() for k in _TAHAP_SELESAI_KEYWORDS):
                _selesai_db.add(str(_r["kode_paket"]))
    except Exception:
        pass

    errors = []
    scraped = 0

    for row in rows:
        id_paket_internal = str(row[0])  # ID paket-level (kolom 0), bukan untuk kirim verifikasi
        nama_paket   = row[1]
        status_spse  = row[2]
        satker       = row[4]
        kode_paket   = str(row[5])   # kode resmi non-tender

        # Paket selesai — skip scrape detail sama sekali (sudah tidak relevan)
        _tahap_skrg = (tahap_map.get(kode_paket) or "").lower()
        if any(k in _tahap_skrg for k in _TAHAP_SELESAI_KEYWORDS) or kode_paket in _selesai_db:
            log(f"  Skip detail {kode_paket} — sudah Penandatanganan Kontrak, update tahap saja")
            try:
                _sb().table("draft_paket_pl").update(
                    {"tahap_spse": tahap_map.get(kode_paket)}
                ).eq("kode_paket", kode_paket).execute()
            except Exception:
                pass
            continue

        log(f"  Scraping {kode_paket} — {nama_paket[:40]}...")

        # Ambil ID peserta dari halaman evaluasi (untuk kirimundanganverifikasi)
        id_nontender = id_paket_internal  # fallback jika belum ada peserta
        is_ulang = False  # badge "Paket Ulang" / "Pengadaan Langsung Ulang" di halaman SPSE
        try:
            import re as _re
            r_eval = requests.get(
                f"{base_url}evaluasinontender/{kode_paket}",
                headers=headers, timeout=15
            )
            ids_peserta = _re.findall(
                r'/evaluasinontender/(\d+)/kirimundanganverifikasi', r_eval.text
            )
            if ids_peserta:
                id_nontender = ids_peserta[0]
            # Deteksi paket ulang dari badge (reuse response, tanpa request tambahan)
            is_ulang = ("Paket Ulang" in r_eval.text) or ("Pengadaan Langsung Ulang" in r_eval.text)
        except Exception:
            pass

        # 2. Fetch detail dari halaman edit
        jenis_kontrak = ""
        hps_str = ""
        _hps_live = {}
        metode_pengadaan = ""
        edit_html = ""
        try:
            r_edit = requests.get(
                f"{base_url}nontender/{kode_paket}/edit",
                headers=headers, timeout=15
            )
            edit_html = r_edit.text
            jenis_kontrak = _parse_jenis_kontrak_dari_edit(edit_html)
            metode_pengadaan = _parse_metode_pengadaan_dari_edit(edit_html)
        except Exception as e:
            errors.append(f"{kode_paket}: gagal fetch edit — {e}")

        # HPS/Pagu wajib berasal dari halaman HPS live, bukan nilai stale di edit.
        try:
            from hps_engine import scrape_hps_pl
            _hps_live = scrape_hps_pl(kode_paket)
            hps_str = _hps_live.get("nilai_hps", "")
        except Exception as e:
            errors.append(f"{kode_paket}: gagal scrape HPS live — {e}")

        # 2b. Fetch nama PPK dari halaman view
        nama_ppk = ""
        try:
            r_view = requests.get(
                f"{base_url}nontender/{kode_paket}",
                headers=headers, timeout=15,
            )
            nama_ppk = _parse_nama_ppk_dari_view(r_view.text)
        except Exception:
            pass

        # 2c. Fetch sumber_anggaran + lokasi authoritative dari viewdraftpl
        viewdraft = _scrape_viewdraftpl(kode_paket, headers, base_url)

        # 3. Deteksi jenis PL (dari metode, fallback nama)
        jenis_pl = _derive_jenis_pl_dari_metode(metode_pengadaan, nama_paket)

        data = {
            "kode_paket":        kode_paket,
            "id_nontender":      id_nontender,
            "nama_paket":        nama_paket,
            "satker":            satker,
            "nilai_hps":         hps_str,
            "jenis_pl":          jenis_pl,
            "jenis_kontrak":     jenis_kontrak,
            "metode_pengadaan":  metode_pengadaan,
            "lokasi":            viewdraft.get("lokasi") or "Kabupaten Tapin",
            "status":            status_spse.lower() if status_spse else "draft",
            "is_ulang":          is_ulang,
            "tahap_spse":        tahap_map.get(kode_paket),  # None jika belum ada tahapan
            "diambil_pada":      datetime.now(timezone.utc).isoformat(),
        }

        # Serap adalah boundary paket baru/refresh. Parser berikutnya mengisi
        # kembali field ini dari KAK/ND/Draft; jangan biarkan nilai donor atau
        # hasil paket lama hidup karena upsert parsial Supabase.
        for _field in (
            "nama_ppk", "nip_ppk", "no_sk_ppk", "sbu_baru", "sbu_lama",
            "jabatan_teknis", "skk_teknis", "jabatan_k3", "skk_k3",
            "dpa_nomor", "sub_kegiatan", "nama_file_uraian", "mak",
            "nama_penyedia", "npwp_penyedia", "personil_json",
            "nomor_nota_dinas", "nomor_rekomendasi", "tgl_rekomendasi",
            "uraian_singkat", "masa_berlaku",
        ):
            data[_field] = None

        if nama_ppk:
            data["nama_ppk"] = _lookup_nama_ppk_lengkap(nama_ppk)
        if _hps_live.get("nilai_pagu"):
            data["nilai_pagu"] = _hps_live["nilai_pagu"]
        if viewdraft.get("sumber_anggaran"):
            data["sumber_anggaran"] = viewdraft["sumber_anggaran"]

        try:
            _sb().table("draft_paket_pl").upsert(data, on_conflict="kode_paket").execute()
            scraped += 1
        except Exception as e:
            errors.append(f"{kode_paket}: gagal upsert — {e}")

    log(f"Selesai: {scraped} paket disimpan, {len(errors)} error")

    # 4. Fallback data dari paket lama (untuk paket ulang yang belum dapat delegasi PPK)
    log("Cek paket ulang — copy data dari paket lama jika ada field kosong...")
    semua_paket = load_draft_pl()  # ambil setelah upsert di atas selesai
    for row in rows:
        kode = str(row[5])
        nama = row[1]
        hasil_copy = _copy_data_dari_paket_lama(kode, nama, semua_paket, log_fn)
        if hasil_copy:
            log(f"  [ulang] {kode} — {nama[:40]}: data disalin dari paket lama")

    # 5. Auto set Usaha Kecil — hanya paket aktif (bukan selesai)
    log("Auto set Usaha Kecil untuk paket aktif...")
    for row in rows:
        kode = str(row[5])
        _tahap = (tahap_map.get(kode) or "").lower()
        if any(k in _tahap for k in _TAHAP_SELESAI_KEYWORDS) or kode in _selesai_db:
            continue
        ok_kual = set_kualifikasi_usaha_pl(kode, headers, base_url)
        log(f"  Set Usaha Kecil {kode}: {'OK' if ok_kual else 'GAGAL'}")

    return {"ok": True, "scraped": scraped, "errors": errors}


def set_kualifikasi_usaha_pl(kode_paket: str, headers: dict, base_url: str) -> bool:
    """
    Set kualifikasi usaha ke Kecil (kualifikasiId=21) via POST /nontender/{kode}/simpan.
    headers: dict Cookie+User-Agent (sudah siap dari serap).
    Return True jika 302 redirect (sukses), False jika gagal.
    """
    import requests
    from bs4 import BeautifulSoup

    try:
        # Ambil authenticityToken dari halaman edit
        r_edit = requests.get(
            f"{base_url}nontender/{kode_paket}/edit",
            headers=headers, timeout=15,
        )
        soup = BeautifulSoup(r_edit.text, "html.parser")
        token_input = soup.find("input", {"name": "authenticityToken"})
        if not token_input:
            return False
        token = token_input.get("value", "")

        # POST simpan
        post_headers = {**headers, "Content-Type": "application/x-www-form-urlencoded"}
        payload = {
            "authenticityToken": token,
            "kualifikasiId": "21",   # 21 = Kecil
            "pl.oap": "1",
        }
        r_post = requests.post(
            f"{base_url}nontender/{kode_paket}/simpan",
            headers=post_headers,
            data=payload,
            timeout=15,
            allow_redirects=False,
        )
        return r_post.status_code in (301, 302)
    except Exception:
        return False


# Mapping metode pengadaan → (kategoriId, pilih)
METODE_PL_MAP = {
    "Pengadaan Barang — PL":              (0, 0),
    "Pekerjaan Konstruksi — PL":          (2, 3),
    "JKK Non-Konstruksi — PL":           (1, 9),
    "JKK Konstruksi — PL":               (5, 17),
    "JKK Perorangan Non-Konstruksi — PL": (4, 13),
    "JKK Perorangan Konstruksi — PL":    (6, 21),
    "Jasa Lainnya — PL":                 (3, 6),
    "PK Terintegrasi — PL":              (7, 25),
}


def ubah_metode_pl(
    kode_paket: str,
    kategori_id: int,
    pilih: int,
    cookie_str: str,
    base_url: str,
    debug: bool = False,
) -> bool:
    """
    Ubah metode pengadaan via POST /nontender/{kode}/metodesubmit.
    Return True jika 302/200, False jika gagal.
    debug=True: print status + body + semua form fields untuk investigasi.
    """
    import requests
    from bs4 import BeautifulSoup

    _UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.5 Safari/605.1.15"
    headers = {"Cookie": cookie_str, "User-Agent": _UA}
    try:
        r_form = requests.get(
            f"{base_url}nontender/{kode_paket}/metode",
            headers=headers, timeout=15,
        )
        soup = BeautifulSoup(r_form.text, "html.parser")

        if debug:
            # Dump semua input fields dari form /metode
            print(f"\n=== DEBUG ubah_metode_pl({kode_paket}) ===")
            print(f"GET /metode status: {r_form.status_code}")
            for inp in soup.find_all(["input", "select", "textarea"]):
                print(f"  field: name={inp.get('name')} type={inp.get('type')} value={inp.get('value','')}")
            # Dump radio/option pilih
            for opt in soup.find_all("option"):
                print(f"  option: value={opt.get('value')} text={opt.get_text(strip=True)}")
            # Dump form action
            for frm in soup.find_all("form"):
                print(f"  form: action={frm.get('action')} method={frm.get('method')}")
            # Dump title + tombol submit (cek onsubmit/onclick)
            title = soup.find("title")
            print(f"  page title: {title.get_text() if title else '(none)'}")
            for btn in soup.find_all(["button", "input"], type=lambda t: t in (None, "submit")):
                print(f"  btn: name={btn.get('name')} onclick={btn.get('onclick','')[:100]} onsubmit={btn.get('onsubmit','')}")
            for frm2 in soup.find_all("form"):
                print(f"  form onsubmit: {frm2.get('onsubmit','')[:100]}")
            # Dump raw HTML di sekitar select kategoriId (800 char)
            import re as _re
            m_sel = _re.search(r'(?s)(kategoriId.{0,800})', r_form.text)
            if m_sel:
                print(f"  HTML snippet kategoriId:\n{m_sel.group(1)[:800]}")
            # Dump POST response body (ikuti redirect)
            print(f"  === IKUTI REDIRECT ===")
            r_follow = requests.get(
                f"{base_url}nontender/{kode_paket}/edit",
                headers=headers, timeout=15,
            )
            soup2 = BeautifulSoup(r_follow.text, "html.parser")
            m2 = _re.search(r"Metode Pengadaan\s+(.+?)(?:Kualifikasi|$)", soup2.get_text(" ", strip=True))
            print(f"  Metode setelah POST: {m2.group(1)[:80] if m2 else '(tidak ketemu)'}")

        token_input = soup.find("input", {"name": "authenticityToken"})
        if not token_input:
            if debug:
                print("  ERROR: authenticityToken tidak ditemukan!")
                print(r_form.text[:500])
            return False
        token = token_input.get("value", "")

        post_headers = {
            **headers,
            "Referer": f"{base_url}nontender/{kode_paket}/edit",
            "Origin": "https://spse.inaproc.id",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "cache-control": "max-age=0",
            "priority": "u=0, i",
        }
        # Wajib multipart/form-data + field simpan=simpan (browser behavior)
        payload = {
            "authenticityToken": token,
            "kategoriId": str(kategori_id),
            "pilih": str(pilih),
            "simpan": "simpan",
        }

        if debug:
            print(f"  POST payload (multipart): {payload}")

        r_post = requests.post(
            f"{base_url}nontender/{kode_paket}/metodesubmit",
            headers=post_headers,
            files={k: (None, v) for k, v in payload.items()},  # multipart/form-data
            timeout=15,
            allow_redirects=False,
        )

        if debug:
            print(f"  POST status: {r_post.status_code}")
            print(f"  POST Location: {r_post.headers.get('Location','')}")
            print(f"  POST body (500 char): {r_post.text[:500]}")

        return r_post.status_code in (200, 301, 302)
    except Exception as e:
        if debug:
            print(f"  EXCEPTION: {e}")
        return False


def ubah_metode_pl_playwright(
    kode_paket: str,
    kategori_id: int,
    pilih: int,
    base_url: str,
) -> bool:
    """Ubah metode via Playwright CDP (handle JS confirm). Preferred over requests."""
    import spse_browser
    hasil = spse_browser.ubah_metode_via_playwright(kode_paket, kategori_id, pilih, base_url)
    return hasil == "OK"


def debug_ubah_metode_pl(kode_paket: str, cookie_str: str, base_url: str) -> str:
    """
    Helper debug: jalankan ubah_metode JKK Konstruksi dengan debug=True.
    Return string log untuk ditampilkan di UI.
    """
    import io, sys
    buf = io.StringIO()
    old_stdout = sys.stdout
    sys.stdout = buf
    ubah_metode_pl(kode_paket, 5, 9, cookie_str, base_url, debug=True)
    sys.stdout = old_stdout
    return buf.getvalue()


def ubah_ke_jkk_konstruksi_pl(kode_paket: str, cookie_str: str, base_url: str) -> bool:
    """Shortcut: ubah metode ke JKK Konstruksi PL (kategoriId=5, pilih=17) via CDP Playwright."""
    return ubah_metode_pl_playwright(kode_paket, 5, 17, base_url)


# ============================================================
# Download Dokumen Paket PL dari SPSE
# ============================================================

# Map label endpoint → subfolder rapi (dibuat on-demand saat ada file)
SUBFOLDER_DOK_PPK = {
    "KAK & Personil":            "1. KAK & Spesifikasi Teknis",
    "Rancangan Kontrak":         "2. Rancangan Kontrak",
    "Uraian Singkat Pekerjaan":  "3. Uraian Singkat Pekerjaan",
    "Informasi Lainnya":         "4. Informasi Lainnya",
    "Nota Dinas PPK":            "4. Informasi Lainnya",
}

_DOWNLOAD_PATH_LIMIT = 240


def _safe_download_name_for_folder(folder: str, filename: str, limit: int = _DOWNLOAD_PATH_LIMIT) -> str:
    """Bound downloaded filenames before Windows rejects an overlong path."""
    name = os.path.basename(str(filename or "").strip()).rstrip(" .") or "dokumen"
    folder_len = len(os.path.abspath(folder))
    available = int(limit) - folder_len - 1
    if available < 1:
        raise OSError(
            f"Folder path terlalu panjang untuk download ({folder_len} karakter; "
            f"batas {int(limit) - 1})"
        )
    if len(name) <= available:
        return name

    base, ext = os.path.splitext(name)
    digest = hashlib.sha1(name.encode("utf-8", "replace")).hexdigest()[:8]
    suffix = f"_{digest}{ext}"
    if available < len(digest) + len(ext):
        raise OSError(
            f"Folder path menyisakan ruang nama file terlalu pendek ({available} karakter)"
        )
    if available <= len(suffix):
        return f"{digest}{ext}"
    keep = available - len(suffix)
    prefix = base[:keep].rstrip(" ._") or "dokumen"
    return f"{prefix}{suffix}"


def buat_subfolder_dokumen(folder_paket: str) -> list:
    """Buat semua subfolder dokumen di folder_paket (0-9, selaras setup_paket_baru.py).
    Return list subfolder yang baru dibuat."""
    dibuat = []
    _semua_subfolder = [
        "0. Draft Dokumen PPK",
        "1. KAK & Spesifikasi Teknis",
        "2. Rancangan Kontrak",
        "3. Uraian Singkat Pekerjaan",
        "4. Informasi Lainnya",
        "5. Evaluator Kualifikasi & Teknis",
        "6. BA Reviu Lengkap",
        "7. Berita Acara + Summary Non Tender",
        "8. Dokumen Kualifikasi",
        "9. Dokumen Teknis Biaya",
    ]
    for sub in _semua_subfolder:
        p = os.path.join(folder_paket, sub)
        if not os.path.isdir(p):
            os.makedirs(p, exist_ok=True)
            dibuat.append(sub)
    return dibuat


def download_dokumen_paket_pl(
    kode_paket: str,
    folder_tujuan: str,
    progress_cb=None,
    cookie_str: str = "",
    skip_merge: bool = False,
    force_clean: bool = False,
    per_file_workers: int = 3,
    download_timeout: int = 15,
) -> dict:
    """
    Download dokumen dari endpoint non-tender PP ke folder_tujuan:
      - /dokumennontender/{kode}/spek  → KAK, Daftar Personil, RAB
      - /dokumennontender/{kode}/docsskk → Rancangan SPK/SPMK/SSUK/SSKK

    Pakai cookie PP via spse_browser.get_spse_cookies() — bisa juga di-pass
    eksplisit lewat parameter cookie_str (untuk paralel: hindari race init Playwright).

    skip_merge=True: lewati gabung PDF (Excel COM tidak thread-safe untuk paralel).
                     Merge dilakukan sequential setelah pool selesai via gabung_draft_pl().
    force_clean=True: hapus semua file di dalam SUBFOLDER_DOK_PPK sebelum download.

    Return: {"ok": [...], "error": [...]}
    """
    import requests
    import urllib.parse
    import glob as _glob
    import threading
    import time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from bs4 import BeautifulSoup
    import spse_browser

    def log(msg):
        if progress_cb:
            progress_cb(msg)

    os.makedirs(folder_tujuan, exist_ok=True)
    hasil = {"ok": [], "error": []}
    _write_lock = threading.Lock()

    if force_clean:
        for sub_name in SUBFOLDER_DOK_PPK.values():
            sub_path = os.path.join(folder_tujuan, sub_name)
            if os.path.isdir(sub_path):
                for f in os.listdir(sub_path):
                    fp = os.path.join(sub_path, f)
                    if os.path.isfile(fp):
                        os.remove(fp)

    if not cookie_str:
        cookie_str = spse_browser.get_spse_cookies()
    if not cookie_str:
        hasil["error"].append("Cookie SPSE kosong — buka Brave SPSE dan login ulang.")
        return hasil

    hdrs = {
        "Cookie": cookie_str,
        "User-Agent": "Mozilla/5.0",
        "Referer": f"{BASE_URL}/admin/pegawai",
    }

    def _unique_dst(folder, fname):
        fname = _safe_download_name_for_folder(folder, fname)
        dst = os.path.join(folder, fname)
        if not os.path.exists(dst):
            return dst
        base, ext = os.path.splitext(fname)
        n = 2
        while True:
            candidate_name = _safe_download_name_for_folder(folder, f"{base}_{n}{ext}")
            candidate = os.path.join(folder, candidate_name)
            if not os.path.exists(candidate):
                return candidate
            n += 1

    def _fix_customhostname_url(url):
        parsed = urllib.parse.urlsplit(url)
        if parsed.hostname != "customhostname":
            return url, hdrs
        path = parsed.path if parsed.path.startswith("/lpse-prod-data/") else "/lpse-prod-data" + parsed.path
        fixed = urllib.parse.urlunsplit((parsed.scheme, "storage.googleapis.com", path, parsed.query, parsed.fragment))
        fixed_hdrs = {k: v for k, v in hdrs.items() if k.lower() not in ("cookie", "host")}
        log("    ↪ customhostname → storage.googleapis.com/lpse-prod-data")
        return fixed, fixed_hdrs

    def _get_download_response(url):
        current = url
        for _ in range(5):
            current, req_hdrs = _fix_customhostname_url(current)
            resp = requests.get(current, headers=req_hdrs, timeout=download_timeout, stream=True, allow_redirects=False)
            if resp.status_code not in (301, 302, 303, 307, 308):
                return resp
            loc = resp.headers.get("Location")
            if not loc:
                return resp
            current = urllib.parse.urljoin(current, loc)
        current, req_hdrs = _fix_customhostname_url(current)
        return requests.get(current, headers=req_hdrs, timeout=download_timeout, stream=True, allow_redirects=False)

    def _get_download_response_retry(url):
        for i, delay in enumerate((0, 0.7, 1.5), start=1):
            if delay:
                time.sleep(delay)
            resp = _get_download_response(url)
            if resp.status_code not in (404, 429, 500, 502, 503, 504):
                return resp
            if i < 3:
                resp.close()
                log(f"    ↻ retry download {i}/2 (HTTP {resp.status_code})")
        return resp

    def _download_links_dari_endpoint(endpoint_url, label):
        """Scrape link /dl/ dari endpoint, download semua file ke subfolder rapi."""
        r = None
        try:
            r = requests.get(endpoint_url, headers=hdrs, timeout=15)
            if r.status_code in (401, 403) or r.status_code >= 500:
                err = f"HTTP {r.status_code} — sesi SPSE tidak valid atau server gagal"
                hasil["error"].append(f"{label}: {err}")
                log(f"  ❌ {label}: {err}")
                return
            r.raise_for_status()
            response_text = r.text
            soup = BeautifulSoup(response_text, "html.parser")
            lowered = response_text.lower()
            response_path = urllib.parse.urlsplit(getattr(r, "url", endpoint_url)).path.lower().rstrip("/")
            login_page = response_path.endswith("/login") or any(
                marker in lowered
                for marker in ("name=\"username\"", "id=\"username\"", "silakan login")
            )
            if login_page:
                err = "Sesi SPSE tidak valid — server mengembalikan halaman login"
                hasil["error"].append(f"{label}: {err}")
                log(f"  ❌ {label}: {err}")
                return
            links = []
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "/dl/" not in href:
                    continue
                fname_raw = a.get_text(strip=True)
                fname_raw = re.sub(r"\s*-\s*\d+\s*[KkMm][Bb]\s*$", "", fname_raw, re.IGNORECASE).strip()
                fname = re.sub(r'[<>:"/\\|?*]', "_", fname_raw).strip() or "dokumen"
                url_dl = f"https://spse.inaproc.id{href}" if href.startswith("/") else href
                links.append((url_dl, fname))

            # Subfolder tujuan (on-demand: dibuat hanya jika ada file)
            sub = SUBFOLDER_DOK_PPK.get(label, "4. Informasi Lainnya")
            folder_dl = os.path.join(folder_tujuan, sub)
            if links:
                os.makedirs(folder_dl, exist_ok=True)

            log(f"  📂 {label}: {len(links)} file")

            def _download_satu(link):
                url_dl, fname = link
                t0 = time.perf_counter()
                r_dl = None
                try:
                    r_dl = _get_download_response_retry(url_dl)
                    r_dl.raise_for_status()
                    ct = r_dl.headers.get("Content-Type", "")
                    if "text/html" in ct:
                        return False, fname, "session expired (server return HTML)", time.perf_counter() - t0
                    cd = r_dl.headers.get("Content-Disposition", "")
                    m_cd = re.search(r'filename[^;=\n]*=["\']?([^"\';\n]+)', cd)
                    if m_cd:
                        clean = re.sub(r'[<>:"/\\|?*]', "_", urllib.parse.unquote_plus(m_cd.group(1).strip())).strip()
                        if clean:
                            fname = clean
                    original_fname = fname
                    fname = _safe_download_name_for_folder(folder_dl, fname)
                    if fname != original_fname:
                        log(f"    ⚠️ nama file dipendekkan ({len(original_fname)}→{len(fname)} karakter): {fname}")
                    with _write_lock:
                        dst = _unique_dst(folder_dl, fname)
                        with open(dst, "wb") as f:
                            for chunk in r_dl.iter_content(65536):
                                f.write(chunk)
                        hasil["ok"].append(dst)
                    return True, os.path.basename(dst), "", time.perf_counter() - t0
                except Exception as e:
                    return False, fname, str(e), time.perf_counter() - t0
                finally:
                    if r_dl is not None:
                        close = getattr(r_dl, "close", None)
                        if callable(close):
                            close()

            t_ep = time.perf_counter()
            max_workers = max(1, min(int(per_file_workers or 1), len(links) or 1))
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                for fut in as_completed([ex.submit(_download_satu, link) for link in links]):
                    ok, fname, err, dur = fut.result()
                    if ok:
                        log(f"    ✅ {fname} ({dur:.1f}s)")
                    else:
                        hasil["error"].append(f"{fname}: {err}")
                        log(f"    ❌ {fname}: {err} ({dur:.1f}s)")
            log(f"  ⏱ {label}: {len(links)} file, {time.perf_counter() - t_ep:.1f}s")
        except Exception as e:
            hasil["error"].append(f"{label}: {e}")
            log(f"  ❌ {label}: {e}")
        finally:
            if r is not None:
                close = getattr(r, "close", None)
                if callable(close):
                    close()

    ENDPOINTS = [
        (f"{BASE_URL}/dokumennontender/{kode_paket}/spek",      "KAK & Personil"),
        (f"{BASE_URL}/dokumennontender/{kode_paket}/docsskk",   "Rancangan Kontrak"),
        (f"{BASE_URL}/dokumennontender/{kode_paket}/docuraian", "Uraian Singkat Pekerjaan"),
        (f"{BASE_URL}/dokumennontender/{kode_paket}/lainnya",   "Informasi Lainnya"),
        (f"{BASE_URL}/nontender/{kode_paket}/edit",             "Nota Dinas PPK"),
    ]

    for url_ep, label_ep in ENDPOINTS:
        _download_links_dari_endpoint(url_ep, label_ep)

    log(f"🏁 Download selesai: {len(hasil['ok'])} file OK, {len(hasil['error'])} error")

    # ── Catat basename PDF uraian singkat ke Supabase
    try:
        for fpath in hasil["ok"]:
            bn = os.path.basename(fpath).lower()
            if "uraian" in bn and bn.endswith(".pdf"):
                _sb().table("draft_paket_pl").update(
                    {"nama_file_uraian": os.path.basename(fpath)}
                ).eq("kode_paket", kode_paket).execute()
                log(f"  📝 nama_file_uraian: {os.path.basename(fpath)}")
                break
    except Exception as e:
        log(f"  ⚠ gagal simpan nama_file_uraian: {e}")

    # ── Gabung semua PDF jadi 1 draft (tiru flow tender)
    if not skip_merge:
        try:
            merged = gabung_draft_pl(kode_paket, folder_tujuan, hasil["ok"], progress_cb)
            if merged:
                hasil["draft_pdf"] = merged
                log(f"📎 Draft PDF gabungan: {os.path.basename(merged)}")
        except Exception as e:
            import traceback as _tb
            log(f"❌ Gagal gabung PDF: {e}")
            log(f"   {_tb.format_exc()[-300:]}")
            hasil["error"].append(f"Gabung PDF: {e}")

    return hasil


def gabung_draft_pl(kode_paket: str, folder_tujuan: str, files_ok: list, progress_cb=None) -> str:
    """Standalone gabung PDF — dipanggil sequential setelah bulk parallel download.

    Excel/Word COM tidak thread-safe → harus serial.
    Return: path Draft_PL_*.pdf atau "" jika gagal.
    """
    from inbox_engine import _gabung_pdf_draft

    def log(msg):
        if progress_cb:
            progress_cb(msg)

    nama_paket_row = _sb().table("draft_paket_pl").select("nama_paket").eq(
        "kode_paket", kode_paket
    ).maybe_single().execute()
    nama_paket = (nama_paket_row.data or {}).get("nama_paket", kode_paket) if nama_paket_row else kode_paket
    nama_clean = re.sub(r'[<>:"/\\|?*]', "_", nama_paket)[:60].strip()
    draft_dir = os.path.join(folder_tujuan, "0. Draft Dokumen PPK")
    os.makedirs(draft_dir, exist_ok=True)
    draft_path = os.path.join(draft_dir, f"Draft_PL_{nama_clean}.pdf")
    ordered = sorted(files_ok, key=lambda p: _pl_pdf_sort_key(os.path.basename(p)))
    return _gabung_pdf_draft(draft_path, ordered, progress_cb)


def _pl_pdf_sort_key(fname: str) -> tuple:
    """Urutan gabung draft PL: KAK → RAB/Personil → Rancangan → Uraian → Lainnya → Nota."""
    f = fname.lower()
    if "kak" in f: return (0, f)
    if "rab" in f: return (1, f)
    if "personil" in f or "personel" in f: return (2, f)
    if "rincian" in f or "prn" in f: return (3, f)
    if "rancangan" in f or "sskk" in f or "ssuk" in f or "spk" in f or "spmk" in f: return (4, f)
    if "uraian" in f: return (5, f)
    if "rekomendasi" in f or "lainnya" in f: return (6, f)
    if "permohonan" in f or "nota" in f: return (7, f)
    return (9, f)


def umumkan_paket_pl(kode_paket: str, cookie_str: str) -> dict:
    """
    Umumkan paket non-tender ke SPSE via POST /nontender/{kode}/pengumumanpp.
    Langkah: GET edit page → ambil authenticityToken → POST pengumumanpp.
    Return: {ok: bool, pesan: str, status_code: int}
    """
    import requests
    from bs4 import BeautifulSoup
    from config import SPSE_BASE_URL

    headers = {
        'Cookie': cookie_str,
        'User-Agent': 'Mozilla/5.0',
    }

    # Step 1: GET halaman edit → ambil authenticityToken
    try:
        resp_get = requests.get(
            f"{SPSE_BASE_URL}nontender/{kode_paket}/edit",
            headers=headers,
            timeout=15
        )
        soup = BeautifulSoup(resp_get.text, 'html.parser')
        token_input = soup.find('input', {'name': 'authenticityToken'})
        if not token_input:
            return {'ok': False, 'pesan': 'authenticityToken tidak ditemukan di halaman edit', 'status_code': resp_get.status_code}
        token = token_input.get('value', '')
    except Exception as e:
        return {'ok': False, 'pesan': f'GET edit gagal: {e}', 'status_code': 0}

    # Step 2: POST pengumumanpp
    try:
        payload = {
            'authenticityToken': token,
            'alasan': '',
            'setuju': 'setuju',
        }
        resp_post = requests.post(
            f"{SPSE_BASE_URL}nontender/{kode_paket}/pengumumanpp",
            data=payload,
            headers={
                **headers,
                'Referer': f"{SPSE_BASE_URL}nontender/{kode_paket}/edit",
                'Content-Type': 'application/x-www-form-urlencoded',
            },
            allow_redirects=False,
            timeout=15
        )
        if resp_post.status_code == 302:
            location = resp_post.headers.get('Location', '')
            # Cek apakah redirect ke halaman error
            if 'error' in location.lower() or 'login' in location.lower() or location == '/':
                return {'ok': False, 'pesan': f'Redirect ke {location} — kemungkinan session expired atau gagal', 'status_code': 302}
            return {'ok': True, 'pesan': f'Paket {kode_paket} berhasil diumumkan → {location}', 'status_code': 302, 'location': location}
        elif resp_post.status_code == 200:
            # Cek apakah ada pesan error di body
            soup_resp = BeautifulSoup(resp_post.text, 'html.parser')
            body_text = soup_resp.get_text(' ', strip=True)[:300]
            return {'ok': False, 'pesan': f'HTTP 200 (bukan redirect) — mungkin error: {body_text}', 'status_code': 200}
        else:
            return {'ok': False, 'pesan': f'POST gagal: HTTP {resp_post.status_code}', 'status_code': resp_post.status_code}
    except Exception as e:
        return {'ok': False, 'pesan': f'POST pengumumanpp gagal: {e}', 'status_code': 0}


def umumkan_pemenang_pl(kode_paket: str, cookie_str: str) -> dict:
    """Kirim Pengumuman Pemenang non-tender via halaman undangan SPSE."""
    import requests
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin
    from config import SPSE_BASE_URL

    base = f"{SPSE_BASE_URL}nontender/{kode_paket}"
    headers = {
        "Cookie": cookie_str,
        "User-Agent": "Mozilla/5.0",
        "Referer": base,
    }
    try:
        resp_get = requests.get(f"{base}/undangan", headers=headers, timeout=15)
        soup = BeautifulSoup(resp_get.text, "html.parser")
        form = next(
            (f for f in soup.find_all("form")
             if f.find("input", {"name": "authenticityToken"})),
            None,
        )
        token = form.find("input", {"name": "authenticityToken"}) if form else None
        if resp_get.status_code != 200 or not form or not token:
            return {
                "ok": False,
                "pesan": f"Halaman pengumuman gagal dibaca (HTTP {resp_get.status_code})",
                "status_code": resp_get.status_code,
            }

        winners = []
        for row in soup.select("form table tr")[1:]:
            cells = row.find_all("td")
            if len(cells) >= 3 and row.select_one(".fa-check-circle"):
                winners.append({
                    "nama": cells[1].get_text(" ", strip=True),
                    "email": cells[2].get_text(" ", strip=True),
                })
        if not winners:
            return {"ok": False, "pesan": "Pemenang tidak ditemukan di halaman pengumuman", "status_code": 200}

        action = urljoin(f"{SPSE_BASE_URL}nontender/{kode_paket}/", form.get("action", ""))
        resp_post = requests.post(
            action,
            data={"authenticityToken": token.get("value", ""), "simpan": "simpan"},
            headers={**headers, "Content-Type": "application/x-www-form-urlencoded"},
            allow_redirects=False,
            timeout=15,
        )
        location = resp_post.headers.get("Location", "")
        if resp_post.status_code == 302 and not any(x in location.lower() for x in ("login", "error")):
            return {"ok": True, "pesan": "Pengumuman pemenang berhasil dikirim", "status_code": 302, "winners": winners}
        return {
            "ok": False,
            "pesan": f"POST pengumuman gagal (HTTP {resp_post.status_code})",
            "status_code": resp_post.status_code,
            "winners": winners,
        }
    except Exception as e:
        return {"ok": False, "pesan": f"Pengumuman pemenang gagal: {e}", "status_code": 0}
