"""Asisten Pokja — SPSE Automation (Streamlit)."""

import os
import glob as _glob_mod
import json
import pathlib
import re
import shutil
import sqlite3
import sys
import threading
import time
from datetime import datetime, timedelta, date
from concurrent.futures import ThreadPoolExecutor, as_completed
import streamlit as st
from ui_state import activate_mode
from pl_data_ui import (
    fetch_peserta_pl_cached as _fetch_peserta_pl_cached,
    fetch_status_semua_paket_cached as _fetch_status_semua_paket_cached,
    load_draft_pl_cached as _load_draft_pl_cached,
    load_verifikasi_pl_rows_cached as _load_verifikasi_pl_rows_cached,
    parse_jadwal_pl_cached as _parse_jadwal_pl_cached,
)
from ui_pl_pk import (
    PLPK_TAB_LABELS,
    active_rows as active_plpk_rows,
    get_engine as get_plpk_engine,
    render_download_actions,
)
from ui_pl_common import render_package_selection

APP_VERSION = "v2026.07.19.a"

_PL_UNDANGAN_DATES_PATH = pathlib.Path(__file__).resolve().parent / "data" / "pl_undangan_dates.json"
_PL_SBU_HISTORY_PATH = pathlib.Path(__file__).resolve().parent / "data" / "pl_sbu_history.json"


def _load_pl_sbu_history() -> list[dict[str, str]]:
    """Load histori SBU custom PL; format lama berupa string tetap didukung."""
    try:
        raw = json.loads(_PL_SBU_HISTORY_PATH.read_text(encoding="utf-8"))
        result = []
        for item in raw if isinstance(raw, list) else []:
            if isinstance(item, dict):
                baru, lama = str(item.get("baru") or "").strip(), str(item.get("lama") or "").strip()
            else:
                baru, lama = str(item).strip(), ""
            if baru and not any(x["baru"] == baru and x["lama"] == lama for x in result):
                result.append({"baru": baru, "lama": lama})
        return result[:20]
    except (OSError, TypeError, ValueError):
        return []


def _save_pl_sbu_history(baru: str, lama: str = "") -> None:
    baru, lama = str(baru or "").strip(), str(lama or "").strip()
    if not baru:
        return
    history = [x for x in _load_pl_sbu_history() if x["baru"] != baru or x["lama"] != lama]
    history.insert(0, {"baru": baru, "lama": lama})
    try:
        _PL_SBU_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        _PL_SBU_HISTORY_PATH.write_text(json.dumps(history[:20], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    except OSError:
        pass


def _load_last_pl_invitation_dates() -> dict[str, str]:
    """Load tanggal undangan terakhir per kode paket; file rusak dianggap kosong."""
    try:
        data = json.loads(_PL_UNDANGAN_DATES_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, ValueError, TypeError):
        return {}


def _last_pl_invitation_date(kode_paket: str) -> date | None:
    raw = _load_last_pl_invitation_dates().get(str(kode_paket))
    try:
        return date.fromisoformat(raw) if raw else None
    except (TypeError, ValueError):
        return None


def _save_last_pl_invitation_date(kode_paket: str, tanggal: date) -> None:
    """Simpan tanggal hanya setelah kirim sukses; gagal simpan tidak mengganggu UI."""
    try:
        _PL_UNDANGAN_DATES_PATH.parent.mkdir(parents=True, exist_ok=True)
        data = _load_last_pl_invitation_dates()
        data[str(kode_paket)] = tanggal.isoformat()
        tmp = _PL_UNDANGAN_DATES_PATH.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        tmp.replace(_PL_UNDANGAN_DATES_PATH)
    except (OSError, TypeError, ValueError):
        pass
from ui_dpa import render_tab_dpa as _render_tab_dpa
# Display labels are numbered from the physical package folders.
from pl_ui_helpers import _baca_master_data_pl, _cari_xlsm_pl, _fmt_elapsed, _fmt_step_seconds, _pl_hint_ulang, _pl_label, _pl_paket_ulang, _pl_proses_io_satu_paket, _proses_excel_paket_pl, _template_dir_pl_jkk, _pl_workflow, mark_workflow_applied, migrate_pl_workflow
from ui_pl_jadwal import _render_ubah_jadwal_pl
from ui_pl_penetapan import _render_tab10_pl
from ui_sidebar import _get_dark_css, _get_light_css, _sidebar_login_form

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Suppress Playwright/CDP/Node stderr noise di console Streamlit
import io as _io
_devnull = open(os.devnull, "w", encoding="utf-8")


class _SuppressPlaywrightStderr:
    """Context manager: mute stderr selama operasi CDP agar terminal tidak kedap-kedip."""
    def __enter__(self):
        self._orig = sys.stderr
        sys.stderr = _devnull
        return self
    def __exit__(self, *_):
        sys.stderr = self._orig


def _pokja_label(p: dict) -> str:
    """Buat label ringkas paket dengan nomor urut folder bila tersedia."""
    pokja_raw = p.get("pokja") or ""
    m = re.search(r"\d+", pokja_raw)
    pokja_no = m.group() if m else p.get("kode", "")
    nomor = str(p.get("nomor_urut") or "").strip()
    if not nomor:
        _folder_m = re.match(r"^\s*(\d+)\s*\.\s*", str(p.get("folder_dibuat") or ""))
        nomor = _folder_m.group(1) if _folder_m else ""
    nama = str(p.get("nama") or "")
    if nomor and not nama.startswith(f"{nomor}."):
        nama = f"{nomor}. {nama}"
    return f"Pokja {pokja_no} · {p['kode']} — {nama}"


def _tender_display_label(row: dict) -> str:
    """Label Tender untuk row API/DB maupun row UI yang sudah dinormalisasi."""
    if "kode" in row and "nama" in row:
        return _pokja_label(row)
    nomor = str(row.get("nomor_urut") or "").strip()
    if not nomor:
        m = re.match(r"^\s*(\d+)\s*\.\s*", str(row.get("folder_dibuat") or ""))
        nomor = m.group(1) if m else ""
    nama = str(row.get("nama_tender") or row.get("nama_paket") or row.get("nama") or row.get("kode_tender") or "-")
    return f"{nomor}. {nama}" if nomor else nama


def _get_paket_gabungan(filter_selesai: bool = True) -> list[dict]:
    """Gabung global_paket_draft + global_paket_aktif, deduplikasi by kode.
    filter_selesai=True (default): sembunyikan paket yang sudah Hasil Evaluasi/Masa Sanggah/dll.
    Status tahap dibaca dari session state 'tender_tahap_map' (di-isi saat enrich_paket_supabase).
    """
    draft_list = st.session_state.get("global_paket_draft", {}).get("paket", [])
    aktif_list = st.session_state.get("global_paket_aktif", {}).get("paket", [])
    tahap_map = st.session_state.get("tender_tahap_map", {})
    # DataTables SPSE tidak membawa nomor folder; ambil metadata lokal/cache DB.
    _folder_meta = {}
    try:
        _folder_meta = {
            str(r.get("kode_tender")): r
            for r in _load_draft_paket_cached()
            if r.get("kode_tender")
        }
    except Exception:
        pass
    seen, result = set(), []
    for p in draft_list + aktif_list:
        if p["kode"] not in seen:
            seen.add(p["kode"])
            if _is_tender_excluded(p):
                continue
            if filter_selesai:
                tahap = tahap_map.get(p["kode"]) or p.get("status") or ""
                if _is_tender_selesai({"status": tahap}):
                    continue
            _meta = _folder_meta.get(str(p["kode"]), {})
            result.append({
                **p,
                "nomor_urut": p.get("nomor_urut") or _meta.get("nomor_urut") or "",
                "folder_dibuat": p.get("folder_dibuat") or _meta.get("folder_dibuat") or "",
            })
    return result


def _enrich_kode_unik_tender_excel(
    paket: dict,
    *,
    folder_dibuat: str | None = None,
    skip_folder_lookup: bool = False,
) -> tuple[str, str, str]:
    """Baca G2 workbook Tender sebagai source of truth, tanpa menyimpan XLSM.

    Return: (kode_excel, kode_db, status). Status dipakai UI untuk memberi
    peringatan jika Excel dan cache Supabase berbeda.
    """
    kode_db = str(paket.get("kode_unik") or "").strip()
    try:
        from config import TENDER_ROOT as _KU_TENDER_ROOT, sb as _KU_SB
        import os as _ku_os
        import re as _ku_re
        if skip_folder_lookup:
            folder_name = str(folder_dibuat or "").strip()
        else:
            # Supabase hanya sumber metadata tambahan. Credential bisa belum
            # tersedia di PC kantor; jangan hentikan resolver Excel karenanya.
            try:
                rows = (
                    _KU_SB().table("draft_paket")
                    .select("folder_dibuat")
                    .eq("kode_tender", str(paket.get("kode") or paket.get("id_lelang") or ""))
                    .limit(1).execute().data or []
                )
                folder_name = str((rows[0] if rows else {}).get("folder_dibuat") or "").strip()
            except Exception:
                folder_name = ""
        # Metadata folder bisa kosong/stale di cache lama. Susun kandidat dari
        # DB terlebih dahulu, lalu fallback ke nama paket/Pokja di disk.
        _folder_names = [folder_name] if folder_name else []
        _folder_names = [str(x).strip() for x in _folder_names if str(x).strip()]
        _nama_norm = _ku_re.sub(r"[^a-z0-9]+", "", str(paket.get("nama") or "").lower())
        _pokja_match = _ku_re.search(r"\d+", str(paket.get("pokja") or paket.get("kode_pokja") or ""))
        _pokja_norm = _pokja_match.group(0).lstrip("0") if _pokja_match else ""
        try:
            _disk_dirs = [
                d for d in _ku_os.listdir(_KU_TENDER_ROOT)
                if _ku_os.path.isdir(_ku_os.path.join(_KU_TENDER_ROOT, d))
                and "backup" not in d.lower()
            ]
            for _d in _disk_dirs:
                _dn = _ku_re.sub(r"[^a-z0-9]+", "", _d.lower())
                if ((_nama_norm and (_nama_norm in _dn or _dn in _nama_norm))
                        or (_pokja_norm and _ku_re.search(r"pokja\s*0*" + _pokja_norm, _d, _ku_re.I))):
                    if _d not in _folder_names:
                        _folder_names.append(_d)
        except OSError:
            pass

        _kode_tender_target = str(paket.get("kode") or paket.get("id_lelang") or "").strip()
        _kode_tender_digits = _ku_re.sub(r"\D", "", _kode_tender_target)
        for _folder_candidate in _folder_names:
            _folder_candidate = _ku_re.sub(r'[/\\:*?"<>|]', "-", _folder_candidate).strip()
            folder = _ku_os.path.join(_KU_TENDER_ROOT, _folder_candidate)
            try:
                candidates = sorted(
                    p for p in _ku_os.listdir(folder)
                    if p.lower().endswith(".xlsm")
                    and "backup" not in p.lower()
                    and ".bak_" not in p.lower()
                )
            except OSError:
                continue
            for _candidate in candidates:
                path = _ku_os.path.join(folder, _candidate)
                stat = _ku_os.stat(path)
                kode_excel = _read_tender_master_data_cached(path, stat.st_mtime_ns, stat.st_size)
                if kode_excel:
                    return kode_excel, kode_db, "excel" if kode_excel == kode_db else "beda"

        # Fallback paling kuat: cari workbook berdasarkan Kode Tender pada
        # @ Master Data!C4. Ini mengatasi metadata folder/cache yang kosong,
        # stale, atau berbeda format; C4 adalah identitas paket yang stabil.
        if _kode_tender_target:
            try:
                for _d in _ku_os.listdir(_KU_TENDER_ROOT):
                    _dir_path = _ku_os.path.join(_KU_TENDER_ROOT, _d)
                    if (not _ku_os.path.isdir(_dir_path)
                            or "backup" in _d.lower()):
                        continue
                    for _candidate in sorted(_ku_os.listdir(_dir_path)):
                        if (not _candidate.lower().endswith(".xlsm")
                                or "backup" in _candidate.lower()
                                or ".bak_" in _candidate.lower()):
                            continue
                        _path = _ku_os.path.join(_dir_path, _candidate)
                        _stat = _ku_os.stat(_path)
                        _kode_file, _kode_g2 = _read_tender_identity_cached(
                            _path, _stat.st_mtime_ns, _stat.st_size
                        )
                        _kode_file_digits = _ku_re.sub(r"\D", "", _kode_file)
                        if ((_kode_file == _kode_tender_target or
                             (_kode_tender_digits and _kode_file_digits == _kode_tender_digits))
                                and _kode_g2):
                            return _kode_g2, kode_db, "excel" if _kode_g2 == kode_db else "beda"
            except OSError:
                pass
        return "", kode_db, "workbook/G2 tidak ditemukan"
    except Exception as exc:
        return "", kode_db, f"gagal baca Excel: {exc}"


@st.cache_data(ttl=300)
def _read_tender_master_data_cached(path: str, mtime_ns: int, size: int) -> str:
    """Baca G2 workbook Tender sekali per versi file."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        return str(wb["@ Master Data"]["G2"].value or "").strip()
    finally:
        wb.close()


@st.cache_data(ttl=300)
def _read_tender_identity_cached(path: str, mtime_ns: int, size: int) -> tuple[str, str]:
    """Baca identitas paket C4 dan Kode Unik G2 dari workbook Tender."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        ws = wb["@ Master Data"]
        return str(ws["C4"].value or "").strip(), str(ws["G2"].value or "").strip()
    finally:
        wb.close()


def _tender_folder_identity_valid(folder_name: str, kode_tender: str) -> bool:
    """True bila workbook utama folder menyimpan kode tender target di C4."""
    try:
        from config import TENDER_ROOT as _root
        _kode_digits = re.sub(r"\D", "", str(kode_tender or ""))
        if not (_kode_digits and folder_name):
            return False
        _folder = os.path.join(_root, folder_name)
        if not os.path.isdir(_folder):
            return False
        for _f in os.listdir(_folder):
            if not (_f.lower().endswith(".xlsm") and ".bak_" not in _f.lower()):
                continue
            _path = os.path.join(_folder, _f)
            _stat = os.stat(_path)
            _kode_file, _ = _read_tender_identity_cached(
                _path, _stat.st_mtime_ns, _stat.st_size
            )
            if re.sub(r"\D", "", _kode_file) == _kode_digits:
                return True
    except Exception:
        pass
    return False


def _find_tender_folder_local(kode_tender: str, kode_pokja: str = "", nama: str = "") -> str:
    """Cari folder Tender valid di disk saat metadata DB belum lengkap."""
    try:
        from config import TENDER_ROOT as _root
        _kode_digits = re.sub(r"\D", "", str(kode_tender or ""))
        _pokja_digits = re.sub(r"\D", "", str(kode_pokja or "")).lstrip("0")
        _nama_norm = re.sub(r"[^a-z0-9]+", "", str(nama or "").lower())
        _dirs = [
            d for d in os.listdir(_root)
            if os.path.isdir(os.path.join(_root, d)) and "backup" not in d.lower()
        ]
        # Nama/Pokja hanya kandidat lokasi. Jangan menganggap folder berhasil
        # bila workbook masih kosong; biarkan create-folder berikutnya retry.
        for _d in _dirs:
            _dn = re.sub(r"[^a-z0-9]+", "", _d.lower())
            _nama_match = _nama_norm and (_nama_norm in _dn or _dn in _nama_norm)
            _pokja_match = _pokja_digits and re.search(r"pokja0*" + _pokja_digits, _dn)
            if (_nama_match or _pokja_match) and (
                not _kode_digits or _tender_folder_identity_valid(_d, kode_tender)
            ):
                return _d
        # Identitas workbook menjadi fallback paling akurat untuk paket tanpa
        # nomor Pokja di metadata SPSE.
        if _kode_digits:
            for _d in _dirs:
                _folder = os.path.join(_root, _d)
                for _f in os.listdir(_folder):
                    if not (_f.lower().endswith(".xlsm") and ".bak_" not in _f.lower()):
                        continue
                    _path = os.path.join(_folder, _f)
                    _stat = os.stat(_path)
                    _kode_file, _ = _read_tender_identity_cached(_path, _stat.st_mtime_ns, _stat.st_size)
                    if re.sub(r"\D", "", _kode_file) == _kode_digits:
                        return _d
    except Exception:
        pass
    return ""


def _enrich_kode_unik_pl_excel(row: dict) -> str:
    """Pastikan alur BA PL memakai F2 workbook sebagai kode unik utama.

    ``_load_draft_pl_cached`` biasanya sudah melakukan enrichment, tetapi tab
    BA juga dipanggil setelah cache lama masih hidup. Karena itu baca ulang F2
    hanya pada saat tab BA dirender. Return status untuk peringatan UI.
    """
    kode_db = str(row.get("kode_unik") or "").strip()
    try:
        kode_excel = str(_baca_master_data_pl(row).get("kode_unik") or "").strip()
        if kode_excel:
            row["kode_unik"] = kode_excel
            return "excel" if kode_excel == kode_db else "beda"
        return "F2 kosong; memakai fallback DB" if kode_db else "F2 kosong"
    except Exception as exc:
        return f"gagal baca Excel: {exc}"

_TENDER_SELESAI_KW = (
    "surat penunjukan penyedia", "penunjukan penyedia",
    "penandatanganan kontrak", "penandatanganan",
    "tender sudah selesai", "selesai",
)

def _is_tender_selesai(p: dict) -> bool:
    """True jika paket Tender sudah di tahap akhir (Masa Sanggah / Penunjukan / Penandatanganan)."""
    status_tahap = str(p.get("status_tahap") or "").lower()
    if status_tahap:
        return any(k in status_tahap for k in _TENDER_SELESAI_KW)
    status = str(p.get("status") or "").lower()
    return any(k in status for k in _TENDER_SELESAI_KW)


_TENDER_SUDAH_PEMBUKAAN_KW = (
    "pembukaan",
    "evaluasi",
    "klarifikasi",
    "negosiasi",
    "masa sanggah",
    "penetapan",
    "pemenang",
    "penandatanganan",
)


def _is_tender_sudah_pembukaan(p: dict) -> bool:
    """True jika paket sudah masuk pembukaan penawaran atau tahap sesudahnya."""
    tahap = str(p.get("status_tahap") or "").strip().lower()
    if tahap:
        return any(k in tahap for k in _TENDER_SUDAH_PEMBUKAAN_KW)
    # Jika status_tahap belum tersimpan, jangan izinkan paket masuk secara
    # spekulatif. Tab ini memindahkan dokumen penawaran; tahap harus terbukti.
    status = str(p.get("status") or "").strip().lower()
    return bool(status) and "draft" not in status

from config import SPSE_BASE_URL
import spse_browser
import ldk_engine
import ldk_config
import checklist_engine
import masa_berlaku_engine
import tender_setup_engine
import penjelasan_engine
import penjelasan_config
import jadwal_engine
import jadwal_config
import kirimpesan_engine
import merge_engine
import bareviu_engine
import ba_engine
import ba_config
import kualifikasi_engine
import kualifikasi_parser
import kk_evaluasi_engine
import pl_engine
_pl_display_engine = pl_engine
import parse_kak_pl
import pl_kirimpesan_engine
import dokpil_engine_pl as _depl_jkk
import dokpil_engine_plpk as _depl_pk
import upload_dokpil_pl as _udpl
import kualifikasi_engine_pl as _ke_pl_jkk
import kualifikasi_engine_plpk as _ke_pl_pk
# Alias resolver default dipakai lintas Tab 6/7/8. Jangan hanya didefinisikan
# di blok render Tab 6 karena user bisa langsung membuka Tab 7/8.
_ke_pl = _ke_pl_jkk
import kualifikasi_parser_pl as _kp_pl
import hasil_evaluasi_pl_engine as _he_pl_jkk
import hasil_evaluasi_plpk_engine as _he_pl_pk
import sbu_picker as _sp_global








@st.cache_data(ttl=30, show_spinner=False)
def _get_jadwal_penjelasan_gcal_cached() -> dict:
    """Cache daftar jadwal penjelasan agar rerun checkbox tidak hit API GCal."""
    try:
        return penjelasan_engine.get_jadwal_dari_gcalendar() or {}
    except Exception:
        return {}


@st.cache_data(ttl=300, show_spinner=False)
def _fetch_peserta_tender_cached(kode_lelang: str) -> list:
    """Cache GET peserta Tender agar rerun tidak mengulang request SPSE."""
    return kualifikasi_engine.fetch_peserta_by_id_lelang(kode_lelang) or []


@st.cache_data(ttl=3600)
def _lookup_singkatan_dinas(satker: str) -> str:
    if not satker:
        return "DPUPR"
    try:
        from config import sb as _sb_f
        r = _sb_f().table("master_dinas").select("singkatan").ilike("nama_dinas", f"%{satker[:30]}%").limit(1).execute()
        if r.data:
            return r.data[0].get("singkatan") or "DPUPR"
    except Exception:
        pass
    return "DPUPR"


_TENDER_EXCLUDE_KODES = {
    "10096884000",  # legacy 2025 Hatungun; jangan muncul/terbuat lagi di semua workflow Tender
}


@st.cache_data(ttl=60)
def _load_draft_paket_cached() -> list:
    """Load semua draft_paket sekali, cache 60 detik. Dipakai lintas tab.
    Invalidasi via _load_draft_paket_cached.clear() setelah mutasi draft_paket."""
    def _cache_path():
        from config import STATE_DIR
        return os.path.join(STATE_DIR, "draft_paket_cache.sqlite3")

    def _save_local(rows):
        try:
            with sqlite3.connect(_cache_path(), timeout=5) as conn:
                conn.execute(
                    "CREATE TABLE IF NOT EXISTS draft_paket_cache "
                    "(id INTEGER PRIMARY KEY CHECK (id = 1), payload TEXT NOT NULL)"
                )
                conn.execute(
                    "INSERT OR REPLACE INTO draft_paket_cache (id, payload) VALUES (1, ?)",
                    (json.dumps(rows, ensure_ascii=False),),
                )
        except (OSError, sqlite3.Error, TypeError, ValueError):
            pass

    def _load_local():
        try:
            with sqlite3.connect(_cache_path(), timeout=5) as conn:
                raw = conn.execute(
                    "SELECT payload FROM draft_paket_cache WHERE id = 1"
                ).fetchone()
            rows = json.loads(raw[0]) if raw else []
            return rows if isinstance(rows, list) else []
        except (OSError, sqlite3.Error, TypeError, ValueError):
            return []

    try:
        rows = inbox_engine._sb().table("draft_paket").select("*").order("diambil_pada", desc=True).execute().data or []
        _save_local(rows)
        return [r for r in rows if not _is_tender_excluded(r)]
    except Exception:
        # Supabase client dicache lintas rerun; koneksi HTTP pool bisa stale
        # setelah idle. Refresh client sekali agar gangguan sesaat tidak
        # membuat seluruh daftar paket terlihat kosong.
        try:
            inbox_engine._sb.clear()
        except Exception:
            pass
        try:
            rows = inbox_engine._sb().table("draft_paket").select("*").order("diambil_pada", desc=True).execute().data or []
            _save_local(rows)
            return [r for r in rows if not _is_tender_excluded(r)]
        except Exception:
            rows = _load_local()
            # Jika Supabase belum dikonfigurasi/terputus dan cache lokal kosong,
            # gunakan daftar SPSE yang sudah berhasil diambil pada sesi ini.
            # Ini menjaga Tab 0 tetap menampilkan paket untuk ditrace.
            if not rows:
                _now_fallback = datetime.now().isoformat()
                _session_paket = (
                    st.session_state.get("global_paket_draft", {}).get("paket", [])
                    + st.session_state.get("global_paket_aktif", {}).get("paket", [])
                )
                rows = [
                    {
                        "kode_tender": str(p.get("kode") or p.get("id_lelang") or ""),
                        "nama_tender": p.get("nama") or "",
                        "kode_pokja": p.get("kode_pokja") or p.get("pokja") or "",
                        "folder_dibuat": p.get("folder_dibuat") or "",
                        "nomor_urut": p.get("nomor_urut") or "",
                        "status_tahap": p.get("status") or "",
                        "diambil_pada": _now_fallback,
                    }
                    for p in _session_paket
                    if p.get("kode") or p.get("id_lelang")
                ]
            return [r for r in rows if not _is_tender_excluded(r)]


@st.cache_data(ttl=3600)
def _lookup_telepon_pp(satker: str) -> str:
    if not satker:
        return ""
    try:
        from config import sb as _sb_f
        r = _sb_f().table("master_dinas").select("telepon_pp").ilike("nama_dinas", f"%{satker[:30]}%").limit(1).execute()
        if r.data:
            return r.data[0].get("telepon_pp") or ""
    except Exception:
        pass
    return ""






# ── Helper modul-level PL: cari xlsm + proses Excel ───────────────────────────













def _proses_excel_paket_tender(target_dir, kode_tender, xl=None, row_data=None):
    """COM: IsiDataByKodeTender → isi @ Master Data Excel Tender saat create folder.

    Identik dengan _proses_excel_paket_pl tapi tanpa HPS (HPS sudah dijalankan
    via _jalankan_aksi_tender sebelum fungsi ini dipanggil).
    Return list[str] log lines.
    """
    import isi_master_data_tender as _imd_t
    logs = []

    # Cari .xlsm di folder paket
    try:
        _xs = [f for f in os.listdir(target_dir) if f.lower().endswith(".xlsm")]
    except Exception as _e:
        return [f"WARN Excel: gagal scan folder — {_e}"]
    if not _xs:
        return ["WARN Excel dilewati — tidak ada .xlsm di folder"]
    _xs.sort(key=lambda f: (not f.lower().startswith("0. bapk"), f))
    xlsm = os.path.join(target_dir, _xs[0])

    def _master_data_terisi(_app):
        """Validasi file yang sudah tersimpan di disk, bukan cache COM Excel."""
        _kode_target = re.sub(r"\D", "", str(kode_tender or ""))
        for _attempt in range(8):
            _wb_check = None
            try:
                from openpyxl import load_workbook as _load_wb_check
                _wb_check = _load_wb_check(
                    xlsm, read_only=True, data_only=True, keep_vba=True
                )
                _ws_check = _wb_check["@ Master Data"]
                _kode_excel = re.sub(r"\D", "", str(_ws_check["C4"].value or ""))
                _nama_excel = str(_ws_check["C5"].value or "").strip()
                if _kode_target and _kode_excel == _kode_target and _nama_excel:
                    return True
            except Exception:
                pass
            finally:
                if _wb_check is not None:
                    try:
                        _wb_check.close()
                    except Exception:
                        pass
            if _attempt < 7:
                time.sleep(0.5)
        return False

    try:
        _res = _imd_t.proses_master_data_tender(kode_tender, xlsm, xl=xl, row_data=row_data)
        if _res.get("ok") and xl is not None and not _master_data_terisi(xl):
            # Excel instance reuse dapat menyimpan workbook tanpa menerapkan
            # perubahan pada sebagian mesin. Ulangi sekali dengan COM instance
            # dedicated agar bulk create tidak menghasilkan Excel kosong.
            logs.append("WARN Master Data: hasil COM bersama kosong — retry Excel dedicated")
            _res = _imd_t.proses_master_data_tender(kode_tender, xlsm, xl=None, row_data=row_data)
        if _res.get("ok"):
            # Verifikasi final memakai instance baru agar tidak membaca cache
            # workbook dari COM bulk yang sebelumnya melihat file kosong.
            if not _master_data_terisi(None):
                logs.append("WARN Master Data: COM sukses tetapi sentinel masih kosong")
            else:
                logs.append("Master Data Tender: terisi otomatis")
        else:
            logs.append(f"WARN Master Data Tender: {_res.get('pesan', '-')}")
    except Exception as _e:
        logs.append(f"WARN COM Master Data: {_e}")

    return logs


def _is_tender_excluded(row_or_kode) -> bool:
    kode = row_or_kode if isinstance(row_or_kode, str) else str(row_or_kode.get("kode_tender") or row_or_kode.get("kode") or "")
    return kode in _TENDER_EXCLUDE_KODES


def _validasi_hasil_evaluasi_tender(path: str) -> list[str]:
    """Validasi minimum output evaluasi Tender sebelum UI menyatakan sukses."""
    try:
        with open(path, encoding="utf-8") as _f:
            _txt = _f.read()
    except Exception as _e:
        return [f"gagal membaca output: {_e}"]
    _required = {
        "ekstraksi Dokpil": ("EKSTRAK PERSYARATAN", "DOKPIL"),
        "evaluasi administrasi": ("ADMINISTRASI",),
        "evaluasi teknis": ("TEKNIS",),
        "evaluasi kualifikasi": ("KUALIFIKASI",),
        "flag audit": ("FLAG AUDIT", "FLAG KLARIFIKASI"),
        "kesimpulan akhir": ("KESIMPULAN AKHIR",),
    }
    _upper = _txt.upper()
    _missing = [label for label, markers in _required.items() if not any(m in _upper for m in markers)]
    if not _txt.strip():
        _missing.insert(0, "file kosong")
    return _missing


def _nomor_folder_tertinggi_tender(output_base: str) -> int:
    """Scan disk sebagai guard jika nomor_urut Supabase stale/duplikat."""
    import re as _re
    try:
        names = os.listdir(output_base)
    except Exception:
        return 0
    max_no = 0
    for name in names:
        path = os.path.join(output_base, name)
        if not os.path.isdir(path):
            continue
        m = _re.match(r"^\s*(\d+)\.", name)
        if m:
            max_no = max(max_no, int(m.group(1)))
    return max_no


def _simpan_tanggal_undangan_dpp(kode_tender: str, tgl_ba_reviu, tgl_kirim) -> tuple[bool, str]:
    """Simpan tanggal hasil kirim undangan DPP ke Supabase.

    Kolom baru perlu ada lebih dulu:
    - tgl_ba_reviu_dpp
    - tgl_kirim_undangan_dpp
    """
    if not kode_tender:
        return False, "kode_tender kosong"
    try:
        from config import sb as _sb_tgl_dpp
        _sb_tgl_dpp().table("draft_paket").update({
            "tgl_ba_reviu_dpp": tgl_ba_reviu.isoformat() if hasattr(tgl_ba_reviu, "isoformat") else str(tgl_ba_reviu),
            "tgl_kirim_undangan_dpp": tgl_kirim.isoformat() if hasattr(tgl_kirim, "isoformat") else str(tgl_kirim),
        }).eq("kode_tender", str(kode_tender)).execute()
        return True, "tanggal tersimpan"
    except Exception as e:
        return False, str(e)[:180]


def _generate_lampiran_undangan_dpp_preview(kode_tender: str) -> tuple[bool, str]:
    """Generate PDF lampiran DPP lokal saat folder Tender baru dibuat.

    Tanggal/jam hanya preview; saat kirim undangan, PDF digenerate ulang dari input final Tab 1.
    """
    try:
        import undangan_pdf_engine as _upe_tender
        _tgl = datetime.now().date()
        _hari_tgl = f"{_HARI_NAMA[_tgl.weekday()]}, {_tgl.day} {_BULAN_NAMA[_tgl.month-1]} {_tgl.year}"
        _res = _upe_tender.generate_undangan_pdf(
            kode_tender=kode_tender,
            tanggal_kirim=_tgl,
            hari_tgl_rapat=_hari_tgl,
            pukul_rapat="09.00 s.d. 11.00 Wita",
            tempat_rapat=kirimpesan_engine.DEFAULT_TEMPAT,
            output_path=None,
        )
        if _res.get("sukses"):
            return True, os.path.basename(_res.get("pdf_path") or "")
        return False, _res.get("pesan", "gagal generate lampiran")
    except Exception as e:
        return False, str(e)[:180]


st.set_page_config(
    page_title="Dashboard",
    layout="wide",
)

# Streamlit meredupkan elemen lama (data-stale) saat script masih berjalan.
# Pertahankan tampilan normal agar progress/log terasa live di semua role.
st.markdown("""
<style>
[data-testid="stElementContainer"][data-stale="true"],
[data-testid="stTab"][data-stale="true"] {
    opacity: 1 !important;
    transition: none !important;
}
</style>
""", unsafe_allow_html=True)

# Baca dark mode preference dari session_state sebelum inject CSS
_is_dark = st.session_state.get("toggle_dark_mode", True)






if _is_dark:
    st.markdown(_get_dark_css(), unsafe_allow_html=True)

# Kontrol login diringkas dalam popover agar header tidak menjadi kolom vertikal.
if "header_login_role" not in st.session_state:
    st.session_state["header_login_role"] = st.session_state.get("sidebar_login_role", "PP")
_login_popover = st.popover(f"{APP_VERSION} · 🔐 Login / SPSE", use_container_width=True)

# ── Mode Switcher ──────────────────────────────────────────────────────────────
# Filter mode berdasarkan role login
_spse_role = st.session_state.get("spse_role", None)  # "PP", "POKJA", atau None

# Auto-detect role dari CDP kalau session baru tapi Brave masih aktif (misal F5 refresh).
# Retry periodik agar kegagalan sesaat saat Brave/SPSE masih loading tidak mematikan watchdog.
_cdp_role_last_check = st.session_state.get("_cdp_role_last_check", 0.0)
if not _spse_role and (time.time() - _cdp_role_last_check >= 30):
    st.session_state["_cdp_role_last_check"] = time.time()
    try:
        import spse_browser as _sb_detect
        import spse_login as _sl_detect
        _cdp_url = _sb_detect.get_url()  # kosong "" kalau CDP tidak aktif
        if _cdp_url:
            _detected = _sl_detect.detect_login_role()
            if _detected:
                st.session_state["spse_role"] = _detected
                _spse_role = _detected
    except Exception:
        pass
# Pastikan auto-refresh thread tetap jalan selama sudah login
if _spse_role:
    import spse_browser as _sb_ar
    _sb_ar.mulai_auto_refresh()

# ============================================================
# Sidebar — Browser Control
# ============================================================



with _login_popover:
    # ── Toggle Light/Dark Mode ──────────────────────────────────────────────
    _dark_mode = st.toggle("🌙 Dark Mode", value=True, key="toggle_dark_mode")
    if not _dark_mode:
        st.markdown(_get_light_css(), unsafe_allow_html=True)
    st.divider()
    st.header("Browser SPSE")

    # Auto-reconnect Playwright hanya saat dibutuhkan (lazy) — sidebar info pakai CDP HTTP saja
    # buka_browser() dipanggil oleh engine saat submit, bukan di sini setiap refresh

    url_aktif = spse_browser.get_url() if spse_browser._cek_cdp_aktif() else None
    if url_aktif:
        _role_label = st.session_state.get("spse_role", None)
        _at_loginpass = "loginpass" in url_aktif
        # URL root /tapinkab/ atau /tapinkab (tanpa path lain) = halaman setelah logout
        _base = SPSE_BASE_URL.rstrip("/")
        _at_logout_root = url_aktif.rstrip("/") == _base

        # Deteksi logout paksa: Brave masih jalan tapi halaman sudah kembali ke login/root
        if _role_label and (_at_loginpass or _at_logout_root):
            st.session_state.pop("spse_role", None)
            _role_label = None

        # Recover role dari file cache setelah Streamlit hot-reload (browser masih login,
        # tapi session_state hilang). Senyap — tanpa tombol "Deteksi Ulang" yang jadi dead-end.
        if not _role_label and not _at_loginpass and not _at_logout_root:
            try:
                import spse_login as _sl_recover
                _rec = _sl_recover.detect_login_role()
                if _rec:
                    st.session_state["spse_role"] = _rec
                    _role_label = _rec
            except Exception:
                pass

        if _role_label:
            # Sudah login berhasil
            st.success("Browser terhubung")
            _spse_home = SPSE_BASE_URL.rstrip("/") + "/home"
            st.markdown(f"[🔗 {_spse_home}]({_spse_home})", unsafe_allow_html=False)
            if st.button("↗️ Buka SPSE di Brave", use_container_width=True, key="btn_buka_spse_brave"):
                try:
                    spse_browser.buka_tab_baru(_spse_home)
                    st.toast("Tab SPSE dibuka di Brave ✅", icon="✅")
                except Exception as _e:
                    st.error(str(_e))
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Refresh", use_container_width=True):
                    ok = spse_browser.refresh_browser()
                    if not ok:
                        st.toast("⚠️ Reload gagal — CDP tidak responsif", icon="⚠️")
                    st.rerun()
            with col2:
                if st.button("❌ Tutup & Ganti Akun", use_container_width=True, help="Logout SPSE → tutup Brave → kembali ke form login"):
                    import spse_login as _sl_logout
                    _sl_logout.logout_spse()
                    spse_browser.tutup_browser()
                    st.session_state.pop("spse_role", None)
                    st.session_state.pop("login_failed", None)
                    # Bersihkan cache per-role (detail paket + pool DPA) agar tidak nempel saat ganti akun
                    for _k in [k for k in list(st.session_state.keys())
                               if k.startswith(("ppk_detail_", "ppk_dpa_", "pl_dpa_"))
                               or k in ("ppk_dpa_pool", "pl_dpa_pool")]:
                        st.session_state.pop(_k, None)
                    _load_draft_paket_cached.clear()
                    _load_draft_pl_cached.clear()
                    _fetch_peserta_pl_cached.clear()
                    _fetch_status_semua_paket_cached.clear()
                    st.rerun()
            _role_emoji = {"PP": "🏛️", "POKJA": "👥", "PPK": "📝", "E-Katalog": "🛒"}.get(_role_label, "🌐")
            st.success(f"{_role_emoji} Login sebagai **{_role_label}**")

        elif _at_loginpass or st.session_state.get("login_failed"):
            # Browser terbuka tapi masih di halaman login / gagal — tampilkan retry
            st.warning("⏳ Menunggu login...")
            st.caption(url_aktif[:60] + "..." if len(url_aktif) > 60 else url_aktif)
            if st.button("❌ Tutup & Mulai Ulang", use_container_width=True):
                spse_browser.tutup_browser()
                st.session_state.pop("login_failed", None)
                st.session_state.pop("login_failed_role", None)
                st.rerun()
        else:
            # Browser hidup tapi tidak dalam sesi login yang dikenali (role recovery di atas
            # sudah gagal). Bukan dead-end: sediakan tutup bersih lalu form login.
            st.warning("⚠️ Sesi SPSE tidak dikenali — silakan tutup lalu login ulang.")
            st.caption(url_aktif[:60] + "..." if len(url_aktif) > 60 else url_aktif)
            if st.button("❌ Tutup Browser & Login Ulang", type="primary", use_container_width=True):
                spse_browser.tutup_browser()
                st.session_state.pop("login_failed", None)
                st.session_state.pop("login_failed_role", None)
                st.rerun()
    else:
        # CDP tidak aktif — browser tutup/belum dibuka; bersihkan semua state login sisa
        st.session_state.pop("spse_role", None)
        st.session_state.pop("login_failed", None)
        st.session_state.pop("login_failed_role", None)
        _sidebar_login_form()

_spse_role = st.session_state.get("spse_role", None)  # re-sync setelah sidebar logic
_ALL_MODES = ["Tender", "PL - Konsultansi", "PL - Konstruksi", "PPK - Upload Dokumen"]
if _spse_role == "PP":
    _MODE_OPTIONS = ["PL - Konsultansi", "PL - Konstruksi"]
elif _spse_role == "POKJA":
    _MODE_OPTIONS = ["Tender"]
elif _spse_role == "PPK":
    _MODE_OPTIONS = ["PPK - Upload Dokumen"]
elif _spse_role == "E-Katalog":
    _MODE_OPTIONS = ["E-Katalog - Survei Pasar"]
else:
    _MODE_OPTIONS = _ALL_MODES

if _spse_role:
    if "app_mode" not in st.session_state:
        st.session_state["app_mode"] = _MODE_OPTIONS[0]

    _mode_col, _ = st.columns([2, 5])
    with _mode_col:
        if st.session_state.get("app_mode") not in _MODE_OPTIONS:
            st.session_state["app_mode"] = _MODE_OPTIONS[0]
            st.rerun()
        _selected_mode = st.radio(
            "Mode:",
            _MODE_OPTIONS,
            index=_MODE_OPTIONS.index(st.session_state["app_mode"]),
            horizontal=True,
            key="radio_app_mode",
        )
        st.session_state["app_mode"] = _selected_mode
        activate_mode(_selected_mode)

    st.divider()
else:
    st.session_state.pop("app_mode", None)
    st.info("🔐 Silakan login via sidebar untuk memulai.")
    st.stop()

# ============================================================
# Tabs
# ============================================================

_HARI_NAMA  = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
_BULAN_NAMA = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
               "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

_LIBUR_2026 = {
    "2026-01-01": "Tahun Baru 2026 Masehi",
    "2026-01-16": "Isra Mikraj Nabi Muhammad S.A.W.",
    "2026-02-16": "Cuti Bersama Tahun Baru Imlek",
    "2026-02-17": "Tahun Baru Imlek 2577 Kongzili",
    "2026-03-18": "Cuti Bersama Hari Suci Nyepi",
    "2026-03-19": "Hari Suci Nyepi (Tahun Baru Saka 1948)",
    "2026-03-20": "Cuti Bersama Idul Fitri 1447 Hijriah",
    "2026-03-21": "Idul Fitri 1447 Hijriah",
    "2026-03-22": "Idul Fitri 1447 Hijriah",
    "2026-03-23": "Cuti Bersama Idul Fitri 1447 Hijriah",
    "2026-03-24": "Cuti Bersama Idul Fitri 1447 Hijriah",
    "2026-04-03": "Wafat Yesus Kristus",
    "2026-04-05": "Kebangkitan Yesus Kristus (Paskah)",
    "2026-05-01": "Hari Buruh Internasional",
    "2026-05-14": "Kenaikan Yesus Kristus",
    "2026-05-15": "Cuti Bersama Kenaikan Yesus Kristus",
    "2026-05-27": "Idul Adha 1447 Hijriah",
    "2026-05-28": "Cuti Bersama Idul Adha 1447 Hijriah",
    "2026-05-31": "Hari Raya Waisak 2570 BE",
    "2026-06-01": "Hari Lahir Pancasila",
    "2026-06-16": "1 Muharam Tahun Baru Islam 1448 Hijriah",
    "2026-08-17": "Proklamasi Kemerdekaan",
    "2026-08-25": "Maulid Nabi Muhammad S.A.W.",
    "2026-12-24": "Cuti Bersama Kelahiran Yesus Kristus",
    "2026-12-25": "Kelahiran Yesus Kristus",
}
_LIBUR_MAP = {datetime.strptime(k, "%Y-%m-%d").date(): v for k, v in _LIBUR_2026.items()}



# Auto-start scheduler saat app dibuka (daemon thread, jalan terus)
if not st.session_state.get("_scheduler_started"):
    penjelasan_engine.start_scheduler()
    st.session_state["_scheduler_started"] = True

def _fmt_rp(angka_str: str) -> str:
    try:
        return f"Rp {int(angka_str):,}".replace(",", ".")
    except:
        return angka_str or "-"

if st.session_state["app_mode"] == "PPK - Upload Dokumen":
    # ============================================================
    # MODE: PPK — Upload Dokumen Persiapan Pengadaan
    # ============================================================
    import ppk_upload_engine as _ppk_up

    # ── Cek login PPK ────────────────────────────────────────────────────────
    if st.session_state.get("spse_role") != "PPK":
        st.warning("⚠️ Browser belum login sebagai PPK. Login via sidebar terlebih dahulu.")
        st.stop()

    _PPK_STRIP = [
        "Belanja Jasa Konsultansi Perencanaan Arsitektur-Jasa Arsitektur Lainnya ",
        "Belanja Jasa Konsultansi Perencanaan Rekayasa-Jasa Desain Rekayasa untuk Konstruksi Pondasi serta Struktur Bangunan ",
        "Belanja Jasa Konsultansi Perencanaan Rekayasa-",
        "Belanja Jasa Konsultansi Perencanaan Arsitektur-",
        "Belanja Jasa Konsultansi Perencanaan ",
        "Belanja Jasa Konsultansi ",
        "Belanja Jasa ",
        "Belanja ",
    ]

    import dpa_engine as _dpa_match  # cari_dpa_di_pool / load_pool_dpa

    def _nama_singkat(nama: str) -> str:
        for pfx in _PPK_STRIP:
            if nama.startswith(pfx):
                return nama[len(pfx):]
        return nama

    @st.cache_data(ttl=120)
    def _load_paket_ppk(_role_key=None):
        return _ppk_up.fetch_paket_ppk()

    _ppk_tab1, _ppk_tab2, _ppk_tab3 = st.tabs([
        "1️⃣ Info Paket",
        "2️⃣ Upload Dokumen",
        "3️⃣ Import DPA",
    ])

    with _ppk_tab1:
        _info_list = _load_paket_ppk(_role_key=st.session_state.get("spse_role"))
        _ppk_folders_info = _ppk_up.list_subfolder_ppk()
        _ppk_next_info = _ppk_up.next_folder_number(_ppk_folders_info)
        _ppk_assigned_info = set()

        def _rp(v):
            try: return f"Rp {int(v):,}".replace(",", ".")
            except: return str(v) if v else "-"

        if not _info_list:
            st.markdown("## ℹ️ Info Paket PPK")
            st.warning("Tidak ada paket ditemukan. Pastikan login PPK aktif di browser.")
            if st.button("🔄 Refresh Paket", key="ppk_refresh_empty"):
                _load_paket_ppk.clear()
                st.rerun()
        else:
            # Auto-load detail yang belum ada di session_state
            _missing = [p for p in _info_list if f"ppk_detail_{p['kode_paket']}" not in st.session_state]
            if _missing:
                _hcol1, _hcol2 = st.columns([3, 2])
                _hcol1.markdown("## ℹ️ Info Paket PPK")
                with _hcol2:
                    with st.status(f"Memuat {len(_missing)} paket...", expanded=False) as _st:
                        from concurrent.futures import ThreadPoolExecutor, as_completed
                        with ThreadPoolExecutor(max_workers=4) as _ex:
                            _futs = {_ex.submit(_ppk_up.fetch_detail_paket, p["kode_paket"]): p for p in _missing}
                            _done = 0
                            for _fut in as_completed(_futs, timeout=120):
                                _done += 1
                                _p = _futs[_fut]
                                _st.update(label=f"[{_done}/{len(_missing)}] selesai...")
                                try:
                                    st.session_state[f"ppk_detail_{_p['kode_paket']}"] = _fut.result()
                                except Exception:
                                    st.session_state[f"ppk_detail_{_p['kode_paket']}"] = {}
                        _st.update(label=f"✅ {len(_missing)} paket dimuat", state="complete")
            else:
                st.markdown("## ℹ️ Info Paket PPK")

            _n_loaded = sum(1 for p in _info_list if ("ppk_detail_" + p["kode_paket"]) in st.session_state)
            st.caption(f"{len(_info_list)} paket · {_n_loaded} detail tersedia")

            for _ii, _ip in enumerate(_info_list, start=1):
                _ik = _ip["kode_paket"]
                _in = _ip["nama_paket"]
                _is = _ip.get("status", "-")
                _det_key = f"ppk_detail_{_ik}"
                _det = st.session_state.get(_det_key)

                # Label expander: nomor + nama singkat + preview nilai pagu jika sudah dimuat
                _pagu_preview = f" — {_rp(_det.get('nilai_pagu'))}" if _det and _det.get("nilai_pagu") else ""
                _matched_info_folder = _ppk_up.auto_match_folder(_in, _ppk_folders_info)
                _folder_no_info = _ppk_up.folder_number(_matched_info_folder)
                if _folder_no_info is None:
                    while _ppk_next_info in _ppk_assigned_info:
                        _ppk_next_info += 1
                    _folder_no_info = _ppk_next_info
                    _ppk_next_info += 1
                _ppk_assigned_info.add(_folder_no_info)
                _exp_label = f"{_folder_no_info}. {_nama_singkat(_in)[:75]}{_pagu_preview}"

                with st.expander(_exp_label, expanded=False):
                    st.caption(f"Status: {_is}")
                    st.markdown("**Nama Paket (Lengkap):**")
                    st.code(_in, language=None)

                    if not _det:
                        if st.button("Muat Detail", key=f"load_det_{_ik}"):
                            st.session_state[_det_key] = _ppk_up.fetch_detail_paket(_ik)
                            st.rerun()
                        st.caption("⏳ Klik untuk muat detail paket.")
                    else:
                        if _det.get("_error_step1") or _det.get("_error_step2"):
                            st.warning(f"Partial error: {_det.get('_error_step1','')} {_det.get('_error_step2','')}")

                        # Field copy-paste
                        for _lbl, _val in [("Kode RUP", _det.get("kode_rup")), ("Kode Rekening (MAK)", _det.get("mak"))]:
                            if _val:
                                st.markdown(f"**{_lbl}:**")
                                st.code(_val, language=None)

                        # Tabel info
                        _info_rows = [
                            ("Nilai Pagu",      _rp(_det.get("nilai_pagu"))),
                            ("Nilai HPS",       _rp(_det.get("nilai_hps"))),
                            ("Jenis Kontrak",   _det.get("jenis_kontrak", "-")),
                            ("Sumber Dana",     _det.get("sumber_dana", "-")),
                            ("Tahun Anggaran",  _det.get("tahun_anggaran", "-")),
                            ("Lokasi Pekerjaan",_det.get("lokasi", "-")),
                            ("Satuan Kerja",    _det.get("satker", "-")),
                            ("Instansi",        _det.get("instansi", "-")),
                            ("Nama PPK",        _det.get("nama_ppk", "-")),
                        ]
                        st.table([{"Field": k, "Nilai": v} for k, v in _info_rows if v and v != "-"])

                        # Ide 3: Tampilkan info DPA di Tab 1 Info Paket PPK
                        if _det:
                            # Pool semua item DPA bernama (di-cache sekali per render).
                            # Match nama paket SPSE → DPA via nama pekerjaan fisik (bukan prefix konsultansi).
                            if "ppk_dpa_pool" not in st.session_state:
                                from config import sb as _sb_dpa
                                st.session_state["ppk_dpa_pool"] = _dpa_match.load_pool_dpa(_sb_dpa())

                            _dpa_key = f"ppk_dpa_{_ik}"
                            if _dpa_key not in st.session_state:
                                st.session_state[_dpa_key] = _dpa_match.cari_dpa_di_pool(
                                    _in, st.session_state["ppk_dpa_pool"]
                                )

                            _dpa_rows = st.session_state.get(_dpa_key, [])
                            if _dpa_rows:
                                st.divider()
                                _dpa_nama = _dpa_rows[0].get("nama_paket") or "-"
                                st.markdown(f"**📊 Info DPA** — _{_dpa_nama}_")
                                _dpa_table_data = []
                                for _it in _dpa_rows:
                                    _subkeg_id = _it.get("subkegiatan_id") or ""
                                    _parts = _subkeg_id.split("|")
                                    _subkeg_code = _parts[-1] if len(_parts) >= 3 else _subkeg_id
                                    _dpa_table_data.append({
                                        "Uraian": _it.get("uraian") or "-",
                                        "Spesifikasi": _it.get("spesifikasi") or "-",
                                        "Sebelum": _rp(_it.get("jumlah_sebelum")),
                                        "Sesudah": _rp(_it.get("jumlah_sesudah")),
                                        "Sub Kegiatan": _subkeg_code
                                    })
                                st.table(_dpa_table_data)
                            else:
                                st.caption("Belum ada data DPA")

    with _ppk_tab3:
        _render_tab_dpa()

    with _ppk_tab2:
        _ppk_col1, _ppk_col2 = st.columns([3, 1])
        _ppk_col1.markdown("## 📤 Upload Dokumen Persiapan Pengadaan")
        if _ppk_col2.button("🔄 Refresh", key="btn_refresh_paket_ppk", use_container_width=True):
            # Hapus semua cache bulk + cache paket PPK
            for _k in list(st.session_state.keys()):
                if _k.startswith("ppk_versi_") or _k.startswith("ppk_bulk"):
                    del st.session_state[_k]
            _load_paket_ppk.clear()
            st.rerun()

        with st.spinner("Memuat daftar paket dari SPSE..."):
            _paket_list = _load_paket_ppk()

        if not _paket_list:
            st.warning("Tidak ada paket ditemukan. Pastikan login PPK aktif di browser.")
            st.stop()

        st.caption(f"{len(_paket_list)} paket ditemukan")

        _UPLOAD_SECTIONS = [
            {"key": "kak",     "label": "KAK / Spesifikasi",   "icon": "📄", "accept": ["doc","docx","xls","xlsx","pdf","jpg","jpeg","png","zip","rar"], "required": True},
            {"key": "kontrak", "label": "Rancangan Kontrak",    "icon": "📋", "accept": ["pdf"], "required": True},
            {"key": "uraian",  "label": "Uraian Singkat",       "icon": "📝", "accept": ["pdf"], "required": True},
            {"key": "lainnya", "label": "Informasi Lainnya",    "icon": "ℹ️", "accept": ["txt","doc","docx","xls","xlsx","pdf","gif","jpeg","jpg","png","zip","rar","rtf"], "required": False},
            {"key": "nd",      "label": "Nota Dinas PPK",       "icon": "📨", "accept": ["pdf","jpg","jpeg","png"], "required": False},
        ]

        # ── Semua paket tampil sekaligus (dokumen dimuat lazy per paket) ───────────
        for _pi, _pk in enumerate(_paket_list, start=1):
            _kode = _pk["kode_paket"]
            _nama = _pk["nama_paket"]
            _status = _pk.get("status", "-")
            _singkat = _nama_singkat(_nama)
            _folder_no_upload = _ppk_up.folder_number(
                _ppk_up.auto_match_folder(_nama, _ppk_folders_info)
            ) or _pi
            _exp_label = f"{_folder_no_upload}. {_singkat}"

            with st.expander(_exp_label, expanded=False):
                st.caption(f"[{_status}] {_kode} — {_nama}")

                # ── Bulk load dokumen existing (1 subprocess, 4 endpoint) ───
                _bulk_key = f"ppk_bulk_{_kode}"
                if _bulk_key not in st.session_state:
                    _bulk = _ppk_up.list_semua_dokumen(_kode)
                    for _jenis_b, _docs_b in _bulk.items():
                        st.session_state[f"ppk_versi_{_kode}_{_jenis_b}"] = _docs_b
                    st.session_state[_bulk_key] = True

                # ── Hapus Semua ─────────────────────────────────────────────
                _hcol1, _hcol2 = st.columns([5, 1])
                if _hcol2.button("🗑️ Hapus Semua", key=f"hapus_semua_{_kode}",
                                 help="Hapus semua dokumen yang sudah terupload di paket ini"):
                    # Kumpulkan versi dari session_state per jenis
                    _to_del = {}
                    for _sec2 in _UPLOAD_SECTIONS:
                        if _sec2["key"] == "nd": continue
                        _vk2 = f"ppk_versi_{_kode}_{_sec2['key']}"
                        for _d in st.session_state.get(_vk2, []):
                            if _d.get("versi") is not None:
                                _to_del.setdefault(_sec2["key"], []).append(_d["versi"])
                    if _to_del:
                        # Ada versi di session → hapus langsung
                        _hasil_hapus = _ppk_up.hapus_semua_dokumen(_kode, versi_map=_to_del)
                    else:
                        # Tidak ada di session → fallback ke Playwright list
                        _hasil_hapus = _ppk_up.hapus_semua_dokumen(_kode)
                    _total_hapus = _hasil_hapus.get("dihapus", 0)
                    _total_err   = _hasil_hapus.get("gagal", 0)
                    if _total_err == 0:
                        # Bersihkan session state agar re-fetch saat render ulang
                        for _sec2 in _UPLOAD_SECTIONS:
                            _vk2 = f"ppk_versi_{_kode}_{_sec2['key']}"
                            if _vk2 in st.session_state: del st.session_state[_vk2]
                        if _bulk_key in st.session_state: del st.session_state[_bulk_key]
                        st.toast(f"✅ {_total_hapus} dokumen dihapus", icon="🗑️")
                    else:
                        st.warning(f"⚠️ {_total_hapus} dihapus, {_total_err} gagal")
                    st.rerun()

                st.divider()

                # ── Upload dari Folder ──────────────────────────────────────
                _subfolder_list = _ppk_up.list_subfolder_ppk()
                _auto_match     = _ppk_up.auto_match_folder(_nama, _subfolder_list)

                if st.checkbox("📁 Upload dari Folder", key=f"chk_folder_{_kode}"):
                    with st.container(border=True):
                        _input_mode = st.radio(
                            "Cara pilih folder:",
                            ["📂 Pilih dari daftar", "⌨️ Paste path manual"],
                            horizontal=True,
                            key=f"foldermode_{_kode}",
                        )

                        _selected_folder = None
                        if _input_mode == "📂 Pilih dari daftar":
                            _folder_options = ["(pilih folder...)"] + _subfolder_list
                            _default_idx = 0
                            if _auto_match and _auto_match in _subfolder_list:
                                _default_idx = _subfolder_list.index(_auto_match) + 1
                                st.caption(f"💡 Auto-match: **{_auto_match}**")
                            _sel = st.selectbox(
                                "Folder paket:",
                                _folder_options,
                                index=_default_idx,
                                key=f"foldersel_{_kode}",
                            )
                            if _sel != "(pilih folder...)":
                                _selected_folder = os.path.join(_ppk_up.PPK_PL_BASE, _sel)
                        else:
                            _path_input = st.text_input(
                                "Paste path folder:",
                                value=(_auto_match and os.path.join(_ppk_up.PPK_PL_BASE, _auto_match)) or "",
                                key=f"folderpath_{_kode}",
                                placeholder=r"D:\path\ke\folder paket",
                            )
                            # Path dari Explorer sering ikut membawa tanda kutip.
                            _path_clean = os.path.expandvars(
                                _path_input.strip().strip('"').strip("'").strip()
                            ) if _path_input else ""
                            if _path_clean and os.path.isdir(_path_clean):
                                _selected_folder = os.path.normpath(_path_clean)
                            elif _path_clean:
                                st.warning("⚠️ Path tidak valid atau bukan folder.")

                        if _selected_folder:
                            _preview = _ppk_up.scan_folder(_selected_folder, pdf_only=True)
                            if _preview:
                                _JENIS_LABEL = {"kak": "KAK / Spesifikasi", "uraian": "Uraian Singkat", "kontrak": "Rancangan Kontrak", "lainnya": "Informasi Lainnya", "nd": "Nota Dinas PPK"}
                                _doc_files = [
                                    p for p in _preview
                                    if re.match(r"^(?:[1-9]|11)\.\s", p["nama"]) and p["jenis"] != "nd"
                                ]
                                _nd_files = [p for p in _preview if p["jenis"] == "nd"]
                                _other_files = [
                                    p for p in _preview
                                    if p not in _doc_files and p not in _nd_files
                                ]
                                st.markdown("**Preview file terdeteksi:**")
                                for _pf in _preview:
                                    st.markdown(f"- `{_pf['nama']}` → **{_JENIS_LABEL.get(_pf['jenis'], _pf['jenis'])}**")

                                if _doc_files:
                                    st.caption(f"{len(_doc_files)} dokumen nomor 1-9/11 (tanpa Nota Dinas) siap diupload")
                                if _other_files:
                                    st.caption(f"⏭ {len(_other_files)} file lain (mis. nomor 10) tidak ikut upload bulk")
                                if st.button(
                                    f"⬆️ Upload Dokumen PPK + Pilih PP ({len(_doc_files)} file)",
                                    key=f"btn_folder_docs_{_kode}",
                                    type="primary",
                                    disabled=not _doc_files,
                                ):
                                    _flog_box = st.container(border=True)
                                    _flog_lines = []
                                    def _flog(msg):
                                        _flog_lines.append(msg)
                                        _flog_box.write(msg)
                                    _fres = _ppk_up.upload_dokumen_dari_folder(
                                        kode_paket=_kode,
                                        folder_path=_selected_folder,
                                        log_fn=_flog,
                                        pdf_only=True,
                                    )
                                    # Simpan versi hasil upload ke session_state
                                    for _fr in _fres.get("results", []):
                                        if _fr.get("ok") and _fr.get("versi") and _fr.get("jenis") != "nd":
                                            _fvk = f"ppk_versi_{_kode}_{_fr['jenis']}"
                                            _fvl = st.session_state.get(_fvk, [])
                                            _fvl.append({"nama_file": _fr["nama"], "versi": _fr["versi"]})
                                            st.session_state[_fvk] = _fvl
                                    if _fres.get("total_err", 0) == 0:
                                        st.success(f"✅ {_fres['total_ok']} dokumen PPK berhasil diupload!")
                                        st.toast(f"✅ {_fres['total_ok']} dokumen diupload", icon="✅")
                                    else:
                                        st.warning(f"⚠️ {_fres['total_ok']} berhasil, {_fres['total_err']} gagal")
                                    if _fres.get("pp_ok") is False:
                                        st.error("❌ PP gagal dipilih; tidak ada dokumen yang diupload.")

                                if _nd_files:
                                    if len(_nd_files) > 1:
                                        st.warning("⚠️ Ditemukan lebih dari satu Nota Dinas. Pilih satu file melalui uploader Nota Dinas di tab khusus.")
                                    else:
                                        st.caption(f"1 Nota Dinas siap diupload dan dikirim melalui email")
                                        if st.button(
                                            "📨 Upload Nota Dinas + Kirim Email",
                                            key=f"btn_folder_nd_{_kode}",
                                        ):
                                            _nd_file = _nd_files[0]
                                            _nd_log_box = st.container(border=True)
                                            def _nd_log(msg):
                                                _nd_log_box.write(msg)
                                            with open(_nd_file["path"], "rb") as _nd_fh:
                                                _nd_bytes = _nd_fh.read()
                                            _nd_res = _ppk_up.upload_nota_dinas_dan_email(
                                                kode_paket=_kode,
                                                file_bytes=_nd_bytes,
                                                file_name=_nd_file["nama"],
                                                mime_type=_nd_file["mime"],
                                                log_fn=_nd_log,
                                            )
                                            if _nd_res.get("upload_ok") and _nd_res.get("email_ok"):
                                                st.success("✅ Nota Dinas terupload dan email berhasil dikirim.")
                                            elif _nd_res.get("upload_ok"):
                                                st.warning("⚠️ Nota Dinas terupload, tetapi email gagal dikirim.")
                                            else:
                                                st.error(_nd_res.get("error", "Flow Nota Dinas gagal."))
                            else:
                                st.info("Tidak ada file PDF yang cocok di folder ini.")

                # 4 tab per paket
                _tab_labels = [f"{s['icon']} {s['label']}{' *' if s['required'] else ''}" for s in _UPLOAD_SECTIONS]
                _tabs = st.tabs(_tab_labels)

                for _tab, _sec in zip(_tabs, _UPLOAD_SECTIONS):
                    with _tab:
                        if _sec["key"] == "nd":
                            st.info("ℹ️ Dokumen Nota Dinas PPK yang terupload tidak ditampilkan di list ini karena perbedaan endpoint. Cek langsung di SPSE jika ingin melihat daftarnya.")
                            _nd_up = st.file_uploader(
                                "Pilih file Nota Dinas PPK:",
                                type=_sec["accept"],
                                key=f"up_{_kode}_nd_dedicated",
                            )
                            if _nd_up and st.button(
                                "📨 Upload Nota Dinas + Kirim Email",
                                key=f"btn_{_kode}_nd_dedicated",
                                type="primary",
                            ):
                                with st.status(f"Memproses Nota Dinas {_nd_up.name}...", expanded=True) as _nd_sts:
                                    def _nd_mklog(sts):
                                        def _log(msg): sts.write(msg)
                                        return _log
                                    _nd_res = _ppk_up.upload_nota_dinas_dan_email(
                                        kode_paket=_kode,
                                        file_bytes=_nd_up.read(),
                                        file_name=_nd_up.name,
                                        mime_type=_nd_up.type or "application/pdf",
                                        log_fn=_nd_mklog(_nd_sts),
                                    )
                                    if _nd_res.get("upload_ok") and _nd_res.get("email_ok"):
                                        _nd_sts.update(label="✅ Nota Dinas + email berhasil", state="complete")
                                    elif _nd_res.get("upload_ok"):
                                        _nd_sts.update(label="⚠️ Nota Dinas terupload, email gagal", state="error")
                                    else:
                                        _nd_sts.update(label="❌ Flow Nota Dinas gagal", state="error")
                                        st.error(_nd_res.get("error", "Flow Nota Dinas gagal."))
                            continue
                        # Dokumen existing
                        try:
                            if _sec["key"] != "nd":
                                # Pakai session_state dulu (versi tersimpan saat upload)
                                _vkey = f"ppk_versi_{_kode}_{_sec['key']}"
                                if _vkey not in st.session_state:
                                    st.session_state[_vkey] = _ppk_up.list_dokumen(_kode, _sec["key"])
                                _existing = st.session_state[_vkey]
                            else:
                                _existing = []

                            if _existing:
                                for _doc in _existing:
                                    _dc1, _dc2 = st.columns([5, 1])
                                    _dc1.markdown(f"📎 `{_doc['nama_file']}`")
                                    if _dc2.button("🗑️", key=f"del_{_kode}_{_sec['key']}_{_doc['versi']}", help="Hapus"):
                                        if _ppk_up.hapus_dokumen(_kode, _sec["key"], _doc["versi"]):
                                            # Hapus dari session_state
                                            _vk3 = f"ppk_versi_{_kode}_{_sec['key']}"
                                            _vl3 = st.session_state.get(_vk3, [])
                                            st.session_state[_vk3] = [d for d in _vl3 if d.get("versi") != _doc["versi"]]
                                            st.toast("✅ Dihapus", icon="✅")
                                            st.rerun()
                                        else:
                                            st.error("Gagal hapus")
                                st.divider()
                        except Exception as _le:
                            st.caption(f"⚠️ {_le}")

                        # Upload
                        _up = st.file_uploader(
                            f"Pilih file {_sec['label']}:",
                            type=_sec["accept"],
                            key=f"up_{_kode}_{_sec['key']}",
                            label_visibility="collapsed",
                        )
                        if _up:
                            if st.button(f"⬆️ Upload **{_up.name}** ke SPSE", key=f"btn_{_kode}_{_sec['key']}", type="primary"):
                                with st.status(f"Mengupload {_up.name}...", expanded=True) as _sts:
                                    def _mklog(sts):
                                        def _log(msg): sts.write(msg)
                                        return _log
                                    _res = _ppk_up.upload_dokumen(
                                        kode_paket=_kode,
                                        jenis=_sec["key"],
                                        file_bytes=_up.read(),
                                        file_name=_up.name,
                                        mime_type=_up.type or "application/octet-stream",
                                        log_fn=_mklog(_sts),
                                    )
                                    if _res.get("ok"):
                                        _sts.update(label="✅ Berhasil!", state="complete")
                                        st.toast(f"✅ {_up.name} diupload", icon="✅")
                                        # Simpan versi untuk hapus nanti
                                        _vkey = f"ppk_versi_{_kode}_{_sec['key']}"
                                        _vlist = st.session_state.get(_vkey, [])
                                        _vlist.append({"nama_file": _up.name, "versi": _res.get("versi")})
                                        st.session_state[_vkey] = _vlist
                                        st.rerun()
                                    else:
                                        _sts.update(label="❌ Gagal", state="error")
                                        st.error(_res.get("error", "Unknown error"))

                st.divider()
                st.markdown("### Finalisasi Paket")
                st.caption("Setelah dokumen PPK, PP, Nota Dinas, dan email selesai, klik tombol ini untuk membuat paket di SPSE.")
                if st.button(
                    "💾 Simpan dan Membuat Paket",
                    key=f"btn_final_ppk_{_kode}",
                    type="primary",
                    use_container_width=True,
                ):
                    _final_box = st.container(border=True)
                    _final_box.info("Menyimpan dan membuat paket...")
                    def _final_log(msg):
                        _final_box.write(msg)
                    _final_res = _ppk_up.simpan_dan_membuat_paket(_kode, log_fn=_final_log)
                    if _final_res.get("ok"):
                        _final_box.success("✅ Paket berhasil disimpan dan dibuat di SPSE.")
                    else:
                        _final_box.error(_final_res.get("error", "Simpan dan Membuat Paket gagal."))

    st.stop()

if st.session_state["app_mode"] == "PL - Konsultansi":
    # ============================================================
    # MODE: PENGADAAN LANGSUNG (PL JKK & PL PK)
    # ============================================================

    if st.session_state.get("spse_role") != "PP":
        st.warning("⚠️ Browser belum login sebagai PP. Login via sidebar terlebih dahulu.")
        st.stop()

    _PL_TAB_LABELS = [
        "1️⃣ Draft Paket PL",
        "2️⃣ Kirim Undangan DPP",
        "3️⃣ Setup Paket",
        "4️⃣ Pilih Penyedia & Umumkan",
        "5️⃣ Buat Jadwal",
        "6️⃣ Download Kualifikasi",
        "7️⃣ Evaluasi & Teknis/Biaya",
        "8️⃣ Kirim Verifikasi",
        "9️⃣ Upload BA PL",
        "🔟 Penetapan Pemenang",
        "📄 Import DPA",
    ]
    _pl_active_tab = st.radio("Tab PL", _PL_TAB_LABELS, horizontal=True, key="pl_active_tab_jkk")

    if _pl_active_tab == "📄 Import DPA":
        _render_tab_dpa()

    # ── Tab 1: Draft Paket PL (JKK) ──────────────────────────────────────────
    if _pl_active_tab == "1️⃣ Draft Paket PL":
        import os as _pl_os, subprocess as _pl_sp
        from config import POKJA_ROOT as _PL_POKJA_ROOT, OUTPUT_DIR_PL_JKK as _PL_DIR_JKK, OUTPUT_DIR_PL_PK as _PL_DIR_PK, V19_ROOT as _V19_ROOT, POKJA_PYTHON as _POKJA_PYTHON
        _TEMPLATE_DIR_PL    = str(__import__("pathlib").Path(_PL_POKJA_ROOT) / "Paket Experiment - Pengadaan Langsung" / "Development - PL - JKK")
        _TEMPLATE_DIR_PL_PK = str(__import__("pathlib").Path(_PL_POKJA_ROOT) / "Paket Experiment - Pengadaan Langsung" / "Development - PL - PK")

        _PL_PY     = _POKJA_PYTHON
        _PL_SCRIPT = str(pathlib.Path(_V19_ROOT) / "setup_paket_baru.py")
        _PL_NO_WIN = 0x08000000

        _pl_rows = _load_draft_pl_cached()

        # ── Buang duplikat row lama (paket di-ulang → kode baru, row lama nyangkut) ──
        _pl_rows, _pl_dup_n = pl_engine.buang_duplikat_paket_lama(_pl_rows)
        if _pl_dup_n:
            st.caption(f"♻️ {_pl_dup_n} row lama duplikat (paket ulang) disembunyikan otomatis.")

        # ── #4: Filter paket selesai (penandatanganan kontrak) ──────────────────
        _pl_show_done = st.checkbox(
            "Tampilkan paket selesai (sudah teken kontrak)",
            value=False,
            key="pl_show_done",
        )
        _pl_done_n = sum(1 for r in _pl_rows if pl_engine.is_paket_selesai(r))
        if not _pl_show_done:
            _pl_rows = [r for r in _pl_rows if not pl_engine.is_paket_selesai(r)]
            if _pl_done_n:
                st.caption(f"🔒 {_pl_done_n} paket selesai (Penandatanganan Kontrak) disembunyikan — centang di atas untuk tampilkan.")

        _pl_col_kiri, _pl_col_kanan = st.columns(2)

        # ══════════════════════════════════════════════════════
        # KOLOM KIRI — Serap Data + Daftar Paket
        # ══════════════════════════════════════════════════════
        with _pl_col_kiri:
            # ── #1: Gabung Serap SPSE + MAK + HPS ──────────────────────────────
            st.markdown("#### 1. Serap Data Paket PL")
            st.caption("Pilih aksi lalu klik tombol — aksi berjalan berurutan sesuai centang.")
            _cb_serap_spse = st.checkbox("Serap dari SPSE (daftar paket + status)", value=True, key="pl_cb_serap_spse")
            _cb_serap_mak  = st.checkbox("Serap MAK dari Inbox PL",               value=True, key="pl_cb_serap_mak")

            if st.button("🚀 Serap Data Paket PL", type="primary", use_container_width=True, key="btn_serap_pl_gabung"):
                # Aksi 1: Serap dari SPSE
                if _cb_serap_spse:
                    import spse_browser as _sb_pl
                    _pl_cookie = _sb_pl.get_spse_cookies()
                    if not _pl_cookie:
                        st.error("Cookie SPSE kosong — buka Brave SPSE dan login sebagai PP.")
                    else:
                        _pl_pb = st.progress(0.0)
                        _pl_st = st.empty()
                        _pl_logs = []
                        def _pl_log(msg):
                            _pl_logs.append(msg)
                            _pl_st.info(msg)
                        from config import SPSE_BASE_URL as _SPSE_BASE
                        _pl_hasil = pl_engine.serap_paket_pl_dari_spse(
                            _pl_cookie, _SPSE_BASE, log_fn=_pl_log
                        )
                        _pl_pb.progress(1.0)
                        _pl_c1, _pl_c2 = st.columns(2)
                        _pl_c1.metric("✅ Tersimpan", _pl_hasil.get("scraped", 0))
                        _pl_c2.metric("❌ Error", len(_pl_hasil.get("errors", [])))
                        if _pl_hasil.get("errors"):
                            with st.expander("Detail Error SPSE"):
                                for _e in _pl_hasil["errors"]:
                                    st.error(_e)
                        with st.expander("📋 Log lengkap serap"):
                            st.text("\n".join(_pl_logs))
                        # Invalidasi cache agar hasil serap SPSE langsung terbaca.
                        _load_draft_pl_cached.clear()
                        # Reload setelah serap SPSE agar data paket terkini
                        # Cache JKK dipisahkan dari cache PK.
                        _pl_rows = _load_draft_pl_cached()
                        if not _pl_show_done:
                            _pl_rows = [r for r in _pl_rows if not pl_engine.is_paket_selesai(r)]

                # Aksi 2: Serap MAK dari Inbox
                if _cb_serap_mak:
                    import inbox_engine as _ibe
                    _pb_mak = st.progress(0.0)
                    _st_mak = st.empty()
                    _logs_mak = []
                    def _cb_mak(p, m):
                        _pb_mak.progress(min(max(p, 0.0), 1.0))
                        _logs_mak.append(m)
                        _st_mak.info(m)
                    try:
                        _r_mak = _ibe.serap_inbox_pl(progress_cb=_cb_mak)
                        _c1, _c2, _c3 = st.columns(3)
                        _c1.metric("Pesan parse", _r_mak.get("scraped", 0))
                        _c2.metric("Paket update", _r_mak.get("matched", 0))
                        _c3.metric("Error", len(_r_mak.get("errors", [])))
                        if _r_mak.get("errors"):
                            with st.expander("Detail Error MAK"):
                                for _e in _r_mak["errors"]:
                                    st.warning(_e)
                    except Exception as _e:
                        st.error(f"Gagal serap MAK: {_e}")

            st.divider()
            st.markdown("#### 2. Daftar Paket PL")

            _pl_filter = st.selectbox(
                "Filter:",
                ["Semua", "JKK", "PK", "Belum Folder", "Sudah Folder"],
                key="pl_filter_jenis",
            )

            def _pl_match(r):
                _jp = (r.get("jenis_pl") or "").upper()
                if _pl_filter == "JKK":    return _jp == "JKK"
                if _pl_filter == "PK":     return _jp == "PK"
                if _pl_filter == "Belum Folder": return not bool(r.get("folder_dibuat"))
                if _pl_filter == "Sudah Folder": return bool(r.get("folder_dibuat"))
                return True

            _pl_filtered = [r for r in _pl_rows if _pl_match(r)]

            _pl_peserta_map = _fetch_status_semua_paket_cached(
                tuple(r.get("kode_paket", "") for r in _pl_filtered if r.get("kode_paket"))
            )

            if not _pl_filtered:
                st.info("Belum ada paket PL. Klik 'Serap dari SPSE' atau tambah manual.")
            else:
                for _pr in _pl_filtered:
                    _pr_kode   = _pr.get("kode_paket", "")
                    _pr_nama   = _pr.get("nama_paket", "-")
                    _pr_jenis  = (_pr.get("jenis_pl") or "").upper()
                    _pr_hps    = _pr.get("nilai_hps", "-")
                    _pr_status = _pr.get("status", "draft")
                    _pr_folder = bool(_pr.get("folder_dibuat"))
                    _pr_icon   = "✅" if _pr_folder else "📋"
                    # Label ringkas: nomor urut folder, bukan nama metode panjang.
                    _pr_folder_raw = str(_pr.get("folder_dibuat") or "")
                    _pr_no_match = re.search(r"^\s*(\d+)", os.path.basename(_pr_folder_raw.rstrip("\\/")))
                    _pr_nomor = (_pr_no_match.group(1) if _pr_no_match else str(_pr.get("nomor_urut") or "").strip()) or _pr_kode
                    _pr_metode_raw = _pr.get("metode_pengadaan", "") or ""
                    _pr_metode_low = _pr_metode_raw.lower()
                    _pr_label  = f"{_pr_icon} {_pl_label(_pr)}"

                    with st.expander(_pr_label):
                        st.caption(f"`{_pr_kode}` | HPS: {_pr_hps} | Status: **{_pr_status}**")
                        if _pr_jenis == "JKK":
                            st.caption(f"Workflow target SPSE: **{_pl_workflow(_pr)}**")
                        # Badge peserta pendaftaran — pakai kode_paket (bukan id_nontender)
                        _pr_id_nt = _pr.get("kode_paket", "") or _pr.get("id_nontender", "")
                        if _pr_id_nt:
                            _pr_jml_peserta = _pl_peserta_map.get(_pr_id_nt, {}).get("jumlah", -1)
                            if _pr_jml_peserta > 0:
                                st.success(f"✅ {_pr_jml_peserta} peserta sudah mendaftar")
                            elif _pr_jml_peserta == 0:
                                st.warning("⚠️ Belum ada peserta mendaftar")
                        if "non konstruksi" in _pr_metode_low:
                            st.warning("⚠️ Metode: Non Konstruksi — minta PPK ubah ke Konstruksi di SPSE.")

                        # ── Info DPA (match nama pekerjaan fisik dari DPA yg sudah diimport di tab DPA) ──
                        import dpa_engine as _dpa_pl
                        if "pl_dpa_pool" not in st.session_state:
                            try:
                                from config import sb as _sb_dpa_pl
                                st.session_state["pl_dpa_pool"] = _dpa_pl.load_pool_dpa(_sb_dpa_pl())
                            except Exception:
                                st.session_state["pl_dpa_pool"] = []
                        _pl_dpa_key = f"pl_dpa_{_pr_kode}"
                        if _pl_dpa_key not in st.session_state:
                            st.session_state[_pl_dpa_key] = _dpa_pl.cari_dpa_di_pool(
                                _pr_nama, st.session_state["pl_dpa_pool"]
                            )
                        _pl_dpa_rows = st.session_state.get(_pl_dpa_key, [])
                        if _pl_dpa_rows:
                            def _rp_pl(v):
                                try: return f"Rp {int(v):,}".replace(",", ".")
                                except: return str(v) if v else "-"
                            _pl_dpa_nama = _pl_dpa_rows[0].get("nama_paket") or "-"
                            # Expander tidak bisa nested — pakai checkbox toggle
                            if st.checkbox(f"📊 Info DPA — {_pl_dpa_nama[:40]}", key=f"pl_dpa_show_{_pr_kode}"):
                                _pl_dpa_tbl = []
                                for _it in _pl_dpa_rows:
                                    _sk = (_it.get("subkegiatan_id") or "").split("|")
                                    _pl_dpa_tbl.append({
                                        "Uraian": _it.get("uraian") or "-",
                                        "Spesifikasi": _it.get("spesifikasi") or "-",
                                        "Sebelum": _rp_pl(_it.get("jumlah_sebelum")),
                                        "Sesudah": _rp_pl(_it.get("jumlah_sesudah")),
                                        "Sub Kegiatan": _sk[-1] if len(_sk) >= 3 else (_it.get("subkegiatan_id") or "-"),
                                    })
                                st.table(_pl_dpa_tbl)

                        # Ubah Metode Pengadaan inline per paket
                        with st.container(border=True):
                            st.markdown("**🔧 Ubah Metode Pengadaan**")
                            _pr_metode_pilihan = st.selectbox(
                                "Target metode:",
                                list(pl_engine.METODE_PL_MAP.keys()),
                                 index=list(pl_engine.METODE_PL_MAP.keys()).index("JKK Konstruksi — PL"),
                                key=f"pl_ubah_metode_target_{_pr_kode}",
                            )
                            _pr_kat_id, _pr_pilih_val = pl_engine.METODE_PL_MAP[_pr_metode_pilihan]
                            if st.button(
                                "🔄 Ubah Metode via CDP",
                                use_container_width=True,
                                key=f"pl_btn_ubah_metode_{_pr_kode}",
                            ):
                                _pl_base_ubah = pl_engine.BASE_URL + "/"
                                if pl_engine.ubah_metode_pl_playwright(_pr_kode, _pr_kat_id, _pr_pilih_val, _pl_base_ubah):
                                    st.success("✅ Berhasil ubah metode. Serap ulang untuk refresh.")
                                else:
                                    st.error("❌ Gagal ubah metode.")

                        # Aksi paket: tiga operasi utama satu baris; aksi panjang di baris kedua.
                        _pr_c1, _pr_c2, _pr_c3 = st.columns([1.5, 1.1, 1.1])
                        if _pr_folder and _pr_c1.button("📊 Isi Excel", key=f"pl_excel_{_pr_kode}", use_container_width=True):
                            import isi_master_data_pl as _imd_pr
                            import kualifikasi_engine_pl as _keng_pr_excel
                            # Status container tidak boleh berada di dalam expander paket.
                            _pr_excel_status = st.empty()
                            _pr_excel_status.info("📊 Menyiapkan isi ulang Excel...")
                            _pr_excel_log = st.empty()
                            _pr_excel_logs = []
                            def _pr_excel_cb(_msg):
                                _pr_excel_logs.append(str(_msg))
                                _pr_excel_log.code("\n".join(_pr_excel_logs[-20:]))
                                _pr_excel_status.info(f"📊 {str(_msg)[:90]}")
                            _pr_excel_cb("Mencari folder paket...")
                            _pr_excel_fr = _keng_pr_excel.resolve_folder_paket_pl(_pr_kode)
                            _pr_excel_root = _pr_excel_fr.get("pesan", "") if _pr_excel_fr.get("ok") else ""
                            _pr_excel_xlsm = _cari_xlsm_pl(_pr_excel_root) if _pr_excel_root else None
                            if not _pr_excel_xlsm:
                                _pr_excel_status.error("❌ Folder/workbook .xlsm tidak ditemukan.")
                            else:
                                _pr_excel_cb(f"Workbook: {_pr_excel_xlsm}")
                                _pr_excel_res = _imd_pr.isi_master_data_pl(_pr_kode, _pr_excel_xlsm, progress_cb=_pr_excel_cb)
                                if _pr_excel_res.get("ok"):
                                    _pr_excel_status.success("✅ Excel selesai diisi.")
                                    st.success(_pr_excel_res.get("pesan", "Excel selesai diisi."))
                                else:
                                    _pr_excel_status.error("❌ Gagal mengisi Excel.")
                                    st.error(_pr_excel_res.get("pesan", "Gagal mengisi Excel."))
                        if _pr_folder and _pr_c2.button("📦 Unduh", key=f"pl_dl_{_pr_kode}", use_container_width=True):
                            import kualifikasi_engine_pl as _keng_pr_dl
                            _pr_dl_fr = _keng_pr_dl.resolve_folder_paket_pl(_pr_kode)
                            _pr_dl_root = _pr_dl_fr.get("pesan", "") if _pr_dl_fr.get("ok") else ""
                            if not _pr_dl_root:
                                st.error("Folder tidak ditemukan.")
                            else:
                                _pr_dl_logs = []
                                _pr_dl_label = st.empty()
                                _pr_dl_area = st.empty()
                                _pr_dl_label.markdown("🔽 **Mengunduh...**")
                                def _pr_dl_cb(msg, _l=_pr_dl_logs, _a=_pr_dl_area, _b=_pr_dl_label):
                                    _l.append(msg); _a.code("\n".join(_l[-15:])); _b.markdown(f"🔽 **{msg[:50]}**")
                                pl_engine.buat_subfolder_dokumen(_pr_dl_root)
                                _pr_dl_res = pl_engine.download_dokumen_paket_pl(_pr_kode, _pr_dl_root, _pr_dl_cb, force_clean=True)
                                _pr_dl_label.markdown(f"✅ **{len(_pr_dl_res['ok'])} file, ❌ {len(_pr_dl_res['error'])} error**")
                                _pr_kak_p = parse_kak_pl.cari_kak_di_folder(_pr_dl_root)
                                if _pr_kak_p:
                                    _pr_kak_d = parse_kak_pl.parse_kak(_pr_kak_p)
                                    _pr_kak_u = {k: v for k, v in _pr_kak_d.items() if v}
                                    if _pr_kak_u:
                                        pl_engine.simpan_paket_pl({"kode_paket": _pr_kode, **_pr_kak_u})
                                        st.info(f"📋 KAK: {', '.join(_pr_kak_u.keys())}")
                                # Unduh hanya memperbarui dokumen SPSE dan metadata KAK.
                                # Pengisian workbook sengaja tetap eksklusif di tombol Isi Excel.
                                _load_draft_pl_cached.clear()
                                st.rerun()
                        if _pr_folder and _pr_c3.button("💰 HPS", key=f"pl_hps_{_pr_kode}", use_container_width=True):
                            import hps_engine as _hps_pr
                            import kualifikasi_engine_pl as _keng_pr
                            _pr_fr = _keng_pr.resolve_folder_paket_pl(_pr_kode)
                            _pr_root = _pr_fr.get("pesan", "") if _pr_fr.get("ok") else ""
                            _pr_xl = _cari_xlsm_pl(_pr_root) if _pr_root else None
                            if _pr_xl:
                                with st.spinner("Scrape HPS..."):
                                    _pr_hr = _hps_pr.scrape_hps_pl_ke_excel(_pr_kode, _pr_xl)
                                if _pr_hr.get("ok"):
                                    st.success(f"✅ {_pr_hr['count']} item HPS")
                                else:
                                    st.error(_pr_hr.get("pesan", "-"))
                            else:
                                st.error("Folder/xlsm tidak ditemukan.")
                        # Aksi lanjutan: dua tombol berdampingan, sinkronisasi full-width
                        # agar label panjang tetap terbaca pada layar sempit.
                        _pr_r1, _pr_r2 = st.columns([1, 1])
                        if _pr_folder and _pr_r2.button("🤖 Pra-Reviu + Isi DOCM", key=f"pl_pra_reviu_{_pr_kode}", use_container_width=True):
                            import ai_evaluator as _heval_one
                            with st.spinner("Menjalankan pra-reviu paket..."):
                                _one_res = _heval_one.evaluasi_bulk(
                                    [{"nomor_urut": _pr.get("nomor_urut"), "nama_paket": _pr_nama}],
                                    jenis="pra_reviu", model=None, max_workers=1,
                                    engine="codex",
                                )
                            if _one_res and _one_res[0].get("status") == "ok":
                                st.success(f"✅ Pra-reviu {_pr_nama[:50]} selesai.")
                            else:
                                st.error((_one_res[0].get("error") if _one_res else "Pra-reviu gagal")[:300])
                        if _pr_folder and _pr_r1.button("🔄 Reset status folder", key=f"pl_reset_{_pr_kode}", use_container_width=True):
                            from config import sb as _sb_reset_one
                            try:
                                _sb_reset_one().table("draft_paket_pl").update({"folder_dibuat": None}).eq("kode_paket", _pr_kode).execute()
                                _load_draft_pl_cached.clear()
                                st.success(f"✅ Status folder {_pr_nama[:45]} direset.")
                                st.rerun()
                            except Exception as _reset_one_e:
                                st.error(f"Reset gagal: {_reset_one_e}")
                            st.rerun()
                        if _pr_folder and st.button("🔄 Sinkronkan Workflow", key=f"pl_migrate_wf_{_pr_kode}", use_container_width=True):
                            import kualifikasi_engine_pl as _keng_pr_wf
                            _wf_fr = _keng_pr_wf.resolve_folder_paket_pl(_pr_kode)
                            _wf_root = _wf_fr.get("pesan", "") if _wf_fr.get("ok") else ""
                            if not _wf_root:
                                st.error("Folder tidak ditemukan.")
                            else:
                                with st.spinner("Memeriksa dan memigrasikan workflow..."):
                                    _wf_res = migrate_pl_workflow(_wf_root, _pr, expected_family="JKK")
                                if _wf_res.get("ok"):
                                    st.success(_wf_res.get("message", "Workflow sudah sesuai."))
                                    if _wf_res.get("status") == "MIGRATED":
                                        st.info(f"Backup template lama: {_wf_res.get('backup')}")
                                    _load_draft_pl_cached.clear()
                                    st.rerun()
                                else:
                                    st.error(f"Workflow belum dimigrasikan: {_wf_res.get('message', '-')}")

        # ══════════════════════════════════════════════════════
        # KOLOM KANAN — Buat Folder + Download Dokumen
        # ══════════════════════════════════════════════════════

            # ── Ubah Metode Bulk (semua paket di daftar) ──────────────
            if _pl_rows:
                st.divider()
                with st.container(border=True):
                    st.markdown("**🔧 Ubah Metode Pengadaan — Semua Paket**")
                    _pl_opsi_ubah_bulk = {r.get("nama_paket", r.get("kode_paket")): r.get("kode_paket") for r in _pl_filtered}
                    _pl_sel_ubah_bulk = list(_pl_opsi_ubah_bulk.keys())
                    _pl_metode_bulk = st.selectbox(
                        "Target metode:",
                        list(pl_engine.METODE_PL_MAP.keys()),
                         index=list(pl_engine.METODE_PL_MAP.keys()).index("JKK Konstruksi — PL"),
                        key="pl_ubah_metode_bulk_target",
                    )
                    _pl_kat_id_bulk, _pl_pilih_val_bulk = pl_engine.METODE_PL_MAP[_pl_metode_bulk]
                    if st.button(
                        f"🔄 Ubah Metode ({len(_pl_sel_ubah_bulk)} paket) via CDP",
                        disabled=not _pl_sel_ubah_bulk,
                        use_container_width=True,
                        key="pl_btn_ubah_metode_bulk",
                    ):
                        _pl_base_ubah_b = pl_engine.BASE_URL + "/"
                        _pl_ok_b = _pl_fail_b = 0
                        for _nm_b in _pl_sel_ubah_bulk:
                            _kd_b = _pl_opsi_ubah_bulk[_nm_b]
                            _hasil_b = spse_browser.ubah_metode_via_playwright(_kd_b, _pl_kat_id_bulk, _pl_pilih_val_bulk, _pl_base_ubah_b)
                            if _hasil_b == "OK":
                                _pl_ok_b += 1
                                st.write(f"✅ {_nm_b[:45]}")
                            else:
                                _pl_fail_b += 1
                                st.write(f"❌ {_nm_b[:45]} — `{_hasil_b}`")
                        st.success(f"Selesai: {_pl_ok_b} OK, {_pl_fail_b} GAGAL.")

        with _pl_col_kanan:
            st.markdown("#### 3. Buat Folder Paket")

            if "pl_folder_just_created" in st.session_state:
                _msg = st.session_state.pop("pl_folder_just_created")
                st.toast(f"✅ {_msg}", icon="📁")
                st.success(f"✅ {_msg}")
                st.balloons()

            # Dropdown pilih paket
            def _pl_no_dari_nama(nama: str, fallback: int) -> int:
                """Ekstrak nomor dari nama paket, misal 'Paket 1' → 1."""
                import re as _re
                m = _re.search(r"Paket\s+(\d+)", nama, _re.IGNORECASE)
                return int(m.group(1)) if m else fallback

            _pl_dl_dokumen = st.checkbox("📦 Download dokumen SPSE (KAK, Personil, Kontrak) saat buat folder", value=True, key="pl_cb_dl")
            _pl_rt_refresh = False
            _pl_extract_teks = False
            if not st.session_state.get("pl_cb_isi_excel_default_v2"):
                st.session_state["pl_cb_isi_excel"] = True
                st.session_state["pl_cb_isi_excel_default_v2"] = True
            _pl_isi_excel = st.checkbox("📊 Isi Excel @ Master Data (wajib jika workbook langsung dipakai)", key="pl_cb_isi_excel")

            # ── Bulk: Buat Semua Folder ──────────────────────────────
            st.divider()

            _pl_rows_belum = [
                r for r in _pl_rows
                if r.get("nama_paket") and not r.get("folder_dibuat")
            ]

            # Plan: pre-compute nama folder per paket
            # Deteksi nomor tertinggi yang sudah ada di masing-masing output_base
            _pl_no_global = max(
                pl_engine.nomor_folder_tertinggi(_PL_DIR_JKK),
                pl_engine.nomor_folder_tertinggi(_PL_DIR_PK),
            )
            _pl_no_offset = {_PL_DIR_JKK: _pl_no_global, _PL_DIR_PK: _pl_no_global}
            _pl_bulk_plan = []
            for _bi0, _br0 in enumerate(_pl_rows_belum, 1):
                _bnm0  = _br0.get("nama_paket", "")
                _bj0   = (_br0.get("jenis_pl") or "JKK").upper()
                _bpfx0 = {"JKK": "PLJKK", "PK": "PLPK"}.get(_bj0, f"PL{_bj0}")
                _bout_base0  = _PL_DIR_JKK if _bj0 == "JKK" else _PL_DIR_PK
                # Nomor = offset (folder tertinggi di disk) + urutan paket ini
                _bno0  = _pl_no_offset[_bout_base0] + _bi0
                _bnm_folder0 = re.sub(r'[/<>:"\|?*]', "-", f"{_bno0}. {_bpfx0} - {_bnm0}").strip()
                _bnm_folder0 = pl_engine.nama_folder_dengan_suffix_ulang(_bout_base0, _bnm_folder0)
                _bnm_folder0 = pl_engine.truncate_nama_folder(_bout_base0, _bnm_folder0)
                _pl_bulk_plan.append({
                    "kode_paket": _br0.get("kode_paket", ""),
                    "nama_folder": _bnm_folder0,
                    "out_base": _bout_base0,
                    "jenis_pl": _bj0,
                    "workflow": _pl_workflow(_br0),
                    "template_dir": _template_dir_pl_jkk(_br0, _TEMPLATE_DIR_PL) if _bj0 == "JKK" else _template_dir_pl_jkk(_br0, _TEMPLATE_DIR_PL_PK),
                })

            
            # ── #2: Checklist pilih paket untuk buat folder ──────────────────
            if _pl_rows_belum:
                st.markdown("**Pilih paket yang akan dibuat foldernya:**")
                _plf_col1, _plf_col2 = st.columns(2)
                # Tombol pilih semua / batal semua (tiru pola Tab 5)
                if _plf_col1.button("✅ Pilih Semua", key="plf_pilih_semua", use_container_width=True):
                    for _br_chk in _pl_rows_belum:
                        st.session_state[f"plf_chk_{_br_chk.get('kode_paket','')}"] = True
                    st.rerun()
                if _plf_col2.button("❌ Batal Semua", key="plf_batal_semua", use_container_width=True):
                    for _br_chk in _pl_rows_belum:
                        st.session_state[f"plf_chk_{_br_chk.get('kode_paket','')}"] = False
                    st.rerun()
                for _br_chk in _pl_rows_belum:
                    _bkp_chk = _br_chk.get("kode_paket", "")
                    _plf_chk_key = f"plf_chk_{_bkp_chk}"
                    if _plf_chk_key not in st.session_state:
                        st.session_state[_plf_chk_key] = True
                    st.checkbox(
                        f"{_pl_label(_br_chk)} — {(_br_chk.get('jenis_pl') or '').upper()}",
                        key=_plf_chk_key,
                    )
                # Hitung yang dicentang untuk label tombol
                _pl_terpilih_plan = [
                    item for item in _pl_bulk_plan
                    if st.session_state.get(f"plf_chk_{item['kode_paket']}", True)
                ]
                if st.button(
                    f"📁 Buat Folder Terpilih ({len(_pl_terpilih_plan)} paket)",
                    disabled=len(_pl_terpilih_plan) == 0,
                    use_container_width=True,
                    key="pl_btn_buat_terpilih",
                    type="primary",
                ):
                    import time as _pl_time
                    _pl_t0 = _pl_time.perf_counter()
                    _pl_bp = st.progress(0.0)
                    _pl_bulk_status = st.status(f"📁 Memproses {len(_pl_terpilih_plan)} paket terpilih... · ⏱ 0m 0d", expanded=True)
                    _pl_bulk_status_line = _pl_bulk_status.empty()
                    _pl_ok, _pl_fail = 0, 0
                    _pl_bulk_semua_log = {}
                    # ── FASE 1: I/O paralel antar-paket (download+parse+serap+HPS.md) ──
                    import queue as _pl_queue
                    _pl_event_q = _pl_queue.Queue()
                    _pl_live_events = ["⏱ Timer mulai"]
                    _pl_cfg_io = {
                        "py": _PL_PY, "script": _PL_SCRIPT, "no_win": _PL_NO_WIN,
                        "pokja_root": _PL_POKJA_ROOT, "dl_dokumen": bool(_pl_dl_dokumen),
                        "event_q": _pl_event_q,
                    }
                    try:
                        import spse_browser as _pl_sb_io
                        _pl_cookie = _pl_sb_io.get_spse_cookies()  # ambil 1× untuk semua paket
                    except Exception as _ck_e:
                        _pl_cookie = ""
                    _pl_io_hasil = []
                    _pl_done_ct = 0
                    _pl_n_total = len(_pl_terpilih_plan)
                    with ThreadPoolExecutor(max_workers=4) as _pl_ex:
                        from concurrent.futures import wait as _pl_wait, FIRST_COMPLETED as _PL_FIRST_COMPLETED
                        _pl_pending = {
                            _pl_ex.submit(_pl_proses_io_satu_paket, _it, _pl_cookie, _pl_cfg_io)
                            for _it in _pl_terpilih_plan
                        }
                        while _pl_pending:
                            _pl_done, _pl_pending = _pl_wait(_pl_pending, timeout=0.5, return_when=_PL_FIRST_COMPLETED)
                            while not _pl_event_q.empty():
                                _pl_live_events.append(_pl_event_q.get())
                                _pl_live_events = _pl_live_events[-12:]
                            if _pl_live_events:
                                _pl_bulk_status_line.code("\n".join(_pl_live_events))
                            for _pl_fut in _pl_done:
                                _pl_res = _pl_fut.result()
                                _pl_io_hasil.append(_pl_res)
                                _pl_done_ct += 1
                                _pl_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                                _pl_bp.progress(_pl_done_ct / max(_pl_n_total, 1))
                                _pl_bulk_status.update(label=f"[{_pl_done_ct}/{_pl_n_total}] selesai: {_pl_res['nama_folder'][:50]} · ⏱ {_pl_elapsed}")
                                _pl_live_events.append(f"✅ SELESAI: {_pl_res['nama_folder'][:55]} · ⏱ {_pl_elapsed}")
                                _pl_live_events = _pl_live_events[-12:]
                                _pl_bulk_status_line.code("\n".join(_pl_live_events))
                    # ── FASE 2: serial (COM/merge/OCR) di main thread ──
                    for _pl_res in _pl_io_hasil:
                        _pl_nf = _pl_res["nama_folder"]
                        _pl_kp_b = _pl_res["kode"]
                        _pl_target_b = _pl_res["target"]
                        _pl_paket_log = _pl_res["log"]
                        if not _pl_res["ok"]:
                            _pl_fail += 1
                            _pl_bulk_semua_log[_pl_nf] = _pl_paket_log
                            continue
                        _pl_ok += 1
                        _pl_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                        _pl_bulk_status.update(label=f"Finalisasi: {_pl_nf[:55]} · ⏱ {_pl_elapsed}")
                        # Merge PDF draft (COM tidak thread-safe → serial)
                        if _pl_res.get("files_ok"):
                            _t_step = _pl_time.perf_counter()
                            try:
                                _pl_merged = pl_engine.gabung_draft_pl(_pl_kp_b, _pl_target_b, _pl_res["files_ok"])
                                if _pl_merged:
                                    _pl_paket_log.append(f"📎 Draft PDF: {_pl_os.path.basename(_pl_merged)}")
                                _pl_paket_log.append(f"⏱ merge draft: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _mg_e:
                                _pl_paket_log.append(f"⚠ Gabung Draft PDF: {_mg_e}")
                                _pl_paket_log.append(f"⏱ merge draft: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ merge draft: skipped (0 file)")
                        # Extract teks kualifikasi (OCR berat → serial dgn timeout)
                        if _pl_extract_teks and _pl_dl_dokumen:
                            _t_step = _pl_time.perf_counter()
                            try:
                                import extract_teks_kualifikasi as _etk
                                import threading as _etk_th
                                _etk_folder = _pl_os.path.join(_pl_target_b, "8. Dokumen Kualifikasi")
                                if _pl_os.path.isdir(_etk_folder):
                                    _etk_result_box = [None]
                                    def _etk_run(_f=_etk_folder, _log=_pl_paket_log):
                                        try:
                                            _etk_result_box[0] = _etk.extract_folder_kualifikasi(
                                                _f, progress_cb=lambda m: _log.append(f"  {m}"),
                                            )
                                        except Exception as _ex:
                                            _etk_result_box[0] = {"ok": False, "error": str(_ex)}
                                    _etk_t = _etk_th.Thread(target=_etk_run, daemon=True)
                                    _etk_t.start()
                                    _etk_t.join(timeout=120)
                                    if _etk_t.is_alive():
                                        _pl_paket_log.append("⚠ Extract teks: timeout 120s, dilewati")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} timeout")
                                    elif _etk_result_box[0] and _etk_result_box[0].get("ok"):
                                        _etk_res = _etk_result_box[0]
                                        _pl_paket_log.append(f"📝 Extract teks: {len(_etk_res.get('penyedia', []))} penyedia, ~{_etk_res.get('total_token_estimasi', 0)} token")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                                    else:
                                        _pl_paket_log.append("⚠ Extract teks: tidak ada penyedia/PDF")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _etk_e:
                                _pl_paket_log.append(f"⚠ Extract teks: {_etk_e}")
                                _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ OCR: skipped")
                        # Refresh + HPS + Master Data: COM Excel (serial)
                        if _pl_isi_excel:
                            _t_step = _pl_time.perf_counter()
                            try:
                                _excel_logs = _proses_excel_paket_pl(
                                    _pl_target_b, _pl_kp_b,
                                    _pl_res["jenis_pl"], _pl_rt_refresh,
                                    _pl_res.get("template_dir") or _TEMPLATE_DIR_PL, _TEMPLATE_DIR_PL_PK,
                                )
                                for _el in _excel_logs:
                                    _icon = "📊" if _el.startswith("HPS:") else (
                                            "📝" if _el.startswith("Master Data") else (
                                            "🔄" if _el.startswith("Refresh") else "⚠"))
                                    _pl_paket_log.append(f"{_icon} {_el}")
                                _pl_paket_log.append(f"⏱ Excel: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _xl_e:
                                _pl_paket_log.append(f"⚠ Excel Master Data: {_xl_e}")
                                _pl_paket_log.append(f"⏱ Excel: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ Excel: skipped")
                        try:
                            _meta_wf = mark_workflow_applied(
                                _pl_target_b, _pl_res.get("workflow") or _pl_workflow({"jenis_pl": _pl_res["jenis_pl"], "nama_paket": _pl_nf}), _pl_res["jenis_pl"]
                            )
                            _pl_paket_log.append(f"🧩 Workflow: {_meta_wf.get('workflow_applied')}")
                        except Exception as _meta_wf_e:
                            _pl_paket_log.append(f"⚠ Metadata workflow: {_meta_wf_e}")
                        _pl_bulk_semua_log[_pl_nf] = _pl_paket_log
                    _pl_total_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                    _pl_live_events.append(f"⏱ Total waktu: {_pl_total_elapsed}")
                    _pl_bulk_status_line.code('\n'.join(_pl_live_events[-12:]))
                    _pl_ringkasan = f"✅ {_pl_ok} folder berhasil, ❌ {_pl_fail} gagal · ⏱ {_pl_total_elapsed}"
                    _pl_bulk_status.update(label=_pl_ringkasan, state="complete", expanded=False)
                    with st.expander("📋 Log detail per paket", expanded=_pl_fail > 0):
                        for _pl_nf, _pl_logs in _pl_bulk_semua_log.items():
                            st.markdown(f"**{_pl_nf[:70]}**")
                            st.code("\n".join(_pl_logs))
                    st.session_state["pl_folder_bulk_created"] = _pl_ringkasan
                    _load_draft_pl_cached.clear()
            else:
                st.info("✅ Semua paket sudah punya folder.")
                st.button("📁 Buat Folder Terpilih (0 paket)", disabled=True, use_container_width=True, key="pl_btn_buat_terpilih_disabled")
            # ── #4: Update Data Folder (Re-download + Reset) ─────────────────
            st.divider()
            st.markdown("#### 4. Sinkronisasi & Pemutakhiran Paket Existing")
            st.caption(
                "⚠️ **Sudah dijalankan otomatis saat Buat Folder** (jika checkbox Download dicentang). "
                "Gunakan tombol ini **hanya** jika ada perubahan dokumen dari PPK atau ingin refresh ulang."
            )
            _cb_dl_dok_bulk = st.checkbox("📦 Re-download Dokumen SPSE (KAK, Personil, Kontrak)", value=False, key="pl_cb_dl_dok_bulk")
            _cb_template_refresh = st.checkbox("🔄 Refresh Template ke folder PL existing", value=False, key="pl_cb_template_refresh_existing")
            _cb_workflow_migrate = st.checkbox("🔄 Sinkronkan workflow PLJKK yang berubah", value=False, key="pl_cb_workflow_migrate_existing")
            _cb_hps_update = st.checkbox("💰 Update HPS semua paket berfolder → Excel + MD", value=False, key="pl_cb_hps_update")
            _cb_excel_existing = st.checkbox(
                "📊 Isi ulang Excel @ Master Data semua paket berfolder",
                value=False,
                key="pl_cb_excel_existing",
            )
            _cb_reset_existing = st.checkbox(
                "🔄 Reset status folder semua paket berfolder",
                value=False,
                key="pl_cb_reset_existing",
            )
            _cb_pra_reviu_bulk = st.checkbox(
                "🤖 Pra-Reviu + Isi DOCM semua paket berfolder",
                value=False,
                key="pl_cb_pra_reviu_bulk",
            )

            if st.button("▶ Jalankan Operasi Terpilih", use_container_width=True, key="btn_update_data_folder"):
                if _cb_workflow_migrate:
                    import kualifikasi_engine_pl as _keng_wf_bulk
                    _wf_rows_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat") and (r.get("jenis_pl") or "JKK").upper() == "JKK"
                    ]
                    _wf_bulk_status = st.status(f"🔄 Sinkronkan workflow {len(_wf_rows_bulk)} paket JKK...", expanded=True)
                    _wf_bulk_line = _wf_bulk_status.empty()
                    _wf_ok = _wf_fail = 0
                    for _wf_i, _wf_row in enumerate(_wf_rows_bulk, 1):
                        try:
                            _wf_fr = _keng_wf_bulk.resolve_folder_paket_pl(_wf_row.get("kode_paket", ""))
                            _wf_root = _wf_fr.get("pesan", "") if _wf_fr.get("ok") else ""
                            _wf_res = migrate_pl_workflow(_wf_root, _wf_row, expected_family="JKK") if _wf_root else {"ok": False, "message": "folder tidak ditemukan"}
                            if _wf_res.get("ok"):
                                _wf_ok += 1
                                _wf_bulk_line.write(f"✅ [{_wf_i}] {_pl_label(_wf_row)} — {_wf_res.get('message', '-')}")
                            else:
                                _wf_fail += 1
                                _wf_bulk_line.write(f"❌ [{_wf_i}] {_pl_label(_wf_row)} — {_wf_res.get('message', '-')}")
                        except Exception as _wf_e:
                            _wf_fail += 1
                            _wf_bulk_line.write(f"❌ [{_wf_i}] {_pl_label(_wf_row)} — {_wf_e}")
                    _wf_bulk_status.update(label=f"🔄 Workflow selesai: ✅ {_wf_ok}, ❌ {_wf_fail}", state="complete", expanded=_wf_fail > 0)
                    _load_draft_pl_cached.clear()

                # Aksi: Refresh template bulk semua paket berfolder
                if _cb_template_refresh:
                    import kualifikasi_engine_pl as _keng_pl_rt
                    from pathlib import Path as _rt_Path
                    from refresh_template import refresh_template_paket as _rt_existing
                    _pl_rows_rt_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    if not _pl_rows_rt_bulk:
                        st.info("Tidak ada paket dengan folder untuk refresh template.")
                    else:
                        _rt_ok, _rt_fail = 0, 0
                        _rt_status = st.status(
                            f"🔄 Refresh template {len(_pl_rows_rt_bulk)} paket existing...", expanded=True
                        )
                        _rt_line = _rt_status.empty()
                        _rt_bp = st.progress(0.0)
                        for _rt_i, _rt_row in enumerate(_pl_rows_rt_bulk):
                            _rt_kp = _rt_row.get("kode_paket", "")
                            _rt_nama = _pl_label(_rt_row)
                            _rt_bp.progress((_rt_i + 1) / len(_pl_rows_rt_bulk))
                            _rt_status.update(label=f"[{_rt_i+1}/{len(_pl_rows_rt_bulk)}] {_rt_nama}")
                            try:
                                _rt_fr = _keng_pl_rt.resolve_folder_paket_pl(_rt_kp)
                                _rt_root = _rt_fr.get("pesan", "") if _rt_fr.get("ok") else ""
                                if not _rt_root or not _pl_os.path.isdir(_rt_root):
                                    raise ValueError("folder tidak ditemukan")
                                _rt_jenis = (_rt_row.get("jenis_pl") or "JKK").upper()
                                _rt_src = _template_dir_pl_jkk(_rt_row, _TEMPLATE_DIR_PL) if _rt_jenis == "JKK" else _template_dir_pl_jkk(_rt_row, _TEMPLATE_DIR_PL_PK)
                                _rt_mode = "pl_jkk" if _rt_jenis == "JKK" else "pl_pk"
                                _rt_existing([_rt_Path(_rt_root)], _rt_Path(_rt_src), _rt_mode, auto_relink=True, dry_run=False)
                                _rt_ok += 1
                                _rt_tpl = "Disdag" if "Disdag" in str(_rt_src) else "Default"
                                _rt_line.write(f"✅ [{_rt_i+1}] {_rt_nama} — template {_rt_tpl}")
                            except Exception as _rt_e:
                                _rt_fail += 1
                                _rt_line.write(f"❌ [{_rt_i+1}] {_rt_nama} — {_rt_e}")
                        _rt_bp.progress(1.0)
                        _rt_line.empty()
                        _rt_status.update(
                            label=f"🔄 Refresh template selesai: ✅ {_rt_ok} sukses, ❌ {_rt_fail} gagal",
                            state="complete", expanded=_rt_fail > 0,
                        )

                # Aksi: Download dokumen bulk semua paket berfolder
                if _cb_dl_dok_bulk:
                    import kualifikasi_engine_pl as _keng_pl_dl
                    _pl_rows_dl_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    if not _pl_rows_dl_bulk:
                        st.info("Tidak ada paket dengan folder untuk download dokumen.")
                    else:
                        _dl_bulk_ok, _dl_bulk_fail = 0, 0
                        _dl_bulk_status = st.status(
                            f"📦 Download dokumen {len(_pl_rows_dl_bulk)} paket...", expanded=True
                        )
                        _dl_bulk_line = _dl_bulk_status.empty()
                        _dl_bulk_bp = st.progress(0.0)
                        for _db_i, _db_row in enumerate(_pl_rows_dl_bulk):
                            _db_kp   = _db_row.get("kode_paket", "")
                            _db_nama = _pl_label(_db_row)
                            _dl_bulk_status.update(label=f"[{_db_i+1}/{len(_pl_rows_dl_bulk)}] {_db_nama}")
                            _dl_bulk_bp.progress((_db_i + 1) / len(_pl_rows_dl_bulk))
                            _db_root = ""
                            try:
                                _db_fr = _keng_pl_dl.resolve_folder_paket_pl(_db_kp)
                                _db_root = _db_fr.get("pesan", "") if _db_fr.get("ok") else ""
                            except Exception:
                                _db_root = ""
                            if not _db_root or not _pl_os.path.isdir(_db_root):
                                _dl_bulk_fail += 1
                                _dl_bulk_line.write(f"⚠ [{_db_i+1}] {_db_nama} — folder tidak ditemukan")
                                continue
                            try:
                                _db_dl_logs = []
                                def _db_dl_cb(msg, _log=_db_dl_logs):
                                    _log.append(msg)
                                _db_dl_res = pl_engine.download_dokumen_paket_pl(_db_kp, _db_root, _db_dl_cb, force_clean=True)
                                _dl_bulk_ok += 1
                                _dl_bulk_line.write(f"✅ [{_db_i+1}] {_db_nama} — {len(_db_dl_res.get('ok',[]))} file")
                                # Parse KAK setelah download
                                _db_kak_p = parse_kak_pl.cari_kak_di_folder(_db_root)
                                if _db_kak_p:
                                    _db_kak_d = parse_kak_pl.parse_kak(_db_kak_p)
                                    _db_kak_u = {k: v for k, v in _db_kak_d.items() if v}
                                    if _db_kak_u:
                                        pl_engine.simpan_paket_pl({"kode_paket": _db_kp, **_db_kak_u})
                            except Exception as _db_e:
                                _dl_bulk_fail += 1
                                _dl_bulk_line.write(f"❌ [{_db_i+1}] {_db_nama} — {_db_e}")
                        _dl_bulk_bp.progress(1.0)
                        _dl_bulk_line.empty()
                        _dl_bulk_status.update(
                            label=f"📦 Download selesai: ✅ {_dl_bulk_ok} sukses, ❌ {_dl_bulk_fail} gagal",
                            state="complete", expanded=_dl_bulk_fail > 0,
                        )


                # Aksi: Update HPS bulk
                if _cb_hps_update:
                    import hps_engine as _hps_upd
                    import kualifikasi_engine_plpk as _keng_hps_upd
                    _pl_rows_hps_upd = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    if not _pl_rows_hps_upd:
                        st.info("Tidak ada paket dengan folder untuk update HPS.")
                    else:
                        _hps_upd_ok, _hps_upd_fail = 0, 0
                        _hps_upd_gagal = []
                        _hps_upd_status = st.status(
                            f"💰 Update HPS {len(_pl_rows_hps_upd)} paket...", expanded=True
                        )
                        _hps_upd_line = _hps_upd_status.empty()
                        _hps_upd_bp = st.progress(0.0)
                        for _hu_i, _hu_row in enumerate(_pl_rows_hps_upd):
                            _hu_kp   = _hu_row.get("kode_paket", "")
                            _hu_nama = _pl_label(_hu_row)
                            _hps_upd_status.update(label=f"[{_hu_i+1}/{len(_pl_rows_hps_upd)}] {_hu_nama}")
                            _hps_upd_bp.progress((_hu_i + 1) / len(_pl_rows_hps_upd))
                            try:
                                _hu_fr = _keng_hps_upd.resolve_folder_paket_pl(_hu_kp)
                                _hu_root = _hu_fr.get("pesan", "") if _hu_fr.get("ok") else ""
                                _hu_xl = _cari_xlsm_pl(_hu_root) if _hu_root and _pl_os.path.isdir(_hu_root) else None
                                if not _hu_xl:
                                    raise ValueError("folder/xlsm tidak ditemukan")
                                _hu_r = _hps_upd.scrape_hps_pl_ke_excel(_hu_kp, _hu_xl)
                                if _hu_r.get("ok"):
                                    _hps_upd_ok += 1
                                    _hps_upd_line.write(f"✅ [{_hu_i+1}] {_hu_nama} — {_hu_r['count']} item")
                                else:
                                    raise ValueError(_hu_r.get("pesan", "-"))
                            except Exception as _hu_e:
                                _hps_upd_fail += 1
                                _hps_upd_gagal.append(f"{_hu_nama}: {_hu_e}")
                                _hps_upd_line.write(f"❌ [{_hu_i+1}] {_hu_nama} — {_hu_e}")
                        _hps_upd_bp.progress(1.0)
                        _hps_upd_line.empty()
                        _hps_upd_status.update(
                            label=f"💰 HPS selesai: ✅ {_hps_upd_ok} sukses, ❌ {_hps_upd_fail} gagal",
                            state="complete", expanded=_hps_upd_fail > 0,
                        )
                        if _hps_upd_gagal:
                            st.warning("Paket gagal:\n" + "\n".join(_hps_upd_gagal))

                # Aksi: Isi ulang @ Master Data + @ Evaluasi pada folder existing
                if _cb_excel_existing:
                    import kualifikasi_engine_pl as _keng_excel_existing
                    import isi_master_data_pl as _imd_excel_existing
                    _pl_excel_existing_rows = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    _excel_existing_ok, _excel_existing_fail = 0, 0
                    _excel_existing_status = st.status(
                        f"📊 Isi ulang Excel {len(_pl_excel_existing_rows)} paket...", expanded=True
                    )
                    _excel_existing_line = _excel_existing_status.empty()
                    _excel_existing_bp = st.progress(0.0)
                    _excel_existing_failures = []
                    for _xe_i, _xe_row in enumerate(_pl_excel_existing_rows):
                        _xe_kp = _xe_row.get("kode_paket", "")
                        _xe_nama = _pl_label(_xe_row)
                        _excel_existing_status.update(
                            label=f"[{_xe_i+1}/{len(_pl_excel_existing_rows)}] {_xe_nama}"
                        )
                        _excel_existing_bp.progress((_xe_i + 1) / len(_pl_excel_existing_rows))
                        try:
                            _xe_fr = _keng_excel_existing.resolve_folder_paket_pl(_xe_kp)
                            _xe_root = _xe_fr.get("pesan", "") if _xe_fr.get("ok") else ""
                            _xe_xlsm = _cari_xlsm_pl(_xe_root) if _xe_root and _pl_os.path.isdir(_xe_root) else None
                            if not _xe_xlsm:
                                raise ValueError("folder atau workbook .xlsm tidak ditemukan")
                            _xe_res = _imd_excel_existing.isi_master_data_pl(_xe_kp, _xe_xlsm)
                            if not _xe_res.get("ok"):
                                raise ValueError(_xe_res.get("pesan", "macro gagal"))
                            _excel_existing_ok += 1
                            _excel_existing_line.write(
                                f"✅ [{_xe_i+1}] {_xe_nama} → {_xe_res.get('pesan', 'selesai')}"
                            )
                        except Exception as _xe_e:
                            _excel_existing_fail += 1
                            _excel_existing_failures.append(f"{_xe_nama}: {_xe_e}")
                            _excel_existing_line.write(f"❌ [{_xe_i+1}] {_xe_nama} → {_xe_e}")
                    _excel_existing_bp.progress(1.0)
                    _excel_existing_line.empty()
                    _excel_existing_status.update(
                        label=f"✅ Isi ulang Excel selesai: {_excel_existing_ok} sukses, {_excel_existing_fail} gagal",
                        state="complete",
                        expanded=_excel_existing_fail > 0,
                    )
                    if _excel_existing_failures:
                        st.warning("Paket gagal:\n" + "\n".join(_excel_existing_failures))

                if _cb_reset_existing:
                    from config import sb as _sb_reset_bulk
                    _reset_rows_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    try:
                        _sb_reset_bulk().table("draft_paket_pl").update({"folder_dibuat": None}).in_(
                            "kode_paket", [r.get("kode_paket") for r in _reset_rows_bulk]
                        ).execute()
                        _load_draft_pl_cached.clear()
                        st.success(f"✅ {len(_reset_rows_bulk)} status folder berhasil direset.")
                        st.rerun()
                    except Exception as _reset_bulk_e:
                        st.error(f"Reset bulk gagal: {_reset_bulk_e}")

                if _cb_pra_reviu_bulk:
                    import ai_evaluator as _heval_bulk
                    _pra_rows_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    _pra_jobs_bulk = [
                        {"nomor_urut": r.get("nomor_urut"), "nama_paket": r.get("nama_paket", "-")}
                        for r in _pra_rows_bulk
                    ]
                    if _pra_jobs_bulk:
                        _pra_status_bulk = st.status(
                            f"🤖 Pra-reviu {len(_pra_jobs_bulk)} paket...", expanded=True
                        )
                        _pra_line_bulk = _pra_status_bulk.empty()
                        def _pra_progress_bulk(_result, _done, _total, _message):
                            _pra_line_bulk.write(f"[{_done}/{_total}] {_message}")
                        _pra_results_bulk = _heval_bulk.evaluasi_bulk(
                            _pra_jobs_bulk, jenis="pra_reviu", model=None, max_workers=1,
                            engine="codex", progress_cb=_pra_progress_bulk,
                        )
                        _pra_ok_bulk = sum(1 for r in _pra_results_bulk if r.get("status") == "ok")
                        _pra_status_bulk.update(
                            label=f"✅ Pra-reviu selesai: {_pra_ok_bulk} sukses, {len(_pra_results_bulk)-_pra_ok_bulk} gagal",
                            state="complete", expanded=_pra_ok_bulk != len(_pra_results_bulk),
                        )

    # ── Tab 2: Kirim Undangan DPP ─────────────────────────────────────────────
    if _pl_active_tab == "2️⃣ Kirim Undangan DPP":
        # Satu kanvas penuh: daftar paket juga menjadi tempat upload BA Reviu.
        _kd_col_list = st.container()

        with _kd_col_list:
            st.markdown("### 1. Pilih Paket")

            _pl_rows_kd = _load_draft_pl_cached()
            _pl_rows_kd, _ = pl_engine.buang_duplikat_paket_lama(_pl_rows_kd)
            _pl_rows_kd = [r for r in _pl_rows_kd if not pl_engine.is_paket_selesai(r)]
            _kd_selected = []
            if not _pl_rows_kd:
                st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
            else:
                _kd_sel_col1, _kd_sel_col2 = st.columns(2)
                with _kd_sel_col1:
                    if st.button("✅ Semua", key="kd_sel_all", use_container_width=True):
                        for _rr in _pl_rows_kd:
                            st.session_state[f"kd_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _kd_sel_col2:
                    if st.button("⬜ Kosong", key="kd_sel_none", use_container_width=True):
                        for _rr in _pl_rows_kd:
                            st.session_state[f"kd_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _kd_selected = []
                import upload_ba_reviu_pl as _ubrpl
                def _do_upload_ba_pl(paket_list, tgl=None):
                    hasil = []
                    prog = st.progress(0, text="Memulai upload...")
                    for _i, _p in enumerate(paket_list):
                        prog.progress((_i + 1) / len(paket_list), text=f"Upload {_p['kode_paket']} ({_i+1}/{len(paket_list)})...")
                        _tgl_ba = _p.get("_tgl_acara") or tgl or datetime.now().date()
                        _res = _ubrpl.upload_ba_reviu_pl(
                            kode_paket=_p["kode_paket"], file_bytes=_p["_ba_file"].getvalue(),
                            file_name=_p["_ba_file"].name, tgl_ba=_tgl_ba.strftime("%d-%m-%Y"),
                        )
                        hasil.append({"kode": _p["kode_paket"], "nama": _p["nama_paket"], "sukses": _res["ok"], "pesan": f"HTTP {_res.get('status','?')}" if _res["ok"] else _res.get("error", "?")})
                    prog.empty()
                    _ok = sum(1 for h in hasil if h["sukses"])
                    st.success(f"✅ {_ok} BA Reviu berhasil diupload!") if _ok == len(hasil) else st.warning(f"⚠️ {_ok} berhasil, {len(hasil)-_ok} gagal.")
                    st.dataframe(hasil, use_container_width=True, hide_index=True)

                for _rr in _pl_rows_kd:
                    _kd_key     = f"kd_chk_{_rr['kode_paket']}_v19"
                    _kd_tgl_key = f"kd_tgl_acara_{_rr['kode_paket']}"
                    _kd_display = _pl_label(_rr)
                    _col_chk, _col_tgl, _col_ba = st.columns([3, 2, 3])
                    with _col_chk:
                        _kd_chk = st.checkbox(
                            "",
                            value=st.session_state.get(_kd_key, True),
                            key=_kd_key,
                            label_visibility="collapsed",
                        )
                        st.markdown(f"**{_kd_display}**")
                    with _col_tgl:
                        st.caption("Tanggal Undangan / BA Reviu")
                        _kd_tgl_acara = st.date_input(
                            "Tanggal Acara",
                            value=st.session_state.get(_kd_tgl_key)
                            or _last_pl_invitation_date(_rr["kode_paket"])
                            or datetime.now().date(),
                            format="DD/MM/YYYY",
                            key=_kd_tgl_key,
                            label_visibility="collapsed",
                        )
                        st.caption(f"{_HARI_NAMA[_kd_tgl_acara.weekday()]}, {_kd_tgl_acara.day} {_BULAN_NAMA[_kd_tgl_acara.month-1]} {_kd_tgl_acara.year}")
                        if _kd_tgl_acara in _LIBUR_MAP:
                            st.caption(f"⚠️ {_LIBUR_MAP[_kd_tgl_acara]}")
                    with _col_ba:
                        _ba_fkey = f"plba_file_{_rr['kode_paket']}"
                        _ba_up = st.file_uploader("BA Reviu PDF (opsional)", type=["pdf"], key=_ba_fkey)
                    if _kd_chk:
                        _kd_selected.append({**_rr, "_tgl_acara": _kd_tgl_acara})

                st.caption(f"**{len(_kd_selected)}** dari **{len(_pl_rows_kd)}** paket dipilih")

            st.divider()
            st.markdown("### 2. Detail Undangan")
            st.caption("Pesan dikirim PP ke PPK — meminta reviu Dokumen Persiapan Pengadaan.")

            st.markdown("**Waktu Acara (berlaku semua paket)**")
            _kd_col_mulai, _kd_col_selesai = st.columns(2)
            with _kd_col_mulai:
                _kd_jam_mulai = st.time_input(
                    "Mulai",
                    value=datetime.strptime("09:00", "%H:%M").time(),
                    key="kd_jam_mulai",
                    step=1800,
                )
            with _kd_col_selesai:
                _kd_jam_selesai = st.time_input(
                    "Selesai",
                    value=datetime.strptime("11:00", "%H:%M").time(),
                    key="kd_jam_selesai",
                    step=1800,
                )

            with st.expander("ℹ️ Libur Nasional Tersisa"):
                _kd_hari_ini = datetime.now().date()
                for _kd_d in sorted(d for d in _LIBUR_MAP if d >= _kd_hari_ini):
                    st.write(f"• {_HARI_NAMA[_kd_d.weekday()]}, {_kd_d.day} {_BULAN_NAMA[_kd_d.month-1]} {_kd_d.year} — {_LIBUR_MAP[_kd_d]}")

            _kd_tempat = pl_kirimpesan_engine.DEFAULT_TEMPAT
            st.caption(f"📍 Tempat: {_kd_tempat}")

            st.divider()
            st.warning("⚠️ Pesan yang terkirim **tidak bisa dihapus** dari SPSE.")
            if not st.session_state.get("kd_konfirmasi"):
                if st.button(
                    f"📨 Kirim Undangan DPP ke {len(_kd_selected)} Paket",
                    key="kd_kirim",
                    type="primary",
                    disabled=len(_kd_selected) == 0,
                    use_container_width=True,
                ):
                    if not _kd_tempat.strip():
                        st.error("❌ Tempat wajib diisi.")
                    else:
                        st.session_state["kd_konfirmasi"] = True
                        st.rerun()
            else:
                _kd_konfirm_lines = "\n".join(
                    f"{_pl_label(p)}  \n"
                    f"   📅 {_HARI_NAMA[p['_tgl_acara'].weekday()]}, {p['_tgl_acara'].day} {_BULAN_NAMA[p['_tgl_acara'].month-1]} {p['_tgl_acara'].year}"
                    for i, p in enumerate(_kd_selected)
                )
                st.warning(
                    f"Kirim ke **{len(_kd_selected)} paket**\n\n"
                    f"{_kd_konfirm_lines}\n\n"
                    f"- Pukul: {_kd_jam_mulai.strftime('%H.%M')} s.d. {_kd_jam_selesai.strftime('%H.%M')} Wita\n"
                    f"- Tempat: {_kd_tempat.strip()[:80]}\n\n"
                    f"**Tidak bisa dibatalkan setelah dikirim.**"
                )
                _kdc1, _kdc2 = st.columns(2)
                with _kdc1:
                    if st.button("✅ Ya, Kirim", key="kd_ya", type="primary", use_container_width=True):
                        st.session_state["kd_konfirmasi"] = False
                        _kd_progress = st.progress(0, text="Memulai pengiriman...")
                        _kd_hasil = []
                        _tgl_kirim_kd = datetime.now().date()

                        for _ki, _kp in enumerate(_kd_selected):
                            _kd_progress.progress(
                                (_ki + 1) / len(_kd_selected),
                                text=f"Mengirim {_ki+1}/{len(_kd_selected)}...",
                            )
                            _kd_tgl_a  = _kp["_tgl_acara"]
                            _kd_hari_tgl = f"{_HARI_NAMA[_kd_tgl_a.weekday()]}, {_kd_tgl_a.day} {_BULAN_NAMA[_kd_tgl_a.month-1]} {_kd_tgl_a.year}"
                            _kd_pukul    = f"{_kd_jam_mulai.strftime('%H.%M')} s.d. {_kd_jam_selesai.strftime('%H.%M')} Wita"

                            # Generate PDF lampiran otomatis
                            import undangan_pdf_engine as _upe
                            _gen = _upe.generate_undangan_pdf_pl(
                                kode_paket=_kp["kode_paket"],
                                tanggal_kirim=_tgl_kirim_kd,
                                hari_tgl_rapat=_kd_hari_tgl,
                                pukul_rapat=_kd_pukul,
                                tempat_rapat=_kd_tempat.strip(),
                            )
                            _lamp_bytes = _gen["pdf_bytes"] if _gen["sukses"] else None
                            _ku_lamp = _kp.get("kode_unik") or _kp["kode_paket"]
                            _lamp_nama  = f"undangan_reviu_{_ku_lamp}.pdf"
                            if _lamp_bytes:
                                st.session_state.setdefault("_kd_pdf_cache", {})[_ku_lamp] = (_lamp_nama, _lamp_bytes)

                            _waktu_str  = datetime.combine(_kd_tgl_a, _kd_jam_mulai).strftime("%d-%m-%Y %H:%M")
                            _sampai_str = datetime.combine(_kd_tgl_a, _kd_jam_selesai).strftime("%d-%m-%Y %H:%M")

                            _res = pl_kirimpesan_engine.kirim_undangan_pl(
                                kode=_kp["kode_paket"],
                                waktu=_waktu_str,
                                sampai=_sampai_str,
                                tempat=_kd_tempat.strip(),
                                dibawa=pl_kirimpesan_engine.DEFAULT_DIBAWA,
                                hadir=pl_kirimpesan_engine.DEFAULT_HADIR,
                                lampiran_bytes=_lamp_bytes,
                                lampiran_nama=_lamp_nama,
                            )
                            _kd_hasil.append({
                                "Paket": _pl_label(_kp),
                                "Penerima (PPK)": _res.get("penerima", "-"),
                                "PDF": "✅" if _gen["sukses"] else f"❌ {_gen['pesan']}",
                                "Kirim": "✅" if _res["sukses"] else f"❌ {_res['pesan']}",
                            })
                            if _res.get("sukses"):
                                _save_last_pl_invitation_date(_kp["kode_paket"], _kd_tgl_a)

                        _kd_progress.empty()
                        _kd_ok = sum(1 for h in _kd_hasil if h["Kirim"] == "✅")
                        if _kd_ok == len(_kd_hasil):
                            st.success(f"✅ Semua {_kd_ok} undangan berhasil dikirim!")
                        else:
                            st.warning(f"⚠️ {_kd_ok} berhasil, {len(_kd_hasil)-_kd_ok} gagal.")
                        st.dataframe(
                            _kd_hasil,
                            use_container_width=True,
                            column_config={
                                "Paket":          st.column_config.TextColumn("Paket", width="large"),
                                "Penerima (PPK)": st.column_config.TextColumn("Penerima (PPK)"),
                                "PDF":            st.column_config.TextColumn("PDF", width="small"),
                                "Kirim":          st.column_config.TextColumn("Kirim", width="small"),
                            },
                            hide_index=True,
                        )
                        # Tombol download per PDF
                        for _ku_dl, (_nm_dl, _by_dl) in st.session_state.get("_kd_pdf_cache", {}).items():
                            st.download_button(
                                f"⬇️ Download {_nm_dl}",
                                data=_by_dl,
                                file_name=_nm_dl,
                                mime="application/pdf",
                                key=f"kd_dl_{_ku_dl}",
                            )

                with _kdc2:
                    if st.button("❌ Batal", key="kd_batal", use_container_width=True):
                        st.session_state["kd_konfirmasi"] = False
                        st.rerun()

            # BA memakai tanggal acara masing-masing paket; unggah setelah
            # memilih paket dan/atau mengirim undangan.
            _ba_pl_valid = [
                {**_pp, "_ba_file": st.session_state.get(f"plba_file_{_pp['kode_paket']}")}
                for _pp in _kd_selected
                if st.session_state.get(f"plba_file_{_pp['kode_paket']}")
            ]
            if st.button(
                f"📄 Upload BA Reviu DPP ({len(_ba_pl_valid)} file)",
                key="plba_upload_selected",
                type="secondary",
                disabled=not _ba_pl_valid,
                use_container_width=True,
            ):
                _do_upload_ba_pl(_ba_pl_valid)

    # ── Tab 4: Buat Jadwal PL (5 tahap, push langsung ke SPSE) ─────────────
    if _pl_active_tab == "5️⃣ Buat Jadwal":
        st.markdown("### Buat Jadwal Pengadaan Langsung")
        st.caption("5 tahap PL: Upload Penawaran → Pembukaan → Evaluasi → Klarifikasi+Nego → Tanda Tangan Kontrak. Push langsung ke SPSE.")

        import jadwal_engine_pl as _jepl
        _libur_map_pl = _LIBUR_MAP

        _pljd_rows = _load_draft_pl_cached()
        _pljd_rows, _ = pl_engine.buang_duplikat_paket_lama(_pljd_rows)
        _pljd_rows = [r for r in _pljd_rows if not pl_engine.is_paket_selesai(r)]
        if not _pljd_rows:
            st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
        else:
            _pljd_col_list, _pljd_col_detail = st.columns([3, 2])

            with _pljd_col_list:
                st.markdown("### 1. Pilih Paket")
                _pljd_a, _pljd_b = st.columns(2)
                with _pljd_a:
                    if st.button("✅ Semua", key="pljd_sel_all", use_container_width=True):
                        for _rr in _pljd_rows:
                            st.session_state[f"pljd_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _pljd_b:
                    if st.button("⬜ Kosong", key="pljd_sel_none", use_container_width=True):
                        for _rr in _pljd_rows:
                            st.session_state[f"pljd_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _pljd_selected = []
                for _rr in _pljd_rows:
                    _key = f"pljd_chk_{_rr['kode_paket']}_v19"
                    _chk = st.checkbox(
                        "",
                        value=st.session_state.get(_key, True),
                        key=_key,
                        label_visibility="collapsed",
                    )
                    st.markdown(f"**{_pl_label(_rr)}** ({_rr.get('jenis_pl','?')})")
                    if _chk:
                        _pljd_selected.append(_rr)

                st.caption(f"**{len(_pljd_selected)}** dari **{len(_pljd_rows)}** paket dipilih")

            with _pljd_col_detail:
                st.markdown("### 2. Tanggal Mulai (T1)")
                _pljd_beda = st.checkbox("Jadwal berbeda per paket", value=False, key="pljd_beda")

                if not _pljd_beda:
                    # Default: baca tgl_batas_penawaran (T1.selesai) dari DB, kurangi 5 hari
                    _pljd_tgl_default = datetime.now().date()
                    if _pljd_selected:
                        _tbp = _pljd_selected[0].get("tgl_batas_penawaran")
                        if _tbp:
                            try:
                                from datetime import date as _date2, timedelta as _td2
                                _pljd_tgl_default = _date2.fromisoformat(str(_tbp)[:10]) - _td2(days=5)
                            except Exception:
                                pass
                    _c1, _c2 = st.columns(2)
                    with _c1:
                        _pljd_tgl_global = st.date_input(
                            "Tanggal",
                            value=_pljd_tgl_default,
                            format="DD/MM/YYYY",
                            key="pljd_tgl_global",
                        )
                        st.markdown(f"**{_HARI_NAMA[_pljd_tgl_global.weekday()]}, {_pljd_tgl_global.day} {_BULAN_NAMA[_pljd_tgl_global.month-1]} {_pljd_tgl_global.year}**")
                    with _c2:
                        _pljd_jam_global = st.time_input(
                            "Jam",
                            value=datetime.strptime("08:00", "%H:%M").time(),
                            key="pljd_jam_global",
                        )
                    if _pljd_tgl_global in _libur_map_pl:
                        st.warning(f"⚠️ **{_libur_map_pl[_pljd_tgl_global]}**")
                else:
                    if not _pljd_selected:
                        st.info("Pilih paket dulu.")
                    else:
                        for _p in _pljd_selected:
                            _ktgl = f"pljd_tgl_{_p['kode_paket']}"
                            _kjam = f"pljd_jam_{_p['kode_paket']}"
                            _cna, _cdt, _cjm = st.columns([3, 2, 1])
                            with _cna:
                                st.markdown(f"**{_pl_label(_p)}**")
                            with _cdt:
                                st.date_input(
                                    "Tgl",
                                    value=st.session_state.get(_ktgl, datetime.now().date()),
                                    format="DD/MM/YYYY",
                                    key=_ktgl,
                                    label_visibility="collapsed",
                                )
                            with _cjm:
                                st.time_input(
                                    "Jam",
                                    value=st.session_state.get(_kjam, datetime.strptime("08:00", "%H:%M").time()),
                                    key=_kjam,
                                    label_visibility="collapsed",
                                )

                import gcal_pl_helper as _gcalpl_g
                _gcal_ok = _gcalpl_g.check_gcal_token()

                st.divider()
                _pljd_mode = st.radio("Mode Jadwal", ["Normal", "Standar", "Cepat"], horizontal=True, key="pljd_mode")

                if _pljd_selected:
                    st.markdown(f"**📅 Preview Jadwal — cek dulu sebelum push**")
                    _pk_prev = st.selectbox(
                        "Pilih paket buat preview",
                        _pljd_selected,
                        format_func=_pl_label,
                        key="pljd_preview_pilih",
                    )
                    _tgl_pv = st.session_state.get(f"pljd_tgl_{_pk_prev['kode_paket']}", datetime.now().date()) if _pljd_beda else _pljd_tgl_global
                    _jam_pv = st.session_state.get(f"pljd_jam_{_pk_prev['kode_paket']}", datetime.strptime("08:00", "%H:%M").time()) if _pljd_beda else _pljd_jam_global
                    _t1_pv = datetime.combine(_tgl_pv, _jam_pv)
                    if _pljd_mode == "Cepat":
                        _jadwal_pv = _jepl.hitung_jadwal_pl_cepat(_t1_pv)
                    elif _pljd_mode == "Standar":
                        _jadwal_pv = _jepl.hitung_jadwal_pl_standar(_t1_pv)
                    else:
                        _jadwal_pv = _jepl.hitung_jadwal_pl(_t1_pv)
                    for _jd in _jadwal_pv:
                        _m, _s = _jd["mulai"], _jd["selesai"]
                        _sama_hari = _m.date() == _s.date()
                        _txt_mulai = f"{_HARI_NAMA[_m.weekday()]}, {_m.day} {_BULAN_NAMA[_m.month-1]} {_m.strftime('%H:%M')}"
                        _txt_selesai = _s.strftime('%H:%M') if _sama_hari else f"{_HARI_NAMA[_s.weekday()]}, {_s.day} {_BULAN_NAMA[_s.month-1]} {_s.strftime('%H:%M')}"
                        with st.container(border=True):
                            st.write(f"**{_jd['nama']}**  \n{_txt_mulai} → {_txt_selesai}")
                    st.divider()

                st.caption("⚠️ Akan menimpa jadwal yang sudah ada di SPSE.")

                _pljd_submit = st.button(
                    f"🚀 Push Jadwal ke SPSE ({len(_pljd_selected)} paket)",
                    type="primary",
                    use_container_width=True,
                    disabled=len(_pljd_selected) == 0,
                    key="pljd_submit_btn",
                )

                if _pljd_submit:
                    _hasil = []
                    _prog = st.progress(0, text="Mulai...")
                    for _i, _p in enumerate(_pljd_selected):
                        _prog.progress((_i + 1) / len(_pljd_selected),
                                       text=f"{_p['kode_paket']} ({_i+1}/{len(_pljd_selected)})...")
                        if _pljd_beda:
                            _tgl = st.session_state.get(f"pljd_tgl_{_p['kode_paket']}", datetime.now().date())
                            _jam = st.session_state.get(f"pljd_jam_{_p['kode_paket']}", datetime.strptime("08:00", "%H:%M").time())
                        else:
                            _tgl = _pljd_tgl_global
                            _jam = _pljd_jam_global
                        _t1 = datetime.combine(_tgl, _jam)

                        _kp = _p.get("kode_paket")
                        if not _kp:
                            _hasil.append({"paket": _pl_label(_p), "ok": False, "pesan": "kode_paket kosong"})
                            continue
                        try:
                            _mode_str = {"Cepat": "cepat", "Standar": "standar"}.get(_pljd_mode, "normal")
                            _r = _jepl.submit_full_pl(_kp, _t1, mode=_mode_str)
                            _sub = _r["submit_result"]
                            _hasil.append({
                                "paket":  _pl_label(_p),
                                "ok":     _sub["ok"],
                                "pesan":  f"HTTP {_sub['status']}",
                                "mulai":  _t1.strftime("%d/%m/%Y %H:%M"),
                            })
                            # Simpan tgl ke Supabase + push GCal
                            if _sub["ok"]:
                                try:
                                    _jad = _r["jadwal_list"]
                                    pl_engine.simpan_paket_pl({
                                        "kode_paket":            _p["kode_paket"],
                                        "tgl_batas_penawaran":   _jad[0]["selesai"].strftime("%Y-%m-%d"),
                                        "tgl_buka_penawaran":    _jad[1]["mulai"].strftime("%Y-%m-%d"),
                                        "tgl_evaluasi":          _jad[2]["selesai"].strftime("%Y-%m-%d"),
                                        "tgl_negosiasi":         _jad[3]["mulai"].strftime("%Y-%m-%d"),
                                        "tgl_penetapan":         _jad[4]["mulai"].strftime("%Y-%m-%d"),
                                    })
                                except Exception:
                                    pass
                                try:
                                    import gcal_pl_helper as _gcalpl
                                    _gcalpl.push_jadwal_pl_ke_gcal(_kp, _p["nama_paket"], _r["jadwal_list"])
                                except Exception:
                                    pass
                        except Exception as _e:
                            _hasil.append({"paket": _pl_label(_p), "ok": False, "pesan": str(_e)[:100]})

                    _prog.empty()
                    _sukses = sum(1 for h in _hasil if h["ok"])
                    _gagal = len(_hasil) - _sukses
                    if _gagal == 0:
                        st.success(f"✅ Semua {_sukses} paket berhasil dijadwalkan!")
                    else:
                        st.warning(f"⚠️ {_sukses} sukses, {_gagal} gagal")
                    for h in _hasil:
                        _ic = "✅" if h["ok"] else "❌"
                        st.markdown(f"{_ic} **{h['paket']}** — {h['pesan']}" + (f" — mulai {h.get('mulai','')}" if h["ok"] else ""))

                    # Expander preview jadwal per paket sukses
                    _hasil_sukses = [h for h in _hasil if h["ok"]]
                    if _hasil_sukses:
                        with st.expander(f"📅 Lihat Detail Jadwal ({len(_hasil_sukses)} paket)", expanded=True):
                            for _ph in _hasil_sukses:
                                _pk_match = next((p for p in _pljd_selected if p["nama_paket"] == _ph["paket"]), None)
                                if not _pk_match:
                                    continue
                                _tgl_preview = st.session_state.get(f"pljd_tgl_{_pk_match['kode_paket']}", _pljd_tgl_global if not _pljd_beda else datetime.now().date())
                                _jam_preview = st.session_state.get(f"pljd_jam_{_pk_match['kode_paket']}", _pljd_jam_global if not _pljd_beda else datetime.strptime("08:00", "%H:%M").time())
                                _t1_preview = datetime.combine(_tgl_preview, _jam_preview)
                                _mode_str_prev = {"Cepat": "cepat", "Standar": "standar"}.get(_pljd_mode, "normal")
                                if _mode_str_prev == "cepat":
                                    _jadwal_preview = _jepl.hitung_jadwal_pl_cepat(_t1_preview)
                                elif _mode_str_prev == "standar":
                                    _jadwal_preview = _jepl.hitung_jadwal_pl_standar(_t1_preview)
                                else:
                                    _jadwal_preview = _jepl.hitung_jadwal_pl(_t1_preview)
                                st.markdown(f"**{_pl_label(_pk_match)}**")
                                import pandas as _pd_jad
                                _jad_rows = []
                                for _idx_jad, _jd in enumerate(_jadwal_preview, 1):
                                    _dur = _jd["selesai"] - _jd["mulai"]
                                    _dur_str = ""
                                    _dur_days = _dur.days
                                    _dur_hours = _dur.seconds // 3600
                                    _dur_mins = (_dur.seconds % 3600) // 60
                                    if _dur_days > 0:
                                        _dur_str = f"{_dur_days} hari"
                                        if _dur_hours > 0:
                                            _dur_str += f" {_dur_hours} jam"
                                    elif _dur_hours > 0:
                                        _dur_str = f"{_dur_hours} jam {_dur_mins} menit"
                                    else:
                                        _dur_str = f"{_dur_mins} menit"
                                    _jad_rows.append({
                                        "No": _idx_jad,
                                        "Tahap": _jd["nama"],
                                        "Mulai": _jd["mulai"].strftime("%d-%m-%Y %H:%M"),
                                        "Selesai": _jd["selesai"].strftime("%d-%m-%Y %H:%M"),
                                        "Durasi": _dur_str,
                                    })
                                st.dataframe(_pd_jad.DataFrame(_jad_rows), use_container_width=True, hide_index=True)

                with st.expander("ℹ️ Libur Nasional Tersisa"):
                    _hari_ini = datetime.now().date()
                    _sisa = sorted(d for d in _libur_map_pl if d >= _hari_ini)
                    for d in _sisa[:15]:
                        st.write(f"• {_HARI_NAMA[d.weekday()]}, {d.day} {_BULAN_NAMA[d.month-1]} {d.year} — {_libur_map_pl[d]}")

        st.divider()
        st.markdown("#### 🔄 Sync Jadwal ke Google Calendar")
        st.caption("Baca jadwal aktual dari SPSE → update GCal + Supabase tgl_evaluasi/tgl_negosiasi/tgl_penetapan. Jalankan setelah ada perubahan jadwal di SPSE.")
        import gcal_pl_helper as _gcalpl_tc
        _sync_gcal_pl_btn = False
        if not _gcalpl_tc.check_gcal_token():
            st.warning("🔐 Token Google Calendar tidak valid atau expired.")
            if st.button("🔑 Login Ulang ke Google Calendar", key="reauth_gcal_btn_jkk", type="primary", use_container_width=True):
                import gcal_helper as _gcalh_ra
                try:
                    _gcalh_ra.generate_token()
                    st.success("✅ Token diperbarui! Klik Sync untuk melanjutkan.")
                    st.rerun()
                except Exception as _e_ra:
                    st.error(f"❌ Gagal reauth: {_e_ra}")
        else:
            _sync_gcal_pl_btn = st.button("🔄 Sync Jadwal ke GCal", key="sync_gcal_pl_btn_jkk", use_container_width=True, type="primary")
        if _sync_gcal_pl_btn:
            import gcal_pl_helper as _gcalpl
            _gcalpl_prog = st.progress(0.0, text="Memulai sync...")
            _gcalpl_results = _gcalpl.sync_semua_paket_pl(
                progress_cb=lambda f, m: _gcalpl_prog.progress(f, text=m)
            )
            _gcalpl_prog.empty()
            _gcalpl_ok = sum(1 for r in _gcalpl_results if r["ok"])
            _gcalpl_skip = sum(1 for r in _gcalpl_results if not r["ok"] and "kosong" in r.get("error", ""))
            _gcalpl_err = len(_gcalpl_results) - _gcalpl_ok - _gcalpl_skip
            if _gcalpl_err == 0:
                st.success(f"✅ {_gcalpl_ok} paket sync OK, {_gcalpl_skip} skip (jadwal belum diisi SPSE).")
            else:
                st.warning(f"⚠️ {_gcalpl_ok} OK, {_gcalpl_skip} skip, {_gcalpl_err} error.")
            _gcalpl_display = [
                {
                    "Paket": _pl_label(r),
                    "Status": "✅" if r["ok"] else ("⏭ Skip" if "kosong" in r.get("error","") else "❌"),
                    "GCal +": r["gcal_inserted"],
                    "GCal -": r["gcal_deleted"],
                    "Tgl Evaluasi": r["tgl_evaluasi"],
                    "Tgl Negosiasi": r["tgl_negosiasi"],
                    "Tgl Penetapan": r["tgl_penetapan"],
                    "Error": r["error"][:60] if r["error"] else "",
                }
                for r in _gcalpl_results
            ]
            st.dataframe(_gcalpl_display, use_container_width=True, hide_index=True)
        st.divider()
        _render_ubah_jadwal_pl(_pljd_rows, _jepl, "pljd_edit_jkk")

    # ── Tab 3: Setup Paket PL (LDK + Masa Berlaku + Checklist + Upload Dokpil) ─
    if _pl_active_tab == "3️⃣ Setup Paket":
        st.markdown("### Setup Paket Pengadaan Langsung")
        st.caption(
            "Submit LDK (Persyaratan Kualifikasi) + Masa Berlaku Penawaran + "
            "Checklist Dokumen Penawaran + Upload Dokumen Pemilihan (Dokpil PDF) ke SPSE. "
            "KAK / Rancangan Kontrak / Uraian Singkat / Informasi Lainnya tugas PPK (bukan PP)."
        )

        _depl = _depl_jkk  # alias JKK — sudah di-import top-level

        _plsp_rows = _load_draft_pl_cached()
        _plsp_rows, _ = pl_engine.buang_duplikat_paket_lama(_plsp_rows)
        _plsp_rows = [r for r in _plsp_rows if not pl_engine.is_paket_selesai(r)]
        if not _plsp_rows:
            st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
        else:
            _plsp_col_list, _plsp_col_kanan = st.columns([2, 3])

            with _plsp_col_list:
                st.markdown("### 1. Pilih Paket + Upload Dokpil")
                _plsp_sel_all, _plsp_sel_none = st.columns(2)
                with _plsp_sel_all:
                    if st.button("✅ Semua", key="plsp_sel_all", use_container_width=True):
                        for _rr in _plsp_rows:
                            st.session_state[f"plsp_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _plsp_sel_none:
                    if st.button("⬜ Kosong", key="plsp_sel_none", use_container_width=True):
                        for _rr in _plsp_rows:
                            st.session_state[f"plsp_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _sp = _sp_global  # alias — sudah di-import top-level

                _plsp_selected = []
                for _rr in _plsp_rows:
                    _kp_key = _rr["kode_paket"]
                    _plsp_chk_key  = f"plsp_chk_{_kp_key}_v19"
                    _plsp_file_key = f"plsp_dokpil_{_kp_key}"

                    _col_chk, _col_file = st.columns([3, 2])
                    with _col_chk:
                        if _plsp_chk_key not in st.session_state:
                            st.session_state[_plsp_chk_key] = True
                        _chk = st.checkbox("", key=_plsp_chk_key, label_visibility="collapsed")
                        st.markdown(f"**{_pl_label(_rr)}** ({_rr.get('jenis_pl','?')})")
                    with _col_file: # JKK tab3 marker
                        _dokpil_up = st.file_uploader(
                            "Dokpil PDF",
                            type=["pdf"],
                            key=_plsp_file_key,
                            label_visibility="collapsed",
                        )
                        if _dokpil_up:
                            _ku_prev = _rr.get("kode_unik") or "?"
                            _sk_prev = _lookup_singkatan_dinas(_rr.get("satker", ""))
                            # Tanggal: dari DB (tgl_dokpil) → fallback session → hari ini
                            _tgl_db = _rr.get("tgl_dokpil")
                            if _tgl_db:
                                try:
                                    from datetime import date as _date
                                    _tgl_prev = _date.fromisoformat(str(_tgl_db))
                                except Exception:
                                    _tgl_prev = datetime.now().date()
                            else:
                                _tgl_prev = datetime.now().date()
                            # Nomor: dari DB → fallback generate
                            _no_prev = _rr.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                nama_paket=_rr["nama_paket"],
                                kode_unik=_ku_prev,
                                skpd_singkat=_sk_prev,
                                tahun=_tgl_prev.year,
                                paket_ulang=_pl_paket_ulang(_rr),
                                nomor_urut=_rr.get("nomor_urut"),
                            )
                            st.caption(f"📄 {_dokpil_up.name}  \n📋 `{_no_prev}`  \n📅 {_tgl_prev.strftime('%d-%m-%Y')}")
                            if st.button("📤 Upload Dokpil", key=f"plsp_upload_only_{_kp_key}", use_container_width=True):
                                with st.spinner("Mengupload dokpil..."):
                                    try:
                                        _r_up_only = _udpl.upload_dokpil_pl(
                                            kode_paket=_kp_key,
                                            file_bytes=_dokpil_up.getvalue(),
                                            file_name=_dokpil_up.name,
                                            nomor_dokpil=_no_prev,
                                            tgl_dokpil=_tgl_prev.strftime("%d-%m-%Y"),
                                        )
                                        if _r_up_only["ok"]:
                                            from config import sb as _sb_up_only
                                            _sb_up_only().table("draft_paket_pl").update({
                                                "nomor_dokpil": _no_prev,
                                            }).eq("kode_paket", _kp_key).execute()
                                            st.success(f"✅ Upload berhasil — {_no_prev}")
                                        else:
                                            st.error(f"❌ HTTP {_r_up_only.get('status','?')} — {_r_up_only.get('error') or _r_up_only.get('body','')[:300]}")
                                            st.json(_r_up_only)
                                    except Exception as _e_up_only:
                                        st.error(f"❌ Exception: {_e_up_only}")

                    if _chk:
                        _plsp_selected.append({
                            **_rr,
                            "_dokpil_file": _dokpil_up,
                        })

                st.caption(f"**{len(_plsp_selected)}** dari **{len(_plsp_rows)}** paket dipilih")

                # Kumpulkan semua paket yang sudah ada file dokpil
                _all_with_file = [
                    {**_rr, "_dokpil_file": st.session_state.get(f"plsp_dokpil_{_rr['kode_paket']}")}
                    for _rr in _plsp_rows
                    if st.session_state.get(f"plsp_dokpil_{_rr['kode_paket']}")
                ]
                if _all_with_file:
                    st.divider()
                    if st.button(f"📤 Upload Semua Dokpil ({len(_all_with_file)} file)", key="plsp_upload_all_dokpil", use_container_width=True, type="primary"):
                        from config import sb as _sb_upall
                        _cl_upall = _sb_upall()
                        for _rr_up in _all_with_file:
                            _kp_up = _rr_up["kode_paket"]
                            _f_up = _rr_up["_dokpil_file"]
                            _ku_up = _rr_up.get("kode_unik") or "?"
                            _sk_up = _lookup_singkatan_dinas(_rr_up.get("satker", ""))
                            # Tanggal dari DB, fallback session
                            _tgl_db_up = _rr_up.get("tgl_dokpil")
                            if _tgl_db_up:
                                try:
                                    from datetime import date as _date2
                                    _tgl_up = _date2.fromisoformat(str(_tgl_db_up))
                                except Exception:
                                    _tgl_up = datetime.now().date()
                            else:
                                _tgl_up = datetime.now().date()
                            # Nomor dari DB, fallback generate
                            _no_up = _rr_up.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                nama_paket=_rr_up["nama_paket"],
                                kode_unik=_ku_up,
                                skpd_singkat=_sk_up,
                                tahun=_tgl_up.year,
                                paket_ulang=_pl_paket_ulang(_rr_up),
                                nomor_urut=_rr_up.get("nomor_urut"),
                            )
                            try:
                                _r_upall = _udpl.upload_dokpil_pl(
                                    kode_paket=_kp_up,
                                    file_bytes=_f_up.getvalue(),
                                    file_name=_f_up.name,
                                    nomor_dokpil=_no_up,
                                    tgl_dokpil=_tgl_up.strftime("%d-%m-%Y"),
                                )
                                if _r_upall["ok"]:
                                    _cl_upall.table("draft_paket_pl").update({"nomor_dokpil": _no_up}).eq("kode_paket", _kp_up).execute()
                                    st.success(f"✅ {_pl_label(_rr_up)} — {_no_up}")
                                else:
                                    st.error(f"❌ {_pl_label(_rr_up)} HTTP {_r_upall.get('status','?')} — {_r_upall.get('body','')[:100]}")
                            except Exception as _e_upall:
                                st.error(f"❌ {_pl_label(_rr_up)}: {_e_upall}")
                        _load_draft_pl_cached.clear()

            with _plsp_col_kanan:
                st.markdown("### 2. Konfigurasi Setup Paket")

                if not _plsp_selected:
                    st.info("Pilih paket di sebelah kiri.")
                else:
                    # ── SEKSI 1: SBU Global ───────────────────────────────────
                    st.markdown("#### 🏗️ Seksi 1 — SBU Global")
                    _sbu_global_aktif = st.toggle(
                        "SBU Global (apply 1 SBU ke semua paket terpilih)",
                        value=st.session_state.get("plsp_sbu_global_aktif", False),
                        key="plsp_sbu_global_aktif",
                    )
                    if _sbu_global_aktif:
                        st.caption("Satu pilihan SBU apply ke semua paket terpilih.")

                        _plsp_klas_list = ["(auto-detect dari paket pertama)"] + _sp.list_klasifikasi()

                        _first_p = _plsp_selected[0]
                        _detected_g = _sp.detect_from_draft(
                            _first_p.get("sbu_baru") or "", _first_p.get("sbu_lama") or ""
                        )
                        _g_kode_baru = _detected_g.get("kode_baru", "")
                        _g_kode_lama = _detected_g.get("kode_lama", "")

                        _g_klas_default = 0
                        if _g_kode_baru:
                            _baru_info_g = _sp.get_sbu_baru_by_kode(_g_kode_baru)
                            _klas_det_g = (_baru_info_g or {}).get("klasifikasi", "")
                            if _klas_det_g in _plsp_klas_list:
                                _g_klas_default = _plsp_klas_list.index(_klas_det_g)

                        if _g_kode_baru:
                            st.caption(f"Auto-detect dari **{_pl_label(_first_p)}**: `{_g_kode_baru}` / `{_g_kode_lama}`")

                        _g_picked_klas = st.selectbox(
                            "Klasifikasi",
                            _plsp_klas_list,
                            index=_g_klas_default,
                            key="plsp_global_klas",
                        )

                        if _g_picked_klas and _g_picked_klas != "(auto-detect dari paket pertama)":
                            _g_baru_options = _sp.list_sbu_baru_by_klasifikasi(_g_picked_klas)
                        else:
                            _g_baru_options = []
                            if _g_kode_baru:
                                _g_baru_options = [_sp.get_sbu_baru_by_kode(_g_kode_baru)]
                        _g_baru_labels = [
                            f"{b['kode']} — {(b.get('nama_singkat') or b.get('nama_full',''))[:70]}"
                            for b in _g_baru_options if b
                        ]
                        _g_baru_default = 0
                        for _gi, _gb in enumerate(_g_baru_options):
                            if _gb and _gb.get("kode") == _g_kode_baru:
                                _g_baru_default = _gi
                                break
                        _g_picked_baru_label = st.selectbox(
                            "SBU Baru (KBLI 2020)",
                            _g_baru_labels or ["(pilih klasifikasi dulu)"],
                            index=_g_baru_default if _g_baru_labels else 0,
                            key="plsp_global_sbu_baru",
                        )
                        _g_picked_baru_kode = (
                            _g_picked_baru_label.split(" — ", 1)[0]
                            if _g_baru_labels and " — " in _g_picked_baru_label else ""
                        )

                        _g_lama_options = _sp.list_sbu_lama_padanan(_g_picked_baru_kode) if _g_picked_baru_kode else []
                        _g_lama_labels = ["(tidak dipersyaratkan / hanya SBU 2020)"] + [
                            f"{l['kode']} — {(l.get('nama_singkat') or l.get('nama_full',''))[:70]}"
                            for l in _g_lama_options
                        ]
                        _g_lama_default = 0
                        for _gli, _gl in enumerate(_g_lama_options):
                            if _gl.get("kode") == _g_kode_lama:
                                _g_lama_default = _gli + 1
                                break
                        _g_picked_lama_label = st.selectbox(
                            "SBU Lama (KBLI 2017) — opsional",
                            _g_lama_labels,
                            index=_g_lama_default,
                            key="plsp_global_sbu_lama",
                        )
                        _g_picked_lama_kode = (
                            _g_picked_lama_label.split(" — ", 1)[0]
                            if " — " in _g_picked_lama_label else ""
                        )

                        _sbu_baru_global = ""
                        _sbu_lama_global = ""
                        if _g_picked_baru_kode:
                            _baru_obj_g = _sp.get_sbu_baru_by_kode(_g_picked_baru_kode)
                            _sbu_baru_global = (_baru_obj_g or {}).get("nama_full", "")
                        if _g_picked_lama_kode:
                            _lama_obj_g = _sp.get_sbu_lama_by_kode(_g_picked_lama_kode)
                            _sbu_lama_global = (_lama_obj_g or {}).get("nama_full", "")
                        if not _sbu_baru_global:
                            _sbu_baru_global = _first_p.get("sbu_baru") or ""

                        if _sbu_baru_global:
                            st.caption(f"🔹 Baru: `{_sbu_baru_global[:80]}`")
                        if _sbu_lama_global:
                            st.caption(f"🔸 Lama: `{_sbu_lama_global[:80]}`")
                        elif _sbu_baru_global:
                            st.caption("ℹ️ SBU Lama tidak dipersyaratkan — hanya SBU 2020 di LDK")

                        if st.button(
                            f"💾 Simpan SBU Global ke {len(_plsp_selected)} paket",
                            key="plsp_save_sbu_btn", use_container_width=True,
                        ):
                            from config import sb as _sb_factory
                            _client_sbu = _sb_factory()
                            _ok_sbu = 0
                            for _p in _plsp_selected:
                                try:
                                    _client_sbu.table("draft_paket_pl").update({
                                        "sbu_baru": _sbu_baru_global,
                                        "sbu_lama": _sbu_lama_global,
                                    }).eq("kode_paket", _p["kode_paket"]).execute()
                                    _ok_sbu += 1
                                except Exception as _e:
                                    st.error(f"❌ {_pl_label(_p)}: {_e}")
                            st.success(f"✅ {_ok_sbu}/{len(_plsp_selected)} paket disimpan ke Supabase")
                    else:
                        st.caption("ℹ️ Mode custom — pilih SBU yang pernah dipakai atau input baru.")
                        _pl_sbu_hist = _load_pl_sbu_history()
                        _pl_sbu_new = "✏️ Input SBU baru"
                        _pl_sbu_opts = [f"{x['baru'][:100]}" for x in _pl_sbu_hist] + [_pl_sbu_new]
                        _pl_sbu_pick = st.selectbox("SBU Custom — histori", _pl_sbu_opts or [_pl_sbu_new], key="plsp_custom_sbu_pick")
                        _pl_sbu_idx = _pl_sbu_opts.index(_pl_sbu_pick) if _pl_sbu_pick in _pl_sbu_opts else len(_pl_sbu_hist)
                        if _pl_sbu_idx < len(_pl_sbu_hist):
                            _pl_sbu_selected = _pl_sbu_hist[_pl_sbu_idx]
                            _sbu_baru_global = _pl_sbu_selected["baru"]
                            _sbu_lama_global = _pl_sbu_selected["lama"] or None
                            st.caption(f"🔹 Baru: `{_sbu_baru_global}`")
                            if _sbu_lama_global:
                                st.caption(f"🔸 Lama: `{_sbu_lama_global}`")
                        else:
                            _sbu_baru_global = st.text_input(
                                "SBU Baru (teks bebas)", key="plsp_custom_sbu_baru",
                                placeholder="Contoh: RE201 — Jasa Desain Rekayasa untuk Konstruksi Pondasi serta Struktur Bangunan",
                            ) or None
                            _sbu_lama_global = st.text_input(
                                "SBU Lama (teks bebas, opsional)", key="plsp_custom_sbu_lama",
                                placeholder="Kosongkan jika tidak dipersyaratkan",
                            ) or None
                        if st.button(
                            f"💾 Simpan SBU Custom + POST LDK ke {len(_plsp_selected)} paket",
                            key="plsp_save_sbu_custom_btn", use_container_width=True,
                            disabled=not _sbu_baru_global,
                        ):
                            from config import sb as _sb_factory_c
                            _client_c = _sb_factory_c()
                            _ok_c = 0
                            for _p in _plsp_selected:
                                try:
                                    _client_c.table("draft_paket_pl").update({
                                        "sbu_baru": _sbu_baru_global or "",
                                        "sbu_lama": _sbu_lama_global or "",
                                    }).eq("kode_paket", _p["kode_paket"]).execute()
                                    _ok_c += 1
                                except Exception as _e:
                                    st.error(f"❌ {_pl_label(_p)}: {_e}")
                            _save_pl_sbu_history(_sbu_baru_global, _sbu_lama_global or "")
                            st.success(f"✅ {_ok_c}/{len(_plsp_selected)} paket disimpan ke Supabase")
                            import ldk_config as _ldk_cfg_custom_jkk
                            for _p in _plsp_selected:
                                try:
                                    _r_custom_ldk = _depl_jkk.submit_ldk_pl(
                                        _p["kode_paket"],
                                        sbu_baru=_sbu_baru_global or "",
                                        sbu_lama=_sbu_lama_global or "",
                                        kinerja_text=_ldk_cfg_custom_jkk.KINERJA_PENYEDIA_JKK,
                                    )
                                    st.write(f"{'✅' if _r_custom_ldk['ok'] else '❌'} {_pl_label(_p)} — POST LDK HTTP {_r_custom_ldk['status']}")
                                except Exception as _e_ldk_custom:
                                    st.error(f"❌ {_pl_label(_p)} — POST LDK gagal: {_e_ldk_custom}")

                    st.divider()

                    # ── SEKSI 2: Masa Berlaku ────────────────────────────────
                    st.markdown("#### 📅 Seksi 2 — Masa Berlaku Penawaran")
                    st.caption("Tanggal Dokpil diisi langsung pada panel Excel @ Master Data → INPUT TANGGAL DOKPIL.")

                    _ldk_masa_berlaku = st.number_input(
                        "Masa Berlaku Penawaran (hari)",
                        min_value=1, max_value=180, value=30,
                        key="plsp_masa_berlaku",
                    )

                    if st.button("💾 Submit Masa Berlaku Penawaran", key="plsp_btn_masa_berlaku", use_container_width=True):
                        from config import sb as _sb_factory_mb
                        _client_mb = _sb_factory_mb()
                        for _p in _plsp_selected:
                            try:
                                _client_mb.table("draft_paket_pl").update({
                                    "masa_berlaku": int(_ldk_masa_berlaku),
                                }).eq("kode_paket", _p["kode_paket"]).execute()
                            except Exception as _e_mb:
                                st.warning(f"⚠️ Gagal simpan masa berlaku {_pl_label(_p)}: {_e_mb}")
                            _r_mb = _depl.submit_masa_berlaku_pl(_p["kode_paket"], int(_ldk_masa_berlaku))
                            st.write(f"{'✅' if _r_mb['ok'] else '❌'} {_pl_label(_p)} — HTTP {_r_mb['status']}")

                    st.divider()

                    # ── SEKSI 3: Dokumen Kualifikasi (LDK) ───────────────────
                    st.markdown("#### 📋 Seksi 3 — Dokumen Kualifikasi (LDK)")
                    st.caption("ℹ️ Di-submit ke SPSE bagian Persyaratan Kualifikasi (LDK).")

                    st.markdown("**Syarat Administrasi** *(default: centang idx 0-3, skip 422/423)*")
                    _ADMIN_LABEL = {
                        "413": "413 — KSWP (Wajib Pajak)",
                        "414": "414 — Kapasitas Hukum (Akta Pendirian)",
                        "415": "415 — Pakta Integritas",
                        "416": "416 — Surat Pernyataan Peserta",
                        "422": "422 — (skip default)",
                        "423": "423 — (skip default)",
                    }
                    _ldk_centang_admin_ckm_ids = []
                    _cols_adm = st.columns(2)
                    for _idx_iter, (_cid, _lbl) in enumerate(_ADMIN_LABEL.items()):
                        with _cols_adm[_idx_iter % 2]:
                            _default_adm = _cid in ("413", "414", "415", "416")
                            if st.checkbox(_lbl, value=_default_adm, key=f"plsp_admin_cid_{_cid}"):
                                _ldk_centang_admin_ckm_ids.append(_cid)

                    st.markdown("**Syarat Teknis JKK Konstruksi** *(default: centang 0+1)*")
                    _TEKNIS_LABEL = {
                        "433": "433 — Pengalaman ≥1 JKK 4thn terakhir",
                        "434": "434 — Pengalaman pekerjaan sejenis",
                        "435": "435 — Pengalaman sejenis 10thn terakhir",
                        "436": "436 — Dispensasi penyedia kecil baru <3thn",
                    }
                    _ldk_teknis_ckm_ids = []
                    _cols_tk = st.columns(2)
                    for _idx_iter, (_cid, _lbl) in enumerate(_TEKNIS_LABEL.items()):
                        with _cols_tk[_idx_iter % 2]:
                            _default = True
                            if st.checkbox(_lbl, value=_default, key=f"plsp_teknis_cid_{_cid}"):
                                _ldk_teknis_ckm_ids.append(_cid)

                    import ldk_config as _ldk_cfg_pl
                    # Kinerja penyedia wajib — text beda JKK vs PK
                    _jenis_pl_paket = _plsp_selected[0].get("jenis_pl", "JKK").upper() if _plsp_selected else "JKK"
                    _ldk_kinerja_text = _ldk_cfg_pl.KINERJA_PENYEDIA_JKK if _jenis_pl_paket == "JKK" else _ldk_cfg_pl.KINERJA_PENYEDIA_PK

                    st.caption(
                        "ℹ️ Default: admin all + teknis idx 0+1 (Pengalaman + Dispensasi). "
                        "NPWP/Akta/Pakta auto by sistem. Kinerja Penyedia wajib dikirim otomatis."
                    )

                    if st.button("📋 Submit Dokumen Kualifikasi (LDK)", key="plsp_btn_ldk", use_container_width=True):
                        from config import sb as _sb_factory_ldk
                        _client_ldk = _sb_factory_ldk()
                        for _p in _plsp_selected:
                            try:
                                _client_ldk.table("draft_paket_pl").update({
                                    "sbu_baru": _sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    "sbu_lama": _sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                }).eq("kode_paket", _p["kode_paket"]).execute()
                            except Exception:
                                pass
                            _r_ldk = _depl.submit_ldk_pl(
                                _p["kode_paket"],
                                sbu_baru=_sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                sbu_lama=_sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                centang_admin_ckm_ids=_ldk_centang_admin_ckm_ids,
                                teknis_centang_ckm_ids=_ldk_teknis_ckm_ids,
                                kinerja_text=_ldk_kinerja_text,
                            )
                            st.write(f"{'✅' if _r_ldk['ok'] else '❌'} {_pl_label(_p)} — HTTP {_r_ldk['status']}")

                    st.divider()

                    # Seksi 4 — Checklist Dokumen Penawaran (hardcode semua wajib)
                    _cd_centang_admin = True
                    _cd_centang_syarat = True
                    _cd_centang_harga = True

                    st.divider()
                    st.caption("⬇️ Atau jalankan semua seksi sekaligus:")

                    # ── Submit All-in-One ─────────────────────────────────────
                    if st.button(
                        f"🚀 Push Setup ke SPSE ({len(_plsp_selected)} paket)",
                        key="plsp_submit_btn",
                        type="primary",
                        use_container_width=True,
                    ):
                        _hasil_sp = []
                        _prog_sp = st.progress(0, text="Mulai...")
                        from config import sb as _sb_factory_sp
                        _client_sp = _sb_factory_sp()
                        for _i, _p in enumerate(_plsp_selected):
                            _kp = _p["kode_paket"]
                            _id_nt = _p.get("id_nontender")
                            _nm = _pl_label(_p)
                            _prog_sp.progress((_i + 1) / len(_plsp_selected),
                                              text=f"{_nm} ({_i+1}/{len(_plsp_selected)})...")

                            # 0. Simpan SBU global ke Supabase; tanggal Dokpil berasal dari Excel
                            try:
                                _client_sp.table("draft_paket_pl").update({
                                    "sbu_baru": _sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    "sbu_lama": _sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                }).eq("kode_paket", _kp).execute()
                            except Exception as _e_save:
                                _hasil_sp.append({"paket": _nm, "step": "Simpan Supabase", "ok": False, "pesan": str(_e_save)[:80]})

                            # 1. Submit LDK (kode_paket, bukan id_nontender)
                            try:
                                _r_ldk = _depl.submit_ldk_pl(
                                    _kp,
                                    sbu_baru=_sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    sbu_lama=_sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                    centang_admin_ckm_ids=_ldk_centang_admin_ckm_ids,
                                    teknis_centang_ckm_ids=_ldk_teknis_ckm_ids,
                                    kinerja_text=_ldk_kinerja_text,
                                )
                                _ijin_note = f" | ijin CDP: {_r_ldk.get('ijin_update','—')}" if _r_ldk.get("ijin_update") else ""
                                _hasil_sp.append({
                                    "paket": _nm, "step": "LDK",
                                    "ok": _r_ldk["ok"], "pesan": f"HTTP {_r_ldk['status']}{_ijin_note}",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "LDK", "ok": False, "pesan": str(_e)[:80]})

                            # 2. Masa berlaku penawaran
                            try:
                                _r_mb = _depl.submit_masa_berlaku_pl(_kp, int(_ldk_masa_berlaku))
                                try:
                                    from config import sb as _sb_mb_bulk
                                    _sb_mb_bulk().table("draft_paket_pl").update({"masa_berlaku": int(_ldk_masa_berlaku)}).eq("kode_paket", _kp).execute()
                                except Exception:
                                    pass
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Masa Berlaku",
                                    "ok": _r_mb["ok"], "pesan": f"HTTP {_r_mb['status']} ({_ldk_masa_berlaku} hari)",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "Masa Berlaku", "ok": False, "pesan": str(_e)[:80]})

                            # 3. Checklist Dokumen Penawaran
                            try:
                                _r_cd = _depl.submit_checklist_pl(
                                    _kp,
                                    centang_admin_all=_cd_centang_admin,
                                    centang_syarat_all=_cd_centang_syarat,
                                    centang_harga_all=_cd_centang_harga,
                                )
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Checklist Dok Penawaran",
                                    "ok": _r_cd["ok"], "pesan": f"HTTP {_r_cd['status']}",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "Checklist Dok Penawaran", "ok": False, "pesan": str(_e)[:80]})

                            # 4. Upload Dokpil PDF (jika ada file)
                            _dokpil_file = _p.get("_dokpil_file")
                            if _dokpil_file and _id_nt:
                                try:
                                    _tgl_excel = _p.get("tgl_dokpil")
                                    if _tgl_excel:
                                        from datetime import date as _date_excel
                                        _tgl_excel = _date_excel.fromisoformat(str(_tgl_excel)[:10])
                                    else:
                                        _tgl_excel = datetime.now().date()
                                    # Generate Nomor Dokpil: 000.3.3/01/PL/PP-NN/{KodeUnik}/{SkpdSingkat}/{Tahun}
                                    _kode_unik = _p.get("kode_unik") or ""
                                    _skpd_singkat = _lookup_singkatan_dinas(_p.get("satker", ""))
                                    _nomor_dokpil = _p.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                        nama_paket=_p["nama_paket"],
                                        kode_unik=_kode_unik,
                                        skpd_singkat=_skpd_singkat,
                                        tahun=_tgl_excel.year,
                                        paket_ulang=_pl_paket_ulang(_p),
                                        nomor_urut=_p.get("nomor_urut"),
                                    )
                                    _r_up = _udpl.upload_dokpil_pl(
                                        kode_paket=_kp,
                                        file_bytes=_dokpil_file.getvalue(),
                                        file_name=_dokpil_file.name,
                                        nomor_dokpil=_nomor_dokpil,
                                        tgl_dokpil=_tgl_excel.strftime("%d-%m-%Y"),
                                    )
                                    _hasil_sp.append({
                                        "paket": _nm, "step": "Upload Dokpil",
                                        "ok": _r_up["ok"],
                                        "pesan": f"HTTP {_r_up.get('status','?')} | {_nomor_dokpil}",
                                    })
                                    if _r_up["ok"]:
                                        try:
                                            _client_sp.table("draft_paket_pl").update({
                                                "nomor_dokpil": _nomor_dokpil,
                                            }).eq("kode_paket", _kp).execute()
                                        except Exception:
                                            pass
                                except Exception as _e:
                                    _hasil_sp.append({
                                        "paket": _nm, "step": "Upload Dokpil",
                                        "ok": False, "pesan": str(_e)[:80],
                                    })
                            elif _dokpil_file and not _id_nt:
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Upload Dokpil",
                                    "ok": False, "pesan": "id_nontender kosong, tidak bisa upload",
                                })

                        _prog_sp.empty()
                        _sukses_sp = sum(1 for h in _hasil_sp if h["ok"])
                        _gagal_sp = len(_hasil_sp) - _sukses_sp
                        if _gagal_sp == 0:
                            st.success(f"✅ Semua {_sukses_sp} operasi sukses!")
                        else:
                            st.warning(f"⚠️ {_sukses_sp} sukses, {_gagal_sp} gagal")
                        _load_draft_pl_cached.clear()

                        # Tampilkan log per paket
                        import pandas as _pd
                        _df_sp = _pd.DataFrame(_hasil_sp)
                        if not _df_sp.empty:
                            _df_sp["status"] = _df_sp["ok"].map({True: "✅", False: "❌"})
                            st.dataframe(
                                _df_sp[["status", "paket", "step", "pesan"]],
                                use_container_width=True, hide_index=True,
                            )


    if _pl_active_tab == "4️⃣ Pilih Penyedia & Umumkan":
        st.divider()
        st.markdown("### 🏢 Pilih Penyedia ke SPSE")
        st.caption(
            "Cari penyedia by NPWP → klik pilih ke SPSE (prioritas kabupaten Tapin, "
            "fallback semua kabupaten Kalsel propinsi 22)."
        )

        if st.button("🔄 Refresh Data Penyedia", key="pp_refresh_penyedia_jkk", help="Parse ulang 4. Informasi Lainnya/8. ND.pdf (fallback Draft_PL) → update nama & NPWP penyedia di Supabase"):
            _ref_bar_jkk = st.progress(0.0, text="Memulai...")
            _ref_log_jkk = st.empty()
            import parse_kak_pl as _pkpl_ref_jkk
            def _ref_cb_jkk(p, m):
                _ref_bar_jkk.progress(min(float(p), 1.0), text=m[:120])
                _ref_log_jkk.caption(m)
            _ref_res_jkk = _pkpl_ref_jkk.serap_penyedia_pl(progress_cb=_ref_cb_jkk)
            _ref_bar_jkk.progress(1.0, text="Selesai")
            st.success(f"✅ {_ref_res_jkk.get('updated',0)} diperbarui, {_ref_res_jkk.get('not_found',0)} tidak ditemukan, {len(_ref_res_jkk.get('errors',[]))} error")
            if _ref_res_jkk.get("errors"):
                st.warning("\n".join(_ref_res_jkk["errors"][:5]))
            _load_draft_pl_cached.clear()
            st.rerun()
        _pp_rows = _load_draft_pl_cached()
        _pp_rows, _ = pl_engine.buang_duplikat_paket_lama(_pp_rows)
        _pp_rows = [r for r in _pp_rows if not pl_engine.is_paket_selesai(r)]
        if _pp_rows:
            import pilih_penyedia_pl as _ppp

            _pp_col_list, _pp_col_act = st.columns([2, 3])

            with _pp_col_list:
                st.markdown("**Pilih paket:**")
                _pp_sel_all, _pp_sel_none = st.columns(2)
                with _pp_sel_all:
                    if st.button("✅ Semua", key="pp_sel_all", use_container_width=True):
                        for _rr in _pp_rows:
                            st.session_state[f"pp_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _pp_sel_none:
                    if st.button("⬜ Kosong", key="pp_sel_none", use_container_width=True):
                        for _rr in _pp_rows:
                            st.session_state[f"pp_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _pp_selected = []
                for _rr in _pp_rows:
                    _kp = _rr["kode_paket"]
                    _npwp_disp = _rr.get("npwp_penyedia") or "—"
                    _nama_disp = _rr.get("nama_penyedia") or "—"
                    _pp_chk_key = f"pp_chk_{_kp}_v19"
                    if _pp_chk_key not in st.session_state:
                        st.session_state[_pp_chk_key] = True
                    _chk = st.checkbox("", key=_pp_chk_key, label_visibility="collapsed", help=f"Penyedia: {_nama_disp} | NPWP: {_npwp_disp}")
                    st.markdown(f"**{_pl_label(_rr)}**")
                    if _chk:
                        _pp_selected.append(_rr)

                st.caption(f"**{len(_pp_selected)}** paket dipilih")

            with _pp_col_act:
                if not _pp_selected:
                    st.info("Pilih paket di sebelah kiri.")
                else:
                    # Tabel ringkas paket terpilih
                    import pandas as _pd2
                    _pp_df = _pd2.DataFrame([{
                        "Paket": _pl_label(r),
                        "Penyedia": r.get("nama_penyedia") or "—",
                        "NPWP": r.get("npwp_penyedia") or "—",
                    } for r in _pp_selected])
                    st.dataframe(_pp_df, use_container_width=True, hide_index=True)

                    _invalid = [r for r in _pp_selected if not r.get("npwp_penyedia")]
                    if _invalid:
                        st.warning(
                            f"⚠️ {len(_invalid)} paket belum ada NPWP penyedia: "
                            + ", ".join(_pl_label(r) for r in _invalid)
                        )

                    _valid_pp = [r for r in _pp_selected if r.get("npwp_penyedia")]
                    if _valid_pp:
                        if st.button(
                            f"🏢 Pilih Semua Penyedia ke SPSE ({len(_valid_pp)} paket)",
                            key="pp_submit_btn",
                            type="primary",
                            use_container_width=True,
                        ):
                            import spse_browser as _spse_br
                            _ck_pp = _spse_br.get_spse_cookies()
                            _base_pp = pl_engine.BASE_URL

                            _pp_hasil = []
                            _pp_prog = st.progress(0, text="Mulai pilih penyedia...")
                            for _i_pp, _pp_r in enumerate(_valid_pp):
                                _pp_nm = _pl_label(_pp_r)
                                _pp_prog.progress(
                                    (_i_pp + 1) / len(_valid_pp),
                                    text=f"{_pp_nm} ({_i_pp+1}/{len(_valid_pp)})...",
                                )
                                try:
                                    _res_pp = _ppp.cari_dan_pilih_penyedia(
                                        kode_paket=_pp_r["kode_paket"],
                                        npwp=_pp_r.get("npwp_penyedia") or "",
                                        cookie_str=_ck_pp,
                                        base_url=_base_pp,
                                        nama_penyedia=_pp_r.get("nama_penyedia") or "",
                                    )
                                    _pp_hasil.append({
                                        "paket": _pp_nm,
                                        "ok": _res_pp["ok"],
                                        "pesan": (
                                            f"✅ {_res_pp.get('nama','?')} (kab {_res_pp.get('kabupaten_id','')})"
                                            if _res_pp["ok"]
                                            else f"❌ {_res_pp.get('pesan','?')}"
                                        ),
                                    })
                                except Exception as _e_pp:
                                    _pp_hasil.append({
                                        "paket": _pp_nm,
                                        "ok": False,
                                        "pesan": f"❌ Error: {str(_e_pp)[:80]}",
                                    })

                            _pp_prog.empty()
                            _pp_ok = sum(1 for h in _pp_hasil if h["ok"])
                            if _pp_ok == len(_pp_hasil):
                                st.success(f"✅ {_pp_ok}/{len(_pp_hasil)} paket berhasil dipilih penyedia!")
                            else:
                                st.warning(f"⚠️ {_pp_ok}/{len(_pp_hasil)} sukses")

                            _df_pp = _pd2.DataFrame(_pp_hasil)
                            if not _df_pp.empty:
                                st.dataframe(
                                    _df_pp[["paket", "pesan"]],
                                    use_container_width=True, hide_index=True,
                                )

    # ── Tab 7: Kirim Verifikasi Penyedia ─────────────────────────────────────
    # ── Tab 4 Section 2: Pilih Penyedia ke SPSE ─────────────────────────────
    # ── Tab 3 Section: Umumkan Paket Non Tender (PL JKK) ─────────────────────
    if _pl_active_tab == "4️⃣ Pilih Penyedia & Umumkan":
        st.divider()
        st.markdown("### 📢 Umumkan Paket Non Tender")
        st.caption("Setujui Pakta Integritas dan umumkan paket ke SPSE. Pastikan browser SPSE sudah terhubung.")
        _paket_berfolder_umum = [r for r in _pp_rows if r.get("kode_paket") and r.get("folder_dibuat")]
        if not _paket_berfolder_umum:
            st.info("Tidak ada paket berfolder yang bisa diumumkan.")
        else:
            _umum_col1, _umum_col2 = st.columns(2)
            with _umum_col1:
                if st.button("✅ Semua", key="umum_sel_all_jkk", use_container_width=True):
                    for _r in _paket_berfolder_umum:
                        st.session_state[f"umum_chk_{_r['kode_paket']}"] = True
                    st.rerun()
            with _umum_col2:
                if st.button("⬜ Kosong", key="umum_sel_none_jkk", use_container_width=True):
                    for _r in _paket_berfolder_umum:
                        st.session_state[f"umum_chk_{_r['kode_paket']}"] = False
                    st.rerun()
            _pilih_umum = []
            for _r in _paket_berfolder_umum:
                _umum_key = f"umum_chk_{_r['kode_paket']}"
                _umum_chk = st.checkbox(
                    _pl_label(_r),
                    value=st.session_state.get(_umum_key, True),
                    key=_umum_key,
                )
                if _umum_chk:
                    _pilih_umum.append(_r["kode_paket"])
            if st.button("📢 Umumkan Paket Terpilih", key="btn_umumkan_pl_jkk", disabled=not _pilih_umum):
                try:
                    import spse_browser as _spse_br_umum
                    _spse_br_umum.buka_browser(navigate=False)
                    _cookie_umum = _spse_br_umum.get_spse_cookies()
                except Exception as _e_br_umum:
                    st.error(f"Browser SPSE tidak terhubung: {_e_br_umum}")
                    _cookie_umum = None
                if _cookie_umum:
                    for _kp_umum in _pilih_umum:
                        _r_umum = next((r for r in _paket_berfolder_umum if r["kode_paket"] == _kp_umum), {"nama_paket": _kp_umum})
                        _nm_umum = _pl_label(_r_umum)
                        _ru = pl_engine.umumkan_paket_pl(_kp_umum, _cookie_umum)
                        if _ru["ok"]:
                            st.success(f"✅ {_nm_umum[:60]} — {_ru['pesan']}")
                        else:
                            st.error(f"❌ {_nm_umum[:60]} — {_ru['pesan']}")

    if _pl_active_tab == "8️⃣ Kirim Verifikasi":
        import verifikasi_penyedia_pl as _verif_pl
        from gcal_helper import get_jadwal_klarifikasi_pl as _gcal_klarifikasi


        # Dropdown paket — gabung draft_pl + aktif_pl (dedup by id_nontender)
        _verif_rows, _verif_load_error = _load_verifikasi_pl_rows_cached()
        if _verif_load_error:
            st.error(f"Gagal load paket PL: {_verif_load_error}")

        st.markdown("## 📨 Kirim Undangan Verifikasi Penyedia")
        st.caption("Centang paket yang ingin dikirim. Hanya paket dengan peserta terdaftar yang tampil.")

        # Buang duplikat row lama (paket di-ulang → kode baru, row lama nyangkut)
        _verif_rows, _verif_dup_n = pl_engine.buang_duplikat_paket_lama(_verif_rows)

        # Load status peserta untuk filter
        _batch_rows = _verif_rows  # sudah di-load di atas
        if _batch_rows:
            # Fetch jumlah peserta semua paket
            _batch_kodes = tuple(r["kode_paket"] for r in _batch_rows if r.get("kode_paket"))
            _batch_mon = _fetch_status_semua_paket_cached(_batch_kodes)

            # Filter hanya paket yang ada peserta
            _batch_eligible = [
                r for r in _batch_rows
                if _batch_mon.get(r["kode_paket"], {}).get("jumlah", 0) > 0
                and r.get("id_nontender")
            ]

            if not _batch_eligible:
                st.info("Belum ada paket dengan peserta terdaftar.")
            else:
                st.caption(f"{len(_batch_eligible)} paket tersedia (sudah ada peserta).")

                # Helper: paket dianggap selesai jika terkirim ATAU sudah di tahap akhir SPSE
                _TAHAP_SELESAI = {"Penandatanganan Kontrak", "Paket Sudah Selesai"}
                def _is_selesai(r):
                    if r.get("status_undangan_verifikasi") == "terkirim":
                        return True
                    if r.get("tahap_spse") in _TAHAP_SELESAI:
                        return True
                    return False

                _n_selesai = sum(1 for r in _batch_eligible if _is_selesai(r))
                _n_belum = len(_batch_eligible) - _n_selesai
                _sembunyikan_selesai = st.checkbox(
                    f"Sembunyikan yang sudah selesai ({_n_selesai} paket)",
                    value=True,
                    key="tab7_sembunyikan_selesai"
                )
                if _sembunyikan_selesai:
                    _batch_tampil = [r for r in _batch_eligible if not _is_selesai(r)]
                else:
                    _batch_tampil = _batch_eligible

                if not _batch_tampil:
                    st.success("✅ Semua paket sudah selesai!")
                else:
                    # Waktu — date_input + time_input (pola Tab 3)
                    import datetime as _dtlb
                    _bwc1, _bwc2 = st.columns(2)
                    with _bwc1:
                        st.caption("**Waktu Mulai**")
                        _bv_tgl = st.date_input("Tanggal", value=_dtlb.date.today(), format="DD/MM/YYYY", key="batch_verif_tgl")
                        _bv_jam = st.time_input("Jam", value=_dtlb.time(9, 0), key="batch_verif_jam")
                        st.markdown(f"**{_HARI_NAMA[_bv_tgl.weekday()]}, {_bv_tgl.day} {_BULAN_NAMA[_bv_tgl.month-1]} {_bv_tgl.year}**")
                    with _bwc2:
                        st.caption("**Waktu Selesai**")
                        _bv_tgl_end = st.date_input("Tanggal", value=_dtlb.date.today(), format="DD/MM/YYYY", key="batch_verif_tgl_end")
                        _bv_jam_end = st.time_input("Jam", value=_dtlb.time(15, 0), key="batch_verif_jam_end")
                    _batch_start = f"{_bv_tgl.strftime('%d-%m-%Y')} {_bv_jam.strftime('%H:%M')}"
                    _batch_end   = f"{_bv_tgl_end.strftime('%d-%m-%Y')} {_bv_jam_end.strftime('%H:%M')}"

                    # Checkbox list paket
                    _sudah_kirim = [r for r in _batch_tampil if r.get("status_undangan_verifikasi") == "terkirim"]
                    _belum_kirim = [r for r in _batch_tampil if r.get("status_undangan_verifikasi") != "terkirim"]
                    _ca, _cb = st.columns([4, 1])
                    _ca.markdown("**Pilih paket yang akan dikirim:**")
                    if _sudah_kirim:
                        _ca.caption(f"✅ {len(_sudah_kirim)} sudah terkirim | ⏳ {len(_belum_kirim)} belum")
                    _centang_semua = _cb.checkbox("Centang Semua", value=False, key="batch_centang_semua")
                    _batch_selected = []
                    _bc1, _bc2 = st.columns(2)
                    for _bi, _br in enumerate(_batch_tampil):
                        _ku = _br.get('kode_unik') or _br['kode_paket']
                        _tgl_kirim = _br.get("tgl_undangan_verifikasi")
                        _sudah = _br.get("status_undangan_verifikasi") == "terkirim"
                        _kontrak = _br.get("tahap_spse") == "Penandatanganan Kontrak"
                        _hint_ulang = _pl_hint_ulang(_br)
                        if _sudah and _tgl_kirim:
                            try:
                                _tgl_fmt = _dtlb.datetime.fromisoformat(_tgl_kirim.replace("Z","+00:00")).strftime("%d-%m-%Y")
                            except Exception:
                                _tgl_fmt = _tgl_kirim[:10]
                            _label = f"{_ku} — {_pl_label(_br)} ✅ {_tgl_fmt}"
                        elif _kontrak:
                            _label = f"{_ku} — {_pl_label(_br)} 🔒 Kontrak"
                        else:
                            _label = f"{_ku} — {_pl_label(_br)}"
                        _col = _bc1 if _bi % 2 == 0 else _bc2
                        _default_chk = _centang_semua or (not _sudah and not _kontrak)
                        if _col.checkbox(_label, value=_default_chk, key=f"batch_chk_{_br['kode_paket']}"):
                            _batch_selected.append(_br)

                    st.markdown(f"**{len(_batch_selected)} paket dipilih**")

                    if st.button("📨 Kirim Undangan Verifikasi", key="btn_batch_kirim_konfirm", type="primary", disabled=not _batch_selected):
                        if not _batch_start or not _batch_end:
                            st.error("Waktu mulai dan selesai wajib diisi.")
                        else:
                                import verifikasi_penyedia_pl as _vpl_batch
                                import evaluasi_admin_kualifikasi_pl as _eval_verif_batch
                                _hasil_batch = []
                                _prog = st.progress(0, text="Mengirim...")
                                for _bi2, _bp2 in enumerate(_batch_selected):
                                    _prog.progress((_bi2 + 1) / len(_batch_selected), text=f"Kirim ke {_bp2.get('kode_unik') or _bp2['kode_paket']}...")
                                    _peserta_res = _eval_verif_batch.scrape_peserta_evaluasi(_bp2["kode_paket"])
                                    if not _peserta_res.get("ok") or not _peserta_res.get("peserta"):
                                        _res = {"ok": False, "msg": _peserta_res.get("pesan", "Peserta tidak ditemukan")}
                                    else:
                                        _hasil_peserta = []
                                        for _peserta in _peserta_res["peserta"]:
                                            _hasil_peserta.append(_vpl_batch.kirim_verifikasi(
                                                id_nontender=_peserta["id_nontender"],
                                                kode_paket=_bp2["kode_paket"],
                                                waktu_start=_batch_start,
                                                waktu_end=_batch_end,
                                            ))
                                        _res = {
                                            "ok": all(_r.get("ok") for _r in _hasil_peserta),
                                            "msg": "; ".join(_r.get("msg", "") for _r in _hasil_peserta),
                                        }
                                    _hasil_batch.append({
                                        "paket": _bp2.get("kode_unik") or _bp2["kode_paket"],
                                        "nama": _pl_label(_bp2),
                                        "ok": _res["ok"],
                                        "msg": _res["msg"],
                                    })
                                _prog.empty()

                                # Tampilkan hasil
                                _ok_list = [h for h in _hasil_batch if h["ok"]]
                                _fail_list = [h for h in _hasil_batch if not h["ok"]]
                                if _ok_list:
                                    st.success(f"✅ Berhasil: {len(_ok_list)} paket")
                                    for h in _ok_list:
                                        st.write(f"  ✅ {h['paket']} — {h['nama']}")
                                if _fail_list:
                                    st.error(f"❌ Gagal: {len(_fail_list)} paket")
                                    for h in _fail_list:
                                        st.write(f"  ❌ {h['paket']} — {h['nama']}: {h['msg']}")

    # ── Tab 8: Upload BA PL ───────────────────────────────────────────────────
    if _pl_active_tab == "9️⃣ Upload BA PL":
        import ba_engine_pl as _ba_pl_engine5
        import ba_config_pl as _ba_cfg_pl
        import os as _os8
        import re as _re8
        import datetime as _dt8

        st.markdown("## Berita Acara — Pengadaan Langsung")

        # Load paket PL — filter paket selesai
        _pl8_rows = []
        try:
            _raw8 = _load_draft_pl_cached()
            _raw8, _ = pl_engine.buang_duplikat_paket_lama(_raw8)
            _pl8_rows = [r for r in _raw8 if not pl_engine.is_paket_selesai(r)]
            _pl8_kode_status = []
            for _r8 in _pl8_rows:
                _s8 = _enrich_kode_unik_pl_excel(_r8)
                if _s8 != "excel":
                    _pl8_kode_status.append(
                        f"{_r8.get('kode_paket')}: {_s8}"
                    )
        except Exception as _e8:
            st.error(f"Gagal load paket PL: {_e8}")
            _pl8_kode_status = []
        if _pl8_kode_status:
            st.warning(
                "⚠️ Status Kode Unik PL (sumber utama `@ Master Data!F2`):\n\n"
                + "\n".join(f"- {x}" for x in _pl8_kode_status)
            )

        # ── Helper auto nomor + tanggal (otomatis seperti mode tender) ───────
        def _skpd_pl8(_satker):
            if not _satker:
                return "DPUPR"
            try:
                from config import sb as _sbf8
                _r = _sbf8().table("master_dinas").select("singkatan").ilike(
                    "nama_dinas", f"%{_satker[:30]}%").limit(1).execute()
                if _r.data:
                    return _r.data[0].get("singkatan") or "DPUPR"
            except Exception:
                pass
            return "DPUPR"

        def _kode_unik_excel_pl8(_row):
            """Baca kode unik manual dari @ Master Data!F2 workbook paket."""
            return _baca_master_data_pl(_row).get("kode_unik", "")

        def _nomor_dokpil_pl8(_row):
            """Nomor dokpil final dari @ Master Data!C20 workbook paket."""
            _md = _baca_master_data_pl(_row)
            _base = str(_md.get("nomor_dokpil") or "").strip()
            _kode = str(_md.get("kode_unik") or _row.get("kode_unik") or "").strip()
            if _base and _kode:
                # Format PL: .../PP-XX/<kode unik>/<SKPD>/2026.
                _base = _re8.sub(r"(/PP-[^/]+/)[^/]+(?=/)", rf"\1{_kode}", _base, count=1)
            return _base

        def _auto_nomor_pl8(_row, _jenis_key):
            """Derive nomor BA per jenis: ganti slot /NN/ pertama."""
            _base = _nomor_dokpil_pl8(_row)
            if not _base:
                return ""
            return ba_engine.derive_nomor_ba(_base, _ba_cfg_pl.NOMOR_URUT_PL[_jenis_key])

        def _auto_tgl_pl8(_row):
            """Tanggal BA = end Evaluasi Penawaran (tgl_evaluasi), fallback live SPSE."""
            _v = _row.get("tgl_evaluasi")
            if _v:
                try:
                    return _dt8.date.fromisoformat(str(_v)[:10])
                except Exception:
                    pass
            try:
                _jd = _parse_jadwal_pl_cached(_row.get("kode_paket", ""))
                if _jd and len(_jd) > 2:
                    return _jd[2]["selesai"].date()
            except Exception:
                pass
            return None

        def _backup_pdf_pl8(_row, _jenis_key, _pdf_bytes):
            """Simpan PDF cetak ke folder masing-masing paket. Return path atau ''."""
            _kp = _row.get("kode_paket", "")
            _dir = None
            # 1) resolver folder paket PL resmi (folder_paket root)
            try:
                import kualifikasi_engine_pl as _kepl8
                _rf = _kepl8.resolve_folder_paket_pl(_kp)
                if _rf.get("ok"):
                    _root = _rf.get("pesan") or ""  # pesan = folder_paket root
                    if _root and _os8.path.isdir(_root):
                        _sub = _os8.path.join(_root, "7. Berita Acara + Summary Non Tender")
                        _os8.makedirs(_sub, exist_ok=True)
                        _dir = _sub
            except Exception:
                pass
            # 2) fallback: Asisten_Pokja_Downloads
            if not _dir:
                try:
                    from config import POKJA_ROOT as _PR8
                    _dir = _os8.path.join(_PR8, "Asisten_Pokja_Downloads", f"Cetak_BA_PL_{_kp}")
                except Exception:
                    return ""
            try:
                _os8.makedirs(_dir, exist_ok=True)
                _fn = f"{_ba_cfg_pl.FILE_LABEL_PL[_jenis_key]}-{_kp}.pdf"
                _fp = _os8.path.join(_dir, _fn)
                with open(_fp, "wb") as _fh:
                    _fh.write(_pdf_bytes)
                return _fp
            except Exception:
                return ""

        def _proses_ba_pl8(_row, _jenis_key, _nomor, _tgl):
            """Cetak → backup → upload satu BA. Return (ok: bool, pesan: str)."""
            # Endpoint cetak/upload BA nontender pakai kode_paket (BUKAN id_nontender)
            _id = _row.get("kode_paket")
            if not _nomor:
                return False, "nomor kosong"
            if not _tgl:
                return False, "tanggal kosong (Evaluasi belum dijadwal)"
            _tgls = _tgl.strftime("%d-%m-%Y")
            try:
                _rc = _ba_pl_engine5.cetak_ba_pl(
                    id_nontender=_id, jenis_key=_jenis_key,
                    nomor_ba=_nomor, tanggal_ba=_tgls,
                )
                if not _rc["ok"]:
                    return False, f"cetak gagal ({_rc.get('status')}) {_rc.get('error','')}"
                _backup_pdf_pl8(_row, _jenis_key, _rc["pdf_bytes"])
                _fn = f"{_ba_cfg_pl.FILE_LABEL_PL[_jenis_key]}-{_row.get('kode_paket','')}.pdf"
                _ru = _ba_pl_engine5.upload_ba_pl(
                    id_nontender=_id, jenis_key=_jenis_key,
                    nomor_ba=_nomor, tanggal_ba=_tgls,
                    file_bytes=_rc["pdf_bytes"], file_name=_fn,
                )
                if _ru.get("ok"):
                    return True, "OK"
                return False, f"upload gagal ({_ru.get('status')})"
            except Exception as _e:
                return False, str(_e)[:100]

        # ── SECTION 1: 2 BA AUTO (Evaluasi + Hasil) ──────────────────────────
        st.markdown("### 1. Cetak + Upload Otomatis")
        st.caption("BA Evaluasi Penawaran dan BA Hasil Non Tender dicetak langsung dari SPSE lalu di-upload kembali.")

        if not _pl8_rows:
            st.info("Tidak ada paket PL di database.")
        else:
            # Tanggal: default OTOMATIS dari tgl_evaluasi per paket (mode tender).
            _pl8_tgl_mode = st.radio(
                "Mode Tanggal",
                ["Otomatis (tgl Evaluasi per paket)", "Satu tanggal semua manual"],
                horizontal=True, key="pl8_tgl_mode",
            )
            _pl8_tgl_global = None
            if _pl8_tgl_mode == "Satu tanggal semua manual":
                _pl8_tgl_global = st.date_input(
                    "Tanggal BA (semua paket)", value=datetime.now().date(),
                    format="DD/MM/YYYY", key="pl8_tgl_global",
                )
                st.caption(f"{_HARI_NAMA[_pl8_tgl_global.weekday()]}, {_pl8_tgl_global.day} "
                           f"{_BULAN_NAMA[_pl8_tgl_global.month-1]} {_pl8_tgl_global.year}")
            else:
                st.caption("Tanggal otomatis = hari terakhir Evaluasi Penawaran (dari jadwal). "
                           "Nomor BA auto-derive dari Nomor Dokpil.")

            # ── Tombol BULK: cetak+upload SEMUA paket × 2 BA tanpa centang ──
            if st.button(
                f"🚀🚀 Cetak + Upload SEMUA Paket ({len(_pl8_rows)} paket × 2 BA)",
                type="primary", key="pl8_bulk_all_paket", use_container_width=True,
            ):
                with st.status(f"Proses {len(_pl8_rows)} paket × 2 BA...", expanded=True) as _stb8:
                    _ok_total, _gagal_total = 0, 0
                    for _pbulk in _pl8_rows:
                        _kb = _pbulk.get("kode_paket", "")
                        _tgl_b = _pl8_tgl_global if _pl8_tgl_mode == "Satu tanggal semua manual" else _auto_tgl_pl8(_pbulk)
                        _stb8.write(f"**{_pl_label(_pbulk)}**")
                        for _jkb, _lblb in [("evaluasi", "Evaluasi"), ("hasil", "Hasil")]:
                            _nob = _auto_nomor_pl8(_pbulk, _jkb)
                            _okb, _pesb = _proses_ba_pl8(_pbulk, _jkb, _nob, _tgl_b)
                            if _okb:
                                _stb8.write(f"  ✅ BA {_lblb}")
                                _ok_total += 1
                            else:
                                _stb8.write(f"  ❌ BA {_lblb} — {_pesb}")
                                _gagal_total += 1
                    _stb8.update(
                        label=f"Selesai — {_ok_total} BA OK, {_gagal_total} gagal.",
                        state="complete" if _gagal_total == 0 else "error",
                    )

            st.divider()

            # Header tabel per paket
            # Daftar paket — 1 baris horizontal per paket
            for _p8 in _pl8_rows:
                _k8  = _p8.get("kode_paket", "")
                _id8 = _k8
                if _pl8_tgl_mode == "Satu tanggal semua manual":
                    _tgl8 = _pl8_tgl_global
                else:
                    _tgl8 = _auto_tgl_pl8(_p8)
                _no8ev = _auto_nomor_pl8(_p8, "evaluasi")
                _no8hs = _auto_nomor_pl8(_p8, "hasil")
                _col_nama, _col_tgl, _col_btn = st.columns([5, 3, 2])
                with _col_nama:
                    st.markdown(f"**{_pl_label(_p8)}**")
                with _col_tgl:
                    if _tgl8:
                        st.caption(f"📅 {_HARI_NAMA[_tgl8.weekday()]}, {_tgl8.day} {_BULAN_NAMA[_tgl8.month-1]} {_tgl8.year}")
                    else:
                        st.caption("⚠️ Tanggal belum ada")
                with _col_btn:
                    if _tgl8 and st.button("🖨️ Cetak + Upload", key=f"pl8_ev_hs_{_k8}", use_container_width=True, type="primary"):
                        _tgl8s = _tgl8.strftime("%d-%m-%Y")
                        for _jk8, _no8, _lbl8 in [("evaluasi", _no8ev, "BA Evaluasi"), ("hasil", _no8hs, "BA Hasil")]:
                            with st.spinner(f"Proses {_lbl8}..."):
                                _rc8x = _ba_pl_engine5.cetak_ba_pl(id_nontender=_id8, jenis_key=_jk8, nomor_ba=_no8, tanggal_ba=_tgl8s)
                                if _rc8x["ok"]:
                                    _backup_pdf_pl8(_p8, _jk8, _rc8x["pdf_bytes"])
                                    _fn8x = f"{_ba_cfg_pl.FILE_LABEL_PL[_jk8]}-{_k8}.pdf"
                                    _ru8x = _ba_pl_engine5.upload_ba_pl(id_nontender=_id8, jenis_key=_jk8, nomor_ba=_no8, tanggal_ba=_tgl8s, file_bytes=_rc8x["pdf_bytes"], file_name=_fn8x)
                                    if _ru8x.get("ok"):
                                        st.success(f"✅ {_lbl8} berhasil")
                                    else:
                                        st.error(f"❌ {_lbl8} upload gagal: {_ru8x.get('status')}")
                                else:
                                    st.error(f"❌ {_lbl8} cetak gagal: {_rc8x.get('error')}")

            # Tombol batch semua

        # ── SECTION 2: BA LAINNYA (upload-only) — HIDDEN ─────────────────────
        if False:
            st.divider()
            st.markdown("### 2. BA Lainnya (Upload Manual)")
            st.caption("BA Penjelasan, Pengumuman Pemenang, dll — upload file PDF langsung ke SPSE.")

            if _pl8_rows:
                _JENIS_LAINNYA = {
                    "penjelasan":   "BA Penjelasan",
                    "pengumuman":   "Pengumuman Pemenang Akhir",
                }
                _pl8_lc1, _pl8_lc2 = st.columns([1, 2])
                with _pl8_lc1:
                    _jenis_lain8 = st.selectbox(
                        "Jenis BA",
                        options=list(_JENIS_LAINNYA.keys()),
                        format_func=lambda k: _JENIS_LAINNYA[k],
                        key="pl8_jenis_lain",
                    )
                with _pl8_lc2:
                    _nomor_lain8  = st.text_input(
                        "Nomor BA",
                        placeholder="000.3.3/06/PL/PP-01/KPP1/DPUPR/2026",
                        key="pl8_nomor_lain",
                    )
                _tgl_lain8 = st.date_input(
                    "Tanggal BA", value=datetime.now().date(), format="DD/MM/YYYY",
                    key="pl8_tgl_lain",
                )
                st.caption(f"{_HARI_NAMA[_tgl_lain8.weekday()]}, {_tgl_lain8.day} "
                           f"{_BULAN_NAMA[_tgl_lain8.month-1]} {_tgl_lain8.year}")
                _tempat_lain8 = ""
                if _jenis_lain8 == "pengumuman":
                    _tempat_lain8 = st.text_input("Tempat", placeholder="Contoh: Kantor RSUD", key="pl8_tempat_lain")
                _info_lain8 = st.text_area("Keterangan Tambahan", value="", key="pl8_info_lain", height=68)

                st.markdown("**Pilih Paket + Upload File:**")
                _pl8_lain_valid = []
                for _p8l in _pl8_rows:
                    _k8l  = _p8l.get("kode_paket", "")
                    _id8l = _k8l  # BA nontender pakai kode_paket
                    _lc1, _lc2, _lc3 = st.columns([3, 3, 1])
                    with _lc1:
                        st.caption(_pl_label(_p8l))
                    with _lc2:
                        _fl8 = st.file_uploader(
                            "PDF", type=["pdf"], key=f"pl8_lain_file_{_k8l}",
                            label_visibility="collapsed",
                        )
                    with _lc3:
                        if _fl8 and st.button("📤", key=f"pl8_lain_up_{_k8l}", use_container_width=True):
                            with st.spinner(f"Upload {_k8l}..."):
                                _rl8 = _ba_pl_engine5.upload_ba_pl(
                                    id_nontender=_id8l, jenis_key=_jenis_lain8,
                                    nomor_ba=_nomor_lain8,
                                    tanggal_ba=_tgl_lain8.strftime("%d-%m-%Y"),
                                    info=_info_lain8,
                                    file_bytes=_fl8.read(), file_name=_fl8.name,
                                    tempat=_tempat_lain8,
                                )
                            if _rl8.get("ok"):
                                st.success(f"✅ {_k8l} — berhasil")
                            else:
                                st.error(f"❌ {_k8l} — status {_rl8.get('status')}")
                    if _fl8:
                        _pl8_lain_valid.append({**_p8l, "_id8l": _id8l, "_fl8": _fl8})

                if _pl8_lain_valid:
                    if st.button(
                        f"📤 Upload Semua ({len(_pl8_lain_valid)} paket)",
                        key="pl8_lain_all", use_container_width=True,
                    ):
                        with st.status(f"Upload {len(_pl8_lain_valid)} paket...", expanded=True) as _st8l:
                            for _pbl in _pl8_lain_valid:
                                _st8l.write(f"⏳ {_pbl['kode_paket']}...")
                                _rbl = _ba_pl_engine5.upload_ba_pl(
                                    id_nontender=_pbl["_id8l"], jenis_key=_jenis_lain8,
                                    nomor_ba=_nomor_lain8,
                                    tanggal_ba=_tgl_lain8.strftime("%d-%m-%Y"),
                                    info=_info_lain8,
                                    file_bytes=_pbl["_fl8"].read(), file_name=_pbl["_fl8"].name,
                                    tempat=_tempat_lain8,
                                )
                                if _rbl.get("ok"):
                                    _st8l.write(f"  ✅ {_pbl['kode_paket']} — berhasil")
                                else:
                                    _st8l.write(f"  ❌ {_pbl['kode_paket']} — status {_rbl.get('status')}")


        # ── Tools: Gabung Semua BA PL JKK (Bulk Print) ────────────────────────
        st.divider()
        st.markdown("### 🖨️ Gabung Semua BA PL JKK")
        st.caption(
            "Scan semua folder paket di **@ Pengadaan Langsung JKK** → ambil file `BA_PLJKK_*.pdf` "
            "tiap folder → gabung jadi satu file untuk print sekali."
        )

        import os as _ba_os
        import re as _ba_re

        # Root folder PL JKK (parent dari semua folder paket)
        from config import OUTPUT_DIR_PL_JKK as _BA_PL_JKK_ROOT
        _ba_root = _BA_PL_JKK_ROOT
        _ba_root_in = st.text_input(
            "Folder root PL JKK:",
            value=_ba_root,
            key="ba_bulk_root",
        )

        def _ba_nomor_folder(nama):
            """Ambil nomor urut dari nama folder (mis. '10. PLJKK...' → 10). Tanpa nomor → 9999."""
            m = _ba_re.match(r"\s*(\d+)", nama)
            return int(m.group(1)) if m else 9999

        # Hanya paket yang belum selesai (reuse _pl8_rows kalau ada, fallback ke semua)
        _ba_kode_aktif = {r.get("kode_paket", "") for r in locals().get("_pl8_rows", []) if r.get("kode_paket")}

        _ba_found = []  # (nomor, nama_folder, path_pdf)
        if _ba_os.path.isdir(_ba_root_in):
            for _entry in _ba_os.listdir(_ba_root_in):
                _sub = _ba_os.path.join(_ba_root_in, _entry)
                if not _ba_os.path.isdir(_sub):
                    continue
                # Skip folder paket yang sudah selesai (kode_paket tidak ada di set aktif)
                if _ba_kode_aktif and not any(k in _entry for k in _ba_kode_aktif):
                    continue
                # Match BA_PLJKK_*.pdf (bukan BA_REVIU_* / BA lain)
                _matches = [
                    f for f in _ba_os.listdir(_sub)
                    if _ba_re.match(r"(?i)^BA_PLJKK_.*\.pdf$", f)
                ]
                for _mf in _matches:
                    _ba_found.append((
                        _ba_nomor_folder(_entry), _entry,
                        _ba_os.path.join(_sub, _mf),
                    ))
            _ba_found.sort(key=lambda x: (x[0], x[1]))
        else:
            st.error(f"Folder tidak ditemukan: {_ba_root_in}")

        if _ba_found:
            st.success(f"✅ Ditemukan {len(_ba_found)} file BA_PLJKK (urut nomor paket):")
            _ba_prev = [
                {"No": n, "Folder": fol, "File": _ba_os.path.basename(p)}
                for n, fol, p in _ba_found
            ]
            st.dataframe(_ba_prev, use_container_width=True, hide_index=True)

            _ba_out_nama = st.text_input(
                "Nama file gabungan:",
                value="_KUMPULAN_BA_PL_JKK.pdf",
                key="ba_bulk_out_nama",
            )

            if st.button("🖨️ Gabung Semua BA", type="primary", key="ba_bulk_run"):
                from pypdf import PdfReader as _BaReader, PdfWriter as _BaWriter

                _writer = _BaWriter()
                _ok, _err = 0, []
                with st.status("Menggabungkan BA...", expanded=True) as _ba_status:
                    for _n, _fol, _path in _ba_found:
                        try:
                            _rdr = _BaReader(_path)
                            for _pg in _rdr.pages:
                                _writer.add_page(_pg)
                            st.write(f"✅ Paket {_n} — {len(_rdr.pages)} hlm")
                            _ok += 1
                        except Exception as _ex:
                            st.write(f"❌ Paket {_n} ({_fol}): {_ex}")
                            _err.append(_fol)

                    if len(_writer.pages) == 0:
                        _ba_status.update(label="❌ Tidak ada halaman tergabung.", state="error")
                    else:
                        # Tulis ke root folder; fallback suffix jika file terkunci (Nitro/dll)
                        _out_nama = _ba_out_nama.strip() or "_KUMPULAN_BA_PL_JKK.pdf"
                        if not _out_nama.lower().endswith(".pdf"):
                            _out_nama += ".pdf"
                        _base, _ext = _ba_os.path.splitext(_out_nama)
                        _target = _ba_os.path.join(_ba_root_in, _out_nama)
                        _written = None
                        for _attempt in range(6):
                            _try_path = _target if _attempt == 0 else \
                                _ba_os.path.join(_ba_root_in, f"{_base}_v{_attempt + 1}{_ext}")
                            try:
                                with open(_try_path, "wb") as _fh:
                                    _writer.write(_fh)
                                _written = _try_path
                                break
                            except PermissionError:
                                continue
                        if _written:
                            _ba_status.update(
                                label=f"✅ {_ok} BA digabung → {len(_writer.pages)} halaman.",
                                state="complete",
                            )
                            st.success(f"📁 Tersimpan: `{_written}`")
                            if _err:
                                st.warning(f"⚠️ Gagal: {', '.join(_err)}")
                        else:
                            _ba_status.update(
                                label="❌ Gagal tulis (file terkunci). Tutup PDF viewer dulu.",
                                state="error",
                            )
        elif _ba_os.path.isdir(_ba_root_in):
            _ba_nama_paket_aktif = ", ".join(
                _pl_label(r)
                for r in _pl8_rows
            )
            st.info(f"Tidak ada file `BA_PLJKK_*.pdf` ditemukan. Paket yang di-scan: {_ba_nama_paket_aktif or '(kosong)'}.")

    # ── Tab 5: Download Dok Kualifikasi PL ───────────────────────────────────
    if _pl_active_tab == "6️⃣ Download Kualifikasi":
        _ke_pl = _ke_pl_jkk  # alias — sudah di-import top-level
        _he_pl = _he_pl_jkk

        st.markdown("## Download Dokumen Kualifikasi — Pengadaan Langsung")
        st.caption("Download dok kualifikasi peserta dari SPSE + populate sheet Hasil Evaluasi di BAPLJKK.")

        # Cache paket list di session state — hindari query Supabase tiap render
        if "pl7_rows" not in st.session_state:
            try:
                _raw7 = _load_draft_pl_cached()
                _raw7, _ = pl_engine.buang_duplikat_paket_lama(_raw7)
                _raw7 = [r for r in _raw7 if not pl_engine.is_paket_selesai(r)]
                st.session_state["pl7_rows"] = _raw7
            except Exception as _e7:
                st.session_state["pl7_rows"] = []
                st.error(f"Gagal load paket PL: {_e7}")
        _pl7_rows = st.session_state["pl7_rows"]

        if not _pl7_rows:
            st.info("Tidak ada paket PL di database.")
            if st.button("🔄 Reload", key="pl7_reload"):
                del st.session_state["pl7_rows"]
                st.rerun()
        else:
            _pl7c1, _pl7c2 = st.columns([1, 1])

            with _pl7c1:
                st.markdown("#### Pilih Paket")
                _pl7_selected_rows = render_package_selection(st, _pl7_rows, _pl_label, prefix="pl7")

            with _pl7c2:
                st.markdown("#### Aksi")

                _n_paket7 = len(_pl7_selected_rows)

                if not _pl7_selected_rows:
                    st.info("Centang minimal 1 paket di kiri.")
                else:
                    st.markdown(f"**{_n_paket7} paket** dipilih")
                    st.caption("Peserta akan di-fetch via CDP saat tombol Jalankan diklik.")

                    _do_download7 = st.checkbox("⬇️ Download dokumen kualifikasi", value=True, key="pl7_do_dl")
                    _do_parse7    = st.checkbox("📋 Parse & populate sheet Hasil Evaluasi", value=True, key="pl7_do_parse")

                    _btn7 = st.button(
                        f"▶ Jalankan — {_n_paket7} paket",
                        type="primary", key="pl7_run", use_container_width=True,
                    )

                    if _btn7:
                        _pb7 = st.progress(0.0, text="Memulai...")
                        _log7_lines = []  # akumulasi log — update sekali per paket, bukan per baris
                        _ringkasan7: list = []  # kumpul status tiap paket untuk ringkasan akhir

                        def _flush7(container):
                            """Render log terakumulasi ke container — 1 update per paket."""
                            container.code("\n".join(_log7_lines[-60:]))  # max 60 baris terakhir

                        for _i7, _rpl7 in enumerate(_pl7_selected_rows):
                            _kpl7    = _rpl7["kode_paket"]
                            _nama7   = _pl_label(_rpl7)
                            _status7 = st.status(f"Paket {_i7+1}/{_n_paket7} — {_nama7}", expanded=True)

                            with _status7:
                                _log7_lines.clear()
                                _log7_box = st.empty()

                                def _lcb7(msg, _box=_log7_box, _lines=_log7_lines):
                                    _lines.append(msg)
                                    _box.code("\n".join(_lines[-40:]))

                                # Fetch peserta
                                _lcb7(f"[{_i7+1}/{_n_paket7}] Fetch peserta SPSE...")
                                _pb7.progress((_i7) / _n_paket7, text=f"{_nama7} — fetch peserta")
                                _fp7 = _ke_pl.fetch_peserta_pl(_kpl7)
                                if not _fp7.get("ok"):
                                    _lcb7(f"[SKIP] Peserta: {_fp7['pesan']}")
                                    _status7.update(label=f"SKIP {_nama7} — {_fp7['pesan']}", state="error", expanded=False)
                                    _ringkasan7.append({"nama": _nama7, "status": "skip", "detail": _fp7["pesan"]})
                                    continue
                                _peserta7 = _fp7["peserta"]
                                _lcb7(f"Peserta ({len(_peserta7)}): {', '.join(p['nama'] for p in _peserta7)}")

                                # Resolve folder
                                _folder7 = _ke_pl.resolve_folder_paket_pl(_kpl7)
                                if not _folder7.get("ok"):
                                    _lcb7(f"[SKIP] Folder: {_folder7['pesan']}")
                                    _status7.update(label=f"SKIP {_nama7} — folder tidak ditemukan", state="error", expanded=False)
                                    _ringkasan7.append({"nama": _nama7, "status": "skip", "detail": _folder7.get("pesan", "folder tidak ditemukan")})
                                    continue
                                _folder_kual7 = _folder7["path"]

                                # Download kualifikasi
                                if _do_download7:
                                    _pb7.progress((_i7 + 0.3) / _n_paket7, text=f"{_nama7} — download kualifikasi")
                                    for _ui7, _p7 in enumerate(_peserta7, 1):
                                        _lcb7(f"--- Download [{_ui7}/{len(_peserta7)}] {_p7['nama']} ---")
                                        _ke_pl.download_kualifikasi_peserta_pl(
                                            _p7, _folder_kual7, _ui7, len(_peserta7), _lcb7,
                                        )

                                    # Serap penyedia (dipindah dari create folder — di sini peserta sudah
                                    # terdaftar & Draft_PL/ND.pdf ada, jadi parse berhasil bukan timeout).
                                    _lcb7("--- Serap penyedia (nama/NPWP/personil) ---")
                                    try:
                                        import parse_kak_pl as _pkpl7
                                        _sp7 = _pkpl7.serap_penyedia_pl(kode_paket_filter=_kpl7)
                                        _lcb7(f"👤 Penyedia: {_sp7.get('updated',0)} diperbarui"
                                              if _sp7.get("updated", 0) > 0 else "👤 Penyedia: tidak ada data baru")
                                    except Exception as _sp7_e:
                                        _lcb7(f"⚠ Serap penyedia: {_sp7_e}")

                                # Parse evaluasi
                                if _do_parse7:
                                    _pb7.progress((_i7 + 0.7) / _n_paket7, text=f"{_nama7} — parse evaluasi")
                                    _lcb7("--- Populate sheet Hasil Evaluasi ---")
                                    _hasil7 = _he_pl.populate_hasil_evaluasi_pl(_kpl7, _peserta7, _lcb7)
                                    _lcb7(f"{'[OK]' if _hasil7.get('ok') else '[GAGAL]'} {_hasil7['pesan']}")

                                    # Refresh @ Master Data agar tgl_pembukaan benar
                                    # (penting untuk paket ulang: kode_paket baru → tanggal baru dari Supabase)
                                    if _hasil7.get("ok"):
                                        _lcb7("--- Refresh @ Master Data ---")
                                        try:
                                            import isi_master_data_pl as _imd7
                                            _xlsm7 = _he_pl._find_xlsm(_kpl7)
                                            if _xlsm7:
                                                _md7 = _imd7.isi_master_data_pl(_kpl7, _xlsm7, progress_cb=_lcb7)
                                                _lcb7(f"{'[OK]' if _md7.get('ok') else '[WARN]'} {_md7['pesan']}")
                                            else:
                                                _lcb7("[WARN] File .xlsm tidak ditemukan untuk refresh @ Master Data")
                                        except Exception as _e_md7:
                                            _lcb7(f"[WARN] Refresh @ Master Data gagal: {_e_md7}")

                                    _ringkasan7.append({
                                        "nama"  : _nama7,
                                        "status": "ok" if _hasil7.get("ok") else "gagal",
                                        "detail": _hasil7.get("pesan", ""),
                                    })
                                else:
                                    # Tidak ada parse → anggap OK (hanya download)
                                    _ringkasan7.append({"nama": _nama7, "status": "ok", "detail": "download saja"})

                                _status7.update(label=f"Selesai — {_nama7}", state="complete", expanded=False)

                            _pb7.progress((_i7 + 1) / _n_paket7, text=f"Selesai {_i7+1}/{_n_paket7} paket")

                        _pb7.progress(1.0, text="Semua paket selesai.")
                        from batch_summary import render_ringkasan_batch as _rrb7
                        _rrb7(st, _ringkasan7)

    # ── Tab 6: Evaluasi SPSE + Download Teknis/Biaya ─────────────────────────
    if _pl_active_tab == "7️⃣ Evaluasi & Teknis/Biaya":
        import evaluasi_admin_kualifikasi_pl as _eval_pl
        import dokumen_teknis_biaya_pl as _dtb_pl
        import penawaran_pl_engine as _penawaran_pl

        st.markdown("## Evaluasi SPSE & Download Teknis/Biaya — Pengadaan Langsung")
        st.caption("Submit evaluasi Admin+Kualifikasi LULUS di SPSE, lalu download dokumen teknis/biaya peserta.")

        # Selalu normalisasi daftar di sini. Tab Evaluasi dapat dibuka langsung,
        # tanpa lebih dulu merender Tab Download Kualifikasi.
        _raw8 = _load_draft_pl_cached()
        _raw8, _pl8_dup_n = pl_engine.buang_duplikat_paket_lama(_raw8)
        _pl8_rows = [r for r in _raw8 if not pl_engine.is_paket_selesai(r)]
        if _pl8_dup_n:
            st.caption(f"♻️ {_pl8_dup_n} row lama duplikat disembunyikan otomatis.")

        if not _pl8_rows:
            st.info("Tidak ada paket PL. Reload di Tab 7.")
        else:
            _pl8c1, _pl8c2 = st.columns([1, 1])

            with _pl8c1:
                st.markdown("#### Pilih Paket")
                if "pl8_checked" not in st.session_state:
                    st.session_state["pl8_checked"] = {}

                _pl8_kodes = [r["kode_paket"] for r in _pl8_rows]
                if st.session_state.get("pl8_selection_default_v19") != True:
                    for _k in _pl8_kodes:
                        st.session_state.pop(f"pl8_chk_{_k}_v19", None)
                    st.session_state["pl8_checked"].clear()
                    st.session_state["pl8_selection_default_v19"] = True
                for _k in _pl8_kodes:
                    if f"pl8_chk_{_k}_v19" not in st.session_state:
                        st.session_state[f"pl8_chk_{_k}_v19"] = True
                _pl8bc1, _pl8bc2 = st.columns(2)
                if _pl8bc1.button("✅ Pilih Semua", key="pl8_select_all", use_container_width=True):
                    for _k in _pl8_kodes:
                        st.session_state[f"pl8_chk_{_k}_v19"] = True
                        st.session_state["pl8_checked"][_k] = True
                if _pl8bc2.button("❌ Batal Semua", key="pl8_deselect_all", use_container_width=True):
                    for _k in _pl8_kodes:
                        st.session_state[f"pl8_chk_{_k}_v19"] = False
                        st.session_state["pl8_checked"][_k] = False

                for _rpl8 in _pl8_rows:
                    _kpl8 = _rpl8["kode_paket"]
                    _chk8 = st.checkbox(
                        "",
                        key=f"pl8_chk_{_kpl8}_v19",
                        label_visibility="collapsed",
                    )
                    st.markdown(f"**{_pl_label(_rpl8)}**")
                    st.session_state["pl8_checked"][_kpl8] = _chk8

            with _pl8c2:
                st.markdown("#### Aksi")

                _pl8_selected_rows = [r for r in _pl8_rows if st.session_state["pl8_checked"].get(r["kode_paket"])]
                _n_paket8 = len(_pl8_selected_rows)

                if not _pl8_selected_rows:
                    st.info("Centang minimal 1 paket di kiri.")
                else:
                    st.markdown(f"**{_n_paket8} paket** dipilih")
                    st.caption("Peserta di-scrape dari SPSE saat Jalankan.")
                    _do_eval_admin = st.checkbox("⚖️ Submit evaluasi Admin + Kualifikasi LULUS di SPSE", value=True, key="pl8_do_eval_admin")
                    _do_eval_teknis = st.checkbox("⚙️ Submit evaluasi Teknis LULUS di SPSE", value=True, key="pl8_do_eval_teknis")
                    _do_eval_harga = st.checkbox("💰 Submit evaluasi Harga LULUS di SPSE", value=True, key="pl8_do_eval_harga")

                    _do_tekbio8  = st.checkbox("⬇️ Download dokumen teknis/biaya + gabung PDF", value=True, key="pl8_do_tekbio")
                    _do_penawaran8 = st.checkbox("📊 Tulis rincian penawaran ke sheet '6. Penawaran' Excel", value=True, key="pl8_do_penawaran")

                    st.warning("Evaluasi LULUS bersifat **permanen** — modifikasi data SPSE production.")
                    _konfirmasi8 = st.checkbox(
                        "Saya paham tindakan ini tidak bisa dibatalkan.",
                        value=False, key="pl8_konfirmasi",
                    )
                    _btn8_disabled = (_do_eval_admin or _do_eval_teknis or _do_eval_harga) and not _konfirmasi8
                    if _btn8_disabled:
                        st.info("Centang konfirmasi untuk mengaktifkan tombol.")
                    _btn8 = st.button(
                        f"▶ Jalankan — {_n_paket8} paket",
                        type="primary", key="pl8_run", use_container_width=True,
                        disabled=_btn8_disabled,
                    )

                    st.divider()
                    st.markdown("#### 🤖 Evaluasi AI")
                    st.caption("AI baca dokumen di folder paket → output `.md`. Paralel per paket.")
                    _do_extract_teks8 = st.checkbox("📝 Extract teks kualifikasi (.txt) dari folder '8. Dokumen Kualifikasi' — hemat token sebelum evaluasi AI", value=False, key="pl8_do_extract_teks")
                    _ai_eval_engine = "codex"
                    _ai_eval_model = None
                    st.caption("Engine: Codex CLI · Model terkunci: gpt-5.6-luna · Reasoning: medium")
                    _do_ai_kualifikasi = st.checkbox("⚖️ Evaluasi Admin+Kualifikasi (Sesi 1) via AI", value=True, key="pl8_do_ai_kual")
                    _do_ai_teknis = st.checkbox("🔬 Evaluasi Teknis (Sesi 2) via AI", value=True, key="pl8_do_ai_teknis")
                    _do_ai_biaya = st.checkbox("💰 Evaluasi Biaya (Sesi 3) via AI", value=True, key="pl8_do_ai_biaya")
                    _btn_ai_eval = st.button(
                        f"🤖 Jalankan Evaluasi AI — {_n_paket8} paket",
                        key="pl8_btn_ai_eval", use_container_width=True,
                        disabled=not (_do_ai_kualifikasi or _do_ai_teknis or _do_ai_biaya),
                    )
                    if _btn_ai_eval and (_do_ai_kualifikasi or _do_ai_teknis or _do_ai_biaya):
                        import ai_evaluator as _heval8
                        if _do_extract_teks8:
                            import extract_teks_kualifikasi as _etk8
                            st.info("📝 Extract teks kualifikasi...")
                            for _r8 in _pl8_selected_rows:
                                try:
                                    _folder8 = _heval8._folder_paket(_r8.get("nomor_urut"), _r8.get("nama_paket", ""), jenis_pl="JKK", is_ulang=bool(_r8.get("is_ulang")))
                                    if not _folder8:
                                        continue
                                    _etk_folder8 = os.path.join(str(_folder8), "8. Dokumen Kualifikasi")
                                    if os.path.isdir(_etk_folder8):
                                        _etk8.extract_folder_kualifikasi(_etk_folder8)
                                except Exception as _etk8_e:
                                    st.warning(f"⚠ Extract teks {_pl_label(_r8)}: {_etk8_e}")
                        _ai_jobs = [{"nomor_urut": r.get("nomor_urut"), "nama_paket": r.get("nama_paket",""), "is_ulang": bool(r.get("is_ulang"))} for r in _pl8_selected_rows]
                        _ai_label_map = {r.get("nama_paket", ""): _pl_label(r) for r in _pl8_selected_rows}
                        if _do_ai_kualifikasi:
                            st.info("⚖️ Menjalankan evaluasi Admin+Kualifikasi...")
                            _res_kual = _heval8.evaluasi_bulk(_ai_jobs, jenis="kualifikasi", model=_ai_eval_model, max_workers=3, engine=_ai_eval_engine)
                            for _rk in _res_kual:
                                _rk_name = _ai_label_map.get(_rk.get("nama"), _rk.get("nama", "-"))
                                if _rk["status"] == "ok":
                                    st.success(f"✅ {_rk_name[:50]}")
                                    with st.expander(f"Output kualifikasi: {_rk_name[:35]}"):
                                        st.markdown(_rk["output"][:3000])
                                else:
                                    st.error(f"❌ {_rk_name[:50]} — {_rk['error'][:200]}")
                        if _do_ai_teknis:
                            st.info("🔬 Menjalankan evaluasi Teknis...")
                            _res_teknis = _heval8.evaluasi_bulk(_ai_jobs, jenis="teknis", model=_ai_eval_model, max_workers=3, engine=_ai_eval_engine)
                            for _rt in _res_teknis:
                                _rt_name = _ai_label_map.get(_rt.get("nama"), _rt.get("nama", "-"))
                                if _rt["status"] == "ok":
                                    st.success(f"✅ {_rt_name[:50]}")
                                    with st.expander(f"Output teknis: {_rt_name[:35]}"):
                                        st.markdown(_rt["output"][:3000])
                                else:
                                    st.error(f"❌ {_rt_name[:50]} — {_rt['error'][:200]}")
                        if _do_ai_biaya:
                            _ai_jobs_biaya = _ai_jobs
                            if _do_ai_teknis:
                                _teknis_ok = {
                                    _r.get("nama")
                                    for _r in _res_teknis
                                    if _r.get("status") == "ok"
                                }
                                _ai_jobs_biaya = [
                                    _j for _j in _ai_jobs
                                    if _j.get("nama_paket") in _teknis_ok
                                ]
                                _n_skip_biaya = len(_ai_jobs) - len(_ai_jobs_biaya)
                                if _n_skip_biaya:
                                    st.warning(
                                        f"⏭️ {_n_skip_biaya} paket tidak masuk evaluasi biaya "
                                        "karena Sesi 2 gagal pada proses ini."
                                    )
                            if _ai_jobs_biaya:
                                st.info("💰 Menjalankan evaluasi Biaya...")
                                _res_biaya = _heval8.evaluasi_bulk(
                                    _ai_jobs_biaya, jenis="biaya",
                                    model=_ai_eval_model, max_workers=3,
                                    engine=_ai_eval_engine,
                                )
                                for _rb in _res_biaya:
                                    _rb_name = _ai_label_map.get(
                                        _rb.get("nama"), _rb.get("nama", "-")
                                    )
                                    if _rb["status"] == "ok":
                                        st.success(f"✅ {_rb_name[:50]}")
                                        with st.expander(
                                            f"Output biaya: {_rb_name[:35]}"
                                        ):
                                            st.markdown(_rb["output"][:3000])
                                    else:
                                        st.error(
                                            f"❌ {_rb_name[:50]} — "
                                            f"{_rb['error'][:200]}"
                                        )


                    if _btn8:
                        _pb8 = st.progress(0.0, text="Memulai...")
                        _ringkasan8: list = []  # kumpul status tiap paket untuk ringkasan akhir

                        for _i8, _rpl8 in enumerate(_pl8_selected_rows):
                            _kpl8  = _rpl8["kode_paket"]
                            _nama8 = _pl_label(_rpl8)
                            _status8 = st.status(f"Paket {_i8+1}/{_n_paket8} — {_nama8}", expanded=True)

                            with _status8:
                                _log8_lines = []
                                _log8_box   = st.empty()

                                def _lcb8(msg, _box=_log8_box, _lines=_log8_lines):
                                    _lines.append(msg)
                                    _box.code("\n".join(_lines[-40:]))

                                # Scrape id_nontender per peserta
                                _lcb8("Scrape peserta evaluasi dari SPSE...")
                                _pb8.progress(_i8 / _n_paket8, text=f"{_nama8} — scrape peserta")
                                _res_peserta8 = _eval_pl.scrape_peserta_evaluasi(_kpl8)
                                if not _res_peserta8.get("ok"):
                                    _lcb8(f"[SKIP] {_res_peserta8['pesan']}")
                                    _status8.update(label=f"SKIP {_nama8} — {_res_peserta8['pesan']}", state="error", expanded=False)
                                    _ringkasan8.append({"nama": _nama8, "status": "skip", "detail": _res_peserta8["pesan"]})
                                    continue
                                _peserta8 = _res_peserta8["peserta"]
                                _lcb8(f"Peserta ({len(_peserta8)}): {', '.join(p['nama'] for p in _peserta8)}")

                                # Resolve folder paket root
                                _folder8      = _ke_pl.resolve_folder_paket_pl(_kpl8)
                                _folder_paket8 = _folder8.get("pesan", "") if _folder8.get("ok") else ""

                                # Evaluasi LULUS
                                if _do_eval_admin or _do_eval_teknis or _do_eval_harga:
                                    _pb8.progress((_i8 + 0.3) / _n_paket8, text=f"{_nama8} — submit evaluasi")
                                    _lcb8(f"--- Submit evaluasi LULUS (Admin:{_do_eval_admin}, Teknis:{_do_eval_teknis}, Harga:{_do_eval_harga}) ---")
                                    _eval8 = _eval_pl.evaluasi_batch_lulus(
                                        _kpl8,
                                        admin=_do_eval_admin,
                                        kualifikasi=_do_eval_admin,
                                        teknis=_do_eval_teknis,
                                        harga=_do_eval_harga,
                                        progress_cb=_lcb8
                                    )
                                    _lcb8(f"{'[OK]' if _eval8.get('ok') else '[SEBAGIAN GAGAL]'} {_eval8['ringkasan']}")

                                # Download teknis/biaya
                                if _do_tekbio8:
                                    if not _folder_paket8:
                                        _lcb8("[SKIP] Folder paket tidak ditemukan")
                                    else:
                                        _pb8.progress((_i8 + 0.6) / _n_paket8, text=f"{_nama8} — download teknis/biaya")
                                        for _ui8, _ep8 in enumerate(_peserta8, 1):
                                            _lcb8(f"--- Download [{_ui8}/{len(_peserta8)}] {_ep8['nama']} ---")
                                            _res_tb8 = _dtb_pl.download_teknis_biaya_peserta(
                                                id_nontender=_ep8["id_nontender"],
                                                nama_peserta=_ep8["nama"],
                                                folder_paket=_folder_paket8,
                                                urutan=_ui8,
                                                progress_cb=_lcb8,
                                            )
                                            _lcb8(f"{'[OK]' if _res_tb8['ok'] else '[GAGAL]'} {_res_tb8['pesan']}")

                                # Tulis penawaran ke sheet 6. Penawaran
                                if _do_penawaran8:
                                    if not _folder_paket8:
                                        _lcb8("[SKIP] Folder paket tidak ditemukan, skip penawaran")
                                    else:
                                        _pb8.progress((_i8 + 0.8) / _n_paket8, text=f"{_nama8} — tulis penawaran")
                                        _lcb8("--- Tulis rincian penawaran ke Excel ---")
                                        for _ep8_p in _peserta8:
                                            _lcb8(f"  Peserta: {_ep8_p['nama']}")
                                            _res_pnw8 = _penawaran_pl.tulis_penawaran_ke_excel(
                                                folder_paket=_folder_paket8,
                                                id_nontender=_ep8_p["id_nontender"],
                                                progress_cb=_lcb8,
                                            )
                                            _lcb8(f"  {'[OK]' if _res_pnw8['ok'] else '[GAGAL]'} {_res_pnw8['pesan']}" +
                                                  (f" — Total Rp {_res_pnw8['total_penawaran']:,.0f}" if _res_pnw8.get('total_penawaran') else ""))

                                _ringkasan8.append({"nama": _nama8, "status": "ok", "detail": ""})
                                _status8.update(label=f"Selesai — {_nama8}", state="complete", expanded=False)

                            _pb8.progress((_i8 + 1) / _n_paket8, text=f"Selesai {_i8+1}/{_n_paket8} paket")

                        _pb8.progress(1.0, text="Semua paket selesai.")
                        from batch_summary import render_ringkasan_batch as _rrb8
                        _rrb8(st, _ringkasan8)

    if _pl_active_tab == "🔟 Penetapan Pemenang":
        _raw10 = _load_draft_pl_cached()
        _raw10, _ = pl_engine.buang_duplikat_paket_lama(_raw10)
        _rows10 = [r for r in _raw10 if not pl_engine.is_paket_selesai(r)]
        _render_tab10_pl(_rows10, "pl10")

    st.stop()  # Jangan render tab Tender jika mode PL

if st.session_state["app_mode"] == "E-Katalog - Survei Pasar":
    import ekatalog_engine
    st.markdown("### 🛍️ E-Katalog — Survei Pasar Inaproc")
    ekatalog_engine.render_survei_pasar()
    st.stop()

if st.session_state["app_mode"] == "PL - Konstruksi":
    # ============================================================
    # MODE: PENGADAAN LANGSUNG — PEKERJAAN KONSTRUKSI (PL PK)
    # ============================================================

    if st.session_state.get("spse_role") != "PP":
        st.warning("⚠️ Browser belum login sebagai PP. Login via sidebar terlebih dahulu.")
        st.stop()

    # Rebind engine PK-specific ke varian _plpk (scope module, mode PK only)
    pl_engine = get_plpk_engine()
    import pl_engine as _pl_engine_utils
    _pl_active_tab = st.radio("Tab PL", PLPK_TAB_LABELS, horizontal=True, key="pl_active_tab_pk")
    _is_plpk_tab6 = _pl_active_tab == "6️⃣ Download Kualifikasi"

    if _pl_active_tab == "📄 Import DPA":
        _render_tab_dpa()

    # ── Tab 1: Draft Paket PL (PK) ───────────────────────────────────────────
    if _pl_active_tab == "1️⃣ Draft Paket PL":
        import os as _pl_os, subprocess as _pl_sp
        from config import POKJA_ROOT as _PL_POKJA_ROOT, OUTPUT_DIR_PL_JKK as _PL_DIR_JKK, OUTPUT_DIR_PL_PK as _PL_DIR_PK, V19_ROOT as _V19_ROOT, POKJA_PYTHON as _POKJA_PYTHON
        _TEMPLATE_DIR_PL    = str(__import__("pathlib").Path(_PL_POKJA_ROOT) / "Paket Experiment - Pengadaan Langsung" / "Development - PL - JKK")
        _TEMPLATE_DIR_PL_PK = str(__import__("pathlib").Path(_PL_POKJA_ROOT) / "Paket Experiment - Pengadaan Langsung" / "Development - PL - PK")

        _PL_PY     = _POKJA_PYTHON
        _PL_SCRIPT = str(pathlib.Path(_V19_ROOT) / "setup_paket_baru.py")
        _PL_NO_WIN = 0x08000000

        _pl_rows = _load_draft_pl_cached("PK")

        # ── Buang duplikat row lama (paket di-ulang → kode baru, row lama nyangkut) ──
        _pl_rows, _pl_dup_n = pl_engine.buang_duplikat_paket_lama(_pl_rows)
        if _pl_dup_n:
            st.caption(f"♻️ {_pl_dup_n} row lama duplikat (paket ulang) disembunyikan otomatis.")

        # ── #4: Filter paket selesai (penandatanganan kontrak) ──────────────────
        _pl_show_done = st.checkbox(
            "Tampilkan paket selesai (sudah teken kontrak)",
            value=False,
            key="pl_show_done",
        )
        _pl_done_n = sum(1 for r in _pl_rows if pl_engine.is_paket_selesai(r))
        if not _pl_show_done:
            _pl_rows = [r for r in _pl_rows if not pl_engine.is_paket_selesai(r)]
            if _pl_done_n:
                st.caption(f"🔒 {_pl_done_n} paket selesai (Penandatanganan Kontrak) disembunyikan — centang di atas untuk tampilkan.")

        _pl_col_kiri, _pl_col_kanan = st.columns(2)

        # ══════════════════════════════════════════════════════
        # KOLOM KIRI — Serap Data + Daftar Paket
        # ══════════════════════════════════════════════════════
        with _pl_col_kiri:
            # ── #1: Gabung Serap SPSE + MAK + HPS ──────────────────────────────
            st.markdown("#### 1. Serap Data Paket PL")
            st.caption("Pilih aksi lalu klik tombol — aksi berjalan berurutan sesuai centang.")
            _cb_serap_spse = st.checkbox("Serap dari SPSE (daftar paket + status)", value=True, key="pl_cb_serap_spse")
            _cb_serap_mak  = st.checkbox("Serap MAK dari Inbox PL",               value=True, key="pl_cb_serap_mak")

            if st.button("🚀 Serap Data Paket PL", type="primary", use_container_width=True, key="btn_serap_pl_gabung"):
                # Aksi 1: Serap dari SPSE
                if _cb_serap_spse:
                    import spse_browser as _sb_pl
                    _pl_cookie = _sb_pl.get_spse_cookies()
                    if not _pl_cookie:
                        st.error("Cookie SPSE kosong — buka Brave SPSE dan login sebagai PP.")
                    else:
                        _pl_pb = st.progress(0.0)
                        _pl_st = st.empty()
                        _pl_logs = []
                        def _pl_log(msg):
                            _pl_logs.append(msg)
                            _pl_st.info(msg)
                        from config import SPSE_BASE_URL as _SPSE_BASE
                        _pl_hasil = pl_engine.serap_paket_pl_dari_spse(
                            _pl_cookie, _SPSE_BASE, log_fn=_pl_log
                        )
                        _pl_pb.progress(1.0)
                        _pl_c1, _pl_c2 = st.columns(2)
                        _pl_c1.metric("✅ Tersimpan", _pl_hasil.get("scraped", 0))
                        _pl_c2.metric("❌ Error", len(_pl_hasil.get("errors", [])))
                        if _pl_hasil.get("errors"):
                            with st.expander("Detail Error SPSE"):
                                for _e in _pl_hasil["errors"]:
                                    st.error(_e)
                        with st.expander("📋 Log lengkap serap"):
                            st.text("\n".join(_pl_logs))
                        # Invalidasi cache agar hasil serap SPSE langsung terbaca.
                        _load_draft_pl_cached.clear()
                        # Reload setelah serap SPSE agar data paket terkini
                        _pl_rows = _load_draft_pl_cached("PK")
                        if not _pl_show_done:
                            _pl_rows = [r for r in _pl_rows if not pl_engine.is_paket_selesai(r)]

                # Aksi 2: Serap MAK dari Inbox
                if _cb_serap_mak:
                    import inbox_engine as _ibe
                    _pb_mak = st.progress(0.0)
                    _st_mak = st.empty()
                    _logs_mak = []
                    def _cb_mak(p, m):
                        _pb_mak.progress(min(max(p, 0.0), 1.0))
                        _logs_mak.append(m)
                        _st_mak.info(m)
                    try:
                        _r_mak = _ibe.serap_inbox_pl(progress_cb=_cb_mak)
                        _c1, _c2, _c3 = st.columns(3)
                        _c1.metric("Pesan parse", _r_mak.get("scraped", 0))
                        _c2.metric("Paket update", _r_mak.get("matched", 0))
                        _c3.metric("Error", len(_r_mak.get("errors", [])))
                        if _r_mak.get("errors"):
                            with st.expander("Detail Error MAK"):
                                for _e in _r_mak["errors"]:
                                    st.warning(_e)
                    except Exception as _e:
                        st.error(f"Gagal serap MAK: {_e}")

            st.divider()
            st.markdown("#### 2. Daftar Paket PL")

            _pl_filter = st.selectbox(
                "Filter:",
                ["Semua", "JKK", "PK", "Belum Folder", "Sudah Folder"],
                key="pl_filter_jenis",
            )

            def _pl_match(r):
                _jp = (r.get("jenis_pl") or "").upper()
                if _pl_filter == "JKK":    return _jp == "JKK"
                if _pl_filter == "PK":     return _jp == "PK"
                if _pl_filter == "Belum Folder": return not bool(r.get("folder_dibuat"))
                if _pl_filter == "Sudah Folder": return bool(r.get("folder_dibuat"))
                return True

            _pl_filtered = [r for r in _pl_rows if _pl_match(r)]

            _pl_peserta_map = _fetch_status_semua_paket_cached(
                tuple(r.get("kode_paket", "") for r in _pl_filtered if r.get("kode_paket"))
            )

            if not _pl_filtered:
                st.info("Belum ada paket PL. Klik 'Serap dari SPSE' atau tambah manual.")
            else:
                for _pr in _pl_filtered:
                    _pr_kode   = _pr.get("kode_paket", "")
                    _pr_nama   = _pr.get("nama_paket", "-")
                    _pr_jenis  = (_pr.get("jenis_pl") or "").upper()
                    _pr_hps    = _pr.get("nilai_hps", "-")
                    _pr_status = _pr.get("status", "draft")
                    _pr_folder = bool(_pr.get("folder_dibuat"))
                    _pr_icon   = "✅" if _pr_folder else "📋"
                    # Label ringkas: nomor urut folder, bukan nama metode panjang.
                    _pr_folder_raw = str(_pr.get("folder_dibuat") or "")
                    _pr_no_match = re.search(r"^\s*(\d+)", os.path.basename(_pr_folder_raw.rstrip("\\/")))
                    _pr_nomor = (_pr_no_match.group(1) if _pr_no_match else str(_pr.get("nomor_urut") or "").strip()) or _pr_kode
                    _pr_metode_raw = _pr.get("metode_pengadaan", "") or ""
                    _pr_metode_low = _pr_metode_raw.lower()
                    _pr_label  = f"{_pr_icon} {_pl_label(_pr)}"

                    with st.expander(_pr_label):
                        st.caption(f"`{_pr_kode}` | HPS: {_pr_hps} | Status: **{_pr_status}**")
                        # Badge peserta pendaftaran — pakai kode_paket (bukan id_nontender)
                        _pr_id_nt = _pr.get("kode_paket", "") or _pr.get("id_nontender", "")
                        if _pr_id_nt:
                            _pr_jml_peserta = _pl_peserta_map.get(_pr_id_nt, {}).get("jumlah", -1)
                            if _pr_jml_peserta > 0:
                                st.success(f"✅ {_pr_jml_peserta} peserta sudah mendaftar")
                            elif _pr_jml_peserta == 0:
                                st.warning("⚠️ Belum ada peserta mendaftar")
                        if "non konstruksi" in _pr_metode_low:
                            st.warning("⚠️ Metode: Non Konstruksi — minta PPK ubah ke Konstruksi di SPSE.")

                        # ── Info DPA (match nama pekerjaan fisik dari DPA yg sudah diimport di tab DPA) ──
                        import dpa_engine as _dpa_pl
                        if "pl_dpa_pool" not in st.session_state:
                            try:
                                from config import sb as _sb_dpa_pl
                                st.session_state["pl_dpa_pool"] = _dpa_pl.load_pool_dpa(_sb_dpa_pl())
                            except Exception:
                                st.session_state["pl_dpa_pool"] = []
                        _pl_dpa_key = f"pl_dpa_{_pr_kode}"
                        if _pl_dpa_key not in st.session_state:
                            st.session_state[_pl_dpa_key] = _dpa_pl.cari_dpa_di_pool(
                                _pr_nama, st.session_state["pl_dpa_pool"]
                            )
                        _pl_dpa_rows = st.session_state.get(_pl_dpa_key, [])
                        if _pl_dpa_rows:
                            def _rp_pl(v):
                                try: return f"Rp {int(v):,}".replace(",", ".")
                                except: return str(v) if v else "-"
                            _pl_dpa_nama = _pl_dpa_rows[0].get("nama_paket") or "-"
                            # Expander tidak bisa nested — pakai checkbox toggle
                            if st.checkbox(f"📊 Info DPA — {_pl_dpa_nama[:40]}", key=f"pl_dpa_show_{_pr_kode}"):
                                _pl_dpa_tbl = []
                                for _it in _pl_dpa_rows:
                                    _sk = (_it.get("subkegiatan_id") or "").split("|")
                                    _pl_dpa_tbl.append({
                                        "Uraian": _it.get("uraian") or "-",
                                        "Spesifikasi": _it.get("spesifikasi") or "-",
                                        "Sebelum": _rp_pl(_it.get("jumlah_sebelum")),
                                        "Sesudah": _rp_pl(_it.get("jumlah_sesudah")),
                                        "Sub Kegiatan": _sk[-1] if len(_sk) >= 3 else (_it.get("subkegiatan_id") or "-"),
                                    })
                                st.table(_pl_dpa_tbl)

                        # Ubah Metode Pengadaan inline per paket
                        with st.container(border=True):
                            st.markdown("**🔧 Ubah Metode Pengadaan**")
                            _pr_metode_pilihan = st.selectbox(
                                "Target metode:",
                                list(pl_engine.METODE_PL_MAP.keys()),
                                 index=list(pl_engine.METODE_PL_MAP.keys()).index("Pekerjaan Konstruksi — PL"),
                                key=f"pl_ubah_metode_target_{_pr_kode}",
                            )
                            _pr_kat_id, _pr_pilih_val = pl_engine.METODE_PL_MAP[_pr_metode_pilihan]
                            if st.button(
                                "🔄 Ubah Metode via CDP",
                                use_container_width=True,
                                key=f"pl_btn_ubah_metode_{_pr_kode}",
                            ):
                                _pl_base_ubah = pl_engine.BASE_URL + "/"
                                if pl_engine.ubah_metode_pl_playwright(_pr_kode, _pr_kat_id, _pr_pilih_val, _pl_base_ubah):
                                    st.success("✅ Berhasil ubah metode. Serap ulang untuk refresh.")
                                else:
                                    st.error("❌ Gagal ubah metode.")

                        # Aksi paket: tiga operasi utama satu baris; aksi panjang di baris kedua.
                        _pr_c1, _pr_c2, _pr_c3 = st.columns([1.5, 1.1, 1.1])
                        # Isi ulang workbook yang sudah ada tanpa membuat folder baru.
                        if _pr_folder and _pr_c1.button("📊 Isi Excel", key=f"pl_excel_{_pr_kode}", use_container_width=True):
                            import isi_master_data_pl as _imd_pr
                            import kualifikasi_engine_pl as _keng_pr_excel
                            # Status container tidak boleh berada di dalam expander paket.
                            _pr_excel_status = st.empty()
                            _pr_excel_status.info("📊 Menyiapkan isi ulang Excel...")
                            _pr_excel_log = st.empty()
                            _pr_excel_logs = []
                            def _pr_excel_cb(_msg):
                                _pr_excel_logs.append(str(_msg))
                                _pr_excel_log.code("\n".join(_pr_excel_logs[-20:]))
                                _pr_excel_status.info(f"📊 {str(_msg)[:90]}")
                            _pr_excel_cb("Mencari folder paket...")
                            _pr_excel_fr = _keng_pr_excel.resolve_folder_paket_pl(_pr_kode)
                            _pr_excel_root = _pr_excel_fr.get("pesan", "") if _pr_excel_fr.get("ok") else ""
                            _pr_excel_xlsm = _cari_xlsm_pl(_pr_excel_root) if _pr_excel_root else None
                            if not _pr_excel_xlsm:
                                _pr_excel_status.error("❌ Folder/workbook .xlsm tidak ditemukan.")
                            else:
                                _pr_excel_cb(f"Workbook: {_pr_excel_xlsm}")
                                _pr_excel_res = _imd_pr.isi_master_data_pl(_pr_kode, _pr_excel_xlsm, progress_cb=_pr_excel_cb)
                                if _pr_excel_res.get("ok"):
                                    _pr_excel_status.success("✅ Excel selesai diisi.")
                                    st.success(_pr_excel_res.get("pesan", "Excel selesai diisi."))
                                else:
                                    _pr_excel_status.error("❌ Gagal mengisi Excel.")
                                    st.error(_pr_excel_res.get("pesan", "Gagal mengisi Excel."))
                        if _pr_folder and _pr_c2.button("📦 Unduh", key=f"pl_dl_{_pr_kode}", use_container_width=True):
                            import kualifikasi_engine_pl as _keng_pr_dl
                            _pr_dl_fr = _keng_pr_dl.resolve_folder_paket_pl(_pr_kode)
                            _pr_dl_root = _pr_dl_fr.get("pesan", "") if _pr_dl_fr.get("ok") else ""
                            if not _pr_dl_root:
                                st.error("Folder tidak ditemukan.")
                            else:
                                _pr_dl_logs = []
                                _pr_dl_label = st.empty()
                                _pr_dl_area = st.empty()
                                _pr_dl_label.markdown("🔽 **Mengunduh...**")
                                def _pr_dl_cb(msg, _l=_pr_dl_logs, _a=_pr_dl_area, _b=_pr_dl_label):
                                    _l.append(msg); _a.code("\n".join(_l[-15:])); _b.markdown(f"🔽 **{msg[:50]}**")
                                pl_engine.buat_subfolder_dokumen(_pr_dl_root)
                                _pr_dl_res = pl_engine.download_dokumen_paket_pl(_pr_kode, _pr_dl_root, _pr_dl_cb, force_clean=True)
                                _pr_dl_label.markdown(f"✅ **{len(_pr_dl_res['ok'])} file, ❌ {len(_pr_dl_res['error'])} error**")
                                _pr_kak_p = parse_kak_pl.cari_kak_di_folder(_pr_dl_root)
                                if _pr_kak_p:
                                    _pr_kak_d = parse_kak_pl.parse_kak(_pr_kak_p)
                                    _pr_kak_u = {k: v for k, v in _pr_kak_d.items() if v}
                                    if _pr_kak_u:
                                        pl_engine.simpan_paket_pl({"kode_paket": _pr_kode, **_pr_kak_u})
                                        st.info(f"📋 KAK: {', '.join(_pr_kak_u.keys())}")
                                # Unduh hanya memperbarui dokumen SPSE dan metadata KAK.
                                # Pengisian workbook sengaja tetap eksklusif di tombol Isi Excel.
                                _load_draft_pl_cached.clear()
                                st.rerun()
                        if _pr_folder and _pr_c3.button("💰 HPS", key=f"pl_hps_{_pr_kode}", use_container_width=True):
                            import hps_engine as _hps_pr
                            import kualifikasi_engine_pl as _keng_pr
                            _pr_fr = _keng_pr.resolve_folder_paket_pl(_pr_kode)
                            _pr_root = _pr_fr.get("pesan", "") if _pr_fr.get("ok") else ""
                            _pr_xl = _cari_xlsm_pl(_pr_root) if _pr_root else None
                            if _pr_xl:
                                with st.spinner("Scrape HPS..."):
                                    _pr_hr = _hps_pr.scrape_hps_pl_ke_excel(_pr_kode, _pr_xl)
                                if _pr_hr.get("ok"):
                                    st.success(f"✅ {_pr_hr['count']} item HPS")
                                else:
                                    st.error(_pr_hr.get("pesan", "-"))
                            else:
                                st.error("Folder/xlsm tidak ditemukan.")
                        _pr_r1, _pr_r2 = st.columns([1, 1])
                        if _pr_folder and _pr_r2.button("🤖 Pra-Reviu + Isi DOCM", key=f"pl_pra_reviu_{_pr_kode}", use_container_width=True):
                            import ai_evaluator as _heval_one
                            with st.spinner("Menjalankan pra-reviu paket..."):
                                _one_res = _heval_one.evaluasi_bulk(
                                    [{"nomor_urut": _pr.get("nomor_urut"), "nama_paket": _pr_nama}],
                                    jenis="pra_reviu", model=None, max_workers=1, engine="codex",
                                )
                            if _one_res and _one_res[0].get("status") == "ok":
                                st.success(f"✅ Pra-reviu {_pr_nama[:50]} selesai.")
                            else:
                                st.error((_one_res[0].get("error") if _one_res else "Pra-reviu gagal")[:300])
                        if _pr_folder and _pr_r1.button("🔄 Reset status folder", key=f"pl_reset_{_pr_kode}", use_container_width=True):
                            from config import sb as _sb_reset_one
                            try:
                                _sb_reset_one().table("draft_paket_pl").update({"folder_dibuat": None}).eq("kode_paket", _pr_kode).execute()
                                _load_draft_pl_cached.clear()
                                st.success(f"✅ Status folder {_pr_nama[:45]} direset.")
                                st.rerun()
                            except Exception as _reset_one_e:
                                st.error(f"Reset gagal: {_reset_one_e}")

        # ══════════════════════════════════════════════════════
        # KOLOM KANAN — Buat Folder + Download Dokumen
        # ══════════════════════════════════════════════════════

            # ── Ubah Metode Bulk (semua paket di daftar) ──────────────
            if _pl_rows:
                st.divider()
                with st.container(border=True):
                    st.markdown("**🔧 Ubah Metode Pengadaan — Semua Paket**")
                    _pl_opsi_ubah_bulk = {r.get("nama_paket", r.get("kode_paket")): r.get("kode_paket") for r in _pl_filtered}
                    _pl_sel_ubah_bulk = list(_pl_opsi_ubah_bulk.keys())
                    _pl_metode_bulk = st.selectbox(
                        "Target metode:",
                        list(pl_engine.METODE_PL_MAP.keys()),
                         index=list(pl_engine.METODE_PL_MAP.keys()).index("Pekerjaan Konstruksi — PL"),
                        key="pk_ubah_metode_bulk_target",
                    )
                    _pl_kat_id_bulk, _pl_pilih_val_bulk = pl_engine.METODE_PL_MAP[_pl_metode_bulk]
                    if st.button(
                        f"🔄 Ubah Metode ({len(_pl_sel_ubah_bulk)} paket) via CDP",
                        disabled=not _pl_sel_ubah_bulk,
                        use_container_width=True,
                        key="pk_btn_ubah_metode_bulk",
                    ):
                        _pl_base_ubah_b = pl_engine.BASE_URL + "/"
                        _pl_ok_b = _pl_fail_b = 0
                        for _nm_b in _pl_sel_ubah_bulk:
                            _kd_b = _pl_opsi_ubah_bulk[_nm_b]
                            _hasil_b = spse_browser.ubah_metode_via_playwright(_kd_b, _pl_kat_id_bulk, _pl_pilih_val_bulk, _pl_base_ubah_b)
                            if _hasil_b == "OK":
                                _pl_ok_b += 1
                                st.write(f"✅ {_nm_b[:45]}")
                            else:
                                _pl_fail_b += 1
                                st.write(f"❌ {_nm_b[:45]} — `{_hasil_b}`")
                        st.success(f"Selesai: {_pl_ok_b} OK, {_pl_fail_b} GAGAL.")

        with _pl_col_kanan:
            st.markdown("#### 3. Buat Folder Paket")

            if "pl_folder_just_created" in st.session_state:
                _msg = st.session_state.pop("pl_folder_just_created")
                st.toast(f"✅ {_msg}", icon="📁")
                st.success(f"✅ {_msg}")
                st.balloons()

            # Dropdown pilih paket
            def _pl_no_dari_nama(nama: str, fallback: int) -> int:
                """Ekstrak nomor dari nama paket, misal 'Paket 1' → 1."""
                import re as _re
                m = _re.search(r"Paket\s+(\d+)", nama, _re.IGNORECASE)
                return int(m.group(1)) if m else fallback

            _pl_dl_dokumen = st.checkbox("📦 Download dokumen SPSE (KAK, Personil, Kontrak) saat buat folder", value=True, key="pl_cb_dl_pk")
            _pl_rt_refresh = False
            _pl_extract_teks = False
            if not st.session_state.get("pl_cb_isi_excel_pk_default_v2"):
                st.session_state["pl_cb_isi_excel_pk"] = True
                st.session_state["pl_cb_isi_excel_pk_default_v2"] = True
            _pl_isi_excel = st.checkbox("📊 Isi Excel @ Master Data (wajib jika workbook langsung dipakai)", key="pl_cb_isi_excel_pk")

            # ── Bulk: Buat Semua Folder ──────────────────────────────
            st.divider()

            _pl_rows_belum = [
                r for r in _pl_rows
                if r.get("nama_paket") and not r.get("folder_dibuat")
            ]

            # Plan: pre-compute nama folder per paket
            _pl_no_global = max(
                _pl_engine_utils.nomor_folder_tertinggi(_PL_DIR_JKK),
                _pl_engine_utils.nomor_folder_tertinggi(_PL_DIR_PK),
            )
            _pl_no_offset = {_PL_DIR_JKK: _pl_no_global, _PL_DIR_PK: _pl_no_global}
            _pl_bulk_plan = []
            for _bi0, _br0 in enumerate(_pl_rows_belum, 1):
                _bnm0  = _br0.get("nama_paket", "")
                _bj0   = (_br0.get("jenis_pl") or "PK").upper()
                _bpfx0 = {"JKK": "PLJKK", "PK": "PLPK"}.get(_bj0, f"PL{_bj0}")
                _bout_base0  = _PL_DIR_JKK if _bj0 == "JKK" else _PL_DIR_PK
                _bno0  = _pl_no_offset[_bout_base0] + _bi0
                _bnm_folder0 = re.sub(r'[/<>:"\|?*]', "-", f"{_bno0}. {_bpfx0} - {_bnm0}").strip()
                _bnm_folder0 = pl_engine.nama_folder_dengan_suffix_ulang(_bout_base0, _bnm_folder0)
                _bnm_folder0 = _pl_engine_utils.truncate_nama_folder(_bout_base0, _bnm_folder0)
                _pl_bulk_plan.append({
                    "kode_paket": _br0.get("kode_paket", ""),
                    "nama_folder": _bnm_folder0,
                    "out_base": _bout_base0,
                    "jenis_pl": _bj0,
                    "workflow": _pl_workflow(_br0),
                    "template_dir": _template_dir_pl_jkk(_br0, _TEMPLATE_DIR_PL) if _bj0 == "JKK" else _template_dir_pl_jkk(_br0, _TEMPLATE_DIR_PL_PK),
                })

            
            # ── #2: Checklist pilih paket untuk buat folder ──────────────────
            if _pl_rows_belum:
                st.markdown("**Pilih paket yang akan dibuat foldernya:**")
                _plf_col1, _plf_col2 = st.columns(2)
                # Tombol pilih semua / batal semua (tiru pola Tab 5)
                if _plf_col1.button("✅ Pilih Semua", key="plf_pilih_semua", use_container_width=True):
                    for _br_chk in _pl_rows_belum:
                        st.session_state[f"plf_chk_{_br_chk.get('kode_paket','')}"] = True
                    st.rerun()
                if _plf_col2.button("❌ Batal Semua", key="plf_batal_semua", use_container_width=True):
                    for _br_chk in _pl_rows_belum:
                        st.session_state[f"plf_chk_{_br_chk.get('kode_paket','')}"] = False
                    st.rerun()
                for _br_chk in _pl_rows_belum:
                    _bkp_chk = _br_chk.get("kode_paket", "")
                    _plf_chk_key = f"plf_chk_{_bkp_chk}"
                    if _plf_chk_key not in st.session_state:
                        st.session_state[_plf_chk_key] = True
                    st.checkbox(
                        f"{_pl_label(_br_chk)} — {(_br_chk.get('jenis_pl') or '').upper()}",
                        key=_plf_chk_key,
                    )
                # Hitung yang dicentang untuk label tombol
                _pl_terpilih_plan = [
                    item for item in _pl_bulk_plan
                    if st.session_state.get(f"plf_chk_{item['kode_paket']}", True)
                ]
                if st.button(
                    f"📁 Buat Folder Terpilih ({len(_pl_terpilih_plan)} paket)",
                    disabled=len(_pl_terpilih_plan) == 0,
                    use_container_width=True,
                    key="pl_btn_buat_terpilih",
                    type="primary",
                ):
                    import time as _pl_time
                    _pl_t0 = _pl_time.perf_counter()
                    _pl_bp = st.progress(0.0)
                    _pl_bulk_status = st.status(f"📁 Memproses {len(_pl_terpilih_plan)} paket terpilih... · ⏱ 0m 0d", expanded=True)
                    _pl_bulk_status_line = _pl_bulk_status.empty()
                    _pl_ok, _pl_fail = 0, 0
                    _pl_bulk_semua_log = {}
                    # ── FASE 1: I/O paralel antar-paket (download+parse+serap+HPS.md) ──
                    import queue as _pl_queue
                    _pl_event_q = _pl_queue.Queue()
                    _pl_live_events = ["⏱ Timer mulai"]
                    _pl_cfg_io = {
                        "py": _PL_PY, "script": _PL_SCRIPT, "no_win": _PL_NO_WIN,
                        "pokja_root": _PL_POKJA_ROOT, "dl_dokumen": bool(_pl_dl_dokumen),
                        "event_q": _pl_event_q,
                    }
                    try:
                        import spse_browser as _pl_sb_io
                        _pl_cookie = _pl_sb_io.get_spse_cookies()  # ambil 1× untuk semua paket
                    except Exception as _ck_e:
                        _pl_cookie = ""
                    _pl_io_hasil = []
                    _pl_done_ct = 0
                    _pl_n_total = len(_pl_terpilih_plan)
                    with ThreadPoolExecutor(max_workers=4) as _pl_ex:
                        from concurrent.futures import wait as _pl_wait, FIRST_COMPLETED as _PL_FIRST_COMPLETED
                        _pl_pending = {
                            _pl_ex.submit(_pl_proses_io_satu_paket, _it, _pl_cookie, _pl_cfg_io)
                            for _it in _pl_terpilih_plan
                        }
                        while _pl_pending:
                            _pl_done, _pl_pending = _pl_wait(_pl_pending, timeout=0.5, return_when=_PL_FIRST_COMPLETED)
                            while not _pl_event_q.empty():
                                _pl_live_events.append(_pl_event_q.get())
                                _pl_live_events = _pl_live_events[-12:]
                            if _pl_live_events:
                                _pl_bulk_status_line.code("\n".join(_pl_live_events))
                            for _pl_fut in _pl_done:
                                _pl_res = _pl_fut.result()
                                _pl_io_hasil.append(_pl_res)
                                _pl_done_ct += 1
                                _pl_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                                _pl_bp.progress(_pl_done_ct / max(_pl_n_total, 1))
                                _pl_bulk_status.update(label=f"[{_pl_done_ct}/{_pl_n_total}] selesai: {_pl_res['nama_folder'][:50]} · ⏱ {_pl_elapsed}")
                                _pl_live_events.append(f"✅ SELESAI: {_pl_res['nama_folder'][:55]} · ⏱ {_pl_elapsed}")
                                _pl_live_events = _pl_live_events[-12:]
                                _pl_bulk_status_line.code("\n".join(_pl_live_events))
                    # ── FASE 2: serial (COM/merge/OCR) di main thread ──
                    for _pl_res in _pl_io_hasil:
                        _pl_nf = _pl_res["nama_folder"]
                        _pl_kp_b = _pl_res["kode"]
                        _pl_target_b = _pl_res["target"]
                        _pl_paket_log = _pl_res["log"]
                        if not _pl_res["ok"]:
                            _pl_fail += 1
                            _pl_bulk_semua_log[_pl_nf] = _pl_paket_log
                            continue
                        _pl_ok += 1
                        _pl_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                        _pl_bulk_status.update(label=f"Finalisasi: {_pl_nf[:55]} · ⏱ {_pl_elapsed}")
                        # Merge PDF draft (COM tidak thread-safe → serial)
                        if _pl_res.get("files_ok"):
                            _t_step = _pl_time.perf_counter()
                            try:
                                _pl_merged = pl_engine.gabung_draft_pl(_pl_kp_b, _pl_target_b, _pl_res["files_ok"])
                                if _pl_merged:
                                    _pl_paket_log.append(f"📎 Draft PDF: {_pl_os.path.basename(_pl_merged)}")
                                _pl_paket_log.append(f"⏱ merge draft: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _mg_e:
                                _pl_paket_log.append(f"⚠ Gabung Draft PDF: {_mg_e}")
                                _pl_paket_log.append(f"⏱ merge draft: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ merge draft: skipped (0 file)")
                        # Extract teks kualifikasi (OCR berat → serial dgn timeout)
                        if _pl_extract_teks and _pl_dl_dokumen:
                            _t_step = _pl_time.perf_counter()
                            try:
                                import extract_teks_kualifikasi as _etk
                                import threading as _etk_th
                                _etk_folder = _pl_os.path.join(_pl_target_b, "8. Dokumen Kualifikasi")
                                if _pl_os.path.isdir(_etk_folder):
                                    _etk_result_box = [None]
                                    def _etk_run(_f=_etk_folder, _log=_pl_paket_log):
                                        try:
                                            _etk_result_box[0] = _etk.extract_folder_kualifikasi(
                                                _f, progress_cb=lambda m: _log.append(f"  {m}"),
                                            )
                                        except Exception as _ex:
                                            _etk_result_box[0] = {"ok": False, "error": str(_ex)}
                                    _etk_t = _etk_th.Thread(target=_etk_run, daemon=True)
                                    _etk_t.start()
                                    _etk_t.join(timeout=120)
                                    if _etk_t.is_alive():
                                        _pl_paket_log.append("⚠ Extract teks: timeout 120s, dilewati")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} timeout")
                                    elif _etk_result_box[0] and _etk_result_box[0].get("ok"):
                                        _etk_res = _etk_result_box[0]
                                        _pl_paket_log.append(f"📝 Extract teks: {len(_etk_res.get('penyedia', []))} penyedia, ~{_etk_res.get('total_token_estimasi', 0)} token")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                                    else:
                                        _pl_paket_log.append("⚠ Extract teks: tidak ada penyedia/PDF")
                                        _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _etk_e:
                                _pl_paket_log.append(f"⚠ Extract teks: {_etk_e}")
                                _pl_paket_log.append(f"⏱ OCR: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ OCR: skipped")
                        # Refresh + HPS + Master Data: COM Excel (serial)
                        if _pl_isi_excel:
                            _t_step = _pl_time.perf_counter()
                            try:
                                _excel_logs = _proses_excel_paket_pl(
                                    _pl_target_b, _pl_kp_b,
                                    _pl_res["jenis_pl"], _pl_rt_refresh,
                                    _pl_res.get("template_dir") or _TEMPLATE_DIR_PL, _TEMPLATE_DIR_PL_PK,
                                )
                                for _el in _excel_logs:
                                    _icon = "📊" if _el.startswith("HPS:") else (
                                            "📝" if _el.startswith("Master Data") else (
                                            "🔄" if _el.startswith("Refresh") else "⚠"))
                                    _pl_paket_log.append(f"{_icon} {_el}")
                                _pl_paket_log.append(f"⏱ Excel: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)}")
                            except Exception as _xl_e:
                                _pl_paket_log.append(f"⚠ Excel Master Data: {_xl_e}")
                                _pl_paket_log.append(f"⏱ Excel: {_fmt_step_seconds(_pl_time.perf_counter() - _t_step)} error")
                        else:
                            _pl_paket_log.append("⏱ Excel: skipped")
                        try:
                            _meta_wf = mark_workflow_applied(
                                _pl_target_b, _pl_res.get("workflow") or _pl_workflow({"jenis_pl": _pl_res["jenis_pl"], "nama_paket": _pl_nf}), _pl_res["jenis_pl"]
                            )
                            _pl_paket_log.append(f"🧩 Workflow: {_meta_wf.get('workflow_applied')}")
                        except Exception as _meta_wf_e:
                            _pl_paket_log.append(f"⚠ Metadata workflow: {_meta_wf_e}")
                        _pl_bulk_semua_log[_pl_nf] = _pl_paket_log
                    _pl_total_elapsed = _fmt_elapsed(_pl_time.perf_counter() - _pl_t0)
                    _pl_live_events.append(f"⏱ Total waktu: {_pl_total_elapsed}")
                    _pl_bulk_status_line.code('\n'.join(_pl_live_events[-12:]))
                    _pl_ringkasan = f"✅ {_pl_ok} folder berhasil, ❌ {_pl_fail} gagal · ⏱ {_pl_total_elapsed}"
                    _pl_bulk_status.update(label=_pl_ringkasan, state="complete", expanded=False)
                    with st.expander("📋 Log detail per paket", expanded=_pl_fail > 0):
                        for _pl_nf, _pl_logs in _pl_bulk_semua_log.items():
                            st.markdown(f"**{_pl_nf[:70]}**")
                            st.code("\n".join(_pl_logs))
                    st.session_state["pl_folder_bulk_created"] = _pl_ringkasan
                    _load_draft_pl_cached.clear()
            else:
                st.info("✅ Semua paket sudah punya folder.")
                st.button("📁 Buat Folder Terpilih (0 paket)", disabled=True, use_container_width=True, key="pl_btn_buat_terpilih_disabled_pk")
            # ── #4: Update Data Folder (Re-download + Reset) ─────────────────
            st.divider()
            st.markdown("#### 4. Sinkronisasi & Pemutakhiran Paket Existing")
            st.caption(
                "⚠️ **Sudah dijalankan otomatis saat Buat Folder** (jika checkbox Download dicentang). "
                "Gunakan tombol ini **hanya** jika ada perubahan dokumen dari PPK atau ingin refresh ulang."
            )
            _cb_dl_dok_bulk = st.checkbox("📦 Re-download Dokumen SPSE (KAK, Personil, Kontrak)", value=False, key="pl_cb_dl_dok_bulk_pk")
            _cb_hps_update = st.checkbox("💰 Update HPS semua paket berfolder → Excel + MD", value=False, key="pl_cb_hps_update_pk")
            _cb_excel_existing = st.checkbox("📊 Isi ulang Excel @ Master Data semua paket berfolder", value=False, key="pl_cb_excel_existing_pk")
            _cb_reset_existing = st.checkbox("🔄 Reset status folder semua paket berfolder", value=False, key="pl_cb_reset_existing_pk")
            _cb_pra_reviu_bulk = st.checkbox("🤖 Pra-Reviu + Isi DOCM semua paket berfolder", value=False, key="pl_cb_pra_reviu_bulk_pk")

            if st.button("▶ Jalankan Operasi Terpilih", use_container_width=True, key="btn_update_data_folder_pk"):
                # Aksi: Download dokumen bulk semua paket berfolder
                if _cb_dl_dok_bulk:
                    import kualifikasi_engine_plpk as _keng_pl_dl
                    _pl_rows_dl_bulk = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    if not _pl_rows_dl_bulk:
                        st.info("Tidak ada paket dengan folder untuk download dokumen.")
                    else:
                        _dl_bulk_ok, _dl_bulk_fail = 0, 0
                        _dl_bulk_status = st.status(
                            f"📦 Download dokumen {len(_pl_rows_dl_bulk)} paket...", expanded=True
                        )
                        _dl_bulk_line = _dl_bulk_status.empty()
                        _dl_bulk_bp = st.progress(0.0)
                        for _db_i, _db_row in enumerate(_pl_rows_dl_bulk):
                            _db_kp   = _db_row.get("kode_paket", "")
                            _db_nama = _pl_label(_db_row)
                            _dl_bulk_status.update(label=f"[{_db_i+1}/{len(_pl_rows_dl_bulk)}] {_db_nama}")
                            _dl_bulk_bp.progress((_db_i + 1) / len(_pl_rows_dl_bulk))
                            _db_root = ""
                            try:
                                _db_fr = _keng_pl_dl.resolve_folder_paket_pl(_db_kp)
                                _db_root = _db_fr.get("pesan", "") if _db_fr.get("ok") else ""
                            except Exception:
                                _db_root = ""
                            if not _db_root or not _pl_os.path.isdir(_db_root):
                                _dl_bulk_fail += 1
                                _dl_bulk_line.write(f"⚠ [{_db_i+1}] {_db_nama} — folder tidak ditemukan")
                                continue
                            try:
                                _db_dl_logs = []
                                def _db_dl_cb(msg, _log=_db_dl_logs):
                                    _log.append(msg)
                                _db_dl_res = pl_engine.download_dokumen_paket_pl(_db_kp, _db_root, _db_dl_cb, force_clean=True)
                                _dl_bulk_ok += 1
                                _dl_bulk_line.write(f"✅ [{_db_i+1}] {_db_nama} — {len(_db_dl_res.get('ok',[]))} file")
                                # Parse KAK setelah download
                                _db_kak_p = parse_kak_pl.cari_kak_di_folder(_db_root)
                                if _db_kak_p:
                                    _db_kak_d = parse_kak_pl.parse_kak(_db_kak_p)
                                    _db_kak_u = {k: v for k, v in _db_kak_d.items() if v}
                                    if _db_kak_u:
                                        pl_engine.simpan_paket_pl({"kode_paket": _db_kp, **_db_kak_u})
                            except Exception as _db_e:
                                _dl_bulk_fail += 1
                                _dl_bulk_line.write(f"❌ [{_db_i+1}] {_db_nama} — {_db_e}")
                        _dl_bulk_bp.progress(1.0)
                        _dl_bulk_line.empty()
                        _dl_bulk_status.update(
                            label=f"📦 Download selesai: ✅ {_dl_bulk_ok} sukses, ❌ {_dl_bulk_fail} gagal",
                            state="complete", expanded=_dl_bulk_fail > 0,
                        )


                # Aksi: Update HPS bulk
                if _cb_hps_update:
                    import hps_engine as _hps_upd
                    import kualifikasi_engine_plpk as _keng_hps_upd
                    _pl_rows_hps_upd = [
                        r for r in _pl_rows
                        if r.get("kode_paket") and r.get("folder_dibuat")
                    ]
                    if not _pl_rows_hps_upd:
                        st.info("Tidak ada paket dengan folder untuk update HPS.")
                    else:
                        _hps_upd_ok, _hps_upd_fail = 0, 0
                        _hps_upd_gagal = []
                        _hps_upd_status = st.status(
                            f"💰 Update HPS {len(_pl_rows_hps_upd)} paket...", expanded=True
                        )
                        _hps_upd_line = _hps_upd_status.empty()
                        _hps_upd_bp = st.progress(0.0)
                        for _hu_i, _hu_row in enumerate(_pl_rows_hps_upd):
                            _hu_kp   = _hu_row.get("kode_paket", "")
                            _hu_nama = _pl_label(_hu_row)
                            _hps_upd_status.update(label=f"[{_hu_i+1}/{len(_pl_rows_hps_upd)}] {_hu_nama}")
                            _hps_upd_bp.progress((_hu_i + 1) / len(_pl_rows_hps_upd))
                            try:
                                _hu_fr = _keng_hps_upd.resolve_folder_paket_pl(_hu_kp)
                                _hu_root = _hu_fr.get("pesan", "") if _hu_fr.get("ok") else ""
                                _hu_xl = _cari_xlsm_pl(_hu_root) if _hu_root and _pl_os.path.isdir(_hu_root) else None
                                if not _hu_xl:
                                    raise ValueError("folder/xlsm tidak ditemukan")
                                _hu_r = _hps_upd.scrape_hps_pl_ke_excel(_hu_kp, _hu_xl)
                                if _hu_r.get("ok"):
                                    _hps_upd_ok += 1
                                    _hps_upd_line.write(f"✅ [{_hu_i+1}] {_hu_nama} — {_hu_r['count']} item")
                                else:
                                    raise ValueError(_hu_r.get("pesan", "-"))
                            except Exception as _hu_e:
                                _hps_upd_fail += 1
                                _hps_upd_gagal.append(f"{_hu_nama}: {_hu_e}")
                                _hps_upd_line.write(f"❌ [{_hu_i+1}] {_hu_nama} — {_hu_e}")
                        _hps_upd_bp.progress(1.0)
                        _hps_upd_line.empty()
                        _hps_upd_status.update(
                            label=f"💰 HPS selesai: ✅ {_hps_upd_ok} sukses, ❌ {_hps_upd_fail} gagal",
                            state="complete", expanded=_hps_upd_fail > 0,
                        )
                        if _hps_upd_gagal:
                            st.warning("Paket gagal:\n" + "\n".join(_hps_upd_gagal))

                if _cb_excel_existing:
                    import isi_master_data_pl as _imd_excel_existing
                    import kualifikasi_engine_plpk as _keng_excel_existing
                    _excel_rows = [r for r in _pl_rows if r.get("kode_paket") and r.get("folder_dibuat")]
                    _ok_x = _fail_x = 0
                    _st_x = st.status(f"📊 Isi ulang Excel {len(_excel_rows)} paket...", expanded=True)
                    _ln_x = _st_x.empty()
                    for _xe_i, _xe_row in enumerate(_excel_rows):
                        _xe_kp = _xe_row.get("kode_paket", ""); _xe_nm = _pl_label(_xe_row)
                        try:
                            _xe_fr = _keng_excel_existing.resolve_folder_paket_pl(_xe_kp)
                            _xe_root = _xe_fr.get("pesan", "") if _xe_fr.get("ok") else ""
                            _xe_xlsm = _cari_xlsm_pl(_xe_root) if _xe_root and _pl_os.path.isdir(_xe_root) else None
                            if not _xe_xlsm: raise ValueError("folder/workbook .xlsm tidak ditemukan")
                            _xe_res = _imd_excel_existing.isi_master_data_pl(_xe_kp, _xe_xlsm)
                            if not _xe_res.get("ok"): raise ValueError(_xe_res.get("pesan", "macro gagal"))
                            _ok_x += 1; _ln_x.write(f"✅ {_xe_nm}")
                        except Exception as _xe_e:
                            _fail_x += 1; _ln_x.write(f"❌ {_xe_nm} — {_xe_e}")
                    _st_x.update(label=f"📊 Isi ulang Excel selesai: ✅ {_ok_x}, ❌ {_fail_x}", state="complete", expanded=_fail_x > 0)

                if _cb_reset_existing:
                    from config import sb as _sb_reset_bulk
                    _reset_rows_bulk = [r for r in _pl_rows if r.get("kode_paket") and r.get("folder_dibuat")]
                    try:
                        _sb_reset_bulk().table("draft_paket_pl").update({"folder_dibuat": None}).in_("kode_paket", [r.get("kode_paket") for r in _reset_rows_bulk]).execute()
                        _load_draft_pl_cached.clear(); st.success(f"✅ {len(_reset_rows_bulk)} status folder berhasil direset."); st.rerun()
                    except Exception as _reset_bulk_e:
                        st.error(f"Reset bulk gagal: {_reset_bulk_e}")

                if _cb_pra_reviu_bulk:
                    import ai_evaluator as _heval_bulk
                    _pra_rows_bulk = [r for r in _pl_rows if r.get("kode_paket") and r.get("folder_dibuat")]
                    _pra_jobs_bulk = [{"nomor_urut": r.get("nomor_urut"), "nama_paket": r.get("nama_paket", "-")} for r in _pra_rows_bulk]
                    if _pra_jobs_bulk:
                        _pra_res = _heval_bulk.evaluasi_bulk(_pra_jobs_bulk, jenis="pra_reviu", model=None, max_workers=1, engine="codex")
                        _pra_ok = sum(1 for r in _pra_res if r.get("status") == "ok")
                        st.success(f"🤖 Pra-reviu selesai: ✅ {_pra_ok}, ❌ {len(_pra_res)-_pra_ok}")

    # ── Tab 2: Kirim Undangan DPP ─────────────────────────────────────────────
    if _pl_active_tab == "2️⃣ Kirim Undangan DPP":
        # Satu kanvas penuh: daftar paket juga menjadi tempat upload BA Reviu.
        _kd_col_list = st.container()

        with _kd_col_list:
            st.markdown("### 1. Pilih Paket")

            _pl_rows_kd = _load_draft_pl_cached("PK")
            _pl_rows_kd, _ = pl_engine.buang_duplikat_paket_lama(_pl_rows_kd)
            _pl_rows_kd = [r for r in _pl_rows_kd if not pl_engine.is_paket_selesai(r)]
            _kd_selected = []
            if not _pl_rows_kd:
                st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
            else:
                _kd_sel_col1, _kd_sel_col2 = st.columns(2)
                with _kd_sel_col1:
                    if st.button("✅ Semua", key="kd_sel_all", use_container_width=True):
                        for _rr in _pl_rows_kd:
                            st.session_state[f"kd_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _kd_sel_col2:
                    if st.button("⬜ Kosong", key="kd_sel_none", use_container_width=True):
                        for _rr in _pl_rows_kd:
                            st.session_state[f"kd_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _kd_selected = []
                import upload_ba_reviu_pl as _ubrpl
                def _do_upload_ba_pl(paket_list, tgl=None):
                    hasil = []
                    prog = st.progress(0, text="Memulai upload...")
                    for _i, _p in enumerate(paket_list):
                        prog.progress((_i + 1) / len(paket_list), text=f"Upload {_p['kode_paket']} ({_i+1}/{len(paket_list)})...")
                        _tgl_ba = _p.get("_tgl_acara") or tgl or datetime.now().date()
                        _res = _ubrpl.upload_ba_reviu_pl(kode_paket=_p["kode_paket"], file_bytes=_p["_ba_file"].getvalue(), file_name=_p["_ba_file"].name, tgl_ba=_tgl_ba.strftime("%d-%m-%Y"))
                        hasil.append({"kode": _p["kode_paket"], "nama": _p["nama_paket"], "sukses": _res["ok"], "pesan": f"HTTP {_res.get('status','?')}" if _res["ok"] else _res.get("error", "?")})
                    prog.empty()
                    _ok = sum(1 for h in hasil if h["sukses"])
                    st.success(f"✅ {_ok} BA Reviu berhasil diupload!") if _ok == len(hasil) else st.warning(f"⚠️ {_ok} berhasil, {len(hasil)-_ok} gagal.")
                    st.dataframe(hasil, use_container_width=True, hide_index=True)
                for _rr in _pl_rows_kd:
                    _kd_key     = f"kd_chk_{_rr['kode_paket']}_v19"
                    _kd_tgl_key = f"kd_tgl_acara_{_rr['kode_paket']}"
                    _kd_display = _pl_label(_rr)
                    _col_chk, _col_tgl, _col_ba = st.columns([3, 2, 3])
                    with _col_chk:
                        _kd_chk = st.checkbox(
                            "",
                            value=st.session_state.get(_kd_key, True),
                            key=_kd_key,
                            label_visibility="collapsed",
                        )
                        st.markdown(f"**{_kd_display}**")
                    with _col_tgl:
                        st.caption("Tanggal Undangan / BA Reviu")
                        _kd_tgl_acara = st.date_input(
                            "Tanggal Acara",
                            value=st.session_state.get(_kd_tgl_key)
                            or _last_pl_invitation_date(_rr["kode_paket"])
                            or datetime.now().date(),
                            format="DD/MM/YYYY",
                            key=_kd_tgl_key,
                            label_visibility="collapsed",
                        )
                        st.caption(f"{_HARI_NAMA[_kd_tgl_acara.weekday()]}, {_kd_tgl_acara.day} {_BULAN_NAMA[_kd_tgl_acara.month-1]} {_kd_tgl_acara.year}")
                        if _kd_tgl_acara in _LIBUR_MAP:
                            st.caption(f"⚠️ {_LIBUR_MAP[_kd_tgl_acara]}")
                    with _col_ba:
                        _ba_fkey = f"plba_file_{_rr['kode_paket']}"
                        _ba_up = st.file_uploader("BA Reviu PDF (opsional)", type=["pdf"], key=_ba_fkey)
                    if _kd_chk:
                        _kd_selected.append({**_rr, "_tgl_acara": _kd_tgl_acara})

                st.caption(f"**{len(_kd_selected)}** dari **{len(_pl_rows_kd)}** paket dipilih")

            st.divider()
            st.markdown("### 2. Detail Undangan")
            st.caption("Pesan dikirim PP ke PPK — meminta reviu Dokumen Persiapan Pengadaan.")

            st.markdown("**Waktu Acara (berlaku semua paket)**")
            _kd_col_mulai, _kd_col_selesai = st.columns(2)
            with _kd_col_mulai:
                _kd_jam_mulai = st.time_input(
                    "Mulai",
                    value=datetime.strptime("09:00", "%H:%M").time(),
                    key="kd_jam_mulai",
                    step=1800,
                )
            with _kd_col_selesai:
                _kd_jam_selesai = st.time_input(
                    "Selesai",
                    value=datetime.strptime("11:00", "%H:%M").time(),
                    key="kd_jam_selesai",
                    step=1800,
                )

            with st.expander("ℹ️ Libur Nasional Tersisa"):
                _kd_hari_ini = datetime.now().date()
                for _kd_d in sorted(d for d in _LIBUR_MAP if d >= _kd_hari_ini):
                    st.write(f"• {_HARI_NAMA[_kd_d.weekday()]}, {_kd_d.day} {_BULAN_NAMA[_kd_d.month-1]} {_kd_d.year} — {_LIBUR_MAP[_kd_d]}")

            _kd_tempat = pl_kirimpesan_engine.DEFAULT_TEMPAT
            st.caption(f"📍 Tempat: {_kd_tempat}")

            st.divider()
            st.warning("⚠️ Pesan yang terkirim **tidak bisa dihapus** dari SPSE.")
            if not st.session_state.get("kd_konfirmasi"):
                if st.button(
                    f"📨 Kirim Undangan DPP ke {len(_kd_selected)} Paket",
                    key="kd_kirim",
                    type="primary",
                    disabled=len(_kd_selected) == 0,
                    use_container_width=True,
                ):
                    if not _kd_tempat.strip():
                        st.error("❌ Tempat wajib diisi.")
                    else:
                        st.session_state["kd_konfirmasi"] = True
                        st.rerun()
            else:
                _kd_konfirm_lines = "\n".join(
                    f"{_pl_label(p)}  \n"
                    f"   📅 {_HARI_NAMA[p['_tgl_acara'].weekday()]}, {p['_tgl_acara'].day} {_BULAN_NAMA[p['_tgl_acara'].month-1]} {p['_tgl_acara'].year}"
                    for i, p in enumerate(_kd_selected)
                )
                st.warning(
                    f"Kirim ke **{len(_kd_selected)} paket**\n\n"
                    f"{_kd_konfirm_lines}\n\n"
                    f"- Pukul: {_kd_jam_mulai.strftime('%H.%M')} s.d. {_kd_jam_selesai.strftime('%H.%M')} Wita\n"
                    f"- Tempat: {_kd_tempat.strip()[:80]}\n\n"
                    f"**Tidak bisa dibatalkan setelah dikirim.**"
                )
                _kdc1, _kdc2 = st.columns(2)
                with _kdc1:
                    if st.button("✅ Ya, Kirim", key="kd_ya", type="primary", use_container_width=True):
                        st.session_state["kd_konfirmasi"] = False
                        _kd_progress = st.progress(0, text="Memulai pengiriman...")
                        _kd_hasil = []
                        _tgl_kirim_kd = datetime.now().date()

                        for _ki, _kp in enumerate(_kd_selected):
                            _kd_progress.progress(
                                (_ki + 1) / len(_kd_selected),
                                text=f"Mengirim {_ki+1}/{len(_kd_selected)}...",
                            )
                            _kd_tgl_a  = _kp["_tgl_acara"]
                            _kd_hari_tgl = f"{_HARI_NAMA[_kd_tgl_a.weekday()]}, {_kd_tgl_a.day} {_BULAN_NAMA[_kd_tgl_a.month-1]} {_kd_tgl_a.year}"
                            _kd_pukul    = f"{_kd_jam_mulai.strftime('%H.%M')} s.d. {_kd_jam_selesai.strftime('%H.%M')} Wita"

                            # Generate PDF lampiran otomatis
                            import undangan_pdf_engine as _upe
                            _gen = _upe.generate_undangan_pdf_pl(
                                kode_paket=_kp["kode_paket"],
                                tanggal_kirim=_tgl_kirim_kd,
                                hari_tgl_rapat=_kd_hari_tgl,
                                pukul_rapat=_kd_pukul,
                                tempat_rapat=_kd_tempat.strip(),
                            )
                            _lamp_bytes = _gen["pdf_bytes"] if _gen["sukses"] else None
                            _ku_lamp = _kp.get("kode_unik") or _kp["kode_paket"]
                            _lamp_nama  = f"undangan_reviu_{_ku_lamp}.pdf"
                            if _lamp_bytes:
                                st.session_state.setdefault("_kd_pdf_cache", {})[_ku_lamp] = (_lamp_nama, _lamp_bytes)

                            _waktu_str  = datetime.combine(_kd_tgl_a, _kd_jam_mulai).strftime("%d-%m-%Y %H:%M")
                            _sampai_str = datetime.combine(_kd_tgl_a, _kd_jam_selesai).strftime("%d-%m-%Y %H:%M")

                            _res = pl_kirimpesan_engine.kirim_undangan_pl(
                                kode=_kp["kode_paket"],
                                waktu=_waktu_str,
                                sampai=_sampai_str,
                                tempat=_kd_tempat.strip(),
                                dibawa=pl_kirimpesan_engine.DEFAULT_DIBAWA,
                                hadir=pl_kirimpesan_engine.DEFAULT_HADIR,
                                lampiran_bytes=_lamp_bytes,
                                lampiran_nama=_lamp_nama,
                            )
                            _kd_hasil.append({
                                "Paket": _pl_label(_kp),
                                "Penerima (PPK)": _res.get("penerima", "-"),
                                "PDF": "✅" if _gen["sukses"] else f"❌ {_gen['pesan']}",
                                "Kirim": "✅" if _res["sukses"] else f"❌ {_res['pesan']}",
                            })
                            if _res.get("sukses"):
                                _save_last_pl_invitation_date(_kp["kode_paket"], _kd_tgl_a)

                        _kd_progress.empty()
                        _kd_ok = sum(1 for h in _kd_hasil if h["Kirim"] == "✅")
                        if _kd_ok == len(_kd_hasil):
                            st.success(f"✅ Semua {_kd_ok} undangan berhasil dikirim!")
                        else:
                            st.warning(f"⚠️ {_kd_ok} berhasil, {len(_kd_hasil)-_kd_ok} gagal.")
                        st.dataframe(
                            _kd_hasil,
                            use_container_width=True,
                            column_config={
                                "Paket":          st.column_config.TextColumn("Paket", width="large"),
                                "Penerima (PPK)": st.column_config.TextColumn("Penerima (PPK)"),
                                "PDF":            st.column_config.TextColumn("PDF", width="small"),
                                "Kirim":          st.column_config.TextColumn("Kirim", width="small"),
                            },
                            hide_index=True,
                        )
                        # Tombol download per PDF
                        for _ku_dl, (_nm_dl, _by_dl) in st.session_state.get("_kd_pdf_cache", {}).items():
                            st.download_button(
                                f"⬇️ Download {_nm_dl}",
                                data=_by_dl,
                                file_name=_nm_dl,
                                mime="application/pdf",
                                key=f"kd_dl_{_ku_dl}",
                            )

                with _kdc2:
                    if st.button("❌ Batal", key="kd_batal", use_container_width=True):
                        st.session_state["kd_konfirmasi"] = False
                        st.rerun()

            _ba_pl_valid = [
                {**_pp, "_ba_file": st.session_state.get(f"plba_file_{_pp['kode_paket']}")}
                for _pp in _kd_selected
                if st.session_state.get(f"plba_file_{_pp['kode_paket']}")
            ]
            if st.button(
                f"📄 Upload BA Reviu DPP ({len(_ba_pl_valid)} file)",
                key="plba_upload_selected",
                type="secondary",
                disabled=not _ba_pl_valid,
                use_container_width=True,
            ):
                _do_upload_ba_pl(_ba_pl_valid)

    # ── Tab 4: Buat Jadwal PL (5 tahap, push langsung ke SPSE) ─────────────
    if _pl_active_tab == "5️⃣ Buat Jadwal":
        st.markdown("### Buat Jadwal Pengadaan Langsung")
        st.caption("5 tahap PL: Upload Penawaran → Pembukaan → Evaluasi → Klarifikasi+Nego → Tanda Tangan Kontrak. Push langsung ke SPSE.")

        import jadwal_engine_pl as _jepl
        _libur_map_pl = _LIBUR_MAP

        _pljd_rows = _load_draft_pl_cached("PK")
        _pljd_rows, _ = pl_engine.buang_duplikat_paket_lama(_pljd_rows)
        _pljd_rows = [r for r in _pljd_rows if not pl_engine.is_paket_selesai(r)]
        if not _pljd_rows:
            st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
        else:
            _pljd_col_list, _pljd_col_detail = st.columns([3, 2])

            with _pljd_col_list:
                st.markdown("### 1. Pilih Paket")
                _pljd_a, _pljd_b = st.columns(2)
                with _pljd_a:
                    if st.button("✅ Semua", key="pljd_sel_all", use_container_width=True):
                        for _rr in _pljd_rows:
                            st.session_state[f"pljd_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _pljd_b:
                    if st.button("⬜ Kosong", key="pljd_sel_none", use_container_width=True):
                        for _rr in _pljd_rows:
                            st.session_state[f"pljd_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _pljd_selected = []
                for _rr in _pljd_rows:
                    _key = f"pljd_chk_{_rr['kode_paket']}_v19"
                    _chk = st.checkbox(
                        "",
                        value=st.session_state.get(_key, True),
                        key=_key,
                        label_visibility="collapsed",
                    )
                    st.markdown(f"**{_pl_label(_rr)}** ({_rr.get('jenis_pl','?')})")
                    if _chk:
                        _pljd_selected.append(_rr)

                st.caption(f"**{len(_pljd_selected)}** dari **{len(_pljd_rows)}** paket dipilih")

            with _pljd_col_detail:
                st.markdown("### 2. Tanggal Mulai (T1)")
                _pljd_beda = st.checkbox("Jadwal berbeda per paket", value=False, key="pljd_beda")

                if not _pljd_beda:
                    # Default: baca tgl_batas_penawaran (T1.selesai) dari DB, kurangi 5 hari
                    _pljd_tgl_default = datetime.now().date()
                    if _pljd_selected:
                        _tbp = _pljd_selected[0].get("tgl_batas_penawaran")
                        if _tbp:
                            try:
                                from datetime import date as _date2, timedelta as _td2
                                _pljd_tgl_default = _date2.fromisoformat(str(_tbp)[:10]) - _td2(days=5)
                            except Exception:
                                pass
                    _c1, _c2 = st.columns(2)
                    with _c1:
                        _pljd_tgl_global = st.date_input(
                            "Tanggal",
                            value=_pljd_tgl_default,
                            format="DD/MM/YYYY",
                            key="pljd_tgl_global",
                        )
                        st.markdown(f"**{_HARI_NAMA[_pljd_tgl_global.weekday()]}, {_pljd_tgl_global.day} {_BULAN_NAMA[_pljd_tgl_global.month-1]} {_pljd_tgl_global.year}**")
                    with _c2:
                        _pljd_jam_global = st.time_input(
                            "Jam",
                            value=datetime.strptime("08:00", "%H:%M").time(),
                            key="pljd_jam_global",
                        )
                    if _pljd_tgl_global in _libur_map_pl:
                        st.warning(f"⚠️ **{_libur_map_pl[_pljd_tgl_global]}**")
                else:
                    if not _pljd_selected:
                        st.info("Pilih paket dulu.")
                    else:
                        for _p in _pljd_selected:
                            _ktgl = f"pljd_tgl_{_p['kode_paket']}"
                            _kjam = f"pljd_jam_{_p['kode_paket']}"
                            _cna, _cdt, _cjm = st.columns([3, 2, 1])
                            with _cna:
                                st.markdown(f"**{_pl_label(_p)}**")
                            with _cdt:
                                st.date_input(
                                    "Tgl",
                                    value=st.session_state.get(_ktgl, datetime.now().date()),
                                    format="DD/MM/YYYY",
                                    key=_ktgl,
                                    label_visibility="collapsed",
                                )
                            with _cjm:
                                st.time_input(
                                    "Jam",
                                    value=st.session_state.get(_kjam, datetime.strptime("08:00", "%H:%M").time()),
                                    key=_kjam,
                                    label_visibility="collapsed",
                                )

                import gcal_pl_helper as _gcalpl_gpk
                _gcal_ok_pk = _gcalpl_gpk.check_gcal_token()

                _pljd_mode = st.radio("Mode Jadwal", ["Normal", "Standar", "Cepat"], horizontal=True, key="pljd_mode_pk")

                if _pljd_selected:
                    st.markdown(f"**📅 Preview Jadwal — cek dulu sebelum push**")
                    _pk_prev = st.selectbox(
                        "Pilih paket buat preview",
                        _pljd_selected,
                        format_func=_pl_label,
                        key="pljd_preview_pilih_pk",
                    )
                    _tgl_pv = st.session_state.get(f"pljd_tgl_{_pk_prev['kode_paket']}", datetime.now().date()) if _pljd_beda else _pljd_tgl_global
                    _jam_pv = st.session_state.get(f"pljd_jam_{_pk_prev['kode_paket']}", datetime.strptime("08:00", "%H:%M").time()) if _pljd_beda else _pljd_jam_global
                    _t1_pv = datetime.combine(_tgl_pv, _jam_pv)
                    if _pljd_mode == "Cepat":
                        _jadwal_pv = _jepl.hitung_jadwal_pl_cepat(_t1_pv)
                    elif _pljd_mode == "Standar":
                        _jadwal_pv = _jepl.hitung_jadwal_pl_standar(_t1_pv)
                    else:
                        _jadwal_pv = _jepl.hitung_jadwal_pl(_t1_pv)
                    for _jd in _jadwal_pv:
                        _m, _s = _jd["mulai"], _jd["selesai"]
                        _sama_hari = _m.date() == _s.date()
                        _txt_mulai = f"{_HARI_NAMA[_m.weekday()]}, {_m.day} {_BULAN_NAMA[_m.month-1]} {_m.strftime('%H:%M')}"
                        _txt_selesai = _s.strftime('%H:%M') if _sama_hari else f"{_HARI_NAMA[_s.weekday()]}, {_s.day} {_BULAN_NAMA[_s.month-1]} {_s.strftime('%H:%M')}"
                        with st.container(border=True):
                            st.write(f"**{_jd['nama']}**  \n{_txt_mulai} → {_txt_selesai}")
                    st.divider()

                st.divider()
                st.caption("⚠️ Akan menimpa jadwal yang sudah ada di SPSE.")

                _pljd_submit = st.button(
                    f"🚀 Push Jadwal ke SPSE ({len(_pljd_selected)} paket)",
                    type="primary",
                    use_container_width=True,
                    disabled=len(_pljd_selected) == 0,
                    key="pljd_submit_btn_pk",
                )

                if _pljd_submit:
                    _hasil = []
                    _prog = st.progress(0, text="Mulai...")
                    for _i, _p in enumerate(_pljd_selected):
                        _prog.progress((_i + 1) / len(_pljd_selected),
                                       text=f"{_p['kode_paket']} ({_i+1}/{len(_pljd_selected)})...")
                        if _pljd_beda:
                            _tgl = st.session_state.get(f"pljd_tgl_{_p['kode_paket']}", datetime.now().date())
                            _jam = st.session_state.get(f"pljd_jam_{_p['kode_paket']}", datetime.strptime("08:00", "%H:%M").time())
                        else:
                            _tgl = _pljd_tgl_global
                            _jam = _pljd_jam_global
                        _t1 = datetime.combine(_tgl, _jam)

                        _kp = _p.get("kode_paket")
                        if not _kp:
                            _hasil.append({"paket": _pl_label(_p), "ok": False, "pesan": "kode_paket kosong"})
                            continue
                        try:
                            _mode_str = {"Cepat": "cepat", "Standar": "standar"}.get(_pljd_mode, "normal")
                            _r = _jepl.submit_full_pl(_kp, _t1, mode=_mode_str)
                            _sub = _r["submit_result"]
                            _hasil.append({
                                "paket":  _pl_label(_p),
                                "ok":     _sub["ok"],
                                "pesan":  f"HTTP {_sub['status']}",
                                "mulai":  _t1.strftime("%d/%m/%Y %H:%M"),
                            })
                            # Simpan tgl ke Supabase + push GCal
                            if _sub["ok"]:
                                try:
                                    _jad = _r["jadwal_list"]
                                    pl_engine.simpan_paket_pl({
                                        "kode_paket":            _p["kode_paket"],
                                        "tgl_batas_penawaran":   _jad[0]["selesai"].strftime("%Y-%m-%d"),
                                        "tgl_buka_penawaran":    _jad[1]["mulai"].strftime("%Y-%m-%d"),
                                        "tgl_evaluasi":          _jad[2]["selesai"].strftime("%Y-%m-%d"),
                                        "tgl_negosiasi":         _jad[3]["mulai"].strftime("%Y-%m-%d"),
                                        "tgl_penetapan":         _jad[4]["mulai"].strftime("%Y-%m-%d"),
                                    })
                                except Exception:
                                    pass
                                try:
                                    import gcal_pl_helper as _gcalpl
                                    _gcalpl.push_jadwal_pl_ke_gcal(_kp, _p["nama_paket"], _r["jadwal_list"])
                                except Exception:
                                    pass
                        except Exception as _e:
                            _hasil.append({"paket": _pl_label(_p), "ok": False, "pesan": str(_e)[:100]})

                    _prog.empty()
                    _sukses = sum(1 for h in _hasil if h["ok"])
                    _gagal = len(_hasil) - _sukses
                    if _gagal == 0:
                        st.success(f"✅ Semua {_sukses} paket berhasil dijadwalkan!")
                    else:
                        st.warning(f"⚠️ {_sukses} sukses, {_gagal} gagal")
                    for h in _hasil:
                        _ic = "✅" if h["ok"] else "❌"
                        st.markdown(f"{_ic} **{h['paket']}** — {h['pesan']}" + (f" — mulai {h.get('mulai','')}" if h["ok"] else ""))

                    # Expander preview jadwal per paket sukses
                    _hasil_sukses = [h for h in _hasil if h["ok"]]
                    if _hasil_sukses:
                        with st.expander(f"📅 Lihat Detail Jadwal ({len(_hasil_sukses)} paket)", expanded=True):
                            for _ph in _hasil_sukses:
                                _pk_match = next((p for p in _pljd_selected if p["nama_paket"] == _ph["paket"]), None)
                                if not _pk_match:
                                    continue
                                _tgl_preview = st.session_state.get(f"pljd_tgl_{_pk_match['kode_paket']}", _pljd_tgl_global if not _pljd_beda else datetime.now().date())
                                _jam_preview = st.session_state.get(f"pljd_jam_{_pk_match['kode_paket']}", _pljd_jam_global if not _pljd_beda else datetime.strptime("08:00", "%H:%M").time())
                                _t1_preview = datetime.combine(_tgl_preview, _jam_preview)
                                _mode_str_prev = {"Cepat": "cepat", "Standar": "standar"}.get(_pljd_mode, "normal")
                                if _mode_str_prev == "cepat":
                                    _jadwal_preview = _jepl.hitung_jadwal_pl_cepat(_t1_preview)
                                elif _mode_str_prev == "standar":
                                    _jadwal_preview = _jepl.hitung_jadwal_pl_standar(_t1_preview)
                                else:
                                    _jadwal_preview = _jepl.hitung_jadwal_pl(_t1_preview)
                                st.markdown(f"**{_pl_label(_pk_match)}**")
                                import pandas as _pd_jad
                                _jad_rows = []
                                for _idx_jad, _jd in enumerate(_jadwal_preview, 1):
                                    _dur = _jd["selesai"] - _jd["mulai"]
                                    _dur_str = ""
                                    _dur_days = _dur.days
                                    _dur_hours = _dur.seconds // 3600
                                    _dur_mins = (_dur.seconds % 3600) // 60
                                    if _dur_days > 0:
                                        _dur_str = f"{_dur_days} hari"
                                        if _dur_hours > 0:
                                            _dur_str += f" {_dur_hours} jam"
                                    elif _dur_hours > 0:
                                        _dur_str = f"{_dur_hours} jam {_dur_mins} menit"
                                    else:
                                        _dur_str = f"{_dur_mins} menit"
                                    _jad_rows.append({
                                        "No": _idx_jad,
                                        "Tahap": _jd["nama"],
                                        "Mulai": _jd["mulai"].strftime("%d-%m-%Y %H:%M"),
                                        "Selesai": _jd["selesai"].strftime("%d-%m-%Y %H:%M"),
                                        "Durasi": _dur_str,
                                    })
                                st.dataframe(_pd_jad.DataFrame(_jad_rows), use_container_width=True, hide_index=True)

                with st.expander("ℹ️ Libur Nasional Tersisa"):
                    _hari_ini = datetime.now().date()
                    _sisa = sorted(d for d in _libur_map_pl if d >= _hari_ini)
                    for d in _sisa[:15]:
                        st.write(f"• {_HARI_NAMA[d.weekday()]}, {d.day} {_BULAN_NAMA[d.month-1]} {d.year} — {_libur_map_pl[d]}")

        st.divider()
        st.markdown("#### 🔄 Sync Jadwal ke Google Calendar")
        st.caption("Baca jadwal aktual dari SPSE → update GCal + Supabase tgl_evaluasi/tgl_negosiasi/tgl_penetapan. Jalankan setelah ada perubahan jadwal di SPSE.")
        import gcal_pl_helper as _gcalpl_tc
        _sync_gcal_pl_btn = False
        if not _gcalpl_tc.check_gcal_token():
            st.warning("🔐 Token Google Calendar tidak valid atau expired.")
            if st.button("🔑 Login Ulang ke Google Calendar", key="reauth_gcal_btn_pk", type="primary", use_container_width=True):
                import gcal_helper as _gcalh_ra
                try:
                    _gcalh_ra.generate_token()
                    st.success("✅ Token diperbarui! Klik Sync untuk melanjutkan.")
                    st.rerun()
                except Exception as _e_ra:
                    st.error(f"❌ Gagal reauth: {_e_ra}")
        else:
            _sync_gcal_pl_btn = st.button("🔄 Sync Jadwal ke GCal", key="sync_gcal_pl_btn_pk", use_container_width=True, type="primary")
        if _sync_gcal_pl_btn:
            import gcal_pl_helper as _gcalpl
            _gcalpl_prog = st.progress(0.0, text="Memulai sync...")
            _gcalpl_results = _gcalpl.sync_semua_paket_pl(
                progress_cb=lambda f, m: _gcalpl_prog.progress(f, text=m)
            )
            _gcalpl_prog.empty()
            _gcalpl_ok = sum(1 for r in _gcalpl_results if r["ok"])
            _gcalpl_skip = sum(1 for r in _gcalpl_results if not r["ok"] and "kosong" in r.get("error", ""))
            _gcalpl_err = len(_gcalpl_results) - _gcalpl_ok - _gcalpl_skip
            if _gcalpl_err == 0:
                st.success(f"✅ {_gcalpl_ok} paket sync OK, {_gcalpl_skip} skip (jadwal belum diisi SPSE).")
            else:
                st.warning(f"⚠️ {_gcalpl_ok} OK, {_gcalpl_skip} skip, {_gcalpl_err} error.")
            _gcalpl_display = [
                {
                    "Paket": _pl_label(r),
                    "Status": "✅" if r["ok"] else ("⏭ Skip" if "kosong" in r.get("error","") else "❌"),
                    "GCal +": r["gcal_inserted"],
                    "GCal -": r["gcal_deleted"],
                    "Tgl Evaluasi": r["tgl_evaluasi"],
                    "Tgl Negosiasi": r["tgl_negosiasi"],
                    "Tgl Penetapan": r["tgl_penetapan"],
                    "Error": r["error"][:60] if r["error"] else "",
                }
                for r in _gcalpl_results
            ]
            st.dataframe(_gcalpl_display, use_container_width=True, hide_index=True)
        st.divider()
        _render_ubah_jadwal_pl(_pljd_rows, _jepl, "pljd_edit_pk")

    # ── Tab 3: Setup Paket PL (LDK + Masa Berlaku + Checklist + Upload Dokpil) ─
    if _pl_active_tab == "3️⃣ Setup Paket":
        st.markdown("### Setup Paket Pengadaan Langsung")
        st.caption(
            "Submit LDK (Persyaratan Kualifikasi) + Masa Berlaku Penawaran + "
            "Checklist Dokumen Penawaran + Upload Dokumen Pemilihan (Dokpil PDF) ke SPSE. "
            "KAK / Rancangan Kontrak / Uraian Singkat / Informasi Lainnya tugas PPK (bukan PP)."
        )

        _depl = _depl_pk  # alias — sudah di-import top-level

        _plsp_rows = _load_draft_pl_cached("PK")
        _plsp_rows, _ = pl_engine.buang_duplikat_paket_lama(_plsp_rows)
        _plsp_rows = [r for r in _plsp_rows if not pl_engine.is_paket_selesai(r)]
        if not _plsp_rows:
            st.info("⚠️ Belum ada paket PL. Serap dari SPSE di Tab 1 terlebih dahulu.")
        else:
            _plsp_col_list, _plsp_col_kanan = st.columns([2, 3])

            with _plsp_col_list:
                st.markdown("### 1. Pilih Paket + Upload Dokpil")
                _plsp_sel_all, _plsp_sel_none = st.columns(2)
                with _plsp_sel_all:
                    if st.button("✅ Semua", key="plsp_sel_all", use_container_width=True):
                        for _rr in _plsp_rows:
                            st.session_state[f"plsp_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _plsp_sel_none:
                    if st.button("⬜ Kosong", key="plsp_sel_none", use_container_width=True):
                        for _rr in _plsp_rows:
                            st.session_state[f"plsp_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _sp = _sp_global  # alias — sudah di-import top-level

                _plsp_selected = []
                for _rr in _plsp_rows:
                    _kp_key = _rr["kode_paket"]
                    _plsp_chk_key  = f"plsp_chk_{_kp_key}_v19"
                    _plsp_file_key = f"plsp_dokpil_{_kp_key}"

                    _col_chk, _col_file = st.columns([3, 2])
                    with _col_chk:
                        if _plsp_chk_key not in st.session_state:
                            st.session_state[_plsp_chk_key] = True
                        _chk = st.checkbox("", key=_plsp_chk_key, label_visibility="collapsed")
                        st.markdown(f"**{_pl_label(_rr)}** ({_rr.get('jenis_pl','?')})")
                    with _col_file:
                        _dokpil_up = st.file_uploader(
                            "Dokpil PDF",
                            type=["pdf"],
                            key=_plsp_file_key,
                            label_visibility="collapsed",
                        )
                        if _dokpil_up:
                            _ku_prev = _rr.get("kode_unik") or "?"
                            _sk_prev = _lookup_singkatan_dinas(_rr.get("satker", ""))
                            # Tanggal: dari DB (tgl_dokpil) → fallback session → hari ini
                            _tgl_db = _rr.get("tgl_dokpil")
                            if _tgl_db:
                                try:
                                    from datetime import date as _date
                                    _tgl_prev = _date.fromisoformat(str(_tgl_db))
                                except Exception:
                                    _tgl_prev = datetime.now().date()
                            else:
                                _tgl_prev = datetime.now().date()
                            # Nomor: dari DB → fallback generate
                            _no_prev = _rr.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                nama_paket=_rr["nama_paket"],
                                kode_unik=_ku_prev,
                                skpd_singkat=_sk_prev,
                                tahun=_tgl_prev.year,
                                paket_ulang=_pl_paket_ulang(_rr),
                                nomor_urut=_rr.get("nomor_urut"),
                            )
                            st.caption(f"📄 {_dokpil_up.name}  \n📋 `{_no_prev}`  \n📅 {_tgl_prev.strftime('%d-%m-%Y')}")
                            if st.button("📤 Upload Dokpil", key=f"plsp_upload_only_{_kp_key}", use_container_width=True):
                                with st.spinner("Mengupload dokpil..."):
                                    try:
                                        _r_up_only = _udpl.upload_dokpil_pl(
                                            kode_paket=_kp_key,
                                            file_bytes=_dokpil_up.getvalue(),
                                            file_name=_dokpil_up.name,
                                            nomor_dokpil=_no_prev,
                                            tgl_dokpil=_tgl_prev.strftime("%d-%m-%Y"),
                                        )
                                        if _r_up_only["ok"]:
                                            from config import sb as _sb_up_only
                                            _sb_up_only().table("draft_paket_pl").update({
                                                "nomor_dokpil": _no_prev,
                                            }).eq("kode_paket", _kp_key).execute()
                                            st.success(f"✅ Upload berhasil — {_no_prev}")
                                        else:
                                            st.error(f"❌ HTTP {_r_up_only.get('status','?')} — {_r_up_only.get('error') or _r_up_only.get('body','')[:300]}")
                                            st.json(_r_up_only)
                                    except Exception as _e_up_only:
                                        st.error(f"❌ Exception: {_e_up_only}")

                    if _chk:
                        _plsp_selected.append({
                            **_rr,
                            "_dokpil_file": _dokpil_up,
                        })

                st.caption(f"**{len(_plsp_selected)}** dari **{len(_plsp_rows)}** paket dipilih")

                # Kumpulkan semua paket yang sudah ada file dokpil
                _all_with_file = [
                    {**_rr, "_dokpil_file": st.session_state.get(f"plsp_dokpil_{_rr['kode_paket']}")}
                    for _rr in _plsp_rows
                    if st.session_state.get(f"plsp_dokpil_{_rr['kode_paket']}")
                ]
                if _all_with_file:
                    st.divider()
                    if st.button(f"📤 Upload Semua Dokpil ({len(_all_with_file)} file)", key="plsp_upload_all_dokpil", use_container_width=True, type="primary"):
                        from config import sb as _sb_upall
                        _cl_upall = _sb_upall()
                        for _rr_up in _all_with_file:
                            _kp_up = _rr_up["kode_paket"]
                            _f_up = _rr_up["_dokpil_file"]
                            _ku_up = _rr_up.get("kode_unik") or "?"
                            _sk_up = _lookup_singkatan_dinas(_rr_up.get("satker", ""))
                            # Tanggal dari DB, fallback session
                            _tgl_db_up = _rr_up.get("tgl_dokpil")
                            if _tgl_db_up:
                                try:
                                    from datetime import date as _date2
                                    _tgl_up = _date2.fromisoformat(str(_tgl_db_up))
                                except Exception:
                                    _tgl_up = datetime.now().date()
                            else:
                                _tgl_up = datetime.now().date()
                            # Nomor dari DB, fallback generate
                            _no_up = _rr_up.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                nama_paket=_rr_up["nama_paket"],
                                kode_unik=_ku_up,
                                skpd_singkat=_sk_up,
                                tahun=_tgl_up.year,
                                paket_ulang=_pl_paket_ulang(_rr_up),
                                nomor_urut=_rr_up.get("nomor_urut"),
                            )
                            try:
                                _r_upall = _udpl.upload_dokpil_pl(
                                    kode_paket=_kp_up,
                                    file_bytes=_f_up.getvalue(),
                                    file_name=_f_up.name,
                                    nomor_dokpil=_no_up,
                                    tgl_dokpil=_tgl_up.strftime("%d-%m-%Y"),
                                )
                                if _r_upall["ok"]:
                                    _cl_upall.table("draft_paket_pl").update({"nomor_dokpil": _no_up}).eq("kode_paket", _kp_up).execute()
                                    st.success(f"✅ {_pl_label(_rr_up)} — {_no_up}")
                                else:
                                    st.error(f"❌ {_pl_label(_rr_up)} HTTP {_r_upall.get('status','?')} — {_r_upall.get('body','')[:100]}")
                            except Exception as _e_upall:
                                st.error(f"❌ {_pl_label(_rr_up)}: {_e_upall}")
                        _load_draft_pl_cached.clear()

            with _plsp_col_kanan:
                st.markdown("### 2. Konfigurasi Setup Paket")

                if not _plsp_selected:
                    st.info("Pilih paket di sebelah kiri.")
                else:
                    # ── SEKSI 1: SBU Global ───────────────────────────────────
                    st.markdown("#### 🏗️ Seksi 1 — SBU Global")
                    _sbu_global_aktif = st.toggle(
                        "SBU Global (apply 1 SBU ke semua paket terpilih)",
                        value=st.session_state.get("pk_sbu_global_aktif", False),
                        key="pk_sbu_global_aktif",
                    )
                    if _sbu_global_aktif:
                        st.caption("Satu pilihan SBU apply ke semua paket terpilih.")

                        _plsp_klas_list = ["(auto-detect dari paket pertama)"] + _sp.list_klasifikasi()

                        _first_p = _plsp_selected[0]
                        _detected_g = _sp.detect_from_draft(
                            _first_p.get("sbu_baru") or "", _first_p.get("sbu_lama") or ""
                        )
                        _g_kode_baru = _detected_g.get("kode_baru", "")
                        _g_kode_lama = _detected_g.get("kode_lama", "")

                        _g_klas_default = 0
                        if _g_kode_baru:
                            _baru_info_g = _sp.get_sbu_baru_by_kode(_g_kode_baru)
                            _klas_det_g = (_baru_info_g or {}).get("klasifikasi", "")
                            if _klas_det_g in _plsp_klas_list:
                                _g_klas_default = _plsp_klas_list.index(_klas_det_g)

                        if _g_kode_baru:
                            st.caption(f"Auto-detect dari **{_pl_label(_first_p)}**: `{_g_kode_baru}` / `{_g_kode_lama}`")

                        _g_picked_klas = st.selectbox(
                            "Klasifikasi",
                            _plsp_klas_list,
                            index=_g_klas_default,
                            key="plsp_global_klas",
                        )

                        if _g_picked_klas and _g_picked_klas != "(auto-detect dari paket pertama)":
                            _g_baru_options = _sp.list_sbu_baru_by_klasifikasi(_g_picked_klas)
                        else:
                            _g_baru_options = []
                            if _g_kode_baru:
                                _g_baru_options = [_sp.get_sbu_baru_by_kode(_g_kode_baru)]
                        _g_baru_labels = [
                            f"{b['kode']} — {(b.get('nama_singkat') or b.get('nama_full',''))[:70]}"
                            for b in _g_baru_options if b
                        ]
                        _g_baru_default = 0
                        for _gi, _gb in enumerate(_g_baru_options):
                            if _gb and _gb.get("kode") == _g_kode_baru:
                                _g_baru_default = _gi
                                break
                        _g_picked_baru_label = st.selectbox(
                            "SBU Baru (KBLI 2020)",
                            _g_baru_labels or ["(pilih klasifikasi dulu)"],
                            index=_g_baru_default if _g_baru_labels else 0,
                            key="plsp_global_sbu_baru",
                        )
                        _g_picked_baru_kode = (
                            _g_picked_baru_label.split(" — ", 1)[0]
                            if _g_baru_labels and " — " in _g_picked_baru_label else ""
                        )

                        _g_lama_options = _sp.list_sbu_lama_padanan(_g_picked_baru_kode) if _g_picked_baru_kode else []
                        _g_lama_labels = ["(tidak dipersyaratkan / hanya SBU 2020)"] + [
                            f"{l['kode']} — {(l.get('nama_singkat') or l.get('nama_full',''))[:70]}"
                            for l in _g_lama_options
                        ]
                        _g_lama_default = 0
                        for _gli, _gl in enumerate(_g_lama_options):
                            if _gl.get("kode") == _g_kode_lama:
                                _g_lama_default = _gli + 1
                                break
                        _g_picked_lama_label = st.selectbox(
                            "SBU Lama (KBLI 2017) — opsional",
                            _g_lama_labels,
                            index=_g_lama_default,
                            key="plsp_global_sbu_lama",
                        )
                        _g_picked_lama_kode = (
                            _g_picked_lama_label.split(" — ", 1)[0]
                            if " — " in _g_picked_lama_label else ""
                        )

                        _sbu_baru_global = ""
                        _sbu_lama_global = ""
                        if _g_picked_baru_kode:
                            _baru_obj_g = _sp.get_sbu_baru_by_kode(_g_picked_baru_kode)
                            _sbu_baru_global = (_baru_obj_g or {}).get("nama_full", "")
                        if _g_picked_lama_kode:
                            _lama_obj_g = _sp.get_sbu_lama_by_kode(_g_picked_lama_kode)
                            _sbu_lama_global = (_lama_obj_g or {}).get("nama_full", "")
                        if not _sbu_baru_global:
                            _sbu_baru_global = _first_p.get("sbu_baru") or ""

                        if _sbu_baru_global:
                            st.caption(f"🔹 Baru: `{_sbu_baru_global[:80]}`")
                        if _sbu_lama_global:
                            st.caption(f"🔸 Lama: `{_sbu_lama_global[:80]}`")
                        elif _sbu_baru_global:
                            st.caption("ℹ️ SBU Lama tidak dipersyaratkan — hanya SBU 2020 di LDK")

                        if st.button(
                            f"💾 Simpan SBU Global ke {len(_plsp_selected)} paket",
                            key="plsp_save_sbu_btn_pk", use_container_width=True,
                        ):
                            from config import sb as _sb_factory
                            _client_sbu = _sb_factory()
                            _ok_sbu = 0
                            for _p in _plsp_selected:
                                try:
                                    _client_sbu.table("draft_paket_pl").update({
                                        "sbu_baru": _sbu_baru_global,
                                        "sbu_lama": _sbu_lama_global,
                                    }).eq("kode_paket", _p["kode_paket"]).execute()
                                    _ok_sbu += 1
                                except Exception as _e:
                                    st.error(f"❌ {_pl_label(_p)}: {_e}")
                            st.success(f"✅ {_ok_sbu}/{len(_plsp_selected)} paket disimpan ke Supabase")
                    else:
                        st.caption("ℹ️ Mode custom — pilih SBU yang pernah dipakai atau input baru.")
                        _pk_sbu_hist = _load_pl_sbu_history()
                        _pk_sbu_new = "✏️ Input SBU baru"
                        _pk_sbu_opts = [f"{x['baru'][:100]}" for x in _pk_sbu_hist] + [_pk_sbu_new]
                        _pk_sbu_pick = st.selectbox("SBU Custom — histori", _pk_sbu_opts or [_pk_sbu_new], key="pk_custom_sbu_pick")
                        _pk_sbu_idx = _pk_sbu_opts.index(_pk_sbu_pick) if _pk_sbu_pick in _pk_sbu_opts else len(_pk_sbu_hist)
                        if _pk_sbu_idx < len(_pk_sbu_hist):
                            _pk_sbu_selected = _pk_sbu_hist[_pk_sbu_idx]
                            _sbu_baru_global = _pk_sbu_selected["baru"]
                            _sbu_lama_global = _pk_sbu_selected["lama"] or None
                            st.caption(f"🔹 Baru: `{_sbu_baru_global}`")
                            if _sbu_lama_global:
                                st.caption(f"🔸 Lama: `{_sbu_lama_global}`")
                        else:
                            _sbu_baru_global = st.text_input(
                                "SBU Baru (teks bebas)", key="pk_custom_sbu_baru",
                                placeholder="Contoh: RE201 — Jasa Desain Rekayasa untuk Konstruksi Pondasi serta Struktur Bangunan",
                            ) or None
                            _sbu_lama_global = st.text_input(
                                "SBU Lama (teks bebas, opsional)", key="pk_custom_sbu_lama",
                                placeholder="Kosongkan jika tidak dipersyaratkan",
                            ) or None
                        if st.button(
                            f"💾 Simpan SBU Custom + POST LDK ke {len(_plsp_selected)} paket",
                            key="pk_save_sbu_custom_btn", use_container_width=True,
                            disabled=not _sbu_baru_global,
                        ):
                            from config import sb as _sb_factory_ck
                            _client_c = _sb_factory_ck()
                            _ok_c = 0
                            for _p in _plsp_selected:
                                try:
                                    _client_c.table("draft_paket_pl").update({
                                        "sbu_baru": _sbu_baru_global or "",
                                        "sbu_lama": _sbu_lama_global or "",
                                    }).eq("kode_paket", _p["kode_paket"]).execute()
                                    _ok_c += 1
                                except Exception as _e:
                                    st.error(f"❌ {_pl_label(_p)}: {_e}")
                            _save_pl_sbu_history(_sbu_baru_global, _sbu_lama_global or "")
                            st.success(f"✅ {_ok_c}/{len(_plsp_selected)} paket disimpan ke Supabase")
                            import ldk_config as _ldk_cfg_custom_pk
                            for _p in _plsp_selected:
                                try:
                                    _r_custom_ldk = _depl_pk.submit_ldk_pl(
                                        _p["kode_paket"],
                                        sbu_baru=_sbu_baru_global or "",
                                        sbu_lama=_sbu_lama_global or "",
                                        centang_admin_ckm_ids=["401", "402", "403", "404"],
                                        teknis_centang_ckm_ids=["437", "438", "439"],
                                        kinerja_text=_ldk_cfg_custom_pk.KINERJA_PENYEDIA_PK,
                                    )
                                    st.write(f"{'✅' if _r_custom_ldk['ok'] else '❌'} {_pl_label(_p)} — POST LDK HTTP {_r_custom_ldk['status']}")
                                except Exception as _e_ldk_custom:
                                    st.error(f"❌ {_pl_label(_p)} — POST LDK gagal: {_e_ldk_custom}")

                    st.divider()

                    # ── SEKSI 2: Masa Berlaku ────────────────────────────────
                    st.markdown("#### 📅 Seksi 2 — Masa Berlaku Penawaran")
                    st.caption("Tanggal Dokpil diisi langsung pada panel Excel @ Master Data → INPUT TANGGAL DOKPIL.")

                    _ldk_masa_berlaku = st.number_input(
                        "Masa Berlaku Penawaran (hari)",
                        min_value=1, max_value=180, value=30,
                        key="plsp_masa_berlaku",
                    )

                    if st.button("💾 Submit Masa Berlaku Penawaran", key="plsp_btn_masa_berlaku", use_container_width=True):
                        from config import sb as _sb_factory_mb
                        _client_mb = _sb_factory_mb()
                        for _p in _plsp_selected:
                            try:
                                _client_mb.table("draft_paket_pl").update({
                                    "masa_berlaku": int(_ldk_masa_berlaku),
                                }).eq("kode_paket", _p["kode_paket"]).execute()
                            except Exception as _e_mb:
                                st.warning(f"⚠️ Gagal simpan masa berlaku {_pl_label(_p)}: {_e_mb}")
                            _r_mb = _depl.submit_masa_berlaku_pl(_p["kode_paket"], int(_ldk_masa_berlaku))
                            st.write(f"{'✅' if _r_mb['ok'] else '❌'} {_pl_label(_p)} — HTTP {_r_mb['status']}")

                    st.divider()

                    # ── SEKSI 3: Dokumen Kualifikasi (LDK) ───────────────────
                    st.markdown("#### 📋 Seksi 3 — Dokumen Kualifikasi (LDK)")
                    st.caption("ℹ️ Di-submit ke SPSE bagian Persyaratan Kualifikasi (LDK).")

                    st.markdown("**Syarat Administrasi PLPK** *(ID mengikuti skema PLPK SPSE)*")
                    _ADMIN_LABEL = {
                        "401": "401 — Persyaratan administrasi",
                        "402": "402 — Persyaratan administrasi",
                        "403": "403 — Persyaratan administrasi",
                        "404": "404 — Persyaratan administrasi",
                        "410": "410 — Persyaratan administrasi tambahan",
                        "411": "411 — Persyaratan administrasi tambahan",
                    }
                    _ldk_centang_admin_ckm_ids = []
                    _cols_adm = st.columns(2)
                    for _idx_iter, (_cid, _lbl) in enumerate(_ADMIN_LABEL.items()):
                        with _cols_adm[_idx_iter % 2]:
                            _default_adm = _cid in ("401", "402", "403", "404")
                            if st.checkbox(_lbl, value=_default_adm, key=f"plsp_admin_cid_{_cid}"):
                                _ldk_centang_admin_ckm_ids.append(_cid)

                    st.markdown("**Syarat Teknis PLPK** *(ID mengikuti skema PLPK SPSE)*")
                    _TEKNIS_LABEL = {
                        "437": "437 — Persyaratan teknis",
                        "438": "438 — Persyaratan teknis",
                        "439": "439 — Persyaratan teknis",
                        "440": "440 — Persyaratan teknis",
                        "441": "441 — Persyaratan teknis",
                    }
                    _ldk_teknis_ckm_ids = []
                    _cols_tk = st.columns(2)
                    for _idx_iter, (_cid, _lbl) in enumerate(_TEKNIS_LABEL.items()):
                        with _cols_tk[_idx_iter % 2]:
                            _default = _cid in ("437", "438", "439")
                            if st.checkbox(_lbl, value=_default, key=f"plsp_teknis_cid_{_cid}"):
                                _ldk_teknis_ckm_ids.append(_cid)

                    import ldk_config as _ldk_cfg_pl
                    # Kinerja penyedia wajib — text beda JKK vs PK
                    _jenis_pl_paket = _plsp_selected[0].get("jenis_pl", "JKK").upper() if _plsp_selected else "JKK"
                    _ldk_kinerja_text = _ldk_cfg_pl.KINERJA_PENYEDIA_JKK if _jenis_pl_paket == "JKK" else _ldk_cfg_pl.KINERJA_PENYEDIA_PK

                    st.caption(
                        "ℹ️ PLPK mengikuti checklist Tender PK: admin inti 401–404, teknis inti 437–439. "
                        "ID tambahan 410–411/440–441 opsional. Kinerja Penyedia wajib dikirim otomatis."
                    )

                    if st.button("📋 Submit Dokumen Kualifikasi (LDK)", key="plsp_btn_ldk", use_container_width=True):
                        from config import sb as _sb_factory_ldk
                        _client_ldk = _sb_factory_ldk()
                        for _p in _plsp_selected:
                            try:
                                _client_ldk.table("draft_paket_pl").update({
                                    "sbu_baru": _sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    "sbu_lama": _sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                }).eq("kode_paket", _p["kode_paket"]).execute()
                            except Exception:
                                pass
                            _r_ldk = _depl.submit_ldk_pl(
                                _p["kode_paket"],
                                sbu_baru=_sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                sbu_lama=_sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                centang_admin_ckm_ids=_ldk_centang_admin_ckm_ids,
                                teknis_centang_ckm_ids=_ldk_teknis_ckm_ids,
                                kinerja_text=_ldk_kinerja_text,
                            )
                            st.write(f"{'✅' if _r_ldk['ok'] else '❌'} {_pl_label(_p)} — HTTP {_r_ldk['status']}")

                    st.divider()

                    # Seksi 4 — Checklist Dokumen Penawaran (hardcode semua wajib)
                    _cd_centang_admin = True
                    _cd_centang_syarat = True
                    _cd_centang_harga = True

                    st.divider()
                    st.caption("⬇️ Atau jalankan semua seksi sekaligus:")

                    # ── Submit All-in-One ─────────────────────────────────────
                    if st.button(
                        f"🚀 Push Setup ke SPSE ({len(_plsp_selected)} paket)",
                        key="plsp_submit_btn",
                        type="primary",
                        use_container_width=True,
                    ):
                        _hasil_sp = []
                        _prog_sp = st.progress(0, text="Mulai...")
                        from config import sb as _sb_factory_sp
                        _client_sp = _sb_factory_sp()
                        for _i, _p in enumerate(_plsp_selected):
                            _kp = _p["kode_paket"]
                            _id_nt = _p.get("id_nontender")
                            _nm = _pl_label(_p)
                            _prog_sp.progress((_i + 1) / len(_plsp_selected),
                                              text=f"{_nm} ({_i+1}/{len(_plsp_selected)})...")

                            # 0. Simpan SBU global ke Supabase; tanggal Dokpil berasal dari Excel
                            try:
                                _client_sp.table("draft_paket_pl").update({
                                    "sbu_baru": _sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    "sbu_lama": _sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                }).eq("kode_paket", _kp).execute()
                            except Exception as _e_save:
                                _hasil_sp.append({"paket": _nm, "step": "Simpan Supabase", "ok": False, "pesan": str(_e_save)[:80]})

                            # 1. Submit LDK (kode_paket, bukan id_nontender)
                            try:
                                _r_ldk = _depl.submit_ldk_pl(
                                    _kp,
                                    sbu_baru=_sbu_baru_global if _sbu_baru_global is not None else (_p.get("sbu_baru") or ""),
                                    sbu_lama=_sbu_lama_global if _sbu_lama_global is not None else (_p.get("sbu_lama") or ""),
                                    centang_admin_ckm_ids=_ldk_centang_admin_ckm_ids,
                                    teknis_centang_ckm_ids=_ldk_teknis_ckm_ids,
                                    kinerja_text=_ldk_kinerja_text,
                                )
                                _ijin_note = f" | ijin CDP: {_r_ldk.get('ijin_update','—')}" if _r_ldk.get("ijin_update") else ""
                                _hasil_sp.append({
                                    "paket": _nm, "step": "LDK",
                                    "ok": _r_ldk["ok"], "pesan": f"HTTP {_r_ldk['status']}{_ijin_note}",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "LDK", "ok": False, "pesan": str(_e)[:80]})

                            # 2. Masa berlaku penawaran
                            try:
                                _r_mb = _depl.submit_masa_berlaku_pl(_kp, int(_ldk_masa_berlaku))
                                try:
                                    from config import sb as _sb_mb_bulk
                                    _sb_mb_bulk().table("draft_paket_pl").update({"masa_berlaku": int(_ldk_masa_berlaku)}).eq("kode_paket", _kp).execute()
                                except Exception:
                                    pass
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Masa Berlaku",
                                    "ok": _r_mb["ok"], "pesan": f"HTTP {_r_mb['status']} ({_ldk_masa_berlaku} hari)",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "Masa Berlaku", "ok": False, "pesan": str(_e)[:80]})

                            # 3. Checklist Dokumen Penawaran
                            try:
                                _r_cd = _depl.submit_checklist_pl(
                                    _kp,
                                    centang_admin_all=_cd_centang_admin,
                                    centang_syarat_all=_cd_centang_syarat,
                                    centang_harga_all=_cd_centang_harga,
                                )
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Checklist Dok Penawaran",
                                    "ok": _r_cd["ok"], "pesan": f"HTTP {_r_cd['status']}",
                                })
                            except Exception as _e:
                                _hasil_sp.append({"paket": _nm, "step": "Checklist Dok Penawaran", "ok": False, "pesan": str(_e)[:80]})

                            # 4. Upload Dokpil PDF (jika ada file)
                            _dokpil_file = _p.get("_dokpil_file")
                            if _dokpil_file and _id_nt:
                                try:
                                    _tgl_excel = _p.get("tgl_dokpil")
                                    if _tgl_excel:
                                        from datetime import date as _date_excel
                                        _tgl_excel = _date_excel.fromisoformat(str(_tgl_excel)[:10])
                                    else:
                                        _tgl_excel = datetime.now().date()
                                    # Generate Nomor Dokpil: 000.3.3/01/PL/PP-NN/{KodeUnik}/{SkpdSingkat}/{Tahun}
                                    _kode_unik = _p.get("kode_unik") or ""
                                    _skpd_singkat = _lookup_singkatan_dinas(_p.get("satker", ""))
                                    _nomor_dokpil = _p.get("nomor_dokpil") or _udpl.generate_nomor_dokpil(
                                        nama_paket=_p["nama_paket"],
                                        kode_unik=_kode_unik,
                                        skpd_singkat=_skpd_singkat,
                                        tahun=_tgl_excel.year,
                                        paket_ulang=_pl_paket_ulang(_p),
                                        nomor_urut=_p.get("nomor_urut"),
                                    )
                                    _r_up = _udpl.upload_dokpil_pl(
                                        kode_paket=_kp,
                                        file_bytes=_dokpil_file.getvalue(),
                                        file_name=_dokpil_file.name,
                                        nomor_dokpil=_nomor_dokpil,
                                        tgl_dokpil=_tgl_excel.strftime("%d-%m-%Y"),
                                    )
                                    _hasil_sp.append({
                                        "paket": _nm, "step": "Upload Dokpil",
                                        "ok": _r_up["ok"],
                                        "pesan": f"HTTP {_r_up.get('status','?')} | {_nomor_dokpil}",
                                    })
                                    if _r_up["ok"]:
                                        try:
                                            _client_sp.table("draft_paket_pl").update({
                                                "nomor_dokpil": _nomor_dokpil,
                                            }).eq("kode_paket", _kp).execute()
                                        except Exception:
                                            pass
                                except Exception as _e:
                                    _hasil_sp.append({
                                        "paket": _nm, "step": "Upload Dokpil",
                                        "ok": False, "pesan": str(_e)[:80],
                                    })
                            elif _dokpil_file and not _id_nt:
                                _hasil_sp.append({
                                    "paket": _nm, "step": "Upload Dokpil",
                                    "ok": False, "pesan": "id_nontender kosong, tidak bisa upload",
                                })

                        _prog_sp.empty()
                        _sukses_sp = sum(1 for h in _hasil_sp if h["ok"])
                        _gagal_sp = len(_hasil_sp) - _sukses_sp
                        if _gagal_sp == 0:
                            st.success(f"✅ Semua {_sukses_sp} operasi sukses!")
                        else:
                            st.warning(f"⚠️ {_sukses_sp} sukses, {_gagal_sp} gagal")
                        _load_draft_pl_cached.clear()

                        # Tampilkan log per paket
                        import pandas as _pd
                        _df_sp = _pd.DataFrame(_hasil_sp)
                        if not _df_sp.empty:
                            _df_sp["status"] = _df_sp["ok"].map({True: "✅", False: "❌"})
                            st.dataframe(
                                _df_sp[["status", "paket", "step", "pesan"]],
                                use_container_width=True, hide_index=True,
                            )


    if _pl_active_tab == "4️⃣ Pilih Penyedia & Umumkan":
        st.divider()
        st.markdown("### 🏢 Pilih Penyedia ke SPSE")
        st.caption(
            "Cari penyedia by NPWP → klik pilih ke SPSE (prioritas kabupaten Tapin, "
            "fallback semua kabupaten Kalsel propinsi 22)."
        )

        if st.button("🔄 Refresh Data Penyedia", key="pp_refresh_penyedia_pk", help="Parse ulang 4. Informasi Lainnya/8. ND.pdf (fallback Draft_PL) → update nama & NPWP penyedia di Supabase"):
            _ref_bar_pk = st.progress(0.0, text="Memulai...")
            _ref_log_pk = st.empty()
            import parse_kak_pl as _pkpl_ref_pk
            def _ref_cb_pk(p, m):
                _ref_bar_pk.progress(min(float(p), 1.0), text=m[:120])
                _ref_log_pk.caption(m)
            _ref_res_pk = _pkpl_ref_pk.serap_penyedia_pl(progress_cb=_ref_cb_pk)
            _ref_bar_pk.progress(1.0, text="Selesai")
            st.success(f"✅ {_ref_res_pk.get('updated',0)} diperbarui, {_ref_res_pk.get('not_found',0)} tidak ditemukan, {len(_ref_res_pk.get('errors',[]))} error")
            if _ref_res_pk.get("errors"):
                st.warning("\n".join(_ref_res_pk["errors"][:5]))
            _load_draft_pl_cached.clear()
            st.rerun()
        _pp_rows = _load_draft_pl_cached("PK")
        _pp_rows, _ = pl_engine.buang_duplikat_paket_lama(_pp_rows)
        _pp_rows = [r for r in _pp_rows if not pl_engine.is_paket_selesai(r)]
        if _pp_rows:
            import pilih_penyedia_pl as _ppp

            _pp_col_list, _pp_col_act = st.columns([2, 3])

            with _pp_col_list:
                st.markdown("**Pilih paket:**")
                _pp_sel_all, _pp_sel_none = st.columns(2)
                with _pp_sel_all:
                    if st.button("✅ Semua", key="pp_sel_all", use_container_width=True):
                        for _rr in _pp_rows:
                            st.session_state[f"pp_chk_{_rr['kode_paket']}_v19"] = True
                        st.rerun()
                with _pp_sel_none:
                    if st.button("⬜ Kosong", key="pp_sel_none", use_container_width=True):
                        for _rr in _pp_rows:
                            st.session_state[f"pp_chk_{_rr['kode_paket']}_v19"] = False
                        st.rerun()

                _pp_selected = []
                for _rr in _pp_rows:
                    _kp = _rr["kode_paket"]
                    _npwp_disp = _rr.get("npwp_penyedia") or "—"
                    _nama_disp = _rr.get("nama_penyedia") or "—"
                    _pp_chk_key = f"pp_chk_{_kp}_v19"
                    if _pp_chk_key not in st.session_state:
                        st.session_state[_pp_chk_key] = True
                    _chk = st.checkbox("", key=_pp_chk_key, label_visibility="collapsed", help=f"Penyedia: {_nama_disp} | NPWP: {_npwp_disp}")
                    st.markdown(f"**{_pl_label(_rr)}**")
                    if _chk:
                        _pp_selected.append(_rr)

                st.caption(f"**{len(_pp_selected)}** paket dipilih")

            with _pp_col_act:
                if not _pp_selected:
                    st.info("Pilih paket di sebelah kiri.")
                else:
                    # Tabel ringkas paket terpilih
                    import pandas as _pd2
                    _pp_df = _pd2.DataFrame([{
                        "Paket": _pl_label(r),
                        "Penyedia": r.get("nama_penyedia") or "—",
                        "NPWP": r.get("npwp_penyedia") or "—",
                    } for r in _pp_selected])
                    st.dataframe(_pp_df, use_container_width=True, hide_index=True)

                    _invalid = [r for r in _pp_selected if not r.get("npwp_penyedia")]
                    if _invalid:
                        st.warning(
                            f"⚠️ {len(_invalid)} paket belum ada NPWP penyedia: "
                            + ", ".join(_pl_label(r) for r in _invalid)
                        )

                    _valid_pp = [r for r in _pp_selected if r.get("npwp_penyedia")]
                    if _valid_pp:
                        if st.button(
                            f"🏢 Pilih Semua Penyedia ke SPSE ({len(_valid_pp)} paket)",
                            key="pp_submit_btn",
                            type="primary",
                            use_container_width=True,
                        ):
                            import spse_browser as _spse_br
                            _ck_pp = _spse_br.get_spse_cookies()
                            _base_pp = pl_engine.BASE_URL

                            _pp_hasil = []
                            _pp_prog = st.progress(0, text="Mulai pilih penyedia...")
                            for _i_pp, _pp_r in enumerate(_valid_pp):
                                _pp_nm = _pl_label(_pp_r)
                                _pp_prog.progress(
                                    (_i_pp + 1) / len(_valid_pp),
                                    text=f"{_pp_nm} ({_i_pp+1}/{len(_valid_pp)})...",
                                )
                                try:
                                    _res_pp = _ppp.cari_dan_pilih_penyedia(
                                        kode_paket=_pp_r["kode_paket"],
                                        npwp=_pp_r.get("npwp_penyedia") or "",
                                        cookie_str=_ck_pp,
                                        base_url=_base_pp,
                                        nama_penyedia=_pp_r.get("nama_penyedia") or "",
                                    )
                                    _pp_hasil.append({
                                        "paket": _pp_nm,
                                        "ok": _res_pp["ok"],
                                        "pesan": (
                                            f"✅ {_res_pp.get('nama','?')} (kab {_res_pp.get('kabupaten_id','')})"
                                            if _res_pp["ok"]
                                            else f"❌ {_res_pp.get('pesan','?')}"
                                        ),
                                    })
                                except Exception as _e_pp:
                                    _pp_hasil.append({
                                        "paket": _pp_nm,
                                        "ok": False,
                                        "pesan": f"❌ Error: {str(_e_pp)[:80]}",
                                    })

                            _pp_prog.empty()
                            _pp_ok = sum(1 for h in _pp_hasil if h["ok"])
                            if _pp_ok == len(_pp_hasil):
                                st.success(f"✅ {_pp_ok}/{len(_pp_hasil)} paket berhasil dipilih penyedia!")
                            else:
                                st.warning(f"⚠️ {_pp_ok}/{len(_pp_hasil)} sukses")

                            _df_pp = _pd2.DataFrame(_pp_hasil)
                            if not _df_pp.empty:
                                st.dataframe(
                                    _df_pp[["paket", "pesan"]],
                                    use_container_width=True, hide_index=True,
                                )

    # ── Tab 7: Kirim Verifikasi Penyedia ─────────────────────────────────────
    # ── Tab 4 Section 2: Pilih Penyedia ke SPSE ─────────────────────────────
    # ── Tab 3 Section: Umumkan Paket Non Tender (PL PK) ─────────────────────
    if _pl_active_tab == "4️⃣ Pilih Penyedia & Umumkan":
        st.divider()
        st.markdown("### 📢 Umumkan Paket Non Tender")
        st.caption("Setujui Pakta Integritas dan umumkan paket ke SPSE. Pastikan browser SPSE sudah terhubung.")
        _paket_berfolder_umum_pk = [r for r in _pp_rows if r.get("kode_paket") and r.get("folder_dibuat")]
        if not _paket_berfolder_umum_pk:
            st.info("Tidak ada paket berfolder yang bisa diumumkan.")
        else:
            _umum_col1_pk, _umum_col2_pk = st.columns(2)
            with _umum_col1_pk:
                if st.button("✅ Semua", key="umum_sel_all_pk", use_container_width=True):
                    for _r in _paket_berfolder_umum_pk:
                        st.session_state[f"umum_chk_{_r['kode_paket']}"] = True
                    st.rerun()
            with _umum_col2_pk:
                if st.button("⬜ Kosong", key="umum_sel_none_pk", use_container_width=True):
                    for _r in _paket_berfolder_umum_pk:
                        st.session_state[f"umum_chk_{_r['kode_paket']}"] = False
                    st.rerun()
            _pilih_umum_pk = []
            for _r in _paket_berfolder_umum_pk:
                _umum_key_pk = f"umum_chk_{_r['kode_paket']}"
                _umum_chk_pk = st.checkbox(
                    _pl_label(_r),
                    value=st.session_state.get(_umum_key_pk, True),
                    key=_umum_key_pk,
                )
                if _umum_chk_pk:
                    _pilih_umum_pk.append(_r["kode_paket"])
            if st.button("📢 Umumkan Paket Terpilih", key="btn_umumkan_pl_pk", disabled=not _pilih_umum_pk):
                try:
                    import spse_browser as _spse_br_umum_pk
                    _spse_br_umum_pk.buka_browser(navigate=False)
                    _cookie_umum_pk = _spse_br_umum_pk.get_spse_cookies()
                except Exception as _e_br_umum_pk:
                    st.error(f"Browser SPSE tidak terhubung: {_e_br_umum_pk}")
                    _cookie_umum_pk = None
                if _cookie_umum_pk:
                    for _kp_umum_pk in _pilih_umum_pk:
                        _r_umum_pk = next((r for r in _paket_berfolder_umum_pk if r["kode_paket"] == _kp_umum_pk), {"nama_paket": _kp_umum_pk})
                        _nm_umum_pk = _pl_label(_r_umum_pk)
                        _ru_pk = pl_engine.umumkan_paket_pl(_kp_umum_pk, _cookie_umum_pk)
                        if _ru_pk["ok"]:
                            st.success(f"✅ {_nm_umum_pk[:60]} — {_ru_pk['pesan']}")
                        else:
                            st.error(f"❌ {_nm_umum_pk[:60]} — {_ru_pk['pesan']}")

    if _pl_active_tab == "8️⃣ Kirim Verifikasi":
        import verifikasi_penyedia_pl as _verif_pl
        from gcal_helper import get_jadwal_klarifikasi_pl as _gcal_klarifikasi


        # Dropdown paket — gabung draft_pl + aktif_pl (dedup by id_nontender)
        _verif_rows, _verif_load_error = _load_verifikasi_pl_rows_cached("PK")
        if _verif_load_error:
            st.error(f"Gagal load paket PL: {_verif_load_error}")

        st.markdown("## 📨 Kirim Undangan Verifikasi Penyedia")
        st.caption("Centang paket yang ingin dikirim. Hanya paket dengan peserta terdaftar yang tampil.")

        # Buang duplikat row lama (paket di-ulang → kode baru, row lama nyangkut)
        _verif_rows, _verif_dup_n = pl_engine.buang_duplikat_paket_lama(_verif_rows)

        # Load status peserta untuk filter
        _batch_rows = _verif_rows  # sudah di-load di atas
        if _batch_rows:
            # Fetch jumlah peserta semua paket
            _batch_kodes = tuple(r["kode_paket"] for r in _batch_rows if r.get("kode_paket"))
            _batch_mon = _fetch_status_semua_paket_cached(_batch_kodes)

            # Filter hanya paket yang ada peserta
            _batch_eligible = [
                r for r in _batch_rows
                if _batch_mon.get(r["kode_paket"], {}).get("jumlah", 0) > 0
                and r.get("id_nontender")
            ]

            if not _batch_eligible:
                st.info("Belum ada paket dengan peserta terdaftar.")
            else:
                st.caption(f"{len(_batch_eligible)} paket tersedia (sudah ada peserta).")

                # Helper: paket dianggap selesai jika terkirim ATAU sudah di tahap akhir SPSE
                _TAHAP_SELESAI = {"Penandatanganan Kontrak", "Paket Sudah Selesai"}
                def _is_selesai(r):
                    if r.get("status_undangan_verifikasi") == "terkirim":
                        return True
                    if r.get("tahap_spse") in _TAHAP_SELESAI:
                        return True
                    return False

                _n_selesai = sum(1 for r in _batch_eligible if _is_selesai(r))
                _n_belum = len(_batch_eligible) - _n_selesai
                _sembunyikan_selesai = st.checkbox(
                    f"Sembunyikan yang sudah selesai ({_n_selesai} paket)",
                    value=True,
                    key="tab7_sembunyikan_selesai"
                )
                if _sembunyikan_selesai:
                    _batch_tampil = [r for r in _batch_eligible if not _is_selesai(r)]
                else:
                    _batch_tampil = _batch_eligible

                if not _batch_tampil:
                    st.success("✅ Semua paket sudah selesai!")
                else:
                    # Waktu — date_input + time_input (pola Tab 3)
                    import datetime as _dtlb
                    _bwc1, _bwc2 = st.columns(2)
                    with _bwc1:
                        st.caption("**Waktu Mulai**")
                        _bv_tgl = st.date_input("Tanggal", value=_dtlb.date.today(), format="DD/MM/YYYY", key="batch_verif_tgl")
                        _bv_jam = st.time_input("Jam", value=_dtlb.time(9, 0), key="batch_verif_jam")
                        st.markdown(f"**{_HARI_NAMA[_bv_tgl.weekday()]}, {_bv_tgl.day} {_BULAN_NAMA[_bv_tgl.month-1]} {_bv_tgl.year}**")
                    with _bwc2:
                        st.caption("**Waktu Selesai**")
                        _bv_tgl_end = st.date_input("Tanggal", value=_dtlb.date.today(), format="DD/MM/YYYY", key="batch_verif_tgl_end")
                        _bv_jam_end = st.time_input("Jam", value=_dtlb.time(15, 0), key="batch_verif_jam_end")
                    _batch_start = f"{_bv_tgl.strftime('%d-%m-%Y')} {_bv_jam.strftime('%H:%M')}"
                    _batch_end   = f"{_bv_tgl_end.strftime('%d-%m-%Y')} {_bv_jam_end.strftime('%H:%M')}"

                    # Checkbox list paket
                    _sudah_kirim = [r for r in _batch_tampil if r.get("status_undangan_verifikasi") == "terkirim"]
                    _belum_kirim = [r for r in _batch_tampil if r.get("status_undangan_verifikasi") != "terkirim"]
                    _ca, _cb = st.columns([4, 1])
                    _ca.markdown("**Pilih paket yang akan dikirim:**")
                    if _sudah_kirim:
                        _ca.caption(f"✅ {len(_sudah_kirim)} sudah terkirim | ⏳ {len(_belum_kirim)} belum")
                    _centang_semua = _cb.checkbox("Centang Semua", value=False, key="batch_centang_semua")
                    _batch_selected = []
                    _bc1, _bc2 = st.columns(2)
                    for _bi, _br in enumerate(_batch_tampil):
                        _ku = _br.get('kode_unik') or _br['kode_paket']
                        _tgl_kirim = _br.get("tgl_undangan_verifikasi")
                        _sudah = _br.get("status_undangan_verifikasi") == "terkirim"
                        _kontrak = _br.get("tahap_spse") == "Penandatanganan Kontrak"
                        _hint_ulang = _pl_hint_ulang(_br)
                        if _sudah and _tgl_kirim:
                            try:
                                _tgl_fmt = _dtlb.datetime.fromisoformat(_tgl_kirim.replace("Z","+00:00")).strftime("%d-%m-%Y")
                            except Exception:
                                _tgl_fmt = _tgl_kirim[:10]
                            _label = f"{_ku} — {_pl_label(_br)} ✅ {_tgl_fmt}"
                        elif _kontrak:
                            _label = f"{_ku} — {_pl_label(_br)} 🔒 Kontrak"
                        else:
                            _label = f"{_ku} — {_pl_label(_br)}"
                        _col = _bc1 if _bi % 2 == 0 else _bc2
                        _default_chk = _centang_semua or (not _sudah and not _kontrak)
                        if _col.checkbox(_label, value=_default_chk, key=f"batch_chk_{_br['kode_paket']}"):
                            _batch_selected.append(_br)

                    st.markdown(f"**{len(_batch_selected)} paket dipilih**")

                    if st.button("📨 Kirim Undangan Verifikasi", key="btn_batch_kirim_konfirm", type="primary", disabled=not _batch_selected):
                        if not _batch_start or not _batch_end:
                            st.error("Waktu mulai dan selesai wajib diisi.")
                        else:
                                import verifikasi_penyedia_pl as _vpl_batch
                                import evaluasi_admin_kualifikasi_pl as _eval_verif_batch
                                _hasil_batch = []
                                _prog = st.progress(0, text="Mengirim...")
                                for _bi2, _bp2 in enumerate(_batch_selected):
                                    _prog.progress((_bi2 + 1) / len(_batch_selected), text=f"Kirim ke {_bp2.get('kode_unik') or _bp2['kode_paket']}...")
                                    _peserta_res = _eval_verif_batch.scrape_peserta_evaluasi(_bp2["kode_paket"])
                                    if not _peserta_res.get("ok") or not _peserta_res.get("peserta"):
                                        _res = {"ok": False, "msg": _peserta_res.get("pesan", "Peserta tidak ditemukan")}
                                    else:
                                        _hasil_peserta = []
                                        for _peserta in _peserta_res["peserta"]:
                                            _hasil_peserta.append(_vpl_batch.kirim_verifikasi(
                                                id_nontender=_peserta["id_nontender"],
                                                kode_paket=_bp2["kode_paket"],
                                                waktu_start=_batch_start,
                                                waktu_end=_batch_end,
                                            ))
                                        _res = {
                                            "ok": all(_r.get("ok") for _r in _hasil_peserta),
                                            "msg": "; ".join(_r.get("msg", "") for _r in _hasil_peserta),
                                        }
                                    _hasil_batch.append({
                                        "paket": _bp2.get("kode_unik") or _bp2["kode_paket"],
                                        "nama": _pl_label(_bp2),
                                        "ok": _res["ok"],
                                        "msg": _res["msg"],
                                    })
                                _prog.empty()

                                # Tampilkan hasil
                                _ok_list = [h for h in _hasil_batch if h["ok"]]
                                _fail_list = [h for h in _hasil_batch if not h["ok"]]
                                if _ok_list:
                                    st.success(f"✅ Berhasil: {len(_ok_list)} paket")
                                    for h in _ok_list:
                                        st.write(f"  ✅ {h['paket']} — {h['nama']}")
                                if _fail_list:
                                    st.error(f"❌ Gagal: {len(_fail_list)} paket")
                                    for h in _fail_list:
                                        st.write(f"  ❌ {h['paket']} — {h['nama']}: {h['msg']}")

    # ── Tab 8: Upload BA PL ───────────────────────────────────────────────────
    if _pl_active_tab == "9️⃣ Upload BA PL":
        import ba_engine_pl as _ba_pl_engine5
        import ba_config_pl as _ba_cfg_pl
        import os as _os8
        import re as _re8
        import datetime as _dt8

        st.markdown("## Berita Acara — Pengadaan Langsung")

        # Load paket PL — filter paket selesai
        _pl8_rows = []
        try:
            _raw8 = _load_draft_pl_cached("PK")
            _raw8, _ = pl_engine.buang_duplikat_paket_lama(_raw8)
            _pl8_rows = [r for r in _raw8 if not pl_engine.is_paket_selesai(r)]
            _pl8_kode_status = []
            for _r8 in _pl8_rows:
                _s8 = _enrich_kode_unik_pl_excel(_r8)
                if _s8 != "excel":
                    _pl8_kode_status.append(
                        f"{_r8.get('kode_paket')}: {_s8}"
                    )
        except Exception as _e8:
            st.error(f"Gagal load paket PL: {_e8}")
            _pl8_kode_status = []
        if _pl8_kode_status:
            st.warning(
                "⚠️ Status Kode Unik PL (sumber utama `@ Master Data!F2`):\n\n"
                + "\n".join(f"- {x}" for x in _pl8_kode_status)
            )

        # ── Helper auto nomor + tanggal (otomatis seperti mode tender) ───────
        def _skpd_pl8(_satker):
            if not _satker:
                return "DPUPR"
            try:
                from config import sb as _sbf8
                _r = _sbf8().table("master_dinas").select("singkatan").ilike(
                    "nama_dinas", f"%{_satker[:30]}%").limit(1).execute()
                if _r.data:
                    return _r.data[0].get("singkatan") or "DPUPR"
            except Exception:
                pass
            return "DPUPR"

        def _kode_unik_excel_pl8(_row):
            """Baca kode unik manual dari @ Master Data!F2 workbook paket."""
            return _baca_master_data_pl(_row).get("kode_unik", "")

        def _nomor_dokpil_pl8(_row):
            """Nomor dokpil final dari @ Master Data!C20 workbook paket."""
            _md = _baca_master_data_pl(_row)
            _base = str(_md.get("nomor_dokpil") or "").strip()
            _kode = str(_md.get("kode_unik") or _row.get("kode_unik") or "").strip()
            if _base and _kode:
                _base = _re8.sub(r"(/PP-[^/]+/)[^/]+(?=/)", rf"\1{_kode}", _base, count=1)
            return _base

        def _auto_nomor_pl8(_row, _jenis_key):
            """Derive nomor BA per jenis: ganti slot /NN/ pertama."""
            _base = _nomor_dokpil_pl8(_row)
            if not _base:
                return ""
            return ba_engine.derive_nomor_ba(_base, _ba_cfg_pl.NOMOR_URUT_PL[_jenis_key])

        def _auto_tgl_pl8(_row):
            """Tanggal BA = end Evaluasi Penawaran (tgl_evaluasi), fallback live SPSE."""
            _v = _row.get("tgl_evaluasi")
            if _v:
                try:
                    return _dt8.date.fromisoformat(str(_v)[:10])
                except Exception:
                    pass
            try:
                _jd = _parse_jadwal_pl_cached(_row.get("kode_paket", ""))
                if _jd and len(_jd) > 2:
                    return _jd[2]["selesai"].date()
            except Exception:
                pass
            return None

        def _backup_pdf_pl8(_row, _jenis_key, _pdf_bytes):
            """Simpan PDF cetak ke folder masing-masing paket. Return path atau ''."""
            _kp = _row.get("kode_paket", "")
            _dir = None
            # 1) resolver folder paket PL resmi (folder_paket root)
            try:
                import kualifikasi_engine_plpk as _kepl8
                _rf = _kepl8.resolve_folder_paket_pl(_kp)
                if _rf.get("ok"):
                    _root = _rf.get("pesan") or ""  # pesan = folder_paket root
                    if _root and _os8.path.isdir(_root):
                        _sub = _os8.path.join(_root, "7. Berita Acara + Summary Non Tender")
                        _os8.makedirs(_sub, exist_ok=True)
                        _dir = _sub
            except Exception:
                pass
            # 2) fallback: Asisten_Pokja_Downloads
            if not _dir:
                try:
                    from config import POKJA_ROOT as _PR8
                    _dir = _os8.path.join(_PR8, "Asisten_Pokja_Downloads", f"Cetak_BA_PL_{_kp}")
                except Exception:
                    return ""
            try:
                _os8.makedirs(_dir, exist_ok=True)
                _fn = f"{_ba_cfg_pl.FILE_LABEL_PL[_jenis_key]}-{_kp}.pdf"
                _fp = _os8.path.join(_dir, _fn)
                with open(_fp, "wb") as _fh:
                    _fh.write(_pdf_bytes)
                return _fp
            except Exception:
                return ""

        def _proses_ba_pl8(_row, _jenis_key, _nomor, _tgl):
            """Cetak → backup → upload satu BA. Return (ok: bool, pesan: str)."""
            # Endpoint cetak/upload BA nontender pakai kode_paket (BUKAN id_nontender)
            _id = _row.get("kode_paket")
            if not _nomor:
                return False, "nomor kosong"
            if not _tgl:
                return False, "tanggal kosong (Evaluasi belum dijadwal)"
            _tgls = _tgl.strftime("%d-%m-%Y")
            try:
                _rc = _ba_pl_engine5.cetak_ba_pl(
                    id_nontender=_id, jenis_key=_jenis_key,
                    nomor_ba=_nomor, tanggal_ba=_tgls,
                )
                if not _rc["ok"]:
                    return False, f"cetak gagal ({_rc.get('status')}) {_rc.get('error','')}"
                _backup_pdf_pl8(_row, _jenis_key, _rc["pdf_bytes"])
                _fn = f"{_ba_cfg_pl.FILE_LABEL_PL[_jenis_key]}-{_row.get('kode_paket','')}.pdf"
                _ru = _ba_pl_engine5.upload_ba_pl(
                    id_nontender=_id, jenis_key=_jenis_key,
                    nomor_ba=_nomor, tanggal_ba=_tgls,
                    file_bytes=_rc["pdf_bytes"], file_name=_fn,
                )
                if _ru.get("ok"):
                    return True, "OK"
                return False, f"upload gagal ({_ru.get('status')})"
            except Exception as _e:
                return False, str(_e)[:100]

        # ── SECTION 1: 2 BA AUTO (Evaluasi + Hasil) ──────────────────────────
        st.markdown("### 1. Cetak + Upload Otomatis")
        st.caption("BA Evaluasi Penawaran dan BA Hasil Non Tender dicetak langsung dari SPSE lalu di-upload kembali.")

        if not _pl8_rows:
            st.info("Tidak ada paket PL di database.")
        else:
            # Tanggal: default OTOMATIS dari tgl_evaluasi per paket (mode tender).
            _pl8_tgl_mode = st.radio(
                "Mode Tanggal",
                ["Otomatis (tgl Evaluasi per paket)", "Satu tanggal semua manual"],
                horizontal=True, key="pl8_tgl_mode",
            )
            _pl8_tgl_global = None
            if _pl8_tgl_mode == "Satu tanggal semua manual":
                _pl8_tgl_global = st.date_input(
                    "Tanggal BA (semua paket)", value=datetime.now().date(),
                    format="DD/MM/YYYY", key="pl8_tgl_global",
                )
                st.caption(f"{_HARI_NAMA[_pl8_tgl_global.weekday()]}, {_pl8_tgl_global.day} "
                           f"{_BULAN_NAMA[_pl8_tgl_global.month-1]} {_pl8_tgl_global.year}")
            else:
                st.caption("Tanggal otomatis = hari terakhir Evaluasi Penawaran (dari jadwal). "
                           "Nomor BA auto-derive dari Nomor Dokpil.")

            # ── Tombol BULK: cetak+upload SEMUA paket × 2 BA tanpa centang ──
            if st.button(
                f"🚀🚀 Cetak + Upload SEMUA Paket ({len(_pl8_rows)} paket × 2 BA)",
                type="primary", key="pl8_bulk_all_paket", use_container_width=True,
            ):
                with st.status(f"Proses {len(_pl8_rows)} paket × 2 BA...", expanded=True) as _stb8:
                    _ok_total, _gagal_total = 0, 0
                    for _pbulk in _pl8_rows:
                        _kb = _pbulk.get("kode_paket", "")
                        _tgl_b = _pl8_tgl_global if _pl8_tgl_mode == "Satu tanggal semua manual" else _auto_tgl_pl8(_pbulk)
                        _stb8.write(f"**{_pl_label(_pbulk)}**")
                        for _jkb, _lblb in [("evaluasi", "Evaluasi"), ("hasil", "Hasil")]:
                            _nob = _auto_nomor_pl8(_pbulk, _jkb)
                            _okb, _pesb = _proses_ba_pl8(_pbulk, _jkb, _nob, _tgl_b)
                            if _okb:
                                _stb8.write(f"  ✅ BA {_lblb}")
                                _ok_total += 1
                            else:
                                _stb8.write(f"  ❌ BA {_lblb} — {_pesb}")
                                _gagal_total += 1
                    _stb8.update(
                        label=f"Selesai — {_ok_total} BA OK, {_gagal_total} gagal.",
                        state="complete" if _gagal_total == 0 else "error",
                    )

            st.divider()

            # Header tabel per paket
            # Daftar paket — 1 baris horizontal per paket
            for _p8 in _pl8_rows:
                _k8  = _p8.get("kode_paket", "")
                _id8 = _k8
                if _pl8_tgl_mode == "Satu tanggal semua manual":
                    _tgl8 = _pl8_tgl_global
                else:
                    _tgl8 = _auto_tgl_pl8(_p8)
                _no8ev = _auto_nomor_pl8(_p8, "evaluasi")
                _no8hs = _auto_nomor_pl8(_p8, "hasil")
                _col_nama, _col_tgl, _col_btn = st.columns([5, 3, 2])
                with _col_nama:
                    st.markdown(f"**{_pl_label(_p8)}**")
                with _col_tgl:
                    if _tgl8:
                        st.caption(f"📅 {_HARI_NAMA[_tgl8.weekday()]}, {_tgl8.day} {_BULAN_NAMA[_tgl8.month-1]} {_tgl8.year}")
                    else:
                        st.caption("⚠️ Tanggal belum ada")
                with _col_btn:
                    if _tgl8 and st.button("🖨️ Cetak + Upload", key=f"pl8_ev_hs_{_k8}", use_container_width=True, type="primary"):
                        _tgl8s = _tgl8.strftime("%d-%m-%Y")
                        for _jk8, _no8, _lbl8 in [("evaluasi", _no8ev, "BA Evaluasi"), ("hasil", _no8hs, "BA Hasil")]:
                            with st.spinner(f"Proses {_lbl8}..."):
                                _rc8x = _ba_pl_engine5.cetak_ba_pl(id_nontender=_id8, jenis_key=_jk8, nomor_ba=_no8, tanggal_ba=_tgl8s)
                                if _rc8x["ok"]:
                                    _backup_pdf_pl8(_p8, _jk8, _rc8x["pdf_bytes"])
                                    _fn8x = f"{_ba_cfg_pl.FILE_LABEL_PL[_jk8]}-{_k8}.pdf"
                                    _ru8x = _ba_pl_engine5.upload_ba_pl(id_nontender=_id8, jenis_key=_jk8, nomor_ba=_no8, tanggal_ba=_tgl8s, file_bytes=_rc8x["pdf_bytes"], file_name=_fn8x)
                                    if _ru8x.get("ok"):
                                        st.success(f"✅ {_lbl8} berhasil")
                                    else:
                                        st.error(f"❌ {_lbl8} upload gagal: {_ru8x.get('status')}")
                                else:
                                    st.error(f"❌ {_lbl8} cetak gagal: {_rc8x.get('error')}")

            # Tombol batch semua

        # ── SECTION 2: BA LAINNYA (upload-only) — HIDDEN ─────────────────────
        if False:
            st.divider()
            st.markdown("### 2. BA Lainnya (Upload Manual)")
            st.caption("BA Penjelasan, Pengumuman Pemenang, dll — upload file PDF langsung ke SPSE.")

            if _pl8_rows:
                _JENIS_LAINNYA = {
                    "penjelasan":   "BA Penjelasan",
                    "pengumuman":   "Pengumuman Pemenang Akhir",
                }
                _pl8_lc1, _pl8_lc2 = st.columns([1, 2])
                with _pl8_lc1:
                    _jenis_lain8 = st.selectbox(
                        "Jenis BA",
                        options=list(_JENIS_LAINNYA.keys()),
                        format_func=lambda k: _JENIS_LAINNYA[k],
                        key="pl8_jenis_lain",
                    )
                with _pl8_lc2:
                    _nomor_lain8  = st.text_input(
                        "Nomor BA",
                        placeholder="000.3.3/06/PL/PP-01/KPP1/DPUPR/2026",
                        key="pl8_nomor_lain",
                    )
                _tgl_lain8 = st.date_input(
                    "Tanggal BA", value=datetime.now().date(), format="DD/MM/YYYY",
                    key="pl8_tgl_lain",
                )
                st.caption(f"{_HARI_NAMA[_tgl_lain8.weekday()]}, {_tgl_lain8.day} "
                           f"{_BULAN_NAMA[_tgl_lain8.month-1]} {_tgl_lain8.year}")
                _tempat_lain8 = ""
                if _jenis_lain8 == "pengumuman":
                    _tempat_lain8 = st.text_input("Tempat", placeholder="Contoh: Kantor RSUD", key="pl8_tempat_lain")
                _info_lain8 = st.text_area("Keterangan Tambahan", value="", key="pl8_info_lain", height=68)

                st.markdown("**Pilih Paket + Upload File:**")
                _pl8_lain_valid = []
                for _p8l in _pl8_rows:
                    _k8l  = _p8l.get("kode_paket", "")
                    _id8l = _k8l  # BA nontender pakai kode_paket
                    _lc1, _lc2, _lc3 = st.columns([3, 3, 1])
                    with _lc1:
                        st.caption(_pl_label(_p8l))
                    with _lc2:
                        _fl8 = st.file_uploader(
                            "PDF", type=["pdf"], key=f"pl8_lain_file_{_k8l}",
                            label_visibility="collapsed",
                        )
                    with _lc3:
                        if _fl8 and st.button("📤", key=f"pl8_lain_up_{_k8l}", use_container_width=True):
                            with st.spinner(f"Upload {_k8l}..."):
                                _rl8 = _ba_pl_engine5.upload_ba_pl(
                                    id_nontender=_id8l, jenis_key=_jenis_lain8,
                                    nomor_ba=_nomor_lain8,
                                    tanggal_ba=_tgl_lain8.strftime("%d-%m-%Y"),
                                    info=_info_lain8,
                                    file_bytes=_fl8.read(), file_name=_fl8.name,
                                    tempat=_tempat_lain8,
                                )
                            if _rl8.get("ok"):
                                st.success(f"✅ {_k8l} — berhasil")
                            else:
                                st.error(f"❌ {_k8l} — status {_rl8.get('status')}")
                    if _fl8:
                        _pl8_lain_valid.append({**_p8l, "_id8l": _id8l, "_fl8": _fl8})

                if _pl8_lain_valid:
                    if st.button(
                        f"📤 Upload Semua ({len(_pl8_lain_valid)} paket)",
                        key="pl8_lain_all", use_container_width=True,
                    ):
                        with st.status(f"Upload {len(_pl8_lain_valid)} paket...", expanded=True) as _st8l:
                            for _pbl in _pl8_lain_valid:
                                _st8l.write(f"⏳ {_pbl['kode_paket']}...")
                                _rbl = _ba_pl_engine5.upload_ba_pl(
                                    id_nontender=_pbl["_id8l"], jenis_key=_jenis_lain8,
                                    nomor_ba=_nomor_lain8,
                                    tanggal_ba=_tgl_lain8.strftime("%d-%m-%Y"),
                                    info=_info_lain8,
                                    file_bytes=_pbl["_fl8"].read(), file_name=_pbl["_fl8"].name,
                                    tempat=_tempat_lain8,
                                )
                                if _rbl.get("ok"):
                                    _st8l.write(f"  ✅ {_pbl['kode_paket']} — berhasil")
                                else:
                                    _st8l.write(f"  ❌ {_pbl['kode_paket']} — status {_rbl.get('status')}")

    # ── Tab 5: Download Dok Kualifikasi PL ───────────────────────────────────
    if _pl_active_tab == "6️⃣ Download Kualifikasi":
        _ke_pl = _ke_pl_pk   # alias — sudah di-import top-level
        _he_pl = _he_pl_pk

        st.markdown("## Download Dokumen Kualifikasi — Pengadaan Langsung")
        st.caption("Download dok kualifikasi peserta dari SPSE + populate sheet Hasil Evaluasi di BAPLJKK.")

        # Cache paket list di session state — hindari query Supabase tiap render
        if "pl7_rows" not in st.session_state:
            try:
                st.session_state["pl7_rows"], _pl7_dup_n = active_plpk_rows(_load_draft_pl_cached, pl_engine)
                if _pl7_dup_n:
                    st.caption(f"♻️ {_pl7_dup_n} row lama duplikat disembunyikan otomatis.")
            except Exception as _e7:
                st.session_state["pl7_rows"] = []
                st.error(f"Gagal load paket PL: {_e7}")
        _pl7_rows = st.session_state["pl7_rows"]

        if not _pl7_rows:
            st.info("Tidak ada paket PL di database.")
            if st.button("🔄 Reload", key="pl7_reload"):
                del st.session_state["pl7_rows"]
                st.rerun()
        else:
            _pl7c1, _pl7c2 = st.columns([1, 1])

            with _pl7c1:
                st.markdown("#### Pilih Paket")
                _pl7_selected_rows = render_package_selection(st, _pl7_rows, _pl_label, prefix="pl7")

            with _pl7c2:
                st.markdown("#### Aksi")

                _n_paket7 = len(_pl7_selected_rows)

                if locals().get("_is_plpk_tab6", False):
                    render_download_actions(
                        st, _pl7_selected_rows, _ke_pl_pk, _he_pl_pk, _pl_label
                    )
                elif not _pl7_selected_rows:
                    st.info("Centang minimal 1 paket di kiri.")
                else:
                    st.markdown(f"**{_n_paket7} paket** dipilih")
                    st.caption("Peserta akan di-fetch via CDP saat tombol Jalankan diklik.")

                    _do_download7 = st.checkbox("⬇️ Download dokumen kualifikasi", value=True, key="pl7_do_dl")
                    _do_parse7    = st.checkbox("📋 Parse & populate sheet Hasil Evaluasi", value=True, key="pl7_do_parse")

                    _btn7 = st.button(
                        f"▶ Jalankan — {_n_paket7} paket",
                        type="primary", key="pl7_run", use_container_width=True,
                    )

                    if _btn7:
                        _pb7 = st.progress(0.0, text="Memulai...")
                        _log7_lines = []  # akumulasi log — update sekali per paket, bukan per baris
                        _ringkasan7: list = []  # kumpul status tiap paket untuk ringkasan akhir

                        def _flush7(container):
                            """Render log terakumulasi ke container — 1 update per paket."""
                            container.code("\n".join(_log7_lines[-60:]))  # max 60 baris terakhir

                        for _i7, _rpl7 in enumerate(_pl7_selected_rows):
                            _kpl7    = _rpl7["kode_paket"]
                            _nama7   = _pl_label(_rpl7)
                            _status7 = st.status(f"Paket {_i7+1}/{_n_paket7} — {_nama7}", expanded=True)

                            with _status7:
                                _log7_lines.clear()
                                _log7_box = st.empty()

                                def _lcb7(msg, _box=_log7_box, _lines=_log7_lines):
                                    _lines.append(msg)
                                    _box.code("\n".join(_lines[-40:]))

                                # Fetch peserta
                                _lcb7(f"[{_i7+1}/{_n_paket7}] Fetch peserta SPSE...")
                                _pb7.progress((_i7) / _n_paket7, text=f"{_nama7} — fetch peserta")
                                _fp7 = _ke_pl.fetch_peserta_pl(_kpl7)
                                if not _fp7.get("ok"):
                                    _lcb7(f"[SKIP] Peserta: {_fp7['pesan']}")
                                    _status7.update(label=f"SKIP {_nama7} — {_fp7['pesan']}", state="error", expanded=False)
                                    _ringkasan7.append({"nama": _nama7, "status": "skip", "detail": _fp7["pesan"]})
                                    continue
                                _peserta7 = _fp7["peserta"]
                                _lcb7(f"Peserta ({len(_peserta7)}): {', '.join(p['nama'] for p in _peserta7)}")

                                # Resolve folder
                                _folder7 = _ke_pl.resolve_folder_paket_pl(_kpl7)
                                if not _folder7.get("ok"):
                                    _lcb7(f"[SKIP] Folder: {_folder7['pesan']}")
                                    _status7.update(label=f"SKIP {_nama7} — folder tidak ditemukan", state="error", expanded=False)
                                    _ringkasan7.append({"nama": _nama7, "status": "skip", "detail": _folder7.get("pesan", "folder tidak ditemukan")})
                                    continue
                                _folder_kual7 = _folder7["path"]

                                # Download kualifikasi
                                if _do_download7:
                                    _pb7.progress((_i7 + 0.3) / _n_paket7, text=f"{_nama7} — download kualifikasi")
                                    for _ui7, _p7 in enumerate(_peserta7, 1):
                                        _lcb7(f"--- Download [{_ui7}/{len(_peserta7)}] {_p7['nama']} ---")
                                        _ke_pl.download_kualifikasi_peserta_pl(
                                            _p7, _folder_kual7, _ui7, len(_peserta7), _lcb7,
                                        )

                                    # Serap penyedia (dipindah dari create folder — di sini peserta sudah
                                    # terdaftar & Draft_PL/ND.pdf ada, jadi parse berhasil bukan timeout).
                                    _lcb7("--- Serap penyedia (nama/NPWP/personil) ---")
                                    try:
                                        import parse_kak_pl as _pkpl7
                                        _sp7 = _pkpl7.serap_penyedia_pl(kode_paket_filter=_kpl7)
                                        _lcb7(f"👤 Penyedia: {_sp7.get('updated',0)} diperbarui"
                                              if _sp7.get("updated", 0) > 0 else "👤 Penyedia: tidak ada data baru")
                                    except Exception as _sp7_e:
                                        _lcb7(f"⚠ Serap penyedia: {_sp7_e}")

                                # Parse evaluasi
                                if _do_parse7:
                                    _pb7.progress((_i7 + 0.7) / _n_paket7, text=f"{_nama7} — parse evaluasi")
                                    _lcb7("--- Populate sheet Hasil Evaluasi ---")
                                    _hasil7 = _he_pl.populate_hasil_evaluasi_pl(_kpl7, _peserta7, _lcb7)
                                    _lcb7(f"{'[OK]' if _hasil7.get('ok') else '[GAGAL]'} {_hasil7['pesan']}")

                                    # Refresh @ Master Data agar tgl_pembukaan benar
                                    # (penting untuk paket ulang: kode_paket baru → tanggal baru dari Supabase)
                                    if _hasil7.get("ok"):
                                        _lcb7("--- Refresh @ Master Data ---")
                                        try:
                                            import isi_master_data_pl as _imd7
                                            _xlsm7 = _he_pl._find_xlsm(_kpl7)
                                            if _xlsm7:
                                                _md7 = _imd7.isi_master_data_pl(_kpl7, _xlsm7, progress_cb=_lcb7)
                                                _lcb7(f"{'[OK]' if _md7.get('ok') else '[WARN]'} {_md7['pesan']}")
                                            else:
                                                _lcb7("[WARN] File .xlsm tidak ditemukan untuk refresh @ Master Data")
                                        except Exception as _e_md7:
                                            _lcb7(f"[WARN] Refresh @ Master Data gagal: {_e_md7}")

                                    _ringkasan7.append({
                                        "nama"  : _nama7,
                                        "status": "ok" if _hasil7.get("ok") else "gagal",
                                        "detail": _hasil7.get("pesan", ""),
                                    })
                                else:
                                    # Tidak ada parse → anggap OK (hanya download)
                                    _ringkasan7.append({"nama": _nama7, "status": "ok", "detail": "download saja"})

                                _status7.update(label=f"Selesai — {_nama7}", state="complete", expanded=False)

                            _pb7.progress((_i7 + 1) / _n_paket7, text=f"Selesai {_i7+1}/{_n_paket7} paket")

                        _pb7.progress(1.0, text="Semua paket selesai.")
                        from batch_summary import render_ringkasan_batch as _rrb7
                        _rrb7(st, _ringkasan7)

    # ── Tab 6: Evaluasi SPSE + Download Teknis/Biaya ─────────────────────────
    if _pl_active_tab == "7️⃣ Evaluasi & Teknis/Biaya":
        import evaluasi_admin_kualifikasi_pl as _eval_pl
        import dokumen_teknis_biaya_pl as _dtb_pl
        import penawaran_pl_engine as _penawaran_pl

        st.markdown("## Evaluasi SPSE & Download Teknis/Biaya — Pengadaan Langsung")
        st.caption("Submit evaluasi Admin+Kualifikasi LULUS di SPSE, lalu download dokumen teknis/biaya peserta.")

        # Selalu normalisasi daftar di sini. Tab Evaluasi dapat dibuka langsung,
        # tanpa lebih dulu merender Tab Download Kualifikasi.
        _pl8_rows, _pl8_dup_n = active_plpk_rows(_load_draft_pl_cached, pl_engine)
        if _pl8_dup_n:
            st.caption(f"♻️ {_pl8_dup_n} row lama duplikat disembunyikan otomatis.")

        if not _pl8_rows:
            st.info("Tidak ada paket PL. Reload di Tab 7.")
        else:
            _pl8c1, _pl8c2 = st.columns([1, 1])

            with _pl8c1:
                st.markdown("#### Pilih Paket")
                if "pl8_checked" not in st.session_state:
                    st.session_state["pl8_checked"] = {}

                _pl8_kodes = [r["kode_paket"] for r in _pl8_rows]
                if st.session_state.get("pl8_selection_default_v19") != True:
                    for _k in _pl8_kodes:
                        st.session_state.pop(f"pl8_chk_{_k}_v19", None)
                    st.session_state["pl8_checked"].clear()
                    st.session_state["pl8_selection_default_v19"] = True
                for _k in _pl8_kodes:
                    if f"pl8_chk_{_k}_v19" not in st.session_state:
                        st.session_state[f"pl8_chk_{_k}_v19"] = True
                _pl8bc1, _pl8bc2 = st.columns(2)
                if _pl8bc1.button("✅ Pilih Semua", key="pl8_select_all", use_container_width=True):
                    for _k in _pl8_kodes:
                        st.session_state[f"pl8_chk_{_k}_v19"] = True
                        st.session_state["pl8_checked"][_k] = True
                if _pl8bc2.button("❌ Batal Semua", key="pl8_deselect_all", use_container_width=True):
                    for _k in _pl8_kodes:
                        st.session_state[f"pl8_chk_{_k}_v19"] = False
                        st.session_state["pl8_checked"][_k] = False

                for _rpl8 in _pl8_rows:
                    _kpl8 = _rpl8["kode_paket"]
                    _chk8 = st.checkbox(
                        "",
                        key=f"pl8_chk_{_kpl8}_v19",
                        label_visibility="collapsed",
                    )
                    st.markdown(f"**{_pl_label(_rpl8)}**")
                    st.session_state["pl8_checked"][_kpl8] = _chk8

            with _pl8c2:
                st.markdown("#### Aksi")

                _pl8_selected_rows = [r for r in _pl8_rows if st.session_state["pl8_checked"].get(r["kode_paket"])]
                _n_paket8 = len(_pl8_selected_rows)

                if not _pl8_selected_rows:
                    st.info("Centang minimal 1 paket di kiri.")
                else:
                    st.markdown(f"**{_n_paket8} paket** dipilih")
                    st.caption("Peserta di-scrape dari SPSE saat Jalankan.")
                    _do_eval_admin = st.checkbox("⚖️ Submit evaluasi Admin + Kualifikasi LULUS di SPSE", value=True, key="pl8_do_eval_admin")
                    _do_eval_teknis = st.checkbox("⚙️ Submit evaluasi Teknis LULUS di SPSE", value=True, key="pl8_do_eval_teknis")
                    _do_eval_harga = st.checkbox("💰 Submit evaluasi Harga LULUS di SPSE", value=True, key="pl8_do_eval_harga")

                    _do_tekbio8  = st.checkbox("⬇️ Download dokumen teknis/biaya + gabung PDF", value=True, key="pl8_do_tekbio")
                    _do_penawaran8 = st.checkbox("📊 Tulis rincian penawaran ke sheet '6. Penawaran' Excel", value=True, key="pl8_do_penawaran")

                    st.warning("Evaluasi LULUS bersifat **permanen** — modifikasi data SPSE production.")
                    _konfirmasi8 = st.checkbox(
                        "Saya paham tindakan ini tidak bisa dibatalkan.",
                        value=False, key="pl8_konfirmasi",
                    )
                    _btn8_disabled = (_do_eval_admin or _do_eval_teknis or _do_eval_harga) and not _konfirmasi8
                    if _btn8_disabled:
                        st.info("Centang konfirmasi untuk mengaktifkan tombol.")
                    _btn8 = st.button(
                        f"▶ Jalankan — {_n_paket8} paket",
                        type="primary", key="pl8_run", use_container_width=True,
                        disabled=_btn8_disabled,
                    )

                    st.divider()
                    st.markdown("#### 🤖 Evaluasi AI")
                    st.caption("AI baca dokumen di folder paket → output `.md`. Paralel per paket.")
                    _do_extract_teks8pk = st.checkbox("📝 Extract teks kualifikasi (.txt) dari folder '8. Dokumen Kualifikasi' — hemat token sebelum evaluasi AI", value=False, key="pl8pk_do_extract_teks")
                    _ai_eval_engine_pk = "codex"
                    _ai_eval_model_pk = None
                    st.caption("Engine: Codex CLI · Model terkunci: gpt-5.6-luna · Reasoning: medium")
                    _do_ai_kualifikasi_pk = st.checkbox("⚖️ Evaluasi Admin+Kualifikasi (Sesi 1) via AI", value=True, key="pl8pk_do_ai_kual")
                    _do_ai_teknis_pk = st.checkbox("🔬 Evaluasi Teknis (Sesi 2) via AI", value=True, key="pl8pk_do_ai_teknis")
                    _btn_ai_eval_pk = st.button(
                        f"🤖 Jalankan Evaluasi AI — {_n_paket8} paket",
                        key="pl8pk_btn_ai_eval", use_container_width=True,
                        disabled=not (_do_ai_kualifikasi_pk or _do_ai_teknis_pk),
                    )
                    if _btn_ai_eval_pk and (_do_ai_kualifikasi_pk or _do_ai_teknis_pk):
                        import ai_evaluator as _heval8pk
                        if _do_extract_teks8pk:
                            import extract_teks_kualifikasi as _etk8pk
                            st.info("📝 Extract teks kualifikasi...")
                            for _r8pk in _pl8_selected_rows:
                                try:
                                    _folder8pk = _heval8pk._folder_paket(_r8pk.get("nomor_urut"), _r8pk.get("nama_paket", ""), jenis_pl="PK", is_ulang=bool(_r8pk.get("is_ulang")))
                                    if not _folder8pk:
                                        continue
                                    _etk_folder8pk = os.path.join(str(_folder8pk), "8. Dokumen Kualifikasi")
                                    if os.path.isdir(_etk_folder8pk):
                                        _etk8pk.extract_folder_kualifikasi(_etk_folder8pk)
                                except Exception as _etk8pk_e:
                                    st.warning(f"⚠ Extract teks {_pl_label(_r8pk)}: {_etk8pk_e}")
                        _ai_jobs_pk = [{"nomor_urut": r.get("nomor_urut"), "nama_paket": r.get("nama_paket",""), "is_ulang": bool(r.get("is_ulang"))} for r in _pl8_selected_rows]
                        if _do_ai_kualifikasi_pk:
                            st.info("⚖️ Menjalankan evaluasi Admin+Kualifikasi...")
                            _res_kual_pk = _heval8pk.evaluasi_bulk(_ai_jobs_pk, jenis="kualifikasi", model=_ai_eval_model_pk, max_workers=3, jenis_pl="PK", engine=_ai_eval_engine_pk)
                            _ai_label_map_pk = {r.get("nama_paket", ""): _pl_label(r) for r in _pl8_selected_rows}
                            for _rk_pk in _res_kual_pk:
                                _rk_pk_name = _ai_label_map_pk.get(_rk_pk.get("nama"), _rk_pk.get("nama", "-"))
                                if _rk_pk["status"] == "ok":
                                    st.success(f"✅ {_rk_pk_name[:50]}")
                                    with st.expander(f"Output kualifikasi: {_rk_pk_name[:35]}"):
                                        st.markdown(_rk_pk["output"][:3000])
                                else:
                                    st.error(f"❌ {_rk_pk_name[:50]} — {_rk_pk['error'][:200]}")
                        if _do_ai_teknis_pk:
                            st.info("🔬 Menjalankan evaluasi Teknis...")
                            _res_teknis_pk = _heval8pk.evaluasi_bulk(_ai_jobs_pk, jenis="teknis", model=_ai_eval_model_pk, max_workers=3, jenis_pl="PK", engine=_ai_eval_engine_pk)
                            for _rt_pk in _res_teknis_pk:
                                _rt_pk_name = _ai_label_map_pk.get(_rt_pk.get("nama"), _rt_pk.get("nama", "-"))
                                if _rt_pk["status"] == "ok":
                                    st.success(f"✅ {_rt_pk_name[:50]}")
                                    with st.expander(f"Output teknis: {_rt_pk_name[:35]}"):
                                        st.markdown(_rt_pk["output"][:3000])
                                else:
                                    st.error(f"❌ {_rt_pk_name[:50]} — {_rt_pk['error'][:200]}")

                    if _btn8:
                        _pb8 = st.progress(0.0, text="Memulai...")
                        _ringkasan8: list = []  # kumpul status tiap paket untuk ringkasan akhir

                        for _i8, _rpl8 in enumerate(_pl8_selected_rows):
                            _kpl8  = _rpl8["kode_paket"]
                            _nama8 = _pl_label(_rpl8)
                            _status8 = st.status(f"Paket {_i8+1}/{_n_paket8} — {_nama8}", expanded=True)

                            with _status8:
                                _log8_lines = []
                                _log8_box   = st.empty()

                                def _lcb8(msg, _box=_log8_box, _lines=_log8_lines):
                                    _lines.append(msg)
                                    _box.code("\n".join(_lines[-40:]))

                                # Scrape id_nontender per peserta
                                _lcb8("Scrape peserta evaluasi dari SPSE...")
                                _pb8.progress(_i8 / _n_paket8, text=f"{_nama8} — scrape peserta")
                                _res_peserta8 = _eval_pl.scrape_peserta_evaluasi(_kpl8)
                                if not _res_peserta8.get("ok"):
                                    _lcb8(f"[SKIP] {_res_peserta8['pesan']}")
                                    _status8.update(label=f"SKIP {_nama8} — {_res_peserta8['pesan']}", state="error", expanded=False)
                                    _ringkasan8.append({"nama": _nama8, "status": "skip", "detail": _res_peserta8["pesan"]})
                                    continue
                                _peserta8 = _res_peserta8["peserta"]
                                _lcb8(f"Peserta ({len(_peserta8)}): {', '.join(p['nama'] for p in _peserta8)}")

                                # Resolve folder paket root
                                _folder8      = _ke_pl.resolve_folder_paket_pl(_kpl8)
                                _folder_paket8 = _folder8.get("pesan", "") if _folder8.get("ok") else ""

                                # Evaluasi LULUS
                                if _do_eval_admin or _do_eval_teknis or _do_eval_harga:
                                    _pb8.progress((_i8 + 0.3) / _n_paket8, text=f"{_nama8} — submit evaluasi")
                                    _lcb8(f"--- Submit evaluasi LULUS (Admin:{_do_eval_admin}, Teknis:{_do_eval_teknis}, Harga:{_do_eval_harga}) ---")
                                    _eval8 = _eval_pl.evaluasi_batch_lulus(
                                        _kpl8,
                                        admin=_do_eval_admin,
                                        kualifikasi=_do_eval_admin,
                                        teknis=_do_eval_teknis,
                                        harga=_do_eval_harga,
                                        progress_cb=_lcb8
                                    )
                                    _lcb8(f"{'[OK]' if _eval8.get('ok') else '[SEBAGIAN GAGAL]'} {_eval8['ringkasan']}")

                                # Download teknis/biaya
                                if _do_tekbio8:
                                    if not _folder_paket8:
                                        _lcb8("[SKIP] Folder paket tidak ditemukan")
                                    else:
                                        _pb8.progress((_i8 + 0.6) / _n_paket8, text=f"{_nama8} — download teknis/biaya")
                                        for _ui8, _ep8 in enumerate(_peserta8, 1):
                                            _lcb8(f"--- Download [{_ui8}/{len(_peserta8)}] {_ep8['nama']} ---")
                                            _res_tb8 = _dtb_pl.download_teknis_biaya_peserta(
                                                id_nontender=_ep8["id_nontender"],
                                                nama_peserta=_ep8["nama"],
                                                folder_paket=_folder_paket8,
                                                urutan=_ui8,
                                                progress_cb=_lcb8,
                                            )
                                            _lcb8(f"{'[OK]' if _res_tb8['ok'] else '[GAGAL]'} {_res_tb8['pesan']}")

                                # Tulis penawaran ke sheet 6. Penawaran
                                if _do_penawaran8:
                                    if not _folder_paket8:
                                        _lcb8("[SKIP] Folder paket tidak ditemukan, skip penawaran")
                                    else:
                                        _pb8.progress((_i8 + 0.8) / _n_paket8, text=f"{_nama8} — tulis penawaran")
                                        _lcb8("--- Tulis rincian penawaran ke Excel ---")
                                        for _ep8_p in _peserta8:
                                            _lcb8(f"  Peserta: {_ep8_p['nama']}")
                                            _res_pnw8 = _penawaran_pl.tulis_penawaran_ke_excel(
                                                folder_paket=_folder_paket8,
                                                id_nontender=_ep8_p["id_nontender"],
                                                progress_cb=_lcb8,
                                            )
                                            _lcb8(f"  {'[OK]' if _res_pnw8['ok'] else '[GAGAL]'} {_res_pnw8['pesan']}" +
                                                  (f" — Total Rp {_res_pnw8['total_penawaran']:,.0f}" if _res_pnw8.get('total_penawaran') else ""))

                                _ringkasan8.append({"nama": _nama8, "status": "ok", "detail": ""})
                                _status8.update(label=f"Selesai — {_nama8}", state="complete", expanded=False)

                            _pb8.progress((_i8 + 1) / _n_paket8, text=f"Selesai {_i8+1}/{_n_paket8} paket")

                        _pb8.progress(1.0, text="Semua paket selesai.")
                        from batch_summary import render_ringkasan_batch as _rrb8
                        _rrb8(st, _ringkasan8)

    if _pl_active_tab == "🔟 Penetapan Pemenang":
        _raw10 = _load_draft_pl_cached("PK")
        _raw10, _ = pl_engine.buang_duplikat_paket_lama(_raw10)
        _rows10 = [r for r in _raw10 if not pl_engine.is_paket_selesai(r)]
        _render_tab10_pl(_rows10, "pl10pk")

    st.stop()  # Jangan render tab Tender jika mode PL

# ============================================================
# MODE: TENDER
# ============================================================
_TENDER_TAB_LABELS = [
    "0️⃣ Persiapan Draft Paket",
    "1️⃣ Kirim Undangan DPP", "2️⃣ Buat Jadwal",
    "3️⃣ Setup Paket", "4️⃣ Pemberian Penjelasan",
    "5️⃣ Undangan Pembuktian Kualifikasi", "6️⃣ Download Kualifikasi",
    "7️⃣ Dokumen Penawaran", "8️⃣ Upload & Cetak 5 BA",
]
_tender_active_tab = st.radio("Tab Tender", _TENDER_TAB_LABELS, horizontal=True, key="tender_active_tab")

# ============================================================
# Tab 0: Persiapan Draft Paket
# ============================================================

if _tender_active_tab == "0️⃣ Persiapan Draft Paket":
    import inbox_engine
    import os as _os, subprocess as _sp
    from config import POKJA_ROOT as _POKJA_ROOT, TENDER_ROOT as _TENDER_ROOT, V19_ROOT as _V19_ROOT, POKJA_PYTHON as _POKJA_PYTHON

    _PY     = _POKJA_PYTHON
    _SCRIPT = str(pathlib.Path(_V19_ROOT) / "setup_paket_baru.py")
    _NO_WIN = 0x08000000  # CREATE_NO_WINDOW — cegah jendela hitam ngeblink di Windows

    # ── Load data draft_paket ──
    _draft_rows = _load_draft_paket_cached()

    # Isi tender_tahap_map dari Supabase kalau session state belum ada (startup/refresh)
    if "tender_tahap_map" not in st.session_state:
        st.session_state["tender_tahap_map"] = {
            str(_r["kode_tender"]): (_r.get("status_tahap") or "")
            for _r in _draft_rows if _r.get("status_tahap")
        }

    # ── Layout: kolom kiri (1) dan kanan (2) ──
    _col_kiri, _col_kanan = st.columns(2)

    # ══════════════════════════════════════════
    # KOLOM KIRI — 1. Scrap Inbox SPSE
    # ══════════════════════════════════════════
    with _col_kiri:
        # ── 1. Serap Data Paket (gabung: pilih aksi + 1 tombol) ─────────────
        st.markdown("#### 1. Serap Data Paket")
        st.caption("Pilih aksi lalu klik tombol — aksi berjalan berurutan sesuai centang.")
        _cb_serap_inbox = st.checkbox("Update Inbox (parse Delegasi Pokja → Supabase)", value=True, key="t_cb_serap_inbox")
        _cb_serap_sync  = st.checkbox("Sinkronkan daftar paket dari SPSE",              value=True, key="t_cb_serap_sync")

        # Auto-load paket SPSE: session baru → baca cache dulu (instan), baru fetch kalau expired
        if "global_paket_draft" not in st.session_state:
            _cache = kirimpesan_engine.load_paket_cache()
            if _cache:
                st.session_state["global_paket_draft"] = _cache["draft"]
                st.session_state["global_paket_aktif"] = _cache["aktif"]
            else:
                with st.spinner("Memuat daftar paket dari SPSE..."):
                    _gd0 = kirimpesan_engine.fetch_paket_draft()
                    _ga0 = kirimpesan_engine.fetch_paket_aktif()
                    _tahap0 = kirimpesan_engine.fetch_tahap_tender(_ga0.get("paket", []))
                    kirimpesan_engine.enrich_paket_supabase(_gd0.get("paket", []))
                    kirimpesan_engine.enrich_paket_supabase(_ga0.get("paket", []), tahap_map=_tahap0)
                    st.session_state["tender_tahap_map"] = _tahap0
                    st.session_state["global_paket_draft"] = _gd0
                    st.session_state["global_paket_aktif"] = _ga0
                    kirimpesan_engine.save_paket_cache(_gd0, _ga0)

        # Supabase fallback baru bisa memakai daftar SPSE setelah global state
        # terisi; refresh sumber Tab 0 pada startup yang sebelumnya kosong.
        if not _draft_rows:
            _load_draft_paket_cached.clear()
            _draft_rows = _load_draft_paket_cached()

        if st.button("🚀 Serap Data Paket", type="primary", use_container_width=True, key="btn_serap_tender_gabung"):
            _tender_rows_dirty = False
            # Aksi 1: Update Inbox
            if _cb_serap_inbox:
                _pb = st.progress(0.0)
                _st = st.empty()
                def cb(pct, msg):
                    _pb.progress(min(pct, 1.0))
                    _st.info(msg)
                try:
                    hasil = inbox_engine.serap_inbox(progress_cb=cb)
                    _pb.progress(1.0)
                    _c1, _c2, _c3, _c4 = st.columns(4)
                    _c1.metric("✅ Baru", hasil["baru"])
                    _c2.metric("🔄 Diperbarui", hasil["diperbarui"])
                    _c3.metric("⏭️ Dilewati", hasil.get("skip", 0))
                    _c4.metric("❌ Error", len(hasil["error"]))
                    if hasil["error"]:
                        with st.expander("Detail Error Inbox"):
                            for e in hasil["error"]:
                                st.error(e)
                    if hasil["data"]:
                        st.success(f"{len(hasil['data'])} paket diproses.")
                    else:
                        _st.warning("Tidak ada pesan Delegasi Pokja baru.")
                    _load_draft_paket_cached.clear()
                    _tender_rows_dirty = True
                except Exception as e:
                    st.error(f"Gagal serap inbox: {e}")

            # Aksi 2: Sinkronkan paket SPSE
            if _cb_serap_sync:
                kirimpesan_engine.clear_paket_cache()
                with st.spinner("Mengambil daftar paket dari SPSE..."):
                    _gd_r = kirimpesan_engine.fetch_paket_draft()
                    _ga_r = kirimpesan_engine.fetch_paket_aktif()
                    _tahap_r = kirimpesan_engine.fetch_tahap_tender(_ga_r.get("paket", []))
                    kirimpesan_engine.enrich_paket_supabase(_gd_r.get("paket", []))
                    kirimpesan_engine.enrich_paket_supabase(_ga_r.get("paket", []), tahap_map=_tahap_r)
                    st.session_state["tender_tahap_map"] = _tahap_r
                    st.session_state["global_paket_draft"] = _gd_r
                    st.session_state["global_paket_aktif"] = _ga_r
                    kirimpesan_engine.save_paket_cache(_gd_r, _ga_r)
                    _load_draft_paket_cached.clear()
                    _tender_rows_dirty = True
                st.toast("✅ Data paket SPSE tersinkronkan!", icon="🔄")
                st.success(f"Draft: {len(_gd_r.get('paket',[]))} paket | Aktif: {len(_ga_r.get('paket',[]))} paket")

            # Sinkronkan sumber daftar folder dengan data yang baru diserap.
            if _tender_rows_dirty:
                _draft_rows = _load_draft_paket_cached()

        # Info cache ringkas
        import os as _os_sync
        _gd2 = st.session_state.get("global_paket_draft", {})
        _ga2 = st.session_state.get("global_paket_aktif", {})
        _cache_info = ""
        if _os_sync.path.exists(kirimpesan_engine._CACHE_FILE):
            import time as _t_sync
            _age = int((_t_sync.time() - _os_sync.path.getmtime(kirimpesan_engine._CACHE_FILE)) / 60)
            _cache_info = f" (cache {_age}m lalu)"
        st.caption(f"✅ Draft: {len(_gd2.get('paket',[]))} | Aktif: {len(_ga2.get('paket',[]))}{_cache_info}")

        st.divider()

    # ══════════════════════════════════════════
    # KOLOM KANAN — 2. Buat Folder Paket
    # ══════════════════════════════════════════
    with _col_kanan:
        st.markdown("#### 2. Buat Folder Paket")

        # ── Notif folder baru dibuat (persist across rerun) ──
        if "_folder_just_created" in st.session_state:
            _just = st.session_state.pop("_folder_just_created")
            st.toast(f"✅ Folder berhasil dibuat: {_just}", icon="📁")
            st.success(f"✅ Folder **{_just}** berhasil dibuat!")
            st.balloons()
        if "_folder_bulk_created" in st.session_state:
            _bulk_msg = st.session_state.pop("_folder_bulk_created")
            _bulk_error = st.session_state.pop("_folder_bulk_had_error", False)
            if _bulk_error:
                st.toast(_bulk_msg, icon="⚠️")
                st.warning(f"⚠️ {_bulk_msg} — buka log detail")
            else:
                st.toast(_bulk_msg, icon="📁")
                st.success(f"✅ {_bulk_msg}")
                st.balloons()

        _tahun_skrg = str(datetime.now().year)
        _rows_tahun_ini = [_r for _r in _draft_rows if _tahun_skrg in str(_r.get("nomor_pp") or "")]

        # Helper: cari .xlsm utama di folder paket tender (utamakan "0. BAPK")
        def _cari_xlsm_tender(folder):
            try:
                _xs = [f for f in _os.listdir(folder) if f.lower().endswith(".xlsm")]
            except Exception:
                return None
            if not _xs:
                return None
            _xs.sort(key=lambda f: (not f.lower().startswith("0. bapk"), f))
            return _os.path.join(folder, _xs[0])

        # Helper: jalankan aksi (download / HPS / penawaran) untuk 1 paket. Return list log.
        def _jalankan_aksi_tender(kode_tender, id_pesan, kode_pokja, target_path,
                                  do_dl, do_hps, do_pen, st_ctx=None, log=None, line_cb=None):
            log = log if log is not None else []
            def _append(m, _l=log):
                _l.append(m)
                if line_cb:
                    line_cb()
            # Download dokumen SPSE
            if do_dl and id_pesan and kode_tender:
                for _dl_attempt in (1, 2):
                    try:
                        _dh = inbox_engine.download_dokumen_paket(
                            kode_tender, str(id_pesan), target_path,
                            kode_pokja=kode_pokja or "",
                            progress_cb=_append,
                            st_ctx=st_ctx,
                        )
                        _cdp_gagal = not _dh["ok"] and not _dh.get("draft_pdf")
                        if _cdp_gagal:
                            log.append("❌ Brave CDP tidak aktif — buka Brave lalu ulangi")
                        else:
                            log.append(
                                f"📎 Download: ✅{len(_dh['ok'])} file"
                                + (f" | Draft: {_os.path.basename(_dh['draft_pdf'])}" if _dh.get('draft_pdf') else " | ⚠ Draft tidak terbuat")
                            )
                        for _e in _dh.get("error", []):
                            log.append(f"  ❌ {_e}")
                        break
                    except Exception as _e:
                        _cdp_putus = "connection closed" in str(_e).lower() or "target page" in str(_e).lower()
                        if _cdp_putus and _dl_attempt == 1:
                            log.append(f"⚠ CDP terputus, tunggu 2 detik lalu retry... ({_e})")
                            import time as _time_retry
                            _time_retry.sleep(2)
                            continue
                        import traceback as _tb_dl
                        log.append(f"❌ Download error: {_e}")
                        log.append(f"   Traceback: {_tb_dl.format_exc()[-500:]}")
                        break
            # Scrape HPS → Excel
            _xl = _cari_xlsm_tender(target_path) if (do_hps or do_pen) else None
            if do_hps and kode_tender:
                if not _xl:
                    log.append("⚠ HPS dilewati — tidak ada .xlsm di folder")
                else:
                    try:
                        import hps_engine as _hps_eng2
                        _hr = _hps_eng2.scrape_hps_ke_excel(kode_tender, _xl)
                        if _hr.get("ok"):
                            log.append(f"📊 HPS: {_hr.get('count',0)} baris → Excel")
                            if _hr.get("md_path"):
                                log.append(f"📝 HPS MD: {os.path.basename(_hr['md_path'])}")
                        else:
                            log.append(f"⚠ HPS: {_hr.get('pesan','-')}")
                    except Exception as _e:
                        log.append(f"⚠ HPS gagal: {_e}")
            # Scrape Penawaran → Excel
            if do_pen and kode_tender and _xl:
                try:
                    import penawaran_engine as _pen_eng2
                    _pr = _pen_eng2.scrape_penawaran_ke_excel(kode_tender, _xl)
                    if _pr.get("peserta"):
                        log.append(f"💰 Penawaran: {_pr['peserta']} peserta → Excel")
                        _pru = _pen_eng2.update_rumus_penawaran_72(_xl)
                        if _pru.get("ok"):
                            log.append(f"  Rumus 7.2: {_pru['rows_updated']} baris")
                    else:
                        log.append(f"⚠ Penawaran: {_pr.get('errors', ['-'])}")
                except Exception as _e:
                    log.append(f"⚠ Penawaran gagal: {_e}")
            return log

        def _serap_satu_paket_tender(kode_tender):
            """Refresh inbox + daftar SPSE hanya untuk satu paket."""
            _hasil_inbox = inbox_engine.serap_inbox(kode_tender=kode_tender)
            _hasil_spse = kirimpesan_engine.fetch_paket_satu(kode_tender)
            _paket_spse = _hasil_spse.get("paket", []) if _hasil_spse.get("sukses") else []
            if _paket_spse:
                _tahap_satu = kirimpesan_engine.fetch_tahap_tender(_paket_spse)
                kirimpesan_engine.enrich_paket_supabase(_paket_spse, tahap_map=_tahap_satu)

                for _cache_key in ("global_paket_draft", "global_paket_aktif"):
                    _cache = st.session_state.get(_cache_key, {})
                    _cache["paket"] = [
                        _p for _p in _cache.get("paket", [])
                        if str(_p.get("kode")) != str(kode_tender)
                    ]
                    st.session_state[_cache_key] = _cache
                _satu = _paket_spse[0]
                _status_satu = str(_satu.get("status") or "").lower()
                _target_cache = "global_paket_draft" if "draft" in _status_satu else "global_paket_aktif"
                st.session_state.setdefault(_target_cache, {}).setdefault("paket", []).append(_satu)
                _tahap_satu_value = _tahap_satu.get(str(kode_tender), "")
                if _tahap_satu_value:
                    st.session_state.setdefault("tender_tahap_map", {})[str(kode_tender)] = _tahap_satu_value
                kirimpesan_engine.save_paket_cache(
                    st.session_state.get("global_paket_draft", {"paket": []}),
                    st.session_state.get("global_paket_aktif", {"paket": []}),
                )
            _load_draft_paket_cached.clear()
            return _hasil_inbox, _hasil_spse

        # ── Checkbox aksi global (jalan untuk tiap paket saat buat folder) ──
        st.caption("Pilih aksi yang dijalankan otomatis untuk tiap paket saat buat folder.")
        _t_cb_dl  = st.checkbox("📦 Download dokumen SPSE + lampiran", value=True, key="t_cb_dl_dokumen")
        _t_cb_hps = st.checkbox("📊 Scrape HPS → Excel", value=True, key="t_cb_hps")
        st.divider()

        # ── Daftar paket: belum-folder (checklist) + sudah-folder (expander aksi) ──
        # ── Filter paket selesai — baca dari status_tahap Supabase (independent dari session SPSE) ──
        _selesai_kodes = {
            str(_r.get("kode_tender", ""))
            for _r in _draft_rows
            if _is_tender_selesai({"status": _r.get("status_tahap") or ""})
        }
        _t_show_done = st.checkbox(
            "Tampilkan paket selesai (Masa Sanggah / Penunjukan / Penandatanganan)",
            value=False,
            key="t_show_done_tender",
        )
        _t_done_n = len(_selesai_kodes)

        # Checkbox hide paket sudah-folder (tiru pola PL hide-selesai)
        _t_hide_done = st.checkbox(
            "🙈 Sembunyikan paket yang sudah punya folder",
            value=True,
            key="t_hide_done_chk",
        )
        def _tender_tahun_cocok(_r):
            _npp = str(_r.get("nomor_pp") or "")
            if _npp:
                return _tahun_skrg in _npp
            # nomor_pp belum keisi (surat PP belum ada) → anggap baru HANYA jika
            # diambil_pada masih segar (bukan bulk-import historis, bisa "tahun ini" palsu)
            _dp = str(_r.get("diambil_pada") or "")
            try:
                _dp_dt = datetime.fromisoformat(_dp.replace("Z", "+00:00"))
                if _dp_dt.tzinfo is not None:
                    _dp_dt = _dp_dt.replace(tzinfo=None)
                return (datetime.now() - _dp_dt).days <= 14
            except (ValueError, TypeError):
                return False

        # Folder di DB belum dianggap sehat bila workbook tidak menyimpan C4
        # paket target. Tampilkan kembali sebagai "belum" agar create-folder
        # dapat retry folder existing tanpa scrub data.
        for _r_local in _draft_rows:
            _folder_db = _r_local.get("folder_dibuat")
            if _folder_db and not _tender_folder_identity_valid(
                str(_folder_db), str(_r_local.get("kode_tender") or "")
            ):
                _r_local["folder_dibuat"] = None
            if not _r_local.get("folder_dibuat"):
                _folder_local = _find_tender_folder_local(
                    _r_local.get("kode_tender"),
                    _r_local.get("kode_pokja"),
                    _r_local.get("nama_tender"),
                )
                if _folder_local:
                    _r_local["folder_dibuat"] = _folder_local

        _rows_valid = [
            _r for _r in _draft_rows
            if _r.get("nama_tender")
            and not str(_r.get("kode_tender", "")).startswith("_err_")
            and not _is_tender_excluded(_r)
            and _tender_tahun_cocok(_r)
            and (_t_show_done or str(_r.get("kode_tender", "")) not in _selesai_kodes)
        ]
        if _t_done_n and not _t_show_done:
            st.caption(f"🔒 {_t_done_n} paket selesai (Masa Sanggah/Penunjukan/Penandatanganan) disembunyikan — centang di atas untuk tampilkan.")
        _rows_belum = [
            _r for _r in _rows_valid
            if not _r.get("folder_dibuat")
        ]
        # Filter rows_sudah: sembunyikan paket selesai kecuali _t_show_done
        _rows_sudah = [
            _r for _r in _rows_valid
            if _r.get("folder_dibuat")
            and (_t_show_done or str(_r.get("kode_tender", "")) not in _selesai_kodes)
        ]

        # Plan nama folder per paket belum-folder (auto-nomor)
        _max_urut_db = max((int(_r.get("nomor_urut") or 0) for _r in _rows_tahun_ini), default=0)
        _max_urut_disk = _nomor_folder_tertinggi_tender(_TENDER_ROOT)
        _max_urut = max(_max_urut_db, _max_urut_disk)
        _t_plan, _ctr = {}, _max_urut
        import kode_unik_engine as _ku_engine
        for _r in sorted(_rows_belum, key=lambda x: x.get("diambil_pada") or ""):
            _n = int(_r["nomor_urut"]) if _r.get("nomor_urut") else (_ctr := _ctr + 1) and _ctr
            _nama_tender = str(_r.get("nama_tender", "")).strip()
            # Pakai kode_unik dari DB kalau sudah ada, generate kalau belum
            _ku = _r.get("kode_unik") or _ku_engine.generate_kode_unik(_nama_tender)
            _t_plan[_r["kode_tender"]] = {
                "kode_tender": _r["kode_tender"],
                "nomor_urut": _n,
                "kode_unik": _ku,
                "nama_folder": re.sub(r'[/<>:"\|?*\\]', "-", f"{_n}. {_nama_tender} - Pokja {str(_r.get('kode_pokja','')).strip()}").strip(),
                "id_pesan": _r.get("id_pesan", ""),
                "kode_pokja": _r.get("kode_pokja", ""),
                "row": _r,
            }

        # ── Checklist pilih paket untuk buat folder (tiru pola PL) ──
        st.markdown("**Pilih paket yang akan dibuat foldernya:**")
        if _rows_belum:
            _tf_c1, _tf_c2 = st.columns(2)
            if _tf_c1.button("✅ Pilih Semua", key="tf_pilih_semua", use_container_width=True):
                for _r in _rows_belum:
                    st.session_state[f"tf_chk_{_r['kode_tender']}"] = True
                st.rerun()
            if _tf_c2.button("❌ Batal Semua", key="tf_batal_semua", use_container_width=True):
                for _r in _rows_belum:
                    st.session_state[f"tf_chk_{_r['kode_tender']}"] = False
                st.rerun()
            for _r in _rows_belum:
                _kt = _r["kode_tender"]
                _ck = f"tf_chk_{_kt}"
                if _ck not in st.session_state:
                    st.session_state[_ck] = True
                _pk = str(_r.get("kode_pokja") or "").strip()
                st.checkbox(
                    f"{_t_plan.get(_kt, {}).get('nomor_urut', '')}. {str(_r.get('nama_tender',''))} — Pokja {_pk}",
                    key=_ck,
                    help=f"Kode unik: {_t_plan.get(_kt, {}).get('kode_unik', '...')}",
                )
            _t_terpilih = [
                _t_plan[_r["kode_tender"]] for _r in _rows_belum
                if st.session_state.get(f"tf_chk_{_r['kode_tender']}", True)
            ]
            with st.expander(f"📋 Preview {len(_t_terpilih)} folder"):
                for _bp in _t_terpilih:
                    st.caption(_bp["nama_folder"])

            if st.button(
                f"📁 Buat Folder Terpilih ({len(_t_terpilih)} paket)",
                disabled=len(_t_terpilih) == 0,
                use_container_width=True,
                key="t_btn_buat_terpilih",
                type="primary",
            ):
                from datetime import timezone as _tz2
                from streamlit.runtime.scriptrunner import get_script_run_ctx as _get_ctx
                _ctx_bulk = _get_ctx()
                _bp2 = st.progress(0.0)
                # Status proses dibuat ringkas. Jika dibuka expanded, seluruh log
                # live tampil di atas sisa layout halaman dan terlihat seperti
                # halaman kedua/"bayangan" saat operasi folder masih berjalan.
                _bulk_status = st.status(
                    f"📁 Memproses {len(_t_terpilih)} paket terpilih...",
                    expanded=False,
                )
                _bulk_status_line = _bulk_status.empty()
                _ok, _fail = 0, 0
                _bulk_semua_log = {}
                try:
                  for _i, _bp in enumerate(_t_terpilih):
                    _bp2.progress((_i + 1) / len(_t_terpilih))
                    _nf = _bp["nama_folder"]
                    _bulk_status.update(label=f"[{_i+1}/{len(_t_terpilih)}] {_nf[:60]}")
                    _paket_log = []
                    try:
                        _r2 = _sp.run([_PY, _SCRIPT, "--output-dir", _TENDER_ROOT, _nf],
                                      capture_output=True, text=True, timeout=60,
                                       creationflags=_NO_WIN)
                        if _r2.returncode == 0:
                            _paket_log.append("✅ Folder dibuat")
                            # Simpan kode_unik ke Supabase
                            _ku_val = _bp.get("kode_unik")
                            if _ku_val:
                                try:
                                    from config import sb as _sb_ku
                                    _sb_ku().table("tender").update(
                                        {"kode_unik": _ku_val}
                                    ).eq("kode_tender", _bp["kode_tender"]).execute()
                                    _paket_log.append(f"🔑 Kode unik: {_ku_val}")
                                except Exception as _ku_err:
                                    _paket_log.append(f"⚠️ Gagal simpan kode_unik: {_ku_err}")
                            _bp_target = os.path.join(_TENDER_ROOT, _nf)
                            # Copy file evaluator AI ke folder paket Tender PK
                            try:
                                import shutil as _t_shutil
                                _t_eval_files = [
                                    "PROTOKOL_EVALUASI_AI.md",
                                    "EVALUATOR_E2E_TENDER_PK_PASCAKUALIFIKASI.md",
                                ]
                                _t_draft_eval_files = [
                                    "EVALUATOR_PRA_REVIU_DPP_TENDER_PK.md",
                                    "PANDUAN_PATCH_MANUAL_EVALUASI.md",
                                ]
                                _t_eval_copied = []
                                _t_eval_src_dir = os.path.join(_POKJA_ROOT, "_SOP Evaluator")
                                for _tef in _t_eval_files:
                                    _tef_src = os.path.join(_t_eval_src_dir, _tef)
                                    if os.path.isfile(_tef_src):
                                        _t_eval_dir = os.path.join(_bp_target, "5. Evaluator Kualifikasi & Teknis")
                                        os.makedirs(_t_eval_dir, exist_ok=True)
                                        _t_shutil.copy2(_tef_src, os.path.join(_t_eval_dir, _tef))
                                        _t_eval_copied.append(_tef)
                                for _tef in _t_draft_eval_files:
                                    _tef_src = os.path.join(_t_eval_src_dir, _tef)
                                    if os.path.isfile(_tef_src):
                                        _t_draft_dir = os.path.join(_bp_target, "0. Draft Dokumen PPK")
                                        os.makedirs(_t_draft_dir, exist_ok=True)
                                        _t_shutil.copy2(_tef_src, os.path.join(_t_draft_dir, _tef))
                                        _t_eval_copied.append(_tef)
                                if _t_eval_copied:
                                    _paket_log.append(f"📄 Evaluator: {len(_t_eval_copied)} file disalin")
                                else:
                                    _paket_log.append("⚠ Evaluator: tidak ada file ditemukan di master _SOP Evaluator")
                            except Exception as _t_eval_e:
                                _paket_log.append(f"⚠ Evaluator copy: {_t_eval_e}")
                            # Download dulu agar Draft_Pokja tersedia untuk parser.
                            # HPS sengaja dijalankan setelah Master Data supaya
                            # dua instance COM tidak berebut workbook.
                            def _line_cb(_log=_paket_log):
                                _bulk_status_line.code("\n".join(_log[-10:]))
                            _jalankan_aksi_tender(
                                _bp["kode_tender"], _bp.get("id_pesan"), _bp.get("kode_pokja"),
                                _bp_target, _t_cb_dl, False, False,
                                st_ctx=_ctx_bulk, log=_paket_log,
                                line_cb=_line_cb,
                            )
                            _bulk_status_line.code("\n".join(_paket_log[-10:]))
                            # Dedicated COM per paket lebih lambat sedikit, tetapi
                            # menghindari cache/lock workbook pada batch panjang.
                            _excel_row = {
                                **(_bp.get("row") or {}),
                                "kode_unik": _bp.get("kode_unik") or "",
                            }
                            _excel_t_logs = _proses_excel_paket_tender(
                                _bp_target,
                                _bp["kode_tender"],
                                xl=None,
                                row_data=_excel_row,
                            )
                            _paket_log.extend(_excel_t_logs)
                            _md_ok = "Master Data Tender: terisi otomatis" in _excel_t_logs
                            if not _md_ok:
                                _paket_log.append(
                                    "❌ Folder belum disahkan: @ Master Data gagal/identitas tidak cocok"
                                )
                            elif _t_cb_hps:
                                _jalankan_aksi_tender(
                                    _bp["kode_tender"], _bp.get("id_pesan"), _bp.get("kode_pokja"),
                                    _bp_target, False, True, False,
                                    st_ctx=_ctx_bulk, log=_paket_log,
                                    line_cb=_line_cb,
                                )
                            _bulk_status_line.code("\n".join(_paket_log[-10:]))
                            _lamp_ok, _lamp_msg = _generate_lampiran_undangan_dpp_preview(_bp["kode_tender"])
                            _paket_log.append(
                                f"📎 Lampiran DPP: {_lamp_msg}" if _lamp_ok else f"⚠ Lampiran DPP: {_lamp_msg}"
                            )
                            _bulk_status_line.code("\n".join(_paket_log[-10:]))
                            # Snapshot dokumen PPK + download awal (folder baru = semua file dianggap "baru")
                            try:
                                import dokumen_ppk_engine as _dpk
                                _snap = _dpk.ambil_snapshot(_bp["kode_tender"])
                                _dpk.simpan_snapshot(_bp["kode_tender"], _snap)
                                _n_snap = sum(len(v) for v in _snap.values())
                                _paket_log.append(f"📸 Snapshot PPK: {_n_snap} file")
                                if _n_snap:
                                    _items_baru0 = [
                                        {"jenis": _jk, **_f}
                                        for _jk, _fl in _snap.items() for _f in _fl
                                    ]
                                    _dl_awal = _dpk.download_update_dokumen(
                                        _bp["kode_tender"], _bp_target, [], _items_baru0, {},
                                        progress_cb=lambda m, _l=_paket_log: _l.append(m),
                                    )
                                    # Pindah dari root ke subfolder sesuai jenis dokumen
                                    _sub_map0 = {
                                        "spek": "1. KAK & Spesifikasi Teknis",
                                        "docsskk": "2. Rancangan Kontrak",
                                        "uploaduraian": "3. Uraian Singkat Pekerjaan",
                                        "ldk": "8. Dokumen Kualifikasi",
                                        "lainnya": "4. Informasi Lainnya",
                                    }
                                    for _it0 in _items_baru0:
                                        _fn0 = re.sub(r'[<>:"/\\|?*]', "_", _it0["nama"]).strip()
                                        _src0 = os.path.join(_bp_target, _fn0)
                                        _sub0 = _sub_map0.get(_it0["jenis"], "4. Informasi Lainnya")
                                        _dst0 = os.path.join(_bp_target, _sub0, _fn0)
                                        if os.path.exists(_src0):
                                            os.makedirs(os.path.dirname(_dst0), exist_ok=True)
                                            shutil.move(_src0, _dst0)
                                    # Snapshot awal: "File Baru/" cuma arsip sementara, tidak relevan (belum ada histori revisi)
                                    _folder_baru0 = os.path.join(_bp_target, "File Baru")
                                    if os.path.isdir(_folder_baru0):
                                        shutil.rmtree(_folder_baru0, ignore_errors=True)
                                    _paket_log.append(
                                        f"📥 Dok PPK: ✅{len(_dl_awal['ok'])} file"
                                        + (f", ❌{len(_dl_awal['error'])}" if _dl_awal["error"] else "")
                                    )
                            except Exception as _se:
                                _paket_log.append(f"⚠ Snapshot/download gagal: {_se}")
                            if _md_ok:
                                try:
                                    inbox_engine._sb().table("draft_paket").update({
                                        "nomor_urut": _bp["nomor_urut"],
                                        "folder_dibuat": _nf,
                                        "folder_dibuat_pada": datetime.now(_tz2.utc).isoformat(),
                                    }).eq("kode_tender", _bp["kode_tender"]).execute()
                                    _load_draft_paket_cached.clear()
                                    _ok += 1
                                except Exception as _db_err:
                                    _fail += 1
                                    _paket_log.append(f"❌ Gagal simpan status folder: {_db_err}")
                            else:
                                _fail += 1
                        else:
                            _fail += 1
                            _stderr_setup = (_r2.stderr or _r2.stdout or "Tanpa detail error").strip()
                            _paket_log.append(f"❌ Gagal buat folder:\n{_stderr_setup[-4000:]}")
                    except _sp.TimeoutExpired:
                        _fail += 1
                        _paket_log.append("❌ Timeout buat folder")
                    _bulk_semua_log[_nf] = _paket_log
                finally:
                    pass

                _bulk_status.update(
                    label=f"✅ {_ok} folder berhasil, ❌ {_fail} gagal",
                    state="error" if _fail else "complete",
                    expanded=True,
                )
                st.session_state["_folder_bulk_created"] = f"{_ok} folder berhasil, {_fail} gagal"
                st.session_state["_folder_bulk_had_error"] = bool(_fail)
                st.session_state["_folder_bulk_last_summary"] = st.session_state["_folder_bulk_created"]
                st.session_state["_folder_bulk_last_error"] = bool(_fail)
                st.session_state["_folder_bulk_log"] = dict(_bulk_semua_log)
                st.rerun()
        else:
            st.info("Semua paket tahun ini sudah punya folder.")
            st.button(
                "📁 Buat Folder Terpilih (0 paket)",
                disabled=True,
                use_container_width=True,
                key="t_btn_buat_terpilih_empty",
                type="primary",
            )

        if st.session_state.get("_folder_bulk_log"):
            _hasil_bulk = st.session_state.get("_folder_bulk_last_summary", "")
            if st.session_state.get("_folder_bulk_last_error"):
                st.warning(f"⚠️ Hasil terakhir: {_hasil_bulk}")
            else:
                st.success(f"✅ Hasil terakhir: {_hasil_bulk}")
            with st.expander("📋 Log detail per paket (hasil terakhir)", expanded=True):
                for _nf, _logs in st.session_state["_folder_bulk_log"].items():
                    st.markdown(f"**{_nf[:70]}**")
                    st.code("\n".join(_logs))
                if st.button("Tutup log", key="t_btn_tutup_log_bulk"):
                    st.session_state["_folder_bulk_log"] = None
                    st.rerun()

        st.divider()
        st.markdown("#### ↩️ Reset Status Folder")
        _opsi_reset_t = {r.get("kode_tender"): r for r in _rows_valid if r.get("folder_dibuat") and r.get("kode_tender")}
        if _opsi_reset_t:
            _label_reset_t = {
                f"{r.get('folder_dibuat')} — {r.get('kode_tender')}": k
                for k, r in _opsi_reset_t.items()
            }
            _pilih_reset_t = st.multiselect(
                "Pilih paket yang direset",
                list(_label_reset_t.keys()),
                key="t_reset_folder_select",
            )
            _c_reset_sel_t, _c_reset_all_t = st.columns(2)
            if _c_reset_sel_t.button("Reset Terpilih", key="t_btn_reset_folder_selected", use_container_width=True, disabled=not _pilih_reset_t):
                from config import sb as _sb_reset_t
                _kodes_reset_t = [_label_reset_t[x] for x in _pilih_reset_t]
                try:
                    _sb_reset_t().table("draft_paket").update({"folder_dibuat": None}).in_("kode_tender", _kodes_reset_t).execute()
                    st.success(f"✅ {len(_kodes_reset_t)} paket berhasil direset.")
                except Exception as _er_t:
                    st.error(f"Reset gagal: {_er_t}")
                _load_draft_paket_cached.clear()
                st.rerun()
            if _c_reset_all_t.button("Reset Semua", key="t_btn_reset_folder_all", use_container_width=True):
                from config import sb as _sb_reset_t
                _kodes_reset_t = list(_opsi_reset_t.keys())
                try:
                    _sb_reset_t().table("draft_paket").update({"folder_dibuat": None}).in_("kode_tender", _kodes_reset_t).execute()
                    st.success(f"✅ {len(_kodes_reset_t)} paket berhasil direset.")
                except Exception as _er_t:
                    st.error(f"Reset gagal: {_er_t}")
                _load_draft_paket_cached.clear()
                st.rerun()
        else:
            st.info("Tidak ada paket dengan status folder untuk direset.")

        # ── Cek Semua Dokumen PPK (batch) ── (Dipindah ke sini agar selalu di bawah seksi 2)
        st.divider()
        if st.button("🔍 Cek Semua Dokumen PPK", use_container_width=True, key="btn_cek_semua_dok"):
            import dokumen_ppk_engine as _dpk_batch
            # Ambil semua paket yg punya dokumen_snapshot dari Supabase
            _snap_rows = inbox_engine._sb().table("draft_paket") \
                .select("kode_tender, folder_dibuat, dokumen_snapshot") \
                .not_.is_("dokumen_snapshot", "null") \
                .execute()
            # Hanya cek paket yang sedang tampil di Tab 0 (tahun/status aktif).
            # Query snapshot sebelumnya mengambil seluruh histori draft_paket,
            # sehingga paket lama ikut muncul kembali.
            _kode_tender_valid = {
                str(r.get("kode_tender", "")) for r in _rows_valid
                if r.get("kode_tender")
            }
            _kode_tender_terkini = {
                str(p.get("kode", "")) for p in _get_paket_gabungan(filter_selesai=False)
                if p.get("kode")
            }
            _snap_paket = [
                r for r in (_snap_rows.data if _snap_rows.data else [])
                if not _is_tender_excluded(r)
                and str(r.get("kode_tender", "")) in _kode_tender_valid
                and str(r.get("kode_tender", "")) in _kode_tender_terkini
            ]
            if not _snap_paket:
                st.info("Belum ada paket dengan snapshot dokumen. Buat folder paket dulu.")
            else:
                _nama_map = {p["kode"]: p["nama"] for p in _get_paket_gabungan()}
                _hasil_batch = []
                with st.status(f"Memeriksa {len(_snap_paket)} paket...", expanded=True) as _cek_st:
                    for _sp in _snap_paket:
                        _kt = _sp["kode_tender"]
                        _nama = _nama_map.get(_kt) or _sp.get("folder_dibuat") or _kt
                        _cek_st.write(f"🔍 {_nama[:50]}...")
                        try:
                            _diff = _dpk_batch.cek_update_dokumen(_kt)
                            _snap_baru_total = sum(len(v) for v in _diff["snapshot_baru"].values())
                            _cookie_invalid = _snap_baru_total == 0
                            _ada_update = (not _cookie_invalid) and bool(
                                _diff["berubah"] or _diff["baru"] or _diff.get("hilang")
                            )
                            _hasil_batch.append({
                                "kode": _kt,
                                "nama": _nama,
                                "berubah": _diff["berubah"],
                                "baru": _diff["baru"],
                                "hilang": _diff.get("hilang", []),
                                "ada_update": _ada_update,
                                "cookie_invalid": _cookie_invalid,
                            })
                        except Exception as _e_cek:
                            _hasil_batch.append({
                                "kode": _kt,
                                "nama": _nama,
                                "error": str(_e_cek),
                                "ada_update": False,
                            })
                    _cek_st.update(label="✅ Selesai cek dokumen PPK", state="complete")
                _folder_map = {r["kode_tender"]: r.get("folder_dibuat", "") for r in _snap_paket}
                st.session_state["_batch_cek_hasil"] = _hasil_batch
                st.session_state["_batch_folder_map"] = _folder_map

        # Tampil hasil batch (persist setelah rerun)
        if "_batch_cek_hasil" in st.session_state:
            _bh = st.session_state["_batch_cek_hasil"]
            _ada_update_list = [x for x in _bh if x.get("ada_update")]
            _error_list = [x for x in _bh if x.get("error")]
            _cookie_invalid_list = [x for x in _bh if x.get("cookie_invalid") and not x.get("error")]
            _sama_list = [
                x for x in _bh
                if not x.get("ada_update") and not x.get("error") and not x.get("cookie_invalid")
            ]
            if _cookie_invalid_list:
                st.error(f"⚠️ Cookie SPSE expired ({len(_cookie_invalid_list)} paket tidak bisa dicek). Login ulang di Brave.")
            if _ada_update_list:
                st.warning(f"⚠️ {len(_ada_update_list)} paket ada update dokumen PPK")
                _folder_map_bh = st.session_state.get("_batch_folder_map", {})
                for _item in _ada_update_list:
                    with st.expander(f"📄 {_item['nama'][:60]}"):
                        if _item.get("cookie_invalid"):
                            st.error("⚠️ Cookie SPSE expired — login ulang di Brave lalu cek lagi")
                        for _b in _item.get("berubah", []):
                            st.markdown(f"- **Berubah** [{_b['jenis']}]: `{_b['nama_lama']}` → `{_b['nama_baru']}`")
                        for _b in _item.get("baru", []):
                            st.markdown(f"- **File Baru** [{_b['jenis']}]: `{_b['nama']}`")
                        for _b in _item.get("hilang", []):
                            st.markdown(f"- **File Hilang** [{_b['jenis']}]: `{_b['nama']}` — mungkin diganti")
                        _kt_dl = _item["kode"]
                        _fd_dl = _folder_map_bh.get(_kt_dl, "")
                        _folder_dl = _os.path.join(_TENDER_ROOT, _fd_dl) if _fd_dl else ""
                        if _folder_dl and _os.path.exists(_folder_dl):
                            st.button("⬇️ Download Update", key=f"btn_dl_upd_{_kt_dl}", type="primary")
                        else:
                            st.caption(f"⚠️ Folder tidak ditemukan: `{_folder_dl or 'tidak diketahui'}`")

                # Proses download di luar expander
                for _item in _ada_update_list:
                    _kt_dl = _item["kode"]
                    if st.session_state.get(f"btn_dl_upd_{_kt_dl}"):
                        _fd_dl = _folder_map_bh.get(_kt_dl, "")
                        _folder_dl = _os.path.join(_TENDER_ROOT, _fd_dl) if _fd_dl else ""
                        import dokumen_ppk_engine as _dpk_dl
                        _sn_r2 = inbox_engine._sb().table("draft_paket").select("dokumen_snapshot").eq("kode_tender", _kt_dl).execute()
                        _sn_lama2 = {}
                        if _sn_r2.data and _sn_r2.data[0].get("dokumen_snapshot"):
                            _raw2 = _sn_r2.data[0]["dokumen_snapshot"]
                            _sn_lama2 = _raw2 if isinstance(_raw2, dict) else __import__("json").loads(_raw2)
                        _diff_dl = _dpk_dl.cek_update_dokumen(_kt_dl)
                        _dl_log4 = []
                        _dl_st4 = st.status(f"⬇️ Mengunduh update {_item['nama'][:40]}...", expanded=True)
                        _dl_area4 = _dl_st4.empty()
                        def _dl_cb4(msg, _log=_dl_log4, _area=_dl_area4, _st=_dl_st4):
                            _log.append(msg)
                            _area.code("\n".join(_log[-15:]))
                            _st.update(label=f"⬇️ {msg[:60]}...")
                        _dl_res4 = _dpk_dl.download_update_dokumen(
                            _kt_dl, _folder_dl,
                            _diff_dl["berubah"], _diff_dl["baru"],
                            _sn_lama2, progress_cb=_dl_cb4,
                        )
                        _download_final_ok = False
                        if _dl_res4["error"]:
                            _dl_st4.update(
                                label=f"❌ {len(_dl_res4['ok'])} file diupdate, {len(_dl_res4['error'])} gagal — snapshot tidak dimajukan",
                                state="error", expanded=True,
                            )
                            for _e6 in _dl_res4["error"]:
                                st.error(_e6)
                        else:
                            # Download sukses penuh: parse Draft lokal + isi @ Master Data
                            # sebelum snapshot dimajukan. Jika Excel sedang terbuka atau
                            # COM gagal, snapshot ditahan agar update bisa dicoba ulang.
                            _row_update = next(
                                (r for r in _draft_rows if str(r.get("kode_tender", "")) == str(_kt_dl)),
                                None,
                            )
                            _md_logs = _proses_excel_paket_tender(
                                _folder_dl, _kt_dl, row_data=_row_update
                            )
                            _md_gagal = any(
                                "WARN" in str(_line).upper()
                                or "ERROR" in str(_line).upper()
                                for _line in _md_logs
                            )
                            _dl_log4.extend(f"🧾 {_line}" for _line in _md_logs)
                            if _md_gagal:
                                _dl_st4.update(
                                    label="⚠️ File terunduh, sinkronisasi @ Master Data gagal — snapshot ditahan",
                                    state="error", expanded=True,
                                )
                                st.warning(
                                    f"⚠️ {_item['nama'][:50]} — file sudah terunduh, "
                                    "tetapi @ Master Data belum tersinkron. Perbaiki Excel lalu ulangi."
                                )
                                _dl_area4.code("\n".join(_dl_log4[-20:]))
                            else:
                                _dpk_dl.simpan_snapshot(_kt_dl, _diff_dl["snapshot_baru"])
                                _download_final_ok = True
                                _dl_st4.update(
                                    label=f"✅ {len(_dl_res4['ok'])} file diupdate + @ Master Data tersinkron",
                                    state="complete", expanded=True,
                                )
                                _dl_area4.code("\n".join(_dl_log4[-20:]))
                                st.success(f"✅ {_item['nama'][:50]} — dokumen dan @ Master Data sudah diperbarui.")
                        if _download_final_ok:
                            st.session_state["_batch_cek_hasil"] = [
                                x for x in st.session_state["_batch_cek_hasil"] if x["kode"] != _kt_dl
                            ]
                        st.rerun()
            else:
                st.success(f"✅ Semua {len(_bh)} paket — tidak ada update dokumen PPK")
            if _sama_list:
                st.caption(
                    f"ℹ️ {_sama_list[0]['nama'][:70]}"
                    if len(_sama_list) == 1
                    else f"ℹ️ {len(_sama_list)} paket diperiksa tanpa perubahan: "
                    + ", ".join(x["nama"][:45] for x in _sama_list)
                )
            if _error_list:
                with st.expander(f"⚠️ {len(_error_list)} paket gagal dicek"):
                    for _item in _error_list:
                        st.caption(f"`{_item['kode']}` — {_item['error'][:80]}")

    # ── Paket sudah-folder: expander per-paket dengan aksi kompak (tiru PL) ──
    if _rows_sudah:
        st.divider()
        st.markdown("#### 3. Status Folder")
        st.write("")
        for _r in _rows_sudah:
            _kt = _r.get("kode_tender", "")
            _pk = str(_r.get("kode_pokja") or "").strip()
            _nm = str(_r.get("nama_tender") or "-")
            _folder_prefix = re.match(r"^\s*(\d+)\s*\.\s*", str(_r.get("folder_dibuat") or ""))
            _nm_display = (
                f"{_folder_prefix.group(1)}. {_nm}"
                if _folder_prefix and not _nm.startswith(f"{_folder_prefix.group(1)}.")
                else _nm
            )
            _fd = _r.get("folder_dibuat", "")
            _tpath = _os.path.join(_TENDER_ROOT, _fd) if _fd else ""
            _ada = bool(_tpath and _os.path.exists(_tpath))
            with st.expander(f"✅ [Pokja {_pk}] {_nm_display}"):
                st.caption(f"`{_kt}` | Folder: {'✅ ada' if _ada else '⚠️ tidak ditemukan'}")
                if not _ada:
                    st.warning(f"Folder fisik tidak ditemukan: `{_tpath}`")
                    continue
                _ac1, _ac2, _ac3, _ac4, _ac5, _ac6 = st.columns(6)
                # 🔄 Serap satu paket
                if _ac1.button(
                    "🔄 Serap Paket", key=f"t_sync_{_kt}", use_container_width=True,
                    help="Update inbox dan sinkronkan daftar SPSE hanya untuk paket ini.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "sync"
                # 📦 Unduh
                if _ac2.button(
                    "📦 Unduh", key=f"t_dl_{_kt}", use_container_width=True,
                    help="Unduh dokumen paket terbaru dari SPSE ke folder paket.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "dl"
                # 💰 HPS
                if _ac3.button(
                    "💰 HPS", key=f"t_hps_{_kt}", use_container_width=True,
                    help="Ambil rincian HPS dari SPSE dan isi ke Excel paket.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "hps"
                # 🔄 Refresh
                if _ac4.button(
                    "🔄 Refresh", key=f"t_ref_{_kt}", use_container_width=True,
                    help="Refresh template Excel paket dari template master.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "ref"
                # 🔁 Parse
                if _ac5.button(
                    "🔁 Parse", key=f"t_parse_{_kt}", use_container_width=True,
                    help="Baca ulang Draft PPK dan sinkronkan metadata paket ke Supabase.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "parse"
                if _ac6.button(
                    "🤖 Pra-Reviu + Isi DOCM", key=f"t_patch_reviu_{_kt}",
                    use_container_width=True,
                    help="Kirim prompt pra-reviu ke Codex, isi jawaban + draft tanggapan ke DOCM, dan buat Markdown audit.",
                ):
                    st.session_state[f"_t_act_{_kt}"] = "patch_reviu"

            # Proses aksi di LUAR expander (hindari nested st.status di expander)
            _act = st.session_state.pop(f"_t_act_{_kt}", None)
            if _act:
                if _act == "sync":
                    with st.status(f"🔄 Menyerap dan menyinkronkan {_nm[:40]}...", expanded=True) as _sync_st:
                        try:
                            _sync_inbox, _sync_spse = _serap_satu_paket_tender(_kt)
                            if _sync_inbox.get("error"):
                                _sync_st.write(f"⚠ Inbox: {len(_sync_inbox['error'])} error")
                            else:
                                _sync_st.write(
                                    f"📥 Inbox: { _sync_inbox.get('baru', 0) } baru / "
                                    f"{ _sync_inbox.get('diperbarui', 0) } diperbarui"
                                )
                            if _sync_spse.get("sukses") and _sync_spse.get("paket"):
                                _sync_st.update(label="✅ Paket berhasil disinkronkan", state="complete")
                            else:
                                _sync_st.update(
                                    label=f"⚠ Paket tidak ditemukan di SPSE: {_sync_spse.get('pesan', '-')}",
                                    state="error",
                                )
                        except Exception as _sync_e:
                            _sync_st.update(label=f"❌ Sinkronisasi gagal: {_sync_e}", state="error")
                    st.rerun()
                elif _act in ("dl", "hps", "pen"):
                    _do_dl  = _act == "dl"
                    _do_hps = _act == "hps"
                    _do_pen = _act == "pen"
                    from streamlit.runtime.scriptrunner import get_script_run_ctx as _grc_act
                    _logA = []
                    _stA = st.status(f"⏳ {_nm[:40]}...", expanded=True)
                    _areaA = _stA.empty()
                    _jalankan_aksi_tender(
                        _kt, _r.get("id_pesan"), _r.get("kode_pokja"),
                        _tpath, _do_dl, _do_hps, _do_pen,
                        st_ctx=_grc_act(), log=_logA,
                    )
                    _areaA.code("\n".join(_logA))
                    _stA.update(label=f"✅ Selesai: {_nm[:40]}", state="complete", expanded=False)

                elif _act == "ref":
                    from refresh_template import refresh_template_paket as _rtt_refresh
                    from pathlib import Path as _rtt_Path
                    with st.spinner(f"🔄 Refresh Template untuk {_nm[:40]}..."):
                        _res = _rtt_refresh(
                            [_rtt_Path(_tpath)],
                            _rtt_Path(_POKJA_ROOT) / "Paket Experiment",
                            "tender",
                            auto_relink=True,
                            dry_run=False
                        )
                        _logs = _res.get(str(_rtt_Path(_tpath)), [])
                        _ok = all("❌" not in l for l in _logs)
                        if _ok:
                            st.success("✅ Refresh Template selesai.")
                        else:
                            st.warning("⚠️ Ada error saat Refresh Template.")
                        with st.expander("Lihat Log"):
                            for _l in _logs:
                                st.caption(_l)

                elif _act == "parse":
                    with st.status(f"⏳ Re-parsing PDF untuk {_nm[:40]}...", expanded=True) as _st_rp:
                        try:
                            _reparse_link = _r.get("link_pdf")
                            if not _reparse_link:
                                _st_rp.write("🔗 Mengambil link PDF dari detail pesan...")
                                _rp_detail = inbox_engine.parse_detail_pesan(str(_r.get("id_pesan", "")))
                                _reparse_link = _rp_detail.get("link_pdf") or ""
                            if not _reparse_link:
                                _st_rp.update(label="❌ Gagal: Link PDF tidak ditemukan", state="error")
                            else:
                                _st_rp.write("📄 Parsing PDF di memory...")
                                _rp_hasil = inbox_engine.parse_pdf_inmemory(_reparse_link)
                                _rp_data  = {k: v for k, v in _rp_hasil.items() if k in inbox_engine._KOLOM_DRAFT_PAKET and v}
                                _rp_data["link_pdf"] = _reparse_link
                                _st_rp.write("💾 Mengupdate Supabase...")
                                inbox_engine._sb().table("draft_paket").update(_rp_data).eq("kode_tender", _kt).execute()
                                _load_draft_paket_cached.clear()
                                _st_rp.update(label=f"✅ Selesai: {_nm[:40]}", state="complete", expanded=True)
                                with st.container(border=True):
                                    for k, v in _rp_hasil.items():
                                        st.text(f"{k}: {v}")
                        except Exception as _rp_e:
                            _st_rp.update(label=f"❌ Gagal re-parse: {_rp_e}", state="error")

# ============================================================
                elif _act == "patch_reviu":
                    import ai_evaluator as _ai_patch
                    _patch_st = st.status(
                        f"🧾 Menjalankan patch manual Isi Reviu {_nm[:40]}...",
                        expanded=True,
                    )
                    _patch_st.write("Membaca SOP dan memvalidasi dokumen paket...")
                    try:
                        _patch_res = _ai_patch.patch_manual_isi_reviu_single(
                            _tpath, _nm, engine="codex"
                        )
                        if _patch_res.get("status") == "ok":
                            _patch_st.write(_patch_res.get("output") or "AI selesai tanpa output tambahan.")
                            _patch_st.update(
                                label=f"✅ Patch Reviu selesai: {_nm[:40]}",
                                state="complete",
                                expanded=True,
                            )
                        else:
                            _patch_st.update(
                                label=f"❌ Patch Reviu gagal: {_nm[:40]}",
                                state="error",
                                expanded=True,
                            )
                            st.error(_patch_res.get("error") or "AI gagal menjalankan patch.")
                    except Exception as _patch_e:
                        _patch_st.update(
                            label=f"❌ Patch Reviu gagal: {_nm[:40]}",
                            state="error",
                            expanded=True,
                        )
                        st.error(str(_patch_e))

# Tab Setup Paket: LDK Auto-fill + Checklist + Masa Berlaku
# ============================================================

if _tender_active_tab == "3️⃣ Setup Paket":
    # ── Layout 2 kolom: kiri = pilih paket + upload dokpil, kanan = konfigurasi ─
    _sp_col_kiri, _sp_col_kanan = st.columns([2, 3])

    with _sp_col_kiri:
        st.markdown("### 1. Pilih Paket")
        col_spfetch, col_spall, col_spnone = st.columns([3, 1, 1])
        with col_spfetch:
            if "global_paket_draft" not in st.session_state and "global_paket_aktif" not in st.session_state:
                st.info("⚠️ Klik **🔄 Sinkronkan Paket** di **Tab 0** dulu.")
            else:
                st.caption(f"📋 {len(_get_paket_gabungan())} paket tersedia (draft + aktif)")

        sp_selected = []
        if "global_paket_draft" in st.session_state or "global_paket_aktif" in st.session_state:
            paket_list_sp = _get_paket_gabungan()
            if not paket_list_sp:
                st.warning("⚠️ Tidak ada paket ditemukan.")
            else:
                with col_spall:
                    if st.button("✅ Semua", key="sp_sel_all", use_container_width=True):
                        for p in paket_list_sp:
                            st.session_state[f"sp_chk_{p['id_lelang']}"] = True
                        st.rerun()
                with col_spnone:
                    if st.button("⬜ Kosong", key="sp_sel_none", use_container_width=True):
                        for p in paket_list_sp:
                            st.session_state[f"sp_chk_{p['id_lelang']}"] = False
                        st.rerun()

                for p in paket_list_sp:
                    key_chk  = f"sp_chk_{p['id_lelang']}"
                    key_dokpil = f"sp_dokpil_{p['id_lelang']}"
                    col_chk, col_dokpil = st.columns([3, 2])
                    with col_chk:
                        checked = st.checkbox(
                            _pokja_label(p),
                            value=st.session_state.get(key_chk, True),
                            key=key_chk,
                        )
                    with col_dokpil:
                        up_dokpil = st.file_uploader(
                            "DOKPIL",
                            type=["pdf"],
                            key=key_dokpil,
                            label_visibility="collapsed",
                        )
                        if up_dokpil:
                            st.caption(f"📄 {up_dokpil.name}")
                    if checked:
                        sp_selected.append({**p, "_dokpil": up_dokpil})

                st.caption(f"**{len(sp_selected)}** dari **{len(paket_list_sp)}** paket dipilih")

    with _sp_col_kiri:
        st.markdown("### 2. Upload Dokumen Pemilihan")
        st.caption("Menggunakan file DOKPIL yang sudah diupload di atas. Tanggal upload, lalu klik Upload.")

        _dp_dengan_file = [p for p in sp_selected if p.get("_dokpil")]
        _dp_tanpa_file  = [p for p in sp_selected if not p.get("_dokpil")]
        dp_selected     = _dp_dengan_file

        if not sp_selected:
            st.info("Pilih paket dan upload DOKPIL di atas terlebih dahulu.")
        else:
            if _dp_dengan_file:
                st.caption(f"✅ **{len(_dp_dengan_file)} paket** siap diupload:")
                for _p in _dp_dengan_file:
                    _ku = _p.get("kode_unik") or "?"
                    _kp = _p.get("kode_pokja") or "?"
                    _nomor_auto = f"000.3.3/01/T/{_ku}/POKJA{_kp}/UKPBJ/2026"
                    st.markdown(f"- {_pokja_label(_p)[:80]}  \n  📄 `{_p['_dokpil'].name}`  \n  📋 `{_nomor_auto}`")
            if _dp_tanpa_file:
                st.caption(f"⚠️ **{len(_dp_tanpa_file)} paket** tanpa DOKPIL (dilewati):")
                for _p in _dp_tanpa_file:
                    st.markdown(f"- {_pokja_label(_p)[:80]}")

        # Cek apakah ada paket terpilih dengan kode_unik null
        _dp_tanpa_kode_unik = [p for p in _dp_dengan_file if not p.get("kode_unik")]
        if _dp_tanpa_kode_unik:
            st.warning(f"⚠️ **{len(_dp_tanpa_kode_unik)} paket** belum punya Kode Unik — generate dulu via tombol 'Kode Unik Surat' di Excel.")

        _dp_col_tgl, _dp_col_btn = st.columns([2, 3])
        with _dp_col_tgl:
            dp_tgl = st.date_input(
                "Tanggal Dokumen",
                value=datetime.now().date(),
                key="dp_tgl",
                format="DD/MM/YYYY",
                label_visibility="collapsed",
            )
            st.caption(f"{_HARI_NAMA[dp_tgl.weekday()]}, {dp_tgl.day} {_BULAN_NAMA[dp_tgl.month-1]} {dp_tgl.year}")
        with _dp_col_btn:
            _dp_n_file = len(dp_selected)
            if st.button(
                f"📤 Upload Dokumen Pemilihan ({_dp_n_file} file)",
                key="dp_upload",
                type="primary",
                disabled=_dp_n_file == 0,
                use_container_width=True,
            ):
                import dokpil_engine
                with st.status(f"Mengupload {_dp_n_file} Dokumen Pemilihan...", expanded=True) as status:
                    sukses_count = 0
                    for _p in dp_selected:
                        _kode = str(_p.get("kode", ""))
                        if not _kode:
                            st.error(f"❌ Paket {_pokja_label(_p)} tidak memiliki kode tender valid.")
                            continue
                            
                        _ku = _p.get("kode_unik") or "?"
                        _kp = _p.get("kode_pokja") or "?"
                        _nomor_auto = f"000.3.3/01/T/{_ku}/POKJA{_kp}/UKPBJ/2026"
                        _tgl_str = dp_tgl.strftime('%d-%m-%Y')
                        _file = _p['_dokpil']
                        
                        st.write(f"⏳ Uploading: `{_file.name}` untuk **{_pokja_label(_p)[:60]}** ...")
                        try:
                            _res = dokpil_engine.upload_dokumen_pemilihan(
                                paket_id=_kode,
                                nomor_sdp=_nomor_auto,
                                tanggal_sdp=_tgl_str,
                                file_bytes=_file.getvalue(),
                                file_name=_file.name
                            )
                            if _res["ok"]:
                                st.success(f"✅ Berhasil upload: `{_file.name}`")
                                sukses_count += 1
                            else:
                                st.error(f"❌ Gagal upload: `{_file.name}` (Status: {_res['status']})")
                        except Exception as e:
                            st.error(f"❌ Error upload `{_file.name}`: {e}")
                    
                    status.update(label=f"Selesai! {sukses_count}/{_dp_n_file} dokumen berhasil diupload.", state="complete" if sukses_count == _dp_n_file else "error")

        with _sp_col_kanan:
            st.markdown("### 3. Konfigurasi")
            st.caption("Upload DOKPIL per paket di sebelah kiri — akan di-extract saat Push.")

        with _sp_col_kanan:
            # ── Izin Usaha rows (fallback jika DOKPIL tidak diupload) ────────────
            st.markdown("**Izin Usaha** *(default — ditimpa oleh DOKPIL jika diupload)*")
            _t_izin_default = (
                "Memiliki perizinan berusaha di bidang Pekerjaan/Jasa Konstruksi. "
                "a) Memiliki Nomor Induk Berusaha (NIB) dan Sertifikat Standar terverifikasi; "
                "b) Dalam hal Sertifikat Standar sebagaimana dimaksud pada huruf a) belum terverifikasi, "
                "peserta menyampaikan NIB, Sertifikat Standar belum terverifikasi dan tangkapan layar laman OSS "
                "yang mencantumkan bahwa Sertifikat Standar sedang menunggu verifikasi"
            )
            if "ijin_rows" not in st.session_state:
                st.session_state["ijin_rows"] = [dict(r) for r in ldk_config.IJIN_USAHA_DEFAULT["rows"]]
            if st.session_state["ijin_rows"] and st.session_state["ijin_rows"][0].get("klasifikasi", "").startswith(
                "Memiliki perizinan berusaha di bidang Jasa Konstruksi."
            ):
                st.session_state["ijin_rows"][0]["klasifikasi"] = _t_izin_default
                if st.session_state.get("ijin_klas_0", "").startswith(
                    "Memiliki perizinan berusaha di bidang Jasa Konstruksi."
                ):
                    st.session_state["ijin_klas_0"] = _t_izin_default
            _t_ijin_custom = st.toggle("Gunakan izin usaha custom", value=True, key="t_ijin_custom")
            st.caption("Mode custom aktif sebagai default. Isi jenis izin dan klasifikasi langsung.")

            # ── Load/save SBU terakhir ke file ───────────────────────────────────
            import json as _json
            from config import RUNTIME_ROOT as _RUNTIME_ROOT
            _SBU_LAST_FILE = os.path.join(_RUNTIME_ROOT, "_sbu_last.json")

            def _load_sbu_last():
                try:
                    with open(_SBU_LAST_FILE, "r", encoding="utf-8") as f:
                        return _json.load(f)
                except Exception:
                    return {"sbu_2020": "", "sbu_2015": ""}

            def _save_sbu_last(sbu_2020, sbu_2015):
                try:
                    with open(_SBU_LAST_FILE, "w", encoding="utf-8") as f:
                        _json.dump({"sbu_2020": sbu_2020, "sbu_2015": sbu_2015}, f)
                except Exception:
                    pass

            _SBU_HISTORY_FILE = os.path.join(_RUNTIME_ROOT, "_sbu_history.json")
            _SBU_PREFIX = "Memiliki Sertifikat Badan Usaha (SBU) dengan Kualifikasi Usaha Kecil, serta disyaratkan:"

            def _load_sbu_history():
                try:
                    with open(_SBU_HISTORY_FILE, "r", encoding="utf-8") as f:
                        values = _json.load(f)
                    return [str(v).strip() for v in values if str(v).strip()]
                except Exception:
                    return []

            def _save_sbu_history(value):
                value = str(value or "").strip()
                if not value:
                    return
                history = [value] + [v for v in _load_sbu_history() if v != value]
                try:
                    with open(_SBU_HISTORY_FILE, "w", encoding="utf-8") as f:
                        _json.dump(history[:20], f, ensure_ascii=False, indent=2)
                except Exception:
                    pass

            if "tender_sbu_history" not in st.session_state:
                st.session_state["tender_sbu_history"] = _load_sbu_history()

            # Inisialisasi default SBU dari file — hanya sekali per session, SEBELUM widget dirender
            if "sbu_last_loaded" not in st.session_state:
                _last = _load_sbu_last()
                # Hanya set jika key belum ada (jangan overwrite pilihan user saat ini)
                if "sbu_2020_1" not in st.session_state:
                    st.session_state["sbu_2020_1"] = _last["sbu_2020"]
                if "sbu_2015_1" not in st.session_state:
                    st.session_state["sbu_2015_1"] = _last["sbu_2015"]
                st.session_state["sbu_last_loaded"] = True

            # Opsi SBU dropdown — dari _SBU_RULES inbox_engine (cached per session)
            if "sbu_opts_cache" not in st.session_state:
                _sbu_cache = ldk_config.load_sbu_dari_rules()
                st.session_state["sbu_opts_cache"] = _sbu_cache
            else:
                _sbu_cache = st.session_state["sbu_opts_cache"]

            _sbu_opts_2020 = [""] + _sbu_cache["kbli_2020"]
            _sbu_opts_2015 = [""] + _sbu_cache["kbli_2015"]

            # Validasi: nilai tersimpan harus ada di opsi — reset jika tidak ada
            if st.session_state.get("sbu_2020_1", "") not in _sbu_opts_2020:
                st.session_state["sbu_2020_1"] = ""
            if st.session_state.get("sbu_2015_1", "") not in _sbu_opts_2015:
                st.session_state["sbu_2015_1"] = ""

            _sbu_sumber = _sbu_cache.get("sumber", "rules")
            st.caption(
                f"📋 {len(_sbu_cache['kbli_2020'])} SBU 2020 · {len(_sbu_cache['kbli_2015'])} SBU 2015 "
                f"(sumber: {_sbu_sumber})"
            )

            for i, row in enumerate(st.session_state["ijin_rows"]):
                if i == 0:
                    continue
                if i == 1:
                    # Row 2 fixed: hanya deskripsi SBU yang terlihat; nama izin + kalimat baku hidden.
                    st.session_state["ijin_rows"][i]["jenis_izin"] = "Sertifikat Badan Usaha SBU"
                    _sbu_new_option = "✏️ Input SBU baru"
                    _sbu_options = st.session_state["tender_sbu_history"] + [_sbu_new_option]
                    _sbu_pick = st.selectbox(
                        "SBU — pilih histori",
                        _sbu_options,
                        index=0,
                        key="tender_sbu_history_pick",
                    )
                    if _sbu_pick == _sbu_new_option:
                        _sbu_desc = st.text_input(
                            "Deskripsi SBU baru",
                            value=st.session_state.get("tender_sbu_new", ""),
                            key="tender_sbu_new",
                            placeholder="Contoh: SBU BS004 Konstruksi Jaringan Irigasi dan Drainase KBLI 42201",
                        )
                    else:
                        _sbu_desc = _sbu_pick
                    st.session_state["ijin_rows"][i]["klasifikasi"] = f"{_SBU_PREFIX}\n{_sbu_desc}" if _sbu_desc else _SBU_PREFIX
                    continue
                st.caption(f"Row {i+1}")

                # Legacy dropdown SBU sengaja tidak dipakai; semua row custom text.
                if False:
                    col_jn, col_del = st.columns([6, 1])
                    with col_jn:
                        st.session_state["ijin_rows"][i]["jenis_izin"] = st.text_input(
                            "Jenis Izin", value=row["jenis_izin"],
                            key=f"ijin_nama_{i}", label_visibility="collapsed", disabled=not _t_ijin_custom,
                        )
                    with col_del:
                        if len(st.session_state["ijin_rows"]) > 1:
                            if st.button("🗑️", key=f"hapus_row_{i}", use_container_width=True):
                                st.session_state["ijin_rows"].pop(i)
                                st.rerun()

                    col_2020, col_2015 = st.columns(2)
                    with col_2020:
                        sbu_2020 = st.selectbox(
                            "SBU KBLI 2020",
                            options=_sbu_opts_2020,
                            key="sbu_2020_1",
                            label_visibility="visible",
                        )
                    with col_2015:
                        sbu_2015 = st.selectbox(
                            "SBU KBLI 2015 (opsional — kosongkan jika hanya SBU 2020)",
                            options=_sbu_opts_2015,
                            key="sbu_2015_1",
                            label_visibility="visible",
                        )

                    # Auto-generate teks klasifikasi dari pilihan dropdown
                    _gen = ldk_config.build_sbu_klasifikasi(sbu_2020, sbu_2015)
                    if _gen:
                        st.session_state["ijin_rows"][i]["klasifikasi"] = _gen
                        st.text_area(
                            "Preview teks SBU",
                            value=_gen,
                            key=f"ijin_klas_{i}_preview",
                            label_visibility="collapsed",
                            height=100,
                            disabled=True,
                        )
                    else:
                        # Fallback: edit manual jika belum pilih SBU
                        st.session_state["ijin_rows"][i]["klasifikasi"] = st.text_area(
                            "Klasifikasi manual",
                            value=row["klasifikasi"],
                            key=f"ijin_klas_{i}",
                            label_visibility="collapsed",
                            height=80,
                        )
                else:
                    col_r1, col_r2, col_r3 = st.columns([2, 5, 1])
                    with col_r1:
                        st.session_state["ijin_rows"][i]["jenis_izin"] = st.text_input(
                            "Jenis Izin", value=row["jenis_izin"],
                            key=f"ijin_nama_{i}", label_visibility="collapsed",
                        )
                    with col_r2:
                        st.session_state["ijin_rows"][i]["klasifikasi"] = st.text_area(
                            "Klasifikasi", value=row["klasifikasi"],
                            key=f"ijin_klas_{i}", label_visibility="collapsed", height=80,
                            disabled=not _t_ijin_custom,
                        )
                    with col_r3:
                        if len(st.session_state["ijin_rows"]) > 1 and _t_ijin_custom:
                            if st.button("🗑️", key=f"hapus_row_{i}", use_container_width=True):
                                st.session_state["ijin_rows"].pop(i)
                                st.rerun()

            # ── Kinerja Penyedia: internal, tidak ditampilkan ────────────────────
            if "sp_syarat_teknis_rows" not in st.session_state:
                st.session_state["sp_syarat_teknis_rows"] = [
                    {"label": "Kinerja Penyedia", "teks": ldk_config.KINERJA_PENYEDIA_DEFAULT}
                ]

            # ── Persyaratan Kualifikasi Tender ───────────────────────────────────
            _t_ldk_admin_ids = ["401", "402", "403", "404"]
            _t_ldk_teknis_ids = ["437", "438", "439"]

            # ── Checklist Dokumen Penawaran ──────────────────────────────────────
            _t_check_admin_ids = []
            _t_check_teknis_ids = ["341", "342", "344"]
            _t_check_harga_ids = ["347", "348"]

            # ── Masa Berlaku ──────────────────────────────────────────────────────
            mb_nilai_hari = 40

            _t_default_ijin = st.session_state.get("ijin_rows", ldk_config.IJIN_USAHA_DEFAULT["rows"])
            _t_sbu_desc = (
                _t_default_ijin[1].get("klasifikasi", "").replace(_SBU_PREFIX, "", 1).strip()
                if len(_t_default_ijin) > 1 else ""
            )
            _t_kinerja_text = "\n".join(
                r["teks"] for i, r in enumerate(st.session_state.get("sp_syarat_teknis_rows", []))
                if st.session_state.get(f"sp_st_chk_{i}", True) and r.get("teks", "").strip()
            )

            if st.button(
                f"📋 Submit Syarat Kualifikasi + Kinerja ({len(sp_selected)} paket)",
                key="sp_submit_ldk_fixed", use_container_width=True, disabled=not sp_selected,
            ):
                _save_sbu_history(_t_sbu_desc)
                for _p in sp_selected:
                    try:
                        _r = tender_setup_engine.submit_ldk(
                            _p["id_lelang"], _t_default_ijin, _t_ldk_admin_ids, _t_ldk_teknis_ids,
                            kinerja_text=_t_kinerja_text,
                        )
                        st.write(f"{'✅' if _r['ok'] else '❌'} {_p['nama'][:45]} — HTTP {_r['status']}")
                    except Exception as _e:
                        st.write(f"❌ {_p['nama'][:45]} — {_e}")

            if st.button(
                f"🛂 Submit 2 Syarat Izin Usaha ({len(sp_selected)} paket)",
                key="sp_submit_ijin_fixed", use_container_width=True, disabled=not sp_selected,
            ):
                _save_sbu_history(_t_sbu_desc)
                for _p in sp_selected:
                    try:
                        _r = tender_setup_engine.submit_izin_usaha(_p["id_lelang"], _t_default_ijin)
                        st.write(f"{'✅' if _r['ok'] else '❌'} {_p['nama'][:45]} — HTTP {_r['status']}")
                    except Exception as _e:
                        st.write(f"❌ {_p['nama'][:45]} — {_e}")

            if st.button(
                f"📄 Submit Checklist Dokumen ({len(sp_selected)} paket)",
                key="sp_submit_checklist_fixed", use_container_width=True, disabled=not sp_selected,
            ):
                _save_sbu_history(_t_sbu_desc)
                for _p in sp_selected:
                    try:
                        _r = tender_setup_engine.submit_checklist(
                            _p["id_lelang"], _t_check_admin_ids, _t_check_teknis_ids, _t_check_harga_ids,
                        )
                        st.write(f"{'✅' if _r['ok'] else '❌'} {_p['nama'][:45]} — HTTP {_r['status']}")
                    except Exception as _e:
                        st.write(f"❌ {_p['nama'][:45]} — {_e}")

            st.caption("Bulk Setup menjalankan LDK + izin usaha + checklist + masa berlaku. Upload DOKPIL tetap tombol terpisah.")
            sp_push = st.button(
                f"🚀 Bulk Setup ke SPSE ({len(sp_selected)} paket)",
                type="primary",
                use_container_width=True,
                disabled=len(sp_selected) == 0,
                key="sp_push_all",
            )

            if sp_push:
                import tempfile

                _sbu_for_history = st.session_state["ijin_rows"][1].get("klasifikasi", "").replace(_SBU_PREFIX, "", 1).strip() if len(st.session_state.get("ijin_rows", [])) > 1 else ""
                _save_sbu_history(_sbu_for_history)

                # Default dari form (dipakai jika paket tidak punya DOKPIL)
                _default_ijin = _t_default_ijin
                mb_hari = int(mb_nilai_hari)

                progress = st.progress(0, text="Memulai...")
                hasil_sp = []

                for i, p in enumerate(sp_selected):
                    pid = p["id_lelang"]
                    progress.progress((i + 1) / len(sp_selected), text=f"Setup {p['kode']} ({i+1}/{len(sp_selected)})...")

                    row_result = {"kode": p["kode"], "nama": p["nama"][:50], "ldk": "—", "checklist": "—", "masa_berlaku": "—"}

                    # Sesuai instruksi: Seluruh konfigurasi LDK (SBU, Izin, Kinerja)
                    # mutlak menggunakan input manual dari UI Streamlit (Tab 3 Konfigurasi).
                    # Tidak ada lagi parsing otomatis dari file DOKPIL PDF.
                    ijin_rows   = _default_ijin
                    row_result["ldk"] = "—"

                    try:
                        r_ldk = tender_setup_engine.submit_ldk(
                            pid,
                            ijin_rows=ijin_rows,
                            admin_ids=_t_ldk_admin_ids,
                            teknis_ids=_t_ldk_teknis_ids,
                            kinerja_text=_t_kinerja_text,
                        )
                        row_result["ldk"] = "✅" if r_ldk["ok"] else f"❌ {r_ldk['status']}"
                    except Exception as e:
                        row_result["ldk"] = f"❌ {e}"

                    try:
                        r_ck = tender_setup_engine.submit_checklist(
                            pid,
                            admin_ids=_t_check_admin_ids,
                            teknis_ids=_t_check_teknis_ids,
                            harga_ids=_t_check_harga_ids,
                        )
                        row_result["checklist"] = "✅" if r_ck["ok"] else f"❌ HTTP {r_ck['status']}"
                    except Exception as e:
                        row_result["checklist"] = f"❌ {e}"

                    try:
                        r_mb = tender_setup_engine.submit_masa_berlaku(pid, mb_hari)
                        row_result["masa_berlaku"] = "✅" if r_mb["ok"] else f"❌ HTTP {r_mb['status']}"
                    except Exception as e:
                        row_result["masa_berlaku"] = f"❌ {e}"

                    hasil_sp.append(row_result)

                progress.empty()

                sukses_n = sum(1 for h in hasil_sp if all(v == "✅" for k, v in h.items() if k not in ("kode", "nama")))
                st.success(f"✅ Selesai! {sukses_n}/{len(hasil_sp)} paket berhasil.")
                st.dataframe(
                    hasil_sp,
                    use_container_width=True,
                    column_config={
                        "kode":         st.column_config.TextColumn("Kode", width="small"),
                        "nama":         st.column_config.TextColumn("Nama Paket", width="large"),
                        "ldk":          st.column_config.TextColumn("LDK", width="small"),
                        "checklist":    st.column_config.TextColumn("Checklist", width="small"),
                        "masa_berlaku": st.column_config.TextColumn("Masa Berlaku", width="small"),
                    },
                    hide_index=True,
                )

# ============================================================
# Tab 4: Pemberian Penjelasan (v2 — auto-post sapaan via GCal)
# ============================================================

if _tender_active_tab == "4️⃣ Pemberian Penjelasan":
    # ── Layout 2 kolom: kiri = pilih paket, kanan = isi pembukaan ───────
    _pj_col_kiri, _pj_col_kanan = st.columns([2, 3])

    with _pj_col_kiri:
        st.markdown("### 1. Pilih Paket")
        col_pjfetch, col_pjall, col_pjnone = st.columns([3, 1, 1])
        with col_pjfetch:
            if "global_paket_draft" not in st.session_state:
                st.info("⚠️ Klik **🔄 Sinkronkan Paket** di **Tab 0** dulu.")
            else:
                st.caption(f"📋 {len(_get_paket_gabungan())} paket tersedia (draft + aktif)")

        pj_selected = []
        if "global_paket_draft" in st.session_state or "global_paket_aktif" in st.session_state:
            paket_list_pj = _get_paket_gabungan()
            if not paket_list_pj:
                st.warning("⚠️ Tidak ada paket ditemukan.")
            else:
                with col_pjall:
                    if st.button("✅ Semua", key="pj_sel_all", use_container_width=True):
                        for p in paket_list_pj:
                            st.session_state[f"pj_chk_{p['id_lelang']}"] = True
                        st.rerun()
                with col_pjnone:
                    if st.button("⬜ Kosong", key="pj_sel_none", use_container_width=True):
                        for p in paket_list_pj:
                            st.session_state[f"pj_chk_{p['id_lelang']}"] = False
                        st.rerun()

                for p in paket_list_pj:
                    key_chk = f"pj_chk_{p['id_lelang']}"
                    checked = st.checkbox(
                        _pokja_label(p),
                        value=st.session_state.get(key_chk, True),
                        key=key_chk,
                    )
                    if checked:
                        pj_selected.append(p)

                st.caption(f"**{len(pj_selected)}** dari **{len(paket_list_pj)}** paket dipilih")

        # ── Status Antrian ─────────────────────────────────────────────────
        st.divider()
        st.markdown("### 2. Status Antrian")
        col_refresh_pj, col_hapus_fired = st.columns(2)
        with col_refresh_pj:
            if st.button("🔄 Refresh", key="pj_refresh_queue", use_container_width=True):
                st.rerun()
        with col_hapus_fired:
            if st.button("🗑️ Hapus yang Selesai", key="pj_hapus_fired", use_container_width=True):
                jobs = penjelasan_engine.get_jobs()
                for j in jobs:
                    if j["status"] in ("fired", "gagal"):
                        penjelasan_engine.hapus_job(j["paket_id"], j["jenis"])
                st.rerun()

        jobs_all = penjelasan_engine.get_jobs()
        if not jobs_all:
            st.info("Antrian kosong. Pilih paket dan klik Jadwalkan.")
        else:
            from penjelasan_engine import TZ_WIB as _TZ_WIB
            now_q = datetime.now(_TZ_WIB)
            rows_q = []
            _tender_labels_q = {
                str(p.get("id_lelang") or p.get("kode") or ""): _pokja_label(p)
                for p in _get_paket_gabungan(filter_selesai=False)
            }
            for j in jobs_all:
                try:
                    wf = datetime.fromisoformat(j["waktu_fire"])
                    secs = int((wf - now_q).total_seconds())
                    if j["status"] == "fired":
                        countdown_q = "✅ Selesai"
                    elif j["status"] == "gagal":
                        countdown_q = "❌ Gagal"
                    elif secs > 0:
                        h, rem = divmod(secs, 3600)
                        m = rem // 60
                        countdown_q = f"⏳ {h//24}h {h%24}j {m}m"
                    else:
                        countdown_q = "🔴 Menunggu scheduler..."
                    waktu_str = wf.strftime("%d/%m %H:%M")
                except Exception:
                    waktu_str = j.get("waktu_fire", "-")
                    countdown_q = "-"
                rows_q.append({
                    "Paket": _tender_labels_q.get(
                        str(j.get("paket_id") or ""),
                        j.get("nama_paket", j["paket_id"]),
                    )[:70],
                    "Jenis": j.get("jenis", "-"),
                    "Waktu": waktu_str,
                    "Countdown": countdown_q,
                })
            st.dataframe(rows_q, use_container_width=True, hide_index=True)

    with _pj_col_kanan:
        from penjelasan_engine import TZ_WIB

        st.markdown("### 3. Jadwalkan Auto-Post")
        st.caption("Engine cari jadwal penjelasan dari Google Calendar lalu auto-post saat waktunya tiba.")

        pj_jenis = st.selectbox(
            "Jenis Penjelasan",
            options=list(penjelasan_config.JENIS_PAKET.keys()),
            format_func=lambda k: penjelasan_config.JENIS_PAKET[k],
            key="pj_jenis",
        )

        with st.expander("✏️ Override teks pembukaan (opsional)"):
            pj_teks_override = st.text_area(
                "Teks custom", value="", height=120,
                placeholder="Kosongkan untuk pakai template bawaan",
                key="pj_teks_override",
            )

        # ── Preview jadwal GCal per paket terpilih ─────────────────────────
        if pj_selected:
            with st.spinner("Baca jadwal dari Google Calendar..."):
                jadwal_gcal = _get_jadwal_penjelasan_gcal_cached()
            now_pj = datetime.now(TZ_WIB)

            for p in pj_selected:
                tgl_mulai = jadwal_gcal.get(p["id_lelang"])
                if tgl_mulai:
                    secs = int((tgl_mulai - now_pj).total_seconds())
                    if secs > 0:
                        h, rem = divmod(secs, 3600)
                        m = rem // 60
                        countdown = f"⏳ {h//24}h {h%24}j {m}m lagi"
                    elif secs > -10800:
                        countdown = "🔴 AKTIF SEKARANG"
                    else:
                        countdown = "✅ Sudah lewat"
                    st.caption(f"**{p['kode']}** — {tgl_mulai.strftime('%d/%m/%Y %H:%M')} WIB | {countdown}")
                else:
                    st.caption(f"**{p['kode']}** — ⚠️ Tidak ditemukan di GCal")

        st.divider()

        # ── Tombol Jadwalkan ───────────────────────────────────────────────
        pj_n = len(pj_selected)
        if st.button(
            f"📅 Jadwalkan {pj_n} Paket ke Antrian",
            key="pj_jadwalkan",
            type="primary",
            disabled=pj_n == 0,
            use_container_width=True,
        ):
            teks_ov = st.session_state.get("pj_teks_override", "").strip() or None
            hasil_jadwal = []
            for p in pj_selected:
                r = penjelasan_engine.jadwalkan_dari_gcal(
                    paket_id=p["id_lelang"],
                    nama_paket=p["nama"],
                    jenis=pj_jenis,
                    teks_override=teks_ov,
                )
                hasil_jadwal.append({
                    "kode": p["kode"],
                    "nama": p["nama"][:50],
                    "status": "✅ Dijadwalkan" if r["ok"] else "❌ Gagal",
                    "waktu": r["waktu_fire"].strftime("%d/%m/%Y %H:%M") if r["waktu_fire"] else "-",
                    "pesan": r["pesan"],
                })
            ok_n = sum(1 for h in hasil_jadwal if h["status"].startswith("✅"))
            if ok_n == pj_n:
                st.success(f"✅ {ok_n} paket berhasil dijadwalkan.")
            else:
                st.warning(f"⚠️ {ok_n}/{pj_n} paket dijadwalkan. Cek paket yang gagal.")
            st.dataframe(hasil_jadwal, use_container_width=True, hide_index=True)

        # ── Log Scheduler ──────────────────────────────────────────────────
        with st.expander("📋 Log Scheduler"):
            log_lines = penjelasan_engine.get_log()
            if log_lines:
                st.code("\n".join(reversed(log_lines[-30:])), language=None)
            else:
                st.caption("Belum ada log.")

        # ── Post Manual (darurat) ──────────────────────────────────────────
        with st.expander("⚡ Post Manual Sekarang (darurat)"):
            st.caption("Post langsung tanpa menunggu scheduler. Gunakan jika scheduler tidak jalan.")
            if st.button(
                f"🚀 Post ke {pj_n} Paket Sekarang",
                key="pj_post_manual",
                disabled=pj_n == 0,
                use_container_width=True,
            ):
                teks_ov = st.session_state.get("pj_teks_override", "").strip() or None
                progress = st.progress(0, text="Memulai...")
                hasil_pj = []
                for i, p in enumerate(pj_selected):
                    progress.progress((i + 1) / len(pj_selected), text=f"Post ke {p['kode']}...")
                    try:
                        result = penjelasan_engine.auto_post_sapaan(p["id_lelang"], pj_jenis, teks_ov)
                        hasil_pj.append({
                            "kode": p["kode"], "nama": p["nama"][:45],
                            "total": result["total"], "sukses": result["sukses"],
                            "gagal": result["gagal"], "pesan": result.get("pesan", ""),
                        })
                    except Exception as e:
                        hasil_pj.append({
                            "kode": p["kode"], "nama": p["nama"][:45],
                            "total": 0, "sukses": 0, "gagal": 1, "pesan": str(e),
                        })
                progress.empty()
                ok_m = sum(1 for h in hasil_pj if h["gagal"] == 0 and h["total"] > 0)
                if ok_m == len(hasil_pj):
                    st.success(f"✅ {ok_m}/{len(hasil_pj)} paket berhasil.")
                else:
                    st.warning(f"⚠️ {ok_m}/{len(hasil_pj)} paket berhasil.")
                st.dataframe(hasil_pj, use_container_width=True, hide_index=True)

# ============================================================
# Tab 8: Auto-Fill Jadwal
# ============================================================

if _tender_active_tab == "2️⃣ Buat Jadwal":

    _libur_map = _LIBUR_MAP

    _jd_col_list, _jd_col_detail = st.columns([3, 2])

    with _jd_col_list:
        st.markdown("### 1. Pilih Paket")
        col_fetch, col_all, col_none = st.columns([3, 1, 1])
        with col_fetch:
            if "global_paket_draft" not in st.session_state:
                st.info("⚠️ Klik **🔄 Sinkronkan Paket** di **Tab 0** dulu.")
            else:
                st.caption(f"📋 {len(_get_paket_gabungan())} paket tersedia (draft + aktif)")

        jd_selected = []
        if "global_paket_draft" in st.session_state or "global_paket_aktif" in st.session_state:
            paket_list = _get_paket_gabungan()
            if not paket_list:
                st.warning("⚠️ Tidak ada paket ditemukan.")
            else:
                with col_all:
                    if st.button("✅ Semua", key="jd_sel_all", use_container_width=True):
                        for p in paket_list:
                            st.session_state[f"jd_chk_{p['id_lelang']}"] = True
                        st.rerun()
                with col_none:
                    if st.button("⬜ Kosong", key="jd_sel_none", use_container_width=True):
                        for p in paket_list:
                            st.session_state[f"jd_chk_{p['id_lelang']}"] = False
                        st.rerun()

                for p in paket_list:
                    key_chk = f"jd_chk_{p['id_lelang']}"
                    checked = st.checkbox(
                        _pokja_label(p),
                        value=st.session_state.get(key_chk, True),
                        key=key_chk,
                    )
                    if checked:
                        jd_selected.append(p)

                st.caption(f"**{len(jd_selected)}** dari **{len(paket_list)}** paket dipilih")
        else:
            st.info("Klik tombol di atas untuk mengambil daftar paket.")

    with _jd_col_detail:
        st.markdown("### 2. Tanggal Mulai")

        jd_beda_jadwal = st.checkbox(
            "Jadwal berbeda per paket",
            value=False,
            key="jd_beda_jadwal",
        )

        if not jd_beda_jadwal:
            col_date, col_time = st.columns(2)
            with col_date:
                jd_tgl_global = st.date_input(
                    "Tanggal",
                    value=datetime.now().date(),
                    format="DD/MM/YYYY",
                    key="jd_tgl_global",
                )
                st.markdown(f"**{_HARI_NAMA[jd_tgl_global.weekday()]}, {jd_tgl_global.day} {_BULAN_NAMA[jd_tgl_global.month-1]} {jd_tgl_global.year}**")
            with col_time:
                jd_jam_global = st.time_input(
                    "Jam",
                    value=datetime.strptime("08:00", "%H:%M").time(),
                    key="jd_jam_global",
                )
            if jd_tgl_global in _libur_map:
                st.warning(f"⚠️ **{_libur_map[jd_tgl_global]}**")
        else:
            if not jd_selected:
                st.info("Pilih paket di sebelah kiri terlebih dahulu.")
            else:
                for p in jd_selected:
                    key_tgl = f"jd_tgl_{p['id_lelang']}"
                    key_jam = f"jd_jam_{p['id_lelang']}"
                    col_nama, col_tgl, col_jam = st.columns([3, 2, 1])
                    with col_nama:
                        st.markdown(f"**{p['kode']}**")
                    with col_tgl:
                        tgl_p = st.date_input(
                            "Tgl",
                            value=st.session_state.get(key_tgl, datetime.now().date()),
                            format="DD/MM/YYYY",
                            key=key_tgl,
                            label_visibility="collapsed",
                        )
                        if tgl_p in _libur_map:
                            st.caption(f"⚠️ {_libur_map[tgl_p]}")
                        else:
                            st.caption(f"{_HARI_NAMA[tgl_p.weekday()]}, {tgl_p.day} {_BULAN_NAMA[tgl_p.month-1]} {tgl_p.year}")
                    with col_jam:
                        st.time_input(
                            "Jam",
                            value=st.session_state.get(key_jam, datetime.strptime("08:00", "%H:%M").time()),
                            key=key_jam,
                            label_visibility="collapsed",
                        )

        with st.expander("ℹ️ Libur Nasional Tersisa"):
            hari_ini = datetime.now().date()
            sisa = sorted(d for d in _libur_map if d >= hari_ini)
            for d in sisa:
                st.write(f"• {_HARI_NAMA[d.weekday()]}, {d.day} {_BULAN_NAMA[d.month-1]} {d.year} — {_libur_map[d]}")

        st.divider()
        st.caption("⚠️ Akan menimpa jadwal yang sudah ada di SPSE.")

        jd_submit = st.button(
            f"🚀 Set Jadwal ke SPSE ({len(jd_selected)} paket)",
            type="primary",
            use_container_width=True,
            disabled=len(jd_selected) == 0,
            key="jd_submit",
        )

        if jd_submit:
            hasil_list = []
            progress = st.progress(0, text="Memulai...")

            for i, p in enumerate(jd_selected):
                progress.progress((i + 1) / len(jd_selected), text=f"Submit {p['kode']} ({i+1}/{len(jd_selected)})...")

                if jd_beda_jadwal:
                    tgl_p = st.session_state.get(f"jd_tgl_{p['id_lelang']}", datetime.now().date())
                    jam_p = st.session_state.get(f"jd_jam_{p['id_lelang']}", datetime.strptime("08:00", "%H:%M").time())
                else:
                    tgl_p = jd_tgl_global
                    jam_p = jd_jam_global

                tgl_mulai = datetime.combine(tgl_p, jam_p)

                try:
                    result = jadwal_engine.auto_fill_jadwal(p["id_lelang"], tgl_mulai)
                    scraped = result["scraped"]
                    payload = result["payload"]

                    if not scraped.get("csrf"):
                        hasil_list.append({"kode": p["kode"], "nama": p["nama"][:50], "sukses": False, "pesan": "CSRF tidak ditemukan", "mulai": ""})
                        continue
                    if not scraped.get("cookie"):
                        hasil_list.append({"kode": p["kode"], "nama": p["nama"][:50], "sukses": False, "pesan": "Cookie tidak ditemukan", "mulai": ""})
                        continue

                    submit_result = jadwal_engine.submit_jadwal(p["id_lelang"], payload, cookie_str=scraped.get("cookie"))
                    hasil_list.append({
                        "kode": p["kode"],
                        "nama": p["nama"][:50],
                        "sukses": submit_result.get("ok", False),
                        "pesan": f"HTTP {submit_result['status']}",
                        "mulai": tgl_mulai.strftime("%d/%m/%Y %H:%M"),
                    })
                except Exception as e:
                    hasil_list.append({"kode": p["kode"], "nama": p["nama"][:50], "sukses": False, "pesan": str(e), "mulai": ""})

            progress.empty()

            sukses_n = sum(1 for h in hasil_list if h["sukses"])
            gagal_n  = len(hasil_list) - sukses_n
            if gagal_n == 0:
                st.success(f"✅ Semua {sukses_n} paket berhasil dijadwalkan!")
            else:
                st.warning(f"⚠️ {sukses_n} berhasil, {gagal_n} gagal.")

            st.dataframe(
                hasil_list,
                use_container_width=True,
                column_config={
                    "kode":   st.column_config.TextColumn("Kode", width="small"),
                    "nama":   st.column_config.TextColumn("Nama Paket", width="large"),
                    "mulai":  st.column_config.TextColumn("Tgl Mulai"),
                    "sukses": st.column_config.CheckboxColumn("Sukses", width="small"),
                    "pesan":  st.column_config.TextColumn("Pesan"),
                },
                hide_index=True,
            )

        st.divider()
        st.markdown("### 3. Sinkronisasi Google Calendar")
        st.caption("Update acara di Google Calendar berdasarkan data jadwal terbaru SPSE.")

        import gcal_pl_helper as _gcal_td
        if not _gcal_td.check_gcal_token():
            st.warning("⚠️ Token Google Calendar tidak valid atau belum login. Klik tombol di bawah untuk login ulang.")
            gcal_login_btn = st.button(
                "🔑 Login Ulang ke Google Calendar",
                type="primary",
                use_container_width=True,
                key="jd_login_gcal"
            )
            gcal_sync_btn = False
        else:
            gcal_sync_btn = st.button(
                "🔄 Sync Jadwal ke GCalendar",
                type="primary",
                use_container_width=True,
                key="jd_sync_gcal"
            )
            gcal_login_btn = False
        
        if gcal_sync_btn:
            if not jd_selected:
                st.warning("Pilih minimal satu paket di daftar sebelah kiri untuk didaftarkan dan disinkronkan.")
            else:
                import subprocess as _sp
                import pathlib as _pathlib
                import os as _os
                import pandas as _pd
                from config import V19_ROOT as _V19_ROOT, POKJA_PYTHON as _POKJA_PYTHON, SPSE_BASE_URL as _SPSE_BASE_URL
                
                _v19_dir = _pathlib.Path(_V19_ROOT)
                _db_path = _v19_dir / "database_tender.csv"
                _py_exe = _pathlib.Path(_POKJA_PYTHON)
                _script = _v19_dir / "sync_jadwal.py"
                _no_win = 0x08000000
                
                _env = _os.environ.copy()
                _env["PYTHONIOENCODING"] = "utf-8"
                
                with st.status("🔄 Menyiapkan data Sinkronisasi...", expanded=True) as sync_status:
                    st.write("Mendaftarkan URL paket ke database V19...")
                    try:
                        if _db_path.exists():
                            df_db = _pd.read_csv(_db_path)
                        else:
                            df_db = _pd.DataFrame(columns=['url', 'members', 'nama_paket', 'last_sync', 'content_hash'])
                            
                        for _col in ['url', 'members', 'nama_paket', 'last_sync', 'content_hash']:
                            if _col not in df_db.columns:
                                df_db[_col] = ''
                                
                        df_db.set_index('url', inplace=True)
                        
                        for p in jd_selected:
                            _url = f"{_SPSE_BASE_URL.rstrip('/')}/lelang/{p['id_lelang']}/jadwal"
                            _members = p.get('pokja') or p.get('kode', 'Pokja')
                            _nama = p.get('nama', f"Paket {p['id_lelang']}")
                            
                            df_db.loc[_url, 'members'] = _members
                            df_db.loc[_url, 'nama_paket'] = _nama
                            if 'content_hash' not in df_db.loc[_url] or _pd.isna(df_db.loc[_url, 'content_hash']):
                                df_db.loc[_url, 'content_hash'] = ''
                            if 'last_sync' not in df_db.loc[_url] or _pd.isna(df_db.loc[_url, 'last_sync']):
                                df_db.loc[_url, 'last_sync'] = ''
                                
                        df_db.reset_index(inplace=True)
                        df_db.to_csv(_db_path, index=False)
                        
                        _list_paket_str = "\n".join([f"{i+1}. {p.get('nama', p['id_lelang'])}" for i, p in enumerate(jd_selected)])
                        st.success(f"✅ Berhasil mendaftarkan {len(jd_selected)} paket ke radar V19:\n\n{_list_paket_str}")
                    except Exception as e:
                        st.error(f"Gagal mengupdate database: {e}")
                
                    st.write("Memanggil script `sync_jadwal.py`...")
                    log_container = st.empty()
                    try:
                        res = _sp.run(
                            [str(_py_exe), str(_script)],
                            cwd=str(_v19_dir),
                            env=_env,
                            capture_output=True,
                            text=True,
                            encoding="utf-8",
                            creationflags=_no_win,
                            timeout=300
                        )
                        
                        if res.stdout:
                            log_container.code(res.stdout, language="text")
                        if res.stderr:
                            if "RefreshError" in res.stderr or "invalid_grant" in res.stderr:
                                st.error("🔐 **Token Google Calendar Kedaluwarsa!**\n\nSilakan klik tombol Login Ulang (refresh halaman dulu agar tombol muncul).")
                            else:
                                st.error(f"Error output:\n{res.stderr}")
                            
                        if res.returncode == 0:
                            sync_status.update(label="✅ Sinkronisasi GCalendar Selesai", state="complete")
                        else:
                            sync_status.update(label="⚠️ Sinkronisasi GCalendar Selesai dengan Error", state="error")
                            
                    except Exception as e:
                        st.error(f"Gagal menjalankan script: {e}")
                        sync_status.update(label="⚠️ Terjadi Kesalahan", state="error")
                    
        if gcal_login_btn:
            import subprocess as _sp
            import pathlib as _pathlib
            from config import V19_ROOT as _V19_ROOT, RUNTIME_ROOT as _RUNTIME_ROOT, POKJA_PYTHON as _POKJA_PYTHON, find_secret as _find_secret
            
            _v19_dir = _pathlib.Path(_V19_ROOT)
            _py_exe = _pathlib.Path(_POKJA_PYTHON)
            _credentials_path = _find_secret("credentials.json")
            
            _auth_code = f"""
import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

from google_auth_oauthlib.flow import InstalledAppFlow

cred_path = r"{_credentials_path}"
token_path = r"{_pathlib.Path(_RUNTIME_ROOT) / 'state' / 'token.json'}"
os.makedirs(os.path.dirname(token_path), exist_ok=True)
scopes = ['https://www.googleapis.com/auth/calendar']

if os.path.exists(token_path):
    os.remove(token_path)

print("Membuka browser... Silakan login di browser Anda.")
try:
    flow = InstalledAppFlow.from_client_secrets_file(cred_path, scopes)
    creds = flow.run_local_server(port=0)
    with open(token_path, 'w') as f:
        f.write(creds.to_json())
    print("Login berhasil! Token tersimpan.")
except Exception as e:
    print(f"Error: {{e}}", file=sys.stderr)
    sys.exit(1)
"""
            with st.status("🔑 Menunggu Otorisasi Browser...", expanded=True) as auth_status:
                st.write("Jendela browser Google Login akan segera terbuka. Silakan login...")
                try:
                    res = _sp.run(
                        [str(_py_exe), "-c", _auth_code], 
                        cwd=str(_v19_dir),
                        capture_output=True, 
                        text=True, 
                        encoding="utf-8",
                        creationflags=0x08000000,
                        timeout=300
                    )
                    
                    if res.returncode == 0:
                        auth_status.update(label="✅ Login Google Calendar berhasil!", state="complete")
                        st.success("Autentikasi selesai. Kamu bisa menggunakan fitur Sync Jadwal sekarang.")
                        st.rerun()
                    else:
                        auth_status.update(label="⚠️ Login Dibatalkan atau Error", state="error")
                        st.error(res.stderr or "Gagal mendapatkan otorisasi.")
                except _sp.TimeoutExpired:
                    auth_status.update(label="⌛ Waktu tunggu login habis", state="error")
                    st.error("Proses login ditutup karena tidak ada aktivitas lebih dari 5 menit.")
                except Exception as e:
                    auth_status.update(label="⚠️ Terjadi Kesalahan", state="error")
                    st.error(str(e))
# Tab 9: Kirim Undangan
# ============================================================

if _tender_active_tab == "1️⃣ Kirim Undangan DPP":
    # ── Layout 2 kolom: kiri = pilih paket, kanan = detail undangan ─────────
    _kp_col_list, _kp_col_detail = st.columns([3, 2])

    with _kp_col_list:
        st.markdown("### 1. Pilih Paket")
        
        kp_selected = []

        if "global_paket_draft" not in st.session_state and "global_paket_aktif" not in st.session_state:
            st.info("⚠️ Data paket belum disinkronkan. Silakan ke **Tab 0** dan klik **🔄 Sinkronkan Paket**.")
        else:
            paket_list = _get_paket_gabungan()
            if not paket_list:
                st.warning("⚠️ Tidak ada paket ditemukan.")
            else:
                st.caption(f"📋 {len(paket_list)} paket tersedia (draft + aktif) — pilih target:")

                _kp_sel_col1, _kp_sel_col2 = st.columns(2)
                with _kp_sel_col1:
                    if st.button("✅ Semua", key="kp_sel_all", use_container_width=True):
                        for p in paket_list:
                            st.session_state[f"kp_chk_{p['id_lelang']}"] = True
                        st.rerun()
                with _kp_sel_col2:
                    if st.button("⬜ Kosong", key="kp_sel_none", use_container_width=True):
                        for p in paket_list:
                            st.session_state[f"kp_chk_{p['id_lelang']}"] = False
                        st.rerun()

                for p in paket_list:
                    key_chk     = f"kp_chk_{p['id_lelang']}"
                    key_tgl_acara = f"kp_tgl_acara_{p['id_lelang']}"

                    col_chk, col_tgl_p = st.columns([3, 2])
                    with col_chk:
                        checked = st.checkbox(
                            _pokja_label(p),
                            value=st.session_state.get(key_chk, True),
                            key=key_chk,
                        )
                    with col_tgl_p:
                        tgl_acara_p = st.date_input(
                            "Tanggal Acara",
                            value=st.session_state.get(key_tgl_acara, datetime.now().date()),
                            format="DD/MM/YYYY",
                            key=key_tgl_acara,
                            label_visibility="collapsed",
                        )
                        st.caption(f"{_HARI_NAMA[tgl_acara_p.weekday()]}, {tgl_acara_p.day} {_BULAN_NAMA[tgl_acara_p.month-1]} {tgl_acara_p.year}")

                    if checked:
                        kp_selected.append({
                            **p,
                            "_tgl_acara": tgl_acara_p,
                        })

                st.caption(f"**{len(kp_selected)}** dari **{len(paket_list)}** paket dipilih")

        # ── Auto Pre-fill dari Excel (jika ada 1 paket terpilih & folder ada) ──
        if len(kp_selected) == 1:
            pkode = kp_selected[0]["kode"]
            if st.button(
                f"📋 Pre-fill Tanggal & Tempat dari Excel ({pkode})",
                key="kp_prefill",
                use_container_width=True,
                help="Baca E26 (tanggal DPP) dan E31 (tempat) dari file Excel paket",
            ):
                with st.spinner("Membaca Excel..."):
                    info = merge_engine.get_draft_info_from_excel(pkode)
                if info["tanggal"]:
                    st.session_state["kp_tgl"] = info["tanggal"]
                if info["tempat"]:
                    st.session_state["kp_tempat"] = info["tempat"]
                if info["pesan"] != "OK":
                    st.warning(f"Pre-fill: {info['pesan']}")
                else:
                    st.success("Tanggal & tempat berhasil diisi dari Excel.")
                st.rerun()

        # ── 2. Detail Undangan ────────────────────────────────────────────────
        st.divider()
        st.markdown("### 2. Detail Undangan")

        _kp_libur_map = _LIBUR_MAP

        st.markdown("**Waktu Acara (berlaku semua paket)**")
        col_mulai, col_selesai = st.columns(2)
        with col_mulai:
            kp_jam_mulai = st.time_input(
                "Mulai",
                value=datetime.strptime("09:00", "%H:%M").time(),
                key="kp_jam_mulai",
                step=1800,
            )
        with col_selesai:
            kp_jam_selesai = st.time_input(
                "Selesai",
                value=datetime.strptime("11:00", "%H:%M").time(),
                key="kp_jam_selesai",
                step=1800,
            )

        with st.expander("ℹ️ Libur Nasional Tersisa"):
            _kp_hari_ini = datetime.now().date()
            _kp_sisa = sorted(d for d in _kp_libur_map if d >= _kp_hari_ini)
            for d in _kp_sisa:
                st.write(f"• {_HARI_NAMA[d.weekday()]}, {d.day} {_BULAN_NAMA[d.month-1]} {d.year} — {_kp_libur_map[d]}")

        kp_tempat = st.text_area(
            "Tempat",
            value=kirimpesan_engine.DEFAULT_TEMPAT,
            key="kp_tempat",
            height=100,
        )

        # Hardcode: Mekanisme = Offline, Dibawa & Hadir pakai default
        kp_is_online = False
        kp_link = ""
        kp_dibawa = kirimpesan_engine.DEFAULT_DIBAWA
        kp_hadir = kirimpesan_engine.DEFAULT_HADIR

        st.divider()
        st.warning("⚠️ Undangan yang sudah terkirim **tidak bisa dihapus** dari sistem SPSE.")

        kirim_disabled = len(kp_selected) == 0

        if not st.session_state.get("kp_konfirmasi"):
            if st.button(
                f"📨 Kirim Undangan ke {len(kp_selected)} Paket",
                key="kp_kirim",
                type="primary",
                disabled=kirim_disabled,
                use_container_width=True,
            ):
                if not kp_tempat.strip():
                    st.error("❌ Tempat wajib diisi.")
                else:
                    st.session_state["kp_konfirmasi"] = True
                    st.rerun()
        else:
            _kp_konfirm_lines = "\n".join(
                f"{i+1}. Pokja {p.get('pokja', p['kode'])} - {p['nama']}  \n"
                f"   📅 {_HARI_NAMA[p['_tgl_acara'].weekday()]}, {p['_tgl_acara'].day} {_BULAN_NAMA[p['_tgl_acara'].month-1]} {p['_tgl_acara'].year}"
                for i, p in enumerate(kp_selected)
            )
            st.warning(
                f"Kirim ke **{len(kp_selected)} paket**\n\n"
                f"{_kp_konfirm_lines}\n\n"
                f"- Pukul: {kp_jam_mulai.strftime('%H.%M')} s.d. {kp_jam_selesai.strftime('%H.%M')} Wita\n"
                f"- Tempat: {kp_tempat.strip()[:80]}\n\n"
                f"**Tidak bisa dibatalkan setelah dikirim.**"
            )

            col_ya, col_batal = st.columns(2)
            with col_ya:
                if st.button("✅ Ya, Kirim", key="kp_ya", type="primary", use_container_width=True):
                    st.session_state["kp_konfirmasi"] = False

                    import undangan_pdf_engine
                    progress = st.progress(0, text="Memulai pengiriman...")
                    hasil_list = []
                    _tgl_kirim = datetime.now().date()

                    for i, paket in enumerate(kp_selected):
                        progress.progress(
                            (i + 1) / len(kp_selected),
                            text=f"Mengirim ke Pokja {paket.get('pokja', paket['kode'])} ({i+1}/{len(kp_selected)})..."
                        )

                        # Generate PDF lampiran otomatis
                        _tgl_acara = paket["_tgl_acara"]
                        _hari_tgl  = f"{_HARI_NAMA[_tgl_acara.weekday()]}, {_tgl_acara.day} {_BULAN_NAMA[_tgl_acara.month-1]} {_tgl_acara.year}"
                        _pukul_str = f"{kp_jam_mulai.strftime('%H.%M')} s.d. {kp_jam_selesai.strftime('%H.%M')} Wita"
                        _kode_pokja = paket.get("pokja", "000")

                        gen_res = undangan_pdf_engine.generate_undangan_pdf(
                            kode_tender=paket["kode"],
                            tanggal_kirim=_tgl_kirim,
                            hari_tgl_rapat=_hari_tgl,
                            pukul_rapat=_pukul_str,
                            tempat_rapat=kp_tempat.strip(),
                            output_path=None,
                        )
                        _lamp_bytes = gen_res["pdf_bytes"] if gen_res["sukses"] else None
                        _lamp_nama  = f"Undangan_{_kode_pokja.zfill(3)}.pdf"

                        waktu_str  = datetime.combine(_tgl_acara, kp_jam_mulai).strftime("%d-%m-%Y %H:%M")
                        sampai_str = datetime.combine(_tgl_acara, kp_jam_selesai).strftime("%d-%m-%Y %H:%M")

                        res = kirimpesan_engine.kirim_undangan(
                            paket_id=paket["id_lelang"],
                            waktu=waktu_str,
                            sampai=sampai_str,
                            tempat=kp_tempat.strip(),
                            dibawa=kp_dibawa.strip(),
                            hadir=kp_hadir.strip(),
                            is_online=False,
                            link_pembuktian="",
                            lampiran_bytes=_lamp_bytes,
                            lampiran_nama=_lamp_nama,
                        )

                        _tgl_save_ok, _tgl_save_msg = False, "-"
                        if res["sukses"]:
                            _tgl_save_ok, _tgl_save_msg = _simpan_tanggal_undangan_dpp(
                                paket["kode"], _tgl_acara, _tgl_kirim
                            )

                        hasil_list.append({
                            "pokja": f"Pokja {_kode_pokja.zfill(3)}",
                            "nama": paket["nama"],
                            "tgl_ba": _tgl_acara.strftime("%d-%m-%Y"),
                            "pdf": "✅" if gen_res["sukses"] else f"❌ {gen_res['pesan']}",
                            "kirim": "✅" if res["sukses"] else f"❌ {res['pesan']}",
                            "supabase": "✅" if _tgl_save_ok else ("-" if not res["sukses"] else f"⚠️ {_tgl_save_msg}"),
                        })

                    progress.empty()

                    sukses_n = sum(1 for h in hasil_list if h["kirim"] == "✅")
                    gagal_n  = len(hasil_list) - sukses_n
                    if gagal_n == 0:
                        st.success(f"✅ Semua {sukses_n} undangan berhasil dikirim!")
                    else:
                        st.warning(f"⚠️ {sukses_n} berhasil, {gagal_n} gagal.")

                    st.dataframe(
                        hasil_list,
                        use_container_width=True,
                        column_config={
                            "pokja": st.column_config.TextColumn("Pokja", width="small"),
                            "nama":  st.column_config.TextColumn("Nama Paket", width="large"),
                            "tgl_ba": st.column_config.TextColumn("Tanggal BA", width="small"),
                            "pdf":   st.column_config.TextColumn("PDF", width="small"),
                            "kirim": st.column_config.TextColumn("Kirim", width="small"),
                            "supabase": st.column_config.TextColumn("Supabase", width="small"),
                        },
                        hide_index=True,
                    )

            with col_batal:
                if st.button("❌ Batal", key="kp_batal", use_container_width=True):
                    st.session_state["kp_konfirmasi"] = False
                    st.rerun()

    with _kp_col_detail:
        st.markdown("### 3. Upload BA Reviu DPP")
        st.caption("Upload BA Hasil Reviu setelah PPK menandatangani.")
        if "global_paket_draft" not in st.session_state:
            st.info("⚠️ Klik **🔄 Sinkronkan Paket** di **Tab 0** dulu.")

        ba_selected = []
        if "global_paket_draft" in st.session_state or "global_paket_aktif" in st.session_state:
            _ba_paket_list = _get_paket_gabungan()
            if not _ba_paket_list:
                st.warning("⚠️ Tidak ada paket ditemukan.")
            else:
                for _p in _ba_paket_list:
                    _key_chk = f"r1_ba_chk_{_p['id_lelang']}"
                    _col_chk, _col_file = st.columns([3, 2])
                    with _col_chk:
                        _checked = st.checkbox(
                            f"**{_p['kode']}** — {_p['nama']}",
                            value=st.session_state.get(_key_chk, True),
                            key=_key_chk,
                        )
                    with _col_file:
                        _ba_up = st.file_uploader(
                            "BA Reviu",
                            type=["pdf"],
                            key=f"r1_ba_file_{_p['id_lelang']}",
                            label_visibility="collapsed",
                        )
                        if _ba_up:
                            st.caption(f"📋 {_ba_up.name}")
                    if _checked:
                        ba_selected.append({**_p, "_ba_file": _ba_up})

        st.divider()
        ba_tgl = st.date_input(
            "Tanggal BA Reviu",
            value=datetime.now().date(),
            key="r1_ba_tgl",
            format="DD/MM/YYYY",
        )
        st.caption(f"{_HARI_NAMA[ba_tgl.weekday()]}, {ba_tgl.day} {_BULAN_NAMA[ba_tgl.month-1]} {ba_tgl.year}")

        _ba_upload_disabled = len(ba_selected) == 0 or all(p.get("_ba_file") is None for p in ba_selected)
        _ba_n_file = len([p for p in ba_selected if p.get("_ba_file")])
        if st.button(
            f"📤 Upload BA Reviu ({_ba_n_file} file)",
            key="ba_upload",
            type="primary",
            disabled=_ba_upload_disabled,
            use_container_width=True,
        ):
            _ba_progress = st.progress(0, text="Memulai upload...")
            _ba_hasil = []
            _ba_valid = [p for p in ba_selected if p.get("_ba_file")]
            for _i, _p in enumerate(_ba_valid):
                _ba_progress.progress(
                    (_i + 1) / len(_ba_valid),
                    text=f"Upload {_p['kode']} ({_i+1}/{len(_ba_valid)})..."
                )
                _res = bareviu_engine.upload_ba_reviu(
                    paket_id=_p["id_lelang"],
                    file_bytes=_p["_ba_file"].getvalue(),
                    file_name=_p["_ba_file"].name,
                    tgl_dok_ba=ba_tgl.strftime("%d-%m-%Y"),
                )
                _ba_hasil.append({
                    "kode": _p["kode"],
                    "nama": _p["nama"][:50],
                    "sukses": _res["sukses"],
                    "pesan": _res["pesan"],
                })
            _ba_progress.empty()

            _ba_ok = sum(1 for h in _ba_hasil if h["sukses"])
            _ba_fail = len(_ba_hasil) - _ba_ok
            if _ba_fail == 0:
                st.success(f"✅ {_ba_ok} BA Reviu berhasil diupload!")
            else:
                st.warning(f"⚠️ {_ba_ok} berhasil, {_ba_fail} gagal.")

            st.dataframe(
                _ba_hasil,
                use_container_width=True,
                column_config={
                    "kode":   st.column_config.TextColumn("Kode", width="small"),
                    "nama":   st.column_config.TextColumn("Nama Paket", width="large"),
                    "sukses": st.column_config.CheckboxColumn("Sukses", width="small"),
                    "pesan":  st.column_config.TextColumn("Pesan"),
                },
                hide_index=True,
            )

# Tab 8: Upload & Cetak 5 BA

# ============================================================

if _tender_active_tab == "8️⃣ Upload & Cetak 5 BA":

    ba_selected = []

    # ── Pilih Paket ──────────────────────────────────────────────────────
    st.markdown("### Pilih Paket")
    if "global_paket_draft" not in st.session_state and "global_paket_aktif" not in st.session_state:
        st.info("⚠️ Data paket belum disinkronkan. Silakan ke **Tab 0** dan klik **🔄 Sinkronkan Paket**.")
    else:
        paket_list_ba = _get_paket_gabungan()
        if not paket_list_ba:
            st.warning("⚠️ Tidak ada paket ditemukan.")
        else:
            st.caption(f"📋 {len(paket_list_ba)} paket tersedia (draft + aktif) — pilih:")
            _ba_sel_c1, _ba_sel_c2 = st.columns(2)
            with _ba_sel_c1:
                if st.button("✅ Semua", key="ba_sel_all", use_container_width=True):
                    for p in paket_list_ba:
                        st.session_state[f"ba_chk_{p['id_lelang']}"] = True
                    st.rerun()
            with _ba_sel_c2:
                if st.button("⬜ Kosong", key="ba_sel_none", use_container_width=True):
                    for p in paket_list_ba:
                        st.session_state[f"ba_chk_{p['id_lelang']}"] = False
                    st.rerun()
            for p in paket_list_ba:
                key_chk = f'ba_chk_{p["id_lelang"]}'
                _chk_col, _super_col = st.columns([3, 1])
                with _chk_col:
                    checked = st.checkbox(
                        _pokja_label(p),
                        value=st.session_state.get(key_chk, True), key=key_chk,
                    )
                with _super_col:
                    if st.button('🚀', key=f'btn_super_{p["id_lelang"]}', use_container_width=True, help='Cetak & Upload SEMUA BA untuk paket ini'):
                        st.session_state["ba_pending_target"] = "SEMUA"
                        st.session_state["ba_pending_paket"] = [p]
                        st.rerun()
                if checked:
                    ba_selected.append(p)
            # Source of truth Tender: baca G2 workbook sebelum nomor BA dibuat.
            _ba_kode_mismatch = []
            for _p_ba in ba_selected:
                # Paket hasil sinkronisasi sudah membawa folder_dibuat. Hindari
                # query Supabase per baris; fallback lookup tetap tersedia untuk
                # row legacy yang belum memiliki metadata folder.
                # _get_paket_gabungan selalu menyertakan key folder_dibuat,
                # termasuk saat nilainya kosong dari cache lama. Yang boleh
                # melewati lookup Supabase hanya metadata folder yang benar-benar
                # berisi; jika kosong, resolver harus mencari folder dari DB.
                _has_folder_meta = bool(str(_p_ba.get("folder_dibuat") or "").strip())
                _kode_xl, _kode_db, _kode_src = _enrich_kode_unik_tender_excel(
                    _p_ba,
                    folder_dibuat=_p_ba.get("folder_dibuat"),
                    skip_folder_lookup=_has_folder_meta,
                )
                if _kode_xl:
                    _p_ba["kode_unik"] = _kode_xl
                if _kode_src == "beda":
                    _ba_kode_mismatch.append(
                        f"{_p_ba.get('kode')}: Excel `{_kode_xl}` ≠ DB `{_kode_db}`"
                    )
                elif _kode_src not in ("excel", "beda") and _kode_db:
                    _ba_kode_mismatch.append(
                        f"{_p_ba.get('kode')}: G2 {_kode_src}; memakai fallback DB `{_kode_db}`"
                    )
            if _ba_kode_mismatch:
                st.warning(
                    "⚠️ Ada status Kode Unik yang perlu diperhatikan:\n\n"
                    + "\n".join(f"- {x}" for x in _ba_kode_mismatch)
                )
            st.caption(f"**{len(ba_selected)}** dari **{len(paket_list_ba)}** paket dipilih")
            st.divider()
            if st.button(
                f"🚀 Cetak & Upload SEMUA BA — {len(ba_selected)} Paket",
                key="ba_super_all",
                disabled=len(ba_selected) == 0,
                type="primary",
                use_container_width=True,
            ):
                st.session_state["ba_pending_target"] = "SEMUA"
                st.session_state["ba_pending_paket"] = ba_selected[:]
                st.rerun()

    # ── Inisialisasi session state BA ─────────────────────────────────────
    for jenis_key in ba_config.JENIS_KEYS:
        if f"ba_tgl_{jenis_key}" not in st.session_state:
            st.session_state[f"ba_tgl_{jenis_key}"] = datetime.today().strftime("%d-%m-%Y")
        if f"ba_info_{jenis_key}" not in st.session_state:
            st.session_state[f"ba_info_{jenis_key}"] = ba_config.DEFAULT_INFO.get(jenis_key, "")

    # ── Auto-generate nomor BA + tanggal GCal untuk semua paket di ba_selected ──
    # Key per-paket: ba_no_{jenis}_{id}, ba_tgl_{jenis}_{id}, ba_tgl_date_{jenis}_{id}
    # Key global (paket pertama): ba_no_{jenis}, ba_tgl_{jenis} — untuk kompatibilitas
    _ba_sel_ids = tuple(p["id_lelang"] for p in ba_selected)
    if _ba_sel_ids and st.session_state.get("_ba_last_sel_ids") != _ba_sel_ids:
        try:
            import gcal_helper as _gcal
        except Exception:
            _gcal = None
        for _px in ba_selected:
            _pid_x = _px["id_lelang"]
            _ku = _px.get("kode_unik") or ""
            _kp = _px.get("kode_pokja") or ""
            if _ku and _kp:
                _nomor_dokpil = f"000.3.3/01/T/{_ku}/POKJA{_kp}/UKPBJ/2026"
                for jenis_key in ba_config.JENIS_KEYS:
                    _urut = ba_config.NOMOR_URUT[jenis_key]
                    _no = ba_engine.derive_nomor_ba(_nomor_dokpil, _urut)
                    st.session_state[f"ba_no_{jenis_key}_{_pid_x}"] = _no
            if _gcal:
                try:
                    _tgl_map = _gcal.get_tanggal_ba_dari_gcal(_px["nama"])
                    for _jk, _d in _tgl_map.items():
                        if _d is not None:
                            st.session_state[f"ba_tgl_date_{_jk}_{_pid_x}"] = _d
                            st.session_state[f"ba_tgl_{_jk}_{_pid_x}"] = _d.strftime("%d-%m-%Y")
                except Exception:
                    pass
        st.session_state["_ba_last_sel_ids"] = _ba_sel_ids

    # Guard terakhir sebelum warning/cetak. Pada rerun Streamlit, object paket
    # dapat berasal dari cache lama; baca ulang Excel sekali lagi agar warning
    # tidak didasarkan pada state DB yang stale.
    _ba_kode_missing = []
    for _p_guard in ba_selected:
        if not str(_p_guard.get("kode_unik") or "").strip():
            _kode_guard, _, _kode_guard_status = _enrich_kode_unik_tender_excel(
                _p_guard,
                folder_dibuat=_p_guard.get("folder_dibuat"),
                skip_folder_lookup=bool(str(_p_guard.get("folder_dibuat") or "").strip()),
            )
            if _kode_guard:
                _p_guard["kode_unik"] = _kode_guard
            else:
                _ba_kode_missing.append((_p_guard, _kode_guard_status))
    if _ba_kode_missing:
        _missing_labels = ", ".join(
            f"{p.get('kode') or '-'} ({status})" for p, status in _ba_kode_missing
        )
        st.warning(
            "⚠️ Paket ini belum punya Kode Unik — generate dulu via Excel. "
            f"Paket yang belum terbaca: {_missing_labels}"
        )

    # ── Konfirmasi sebelum cetak ──────────────────────────────────────────
    if st.session_state.get("ba_pending_target"):
        st.divider()
        _pending_target = st.session_state["ba_pending_target"]
        _pending_paket  = st.session_state.get("ba_pending_paket") or ba_selected
        _JENIS_AUTO = [k for k in ba_config.JENIS_KEYS if k != "lainnya"]
        _jenis_konfirm = _JENIS_AUTO if _pending_target == "SEMUA" else [_pending_target]
        _label_target = "Semua BA" if _pending_target == "SEMUA" else ba_config.JENIS_BA[_pending_target]

        _total_kosong = []
        st.markdown(f"### 📋 Konfirmasi Cetak & Upload — {_label_target}")
        st.caption("Tanggal sudah terisi dari GCal. Ubah langsung jika perlu, lalu **Ya, Cetak & Upload**.")

        for _pp in _pending_paket:
            _pid_k = _pp["id_lelang"]
            with st.expander(f"📁 {_pokja_label(_pp)}", expanded=True):
                for _jk in _jenis_konfirm:
                    _no   = st.session_state.get(f"ba_no_{_jk}_{_pid_k}", "")
                    _label = ba_config.JENIS_LABEL[_jk]
                    if not _no:
                        st.caption(f"⚠️ {_label}: nomor BA tidak ada — dilewati")
                        continue
                    _tgl_date_cur = st.session_state.get(f"ba_tgl_date_{_jk}_{_pid_k}")
                    _default_date = _tgl_date_cur if isinstance(_tgl_date_cur, date) else date.today()
                    _edit_key = f"ba_edit_date_{_jk}_{_pid_k}"

                    _col_label, _col_tgl = st.columns([3, 3])
                    with _col_label:
                        st.markdown(f"**{_label}**")
                        st.caption(f"`{_no}`")
                    with _col_tgl:
                        # Label di atas date_input = teks Indonesia (tidak rerun)
                        if isinstance(_default_date, date):
                            _tgl_id = f"{_HARI_NAMA[_default_date.weekday()]}, {_default_date.day} {_BULAN_NAMA[_default_date.month-1]} {_default_date.year}"
                            st.caption(f"📅 {_tgl_id}")
                        _new_date = st.date_input(
                            "Ganti tanggal",
                            value=_default_date,
                            key=_edit_key,
                            label_visibility="visible",
                            format="DD/MM/YYYY",
                        )
                    # Update session state setiap render (reactive, no rerun needed)
                    if isinstance(_new_date, date):
                        st.session_state[f"ba_tgl_date_{_jk}_{_pid_k}"] = _new_date
                        st.session_state[f"ba_tgl_{_jk}_{_pid_k}"] = _new_date.strftime("%d-%m-%Y")
                    else:
                        _total_kosong.append(f"{_pokja_label(_pp)} / {_label}")

        if _total_kosong:
            st.warning(f"⚠️ {len(_total_kosong)} jenis akan dilewati (tanggal kosong)")

        _konfirm_c1, _konfirm_c2 = st.columns(2)
        with _konfirm_c1:
            if st.button("✅ Ya, Cetak & Upload", key="ba_konfirm_ya", type="primary", use_container_width=True):
                st.session_state["ba_auto_target"] = st.session_state.pop("ba_pending_target")
                st.session_state["ba_super_paket"]  = st.session_state.pop("ba_pending_paket", None)
                st.rerun()
        with _konfirm_c2:
            if st.button("❌ Batal", key="ba_konfirm_batal", use_container_width=True):
                st.session_state.pop("ba_pending_target", None)
                st.session_state.pop("ba_pending_paket", None)
                st.rerun()

    # ── BA Lainnya (per paket) ────────────────────────────────────────────
    if ba_selected:
        st.divider()
        st.markdown("#### 📁 BA Lainnya")
        st.caption("Upload scan manual — spesifik per paket.")
        for _pl in ba_selected:
            _pid = _pl["id_lelang"]
            with st.expander(f"📁 {_pokja_label(_pl)}", expanded=False):
                _file_key = f"ba_file_lainnya_{_pid}"
                st.file_uploader("File PDF", type=["pdf"], key=_file_key)
                _file_l = st.session_state.get(_file_key)
                if st.button(
                    "🚀 Upload BA Lainnya",
                    key=f"ba_lainnya_upload_{_pid}",
                    disabled=not _file_l,
                    use_container_width=True,
                ):
                    _tgl_str_l = date.today().strftime("%d-%m-%Y")
                    try:
                        _r = ba_engine.upload_ba(
                            paket_id=_pid, jenis_key="lainnya",
                            nomor_ba="", tanggal_ba=_tgl_str_l,
                            file_bytes=_file_l.getvalue(), file_name=_file_l.name,
                            info="",
                        )
                        if _r["ok"]:
                            st.success(f"✅ BA Lainnya berhasil di-upload.")
                        else:
                            st.error(f"❌ Upload gagal: {_r.get('status')}")
                    except Exception as _e:
                        st.error(f"❌ {_e}")

    # ── Proses Cetak & Auto-Upload ───────────────────────────────────────
    _FILE_LABEL_BA = {
        "penjelasan":      "2. Berita Acara Pemberian Penjelasan",
        "evaluasi":        "4. Berita Acara Evaluasi Penawaran",
        "hasil_pemilihan": "10. Berita Acara Hasil Pemilihan",
        "negosiasi":       "8. Berita Acara Negosiasi",
    }
    if st.session_state.get("ba_auto_target"):
        import os as _os
        import re as _re
        from config import POKJA_ROOT as _POKJA_ROOT, TENDER_ROOT as _TENDER_ROOT
        from config import sb as _sb_ba
        jenis_target = st.session_state["ba_auto_target"]
        target_paket = st.session_state.pop("ba_super_paket", None) or ba_selected
        _JENIS_AUTO = [k for k in ba_config.JENIS_KEYS if k != "lainnya"]
        jenis_list = _JENIS_AUTO if jenis_target == "SEMUA" else [jenis_target]
        label_target = "Semua BA" if jenis_target == "SEMUA" else ba_config.JENIS_BA[jenis_target]
        progress = st.progress(0, text=f"Memulai Cetak & Upload {label_target}...")
        hasil_auto = []
        total_ops = len(target_paket) * len(jenis_list)
        op_idx = 0
        for i, p in enumerate(target_paket):
            pid = p["id_lelang"]
            paket_hasil = {"kode": p["kode"], "nama": p["nama"][:50], "ba": []}
            # Resolve folder paket dari Supabase → subfolder "BA + Summary"
            try:
                _sb_row = _sb_ba().table("draft_paket").select("folder_dibuat").eq("kode_tender", p["kode"]).maybe_single().execute()
                _folder_dibuat = (_sb_row.data or {}).get("folder_dibuat", "")
                if _folder_dibuat:
                    _folder_safe = _re.sub(r'[/\\:*?"<>|]', "-", _folder_dibuat).strip()
                    target_dir = _os.path.join(_TENDER_ROOT, _folder_safe, "BA + Summary")
                else:
                    target_dir = _os.path.join(_TENDER_ROOT, "Asisten_Pokja_Downloads", f"Cetak_BA_{p['kode']}")
            except Exception:
                target_dir = _os.path.join(_TENDER_ROOT, "Asisten_Pokja_Downloads", f"Cetak_BA_{p['kode']}")
            _os.makedirs(target_dir, exist_ok=True)
            import time as _time
            for jenis_key in jenis_list:
                op_idx += 1
                progress.progress(op_idx / total_ops, text=f"Proses {p['kode']} — {ba_config.JENIS_BA[jenis_key]} ({op_idx}/{total_ops})...")
                nomor = st.session_state.get(f"ba_no_{jenis_key}_{pid}", "").strip()
                tanggal = st.session_state.get(f"ba_tgl_{jenis_key}_{pid}", "").strip()
                info = ba_config.DEFAULT_INFO.get(jenis_key, "")
                ba_result = {"jenis": ba_config.JENIS_BA[jenis_key], "status": "⏭️ Lewati (nomor/tanggal kosong)"}
                if nomor and tanggal:
                    # Retry hingga 3x jika SPSE 503/timeout
                    for _attempt in range(3):
                        try:
                            r_cetak = ba_engine.cetak_ba(paket_id=pid, jenis_key=jenis_key, nomor_ba=nomor, tanggal_ba=tanggal, info=info)
                            if r_cetak["ok"]:
                                fn = f"{_FILE_LABEL_BA.get(jenis_key, jenis_key)}-{p['kode']}.pdf"
                                with open(_os.path.join(target_dir, fn), "wb") as f:
                                    f.write(r_cetak["pdf_bytes"])
                                r_up = ba_engine.upload_ba(paket_id=pid, jenis_key=jenis_key, nomor_ba=nomor, tanggal_ba=tanggal, file_bytes=r_cetak["pdf_bytes"], file_name=fn, info=info)
                                if r_up["ok"]:
                                    ba_result["status"] = f"✅ Sukses — `{fn}`"
                                elif r_up["status"] in (503, 429, 502) and _attempt < 2:
                                    progress.progress(op_idx / total_ops, text=f"⏳ Server {r_up['status']}, retry {_attempt+2}/3...")
                                    _time.sleep(3)
                                    continue
                                else:
                                    ba_result["status"] = f"❌ Upload Error {r_up['status']}"
                            else:
                                if r_cetak["status"] in (503, 429, 502) and _attempt < 2:
                                    progress.progress(op_idx / total_ops, text=f"⏳ Cetak {r_cetak['status']}, retry {_attempt+2}/3...")
                                    _time.sleep(3)
                                    continue
                                ba_result["status"] = f"❌ Cetak Error {r_cetak['status']}: {r_cetak.get('error')}"
                            break  # sukses atau error non-retryable
                        except Exception as e:
                            if _attempt < 2:
                                _time.sleep(3)
                                continue
                            ba_result["status"] = f"❌ {e}"
                    # Jeda antar BA biar SPSE ga throttle
                    _time.sleep(2)
                paket_hasil["ba"].append(ba_result)
            hasil_auto.append(paket_hasil)
        progress.empty()
        st.success(f"✅ Selesai! {label_target} telah dikirim ke SPSE dan backup PDF disimpan ke folder paket.")
        del st.session_state["ba_auto_target"]
        for h in hasil_auto:
            st.markdown(f"**{h['kode']}** — {h['nama']}")
            for b in h["ba"]:
                st.caption(f"{b['status']} — {b['jenis']}")
            st.divider()

# ============================================================
# Tab 5: Undangan Pembuktian Kualifikasi Tender
# ============================================================

if _tender_active_tab == "5️⃣ Undangan Pembuktian Kualifikasi":
    import undangan_pembuktian_tender as _upt

    st.markdown("### 5️⃣ Undangan Pembuktian Kualifikasi")
    st.caption(
        "Pilih peserta dari hasil evaluasi SPSE. Undangan hanya dikirim setelah "
        "Administrasi, Kualifikasi, Teknis, dan Harga berstatus LULUS."
    )

    # Sumber daftar mengikuti Tab 6: baca metadata lokal/DB draft_paket,
    # bukan global session cache yang bisa tertinggal setelah Sinkronkan Paket.
    _upt_all_rows = None
    try:
        from config import sb as _sb_upt
        _upt_db_rows = _sb_upt().table("draft_paket").select(
            "kode_tender,nama_tender,kode_pokja,nomor_urut,folder_dibuat"
        ).order("nomor_urut").execute().data or []
        _upt_all_rows = [
            {
                "kode": r.get("kode_tender"),
                "nama": r.get("nama_tender") or r.get("kode_tender"),
                "id_lelang": r.get("kode_tender"),
                "pokja": r.get("kode_pokja") or "",
                "nomor_urut": r.get("nomor_urut") or "",
                "folder_dibuat": r.get("folder_dibuat") or "",
                "status": "aktif",
                "tanggal": "",
            }
            for r in _upt_db_rows
            if r.get("kode_tender")
        ]
    except Exception:
        _upt_all_rows = _get_paket_gabungan()
    _upt_all_rows = [p for p in (_upt_all_rows or []) if p.get("kode") and p.get("kode") != "00000000000"]
    _upt_all_rows = [p for p in _upt_all_rows if not _is_tender_excluded(p)]
    _upt_all_rows = sorted(_upt_all_rows, key=lambda p: str(p.get("tanggal", "")), reverse=True)

    # Tab 5 hanya menampilkan paket yang sudah memiliki event tahap evaluasi
    # di GCal. Query event dilakukan sekali per sesi, bukan satu request per paket.
    _upt_source_key = tuple(str(p.get("kode")) for p in _upt_all_rows)
    if (
        st.session_state.get("upt_gcal_source_key") != _upt_source_key
        or st.session_state.get("upt_gcal_filter_version") != 3
    ):
        with st.spinner("Membaca jadwal evaluasi dari Google Calendar..."):
            st.session_state["upt_gcal_result"] = _upt.fetch_paket_evaluasi_gcal(_upt_all_rows)
            st.session_state["upt_gcal_rows"] = st.session_state["upt_gcal_result"].get("paket", [])
            st.session_state["upt_gcal_source_key"] = _upt_source_key
            st.session_state["upt_gcal_filter_version"] = 3
    _upt_gcal_result = st.session_state.get("upt_gcal_result", {})
    _upt_rows = st.session_state.get("upt_gcal_rows", [])

    # Auto-check semua paket yang lolos filter GCal saat sumber berubah.
    # Tombol Kosongkan tetap memungkinkan user membatalkan pilihan manual.
    _upt_selection_key = tuple(str(p.get("kode")) for p in _upt_rows)
    if st.session_state.get("upt_selected_source_key") != _upt_selection_key:
        st.session_state["upt_selected_packages"] = {str(p["kode"]) for p in _upt_rows}
        st.session_state["upt_selected_source_key"] = _upt_selection_key
    elif "upt_selected_packages" not in st.session_state:
        st.session_state["upt_selected_packages"] = {str(p["kode"]) for p in _upt_rows}

    with st.expander("1. Pilih Paket", expanded=True):
        _urg1, _urg2 = st.columns([1, 3])
        with _urg1:
            if st.button("🔄 Muat ulang GCal", key="upt_reload_gcal", use_container_width=True):
                with st.spinner("Membaca ulang jadwal evaluasi GCal..."):
                    st.session_state["upt_gcal_result"] = _upt.fetch_paket_evaluasi_gcal(_upt_all_rows)
                    st.session_state["upt_gcal_rows"] = st.session_state["upt_gcal_result"].get("paket", [])
                    st.session_state["upt_gcal_filter_version"] = 3
                st.rerun()
        with _urg2:
            if _upt_gcal_result.get("ok"):
                st.caption(f"📅 {_upt_gcal_result.get('pesan', '')}")
            else:
                st.warning(f"GCal tidak dapat dibaca: {_upt_gcal_result.get('pesan', 'error')}")
        if not _upt_rows:
            st.info("Belum ada paket yang memiliki event tahap evaluasi di Google Calendar.")
        else:
            st.caption(f"📋 {len(_upt_rows)} paket — centang satu atau lebih:")
            _u1, _u2 = st.columns(2)
            with _u1:
                if st.button("✅ Pilih semua paket", key="upt_all_pkg", use_container_width=True):
                    st.session_state["upt_selected_packages"] = {str(p["kode"]) for p in _upt_rows}
                    st.rerun()
            with _u2:
                if st.button("⬜ Kosongkan", key="upt_none_pkg", use_container_width=True):
                    st.session_state["upt_selected_packages"] = set()
                    st.rerun()
            for _up in _upt_rows:
                _uk = str(_up["kode"])
                _uc = st.checkbox(
                    f"{_pokja_label(_up)[:70]}  \n_📅 Evaluasi: {_up.get('tgl_evaluasi_gcal') or '-'} · Pembuktian: {_up.get('tgl_pembuktian_gcal') or '-'}_",
                    value=_uk in st.session_state["upt_selected_packages"],
                    key=f"upt_pkg_{_uk}",
                )
                if _uc:
                    st.session_state["upt_selected_packages"].add(_uk)
                else:
                    st.session_state["upt_selected_packages"].discard(_uk)

    _upt_selected = [_p for _p in _upt_rows if str(_p["kode"]) in st.session_state["upt_selected_packages"]]
    # Seperti Tab 6, status peserta dimuat otomatis untuk paket yang terpilih.
    # Tetap tersedia tombol manual untuk refresh setelah perubahan di SPSE.
    _upt_status_missing = [
        _p for _p in _upt_selected
        if f"upt_status_{_p['kode']}" not in st.session_state
        or st.session_state.get("upt_status_filter_version") != 2
    ]
    if _upt_status_missing:
        with st.spinner(f"Memuat peserta {len(_upt_status_missing)} paket dari SPSE..."):
            for _up in _upt_status_missing:
                st.session_state[f"upt_status_{_up['kode']}"] = _upt.fetch_evaluasi_paket(str(_up["kode"]))
            st.session_state["upt_status_filter_version"] = 2
    _u3, _u4 = st.columns([1, 3])
    with _u3:
        if st.button(
            f"🔄 Muat status SPSE ({len(_upt_selected)})",
            key="upt_load_status",
            type="primary",
            use_container_width=True,
            disabled=not _upt_selected,
        ):
            with st.spinner("Membaca status evaluasi dari SPSE..."):
                for _up in _upt_selected:
                    st.session_state[f"upt_status_{_up['kode']}"] = _upt.fetch_evaluasi_paket(str(_up["kode"]))
                st.session_state["upt_status_filter_version"] = 2
            st.rerun()
    with _u4:
        st.caption("Status tidak dimuat massal. Pilih paket lalu klik Muat status SPSE.")

    _upt_targets = []
    if _upt_selected:
        st.markdown("#### 2. Peserta & Status Evaluasi")
        st.caption("Paket terpilih otomatis masuk target. Atur jadwal undangan di sini.")
        _ud1, _ud2 = st.columns(2)
        with _ud1:
            _upt_tgl = st.date_input("Tanggal mulai", value=datetime.now().date(), format="DD/MM/YYYY", key="upt_tgl_mulai")
            _upt_jam_mulai = st.time_input("Jam mulai", value=datetime.strptime("09:00", "%H:%M").time(), key="upt_jam_mulai")
        with _ud2:
            _upt_tgl_selesai = st.date_input("Tanggal selesai", value=datetime.now().date(), format="DD/MM/YYYY", key="upt_tgl_selesai")
            _upt_jam_selesai = st.time_input("Jam selesai", value=datetime.strptime("15:00", "%H:%M").time(), key="upt_jam_selesai")
        for _up in _upt_selected:
            _uk = str(_up["kode"])
            _ur = st.session_state.get(f"upt_status_{_uk}")
            if not _ur:
                st.info(f"{_pokja_label(_up)} — klik Muat status SPSE.")
                continue
            if not _ur.get("ok"):
                st.error(f"{_pokja_label(_up)}: {_ur.get('pesan')}")
                continue
            with st.expander(f"{_pokja_label(_up)} — {_ur.get('pesan')}", expanded=True):
                for _ui, _ps in enumerate(_ur.get("peserta", []), 1):
                    _st = _ps.get("status", {})
                    _eligible = all(_st.get(x) == "lulus" for x in ("A", "K", "T", "H"))
                    _status_text = " | ".join(f"{x}: {_st.get(x, 'belum').upper()}" for x in ("A", "K", "T", "H"))
                    _pc1, _pc2 = st.columns([5, 1])
                    with _pc1:
                        _pick = st.checkbox(
                            f"{_ui}. {_ps.get('nama')} — {_status_text}",
                            value=True,
                            key=f"upt_ps_v2_{_uk}_{_ps.get('peserta_id')}",
                        )
                    with _pc2:
                        if st.button(
                            "✅ Sudah LULUS" if _eligible else "▶ Evaluasi",
                            key=f"upt_eval_one_{_uk}_{_ps.get('peserta_id')}",
                            use_container_width=True,
                            disabled=_eligible,
                            help="Evaluasi Administrasi → Kualifikasi → Teknis → Harga melalui endpoint SPSE.",
                        ):
                            st.session_state["upt_eval_pending"] = {"paket": _up, "peserta": _ps}
                            st.rerun()
                    if _pick:
                        _upt_targets.append({"paket": _up, "peserta": _ps})

    _upt_pending = st.session_state.get("upt_eval_pending")
    if _upt_pending:
        _pp = _upt_pending["paket"]
        _pps = _upt_pending["peserta"]
        st.warning(
            f"Konfirmasi evaluasi **{_pps.get('nama')}** pada {_pokja_label(_pp)} "
            "untuk tahap Administrasi → Kualifikasi → Teknis → Harga."
        )
        _py, _pn = st.columns(2)
        with _py:
            if st.button("✅ Ya, evaluasi peserta ini", key="upt_eval_one_yes", type="primary", use_container_width=True):
                _logs_one = [f"=== {_pp.get('kode')} | {_pps.get('nama')} ==="]
                _log_one_area = st.empty()
                _log_one_area.code("\n".join(_logs_one))
                _ev_one = _upt.evaluasi_lulus_otomatis(
                    str(_pp["kode"]), str(_pps["peserta_id"]),
                    progress_cb=lambda msg: (_logs_one.append(msg), _log_one_area.code("\n".join(_logs_one))),
                )
                _logs_one.append(("✅ " if _ev_one.get("ok") else "❌ ") + _ev_one.get("pesan", ""))
                st.session_state[f"upt_status_{_pp['kode']}"] = _upt.fetch_evaluasi_paket(str(_pp["kode"]))
                st.session_state["upt_eval_pending"] = None
                st.session_state["upt_eval_one_log"] = _logs_one
                st.rerun()
        with _pn:
            if st.button("Batal", key="upt_eval_one_no", use_container_width=True):
                st.session_state["upt_eval_pending"] = None
                st.rerun()

    if st.session_state.get("upt_eval_one_log"):
        with st.expander("📋 Log evaluasi peserta terakhir", expanded=True):
            st.code("\n".join(st.session_state["upt_eval_one_log"]))

    if _upt_targets:
        _upt_tempat = (
            "Kantor UKPBJ Kabupaten Tapin, Jl. Datu Suban RT. 01, Kelurahan Rangda Malingkung, "
            "Kecamatan Tapin Utara, Rantau, Kabupaten Tapin. Kode Pos : 71111"
        )
        _upt_dibawa = (
            "Dokumen (asli dan copy) sebagaimana disampaikan dalam penawaran dan pada isian "
            "kualifikasi serta dokumen/data dukungnya"
        )
        _upt_hadir = (
            "1. Direksi yang namanya ada dalam akta pendirian/perubahan atau pihak yang sah "
            "menurut akta pendirian/perubahan; 2. Penerima kuasa dari direksi yang nama penerima "
            "kuasanya tercantum dalam akta pendirian/perubahan; 3. Pihak lain yang bukan direksi "
            "dapat menghadiri pembuktian kualifikasi selama berstatus sebagai tenaga kerja tetap "
            "(yang dibuktikan dengan bukti lapor/potong pajak PPh Pasal 21 Form 1721 atau Form "
            "1721-A1) dan memperoleh kuasa dari Direksi yang namanya ada dalam akta "
            "pendirian/perubahan atau pihak yang sah menurut akta pendirian/perusahaan; 4. Kepala "
            "Cabang perusahaan yang diangkat oleh kantor pusat yang dibuktikan dengan dokumen otentik."
        )
        # Evaluasi sekarang dilakukan lewat tombol per peserta; pengiriman undangan
        # tidak lagi menjalankan evaluasi otomatis sebagai efek samping.
        _upt_auto = False
        st.warning("Undangan yang sudah terkirim tidak dapat dihapus dari SPSE.")
        if st.button(
            f"📨 Siapkan {len(_upt_targets)} undangan",
            key="upt_prepare_send",
            type="primary",
            use_container_width=True,
        ):
            st.session_state["upt_confirm_send"] = True
            st.rerun()

        if st.session_state.get("upt_confirm_send"):
            st.error(
                f"Konfirmasi: {len(_upt_targets)} peserta akan diproses. "
                "Jika opsi otomatis aktif, checklist evaluasi dapat ditulis ke SPSE terlebih dahulu."
            )
            _yc, _nc = st.columns(2)
            with _yc:
                if st.button("✅ Ya, proses & kirim", key="upt_confirm_yes", type="primary", use_container_width=True):
                    st.session_state["upt_confirm_send"] = False
                    _logs = []
                    _log_area = st.empty()
                    def _upt_log(msg):
                        _logs.append(msg)
                        _log_area.code("\n".join(_logs))
                    for _ti in _upt_targets:
                        _tp = _ti["paket"]
                        _ps = _ti["peserta"]
                        _uk = str(_tp["kode"])
                        _pid = str(_ps["peserta_id"])
                        _upt_log(f"=== {_uk} | {_ps.get('nama')} ===")
                        if _upt_auto:
                            _ev = _upt.evaluasi_lulus_otomatis(_uk, _pid, progress_cb=_upt_log)
                            if not _ev.get("ok"):
                                _upt_log(f"❌ Evaluasi gagal: {_ev.get('pesan')}")
                                continue
                        _check = _upt.fetch_evaluasi_paket(_uk)
                        _fresh = next((x for x in _check.get("peserta", []) if x["peserta_id"] == _pid), None)
                        if not _fresh or not all(_fresh.get("status", {}).get(x) == "lulus" for x in ("A", "K", "T", "H")):
                            _upt_log("⏭️ Skip: A/K/T/H belum seluruhnya LULUS di SPSE.")
                            continue
                        _waktu = f"{_upt_tgl:%d-%m-%Y} {_upt_jam_mulai:%H:%M}"
                        _sampai = f"{_upt_tgl_selesai:%d-%m-%Y} {_upt_jam_selesai:%H:%M}"
                        _iv = _upt.kirim_undangan_pembuktian(_pid, _waktu, _sampai, _upt_tempat, _upt_dibawa, _upt_hadir)
                        _upt_log(("✅ " if _iv.get("ok") else "❌ ") + _iv.get("pesan", ""))
                    st.success("Proses undangan selesai. Muat ulang status SPSE untuk memverifikasi hasil.")
            with _nc:
                if st.button("Batal", key="upt_confirm_no", use_container_width=True):
                    st.session_state["upt_confirm_send"] = False
                    st.rerun()

# ============================================================
# Tab 6: Download Dokumen Kualifikasi
# ============================================================

if _tender_active_tab == "6️⃣ Download Kualifikasi":
    # ── Auto-fetch paket dari Supabase (tanpa CDP/Chrome) ────────────────────
    _kl_source_paket = None
    if "global_paket_draft" not in st.session_state and "global_paket_aktif" not in st.session_state:
        try:
            from config import sb as _sb_kl
            _kl_rows = _sb_kl().table("draft_paket").select(
                "kode_tender,nama_tender,kode_pokja,nomor_urut"
            ).order("nomor_urut").execute().data or []
            _kl_rows = [r for r in _kl_rows if not _is_tender_excluded(r)]
            # Adapter: sesuaikan format dengan yang diharapkan _get_paket_gabungan()
            # field: kode, nama, id_lelang, pokja, status, tanggal
            _kl_paket_mapped = [
                {
                    "kode": r["kode_tender"],
                    "nama": r["nama_tender"] or r["kode_tender"],
                    "id_lelang": r["kode_tender"],
                    "pokja": r.get("kode_pokja") or "",
                    "nomor_urut": r.get("nomor_urut") or "",
                    "status": "aktif",
                    "tanggal": "",
                }
                for r in _kl_rows
            ]
            # Sumber lokal Tab 5; jangan menimpa global state yang dipakai Tab 0/tab lain.
            _kl_source_paket = _kl_paket_mapped
        except Exception as _kl_fe:
            st.warning(f"⚠️ Gagal memuat paket dari Supabase: {_kl_fe}. Coba klik 'Sinkronkan Paket' di Tab 0.")

    def _kl_get_paket_gabungan() -> list[dict]:
        return _kl_source_paket if _kl_source_paket is not None else _get_paket_gabungan()

    _kl_has_source = _kl_source_paket is not None or (
        "global_paket_draft" in st.session_state or "global_paket_aktif" in st.session_state
    )
    _kl_source_list = [
        p for p in _kl_get_paket_gabungan() if p.get("kode") != "00000000000"
    ]
    _kl_auto_check = 0 < len(_kl_source_list) <= 20
    if _kl_auto_check:
        for _kl_auto_p in _kl_source_list:
            st.session_state.setdefault(f"kl_chk_v3_{_kl_auto_p['kode']}", True)

    # ── Pre-render: fetch hanya paket yang dipilih user ────────────────────────
    if _kl_has_source:
        _kl_perlu_fetch = [
            p for p in _kl_get_paket_gabungan()
            if p.get("kode") != "00000000000"
            and st.session_state.get(f"kl_chk_v3_{p['kode']}", False)
            and f"kl_peserta_{p['kode']}" not in st.session_state
        ]
        if _kl_perlu_fetch:
            with st.spinner(f"Memuat peserta {len(_kl_perlu_fetch)} paket..."):
                for _kl_fp in _kl_perlu_fetch:
                    _kl_id = _kl_fp.get("id_lelang") or _kl_fp["kode"]
                    st.session_state[f"kl_peserta_{_kl_fp['kode']}"] = _fetch_peserta_tender_cached(_kl_id)

    _kl_col1, _kl_col2 = st.columns([2, 3])

    def _kl_paket_label(p: dict) -> str:
        """Label Seksi 2: nomor folder/urut + nama paket penuh; tanpa kode tender."""
        nomor = str(p.get("nomor_urut") or "").strip()
        if not nomor:
            _folder_match = re.match(r"^\s*(\d+)\s*\.\s*", str(p.get("folder_dibuat") or ""))
            nomor = _folder_match.group(1) if _folder_match else ""
        nama = str(p.get("nama") or p.get("nama_tender") or p.get("kode") or "-").strip()
        return f"{nomor}. {nama}" if nomor else nama

    with _kl_col1:
        st.markdown("#### 1. Pilih Paket")
        if not _kl_has_source:
            st.info("⚠️ Data paket belum disinkronkan. Silakan ke **Tab 0** dan klik **🔄 Sinkronkan Paket**.")
        else:
            _kl_paket_list = sorted(
                [p for p in _kl_get_paket_gabungan() if p.get("kode") != "00000000000"],
                key=lambda p: p.get("tanggal", ""),
                reverse=True,
            )
            if not _kl_paket_list:
                st.warning("⚠️ Tidak ada paket aktif ditemukan.")
            else:
                st.caption(f"📋 {len(_kl_paket_list)} paket — centang satu atau lebih:")
                for p in _kl_paket_list:
                    _kl_chk_key = f"kl_chk_v3_{p['kode']}"
                    if _kl_chk_key not in st.session_state:
                        st.session_state[_kl_chk_key] = _kl_auto_check
                    _checked = st.checkbox(
                        f"{_pokja_label(p)[:70]}  \n_{p.get('status', '')}_",
                        key=_kl_chk_key,
                    )

        st.divider()
        _kl_top3_col, _kl_top3_spacer = st.columns([3, 1])
        with _kl_top3_col:
            st.markdown("#### 2. Peserta per Paket")
        _kl_hanya_top3 = st.checkbox("✅ Hanya 3 peserta teratas (harga terendah)", value=True, key="kl_hanya_top3")

        if _kl_has_source:
            _kl_paket_list2 = sorted(
                [p for p in _kl_get_paket_gabungan() if p.get("kode") != "00000000000"],
                key=lambda p: p.get("tanggal", ""),
                reverse=True,
            )
            _kl_paket_aktif_opsi = [p for p in _kl_paket_list2 if st.session_state.get(f"kl_chk_v3_{p['kode']}", False)]
            _kl_kode_aktif = None
            if _kl_paket_aktif_opsi:
                st.caption("Paket tercentang ditampilkan sekaligus. Pilih peserta di masing-masing paket.")
            else:
                st.caption("← Centang paket di Seksi 1 dulu.")

            _kl_ada_terpilih = False
            for p in _kl_paket_list2:
                if not st.session_state.get(f"kl_chk_v3_{p['kode']}", False):
                    continue
                _kl_ada_terpilih = True
                kl_res_p = st.session_state.get(f"kl_peserta_{p['kode']}")
                if kl_res_p is None:
                    st.caption(f"⏳ {_kl_paket_label(p)} — menunggu fetch...")
                elif not kl_res_p["ok"]:
                    st.warning(f"❌ {_kl_paket_label(p)}: {kl_res_p['pesan']}")
                    if st.button("🔄 Retry", key=f"kl_retry_{p['kode']}"):
                        _kl_id = p.get("id_lelang") or p["kode"]
                        with st.spinner("..."):
                            st.session_state[f"kl_peserta_{p['kode']}"] = kualifikasi_engine.fetch_peserta_by_id_lelang(_kl_id)
                        st.rerun()
                else:
                    n_p = len(kl_res_p["peserta"])
                    _kl_limit = min(3, n_p) if _kl_hanya_top3 else n_p
                    _kl_badge = f"Top {_kl_limit} dari {n_p}" if (_kl_hanya_top3 and n_p > 3) else str(n_p)
                    with st.expander(f"**{_kl_paket_label(p)}** — {_kl_badge} peserta", expanded=True):
                        c1, c2, c3, c4 = st.columns(4)
                        with c1:
                            if st.button("👀 Peserta", key=f"kl_refresh_peserta_{p['kode']}", use_container_width=True,
                                         help="Ambil ulang daftar peserta paket ini dari SPSE."):
                                _kl_id = p.get("id_lelang") or p["kode"]
                                with st.spinner("Mengambil peserta dari SPSE..."):
                                    st.session_state[f"kl_peserta_{p['kode']}"] = kualifikasi_engine.fetch_peserta_by_id_lelang(_kl_id)
                                st.rerun()
                        with c2:
                            if st.button("🏆 Top 3", key=f"kl_top3_{p['kode']}", use_container_width=True):
                                for j, ps in enumerate(kl_res_p["peserta"], 1):
                                    st.session_state[f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}"] = (j <= 3)
                                st.rerun()
                        with c3:
                            if st.button("✅ Semua", key=f"kl_all_{p['kode']}", use_container_width=True):
                                for ps in kl_res_p["peserta"]:
                                    st.session_state[f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}"] = True
                                st.rerun()
                        with c4:
                            if st.button("⬜ Batal", key=f"kl_none_{p['kode']}", use_container_width=True):
                                for ps in kl_res_p["peserta"]:
                                    st.session_state[f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}"] = False
                                st.rerun()
                        for i, ps in enumerate(kl_res_p["peserta"], 1):
                            # Default: centang top-3 saja jika toggle aktif
                            _kl_default_cek = (i <= 3) if _kl_hanya_top3 else True
                            st.checkbox(
                                f"{i}. {ps['nama']}",
                                key=f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}",
                                value=st.session_state.get(f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}", _kl_default_cek),
                            )
            if not _kl_ada_terpilih:
                st.caption("← Centang paket di atas untuk memuat peserta.")

    with _kl_col2:
        st.markdown("#### 3. Folder & Aksi")

        # ── Ringkasan paket terpilih + status folder ───────────────────────────
        _kl_paket_dipilih = []
        if _kl_has_source:
            _kl_all_list = sorted(
                [p for p in _kl_get_paket_gabungan() if p.get("kode") != "00000000000"],
                key=lambda p: p.get("tanggal", ""),
                reverse=True,
            )
            for p in _kl_all_list:
                if not st.session_state.get(f"kl_chk_v3_{p['kode']}", False):
                    continue
                kl_res_p = st.session_state.get(f"kl_peserta_{p['kode']}")
                if not kl_res_p or not kl_res_p["ok"]:
                    continue
                peserta_terpilih = [
                    ps for ps in kl_res_p["peserta"]
                    if st.session_state.get(f"kl_cek_{p['kode']}_{ps['kualifikasi_id']}", True)
                ]
                resolve = kualifikasi_engine.resolve_folder_paket(p["kode"])
                _kl_paket_dipilih.append({
                    "paket": p,
                    "peserta": peserta_terpilih,
                    "folder": resolve["path"] if resolve["ok"] else None,
                    "folder_info": resolve["pesan"] if resolve["ok"] else resolve["pesan"],
                    "folder_ok": resolve["ok"],
                })

        if not _kl_paket_dipilih:
            st.info("← Centang paket dan tunggu peserta dimuat.")
        else:
            # ── Publish Paket: finalisasi metadata setelah data tersimpan ────
            # Tidak scrape ulang. Hanya validasi tabel Supabase + tulis manifest.
            import publish_workflow as _pub_wf
            with st.expander("📤 Publish Paket", expanded=True):
                st.caption(
                    "Klik setelah proses Download/Parse selesai. Publish hanya menandai "
                    "snapshot terbaru untuk dibaca Excel."
                )
                _pub_preview = []
                for _pub_item in _kl_paket_dipilih:
                    _pub_p = _pub_item["paket"]
                    _pub_root = os.path.dirname(_pub_item["folder"]) if _pub_item.get("folder") else None
                    _pub_m = _pub_wf.get_manifest(_pub_p["kode"], _pub_root)
                    _pub_preview.append({
                        "Paket": _pub_p["kode"],
                        "Publish terakhir": _pub_m.get("last_published_at", "Belum pernah"),
                        "Status": ", ".join(
                            f"{k}: {v}" for k, v in (_pub_m.get("status") or {}).items()
                        ) or "Belum pernah publish",
                    })
                if _pub_preview:
                    st.dataframe(_pub_preview, use_container_width=True, hide_index=True)
                if st.button(
                    f"📤 Publish {len(_kl_paket_dipilih)} paket",
                    key="kl_publish_paket",
                    type="primary",
                    use_container_width=True,
                    disabled=not _kl_paket_dipilih,
                ):
                    _pub_ok = 0
                    for _pub_item in _kl_paket_dipilih:
                        _pub_p = _pub_item["paket"]
                        _pub_root = os.path.dirname(_pub_item["folder"]) if _pub_item.get("folder") else None
                        _pub_res = _pub_wf.publish_paket(
                            _pub_p["kode"], _pub_root, _pub_item["peserta"]
                        )
                        if _pub_res.get("ok"):
                            _pub_ok += 1
                            _pub_warn = (_pub_res.get("manifest") or {}).get("warnings") or []
                            if _pub_warn:
                                st.warning(f"⚠️ {_pub_p['kode']}: publish berhasil dengan catatan: {'; '.join(_pub_warn)}")
                            else:
                                st.success(f"✅ {_pub_p['kode']}: snapshot berhasil dipublish.")
                        else:
                            st.error(f"❌ {_pub_p['kode']}: {_pub_res.get('error', 'gagal publish')}")
                    if _pub_ok:
                        st.info(f"{_pub_ok}/{len(_kl_paket_dipilih)} paket siap di-refresh dari Excel.")

            # ── Fungsi proses (dipakai tombol global + per-paket) ──────────────
            def _proses_paket_kk(items_to_run, do_download, do_kk, do_excel):
                log_area = st.empty()
                log_lines = []

                def _log_cb(msg):
                    log_lines.append(msg)
                    log_area.code("\n".join(log_lines))

                progress = st.progress(0, text="Memulai...")
                _total_paket = len(items_to_run)
                _total_semua = sum(len(it["peserta"]) for it in items_to_run)
                _total_steps = _total_paket * (2 if (do_download and do_kk) else 1)
                _step = 0

                for item in items_to_run:
                    p = item["paket"]
                    folder_out = item["folder"]
                    peserta_list = item["peserta"]
                    kode_tender = p["kode"]
                    n_ps = len(peserta_list)

                    _log_cb(f"=== Paket {kode_tender}: {p['nama'][:50]} ({n_ps} peserta) ===")

                    # ── Download dokumen ───────────────────────────────────────
                    if do_download and folder_out:
                        kualifikasi_engine.save_last_dir(folder_out)
                        for i, ps in enumerate(peserta_list):
                            progress.progress(
                                _step / _total_steps + (i / n_ps) / _total_steps * (0.5 if do_kk else 1.0),
                                text=f"[{kode_tender}] Download {i+1}/{n_ps}: {ps['nama'][:35]}...",
                            )
                            kualifikasi_engine.download_kualifikasi_peserta(
                                peserta=ps,
                                folder_output=folder_out,
                                urutan=i + 1,
                                total_peserta=n_ps,
                                progress_cb=_log_cb,
                            )
                        _log_cb(f"--- [{kode_tender}] Download selesai ---")
                        _step += 1

                    # ── Parse & simpan KK Evaluasi ─────────────────────────────
                    if do_kk:
                        _log_cb(f"--- [{kode_tender}] Parse KK Evaluasi ---")
                        semua_data = []
                        for i, ps in enumerate(peserta_list):
                            progress.progress(
                                _step / _total_steps + (i / n_ps) / _total_steps,
                                text=f"[{kode_tender}] Parse KK {i+1}/{n_ps}: {ps['nama'][:35]}...",
                            )
                            slug = re.sub(r'[\\/:*?"<>|]', "", ps["nama"]).strip()[:80]
                            folder_p = os.path.join(folder_out or "", f"{i+1}. {slug}")
                            data_p = kualifikasi_parser.parse_peserta_lengkap(
                                kualifikasi_id=ps["kualifikasi_id"],
                                folder_peserta=folder_p,
                                progress_cb=_log_cb,
                                kode_tender=kode_tender,
                            )
                            if data_p.get("skp_berbeda"):
                                _log_cb(f"  ⚠️ {ps['nama']}: SKP berbeda")
                            if data_p.get("sbu_tidak_sesuai"):
                                _log_cb(f"  ⚠️ {ps['nama']}: SBU tidak sesuai syarat paket (ambil baris pertama)")
                            if data_p.get("ss_tidak_sesuai"):
                                _log_cb(f"  ⚠️ {ps['nama']}: SS tidak sesuai syarat paket (ambil baris pertama)")
                            semua_data.append(data_p)

                        try:
                            from config import sb as _sb_kk
                            from datetime import datetime, timezone
                            _db_kk = _sb_kk()
                            rows = []
                            for i, d in enumerate(semua_data):
                                pgl = d.get("pengalaman", [])
                                p1 = pgl[0] if len(pgl) > 0 else {}
                                p2 = pgl[1] if len(pgl) > 1 else {}
                                pemilik = d.get("pemilik", [])
                                akta_p = d.get("akta_pendirian", {})
                                akta_k = d.get("akta_perubahan", {})
                                rows.append({
                                    "kode_tender": kode_tender,
                                    "urutan": i + 1,
                                    "nama": d.get("nama"),
                                    "npwp": kk_evaluasi_engine._format_npwp(d.get("npwp", "")),
                                    "nib_nomor": d.get("nib_nomor"),
                                    "nib_berlaku": d.get("nib_berlaku"),
                                    "ss_nomor": d.get("ss_nomor"),
                                    "ss_berlaku": d.get("ss_berlaku"),
                                    "ss_terverifikasi": d.get("ss_terverifikasi"),
                                    "sbu_nomor": d.get("sbu_nomor"),
                                    "sbu_berlaku": d.get("sbu_berlaku"),
                                    "sbu_kualifikasi": d.get("sbu_kualifikasi"),
                                    "sbu_klasifikasi": d.get("sbu_klasifikasi"),
                                    "sbu_subklas_label": d.get("sbu_subklas_label"),
                                    "pgl1_nama": p1.get("nama"),
                                    "pgl1_instansi": p1.get("instansi"),
                                    "pgl1_nilai": p1.get("nilai"),
                                    "pgl1_tanggal": (f"{p1.get('tgl_mulai','')} s/d {p1.get('tgl_selesai','')}"
                                                     if p1.get("tgl_mulai") else p1.get("tgl_selesai", "")),
                                    "pgl1_nomor": p1.get("nomor"),
                                    "pgl2_nama": p2.get("nama"),
                                    "pgl2_instansi": p2.get("instansi"),
                                    "pgl2_nilai": p2.get("nilai"),
                                    "pgl2_tanggal": (f"{p2.get('tgl_mulai','')} s/d {p2.get('tgl_selesai','')}"
                                                     if p2.get("tgl_mulai") else p2.get("tgl_selesai", "")),
                                    "pgl2_nomor": p2.get("nomor"),
                                    "skp": d.get("skp"),
                                    "skp_catatan": d.get("skp_catatan"),
                                    "skp_berbeda": bool(d.get("skp_berbeda")),
                                    "kswp_status": d.get("kswp_status"),
                                    "akta_p_nomor": akta_p.get("nomor"),
                                    "akta_p_tanggal": akta_p.get("tanggal"),
                                    "akta_p_notaris": akta_p.get("notaris"),
                                    "akta_k_nomor": akta_k.get("nomor"),
                                    "akta_k_tanggal": akta_k.get("tanggal"),
                                    "akta_k_notaris": akta_k.get("notaris"),
                                    "pemilik_1": pemilik[0] if len(pemilik) > 0 else None,
                                    "pemilik_2": pemilik[1] if len(pemilik) > 1 else None,
                                    "pemilik_3": pemilik[2] if len(pemilik) > 2 else None,
                                    "pemilik_4": pemilik[3] if len(pemilik) > 3 else None,
                                    "kinerja_ada": bool(d.get("kinerja_ada")),
                                    "kinerja_nilai": d.get("kinerja_nilai"),
                                    "kinerja_kategori": d.get("kinerja_kategori"),
                                    "updated_at": datetime.now(timezone.utc).isoformat(),
                                })
                            _db_kk.table("kk_evaluasi_peserta").upsert(rows).execute()
                            _log_cb(f"✅ [{kode_tender}] {len(rows)} peserta tersimpan ke Supabase.")

                            # Tulis langsung ke Excel folder paket
                            _excel_path = None
                            if do_excel and folder_out:
                                try:
                                    # folder_out = .../{folder_paket}/1. Dokumen Kualifikasi
                                    # xlsm ada di folder paket (parent), bukan subfolder kualifikasi
                                    _folder_paket = os.path.dirname(folder_out)
                                    _xlsm_list = (
                                        _glob_mod.glob(os.path.join(_folder_paket, "0. BA*.xlsm")) or
                                        _glob_mod.glob(os.path.join(_folder_paket, "*.xlsm"))
                                    )
                                    if _xlsm_list:
                                        _excel_path = _xlsm_list[0]
                                        _log_cb(f"  📝 [{kode_tender}] Menulis KK Evaluasi ke Excel: {os.path.basename(_excel_path)}")
                                        _res_xl = kk_evaluasi_engine.fill_kk_evaluasi(
                                            _excel_path, semua_data, _log_cb)
                                        if _res_xl["ok"]:
                                            _log_cb(f"✅ [{kode_tender}] KK Evaluasi ditulis ke Excel: {_res_xl['pesan']}")
                                        else:
                                            _log_cb(f"⚠️ [{kode_tender}] KK Evaluasi gagal tulis Excel: {_res_xl['pesan']}")
                                    else:
                                        _log_cb(f"⚠️ [{kode_tender}] File .xlsm tidak ditemukan di folder paket, skip tulis Excel.")
                                except Exception as _e_xl:
                                    _log_cb(f"⚠️ [{kode_tender}] Error tulis Excel: {_e_xl}")

                            # Harga Penawaran — scrape SPSE langsung ke Sheet 6 Excel.
                            if do_excel and _excel_path:
                                try:
                                    import penawaran_engine
                                    _hp_peserta = [
                                        {"peserta_id": ps.get("kualifikasi_id", ""),
                                         "nama_peserta": ps.get("nama", "")}
                                        for ps in peserta_list if ps.get("kualifikasi_id")
                                    ]
                                    hasil_hp = penawaran_engine.scrape_penawaran_ke_excel(
                                        kode_tender, _excel_path, progress_cb=_log_cb,
                                        peserta_override=_hp_peserta or None,
                                    )
                                    if hasil_hp["peserta"] > 0:
                                        _log_cb(
                                            f"✅ [{kode_tender}] Penawaran harga langsung ke Sheet 6: "
                                            f"{hasil_hp['peserta']} peserta"
                                        )
                                        _pru = penawaran_engine.update_rumus_penawaran_72(
                                            _excel_path, progress_cb=_log_cb
                                        )
                                        if _pru.get("ok"):
                                            _log_cb(f"  Rumus 7.2 diperbarui: {_pru['rows_updated']} baris")
                                    else:
                                        _log_cb(f"⚠️ [{kode_tender}] Penawaran harga: {hasil_hp.get('errors', ['kosong'])}")
                                except Exception as e_hp:
                                    _log_cb(f"⚠️ [{kode_tender}] Penawaran harga → Excel error: {e_hp}")

                            try:
                                import identitas_engine
                                _po = [{"peserta_id": ps.get("kualifikasi_id", ""), "nama_peserta": ps.get("nama", "")}
                                       for ps in peserta_list if ps.get("kualifikasi_id")]
                                hasil_id = identitas_engine.scrape_dan_upsert_semua(kode_tender, progress_cb=_log_cb, peserta_override=_po or None)
                                _log_cb(f"✅ [{kode_tender}] Identitas: {hasil_id['peserta']} peserta"
                                        if hasil_id["peserta"] > 0 else f"⚠️ [{kode_tender}] Identitas: kosong")
                            except Exception as e_id:
                                _log_cb(f"⚠️ [{kode_tender}] Identitas error: {e_id}")

                            # ── Conflict detection: sync personil & alat dari PDF
                            try:
                                import conflict_engine
                                for i, d in enumerate(semua_data):
                                    pid  = peserta_list[i].get("kualifikasi_id", "")
                                    nama = d.get("nama", "")
                                    conflict_engine.sync_from_pdf(
                                        kode_tender, pid, nama,
                                        d.get("personel_list", []),
                                        d.get("peralatan_list", []),
                                        log=_log_cb,
                                    )
                                _log_cb(f"✅ [{kode_tender}] Conflict sync selesai.")
                            except Exception as e_cf:
                                _log_cb(f"⚠️ [{kode_tender}] Conflict sync error: {e_cf}")

                        except Exception as e_sb:
                            _log_cb(f"ERROR [{kode_tender}] Supabase: {e_sb}")

                        _step += 1

                progress.progress(1.0, text="Selesai!")
                _parts = []
                if do_download: _parts.append("dokumen didownload")
                if do_kk: _parts.append("KK Evaluasi tersimpan")
                st.success(f"✅ Selesai: {' + '.join(_parts)} — {_total_paket} paket, {_total_semua} peserta. Harga penawaran sudah ditulis langsung ke Sheet 6 Excel; lanjutkan **Muat Input BA** jika perlu.")

            # Tampilkan status folder tiap paket
            _kl_semua_folder_ok = True
            for item in _kl_paket_dipilih:
                p = item["paket"]
                fc1, fc2 = st.columns([4, 1])
                with fc1:
                    if item["folder_ok"]:
                        st.success(f"📁 **{p['kode']}** → `...\\{item['folder_info']}\\1. Dokumen Kualifikasi`")
                    else:
                        st.error(f"❌ **{p['kode']}** — {item['folder_info']}")
                        _kl_semua_folder_ok = False
                with fc2:
                    if item["folder_ok"] and st.button("📂", key=f"kl_open_{p['kode']}", help="Buka folder", use_container_width=True):
                        os.startfile(item["folder"])

                n_ps = len(item["peserta"])
                st.caption(f"  → {n_ps} peserta dipilih" if n_ps else "  → ⚠️ Tidak ada peserta dipilih")

                if st.button(
                    f"▶ Jalankan paket ini",
                    key=f"kl_run_{p['kode']}",
                    use_container_width=True,
                    disabled=(n_ps == 0) or (item.get("folder_ok") is False),
                ):
                    _proses_paket_kk(
                        [item],
                        st.session_state.get("kl_opt_download", True),
                        st.session_state.get("kl_opt_kk", True),
                        st.session_state.get("kl_opt_kk", True),  # excel ikut parse
                    )

            st.divider()

            # Hitung total peserta
            _kl_total_semua = sum(len(item["peserta"]) for item in _kl_paket_dipilih)
            _kl_total_paket = len(_kl_paket_dipilih)

            # ── Opsi aksi ──────────────────────────────────────────────────────
            _kl_do_download = st.checkbox("⬇️ Download dokumen kualifikasi", value=True, key="kl_opt_download")
            _kl_do_kk = st.checkbox(
                "📝 Parse KK + Penawaran Harga → tulis langsung ke Excel paket",
                value=True,
                key="kl_opt_kk",
                help="Ambil data KK dan rincian harga dari SPSE lalu tulis langsung ke workbook paket; tidak memakai Supabase sebagai perantara harga.",
            )
            _kl_do_excel = _kl_do_kk  # Excel selalu ikut parse; harga ditulis langsung ke Sheet 6

            _kl_btn_label = []
            if _kl_do_download: _kl_btn_label.append("Download")
            if _kl_do_kk: _kl_btn_label.append("Parse KK + Harga → Excel")
            _kl_btn_text = " + ".join(_kl_btn_label) if _kl_btn_label else "Pilih minimal satu aksi"

            # Tombol global memproses semua paket terpilih; tombol per paket tetap tersedia.
            _kl_items_run = _kl_paket_dipilih
            _kl_n_aktif = sum(len(it["peserta"]) for it in _kl_items_run)

            _kl_disabled = (
                (not _kl_btn_label) or (not _kl_items_run) or (_kl_n_aktif == 0)
                or (_kl_do_download and any(not it["folder_ok"] for it in _kl_items_run))
            )
            if not _kl_semua_folder_ok and _kl_do_download:
                st.warning("⚠️ Ada paket yang foldernya belum ditemukan — buat folder di Tab 0 terlebih dahulu.")

            st.divider()

            if st.button(
                f"▶ Jalankan: {_kl_btn_text} — {_kl_total_paket} paket, {_kl_n_aktif} peserta",
                key="kl_jalankan",
                type="primary",
                use_container_width=True,
                disabled=_kl_disabled,
            ):
                _proses_paket_kk(_kl_items_run, _kl_do_download, _kl_do_kk, _kl_do_excel)

                # ── Tampilkan konflik personil & alat lintas paket
                if _kl_do_kk:
                    try:
                        import conflict_engine as _ce
                        for _kt in [item["paket"]["kode"] for item in _kl_paket_dipilih]:
                            _kf_p = _ce.get_konflik_personil(_kt)
                            _kf_a = _ce.get_konflik_alat(_kt)
                            if _kf_p or _kf_a:
                                with st.expander(f"⚠️ Konflik Lintas Paket — {_kt}", expanded=True):
                                    if _kf_p:
                                        st.markdown("**Personil digunakan di >1 paket:**")
                                        for k in _kf_p:
                                            paket_str = ", ".join(
                                                f"{e['kode_tender']} ({e['nama_penyedia']})"
                                                for e in k["paket"]
                                            )
                                            st.error(f"🔴 {k['nama_personil']} → {paket_str}")
                                    if _kf_a:
                                        st.markdown("**Alat digunakan di >1 paket:**")
                                        for k in _kf_a:
                                            paket_str = ", ".join(
                                                f"{e['kode_tender']} ({e['nama_penyedia']})"
                                                for e in k["paket"]
                                            )
                                            st.warning(f"🟡 {k['nama_alat']} → {paket_str}")
                    except Exception as _e_kf:
                        st.caption(f"Conflict check error: {_e_kf}")

        # ── Dashboard Konflik Personil & Alat (semua paket) ──────────────────
        st.divider()
        st.markdown("### ⚠️ Konflik Personil & Alat Lintas Paket")
        st.caption("Personil atau alat yang diajukan penyedia di >1 paket aktif. Data yang belum tersinkron ditampilkan sebagai coverage, bukan dianggap tidak ada konflik.")

        def _render_konflik_dashboard(trigger_sync_doktek: bool = False):
            """Query + tampilkan dashboard konflik. Ringan — hanya baca Supabase."""
            try:
                import conflict_engine as _ce_dash
                _cov = _ce_dash.get_sync_coverage()
                st.info(
                    f"Coverage data: {_cov['lengkap']}/{_cov['aktif']} paket lengkap "
                    f"(personil {_cov['personil']}, alat {_cov['alat']}). "
                    f"Belum lengkap: {_cov['belum_lengkap']} paket."
                )
                if trigger_sync_doktek:
                    # Sinkronkan paket yang personil atau alatnya belum lengkap.
                    _ce_dash.sync_new_paket()
                # Lookup nama paket
                from config import sb as _sb_kf
                _nama_map = {
                    r["kode_tender"]: r.get("nama_tender") or r["kode_tender"]
                    for r in (_sb_kf().table("draft_paket").select("kode_tender,nama_tender").execute().data or [])
                }

                # Query konflik (selalu dari data yang sudah tersimpan di Supabase)
                _kf_p_all = _ce_dash.get_konflik_personil()
                _kf_a_all = _ce_dash.get_konflik_alat()
                if not _kf_p_all and not _kf_a_all:
                    st.success("✅ Tidak ada konflik ditemukan.")
                else:
                    if _kf_p_all:
                        st.markdown(f"**Personil konflik: {len(_kf_p_all)} nama**")
                        _rows_kf = []
                        _BULAN_ID = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
                        def _fmt_tgl(d):
                            if d is None:
                                return "-"
                            return f"{d.day:02d} {_BULAN_ID[d.month-1]} {d.year}"
                        for k in _kf_p_all:
                            _seen = set()
                            for e in k["paket"]:
                                kt = e["kode_tender"]
                                if kt in _seen:
                                    continue
                                _seen.add(kt)
                                mulai = e.get("tgl_mulai")
                                selesai = e.get("tgl_selesai")
                                periode = f"{_fmt_tgl(mulai)} – {_fmt_tgl(selesai)}" if mulai else "-"
                                _rows_kf.append({
                                    "Nama Personil": k.get("nama_personil_display") or k["nama_personil"],
                                    "Paket": _nama_map.get(kt, kt),
                                    "Penyedia": e["nama_penyedia"] or "-",
                                    "Periode": periode,
                                })
                        st.dataframe(_rows_kf, use_container_width=True, hide_index=True)
                    if _kf_a_all:
                        st.markdown(f"**Alat konflik: {len(_kf_a_all)} nama**")
                        _rows_ka = []
                        for k in _kf_a_all:
                            for e in k["paket"]:
                                kt = e["kode_tender"]
                                _rows_ka.append({
                                    "Nama Alat": k["nama_alat"],
                                    "Paket": _nama_map.get(kt, kt),
                                    "Penyedia": e["nama_penyedia"] or "-",
                                })
                        st.dataframe(_rows_ka, use_container_width=True, hide_index=True)
            except Exception as _e_dash:
                st.error(f"Error: {_e_dash}")

        # Tombol refresh manual — juga trigger sync dari folder doktek
        if st.button("🔄 Cek Konflik + Sinkronkan Data Belum Lengkap", key="kual_cek_konflik"):
            _render_konflik_dashboard(trigger_sync_doktek=True)
        else:
            # Auto-tampil dari data Supabase yang sudah ada (tanpa sync folder)
            _render_konflik_dashboard(trigger_sync_doktek=False)

    # Evaluasi AI dipusatkan di Tab 6 agar DokFull menjadi sumber tunggal.
    if False:
        st.divider()
        st.markdown("#### 🤖 Evaluasi AI (Codex) — Tender")
        st.caption("Codex CLI · Model terkunci: gpt-5.6-luna · Reasoning: medium · Membaca dokumen Tender dan menulis output `.md`. Paralel per paket.")

        import ai_evaluator as _heval_t

        def _prompt_evaluasi_tender(folder_paket, nama_paket):
            return f"""Lakukan evaluasi penawaran Tender Pekerjaan Konstruksi Pascakualifikasi Sistem Gugur.

IDENTITAS PAKET
- Nama paket: {nama_paket}
- Folder paket: {folder_paket}

PREFLIGHT WAJIB — berhenti dengan ERROR jika gagal
1. Pastikan folder paket di atas benar dan dapat dibaca.
2. Baca PROTOKOL_EVALUASI_AI.md dari folder paket/5. Evaluator Kualifikasi & Teknis/.
   Jika tidak ada, gunakan master root _SOP Evaluator/PROTOKOL_EVALUASI_AI.md.
3. Baca evaluator spesifik: EVALUATOR_E2E_TENDER_PK_PASCAKUALIFIKASI.md,
   prioritaskan salinan di folder paket/5. Evaluator Kualifikasi & Teknis/.
4. Pastikan Dokpil tersedia. Prioritaskan `3. Dokpil Full PK v1.docx`, `dokpil_*.pdf`,
   dan dokumen pemilihan di root paket/subfolder `2. Rancangan Kontrak/`.
5. Cari penyedia pada struktur aktual Tender: `1. Dokumen Penawaran/`,
   `1. Dokumen Gabungan/`, dan `8. Dokumen Kualifikasi/`. File gabungan per peserta
   adalah sumber baca utama; file pecahan hanya fallback jika gabungan rusak/tidak terbaca.
   Jika struktur berbeda, lakukan Glob seluruh root paket dan jelaskan struktur aktual.
   Jangan mengarang penyedia atau dokumen.

ATURAN EVALUASI WAJIB
1. Baca Dokpil lebih dahulu; ekstrak syarat LDP/LDK sebelum membaca penawaran.
2. Evaluasi setiap penyedia satu per satu, maksimal 3 peserta utama jika tersedia.
   Sebutkan daftar penyedia yang ditemukan dan jangan diam-diam melewati peserta.
   Jangan mencampur nama, nilai, atau bukti antar peserta.
3. Gunakan bukti dokumen lokal saja untuk administrasi/teknis/kualifikasi; untuk angka
   harga, Sheet `6. Harga Penawaran` adalah sumber rekonsiliasi resmi dan wajib
   dicocokkan dengan SPSE serta DokFull. Jika dokumen/data tidak ada, tulis TIDAK ADA.
4. Ikuti protokol anti-false-positive: cari seluruh dokumen sebelum menyimpulkan tidak ada;
   ragu = FLAG KLARIFIKASI, bukan gugur otomatis.
5. Untuk setiap temuan, cite nama file + halaman/section + kutipan singkat.
6. Jangan memberi skor jika evaluator menetapkan PASS/FAIL. Jangan membuat fakta,
   halaman, dokumen, atau kutipan.
7. Jangan mengubah Excel, DOCM, Supabase, SPSE, atau dokumen sumber. Output otomatis
   hanya `_HASIL_EVALUASI_PK.md` di root folder paket.

OUTPUT WAJIB
- Tulis _HASIL_EVALUASI_PK.md di ROOT folder paket (bukan subfolder).
- Isi hasil per peserta: Administrasi, Teknis, Kualifikasi, Harga, Flag Audit, dan
  Kesimpulan Akhir E2E.
- Pertahankan bagian `<!-- ANALISIS MANUAL -->` jika file output sudah ada; jangan timpa
  analisis manual user.
- Setelah menulis, buka ulang/cek file output dan pastikan tidak kosong.
- Pastikan output memuat: ekstraksi persyaratan Dokpil, evaluasi Administrasi, Teknis,
  Kualifikasi, Harga, Flag Audit/Klarifikasi, dan Kesimpulan Akhir E2E per peserta.
- Jika file sumber, protokol, atau output gagal dibuat, tulis ERROR yang spesifik dan
  jangan menyatakan evaluasi berhasil.

Mulai evaluasi sekarang."""

        _ai_eval_engine_t = "codex"
        _ai_eval_model_t = None
        _ai_t_selected_kode = [p["kode"] for p in _kl_paket_list if st.session_state.get(f"kl_chk_v3_{p['kode']}", False)] if "_kl_paket_list" in dir() else []
        _btn_ai_eval_t = st.button(
            f"🤖 Jalankan Evaluasi AI — {len(_ai_t_selected_kode)} paket",
            key="tkual_btn_ai_eval", use_container_width=True,
            disabled=len(_ai_t_selected_kode) == 0,
        )
        if _btn_ai_eval_t and _ai_t_selected_kode:
            from config import sb as _sb_ait, TENDER_ROOT as _TENDER_ROOT_AIT
            import os as _os_ait
            _ait_rows = _sb_ait().table("draft_paket").select("kode_tender,nama_tender,folder_dibuat").in_("kode_tender", _ai_t_selected_kode).execute().data or []
            _ait_jobs = []
            for _r in _ait_rows:
                _fd = _r.get("folder_dibuat") or ""
                if not _fd:
                    st.warning(f"⚠️ {_tender_display_label(_r)[:60]} — folder belum dibuat, skip.")
                    continue
                _folder_full = _os_ait.path.join(_TENDER_ROOT_AIT, _fd)
                if not _os_ait.path.isdir(_folder_full):
                    st.warning(f"⚠️ {_tender_display_label(_r)[:60]} — folder tidak ditemukan di disk, skip.")
                    continue
                _ait_jobs.append({"nama": _r.get("nama_tender", _fd), "folder": _folder_full})

            if not _ait_jobs:
                st.error("Tidak ada paket dengan folder valid untuk dievaluasi.")
            else:
                st.info(f"🤖 Menjalankan evaluasi AI untuk {len(_ait_jobs)} paket...")
                def _run_ait(job):
                    try:
                        _prompt = _prompt_evaluasi_tender(job["folder"], job["nama"])
                        _out = _heval_t._run_evaluator(_prompt, model=_ai_eval_model_t, add_dirs=[job["folder"]], engine=_ai_eval_engine_t)
                        _hasil_path = _os_ait.path.join(job["folder"], "_HASIL_EVALUASI_PK.md")
                        if not _os_ait.path.isfile(_hasil_path) or _os_ait.path.getsize(_hasil_path) == 0:
                            raise RuntimeError("Codex selesai, tetapi _HASIL_EVALUASI_PK.md tidak dibuat atau kosong.")
                        _missing_eval = _validasi_hasil_evaluasi_tender(_hasil_path)
                        if _missing_eval:
                            raise RuntimeError(
                                "Output evaluasi belum lengkap; bagian hilang: " + ", ".join(_missing_eval)
                            )
                        return {"nama": job["nama"], "status": "ok", "output": _out, "error": ""}
                    except Exception as _e:
                        return {"nama": job["nama"], "status": "error", "output": "", "error": str(_e)}

                _ait_results = []
                with ThreadPoolExecutor(max_workers=3) as _pool_ait:
                    _futures_ait = {_pool_ait.submit(_run_ait, j): j for j in _ait_jobs}
                    for _fut in as_completed(_futures_ait):
                        _ait_results.append(_fut.result())

                for _rt in _ait_results:
                    if _rt["status"] == "ok":
                        st.success(f"✅ {_rt['nama'][:50]}")
                        with st.expander(f"Output evaluasi: {_rt['nama'][:35]}"):
                            st.markdown(_rt["output"][:3000])
                    else:
                        st.error(f"❌ {_rt['nama'][:50]} — {_rt['error'][:200]}")

# ============================================================
# Tab 7: Dokumen Penawaran — Pindah File ke Folder Paket
# ============================================================

if _tender_active_tab == "7️⃣ Dokumen Penawaran":
    import pindah_penawaran_engine as _pe

    st.markdown("### 7️⃣ Dokumen Penawaran")
    st.caption(
        "Alur kerja: scan hasil decrypt Apendo → pindahkan/gabung dokumen → evaluasi dan input BA."
    )

    st.markdown("#### 1. Scan Dokumen Apendo")
    st.caption("Sumber: `D:\\data\\biddings`. Data dicocokkan dengan paket dan peserta di Supabase.")
    _dp_col_scan, _dp_col_info = st.columns([1, 3])
    with _dp_col_scan:
        if st.button("🔍 Scan Ulang Apendo", key="dp_scan", type="primary", use_container_width=True):
            st.session_state.pop("dp_scan_result", None)
            st.rerun()

    if "dp_scan_result" not in st.session_state:
        with st.spinner("Scanning D:\\data\\biddings ..."):
            _raw = _pe.scan_apendo()
            if _raw:
                _enriched = _pe.lookup_supabase(_raw)
            else:
                _enriched = []
            st.session_state["dp_scan_result"] = _enriched

    _dp_items_raw = st.session_state.get("dp_scan_result", [])
    # Cache sesi lama mungkin belum membawa status_tahap; enrich sekali agar
    # paket selesai tetap tersaring tanpa mewajibkan restart aplikasi.
    if _dp_items_raw and any("status_tahap" not in _item for _item in _dp_items_raw):
        _dp_items_raw = _pe.lookup_supabase(_dp_items_raw)
        st.session_state["dp_scan_result"] = _dp_items_raw
    _dp_items = [
        _item for _item in _dp_items_raw
        if not _is_tender_selesai(_item)
        and _is_tender_sudah_pembukaan(_item)
    ]

    with _dp_col_info:
        if _dp_items:
            _dp_paket_n = len({x.get("kode_tender") for x in _dp_items})
            _dp_peserta_n = len(_dp_items)
            _dp_m1, _dp_m2, _dp_m3 = st.columns(3)
            _dp_m1.metric("Paket", _dp_paket_n)
            _dp_m2.metric("Peserta", _dp_peserta_n)
            _dp_m3.metric("Status", "Siap diproses")
        else:
            st.caption("Belum ada hasil scan.")

    if "dp_notif" in st.session_state:
        st.success(st.session_state.pop("dp_notif"))

    # ── Daftar paket bersama untuk seluruh workflow Tab 6 ────────────────────
    _dp_paket_candidates = [
        _r for _r in _load_draft_paket_cached()
        if _r.get("folder_dibuat")
        and not _is_tender_excluded(_r)
        and not _is_tender_selesai(_r)
    ]
    _dp_paket_rows = [
        _r for _r in _dp_paket_candidates
        if _is_tender_sudah_pembukaan(_r)
    ]
    _dp_paket_pra_pembukaan = [
        _r for _r in _dp_paket_candidates
        if not _is_tender_sudah_pembukaan(_r)
    ]
    _dp_paket_rows.sort(key=lambda _r: _r.get("folder_dibuat", ""))

    st.divider()
    st.markdown("#### 2. Pilih Paket")
    if _dp_paket_pra_pembukaan:
        st.info(
            f"⏭️ {len(_dp_paket_pra_pembukaan)} paket disembunyikan karena belum sampai "
            "Pembukaan Dokumen Penawaran."
        )
    if _dp_paket_rows:
        _dp_sel_c1, _dp_sel_c2, _dp_sel_c3 = st.columns([1, 1, 3])
        if _dp_sel_c1.button("✅ Pilih Semua", key="dp_pilih_semua", use_container_width=True):
            for _r in _dp_paket_rows:
                st.session_state[f"dp_chk_{_r['kode_tender']}"] = True
            st.rerun()
        if _dp_sel_c2.button("❌ Batal Semua", key="dp_batal_semua", use_container_width=True):
            for _r in _dp_paket_rows:
                st.session_state[f"dp_chk_{_r['kode_tender']}"] = False
            st.rerun()
        _dp_sel_n = 0
        _dp_check_cols = st.columns(2)
        for _dp_i, _r in enumerate(_dp_paket_rows):
            _dp_kt = _r["kode_tender"]
            _dp_ck = f"dp_chk_{_dp_kt}"
            if _dp_ck not in st.session_state:
                st.session_state[_dp_ck] = True
            if _dp_check_cols[_dp_i % 2].checkbox(
                f"{_r.get('folder_dibuat', _r.get('nama_tender', _dp_kt))}",
                key=_dp_ck,
                help=f"{_dp_kt} · status: {_r.get('status_tahap') or '-'}",
            ):
                _dp_sel_n += 1
        _dp_selected_kodes = {
            _r["kode_tender"] for _r in _dp_paket_rows
            if st.session_state.get(f"dp_chk_{_r['kode_tender']}", False)
        }
        _dp_sel_c3.caption(f"{_dp_sel_n} dari {len(_dp_paket_rows)} paket terpilih untuk aksi bulk.")
    else:
        _dp_selected_kodes = set()
        st.info("Tidak ada paket aktif dengan folder yang dapat diproses.")

    if not _dp_items:
        st.info("Tidak ada data di `D:\\data\\biddings`. Download dulu via Apendo.")
    else:
        # Hitung total peserta per paket untuk resolve_dest
        _dp_total: dict[str, int] = {}
        for _it in _dp_items:
            _dp_total[_it["kode_tender"]] = _dp_total.get(_it["kode_tender"], 0) + 1

        # Kelompokkan per paket untuk tampilan
        _dp_by_paket: dict[str, list] = {}
        for _it in _dp_items:
            _dp_by_paket.setdefault(_it["kode_tender"], []).append(_it)

        _dp_scan_selected = {
            _kt for _kt in _dp_by_paket if _kt in _dp_selected_kodes
        }
        _dp_bulk_c1, _dp_bulk_c2 = st.columns([1, 3])
        with _dp_bulk_c1:
            if st.button(
                "🚚 Pindahkan & Gabung Terpilih",
                key="dp_bulk_pindah",
                type="primary",
                use_container_width=True,
                disabled=not _dp_scan_selected,
            ):
                _bulk_log = []
                _bulk_ok, _bulk_fail = [], []
                for _bulk_kt in _dp_scan_selected:
                    for _bulk_ps in _dp_by_paket[_bulk_kt]:
                        _bulk_dest = _pe.resolve_dest(_bulk_ps, _dp_total)
                        _bulk_res = _pe.pindah_dan_gabung(_bulk_ps, _bulk_dest, log=_bulk_log.append)
                        _bulk_ok.extend(_bulk_res["sukses"])
                        _bulk_fail.extend(_bulk_res["gagal"])
                if _bulk_ok:
                    st.session_state["dp_notif"] = (
                        f"✅ {len(_bulk_ok)} file berhasil dipindah dari {len(_dp_scan_selected)} paket."
                    )
                    st.session_state.pop("dp_scan_result", None)
                    st.rerun()
                if _bulk_fail:
                    st.error(f"❌ {len(_bulk_fail)} file gagal dipindah.")
                    for _e in _bulk_fail:
                        st.caption(f"• {_e}")
        with _dp_bulk_c2:
            st.caption(
                f"{len(_dp_scan_selected)} paket terpilih memiliki hasil scan Apendo. "
                "Aksi per-paket tetap tersedia di expander bawah."
            )

        _dp_card_cols = st.columns(2)
        for _dp_card_i, (_kt, _peserta_list) in enumerate(_dp_by_paket.items()):
            _folder_dibuat = _peserta_list[0].get("folder_dibuat", "")
            _paket_label = _folder_dibuat if _folder_dibuat else _peserta_list[0].get("nama_tender", _kt)
            _folder_paket = _peserta_list[0].get("folder_paket", "")
            _folder_ada = bool(_folder_paket and os.path.isdir(_folder_paket))

            with _dp_card_cols[_dp_card_i % 2].expander(
                f"**{_paket_label}** · {len(_peserta_list)} peserta", expanded=False
            ):
                _dp_s1, _dp_s2, _dp_s3 = st.columns(3)
                _dp_s1.metric("Peserta", len(_peserta_list))
                _dp_s2.metric("Folder paket", "✅ Ada" if _folder_ada else "❌ Tidak ada")
                _dp_s3.metric("Status", "Siap" if _folder_ada else "Perlu Tab 0")
                if not _folder_ada:
                    st.warning("Folder paket belum ditemukan — buat folder di Tab 0 dulu.")
                else:
                    st.caption(f"📂 `{_folder_paket}`")

                for _ps in _peserta_list:
                    _nama = _ps["nama_perusahaan"]
                    _n_teknis = len(_pe._collect_files(_ps["path_teknis"])) if _ps.get("path_teknis") else 0
                    _n_harga  = len(_pe._collect_files(_ps["path_harga"]))  if _ps.get("path_harga")  else 0
                    st.markdown(f"**Peserta {_ps['urutan']} = {_nama}** — {_n_teknis} file teknis, {_n_harga} file harga")

                if _folder_ada:
                    _dp_btn_col1, _dp_btn_col2 = st.columns(2)
                    with _dp_btn_col1:
                        _dp_run_key = f"dp_run_{_kt}"
                        if st.button(
                            f"🚚 Pindahkan & Gabung PDF",
                            key=_dp_run_key,
                            type="primary",
                            use_container_width=True,
                        ):
                            _log_msgs = []
                            _semua_sukses, _semua_gagal = [], []
                            _dest_dirs = []
                            for _ps in _peserta_list:
                                _dest = _pe.resolve_dest(_ps, _dp_total)
                                _dest_dirs.append(_dest)
                                _hasil = _pe.pindah_dan_gabung(_ps, _dest, log=_log_msgs.append)
                                _semua_sukses.extend(_hasil["sukses"])
                                _semua_gagal.extend(_hasil["gagal"])
                            if _semua_sukses:
                                _notif = (
                                    f"✅ {len(_semua_sukses)} file dipindah dari **{_paket_label}** "
                                    f"→ `{_dest_dirs[0]}`"
                                )
                                st.session_state["dp_notif"] = _notif
                                st.session_state.pop("dp_scan_result", None)
                                st.rerun()
                            for _msg in _log_msgs:
                                st.caption(_msg)
                            if _semua_gagal:
                                st.error(f"❌ {len(_semua_gagal)} gagal:")
                                for _e in _semua_gagal:
                                    st.caption(f"• {_e}")
                    with _dp_btn_col2:
                        if st.button(
                            f"📎 Gabung Dok Lengkap",
                            key=f"dp_gabung_{_kt}",
                            use_container_width=True,
                            help="Gabung DoktekFull + DokkualifFull per peserta → 1. Dokumen Gabungan/",
                        ):
                            _gab_log = []
                            _gab_hasil = _pe.gabung_dokumen_lengkap(_folder_paket, log=_gab_log.append)
                            if _gab_hasil["ok"] > 0:
                                st.success(f"✅ {_gab_hasil['ok']} peserta digabung → `1. Dokumen Gabungan/`")
                            if _gab_hasil["gagal"]:
                                st.error(f"❌ {len(_gab_hasil['gagal'])} gagal:")
                                for _e in _gab_hasil["gagal"]:
                                    st.caption(f"• {_e}")
                            for _m in _gab_log:
                                st.caption(_m)

    # ── Seksi 3: Gabung Dok Lengkap (independen dari Scan Apendo) ─────────────
    st.divider()
    st.markdown("#### 3. Gabung Dokumen Lengkap")
    st.caption("Gabung `DoktekFull` + `DokkualifFull` → `1. Dokumen Gabungan/1. DokFull_*.pdf`. Evaluasi AI membaca DokFull ini sebagai sumber gabungan kualifikasi + penawaran.")

    # Gabung/evaluasi dokumen penawaran hanya untuk paket yang sudah masuk
    # tahap Pembukaan atau sesudahnya. Draft tidak boleh ikut aksi bulk.
    _gab_paket_list = _dp_paket_rows

    if _gab_paket_list:
        from config import TENDER_ROOT as _TENDER_ROOT_GAB
        _gab_valid = [
            _gp for _gp in sorted(_gab_paket_list, key=lambda x: x.get("folder_dibuat", ""))
            if os.path.isdir(os.path.join(_TENDER_ROOT_GAB, _gp["folder_dibuat"], "1. Dokumen Penawaran"))
            or os.path.isdir(os.path.join(_TENDER_ROOT_GAB, _gp["folder_dibuat"], "8. Dokumen Kualifikasi"))
        ]
        # Pilihan Section 3 independen dari pilihan Section 2, tetapi daftar
        # paketnya tetap dibatasi ke tahap pasca-pembukaan.
        for _gp in _gab_valid:
            _gab_ck_key = f"gab2_chk_{_gp['kode_tender']}"
            if _gab_ck_key not in st.session_state:
                st.session_state[_gab_ck_key] = True
        _gab_pick_all, _gab_pick_none = st.columns(2)
        if _gab_pick_all.button("✅ Pilih Semua", key="gab2_pilih_semua", use_container_width=True):
            for _gp in _gab_valid:
                st.session_state[f"gab2_chk_{_gp['kode_tender']}"] = True
            st.rerun()
        if _gab_pick_none.button("❌ Batal Semua", key="gab2_batal_semua", use_container_width=True):
            for _gp in _gab_valid:
                st.session_state[f"gab2_chk_{_gp['kode_tender']}"] = False
            st.rerun()
        _gab_selected_kodes = {
            _gp["kode_tender"] for _gp in _gab_valid
            if st.session_state.get(f"gab2_chk_{_gp['kode_tender']}", False)
        }
        _gab_c1, _gab_c2 = st.columns([1, 3])
        _gab_bulk = [r for r in _gab_valid if r["kode_tender"] in _gab_selected_kodes]
        with _gab_c1:
            _gab_run_bulk = st.button(
                "📎 Gabung Paket Terpilih", key="gab2_semua", type="primary",
                use_container_width=True, disabled=not _gab_bulk,
            )
        with _gab_c2:
            st.caption(f"{len(_gab_valid)} paket valid · {len(_gab_bulk)} terpilih untuk aksi bulk.")
        if _gab_run_bulk:
            _gab_all_log = st.empty()
            _gab_all_lines = []
            _gab_all_ok = 0
            for _gp in _gab_bulk:
                _gp_folder = os.path.join(_TENDER_ROOT_GAB, _gp["folder_dibuat"])
                _gab_all_lines.append(f"📦 Paket: {_gp.get('folder_dibuat', _gp['kode_tender'])}")
                _gab_all_log.code("\n".join(_gab_all_lines[-40:]))
                def _log_gab(m, _lines=_gab_all_lines, _area=_gab_all_log):
                    _lines.append(f"  {m}")
                    _area.code("\n".join(_lines[-40:]))
                _r = _pe.gabung_dokumen_lengkap(_gp_folder, log=_log_gab)
                _gab_all_ok += _r["ok"]
            st.success(f"✅ Selesai — {_gab_all_ok} peserta digabung dari {len(_gab_bulk)} paket.")
        st.divider()

        def _harga_sheet6_tender(folder_paket):
            """Baca total final Sheet 6 via Excel COM; read-only, tanpa mengubah XLSM."""
            import glob as _glob_harga
            _xlsm = (
                _glob_harga.glob(os.path.join(folder_paket, "0. BA*.xlsm"))
                or _glob_harga.glob(os.path.join(folder_paket, "*.xlsm"))
            )
            if not _xlsm:
                return ["- TIDAK DITEMUKAN: workbook .xlsm"]
            _xl_harga = _wb_harga = None
            _out_harga = []
            try:
                import win32com.client as _wc_harga, pythoncom as _pyc_harga
                _pyc_harga.CoInitialize()
                _xl_harga = _wc_harga.DispatchEx("Excel.Application")
                _xl_harga.Visible = False
                _xl_harga.DisplayAlerts = False
                _wb_harga = _xl_harga.Workbooks.Open(
                    _xlsm[0], ReadOnly=True, UpdateLinks=0,
                    IgnoreReadOnlyRecommended=True,
                )
                _ws_harga = _wb_harga.Worksheets("6. Harga Penawaran")
                _used_rows = _ws_harga.UsedRange.Rows.Count
                for _start in (1, 10, 19):
                    _nama = str(_ws_harga.Cells(1, _start).Value or "").strip()
                    if not _nama:
                        continue
                    _total = None
                    for _row in range(1, _used_rows + 1):
                        _label = str(_ws_harga.Cells(_row, _start + 1).Value or "").strip().upper()
                        if _label == "TOTAL PENAWARAN SPSE":
                            _total = _ws_harga.Cells(_row, _start + 7).Value
                            break
                    if _total is None:
                        _out_harga.append(f"- {_nama}: TOTAL PENAWARAN SPSE TIDAK DITEMUKAN")
                    else:
                        _out_harga.append(f"- {_nama}: Rp {_total:,.2f} (Sheet 6, total penawaran SPSE)")
                return _out_harga or ["- Sheet 6 kosong"]
            except Exception as _harga_err:
                return [f"- GAGAL MEMBACA SHEET 6 VIA COM: {_harga_err}"]
            finally:
                try:
                    if _wb_harga is not None:
                        _wb_harga.Close(False)
                    if _xl_harga is not None:
                        _xl_harga.Quit()
                except Exception:
                    pass
                try:
                    import pythoncom as _pyc_harga_cleanup
                    _pyc_harga_cleanup.CoUninitialize()
                except Exception:
                    pass

        def _prompt_evaluasi_tender_apendo(folder_paket, nama_paket, dokfull_paths, checklist_paths):
            _dokfull_list = "\n".join(f"- `{p}`" for p in dokfull_paths)
            _checklist_list = "\n".join(f"- `{p}`" for p in checklist_paths) or "- TIDAK DITEMUKAN"
            _harga_sheet6 = "\n".join(_harga_sheet6_tender(folder_paket))
            return f"""Lakukan evaluasi penawaran (pascakualifikasi) untuk paket tender berikut.

Nama paket: {nama_paket}
Folder paket: {folder_paket}

SUMBER DOKUMEN UTAMA — WAJIB
Untuk administrasi dan kualifikasi, gunakan checklist SPSE berikut:
{_checklist_list}
Untuk teknis dan harga, gunakan PDF gabungan berikut:
{_dokfull_list}
ANGKA HARGA OTORITATIF DARI EXCEL — JANGAN DIGANTI DENGAN ANGKA LAIN
Sheet `6. Harga Penawaran` sudah diisi langsung dari SPSE. Gunakan total berikut
untuk evaluasi harga dan perhitungan persentase terhadap HPS:
{_harga_sheet6}
Jika angka pada DokFull atau sumber lain berbeda, catat sebagai konflik data dan
gunakan angka Sheet 6 untuk evaluasi harga; jangan mengambil angka parser lama.
Ikuti strategi hybrid dalam PROTOKOL_EVALUASI_AI.md. Buka PDF spesifik/pecahan
hanya untuk verifikasi tertarget jika ada konflik, scan buruk, atau bukti detail
belum cukup. Jika sumber tidak ada, tandai TIDAK ADA dan jangan mengarang.

Langkah:
1. Baca file PROTOKOL_EVALUASI_AI.md di folder paket (atau subfolder evaluator jika ada).
2. Ikuti seluruh instruksi dalam protokol tersebut — evaluator yang dipakai: EVALUATOR_E2E_TENDER_PK_PASCAKUALIFIKASI.md.
3. Baca setiap checklist dan DokFull, satu peserta per file, tanpa mencampur data.
4. Evaluasi harga berdasarkan angka Sheet 6 yang disediakan di atas; tampilkan
   nominal, persentase HPS, dan status harga terpisah dari administrasi/teknis/kualifikasi.
5. Output: _HASIL_EVALUASI_PK.md di ROOT folder paket.

Jangan mengubah Excel, DOCM, Supabase, SPSE, atau dokumen sumber. Pertahankan bagian
`<!-- ANALISIS MANUAL -->` jika output sudah ada. Output wajib memuat ekstraksi Dokpil,
Administrasi, Teknis, Kualifikasi, Flag Audit/Klarifikasi, dan Kesimpulan Akhir.

Mulai sekarang."""

        _gab_ai_bulk = [r for r in _gab_valid if r["kode_tender"] in _gab_selected_kodes]
        if st.button(
            "🤖 Evaluasi AI Paket Terpilih",
            key="gab2_ai_bulk",
            type="primary",
            use_container_width=True,
            disabled=not _gab_ai_bulk,
        ):
            import ai_evaluator as _heval_bulk
            _bulk_ai_ok, _bulk_ai_fail = 0, 0
            with st.status(f"🤖 Mengevaluasi {len(_gab_ai_bulk)} paket terpilih...", expanded=True) as _bulk_ai_st:
                for _gp_ai in _gab_ai_bulk:
                    _gp_ai_folder = os.path.join(_TENDER_ROOT_GAB, _gp_ai["folder_dibuat"])
                    try:
                        _dokfull_ai = _pe.cari_dokumen_lengkap(_gp_ai_folder)
                        _checklist_ai = _pe.cari_checklist_kualifikasi(_gp_ai_folder)
                        if not _dokfull_ai:
                            raise RuntimeError("DokFull belum tersedia di 1. Dokumen Gabungan; jalankan Gabung Dokumen Lengkap terlebih dahulu.")
                        _prompt_ai = _prompt_evaluasi_tender_apendo(
                            _gp_ai_folder, _gp_ai.get("folder_dibuat", _gp_ai["kode_tender"]), _dokfull_ai, _checklist_ai
                        )
                        _out_ai = _heval_bulk._run_evaluator(
                            _prompt_ai, model=None, add_dirs=[_gp_ai_folder], engine="codex"
                        )
                        _hasil_ai = os.path.join(_gp_ai_folder, "_HASIL_EVALUASI_PK.md")
                        _missing_ai = _validasi_hasil_evaluasi_tender(_hasil_ai)
                        if _missing_ai:
                            raise RuntimeError("Output belum lengkap: " + ", ".join(_missing_ai))
                        _bulk_ai_ok += 1
                        _bulk_ai_st.write(f"✅ {_gp_ai.get('folder_dibuat', '')[:70]}")
                    except Exception as _bulk_ai_e:
                        _bulk_ai_fail += 1
                        _bulk_ai_st.write(f"❌ {_gp_ai.get('folder_dibuat', '')[:60]} — {_bulk_ai_e}")
                _bulk_ai_st.update(
                    label=f"✅ { _bulk_ai_ok } berhasil · ❌ { _bulk_ai_fail } gagal",
                    state="complete" if _bulk_ai_fail == 0 else "error",
                )

        for _gp in _gab_valid:
            _gp_folder = os.path.join(_TENDER_ROOT_GAB, _gp["folder_dibuat"])
            _gp_label = _gp.get("folder_dibuat", _gp["kode_tender"])
            _gp_c1, _gp_c2, _gp_c3 = st.columns([3, 1, 1])
            with _gp_c1:
                st.checkbox(_gp_label, key=f"gab2_chk_{_gp['kode_tender']}")
            with _gp_c2:
                if st.button("📎 Gabung", key=f"gab2_{_gp['kode_tender']}", use_container_width=True):
                    _gab2_log = []
                    _gab2_hasil = _pe.gabung_dokumen_lengkap(_gp_folder, log=_gab2_log.append)
                    if _gab2_hasil["ok"] > 0:
                        st.success(f"✅ {_gab2_hasil['ok']} peserta digabung → `1. Dokumen Gabungan/`")
                    elif not _gab2_hasil["gagal"]:
                        st.info("ℹ️ Tidak ada DoktekFull ditemukan di folder Dokumen Penawaran.")
                    if _gab2_hasil["gagal"]:
                        st.error(f"❌ {len(_gab2_hasil['gagal'])} gagal:")
                        for _ge in _gab2_hasil["gagal"]:
                            st.caption(f"• {_ge}")
                    for _gm in _gab2_log:
                        st.caption(_gm)
            with _gp_c3:
                if st.button("🤖 Eval AI", key=f"gab2_ai_{_gp['kode_tender']}", use_container_width=True):
                    import ai_evaluator as _heval_ap
                    with st.spinner(f"Evaluasi AI {_gp_label[:40]}..."):
                        try:
                            _dokfull_ap = _pe.cari_dokumen_lengkap(_gp_folder)
                            _checklist_ap = _pe.cari_checklist_kualifikasi(_gp_folder)
                            if not _dokfull_ap:
                                raise RuntimeError("DokFull belum tersedia di 1. Dokumen Gabungan; jalankan Gabung Dokumen Lengkap terlebih dahulu.")
                            _prompt_ap = _prompt_evaluasi_tender_apendo(_gp_folder, _gp_label, _dokfull_ap, _checklist_ap)
                            _out_ap = _heval_ap._run_evaluator(_prompt_ap, model=None, add_dirs=[_gp_folder], engine="codex")
                            _hasil_ap = os.path.join(_gp_folder, "_HASIL_EVALUASI_PK.md")
                            _missing_ap = _validasi_hasil_evaluasi_tender(_hasil_ap)
                            if _missing_ap:
                                raise RuntimeError(
                                    "Output evaluasi belum lengkap; bagian hilang: " + ", ".join(_missing_ap)
                                )
                            st.success(f"✅ Evaluasi AI selesai — {_gp_label[:40]}")
                            with st.expander(f"Output evaluasi: {_gp_label[:35]}"):
                                st.markdown(_out_ap[:3000])
                        except Exception as _e_ap:
                            st.error(f"❌ Evaluasi AI gagal: {_e_ap}")
    else:
        st.info("Tidak ada paket ditemukan di Supabase.")

    # ── Seksi 4: Input BA → tulis Excel langsung via COM ─────────────────────
    st.divider()
    st.markdown("#### 4. Input BA ke Excel")
    st.caption(
        "Isi sheet '0. Input BA' (identitas peserta, dokumen penawaran, SKP) "
        "langsung ke file .xlsm via COM. Urutan peserta dari KK Evaluasi."
    )

    _iba_paket_list = _dp_paket_rows

    if not _iba_paket_list:
        st.info("Tidak ada paket dengan folder dibuat.")
    else:
        import input_ba_engine as _iba_eng

        # Opsi global
        _iba_col_opt1, _iba_col_opt2 = st.columns(2)
        with _iba_col_opt1:
            _iba_do_teknis = st.checkbox(
                "Parse Dok Teknis (override alat/personel)",
                value=True,
                key="iba_do_teknis",
            )
        with _iba_col_opt2:
            _iba_do_gcal = st.checkbox(
                "Sync tanggal dari Google Calendar",
                value=True,
                key="iba_do_gcal",
            )
            if _iba_do_gcal:
                try:
                    import gcal_pl_helper as _iba_gcal_check
                    if not _iba_gcal_check.check_gcal_token():
                        st.warning("🔐 Token Google Calendar perlu login ulang.")
                        if st.button("🔑 Login Ulang Google Calendar", key="iba_gcal_reauth", use_container_width=True):
                            import gcal_helper as _iba_gcal_helper
                            with st.spinner("Menunggu login Google..."):
                                _iba_gcal_helper.generate_token()
                            st.success("✅ Token Google Calendar diperbarui.")
                            st.rerun()
                except Exception as _iba_gcal_check_err:
                    st.caption(f"⚠️ Cek token GCal gagal: {_iba_gcal_check_err}")

        # Fungsi proses (dipakai per-paket maupun global)
        def _proses_input_ba(
            kode_tender, nama_tender, folder_kualifikasi, xlsm_path,
            do_teknis, do_gcal, log_area=None, shared_log_lines=None,
        ):
            """Isi sheet '0. Input BA' untuk satu paket."""
            log_area = log_area if log_area is not None else st.empty()
            log_lines = shared_log_lines if shared_log_lines is not None else []

            def _log_cb(msg):
                log_lines.append(msg)
                try:
                    log_area.code("\n".join(log_lines))
                except Exception:
                    pass

            _log_cb(f"=== Input BA: {kode_tender} ===")

            # Folder dok teknis ada di "1. Dokumen Penawaran" (sibling folder kualifikasi),
            # BUKAN di "8. Dokumen Kualifikasi". Resolve dari parent folder paket.
            _folder_paket_root = os.path.dirname(folder_kualifikasi) if folder_kualifikasi else ""
            _folder_penawaran = os.path.join(_folder_paket_root, "1. Dokumen Penawaran") if _folder_paket_root else ""

            # ── 1. (Opsional) parse dok teknis untuk update alat/personel ke Supabase ──
            if do_teknis:
                try:
                    import dokumen_teknis_engine as _dte
                    from config import sb as _sb_dte
                    _ids_r = _sb_dte().table("peserta_identitas").select(
                        "peserta_id,nama_perusahaan"
                    ).eq("kode_tender", kode_tender).execute()
                    _ps_ids = _ids_r.data or []
                    # Daftar subfolder peserta di Dokumen Penawaran (mis. "1. CV SAMATA")
                    _sub_penawaran = []
                    if os.path.isdir(_folder_penawaran):
                        _sub_penawaran = [
                            d for d in os.listdir(_folder_penawaran)
                            if os.path.isdir(os.path.join(_folder_penawaran, d))
                        ]
                    else:
                        _log_cb(f"  ⚠️ Folder '1. Dokumen Penawaran' tidak ada — skip parse teknis")

                    if not _sub_penawaran:
                        _log_cb(
                            "  ℹ️ Struktur Dokumen Penawaran flat; resolver akan mencari "
                            "file teknis di root karena peserta tunggal."
                        )

                    def _cari_folder_peserta(nama):
                        """Cari subfolder yang namanya cocok peserta (abaikan prefix nomor)."""
                        _nrm = re.sub(r'[\s.,]+', '', re.sub(r'[\\/:*?"<>|]', "", nama)).lower()
                        for _d in _sub_penawaran:
                            # Buang prefix "N. " lalu normalisasi
                            _dn = re.sub(r'^\s*\d+\.\s*', '', _d)
                            _dnrm = re.sub(r'[\s.,]+', '', _dn).lower()
                            if _nrm and (_nrm in _dnrm or _dnrm in _nrm):
                                return os.path.join(_folder_penawaran, _d)
                        return None

                    for _ps_id_row in _ps_ids:
                        _pid = _ps_id_row.get("peserta_id", "")
                        _pnama = _ps_id_row.get("nama_perusahaan", "")
                        _folder_ps = _cari_folder_peserta(_pnama)
                        if not _folder_ps and not _sub_penawaran and len(_ps_ids) == 1:
                            _folder_ps = _folder_penawaran
                        if not _folder_ps:
                            if _sub_penawaran:
                                _log_cb(f"  ⚠️ Folder teknis {_pnama} tidak cocok, skip parse")
                            continue
                        try:
                            _res_dt = _dte.parse_dan_upsert(
                                kode_tender, _pid, _folder_ps, _log_cb
                            )
                            if not _res_dt.get("ok"):
                                _log_cb(f"  ⚠️ Teknis {_pnama}: {_res_dt.get('pesan','tidak ada file teknis')}")
                        except Exception as _e_dt:
                            _log_cb(f"  ⚠️ Teknis error {_pnama}: {_e_dt}")
                except Exception as _e_dte:
                    _log_cb(f"  ⚠️ Modul teknis error: {_e_dte}")

            # ── 2. Ambil urutan peserta dari kk_evaluasi_peserta ─────────────────────
            peserta_rows = []
            skp_rows = []
            try:
                from config import sb as _sb_kk2
                _kk_r = _sb_kk2().table("kk_evaluasi_peserta").select(
                    "*"
                ).eq("kode_tender", kode_tender).order("urutan").execute()
                _kk_rows = _kk_r.data or []

                # Dict identitas by nama lowercase
                _id_r = _sb_kk2().table("peserta_identitas").select(
                    "nama_perusahaan,npwp_raw,alamat,nama_direktur,"
                    "personel_1,personel_2,alat_1,alat_2,alat_3,alat_4,alat_5,alat_6"
                ).eq("kode_tender", kode_tender).execute()
                _id_by_nama = {
                    (r.get("nama_perusahaan") or "").lower(): r
                    for r in (_id_r.data or [])
                }

                if _kk_rows:
                    # Susun peserta sesuai urutan KK Evaluasi (maks 3)
                    for _kr in _kk_rows[:3]:
                        _nama_kk = (_kr.get("nama") or "").lower()
                        _id_data = _id_by_nama.get(_nama_kk, {})
                        # Fallback: partial match
                        if not _id_data:
                            for _k, _v in _id_by_nama.items():
                                if _nama_kk and (_nama_kk in _k or _k in _nama_kk):
                                    _id_data = _v
                                    break
                        peserta_rows.append({
                            "nama_perusahaan": _id_data.get("nama_perusahaan") or _kr.get("nama", ""),
                            "npwp":         kk_evaluasi_engine._format_npwp(_id_data.get("npwp_raw") or ""),
                            "alamat":       _id_data.get("alamat", ""),
                            "nama_direktur": _id_data.get("nama_direktur", ""),
                            "personel_1":   _id_data.get("personel_1", ""),
                            "personel_2":   _id_data.get("personel_2", ""),
                            "alat_1":       _id_data.get("alat_1", ""),
                            "alat_2":       _id_data.get("alat_2", ""),
                            "alat_3":       _id_data.get("alat_3", ""),
                            "alat_4":       _id_data.get("alat_4", ""),
                            "alat_5":       _id_data.get("alat_5", ""),
                            "alat_6":       _id_data.get("alat_6", ""),
                        })
                    _log_cb(f"  Peserta (urutan KK): {len(peserta_rows)} dari {len(_kk_rows)}")
                else:
                    # Fallback: urut peserta_identitas apa adanya
                    _log_cb("  ⚠️ kk_evaluasi_peserta kosong, fallback urut peserta_identitas")
                    for _id_row in (_id_r.data or [])[:3]:
                        peserta_rows.append({
                            "nama_perusahaan": _id_row.get("nama_perusahaan", ""),
                            "npwp":         kk_evaluasi_engine._format_npwp(_id_row.get("npwp_raw") or ""),
                            "alamat":       _id_row.get("alamat", ""),
                            "nama_direktur": _id_row.get("nama_direktur", ""),
                            "personel_1":   _id_row.get("personel_1", ""),
                            "personel_2":   _id_row.get("personel_2", ""),
                            "alat_1":       _id_row.get("alat_1", ""),
                            "alat_2":       _id_row.get("alat_2", ""),
                            "alat_3":       _id_row.get("alat_3", ""),
                            "alat_4":       _id_row.get("alat_4", ""),
                            "alat_5":       _id_row.get("alat_5", ""),
                            "alat_6":       _id_row.get("alat_6", ""),
                        })

                # SKP + Hasil Pembuktian per peserta (urut sama dgn peserta_rows)
                def _hitung_hasil_ms(_kr):
                    """Tentukan Memenuhi/Tidak Memenuhi dari field KK Evaluasi."""
                    if (_kr.get("kswp_status") or "").upper() in ("TIDAK VALID", "INVALID"):
                        return "Tidak Memenuhi"
                    if not _kr.get("nib_nomor"):
                        return "Tidak Memenuhi"
                    if not _kr.get("sbu_nomor"):
                        return "Tidak Memenuhi"
                    if not _kr.get("pgl1_nama"):
                        return "Tidak Memenuhi"
                    return "Memenuhi"

                skp_rows = []
                for _kr in _kk_rows[:3]:
                    _skp_int = None
                    try:
                        _skp_int = int(_kr.get("skp")) if _kr.get("skp") is not None else None
                    except (ValueError, TypeError):
                        pass
                    skp_rows.append({
                        "skp": _skp_int,
                        "skp_catatan": _kr.get("skp_catatan") if _kr.get("skp_catatan") is not None else (5 - _skp_int if _skp_int is not None else None),
                        "hasil": _hitung_hasil_ms(_kr),
                    })

            except Exception as _e_kk:
                _log_cb(f"  ⚠️ Gagal ambil peserta dari Supabase: {_e_kk}")
                skp_rows = []

            # ── 3. Dokumen penawaran ─────────────────────────────────────────────────
            dokpen = None
            try:
                # Source of truth untuk angka BA harus snapshot SPSE terbaru.
                # Supabase hanya cache; memakai cache langsung pernah membuat
                # C26:C29 berisi angka dari scrape lama.
                from identitas_engine import scrape_dokumen_penawaran
                dokpen = scrape_dokumen_penawaran(kode_tender)
                _log_cb(
                    "  Dokpen SPSE live: "
                    f"daftar={dokpen.get('jml_daftar')} "
                    f"kirim={dokpen.get('jml_kirim')} "
                    f"tidak_kirim={dokpen.get('jml_tidak_kirim')} "
                    f"tidak_lengkap={dokpen.get('jml_tidak_lengkap')} "
                    f"tidak_dibuka={dokpen.get('jml_tidak_dapat_dibuka')}"
                )
                # Cache hanya untuk histori/layar lain; kegagalan cache tidak
                # boleh menggagalkan penulisan Excel yang sudah punya snapshot.
                try:
                    from identitas_engine import upsert_dokumen_penawaran
                    upsert_dokumen_penawaran(kode_tender, dokpen)
                except Exception as _e_cache:
                    _log_cb(f"  ⚠️ Cache dokpen gagal diperbarui: {_e_cache}")
            except Exception as _e_dp:
                _log_cb(f"  ❌ Gagal mengambil dokumen penawaran langsung dari SPSE: {_e_dp}")
                _log_cb("  ⚠️ Excel tidak diisi dengan cache lama agar tidak menghasilkan BA stale.")
            if dokpen is None:
                _log_cb("  ⛔ Input BA dihentikan untuk paket ini; ulangi setelah sesi SPSE tersedia.")
                return

            # ── 4. (Opsional) Tanggal dari Google Calendar ───────────────────────────
            tgl_pembukaan = None
            tgl_pembuktian = None
            tgl_penetapan = None
            if do_gcal:
                try:
                    import gcal_helper
                    _gcal_hasil = gcal_helper.get_tanggal_ba_dari_gcal(nama_tender)
                    tgl_pembukaan  = _gcal_hasil.get("pembukaan")
                    tgl_pembuktian = _gcal_hasil.get("negosiasi")
                    tgl_penetapan  = _gcal_hasil.get("hasil_pemilihan") or _gcal_hasil.get("evaluasi")
                    _log_cb(
                        "  GCal: "
                        f"pembukaan={tgl_pembukaan or 'tidak ditemukan'}, "
                        f"pembuktian={tgl_pembuktian or 'tidak ditemukan'}, "
                        f"penetapan={tgl_penetapan or 'tidak ditemukan'}"
                    )
                except Exception as _e_gc:
                    _log_cb(f"  ⚠️ GCal error (token expired?): {_e_gc} — tanggal diisi manual")

            # ── 5. Tulis ke Excel ─────────────────────────────────────────────────────
            if not peserta_rows:
                st.warning(f"Tidak ada peserta untuk {kode_tender} — skip tulis Excel.")
                return

            _res = _iba_eng.fill_input_ba(
                xlsm_path,
                peserta_rows,
                dokpen,
                tgl_pembukaan,
                tgl_pembuktian,
                tgl_penetapan,
                skp_rows,
                _log_cb,
            )
            if _res["ok"]:
                st.success(f"Input BA {kode_tender}: {_res['pesan']}")
            else:
                st.warning(f"Input BA {kode_tender} gagal: {_res['pesan']}")

        def _resolve_input_ba_target(_iba_p):
            """Resolve folder kualifikasi + workbook BA untuk satu paket."""
            _kode = _iba_p["kode_tender"]
            _folder_res = kualifikasi_engine.resolve_folder_paket(_kode)
            _folder_kual = _folder_res.get("path", "")
            _folder_paket = os.path.dirname(_folder_kual) if _folder_kual else ""
            _xlsm = ""
            if _folder_paket and os.path.isdir(_folder_paket):
                _candidates = (
                    _glob_mod.glob(os.path.join(_folder_paket, "0. BA*.xlsm")) or
                    _glob_mod.glob(os.path.join(_folder_paket, "*.xlsm"))
                )
                if _candidates:
                    _xlsm = _candidates[0]
            return _folder_kual, _xlsm

        _iba_bulk = [r for r in _iba_paket_list if r["kode_tender"] in _dp_selected_kodes]
        _iba_bulk_c1, _iba_bulk_c2 = st.columns([1, 3])
        with _iba_bulk_c1:
            _iba_run_bulk = st.button(
                "▶ Input BA Paket Terpilih",
                key="iba_bulk_selected",
                type="primary",
                use_container_width=True,
                disabled=not _iba_bulk,
            )
        with _iba_bulk_c2:
            st.caption(f"{len(_iba_bulk)} paket terpilih untuk input BA.")
        if _iba_run_bulk:
            _iba_bulk_log_area = st.empty()
            _iba_bulk_log_lines = []
            _iba_bulk_ok, _iba_bulk_skip = 0, 0
            for _iba_selected in _iba_bulk:
                _iba_folder_kual, _iba_xlsm = _resolve_input_ba_target(_iba_selected)
                if not _iba_xlsm:
                    _iba_bulk_skip += 1
                    st.warning(f"⚠️ {_tender_display_label(_iba_selected)[:60]} — xlsm tidak ditemukan.")
                    continue
                _proses_input_ba(
                    _iba_selected["kode_tender"],
                    _iba_selected.get("nama_tender", _iba_selected["kode_tender"]),
                    _iba_folder_kual,
                    _iba_xlsm,
                    st.session_state.get("iba_do_teknis", True),
                    st.session_state.get("iba_do_gcal", True),
                    log_area=_iba_bulk_log_area,
                    shared_log_lines=_iba_bulk_log_lines,
                )
                _iba_bulk_ok += 1
            st.info(f"Input BA selesai: {_iba_bulk_ok} diproses, {_iba_bulk_skip} dilewati.")

        # Tombol per-paket
        st.divider()
        for _iba_p in _iba_paket_list:
            _iba_kode  = _iba_p["kode_tender"]
            _iba_nama  = _iba_p.get("nama_tender", _iba_kode)
            _iba_label = _iba_p.get("folder_dibuat", _iba_kode)

            # Resolve folder & xlsm
            _iba_folder_kual, _iba_xlsm = _resolve_input_ba_target(_iba_p)

            _iba_c1, _iba_c2 = st.columns([4, 1])
            with _iba_c1:
                _iba_info = f"**{_iba_label}**"
                if _iba_xlsm:
                    _iba_info += f"  \n`{os.path.basename(_iba_xlsm)}`"
                else:
                    _iba_info += "  \n⚠️ xlsm tidak ditemukan"
                st.markdown(_iba_info)
            with _iba_c2:
                if st.button(
                    "▶ Input BA",
                    key=f"iba_{_iba_kode}",
                    type="primary",
                    use_container_width=True,
                    disabled=not bool(_iba_xlsm),
                ):
                    if _iba_xlsm:
                        _proses_input_ba(
                            _iba_kode,
                            _iba_nama,
                            _iba_folder_kual,
                            _iba_xlsm,
                            st.session_state.get("iba_do_teknis", True),
                            st.session_state.get("iba_do_gcal", True),
                        )

