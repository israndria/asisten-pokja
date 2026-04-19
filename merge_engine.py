"""
merge_engine.py — Jembatan antara Asisten Pokja dan Mail Merge Excel+Word

Flow:
  1. Terima kode_tender dari Tab 1 (Kirim Undangan)
  2. Lookup folder_dibuat dari Supabase tabel draft_paket
  3. Cari file Excel .xlsm di folder tsb (pola: @ BA PK*.xlsm atau @ BA PLPK*.xlsm)
  4. Cari file Word "1. Full Dokumen BA PK*.docx" di folder yang sama
  5. Jalankan word_merge.py mode "pdf" → hasilkan Undangan_{kode}.pdf (hal 1-2)
  6. Kembalikan bytes PDF untuk dikirim sebagai lampiran

Dipanggil dari app.py Tab 1 via tombol "Generate Lampiran".
"""

import os
import sys
import subprocess
import glob
import pathlib
from typing import Optional

# ── Lokasi word_merge.py (satu level di atas Asisten_Pokja) ──────────────────
_THIS_DIR = pathlib.Path(__file__).parent          # Asisten_Pokja/
_WPY_DIR  = _THIS_DIR.parent / "V19_Scheduler" / "WPy64-313110"
_PYTHON   = _WPY_DIR / "python" / "python.exe"
_WORD_MERGE = _WPY_DIR / "word_merge.py"


# ── Supabase helper (sama persis dgn inbox_engine._sb()) ─────────────────────

def _sb():
    import pathlib as _pl
    from supabase import create_client
    from dotenv import load_dotenv
    env_path = _pl.Path(__file__).parent.parent / "V19_Scheduler/WPy64-313110/secret_supabase.env"
    if not env_path.exists():
        env_path = _pl.Path(__file__).parent / "secret_supabase.env"
    load_dotenv(env_path)
    import os as _os
    return create_client(_os.environ["SUPABASE_URL"], _os.environ["SUPABASE_KEY"])


def get_folder_by_kode_tender(kode_tender: str) -> Optional[str]:
    """
    Lookup folder_dibuat dari Supabase berdasarkan kode_tender.
    Return: path folder absolut, atau None jika tidak ditemukan / belum dibuat.
    """
    try:
        sb = _sb()
        r = sb.table("draft_paket").select("folder_dibuat").eq("kode_tender", kode_tender).single().execute()
        folder = (r.data or {}).get("folder_dibuat", "")
        if folder and os.path.isdir(folder):
            return folder
    except Exception:
        pass
    return None


def find_excel_file(folder: str) -> Optional[str]:
    """Cari file Excel .xlsm dengan pola '@ BA PK*.xlsm' atau '@ BA PLPK*.xlsm'."""
    for pattern in ["@ BA PK*.xlsm", "@ BA PLPK*.xlsm"]:
        matches = glob.glob(os.path.join(folder, pattern))
        if matches:
            return matches[0]
    return None


def find_word_ba(folder: str) -> Optional[str]:
    """Cari file Word '1. Full Dokumen BA PK*.docx' di folder."""
    matches = glob.glob(os.path.join(folder, "1. Full Dokumen BA PK*.docx"))
    if matches:
        return matches[0]
    return None


def generate_undangan_pdf(kode_tender: str) -> dict:
    """
    Generate Undangan_{kode}.pdf (hal 1-2 BA PK) untuk lampiran undangan.

    Return:
        {
            "sukses": bool,
            "pdf_path": str,      # path file PDF jika sukses
            "pdf_bytes": bytes,   # bytes PDF jika sukses
            "pesan": str,
        }
    """
    # 1. Lookup folder
    folder = get_folder_by_kode_tender(kode_tender)
    if not folder:
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": (
                f"Folder paket untuk kode tender '{kode_tender}' belum dibuat. "
                "Buat folder terlebih dahulu di Tab 0."
            ),
        }

    # 2. Cari Excel
    excel_path = find_excel_file(folder)
    if not excel_path:
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": f"File Excel (.xlsm) tidak ditemukan di folder: {folder}",
        }

    # 3. Cari Word
    word_path = find_word_ba(folder)
    if not word_path:
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": f"File Word '1. Full Dokumen BA PK*.docx' tidak ditemukan di folder: {folder}",
        }

    # 4. Tentukan nama PDF output
    safe_kode = kode_tender.replace("/", "-").replace("\\", "-")
    pdf_expected = os.path.join(folder, f"Undangan_{safe_kode}.pdf")

    # 5. Jalankan word_merge.py via subprocess (COM bridge pattern — hindari conflict)
    cmd = [
        str(_PYTHON),
        str(_WORD_MERGE),
        "pdf",
        word_path,
        excel_path,
        "satu_data",
        safe_kode,
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120,
            cwd=str(_WPY_DIR),
        )
    except subprocess.TimeoutExpired:
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": "word_merge.py timeout (> 2 menit). Coba lagi atau buka Word manual.",
        }
    except Exception as e:
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": f"Gagal menjalankan word_merge.py: {e}",
        }

    # 6. Cek apakah PDF berhasil dibuat
    if not os.path.exists(pdf_expected):
        stderr = result.stderr.strip() if result.stderr else ""
        stdout = result.stdout.strip() if result.stdout else ""
        detail = stderr or stdout or f"Return code: {result.returncode}"
        return {
            "sukses": False,
            "pdf_path": "",
            "pdf_bytes": b"",
            "pesan": f"PDF tidak terbuat setelah word_merge. Detail: {detail[:300]}",
        }

    # 7. Baca bytes PDF
    try:
        with open(pdf_expected, "rb") as f:
            pdf_bytes = f.read()
    except Exception as e:
        return {
            "sukses": False,
            "pdf_path": pdf_expected,
            "pdf_bytes": b"",
            "pesan": f"PDF berhasil dibuat tapi gagal dibaca: {e}",
        }

    return {
        "sukses": True,
        "pdf_path": pdf_expected,
        "pdf_bytes": pdf_bytes,
        "pesan": f"PDF berhasil: {os.path.basename(pdf_expected)} ({len(pdf_bytes)//1024} KB)",
    }


def get_draft_info_from_excel(kode_tender: str) -> dict:
    """
    Baca field dari Excel via openpyxl (read-only) untuk pre-fill form undangan.
    Field yang dibaca dari sheet 'satu_data' (row 2):
      - tanggal_rapat (untuk pre-fill tanggal undangan)
      - tempat_rapat  (untuk pre-fill tempat undangan)

    Return: {"tanggal": date | None, "tempat": str, "pesan": str}
    """
    folder = get_folder_by_kode_tender(kode_tender)
    if not folder:
        return {"tanggal": None, "tempat": "", "pesan": "Folder tidak ditemukan"}

    excel_path = find_excel_file(folder)
    if not excel_path:
        return {"tanggal": None, "tempat": "", "pesan": "File Excel tidak ditemukan"}

    try:
        import shutil, tempfile
        from openpyxl import load_workbook
        from datetime import date as _date

        # Copy ke temp (file mungkin terbuka di Excel)
        tmp = tempfile.mktemp(suffix=".xlsm")
        shutil.copy2(excel_path, tmp)
        wb = load_workbook(tmp, read_only=True, data_only=True, keep_links=False)
        os.remove(tmp)

        hasil = {"tanggal": None, "tempat": "", "pesan": "OK"}

        # Sheet "1. Input Data" — E26 = tanggal DPP, E31 = tempat
        SHEET_INPUT = "1. Input Data"
        if SHEET_INPUT in wb.sheetnames:
            ws = wb[SHEET_INPUT]
            val_tgl = ws["E26"].value
            val_tempat = ws["E31"].value
            if val_tgl:
                if isinstance(val_tgl, _date):
                    hasil["tanggal"] = val_tgl
                else:
                    import datetime as _dt
                    try:
                        hasil["tanggal"] = _dt.datetime.strptime(str(val_tgl).strip(), "%Y-%m-%d").date()
                    except Exception:
                        pass
            if val_tempat:
                hasil["tempat"] = str(val_tempat).strip()

        wb.close()
        return hasil
    except Exception as e:
        return {"tanggal": None, "tempat": "", "pesan": f"Error baca Excel: {e}"}
